import datetime
from django.views import View
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
import json
from utils.exception_helper import to_miladi
from django.template.response import TemplateResponse
from rest_framework.permissions import IsAuthenticated
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from accounts.logger import add_to_log
from accounts.models import Captcha
from base.forms import SearchForm, open_excel, open_excel_card
from base.models import Area, GsList, Baje, UploadExcel, Zone, Parametrs, AutoExcel
from base.permission_decoder import cache_permission
from util import HOME_PAGE, DENY_PAGE
from .models import PanModels, PanHistory, ValidPan, phoneverify, StatusCardAzad, CardAzad, CardHistory, StatusPan, \
    CardLog
from django.middleware.csrf import get_token
from django.contrib import messages
import jalali_date
from .filters import CardFilter, AzadFilter
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from base.models import GsModel, UserPermission, DefaultPermission
import openpyxl
from openpyxl import load_workbook
from rest_framework.decorators import api_view, permission_classes
from jalali.Jalalian import jdate
import jdatetime
from base.views import createotp
import redis
from random import randint
import requests
import pyodbc
from django.conf import settings

today = str(jdatetime.date.today())
today = today.replace("-", "/")
from base.views import checkxss


@cache_permission('insert_car')
def cartinsert(request):

    gs = GsList.objects.filter(owner__user_id=request.user.id)
    return TemplateResponse(request, 'cartinsert.html', {'gs': gs, 'today': today})


@cache_permission('postyaft')
def postyafte(request):
    add_to_log(request, 'مشاهده پست یافته', 0)
    return TemplateResponse(request, 'postyafte.html', {'today': today})


def checknumber(serial):
    serial = serial.replace('۰', '0')
    serial = serial.replace('۱', '1')
    serial = serial.replace('۲', '2')
    serial = serial.replace('۳', '3')
    serial = serial.replace('۴', '4')
    serial = serial.replace('۵', '5')
    serial = serial.replace('۶', '6')
    serial = serial.replace('۷', '7')
    serial = serial.replace('۸', '8')
    serial = serial.replace('۹', '9')

    return serial


@login_required(login_url='accounts:login')
def addpan(request):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        pan = request.POST.get('pan')
        pan = checkxss(pan)
        pan = checknumber(pan)

        id_gs = request.POST.get('id_gs')
        idstatus = request.POST.get('idstatus')
        tarikh = request.POST.get('tarikh')
        tar = tarikh.replace("/", "-")
        unq = str(pan) + str(datetime.date.today())

        _validpan = is_valid_pan(pan)
        if not _validpan:
            return JsonResponse({"message": "error",
                                 'info': 'ساختار شماره پن مشکل دارد'})
        try:
            vin = ValidPan.objects.get(pan=pan)
            vin = vin.vin
        except ObjectDoesNotExist:
            vin = "11111111111111111"

        if PanModels.objects.filter(pan=pan, gs_id=id_gs, statuspan_id=1).count() > 0:
            messages.success(request, 'این شماره سریال قبلا در این جایگاه ثبت شده است', 'warning')
            return JsonResponse({"message": "error", 'info': 'این شماره سریال قبلا در این جایگاه ثبت شده است'})
        else:
            try:

                a = PanModels.objects.create(pan=pan, user_id=request.user.id, statuspan_id=1, tarikhShamsi=tarikh,
                                             tarikh=tar, uniq=unq, vin=vin, gs_id=id_gs,
                                             status=idstatus)
            except IntegrityError:
                return JsonResponse({"message": "error",
                                     'info': 'این کارت در روز جاری در جایگاه دیگری ثبت شده از سامانه scs.niopdc.ir پیگیری بفرمایید'})

        PanHistory.objects.create(user=request.user, pan_id=a.id, status_id=1,
                                  detail='ADD Pan')

        messages.success(request, 'با موفقیت ثبت شد', 'success')
        return JsonResponse({"message": "success"})


def is_valid_pan(PAN):
    PAN_arr = list(PAN)  # شماره کارت رو به لیست ارقام تبدیل میکنه
    if len(PAN_arr) < 16:  # اگر طول شماره کارت کمتر از 16 بود
        return False  # شماره کارت نامعتبره

    for char in PAN_arr:  # بررسی میکنه همه کاراکترها رقم باشند
        if not char.isdigit():  # اگر کاراکتری غیر عددی پیدا شد
            return False  # شماره کارت نامعتبره

    check_sum = 0  # اینجا الگوریتم لان (Luhn) رو اجرا میکنیم
    for i in range(len(PAN_arr)):
        p = int(PAN_arr[i])  # رقم فعلی رو به عدد تبدیل میکنیم

        if i % 2 != 0:  # ارقام در موقعیت فرد (با ایندکس زوج)
            check_sum += p  # بدون تغییر جمع میشن
        else:  # ارقام در موقعیت زوج (با ایندکس فرد)
            doubled = p * 2  # دوبرابر میشن
            check_sum += (doubled - 9 if doubled >= 10 else doubled)  # اگر دوبرابر شد >=10، 9 تا ازش کم میکنیم

    return check_sum % 10 == 0  # اگر جمع نهایی مضرب 10 بود، شماره کارت معتبره


@login_required(login_url='accounts:login')
def addpanpost(request):
    stpan = 2
    if request.method == 'POST':
        pan = request.POST.get('pan')
        idstatus = request.POST.get('idstatus')
        tarikh = request.POST.get('tarikh')
        tar = tarikh.replace("/", "-")
        unq = str(pan) + str(datetime.date.today())
        gsmodel = GsModel.objects.filter(area_id=request.user.owner.area_id, active=True).last()
        id_gs = gsmodel.id
        if request.user.owner.role.role == 'area':
            stpan = 2
        if request.user.owner.role.role == 'zone':
            stpan = 3

        try:
            vin = ValidPan.objects.get(pan=pan)
            vin = vin.vin
        except ObjectDoesNotExist:
            vin = "11111111111111111"

        a = PanModels.objects.create(pan=pan, user_id=request.user.id, statuspan_id=stpan, tarikhShamsi=tarikh,
                                     tarikh=tar, uniq=unq, vin=vin, gs_id=id_gs, SecondCode='پست', status=idstatus)
        PanHistory.objects.create(user=request.user, pan_id=a.id, status_id=1,
                                  detail='ADD Post Pan')

        messages.success(request, 'با موفقیت ثبت شد', 'success')
        return JsonResponse({"message": "success"})


@api_view(['POST'])
def areazone(request, *args, **kwargs):
    if request.method == 'POST':
        mytag = request.POST.get('myTag')
        area = Area.objects.filter(zone_id=mytag)
        if request.user.owner.role.role == 'zone':
            area = Area.objects.filter(zone_id=request.user.owner.zone_id)

        thislist = []
        for q in area:
            thisdict = {
                "id": q.id,
                "name": q.name,
            }
            thislist.append(thisdict)
        return JsonResponse({"mylist": thislist})


@api_view(['POST'])
def areags(request, *args, **kwargs):
    if request.method == 'POST':
        mytag = request.POST.get('myTag')

        gs = GsModel.objects.filter(area_id=mytag)
        if request.user.owner.role.role == 'area':
            gs = GsModel.objects.filter(area_id=request.user.owner.area_id)
        thislist = []
        for q in gs:
            thisdict = {
                "id": q.id,
                "name": q.name,
            }
            thislist.append(thisdict)

        return JsonResponse({"mylist": thislist})


@cache_permission('list_card')
def cartview(request):
    add_to_log(request, 'مشاهده لیست کارت جامانده', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))

    datest = request.GET.get('datest')
    carts = None
    page_obj = None
    if len(datein) < 10:
        datein = "1400-01-01"
        dateout = "2100-12-29"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")
        if len(datein) > 2 > len(dateout):
            dateout = datetime.date.today()

        else:
            d2 = dateout[0] + "-" + dateout[1] + "-" + dateout[2]
            dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        d1 = datein[0] + "-" + datein[1] + "-" + datein[2]

        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
    isfilter = request.GET.get('isfilter')
    myrole = request.user.owner.role.role
    myid = 0
    form = SearchForm()

    panmodels = PanModels.objects.filter(tarikh__gte=datein, tarikh__lte=dateout)
    if datest == "0":
        panmodels = PanModels.objects.filter(tarikh__gte=datein, tarikh__lte=dateout)
    elif datest == "1":
        panmodels = PanModels.objects.filter(create__gte=datein, create__lte=dateout)
    elif datest == "2":
        panmodels = PanModels.objects.filter(tarikhnahye__gte=d1, tarikhnahye__lte=d2)
    elif datest == "3":
        panmodels = PanModels.objects.filter(tarikhsetad__gte=d1, tarikhsetad__lte=d2)
    elif datest == "4":
        panmodels = PanModels.objects.filter(tarikhemha__gte=d1, tarikhemha__lte=d2)
    elif datest == "5":
        panmodels = PanModels.objects.filter(tarikhmalek__gte=d1, tarikhmalek__lte=d2)

    match request.user.owner.role.role:
        case 'setad':
            carts = panmodels.all().order_by('-tarikh')
        case 'mgr':
            carts = panmodels.all().order_by('-tarikh')
        case 'zone':
            myid = request.user.owner.zone_id
            carts = panmodels.filter(
                gs__area__zone_id=request.user.owner.zone_id).order_by('-tarikh')
        case 'area':
            myid = request.user.owner.area_id
            carts = panmodels.filter(
                gs__area_id=request.user.owner.area_id,
                statuspan_id__in=[2, 1, 3, 5]).order_by('-tarikh')
        case 'gs':
            carts = panmodels.filter(
                gs__gsowner__owner_id=request.user.owner.id,
                statuspan_id__in=[1, 2, 5]).order_by(
                'statuspan_id').order_by('-tarikh')

    _filter = CardFilter(request.GET, queryset=carts)
    carts = _filter.qs
    if 'toexcel' in request.GET:
       toexcel = request.GET['toexcel']
       if toexcel == "1":
            url = request.META.get('HTTP_REFERER')
            if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
                messages.warning(request,
                                 'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
                return redirect(url)
            AutoExcel.objects.create(
                datein=str(request.POST.get('select')),
                dateout=str(request.POST.get('select2')),
                titr=str(request.GET.get('titr')),
                fields=request.GET.getlist('fields'),
                owner_id=request.user.owner.id,
                req_id=request.GET,
                reportmodel=4,
                description=str(request.GET.get('datest'))
            )
            messages.warning(request, 'نتیجه عملیات  مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')
            return redirect(url)

    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = form.cleaned_data['search']
            carts = carts.filter(Q(pan__exact=cd) | Q(vin__exact=cd))

    counter = carts.count()
    paginator = Paginator(carts, 10)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_obj = paginator.num_pages
    bypass = Parametrs.objects.first().bypass_sms
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    return TemplateResponse(request, 'LostCardInfo.html',
                            {'carts': page_object, 'query_string': query_string, 'filter': _filter, 'form': form,
                             'bypass': bypass, 'counter': counter,
                             'isfilter': isfilter, 'myrole': myrole, 'myid': myid, 'page_obj': page_obj,
                             })


@login_required(login_url='accounts:login')
def cartdaryaftings(request):
    csrf_token = get_token(request)

    carts = None
    gss = GsModel.objects.filter(area_id=request.user.owner.area_id)
    if request.method == 'POST':
        gsid = request.POST.get('select3')
        if gsid == '0':
            carts = PanModels.objects.filter(gs__nahye_id=request.user.profile.nahye_id, statuspan=1).order_by(
                '-id')
        else:
            carts = PanModels.objects.filter(gs__nahye_id=request.user.profile.nahye_id, statuspan=1,
                                             gs__exact=gsid).order_by(
                '-id')

    return render(request, 'daryaftings.html', {'carts': carts, 'csrf_token': csrf_token, 'gss': gss})


@login_required(login_url='accounts:login')
@transaction.atomic
def carttozone(request):
    if request.method == 'POST':
        add_to_log(request, 'ارسال کارت جامانده به منطقه', 0)
        mylist = request.POST.get('strIds')
        x = mylist.split(',')
        for item in x:
            updatepan = PanModels.objects.get(id=item)
            updatepan.statuspan_id = 3
            updatepan.tarikhsetad = jalali_date.date2jalali(datetime.date.today())
            updatepan.save()
            PanHistory.objects.create(user=request.user, pan_id=updatepan.id, status_id=updatepan.statuspan_id,
                                      detail='ارسال یه منطقه شد')
        return JsonResponse({"message": "success"})


@login_required(login_url='accounts:login')
@transaction.atomic
def carttoemha(request):
    if request.method == 'POST':
        add_to_log(request, 'امحا کارت', 0)
        mylist = request.POST.get('strIds')
        x = mylist.split(',')
        for item in x:
            updatepan = PanModels.objects.get(id=item)
            updatepan.statuspan_id = 4
            updatepan.tarikhemha = jalali_date.date2jalali(datetime.date.today())
            updatepan.save()
            PanHistory.objects.create(user=request.user, pan_id=updatepan.id, status_id=updatepan.statuspan_id,
                                      detail=' امحا شد')
        return JsonResponse({"message": "success"})


@login_required(login_url='accounts:login')
@transaction.atomic
def carttogs(request):
    if request.method == 'POST':
        mylist = request.POST.get('strIds')
        x = mylist.split(',')
        for item in x:
            updatepan = PanModels.objects.get(id=item)
            updatepan.statuspan_id = 1
            updatepan.save()
            PanHistory.objects.create(user=request.user, pan_id=updatepan.id, status_id=updatepan.statuspan_id,
                                      detail='Send To Gs')
        return JsonResponse({"message": "success"})


@login_required(login_url='accounts:login')
@transaction.atomic
def carttonahye(request):
    if request.method == 'POST':
        add_to_log(request, 'ارسال کارت جامانده به ناحیه', 0)
        mylist = request.POST.get('strIds')
        x = mylist.split(',')
        for item in x:
            updatepan = PanModels.objects.get(id=item)
            updatepan.statuspan_id = 2
            updatepan.tarikhnahye = jalali_date.date2jalali(datetime.date.today())
            updatepan.save()
            PanHistory.objects.create(user=request.user, pan_id=updatepan.id, status_id=updatepan.statuspan_id,
                                      detail='تحویل ناحیه شد')
        return JsonResponse({"message": "success"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def carttomalek(request):
    if request.method == 'POST':
        mylist = request.POST.get('strIds')
        codemelimalek = request.POST.get('codemelimalek')
        namemalek = request.POST.get('namemalek')
        mobailmalek = request.POST.get('mobailmalek')
        otp = request.POST.get('otp')
        otp = checknumber(str(otp))
        if Parametrs.objects.first().bypass_sms == False:
            rd = redis.Redis(host=settings.REDIS_HOST, port=settings.REDIS_PORT, db=settings.REDIS_DB,
                             password=settings.REDIS_PASS)
            phone_code = rd.hget(mobailmalek, 'code')
            if phone_code:
                phone_code = phone_code.decode()
        else:
            phone_code = "1111"

        if int(phone_code) == int(otp):
            updatepan = PanModels.objects.get(id=mylist)
            updatepan.statuspan_id = 5
            updatepan.codemelimalek = codemelimalek
            updatepan.showcardmeli = True
            updatepan.showcardcar = True
            updatepan.malek = namemalek
            updatepan.mobailmalek = mobailmalek
            updatepan.tarikhmalek = jalali_date.date2jalali(datetime.date.today())
            updatepan.save()
            # phoneverify.objects.get(phone=mobailmalek).delete()
            PanHistory.objects.create(user=request.user, pan_id=updatepan.id, status_id=updatepan.statuspan_id,
                                      detail='تحویل مالک شد')
            return JsonResponse({"message": "success"})
        else:

            return JsonResponse({"message": "error"})


def pansearch(request):
    templatepage = 'pansearch.html'
    if request.method == 'POST':
        pan = request.POST.get('search')

        result = PanModels.objects.filter(pan=pan, statuspan_id__in=[1, 2]).last()
        if result:
            panhistory = PanHistory.objects.filter(pan_id=result.id)
            return render(request, templatepage, {'result': result, 'panhistory': panhistory})
        else:
            result1 = 'برای این کارت سوخت اطلاعاتی ثبت نشد'
            messages.success(request, result, 'warning')
            return render(request, templatepage, {'result1': result1})
    return render(request, templatepage)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getworkflowcard(request):
    myticket = []
    ok = 0
    name = ''
    if request.method == 'POST':
        obj = request.POST.get('obj')

        pans = PanHistory.objects.filter(pan_id=obj)

        for pan in pans:
            ok += 1
            if pan.status_id == 1:
                name = ' ' + str(pan.pan.gs.name)
            elif pan.status_id == 2:
                name = 'ناحیه ' + str(pan.pan.gs.area.name)
            elif pan.status_id == 3:
                name = 'منطقه ' + str(pan.pan.gs.area.zone.name)

            if pan.status_id == 5:
                name = str(pan.pan.malek) + ' - ' + str(pan.pan.codemelimalek) + ' - ' + str(pan.pan.mobailmalek)
            thisdict = {
                "id": pan.id,
                "info": pan.detail,
                "user": str(pan.user.first_name) + ' ' + str(pan.user.last_name),
                "date": pan.persiandate(),
                "count": ok,
                "name": name,
            }

            myticket.append(thisdict)
    return JsonResponse({"mylist": myticket})


def searchcard(request):
    templatepage = 'SearchCard.html'

    if request.method == 'POST' and 'search' in request.POST:
        captcha_value = request.POST.get('captcha_value')
        captcha_key = request.POST.get('captcha_key')
        captcha_input = request.POST.get('captcha_value')
        captcha_text = request.session.get('captcha_text')

        form = SearchForm(request.POST)
        if form.is_valid():
            try:
                cd = form.cleaned_data['search']

                rd = redis.Redis(host=settings.REDIS_HOST, port=settings.REDIS_PORT, db=settings.REDIS_DB,
                                 password=settings.REDIS_PASS)
                stored_captcha = rd.get(f"captcha:{request.session.session_key}")
                if not stored_captcha:
                    messages.error(request, 'کد امنیتی منقضی شده است')
                    return render(request, templatepage, {'cd': cd, 'status': 3})
                if len(str(captcha_input)) < 4 or not captcha_input:
                    messages.error(request, 'کد امنیتی اشتباه وارد شده است')
                    return render(request, templatepage, {'cd': cd, 'status': 3})
                if captcha_input and captcha_input != captcha_text:
                    messages.error(request, 'کد امنیتی اشتباه وارد شده است')
                    status = 2
                    return render(request, templatepage, {'cd': cd, 'status': 3})
                panmodel = PanModels.objects.filter(statuspan_id__in=[1, 2])
                carts = panmodel.filter(Q(pan__exact=cd) | Q(vin__exact=cd)).last()

                if carts:
                    try:
                        CardLog.objects.create(card_id=cd, ip_address=request.META['REMOTE_ADDR'], status=True)
                    except:
                        pass
                    messages.success(request, 'کارت جامانده شده ، یافت شد.')
                    return render(request, templatepage, {'carts': carts, 'status': 1})
                # else:
                #
                #     carts = Baje.objects.filter(Q(pan__exact=cd) | Q(vin__exact=cd)).last()
                #     if carts:
                #         messages.success(request, 'کارت در باجه معطله موجود است.')
                #         return render(request, templatepage, {'carts': carts, 'status': 2})
                #     else:
                #         try:
                #             CardLog.objects.create(card_id=cd, ip_address=request.META['REMOTE_ADDR'], status=False)
                #         except:
                #             pass
                #         messages.warning(request, "برای این شماره سابقه ایی وجود ندارد.")
            except ObjectDoesNotExist:
                try:
                    CardLog.objects.create(card_id=cd, ip_address=request.META['REMOTE_ADDR'], status=False)
                except:
                    pass
                messages.error(request, "برای این شماره سابقه ایی وجود ندارد.")
            try:
                CardLog.objects.create(card_id=cd, ip_address=request.META['REMOTE_ADDR'], status=False)
            except:
                pass
        return render(request, templatepage, {'cd': cd, 'status': 4})
    return render(request, templatepage, {'status': 0})


@cache_permission('cardazad')
def card_azad(request):
    statuss = StatusCardAzad.objects.all()
    isfilter = request.GET.get('isfilter')
    tedad = 0

    myrole = request.user.owner.role.role
    myid = 0
    cards = None
    if request.user.owner.role.role == 'setad':
        cards = CardAzad.objects.all()
    if request.user.owner.role.role == 'mgr':
        cards = CardAzad.objects.all()
    elif request.user.owner.role.role == 'zone':
        myid = request.user.owner.zone_id
        cards = CardAzad.objects.filter(gs__area__zone_id=request.user.owner.zone_id)
    elif request.user.owner.role.role == 'area':
        myid = request.user.owner.area_id
        cards = PanModels.objects.filter(gs__area_id=request.user.owner.area_id)
    elif request.user.owner.role.role == 'gs':
        cards = PanModels.objects.filter(gs__gsowner__owner_id=request.user.owner.id, status_id=3)
    _filter = AzadFilter(request.GET, queryset=cards)
    cards = _filter.qs
    form = SearchForm()
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = form.cleaned_data['search']
            cards = cards.filter(Q(pan__exact=cd) | Q(vin__exact=cd))
    if isfilter == '0':
        page_object = cards
        query_string = None
    else:
        paginator = Paginator(cards, 50)
        tedad = paginator.count
        page_num = request.GET.get('page')
        data = request.GET.copy()
        if 'page' in data:
            del data['page']
        query_string = request.META.get("QUERY_STRING", "")

        if query_string.startswith("page"):
            query_string = query_string.split("&", 1)
            query_string = query_string[1]
        page_object = paginator.get_page(page_num)

    if request.method == 'POST':
        ststus = request.POST.get('action')
        mylist = request.POST.get('select_across')
        x = mylist.split(',')

        for item in x:
            updatecard = CardAzad.objects.get(id=item)
            updatecard.status_id = ststus
            updatecard.save()
            CardHistory.objects.create(owner_id=request.user.owner.id, card_id_id=updatecard.id, status_id=ststus
                                       )
    context = {'statuss': statuss, 'cards': page_object, 'query_string': query_string, 'filter': _filter, 'form': form,
               'isfilter': isfilter, 'tedad': tedad, 'myrole': myrole, 'myid': myid}
    return TemplateResponse(request, 'cardazad.html', context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getworkflowcardazad(request):
    myticket = []
    ok = 0
    if request.method == 'POST':
        obj = request.POST.get('obj')
        pans = CardHistory.objects.filter(card_id_id=obj)
        for pan in pans:
            ok += 1
            thisdict = {
                "id": pan.id,
                "info": pan.status.name,
                "user": str(pan.owner.name) + ' ' + str(pan.owner.lname),
                "date": pan.persiandate(),
                "count": ok,
                "name": pan.status.name,
            }

            myticket.append(thisdict)

    return JsonResponse({"mylist": myticket})


@cache_permission('0')
def import_excel(request):
    url = request.META.get('HTTP_REFERER')
    form = open_excel(request.POST)
    if request.method == 'POST':
        add_to_log(request, 'دریافت اکسل کارت جامانده', 0)
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            try:
                form.save()
            except ValidationError as e:
                print(44)
                messages.error(request,
                               f'عملیات شکست خورد ، نوع یا سایز فایل مشکل دارد. نام فایل باید کارکتر انگلیسی باشد ')
                return redirect(url)
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                pan = sheet_obj.cell(row=i, column=1).value
                vin = sheet_obj.cell(row=i, column=2).value
                gs_id = sheet_obj.cell(row=i, column=5).value
                noe = sheet_obj.cell(row=i, column=3).value
                product = sheet_obj.cell(row=i, column=4).value
                if noe == 'کارت آزاد':
                    noe = 1
                if noe == 'کارت حواله ای':
                    noe = 2
                product_id = 4
                if product == 'بنزین':
                    product_id = 2

                try:
                    gs = GsModel.objects.get(gsid=gs_id)
                    updatecard = CardAzad.objects.create(pan=pan, vin=vin, gs_id=gs.id, cardst=noe,
                                                         product_id=product_id, status_id=1)
                    CardHistory.objects.create(owner_id=request.user.owner.id, card_id_id=updatecard.id, status_id=1)
                except IntegrityError:
                    continue

        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


@cache_permission('0')
def import_excel_baje(request, _id):
    if _id == 1:
        Baje.objects.all().delete()
    form = open_excel(request.POST)
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                pan = sheet_obj.cell(row=i, column=2).value
                # tarikh = sheet_obj.cell(row=i, column=4).value
                barcode = sheet_obj.cell(row=i, column=1).value
                mobail = sheet_obj.cell(row=i, column=3).value
                if pan:
                    try:
                        vin = ValidPan.objects.get(pan=pan).vin

                    except:
                        vin = pan

                    try:
                        Baje.objects.create(pan=pan, vin=vin, status=0, barcode=barcode, mobail=mobail)
                    except:
                        continue

        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


def import_access_baje(request):
    db_driver = '{Microsoft Access Driver (*.mdb, *.accdb)}'
    db_path = 'F:\\123.accdb'
    conn_str = (rf'DRIVER={db_driver};'
                rf'DBQ={db_path};')

    conn = pyodbc.connect(conn_str)
    conn = pyodbc.connect(
        r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=F:\\123.accdb;')
    cursor = conn.cursor()
    cursor.execute("select * from table1")
    for item in cursor.fetchall():

        pan = item.Field1

        tarikh = item.Field4
        barcode = item.Field2
        mobail = item.Field3

        if item.Field1:
            pan = pan.split("(")
            pan = pan[0]

            if pan:
                try:
                    vin = ValidPan.objects.get(pan=pan).vin
                    print(vin)
                except:
                    vin = pan

                try:
                    Baje.objects.create(pan=pan, vin=vin, tarikh=tarikh, barcode=barcode, mobail=mobail)
                except:
                    continue

    conn.close()
    return HttpResponse('1')


def addvin():
    result = PanModels.objects.filter(vin='11111111111111111')
    for item in result:
        try:
            get = ValidPan.objects.get(pan=item.pan)
            item.vin = get.vin
            item.save()
        except ValidPan.DoesNotExist:
            continue
    return HttpResponse('ok')


def import_pan_vin():
    path = "media/post/pan.txt"
    i = 0
    with open(path, 'r') as f:
        for line in f:
            _line = line.split(',')
            pan = _line[1]
            vin = _line[0]
            ValidPan.objects.create(pan=pan, vin=vin)  # i += 1

    return True


@cache_permission('expirecard')
def repcardexpire(request):
    add_to_log(request, 'مشاهده گزارش کارت جامانده', 0)
    _list = []
    zone_id = 0
    zones = Zone.objects_limit.all()
    gss = None
    if request.user.owner.role.role == 'zone':
        gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    if request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area_id)

    if request.method == 'POST':
        zone_id = request.POST.get('zone')
        gss = GsModel.objects.filter(area__zone_id=int(zone_id))

    if gss:
        for gs in gss:
            cards_expire = PanModels.objects.filter(statuspan_id=1, gs_id=gs.id)
            card_expire_count = 0
            for item in cards_expire:
                if item.expire_date() > 0:
                    card_expire_count += 1
            _dict = {
                'gsid': gs.gsid,
                'name': gs.name,
                'area': gs.area.name,
                'zone': gs.area.zone.name,
                'tedad': cards_expire.count(),
                'expire': card_expire_count,
            }
            _list.append(_dict)

        context = {'list': _list, 'zones': zones, 'zone_id': int(zone_id)}
        return render(request, 'report/report_card_expire.html', context)
    context = {'zones': zones}
    return TemplateResponse(request, 'report/report_card_expire.html', context)


@cache_permission('expirecard')
def repcardexpirearea(request):
    add_to_log(request, 'مشاهده گزارش کارت جامانده نواحی', 0)
    _list = []
    zone_id = 0
    zones = Area.objects.all()
    gss = None
    if request.user.owner.role.role in ['zone', 'area']:
        gss = Area.objects.filter(zone_id=request.user.owner.zone_id)

    if request.method == 'POST':
        zone_id = request.POST.get('zone')
        gss = Area.objects.filter(zone_id=int(zone_id))

    if gss:
        for area in gss:
            cards_expire = PanModels.objects.filter(statuspan_id=2, gs__area_id=area.id)
            card_expire_area_count = 0
            for item in cards_expire:
                if item.expire_date_area() > 0:
                    card_expire_area_count += 1
            _dict = {

                'area': area.name,
                'zone': area.zone.name,
                'tedad': cards_expire.count(),
                'expire': card_expire_area_count,
            }
            _list.append(_dict)

        context = {'list': _list, 'zones': zones, 'zone_id': int(zone_id)}
        return TemplateResponse(request, 'report/report_card_expire.html', context)
    context = {'zones': zones}
    return TemplateResponse(request, 'report/report_card_expire.html', context)


def delcart(request):
    gs = GsModel.objects.all()
    for item in gs:
        carts = PanModels.objects.values('pan').filter(gs_id=item.id, statuspan_id=1).annotate(gs1=Count('id'))
        for cart in carts:
            if int(cart['gs1']) > 1:
                delcart = PanModels.objects.filter(pan=cart['pan']).order_by('-create')
                i = 0
                for _del in delcart:
                    i += 1
                    if i > 1:
                        # print(_del.id)
                        _del.delete()
    return HttpResponse('ok')


@cache_permission('list_card')
def pan_report(request):
    add_to_log(request, 'گزارش کارت جامانده ثبت شده جایگاه', 0)
    context = {}
    report_data = None
    show_details = False
    details = None
    status_filter = None
    gs_filter = None
    total_emha = 0
    total_zone = 0
    total_gs = 0
    total_owner = 0
    total_area = 0

    if request.method == 'POST':
        if 'start_date' in request.POST:  # فیلتر اصلی گزارش
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            start_date = start_date.replace('/', '-')
            end_date = end_date.replace('/', '-')

            all_gs = GsModel.object_role.c_gsmodel(request).all()
            if start_date and end_date:
                pans = PanModels.object_role.c_gs(request, 0).filter(
                    create__gte=start_date,
                    create__lte=end_date
                )
            else:
                pans = None

            report_data = {}
            status_gs = StatusPan.objects.filter(id=1).first()
            status_owner = StatusPan.objects.filter(id=5).first()
            status_area = StatusPan.objects.filter(id=2).first()
            status_zone = StatusPan.objects.filter(id=3).first()
            status_emha = StatusPan.objects.filter(id=4).first()

            for gs in all_gs:
                gs_name = str(gs.gsid) + " - " + str(gs.name)
                report_data[gs_name] = {
                    'total': 0,
                    'owner': 0,
                    'area': 0,
                    'gs': 0,
                    'zone': 0,
                    'emha': 0,
                    'other': 0,
                    'gs_id': gs.id
                }

            for pan in pans.select_related('gs', 'statuspan'):
                gs_name = str(pan.gs.gsid) + " - " + str(pan.gs.name) if pan.gs else "بدون جایگاه"

                if gs_name not in report_data:
                    report_data[gs_name] = {
                        'total': 0,
                        'owner': 0,
                        'area': 0,
                        'gs': 0,
                        'zone': 0,
                        'emha': 0,
                        'other': 0,
                        'gs_id': pan.gs.id if pan.gs else None
                    }

                report_data[gs_name]['total'] += 1

                if pan.statuspan == status_owner:
                    report_data[gs_name]['owner'] += 1
                    total_owner += 1
                elif pan.statuspan == status_area:
                    report_data[gs_name]['area'] += 1
                    total_area += 1
                elif pan.statuspan == status_gs:
                    report_data[gs_name]['gs'] += 1
                    total_gs += 1
                elif pan.statuspan == status_zone:
                    report_data[gs_name]['zone'] += 1
                    total_zone += 1
                elif pan.statuspan == status_emha:
                    report_data[gs_name]['emha'] += 1
                    total_emha += 1
                else:
                    report_data[gs_name]['other'] += 1

            # Convert report_data to a list of tuples for easier iteration in template
            report_data = sorted(report_data.items(), key=lambda x: x[1]['total'], reverse=True)

        elif 'status_filter' in request.POST:  # فیلتر برای نمایش جزئیات
            show_details = True
            start_date = request.POST.get('detail_start_date')
            end_date = request.POST.get('detail_end_date')
            status_filter = request.POST.get('status_filter')
            gs_filter = request.POST.get('gs_filter')

            start_date = start_date.replace('/', '-')
            end_date = end_date.replace('/', '-')

            details = PanModels.object_role.c_gs(request, 0).filter(
                create__gte=start_date,
                create__lte=end_date
            )

            if gs_filter and gs_filter != 'all':
                details = details.filter(gs_id=gs_filter)

            if status_filter == 'gs':
                details = details.filter(statuspan_id=1)
            elif status_filter == 'owner':
                details = details.filter(statuspan_id=5)
            elif status_filter == 'area':
                details = details.filter(statuspan_id=2)
            elif status_filter == 'zone':
                details = details.filter(statuspan_id=3)
            elif status_filter == 'emha':
                details = details.filter(statuspan_id=4)

            # برای ارسال به اکسل
            if 'export' in request.POST:

                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename="cards_list.xlsx"'

                wb = Workbook()
                ws = wb.active
                ws.title = "کارت ها"

                # عنوان ستون ها
                columns = ['ردیف', 'شماره پن', 'وضعیت', 'جایگاه', 'تاریخ ثبت']
                ws.append(columns)

                # داده ها
                for idx, card in enumerate(details, 1):
                    ws.append([
                        idx,
                        card.pan,
                        card.statuspan.info,
                        f"{card.gs.gsid} - {card.gs.name}" if card.gs else "بدون جایگاه",
                        str(idx)
                    ])

                wb.save(response)
                return response

    context = {
        'report_data': report_data,
        'show_details': show_details,
        'details': details,
        'status_filter': status_filter,
        'gs_filter': gs_filter,
        'gs_list': GsModel.objects.all() if report_data else None,
        'total_gs': total_gs,
        'total_owner': total_owner,
        'total_area': total_area,
        'total_zone': total_zone,
        'total_emha': total_emha,
    }
    return render(request, 'pan_report.html', context)

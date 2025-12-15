import re
from datetime import datetime
import json
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum, F, ExpressionWrapper, IntegerField
from base.forms import open_excel
from util import EXCEL_MODE
from utils.exception_helper import checknumber, checkxss, to_miladi
import jdatetime
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from accounts.logger import add_to_log
from base.permission_decoder import cache_permission
from .forms import InstallLockForm, SoratJalaseForm
from .models import Seris, LockModel, SendPoshtiban, InsertLock, Peymankar, LockLogs, Position, GetLockPeymankar, Status
from django.contrib import messages
from base.models import Owner, Zone, UploadExcel, GsModel, Pump
from django.db import transaction
from django.contrib.auth.models import User
from .filters import LockFilter
import random
import string
from django.db import transaction


@cache_permission('lockunit')
def inputlockzonelist(request):
    _role = 'smart'
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    add_to_log(request, f'مشاهده فرم تخصیص پلمب به واحد ', 0)

    # form_class = UserLoginForm
    template_file = 'inputlockzonelist.html'
    _list = InsertLock.objects.filter(zone_id=request.user.owner.zone_id, peymankar__ename=_role).order_by(
        '-tarikh').annotate(
        count_status_6=Count('lockmodel', filter=Q(lockmodel__status_id__in=[6, 9, 10, 11])),
        count_status_5=Count('lockmodel', filter=Q(lockmodel__status_id=5)),
        remaining=ExpressionWrapper(
            F('tedad') - (F('count_status_6') + F('count_status_5')),
            output_field=IntegerField())

    )
    _sum = _list.aggregate(tedad=Sum('tedad'), sum_count_status_6=Sum('count_status_6'),
                           sum_count_status_5=Sum('count_status_5'), sum_remaining=Sum('remaining'))

    context = {'list': _list, 'sum': _sum}
    return TemplateResponse(request, template_file, context)


@transaction.atomic
@cache_permission('lockunit')
def inputlockzoneadd(request):
    _role = 'smart'
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    template_file = 'inputlockzoneadd.html'
    context = {
        'seris': Seris.objects.all(),
        'peymankars': Peymankar.objects.filter(active=True, ename=_role),
        'positions': Position.objects.all()
    }

    if request.method == "POST":
        try:
            # اعتبارسنجی و پردازش داده‌های ورودی
            tarikh = request.POST.get('select')
            if not tarikh:
                raise ValueError("تاریخ وارد نشده است")

            # تبدیل تاریخ
            try:
                tarikh = tarikh.split("/")
                tarikh = jdatetime.date(day=int(tarikh[2]), month=int(tarikh[1]), year=int(tarikh[0])).togregorian()
            except (ValueError, IndexError):
                raise ValueError("فرمت تاریخ نامعتبر است")

            # اعتبارسنجی سریال‌ها
            serial_in = checknumber(checkxss(request.POST.get('serial_in', '')))
            serial_out = checknumber(checkxss(request.POST.get('serial_out', '')))

            if not all([serial_in, serial_out]):
                raise ValueError("شماره سریال‌ها وارد نشده است")

            tedad = (int(serial_out) - int(serial_in)) + 1
            if tedad <= 0:
                raise ValueError("محدوده سریال‌ها نامعتبر است")
            if tedad > 500:
                messages.warning(request, 'تعداد پلمب در هر بسته حداکثر می‌تواند 500 عدد باشد')
                return TemplateResponse(request, template_file, context)

            # بررسی تکراری نبودن سریال‌ها قبل از ذخیره
            seri_id = int(request.POST.get('seri', 0))
            peymankar_id = int(request.POST.get('peymankar', 0))

            if not all([seri_id, peymankar_id]):
                raise ValueError("سری یا پیمانکار انتخاب نشده است")

            # ایجاد رکورد
            InsertLock.objects.create(
                tarikh=tarikh,
                seri_id=seri_id,
                serial_in=serial_in,
                serial_out=serial_out,
                tedad=tedad,
                peymankar_id=peymankar_id,
                zone_id=request.user.owner.zone_id,
                user_id=request.user.id
            )

            add_to_log(request, f'اضافه کردن پلمب به واحد از {serial_in} تا {serial_out}', 0)
            messages.success(request, 'عملیات با موفقیت انجام شد')
            return redirect('lock:input_lock_zone__list')

        except Exception as e:
            transaction.set_rollback(True)
            messages.error(request, f'خطا در انجام عملیات: {str(e)}')
            context['msg'] = str(e)
            return TemplateResponse(request, template_file, context)

    return TemplateResponse(request, template_file, context)


@cache_permission('lockunit')
def residpeymankarlist(request):
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    add_to_log(request, f'مشاهده فرم رسید پلمب توسط منطقه ', 0)

    # form_class = UserLoginForm
    template_file = 'residinpeymankar.html'
    _list = GetLockPeymankar.objects.filter(zone_id=request.user.owner.zone_id, resid=False, ename=_role).order_by(
        '-created_at')
    context = {'list': _list}
    return TemplateResponse(request, template_file, context)


@cache_permission('lockunit')
def backresidpeymankarlist(request, _id):
    url = request.META.get('HTTP_REFERER')
    add_to_log(request, f'برگرداندن از رسید پلمب توسط منطقه ', 0)
    result = GetLockPeymankar.objects.get(id=_id)
    _serials = LockModel.objects.filter(getlockpeymankar_id=result.id)
    for item in _serials:
        item.status_id = 9
        item.getlockpeymankar_id = None
        item.input_date_poshtiban = None
        item.idp_user_id = None
        item.save()
    result.delete()
    return redirect(url)


@cache_permission('locktek')
def sendtoposhtibanlist(request):
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    add_to_log(request, f'مشاهده فرم تخصیص پلمب به تکنسین ', 0)
    template_file = 'sendtoposhtibanlist.html'
    _list = SendPoshtiban.objects.filter(zone_id=request.user.owner.zone_id, ename=_role).order_by('-tarikh',
                                                                                                   '-id').annotate(
        count_status_6=Count('lockmodel', filter=Q(lockmodel__status_id__in=[6, 9, 10, 11])),
        count_status_5=Count('lockmodel', filter=Q(lockmodel__status_id=5))
    )
    context = {'list': _list}
    return TemplateResponse(request, template_file, context)


@cache_permission('locktek')
def sendtoposhtibanadd(request):
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    template_file = 'sendtoposhtibanadd.html'
    from django.db.models import F

    _list2 = InsertLock.objects.filter(
        zone_id=request.user.owner.zone_id,
        resid=True,
        peymankar__ename=_role
    ).annotate(
        # تعداد سریال‌هایی که به تکنسین‌ها اختصاص داده شده‌اند (status_id=4)
        count_assigned_to_tech=Count('lockmodel', filter=Q(lockmodel__status_id=4)),

        # تعداد کل سریال‌های موجود در بسته
        total_serials=F('tedad')
    ).filter(
        # فقط بسته‌هایی که تعداد اختصاص داده شده کمتر از کل سریال‌هاست
        count_assigned_to_tech__lt=F('total_serials')
    ).order_by('-serial_in', '-tarikh')

    seris = Seris.objects.all()
    if _role == 'smart':
        owners = Owner.objects.filter(
            Q(role__role='tek', active=True, zone_id=request.user.owner.zone_id) |
            Q(refrence__ename__in=['tek', 'ispolomb'], active=True, zone_id=request.user.owner.zone_id)
        ).distinct()
    else:
        owners = Owner.objects.filter(
            Q(role__role='engin', active=True, zone_id=request.user.owner.zone_id) |
            Q(refrence__ename__in='ispolomb', active=True, zone_id=request.user.owner.zone_id)
        ).distinct()

    context = {'seris': seris, 'owners': owners, 'list2': _list2}
    if request.method == "POST":
        tarikh = request.POST.get('select')
        tarikh = tarikh.split("/")
        tarikh = jdatetime.date(day=int(tarikh[2]), month=int(tarikh[1]), year=int(tarikh[0])).togregorian()
        package_id = int(request.POST.get('package', 0))
        seri_id = int(request.POST.get('seri', 0))
        serial_in = checknumber(checkxss(request.POST.get('serial_in', '')))
        serial_out = checknumber(checkxss(request.POST.get('serial_out', '')))
        owner_id = int(request.POST.get('owner', 0))
        if not all([package_id, seri_id, serial_in, serial_out, owner_id]):
            messages.error(request, 'لطفاً تمام فیلدهای ضروری را پر کنید')
            return redirect('lock:send_to_poshtiban_add')

        try:
            package = InsertLock.objects.get(id=package_id, zone_id=request.user.owner.zone_id)
            if package.updated_at.date() > tarikh:
                messages.error(request, 'تاریخ ارسال به تکنسین نباید قبل از تاریخ رسید باشد')
                return redirect('lock:send_to_poshtiban_add')
        except InsertLock.DoesNotExist:
            messages.error(request, 'بسته انتخابی معتبر نیست')
            return redirect('lock:send_to_poshtiban_add')

        # بررسی تطابق سری با بسته
        if package.seri_id != seri_id:
            messages.error(request, 'سری انتخاب شده با سری بسته مطابقت ندارد')
            return redirect('lock:send_to_poshtiban_add')

        # بررسی محدوده سریال‌ها
        if int(serial_in) < int(package.serial_in) or int(serial_out) > int(package.serial_out):
            messages.error(request, f"محدوده سریال‌ها باید بین {package.serial_in} تا {package.serial_out} باشد")
            return redirect('lock:send_to_poshtiban_add')

        _owner = Owner.objects.get(id=owner_id)
        tedad = (int(serial_out) - int(serial_in)) + 1

        if LockModel.objects.filter(serial_number=serial_in, seri=seri_id, zone_id=request.user.owner.zone.id,
                                    status_id=7).count() != 0:
            SendPoshtiban.objects.create(tarikh=tarikh, seri_id=seri_id, serial_in=serial_in, serial_out=serial_out,
                                         tedad=tedad, owner_id=int(owner_id), zone_id=request.user.owner.zone_id,
                                         user_id=request.user.id, ename=_role)

            add_to_log(request,
                       f'اضافه کردن  پلمب به تکنسین {_owner.name} {_owner.lname} از {serial_in} تا {serial_out} ', 0)
            messages.success(request, 'عملیات با موفقیت انجام شد')
        else:
            messages.error(request,
                           'عملیات شکست خورد باید ابتدا توسط سامانه منطقه تعریف گردد یا توسط پیمانکار رسید نشده است')
        return redirect('lock:send_to_poshtiban_list')
    return TemplateResponse(request, template_file, context)


@cache_permission('locktek')
def residposhtibanlist(request):
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    add_to_log(request, f'مشاهده فرم رسید پلمب توسط پیمانکار ', 0)

    # form_class = UserLoginForm
    template_file = 'residpeymankar.html'
    _list = InsertLock.objects.filter(zone_id=request.user.owner.zone_id, resid=False, peymankar__ename=_role).order_by(
        '-tarikh', '-serial_in')
    _list2 = InsertLock.objects.filter(zone_id=request.user.owner.zone_id, resid=True, peymankar__ename=_role).order_by(
        '-tarikh', '-serial_in').annotate(
        count_status_6=Count('lockmodel', filter=Q(lockmodel__status_id__in=[6, 9, 10, 11])),
        count_status_5=Count('lockmodel', filter=Q(lockmodel__status_id=5))
    )

    paginator = Paginator(_list2, 10)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count
    context = {'list2': _list, 'list': page_object, 'page_obj': page_obj, 'tedad': tedad,
               'query_string': query_string, }
    return TemplateResponse(request, template_file, context)


@cache_permission('locktek2')
def residteklist(request):
    add_to_log(request, f'مشاهده فرم رسید پلمب توسط تکنسین ', 0)

    # form_class = UserLoginForm
    template_file = 'residlocktek.html'
    _list = SendPoshtiban.objects.filter(zone_id=request.user.owner.zone_id, owner_id=request.user.owner.id,
                                         resid=False).order_by('-tarikh')
    _list2 = SendPoshtiban.objects.filter(zone_id=request.user.owner.zone_id, owner_id=request.user.owner.id,
                                          resid=True).order_by('-tarikh').annotate(
        count_status_6=Count('lockmodel', filter=Q(lockmodel__status_id__in=[6, 9, 10, 11])),
        count_status_5=Count('lockmodel', filter=Q(lockmodel__status_id=5))
    )
    _list3 = LockModel.objects.filter(status_id=4, owner_id=request.user.owner.id, sendposhtiban__resid=True)

    context = {'list': _list, 'list2': _list2, 'list3': _list3}
    return TemplateResponse(request, template_file, context)


def updateresid(request, _id, _st):
    url = request.META.get('HTTP_REFERER')

    if _st == 1:
        add_to_log(request, f' رسید پلمب توسط پیمانکار ', 0)
        result = InsertLock.objects.get(id=_id, zone_id=request.user.owner.zone.id)
        result.resid = True
        result.save()
        LockModel.objects.filter(insertlock_id=_id).update(status_id=7,
                                                           resid_date_unit=datetime.datetime.today(),
                                                           rdu_user_id=request.user.id)
        for item in LockModel.objects.filter(insertlock_id=_id):
            LockLogs.objects.create(
                status_id=7,
                owner_id=request.user.id,
                lockmodel_id=item.id,
                gs_id=item.gs_id,
                pump_id=item.pump_id)

    elif _st == 3:
        add_to_log(request, f' رسید پلمب توسط منطقه ', 0)
        result = GetLockPeymankar.objects.get(id=_id, zone_id=request.user.owner.zone.id)
        result.resid = True
        result.save()
        LockModel.objects.filter(getlockpeymankar_id=_id).update(status_id=11,
                                                                 input_date_poshtiban=datetime.datetime.today(),
                                                                 idp_user_id=request.user.id)
        for item in LockModel.objects.filter(getlockpeymankar_id=_id):
            LockLogs.objects.create(
                status_id=11,
                owner_id=request.user.id,
                lockmodel_id=item.id,
                gs_id=item.gs_id,
                pump_id=item.pump_id)
    else:
        add_to_log(request, f' رسید پلمب توسط تکنسین ', 0)
        result = SendPoshtiban.objects.get(id=_id, zone_id=request.user.owner.zone.id)
        result.resid = True
        result.save()

        LockModel.objects.filter(sendposhtiban_id=_id).update(status_id=8,
                                                              resid_date_tek=datetime.datetime.today(),
                                                              rdt_user_id=request.user.id)

        for item in LockModel.objects.filter(sendposhtiban_id=_id):
            LockLogs.objects.create(
                status_id=8,
                owner_id=request.user.id,
                lockmodel_id=item.id,

            )

    messages.success(request, 'عملیات با موفقیت انجام شد')
    return redirect(url)


def residnewserial(request, _id):
    url = request.META.get('HTTP_REFERER')
    add_to_log(request, f' رسید پلمب توسط تکنسین ', 0)
    result = LockModel.objects.get(id=_id)
    result.status_id = 8
    result.save()
    LockLogs.objects.create(
        status_id=8,
        owner_id=request.user.id,
        lockmodel_id=_id,
        info=f'رسید پلمب'
    )

    messages.success(request, 'عملیات با موفقیت انجام شد')
    return redirect(url)


@cache_permission('locktek')
def locktokargah(request):
    _role = 'smart'
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    sid = 0
    ownerid = 0
    locks = None
    stores = None
    posts = None
    if request.user.owner.role.role == 'engin':
        owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='engin')
    else:
        owners = Owner.objects.filter(
            Q(role__role='tek', zone_id=request.user.owner.zone_id) |
            Q(refrence__ename='tek', zone_id=request.user.owner.zone_id)
        ).distinct()
    if request.method == 'POST':
        add_to_log(request, f'مشاهده فرم دریافت داغی از تکنسین ', 0)
        ownerid = request.POST.get('ownerid')
        if not ownerid or ownerid == '-1':
            messages.error(request, 'لطفاً یک تکنسین را انتخاب کنید')
            return redirect('lock:locktokargah')
        # sid = request.POST.get('sid')

        user = User.objects.get(id=ownerid)
        locks = LockModel.objects.filter(status_id=6,
                                         zone_id=request.user.owner.zone_id, idg_user_id=user.id).order_by('-updated')
        if _role == 'engin':
            _stores = LockModel.objects.filter(status_id=9, zone_id=request.user.owner.zone_id, ename='engin')
        else:
            _stores = LockModel.objects.filter(status_id=9, zone_id=request.user.owner.zone_id, ename='smart')
        stores = []
        for store in _stores:
            try:
                user = User.objects.get(id=store.idg_user_id)
                stores.append({
                    'id': store.id,
                    'serial': store.serial,
                    'user': user.get_full_name()

                })
            except User.DoesNotExist:
                pass
    return TemplateResponse(request, 'locktokargah.html',
                            {'sid': int(sid), 'masters': locks, 'stores': stores, 'owners': owners,

                             'ownerid': int(ownerid), 'posts': posts})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def addlocktek(request):
    thislist = []
    if request.method == 'POST':

        userid = request.POST.get('id_tek')
        mylist = request.POST.get('strIds')
        val = request.POST.get('val')

        x = mylist.split(',')
        for item in x:
            gsinfo = LockModel.objects.get(id=item)
            gsinfo.status_id = 9
            gsinfo.input_date_tek = datetime.datetime.today()
            gsinfo.idt_user_id = request.user.id
            gsinfo.save()
            LockLogs.objects.create(
                status_id=9,
                owner_id=gsinfo.owner.user.id,
                lockmodel_id=gsinfo.id,
                info=f' دریافت داغی از تکنسین',
            )

            user = User.objects.get(id=gsinfo.idg_user_id)
            _user = user.get_full_name()

            thisdict = {
                "id": gsinfo.id,
                "serial": gsinfo.serial,
                "st": "",
                "user": _user,
            }
            thislist.append(thisdict)

    return JsonResponse({"message": "success", 'list': thislist})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def removelocktek(request):
    thislist = []
    val = 0
    if request.method == 'POST':
        mylist = request.POST.get('strIds')
        userid = request.POST.get('userid')
        val = request.POST.get('val')

        x = mylist.split(',')
        for item in x:
            gsinfo = LockModel.objects.get(id=item)
            gsinfo.status_id = 6
            gsinfo.input_date_tek = None
            gsinfo.idt_user_id = None
            gsinfo.save()
            LockLogs.objects.create(
                status_id=6,
                owner_id=gsinfo.owner.user.id,
                lockmodel_id=gsinfo.id,
                info=f'برگرداندن از دریافت داغی از تکنسین',
            )
            user = User.objects.get(id=gsinfo.idg_user_id)
            _user = user.get_full_name()
            thisdict = {
                "id": gsinfo.id,
                "serial": gsinfo.serial,
                "st": "",
                "user": _user,
            }
            thislist.append(thisdict)

    return JsonResponse({"message": "success", "val": val, 'list': thislist})


@cache_permission('locktek')
def getlockspeymankar(request):
    add_to_log(request, f'مشاهده فرم ایجاد بسته پلمپ داغی به منطقه', 0)

    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'

    url = request.META.get('HTTP_REFERER')

    if request.method == 'POST':
        try:
            selected_items = json.loads(request.POST.get('selected_items', '[]'))

            if not selected_items:
                messages.error(request, 'هیچ پلمبی انتخاب نشده است')
                return redirect(url)

            # فیلتر کردن فقط پلمب‌های موجود و دارای وضعیت مناسب
            locks = LockModel.objects.filter(
                serial__in=selected_items,
                status_id=9,
                zone_id=request.user.owner.zone_id,
                ename=_role
            )

            if locks.count() > 0:
                a = GetLockPeymankar.objects.create(
                    zone_id=request.user.owner.zone_id,
                    peymankar_id=1,
                    tedad=locks.count(),
                    user_id=request.user.id,
                    ename=_role
                )

                # به‌روزرسانی وضعیت فقط پلمب‌های انتخاب شده
                locks.update(
                    status_id=10,
                    getlockpeymankar_id=a.id
                )

                messages.success(request, f'بسته با {locks.count()} پلمب ایجاد شد')
            else:
                messages.error(request, 'هیچ پلمب معتبری برای ایجاد بسته یافت نشد')

        except Exception as e:
            messages.error(request, f'خطا در ایجاد بسته: {str(e)}')

    return redirect(url)


def sendtoexcel(request, _id, _st):
    add_to_log(request, f'ارسال به اکسل لیست پلمب 1 ', 0)
    my_path = 'daghiserial.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "لیست پلمپ های داغی"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:G1')
    ws1["A1"] = "لیست بسته پلمپ های داغی"
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:G2')
    ws1["A2"] = ''
    ws1["A2"].font = fonttitr

    ws1.merge_cells('A3:A3')
    ws1["A3"] = "ردیف"
    ws1["A3"].font = font

    i = 0
    ws1["B3"] = "شماره سریال"
    ws1["B3"].font = font
    ws1["C3"] = "GSID"
    ws1["C3"].font = font
    ws1["D3"] = "نام جایگاه"
    ws1["D3"].font = font
    ws1["E3"] = "تاریخ داغی"
    ws1["E3"].font = font
    ws1["F3"] = "تاریخ نصب"
    ws1["F3"].font = font
    ws1["G3"] = "نام تکنسین"
    ws1["G3"].font = font

    ws1.column_dimensions['A'].width = float(20.25)
    ws1.column_dimensions['B'].width = float(20.25)
    ws1.column_dimensions['C'].width = float(20.25)
    ws1.column_dimensions['D'].width = float(20.25)
    ws1.column_dimensions['E'].width = float(20.25)
    ws1.column_dimensions['F'].width = float(20.25)
    ws1.column_dimensions['G'].width = float(20.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')  # Background color
    i = 0

    for row in LockModel.objects.filter(getlockpeymankar_id=_id):
        i += 1
        gsid = row.gs.gsid if row.gs else ""
        name = row.gs.name if row.gs else ""

        d = [i, row.serial, gsid, name, row.fakdate(), row.installdate(),
             row.owner.name + " " + row.owner.lname]

        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    for cell in ws1["3:3"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border

    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''

    wb.save(response)

    return response


def sendtoexcel2(request, _date, _tedad, _id):
    add_to_log(request, f'ارسال به اکسل لیست پلمب 2 ', 0)
    _role = 'engin'

    if _id == 1:
        if request.user.owner.role.role in ['zone', 'tek']:
            _role = 'smart'
        elif request.user.owner.role.role in ['engin']:
            _role = 'engin'
        elif request.user.owner.refrence.ename == 'tek':
            _role = 'smart'
    else:
        if _id == 2:
            _role = 'smart'
    _daryaftdate = None
    _date = _date.split("-")

    _date = jdatetime.date(day=int(_date[2]), month=int(_date[1]), year=int(_date[0])).togregorian()
    _daryaft = jdatetime.datetime.fromgregorian(datetime=_date)
    _daryaft = _daryaft.strftime('%Y/%m/%d')
    gs = ''
    pump = ''
    position = ''
    daryaftdate = ''
    my_path = 'daghiserial.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "لیست پلمپ های داغی"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"

    ws1.merge_cells('A1:H1')
    ws1["A1"] = "دریافت و مصرف پلمب توسط پیمانکاران پشتیبان و تعمیرات"
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:E2')
    ws1["A2"] = f'تاریخ دریافت {_daryaft}'
    ws1["A2"].font = fonttitr
    ws1.merge_cells('F2:H2')
    ws1["F2"] = f'تعداد {_tedad}'
    ws1["F2"].font = fonttitr

    ws1.merge_cells('A3:A3')
    ws1["A3"] = "ردیف"
    ws1["A3"].font = font

    i = 0
    ws1["B3"] = "تاریخ"
    ws1["B3"].font = font
    ws1["C3"] = "شماره گزارش"
    ws1["C3"].font = font
    ws1["D3"] = "نام جایگاه"
    ws1["D3"].font = font
    ws1["E3"] = "شماره نازل"
    ws1["E3"].font = font
    ws1["F3"] = "نقطه پلمب"
    ws1["F3"].font = font
    ws1["G3"] = "پلمب نصب شده"
    ws1["G3"].font = font
    ws1["H3"] = "پلمب فک شده"
    ws1["H3"].font = font

    ws1.column_dimensions['A'].width = float(20.25)
    ws1.column_dimensions['B'].width = float(20.25)
    ws1.column_dimensions['C'].width = float(20.25)
    ws1.column_dimensions['D'].width = float(20.25)
    ws1.column_dimensions['E'].width = float(20.25)
    ws1.column_dimensions['F'].width = float(20.25)
    ws1.column_dimensions['G'].width = float(20.25)
    ws1.column_dimensions['H'].width = float(20.25)
    ws1.row_dimensions[1].height = 40
    ws1.row_dimensions[2].height = 50

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')  # Background color
    i = 0

    results = LockModel.objects.values('ticket2').filter(input_date_poshtiban=_date, ename=_role,
                                                         zone_id=request.user.owner.zone_id).annotate(
        count=Count('ticket_id'))

    _list = []

    for result in results:
        _locks = LockModel.objects.filter(ticket_id=result['ticket2'], send_date_gs__isnull=False,
                                          zone_id=request.user.owner.zone_id)
        serialin = ""
        gs = ''
        pump = ''
        position = ''
        daryaftdate = ''
        ir = 0
        for _lock in _locks:
            position = 'تلمبه' if _lock.pump else 'رک'
            if len(str(serialin)) > 1:
                serialin = str(serialin) + " " + str(_lock.serial)
            else:
                serialin = str(_lock.serial)

        _locks = LockModel.objects.filter(ticket2=result['ticket2'], input_date_gs__isnull=False,
                                          input_date_poshtiban=_date, ename=_role,
                                          zone_id=request.user.owner.zone_id)

        serialout = ""
        for _lock in _locks:
            pump = _lock.pump.number if _lock.pump else ""
            if ir == 0:
                gs = _lock.gs.name
                pump = pump

                ir += 1
            _daryaftdate = _lock.input_date_gs if _lock.input_date_gs else ""
            if len(str(serialout)) > 1:
                serialout = str(serialout) + " " + str(_lock.serial)
            else:
                serialout = str(_lock.serial)
            if len(str(_daryaftdate)) > 5:
                _daryaftdate = jdatetime.datetime.fromgregorian(datetime=_daryaftdate)
                _daryaftdate = _daryaftdate.strftime('%Y-%m-%d')
        _list.append({
            'tarikh': _daryaftdate,
            'ticket': result['ticket2'],
            'gs': gs,
            'pump': pump,
            'position': position,
            'serialin': serialin,
            'serialout': serialout,
        })
    for row in _list:
        i += 1
        serialin = row['serialin'].replace(" ", "\n")  # جایگزینی فاصله با خط جدید
        serialout = row['serialout'].replace(" ", "\n")  # جایگزینی فاصله با خط جدید
        d = [i, row['tarikh'], row['ticket'], row['gs'], row['pump'], row['position'], serialin,
             serialout]

        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center', wrap_text=True)
            cell.alignment = alignment_obj
            cell.border = thin_border

    for cell in ws1["3:3"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border

    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''

    wb.save(response)

    return response


def sendtoexcel4(request, _date, _tedad, _id):
    add_to_log(request, f'ارسال به اکسل لیست پلمب عملیات ', 0)
    _role = 'smart'

    if _id == 1:
        if request.user.owner.role.role in ['zone', 'tek']:
            _role = 'smart'
        elif request.user.owner.role.role in ['engin']:
            _role = 'engin'
        elif request.user.owner.refrence.ename == 'tek':
            _role = 'smart'
    else:
        if _id == 2:
            _role = 'smart'
    _daryaftdate = None
    _date = _date.split("-")

    _date = jdatetime.date(day=int(_date[2]), month=int(_date[1]), year=int(_date[0])).togregorian()
    _daryaft = jdatetime.datetime.fromgregorian(datetime=_date)
    _daryaft = _daryaft.strftime('%Y/%m/%d')
    gs = ''
    pump = ''
    position = ''
    daryaftdate = ''
    my_path = 'daghiserial.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "لیست پلمپ های داغی"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"

    # ws1.merge_cells('A3:A3')
    ws1["A1"] = "gsid"
    ws1["A1"].font = font
    ws1["B1"] = "شماره نازل"
    ws1["B1"].font = font
    ws1["C1"] = "پیشوند پلمب"
    ws1["C1"].font = font
    ws1["D1"] = "شماره پلمب"
    ws1["D1"].font = font
    ws1["E1"] = "تاریخ نصب"
    ws1["E1"].font = font
    ws1["F1"] = "صورتجلسه"
    ws1["F1"].font = font
    ws1["G1"] = "پیشوند پلمب داغی"
    ws1["G1"].font = font
    ws1["H1"] = "شماره پلمب داغی"
    ws1["H1"].font = font

    ws1.column_dimensions['A'].width = float(20.25)
    ws1.column_dimensions['B'].width = float(20.25)
    ws1.column_dimensions['C'].width = float(20.25)
    ws1.column_dimensions['D'].width = float(20.25)
    ws1.column_dimensions['E'].width = float(20.25)
    ws1.column_dimensions['F'].width = float(20.25)
    ws1.column_dimensions['G'].width = float(20.25)
    ws1.column_dimensions['H'].width = float(20.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')  # Background color
    i = 0

    results = LockModel.objects.filter(input_date_poshtiban=_date, ename=_role,
                                       zone_id=request.user.owner.zone_id)
    print(len(results))
    _list = []
    items_dict = {}

    for result in results:
        meeting_key = result.meeting_number

        # اگر این meeting_number قبلاً ثبت نشده، یک آیتم اولیه ایجاد کن
        if meeting_key not in items_dict:
            items_dict[meeting_key] = {
                'tarikh': '',
                'meeting_number': result.meeting_number,
                'gs': result.gs.gsid,
                'pump': '',
                'serialin': '',
                'serialout': ''
            }

        # بررسی لاگ نصب (status_id=5)
        logs_install = LockLogs.objects.filter(lockmodel__meeting_number=result.meeting_number, status_id=5).last()
        if logs_install:
            try:
                items_dict[meeting_key]['pump'] = logs_install.lockmodel.pump.number
            except:
                items_dict[meeting_key]['pump'] = ''
            items_dict[meeting_key]['tarikh'] = logs_install.lockmodel.send_date_gs
            items_dict[meeting_key]['gs'] = logs_install.lockmodel.gs.gsid
            items_dict[meeting_key]['serialin'] = logs_install.lockmodel.serial
        else:
            try:
                items_dict[meeting_key]['pump'] = result.pump.number
            except:
                items_dict[meeting_key]['pump'] = ''
            items_dict[meeting_key]['tarikh'] = result.send_date_gs
            items_dict[meeting_key]['gs'] = result.gs.gsid
            items_dict[meeting_key]['serialin'] = result.serial

        # بررسی لاگ داغی (status_id=6)
        logs_daghi = LockLogs.objects.filter(lockmodel__meeting_number=result.meeting_number, status_id=6).last()
        if logs_daghi:
            try:
                items_dict[meeting_key]['pump'] = logs_daghi.lockmodel.pump.number
            except:
                items_dict[meeting_key]['pump'] = ''

            items_dict[meeting_key]['gs'] = logs_daghi.lockmodel.gs.gsid
            items_dict[meeting_key]['serialout'] = logs_daghi.lockmodel.serial
        else:
            try:
                items_dict[meeting_key]['pump'] = result.pump.number
            except:
                items_dict[meeting_key]['pump'] = ''

            items_dict[meeting_key]['gs'] = result.gs.gsid
            items_dict[meeting_key]['serialout'] = result.serial

        # تبدیل دیکشنری به لیست
    _list = list(items_dict.values())
    print(_list)
    for row in _list:
        _x1, _x2 = separate_letters_numbers(row['serialin'])
        _y1, _y2 = separate_letters_numbers(row['serialout'])

        d = [row['gs'], row['pump'], _x1, _x2, row['tarikh'], row['meeting_number'],
             _y1, _y2]

        ws1.append(d)
    for col in ws1.columns:
        for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center', wrap_text=True)
            cell.alignment = alignment_obj
            cell.border = thin_border

    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''

    wb.save(response)

    return response


def separate_letters_numbers(text):
    """
    جدا کردن حروف و اعداد از یک رشته
    Returns: (حروف, اعداد)
    """
    # حذف فضاهای خالی و کاراکترهای غیرضرور
    text = str(text).strip()

    # پیدا کردن همه حروف (انگلیسی و فارسی)
    letters = ''.join(re.findall(r'[a-zA-Zآ-ی]', text))

    # پیدا کردن همه اعداد
    numbers = ''.join(re.findall(r'\d', text))

    return letters, numbers


@cache_permission('history')
def history(request):
    items = ''
    st_items = None
    idstore = None
    zoneid = None
    _list = None
    namestore = ''
    serial = ''
    ok = False
    if request.method == 'POST':
        serial = request.POST.get('search')
        serial = checkxss(serial)
        serial = checknumber(serial)
        add_to_log(request, f' جستجوی شماره پلمب {serial}', 0)

        try:
            ok = True
            idstore = LockModel.objects.get(serial=serial)
            namestore = idstore.status.info
            zoneid = idstore.zone_id
            _list = LockLogs.objects.filter(lockmodel_id=idstore.id).order_by('-id')
        except ObjectDoesNotExist:
            ok = False
            messages.error(request, 'این شماره سریال یافت نشد')
    return TemplateResponse(request, 'history.html',
                            {'list': _list, 'items': items, 'ok': ok, 'serial': serial,
                             'idstore': idstore, 'namestore': namestore, 'zoneid': zoneid, 'st_items': st_items})


@cache_permission('lockunit')
def changerange(request):
    template_file = 'changerange.html'
    seris = Seris.objects.all()
    status = Status.objects.all()
    positions = Position.objects.all()
    context = {'seris': seris, 'status': status, 'positions': positions}

    if request.method == "POST":

        seri = int(request.POST.get('seri'))
        # position = int(request.POST.get('position'))
        serial_in = checkxss(request.POST.get('serial_in'))
        serial_in = checknumber(serial_in)
        serial_out = checkxss(request.POST.get('serial_out'))
        serial_out = checknumber(serial_out)
        status = request.POST.get('status')
        tedad = (int(serial_out) - int(serial_in)) + 1

        serial_out = int(serial_out) + 1
        for item in range(int(serial_in), int(serial_out)):

            lenserial = len(str(serial_out))
            _len = lenserial - len(str(item))
            _zero = ''
            for i in range(_len):
                _zero = str(_zero) + '0'
            _serial = str(_zero) + str(item)
            lock = LockModel.objects.get(serial_number=_serial, seri_id=seri)
            lock.status_id = status
            lock.save()

        add_to_log(request, f'اضافه کردن  پلمب به واحد از {serial_in} تا {serial_out} ', 0)
        messages.success(request, 'عملیات با موفقیت انجام شد')
        return redirect('lock:input_lock_zone__list')
    return TemplateResponse(request, template_file, context)


@cache_permission('report')
def reportlock(request):
    add_to_log(request, f' مشاهده فرم گزارش سریال پلمب ها', 0)
    _role = 'smart'
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    template_file = 'reportlock.html'
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    datestatus = str(request.GET.get('status'))

    if len(datein) < 10:
        datein = "2020-10-10"
        dateout = "2100-12-20"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
    _list = LockModel.objects.filter(ename=_role)
    if request.user.owner.role.role == 'zone':
        _list = _list.filter(zone_id=request.user.owner.zone.id)
    if datestatus == 'None':
        _list = _list.filter(input_date_unit__range=(datein, dateout)).order_by(
            '-id')
    elif datestatus == "3":
        _list = _list.filter(input_date_unit__range=(datein, dateout)).order_by(
            '-id')
    elif datestatus == "4":
        _list = _list.filter(send_date_poshtiban__range=(datein, dateout)).order_by(
            '-id')
    elif datestatus == "5":
        _list = _list.filter(send_date_gs__range=(datein, dateout)).order_by('-id')
    elif datestatus == "6":
        _list = _list.filter(input_date_gs__range=(datein, dateout)).order_by('-id')
    elif datestatus == "9":
        _list = _list.filter(input_date_tek__range=(datein, dateout)).order_by(
            '-id')

    _filter = LockFilter(request.GET, queryset=_list, request=request)
    _list = _filter.qs
    if _filter.data:
        pass
    else:
        _list = LockModel.objects.none()
    context = {'filter': _filter, 'list': _list}
    return TemplateResponse(request, template_file, context)


@cache_permission('lockunit')
def listlocks(request):
    add_to_log(request, f' مشاهده فرم گزارش پلمب داغی های دریافت شده', 0)
    _role = 'smart'
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    _result = LockModel.objects.values('input_date_poshtiban', 'zone__name').filter(zone_id=request.user.owner.zone_id,
                                                                                    ename=_role,
                                                                                    input_date_poshtiban__isnull=False).annotate(
        count=Count('id'))
    context = {'result': _result}
    return TemplateResponse(request, 'list_locks.html', context)


@cache_permission('lockunit')
def report(request):
    add_to_log(request, f' مشاهده فرم گزارش پلمب ها', 0)
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    _list = []
    add_to_log(request, f'مشاهده فرم وضعیت پلمب تکنسین ', 0)

    if _role == 'smart':
        owner = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True)
    else:
        owner = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='engin', active=True)

    sum_locl_rock = 0
    sum_locl_pomp = 0
    sum_locl_daghi = 0
    salem = 0
    daghi = 0
    for item in owner:
        locl_pomp = LockModel.objects.filter(status_id=8, owner_id=item.id).count()
        locl_daghi = LockModel.objects.filter(status_id=6, idg_user_id=item.user_id).count()
        sum_locl_pomp += locl_pomp
        sum_locl_daghi += locl_daghi

        _dict = {
            'id': item.id,
            'name': item.name + ' ' + item.lname,
            'zone': item.zone.name,

            'locl_pomp': locl_pomp,
            'locl_daghi': locl_daghi,
        }
        _list.append(_dict)
    salem = LockModel.objects.filter(status_id=7, zone_id=request.user.owner.zone_id, ename=_role).count()
    daghi = LockModel.objects.filter(status_id__in=[9, 10], zone_id=request.user.owner.zone_id, ename=_role).count()
    _dict = {
        'id': 0,
        'name': 'پیمانکار',
        'zone': request.user.owner.zone.name,
        'locl_pomp': salem,
        'locl_daghi': daghi,
    }
    _list.append(_dict)
    sum_locl_pomp += salem
    sum_locl_daghi += daghi
    context = {'list': _list, 'sum_locl_pomp': sum_locl_pomp,
               'sum_locl_daghi': sum_locl_daghi, }
    return TemplateResponse(request, 'report.html', context)


@cache_permission('locktek')
def changetek(request):
    sid = 0
    ownerid = 0
    ownerid2 = 0
    locks = None
    stores = None
    posts = None
    if request.user.owner.role.role == 'engin':
        owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='engin')
    else:
        owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
    if request.method == 'POST':
        add_to_log(request, f' مشاهده فرم جابجایی  پلمب ها', 0)
        # sid = request.POST.get('sid')
        ownerid = request.POST.get('ownerid')
        ownerid2 = request.POST.get('ownerid2')

        locks = LockModel.objects.filter(status_id=8,
                                         zone_id=request.user.owner.zone_id, owner_id=ownerid).order_by('-updated')
        _stores = LockModel.objects.filter(status_id=8, zone_id=request.user.owner.zone_id, owner_id=ownerid2)
        stores = []
        for store in _stores:
            stores.append({
                'id': store.owner_id,
                'serial': store.serial,

            })
    return TemplateResponse(request, 'changetek.html',
                            {'sid': int(sid), 'masters': locks, 'stores': stores, 'owners': owners,
                             'ownerid2': int(ownerid2),

                             'ownerid': int(ownerid), 'posts': posts})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def addlocktek2(request):
    thislist = []
    if request.method == 'POST':

        userid = request.POST.get('id_tek')
        userid2 = request.POST.get('id_tek2')
        mylist = request.POST.get('strIds')
        val = request.POST.get('val')

        x = mylist.split(',')
        for item in x:
            gsinfo = LockModel.objects.get(id=item)
            _owner1 = gsinfo.owner.get_full_name()
            gsinfo.status_id = 4
            gsinfo.input_date_tek = datetime.datetime.today()
            gsinfo.idt_user_id = request.user.id
            gsinfo.owner_id = userid2
            _user = gsinfo.owner.lname
            gsinfo.save()
            _owner2 = gsinfo.owner.get_full_name()
            LockLogs.objects.create(
                status_id=4,
                owner_id=gsinfo.owner.user.id,
                lockmodel_id=gsinfo.id,
                info=f'جابجای پلمب از  تکنسین {_owner1} به {_owner2}',
            )

            thisdict = {
                "id": gsinfo.id,
                "serial": gsinfo.serial,
                "st": "",
                "user": _user,
            }
            thislist.append(thisdict)

    return JsonResponse({"message": "success", 'list': thislist})


@cache_permission('history')
def edit_serial(request, serial):
    try:
        lock = LockModel.objects.get(serial=serial)
        add_to_log(request, f' ویرایش سریال پلمب  {serial}', 0)
        statuses = Status.objects.all()

        if request.method == 'POST':
            form_data = request.POST
            new_serial = form_data.get('serial')
            status_id = form_data.get('status')

            # اعتبارسنجی داده‌ها
            if new_serial and new_serial != lock.serial:
                if LockModel.objects.filter(serial=new_serial).exists():
                    messages.error(request, 'این شماره سریال قبلاً ثبت شده است')
                    return redirect('lock:edit_serial', serial=serial)
                lock.serial = new_serial

            if status_id:
                lock.status_id = status_id

            lock.save()
            messages.success(request, 'تغییرات با موفقیت ذخیره شد')
            return redirect('lock:history')

        context = {
            'lock': lock,
            'statuses': statuses,
            'serial': serial
        }
        return TemplateResponse(request, 'edit_serial.html', context)

    except LockModel.DoesNotExist:
        messages.error(request, 'سریال مورد نظر یافت نشد')
        return redirect('lock:history')


@cache_permission('reportzonelock')
def zone_locks_report(request):
    _statusreport = request.GET.get('statusreport')
    if not _statusreport:
        _statusreport = '1'

    add_to_log(request, f' مشاهده فرم عملکرد  پلمب مناطق', 0)
    _role = 'smart'
    if request.user.owner.role.role == 'engin':
        _role = 'engin'
    else:
        _role = 'smart'
    # دریافت تمام مناطق
    if request.user.owner.role.role in ['zone', 'tek', 'area']:
        zones = Zone.objects.filter(id=request.user.owner.zone.id)
    else:
        zones = Zone.objects_limit.all()

    # دریافت تمام وضعیت‌های ممکن
    statuses = Status.objects.all()

    # ایجاد لیستی برای ذخیره اطلاعات هر منطقه
    zone_data = []

    for zone in zones:
        _countall = InsertLock.objects.filter(zone=zone, peymankar__ename=_role).aggregate(tedad=Sum('tedad'))
        _count = LockModel.objects.filter(zone=zone, input_date_unit__gt='2023-01-01', ename=_role,
                                          insertlock_id__isnull=False).count()
        zone_info = {
            'zone_id': zone.id,
            'zone_name': zone.name,
            'polombs': _count,
            'allsendtopeymankar': _countall['tedad'],
            'status_counts': {}
        }

        # شمارش تعداد پلمپ‌ها برای هر وضعیت در این منطقه
        for status in statuses:
            if _statusreport == "1":
                count = LockModel.objects.filter(zone=zone, status=status, ename=_role).count()
            elif _statusreport == "2":
                count = LockModel.objects.filter(zone=zone, input_date_unit__gt='2020-01-01', status=status,
                                                 ename=_role).count()
            zone_info['status_counts'][status.id] = {
                'status_name': status.info,
                'count': count
            }
        zone_data.append(zone_info)
    context = {
        'zone_data': zone_data,
        'statuses': statuses,
        'status': _statusreport
    }

    return TemplateResponse(request, 'zone_locks_report.html', context)


@cache_permission('locktek')
def senddaghitozone(request):
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    add_to_log(request, f'مشاهده فرم ارسالی داغی به منطقه ', 0)

    # form_class = UserLoginForm
    template_file = 'senddaghlist.html'
    _list = GetLockPeymankar.objects.filter(zone_id=request.user.owner.zone_id, ename=_role).order_by('-created_at')
    context = {'list': _list}
    return TemplateResponse(request, template_file, context)


def sendtoexcel3(request, _id):
    add_to_log(request, f' ارسال به اکسل پلمب 3', 0)
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'
    _daryaftdate = None
    my_path = 'daghiserial.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)

    wb = Workbook()
    results = LockModel.objects.values('ticket2').filter(getlockpeymankar_id=_id, ename=_role,
                                                         zone_id=request.user.owner.zone_id).annotate(
        count=Count('ticket_id'))
    ws1 = wb.active
    ws1.title = "لیست پلمپ های داغی"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"

    ws1.merge_cells('A1:H1')
    ws1["A1"] = "دریافت و مصرف پلمب توسط پیمانکاران پشتیبان و تعمیرات"
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:E2')
    ws1["A2"] = f'تاریخ دریافت '
    ws1["A2"].font = fonttitr
    ws1.merge_cells('F2:H2')
    ws1["F2"] = f'تعداد {results.count()}'
    ws1["F2"].font = fonttitr

    ws1.merge_cells('A3:A3')
    ws1["A3"] = "ردیف"
    ws1["A3"].font = font

    i = 0
    ws1["B3"] = "تاریخ"
    ws1["B3"].font = font
    ws1["C3"] = "شماره گزارش"
    ws1["C3"].font = font
    ws1["D3"] = "نام جایگاه"
    ws1["D3"].font = font
    ws1["E3"] = "شماره نازل"
    ws1["E3"].font = font
    ws1["F3"] = "نقطه پلمب"
    ws1["F3"].font = font
    ws1["G3"] = "پلمب نصب شده"
    ws1["G3"].font = font
    ws1["H3"] = "پلمب فک شده"
    ws1["H3"].font = font

    ws1.column_dimensions['A'].width = float(20.25)
    ws1.column_dimensions['B'].width = float(20.25)
    ws1.column_dimensions['C'].width = float(20.25)
    ws1.column_dimensions['D'].width = float(20.25)
    ws1.column_dimensions['E'].width = float(20.25)
    ws1.column_dimensions['F'].width = float(20.25)
    ws1.column_dimensions['G'].width = float(20.25)
    ws1.column_dimensions['H'].width = float(20.25)
    ws1.row_dimensions[1].height = 40
    ws1.row_dimensions[2].height = 50

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')  # Background color
    i = 0

    _list = []

    for result in results:
        _locks = LockModel.objects.filter(ticket_id=result['ticket2'], send_date_gs__isnull=False,
                                          zone_id=request.user.owner.zone_id)
        serialin = ""
        gs = ''
        pump = ''
        position = ''
        daryaftdate = ''
        ir = 0
        for _lock in _locks:
            position = 'تلمبه' if _lock.pump else 'رک'

            if len(str(serialin)) > 1:
                serialin = str(serialin) + " " + str(_lock.serial)
            else:
                serialin = str(_lock.serial)

        _locks = LockModel.objects.filter(ticket2=result['ticket2'], input_date_gs__isnull=False,
                                          getlockpeymankar_id=_id,
                                          zone_id=request.user.owner.zone_id)

        serialout = ""
        for _lock in _locks:
            pump = _lock.pump.number if _lock.pump else ""
            if ir == 0:
                gs = _lock.gs.name
                pump = pump

                ir += 1
            _daryaftdate = _lock.input_date_gs if _lock.input_date_gs else ""
            if len(str(serialout)) > 1:
                serialout = str(serialout) + " " + str(_lock.serial)
            else:
                serialout = str(_lock.serial)
            if len(str(_daryaftdate)) > 5:
                _daryaftdate = jdatetime.datetime.fromgregorian(datetime=_daryaftdate)
                _daryaftdate = _daryaftdate.strftime('%Y-%m-%d')
        _list.append({
            'tarikh': _daryaftdate,
            'ticket': result['ticket2'],
            'gs': gs,
            'pump': pump,
            'position': position,
            'serialin': serialin,
            'serialout': serialout,
        })
    for row in _list:
        i += 1
        serialin = row['serialin'].replace(" ", "\n")  # جایگزینی فاصله با خط جدید
        serialout = row['serialout'].replace(" ", "\n")  # جایگزینی فاصله با خط جدید
        d = [i, row['tarikh'], row['ticket'], row['gs'], row['pump'], row['position'], serialin,
             serialout]

        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center', wrap_text=True)
            cell.alignment = alignment_obj
            cell.border = thin_border

    for cell in ws1["3:3"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border

    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''

    wb.save(response)

    return response


@cache_permission('bulk_chstore')
@transaction.atomic
def import_excel_changest(request):
    form = open_excel(request.POST)
    zone = Zone.objects.all()
    statusref = Status.objects.all()
    if request.method == 'POST':
        add_to_log(request, f' دریافت پلمب با اکسل', 0)
        _statusref = request.POST.get('statusref')
        _info = request.POST.get('info')
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                master = sheet_obj.cell(row=i, column=1).value

                master = checknumber(str(master))
                if master:

                    try:
                        _list = LockModel.objects.get(serial=master)
                        _newid = _list.id
                        _list.status_id = _statusref

                        _list.save()
                        LockLogs.objects.create(
                            status_id=_statusref,
                            owner_id=request.user.id,
                            lockmodel_id=_newid,
                            info="تغییر وضعیت با فایل اکسل  "
                        )

                    except ObjectDoesNotExist:
                        a = 1

            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        return redirect('base:home')

    return TemplateResponse(request, 'import_excel_changest.html', {'form': form, 'statusref': statusref, 'zone': zone})


def check_serial_polob(request):
    add_to_log(request, f' بررسی بسته  پلمب ها', 0)
    # دریافت و اعتبارسنجی پارامترهای ورودی
    val1 = request.GET.get('val1')
    val2 = request.GET.get('val2')
    seri = request.GET.get('seri')

    if not all([val1, val2, seri]):
        return JsonResponse({'error': 'پارامترهای ورودی ناقص هستند'}, status=400)

    try:
        # دریافت پیشوند سری
        serial_prefix = Seris.objects.filter(id=seri).values_list('info', flat=True).first()
        if not serial_prefix:
            return JsonResponse({'error': 'سری مورد نظر یافت نشد'}, status=404)

        # تبدیل به عدد و اعتبارسنجی
        try:
            start_num = int(val1)
            end_num = int(val2)
        except ValueError:
            return JsonResponse({'error': 'مقادیر سریال باید عددی باشند'}, status=400)

        # ایجاد الگوی سریال‌ها
        serial_pattern = f"{serial_prefix}%"

        # جستجوی سریال‌های موجود در محدوده با یک کوئری
        existing_serial = LockModel.objects.filter(
            serial__startswith=serial_prefix,
            serial__regex=r'^{}\d+$'.format(serial_prefix),
            serial__gte=f"{serial_prefix}{start_num}",
            serial__lte=f"{serial_prefix}{end_num}"
        ).order_by('serial').first()

        if existing_serial:
            return JsonResponse({
                'result': f'شماره سریال {existing_serial.serial} قبلا در سیستم ثبت شده است',
                'duplicate': existing_serial.serial
            })

        return JsonResponse({'result': None})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@cache_permission('package')
@transaction.atomic
def add_serial_to_package(request):
    if request.method == 'POST':
        package_id = request.POST.get('package_id')
        serial_number = request.POST.get('serial_number')
        status_id = request.POST.get('status', '10')  # پیش‌فرض وضعیت 10 (ارسال به منطقه)
        add_to_log(request, f' تزریق پلمب به پکیج {package_id} - {serial_number}', 0)
        try:
            package = GetLockPeymankar.objects.get(id=package_id, zone_id=request.user.owner.zone_id)

            # بررسی وجود سریال در سیستم
            try:
                lock = LockModel.objects.get(serial=serial_number)
                # اگر سریال وجود داشت، وضعیت و بسته آن را به‌روزرسانی کنید

                lock.status_id = status_id
                lock.getlockpeymankar_id = package_id
                lock.input_date_poshtiban = datetime.datetime.today()
                lock.idp_user_id = request.user.id
                lock.save()

                # ایجاد لاگ
                LockLogs.objects.create(
                    status_id=status_id,
                    owner_id=request.user.id,
                    lockmodel_id=lock.id,
                    info=f'افزودن دستی به بسته {package_id}'
                )

                messages.success(request, 'سریال موجود با موفقیت به بسته اضافه شد')
            except LockModel.DoesNotExist:
                # اگر سریال وجود نداشت، یک رکورد جدید ایجاد کنید
                # ابتدا پیشوند سریال را از بسته بگیرید

                # ایجاد رکورد جدید
                new_lock = LockModel.objects.create(
                    owner_id=request.user.owner.id,
                    serial=serial_number,
                    zone_id=package.zone_id,
                    status_id=status_id,
                    getlockpeymankar_id=package_id,
                    input_date_poshtiban=datetime.datetime.today(),
                    idp_user_id=request.user.id,
                    ename=package.ename,
                    # سایر فیلدهای مورد نیاز
                )

                # ایجاد لاگ
                LockLogs.objects.create(
                    status_id=status_id,
                    owner_id=request.user.id,
                    lockmodel_id=new_lock.id,
                    info=f'سریال جدید به بسته {package_id} اضافه شد'
                )

                # افزایش تعداد بسته
                package.tedad = LockModel.objects.filter(getlockpeymankar_id=package_id).count()
                package.save()

                messages.success(request, 'سریال جدید با موفقیت ایجاد و به بسته اضافه شد')

            return redirect('lock:senddaghitozone')

        except GetLockPeymankar.DoesNotExist:
            messages.error(request, 'بسته مورد نظر یافت نشد')
            return redirect('lock:senddaghitozone')

    return redirect('lock:senddaghitozone')


@cache_permission('lockunit')
def filtered_list_locks(request):
    add_to_log(request, 'مشاهده فرم گزارش پلمب داغی های دریافت شده با فیلتر', 0)

    # دریافت پارامترهای فیلتر
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    system_type = request.GET.get('system_type', 'smart')  # پیش‌فرض سامانه

    # تبدیل تاریخ‌ها به فرمت گرگوری
    try:
        if start_date:
            start_date = start_date.split("/")
            start_date = jdatetime.date(day=int(start_date[2]), month=int(start_date[1]),
                                        year=int(start_date[0])).togregorian()

        if end_date:
            end_date = end_date.split("/")
            end_date = jdatetime.date(day=int(end_date[2]), month=int(end_date[1]), year=int(end_date[0])).togregorian()
    except:
        messages.error(request, 'فرمت تاریخ نامعتبر است')
        return redirect('lock:filtered_list_locks')

    # ایجاد کوئری پایه
    queryset = LockModel.objects.filter(
        zone_id=request.user.owner.zone_id,
        ename=system_type,
        input_date_poshtiban__isnull=False
    )

    # اعمال فیلتر تاریخ
    if start_date and end_date:
        queryset = queryset.filter(input_date_poshtiban__range=(start_date, end_date))
    elif start_date:
        queryset = queryset.filter(input_date_poshtiban__gte=start_date)
    elif end_date:
        queryset = queryset.filter(input_date_poshtiban__lte=end_date)

    # گروه‌بندی نتایج
    result = queryset.values('input_date_poshtiban', 'zone__name').annotate(count=Count('id'))

    context = {
        'result': result,
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'system_type': system_type
    }

    return TemplateResponse(request, 'filtered_list_locks.html', context)


def updateserialpolomb(request, _id):
    _role = 'smart'

    result = InsertLock.objects.filter(zone_id=_id, peymankar__ename=_role)
    _diclist = []
    for item in result:
        for i in range(item.tedad):
            _diclist.append(str(item.seri.info) + str(item.serial_in + i))
    _list = LockModel.objects.filter(zone_id=_id, input_date_unit__gt='2023-01-01', ename=_role)
    missing_serials = [serial for serial in _diclist if not _list.filter(serial=serial).exists()]
    lists = []
    for serial in missing_serials:

        numbers_int = int(re.findall(r'\d+', serial)[0])
        letters = ''.join(re.findall(r'[a-zA-Z]+', serial))

        insertlock = InsertLock.objects.filter(
            zone_id=_id,
            serial_in__lte=numbers_int,
            serial_out__gte=numbers_int
        ).first()

        try:
            _ = LockModel.objects.get(serial=serial)
            oldserial = LockModel.objects.filter(insertlock_id=insertlock.id,
                                                 ).first()
            if oldserial.input_date_unit:
                _.input_date_unit = oldserial.input_date_unit
            if oldserial.idu_user_id:
                _.idu_user_id = oldserial.idu_user_id
            if oldserial.send_date_poshtiban:
                _.send_date_poshtiban = oldserial.send_date_poshtiban
            if oldserial.sdp_user_id:
                _.sdp_user_id = oldserial.sdp_user_id
            if oldserial.send_date_tek:
                _.send_date_tek = oldserial.send_date_tek
            if oldserial.sdt_user_id:
                _.sdt_user_id = oldserial.sdt_user_id
            if oldserial.sendposhtiban_id:
                _.sendposhtiban_id = oldserial.sendposhtiban_id
            if oldserial.insertlock_id:
                _.insertlock_id = oldserial.insertlock_id
            _.save()
            lists.append({
                'serial': str(serial),
                'number': str(numbers_int),
                'error': 'ok'
            })

        except Exception as e:
            lists.append({
                'serial': str(serial),
                'number': str(numbers_int),
                'error': str(e)
            })
            continue

    return JsonResponse({'message': lists})


def import_excel_ckpolomb(request):
    add_to_log(request, 'دریافت اکسل پلمب', 0)
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                serial = str(sheet_obj.cell(row=i, column=1).value)
                _zoneid = str(sheet_obj.cell(row=i, column=2).value)
                status_id = str(sheet_obj.cell(row=i, column=3).value)
                numbers_int = int(re.findall(r'\d+', serial)[0])
                letters = ''.join(re.findall(r'[a-zA-Z]+', serial))

                insertlock = InsertLock.objects.filter(
                    zone_id=_zoneid,
                    serial_in__lte=numbers_int,
                    serial_out__gte=numbers_int
                ).first()
                if insertlock:
                    try:
                        oldserial = LockModel.objects.filter(insertlock_id=insertlock.id,
                                                             status_id__gte=status_id).first()
                        _ = LockModel.objects.create(serial=serial,
                                                     input_date_unit=oldserial.input_date_unit,
                                                     idu_user_id=oldserial.idu_user_id,
                                                     send_date_poshtiban=oldserial.send_date_poshtiban,
                                                     sdp_user_id=oldserial.sdp_user_id,
                                                     send_date_tek=oldserial.send_date_tek,
                                                     sdt_user_id=oldserial.sdt_user_id,
                                                     insertlock_id=oldserial.insertlock_id,
                                                     sendposhtiban_id=oldserial.sendposhtiban_id,
                                                     zone_id=oldserial.zone_id,
                                                     status_id=status_id,
                                                     ename='smart'
                                                     )

                    except LockModel.DoesNotExist:
                        pass

    return TemplateResponse(request, 'importexcel.html', {'form': form})


def export_lock_list_to_excel(request):
    add_to_log(request, 'ارسال آمار خرابی به اکسل  ', 0)
    my_path = "ipclog.xlsx"
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)
    fonttitr2 = Font(bold=True, size=20)
    wb = Workbook()

    ws1 = wb.active  # work with default worksheet
    ws1.title = "گزارش تخصیص پلمب های سالم و پلمب های داغی بر اساس صورتجلسه های پلمب ( فایل نظام)"

    ws1.sheet_view.rightToLeft = True
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:R1')

    ws1["A1"] = "گزارش تخصیص پلمب های سالم و پلمب های داغی بر اساس صورتجلسه های پلمب "
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:A3')
    ws1["A2"] = "ردیف"
    ws1["A2"].font = font

    ws1.merge_cells('B2:E2')
    ws1["B2"] = "مشخصات پارت ورودی پلمب های سالم"
    ws1["B2"].font = fonttitr2

    ws1.merge_cells('F2:P2')
    ws1["F2"] = "اطلاعات مربوط به صورتجلسه پلمب  "
    ws1["F2"].font = font

    ws1.merge_cells('Q2:Q3')
    ws1["Q2"] = " توضیحات"
    ws1["Q2"].font = font

    ws1.merge_cells('R2:R3')
    ws1["R2"] = "شماره فایل یا نامه امحا"
    ws1["R2"].font = font

    ws1.merge_cells('B3:B3')
    ws1["B3"] = "شماره پارت دریافت  شده از سامانه"
    ws1["B3"].font = font

    ws1.merge_cells('C3:C3')
    ws1["C3"] = "تاریخ دریافت"
    ws1["C3"].font = font

    ws1.merge_cells('D3:D3')
    ws1["D3"] = "سریال آغازی پارت"
    ws1["D3"].font = font

    ws1.merge_cells('E3:E3')
    ws1["E3"] = "سریال پایانی پارت"
    ws1["E3"].font = font

    ws1.merge_cells('F3:F3')
    ws1["F3"] = "تاریخ تحویل به تکنسین"
    ws1["F3"].font = font

    ws1.merge_cells('G3:G3')
    ws1["G3"] = " نام و نام خانوادگی تکنسین "
    ws1["G3"].font = font

    ws1.merge_cells('H3:H3')
    ws1["H3"] = "شماره پلمب نصب شده "
    ws1["H3"].font = font

    ws1.merge_cells('I3:I3')
    ws1["I3"] = "تاریخ نصب پلمب"
    ws1["I3"].font = font

    ws1.merge_cells('J3:J3')
    ws1["J3"] = "شماره صورتجلسه"
    ws1["J3"].font = font

    ws1.merge_cells('K3:K3')
    ws1["K3"] = "GSID"
    ws1["K3"].font = font

    ws1.merge_cells('L3:L3')
    ws1["L3"] = "نام جایگاه"
    ws1["L3"].font = font
    ws1.merge_cells('M3:M3')
    ws1["M3"] = "ناحیه"
    ws1["M3"].font = font

    ws1.merge_cells('N3:N3')
    ws1["N3"] = "شماره نازل"
    ws1["N3"].font = font

    ws1.merge_cells('O3:O3')
    ws1["O3"] = "نقطه پلمب شده"
    ws1["O3"].font = font

    ws1.merge_cells('P3:P3')
    ws1["P3"] = "شماره پلمب فک شده"
    ws1["P3"].font = font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(15.25)
    ws1.column_dimensions['D'].width = float(25.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='ffd700')  # Background color
    my_fill2 = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    i = 0
    _locks = LockModel.objects.filter(status__gte=5)
    for item in _locks:
        try:
            i += 1
            d = [i, "", str(item.resid_date_unit),
                 str(item.sendposhtiban.seri) + str(item.sendposhtiban.serial_in),
                 str(item.sendposhtiban.seri) + str(item.sendposhtiban.serial_out),
                 str(item.resid_date_tek),
                 str(item.sendposhtiban.owner),
                 str(item.serial),
                 str(item.send_date_gs),
                 str(item.meeting_number),
                 str(item.gs.gsid),
                 str(item.gs.name),
                 str(item.gs.area.name),
                 str(item.pump.number), "",
                 ""

                 ]

            ws1.append(d)
        except Exception as e:
            pass

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    i += 3
    for cell in ws1[i:i]:  # Last row
        cell.font = myfont
        cell.fill = my_fill2
        cell.border = thin_border

    for cell in ws1["2:2"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border
    wb.save(response)
    return response


def change_gsid(request, oldgsid, newgsid):
    oldid = GsModel.objects.get(gsid=oldgsid).id
    newid = GsModel.objects.get(gsid=newgsid).id
    LockModel.objects.filter(gs_id=oldid).update(gs_id=newid)

    return HttpResponse('ok')


@cache_permission('locktek2')
@transaction.atomic
def install_lock(request, code):
    if code == 'add':
        template_file = 'install_lock.html'
    elif code == 'remove':
        template_file = 'remove_lock.html'
    elif code == 'soratjalase':
        template_file = 'soratjalase.html'
    _role = 'smart'
    if request.user.owner.role.role in ['zone', 'tek']:
        _role = 'smart'
    elif request.user.owner.role.role in ['engin']:
        _role = 'engin'
    elif request.user.owner.refrence.ename == 'tek':
        _role = 'smart'

    if request.method == 'POST':
        form = InstallLockForm(request.POST, user=request.user)

        # تنظیم کوئری‌ست جایگاه‌ها برای فرم POST
        form.fields['gs'].queryset = GsModel.object_role.c_gsmodel(request)
        if form.is_valid():
            try:
                gs = form.cleaned_data['gs']
                lock_type = form.cleaned_data['lock_type']
                date = request.POST.get('date')
                serial = form.cleaned_data['serial']
                meeting_number = form.cleaned_data['meeting_number']
                description = form.cleaned_data.get('description', '')
                pump_id = request.POST.get('pump')
                date = to_miladi(date)
                if not pump_id:
                    messages.error(request, 'لطفاً شماره نازل را انتخاب کنید.')
                    context = {'form': form}
                    return TemplateResponse(request, template_file, context)

                try:
                    pump = Pump.objects.get(id=pump_id)
                except Pump.DoesNotExist:
                    messages.error(request, 'نازل انتخاب شده معتبر نیست.')
                    context = {'form': form}
                    return TemplateResponse(request, template_file, context)

                if code == 'add':
                    # پلمپ نصبی - فقط آپدیت
                    if isinstance(serial, LockModel):
                        # آپدیت پلمپ موجود
                        serial.status_id = 5  # نصب شده
                        serial.send_date_gs = date
                        serial.sdg_user = request.user
                        serial.gs = gs
                        serial.pump = pump
                        serial.meeting_number = meeting_number
                        serial.description = description

                        serial.sdg_user_id = request.user.id
                        serial.ename = _role
                        serial.save()

                        LockLogs.objects.create(
                            lockmodel=serial,
                            owner=request.user,
                            status=serial.status,
                            gs=gs,
                            pump=pump,
                            info=f'ثبت نصبی - {description}'
                        )

                        add_to_log(request, f'ثبت پلمپ نصبی برای جایگاه {gs.name}', 0)
                        messages.success(request, 'پلمپ نصبی با موفقیت ثبت شد')

                elif code == 'remove':  # داغی

                    # پلمپ داغی - چک کردن وجود و آپدیت یا ایجاد
                    try:
                        # اگر پلمپ وجود دارد، آپدیت کن
                        lock_model = LockModel.objects.get(serial=serial)
                        lock_model.status_id = 6  # داغی
                        lock_model.input_date_gs = date
                        lock_model.idg_user = request.user
                        lock_model.meeting_number = meeting_number
                        lock_model.description = description
                        lock_model.gs = gs
                        lock_model.pump = pump
                        lock_model.ename = _role
                        lock_model.save()

                        action = 'آپدیت'
                    except LockModel.DoesNotExist:
                        # اگر پلمپ وجود ندارد، ایجاد کن
                        lock_model = LockModel.objects.create(
                            serial=serial,
                            status_id=6,  # داغی
                            input_date_gs=date,
                            idg_user=request.user,
                            meeting_number=meeting_number,
                            description=description,
                            gs=gs,
                            pump=pump,
                            zone=gs.area.zone,
                            owner=request.user.owner,
                            ename=_role  # یا بر اساس منطق شما
                        )
                        action = 'ایجاد'

                    LockLogs.objects.create(
                        lockmodel=lock_model,
                        owner=request.user,
                        status=lock_model.status,
                        gs=gs,
                        pump=pump,
                        info=f'{action} داغی - {description}'
                    )

                    add_to_log(request, f'{action} پلمپ داغی برای جایگاه {gs.name}', 0)
                    messages.success(request, f'پلمپ داغی با موفقیت {action} شد')

                return redirect('lock:install_lock', code)



            except Exception as e:
                transaction.set_rollback(True)
                messages.error(request, f'خطا در ثبت پلمپ: {str(e)}')
    else:
        form = InstallLockForm(user=request.user)
        form.fields['gs'].queryset = GsModel.object_role.c_gsmodel(request)

    context = {'form': form}
    return TemplateResponse(request, template_file, context)


# ویو برای دریافت نازل‌های جایگاه
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_pumps_by_gs(request):
    gs_id = request.GET.get('gs_id')
    if not gs_id:
        return JsonResponse({'error': 'شناسه جایگاه ارسال نشده است'}, status=400)

    try:
        pumps = Pump.objects.filter(
            gs_id=gs_id,
            active=True
        ).values('id', 'number', 'product__name')

        pumps_list = list(pumps)
        return JsonResponse({'pumps': pumps_list})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@cache_permission('locktek2')
@transaction.atomic
def upload_soratjalase(request):
    """
    ویو برای آپلود صورتجلسه وقتی تکنسین می‌بیند قطعه داغی وجود ندارد
    و باید صورتجلسه آپلود کند
    """
    template_file = 'soratjalase.html'

    def generate_unique_serial():
        """تولید شماره سریال یکتا با پیشوند SJ و 6 رقم عددی"""
        while True:
            # تولید 6 رقم تصادفی
            numbers = ''.join(random.choices(string.digits, k=6))
            serial = f"SJ{numbers}"

            # بررسی یکتایی
            if not LockModel.objects.filter(serial=serial).exists():
                return serial

    if request.method == 'POST':
        form = SoratJalaseForm(request.POST, request.FILES, user=request.user)
        form.fields['gs'].queryset = GsModel.object_role.c_gsmodel(request)

        if form.is_valid():
            try:
                # تولید شماره سریال یکتا
                serial_number = generate_unique_serial()

                # ذخیره اطلاعات
                lock_model = LockModel.objects.create(
                    serial=serial_number,
                    meeting_number=form.cleaned_data['meeting_number'],
                    gs=form.cleaned_data['gs'],
                    pump_id=request.POST.get('pump'),
                    input_date_gs=form.cleaned_data['date'],  # تاریخ صورتجلسه
                    idg_user=request.user,
                    soratjalase=form.cleaned_data['soratjalase'],
                    description=form.cleaned_data.get('description', ''),
                    status_id=6,
                    zone=form.cleaned_data['gs'].area.zone,
                    owner=request.user.owner,
                    ename='smart'  # یا بر اساس نقش کاربر
                )

                # ایجاد لاگ
                LockLogs.objects.create(
                    lockmodel=lock_model,
                    owner=request.user,
                    status=lock_model.status,
                    gs=lock_model.gs,
                    info=f'ثبت صورتجلسه - {form.cleaned_data.get("description", "")}'
                )

                add_to_log(request, f'ثبت صورتجلسه برای جایگاه {lock_model.gs.name} با سریال {serial_number}', 0)
                messages.success(request, f'صورتجلسه با موفقیت ثبت شد. شماره سریال: {serial_number}')
                return redirect('lock:history')

            except Exception as e:
                print(e)
                transaction.set_rollback(True)
                messages.error(request, f'خطا در ثبت صورتجلسه: {str(e)}')
    else:
        form = SoratJalaseForm(user=request.user)
        form.fields['gs'].queryset = GsModel.object_role.c_gsmodel(request)
        form.fields['gs'].label_from_instance = lambda obj: f"{obj.gsid} - {obj.name}"

    context = {'form': form}
    return TemplateResponse(request, template_file, context)

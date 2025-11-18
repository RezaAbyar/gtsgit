import openpyxl
import os
from django.db.models import Q
from django.core.files import File
from django.template.response import TemplateResponse
from django.views.generic import ListView, CreateView
from django.utils.decorators import method_decorator
from base.permission_decoder import cache_permission
from .models import BazrasNegar, PngImage
from .forms import BazrasNegarSearchForm, BazrasNegarForm
from django.core.exceptions import ValidationError, ObjectDoesNotExist
import pytesseract
from PIL import Image
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404


@cache_permission('bazrasnegar_search')
def search_bazras_negar(request):
    form = BazrasNegarSearchForm(request.GET or None)
    results = None

    if form.is_valid():
        results = BazrasNegar.objects.all().order_by('-id')
        query = form.cleaned_data.get('query')
        if query:
            results = results.filter(
                Q(title__icontains=query) |
                Q(number__icontains=query) |
                Q(tarikh__icontains=query) |
                Q(tarikh__icontains=query) |
                Q(extracted_text__icontains=query)
            )

    return TemplateResponse(request, 'search_results.html', {
        'form': form,
        'results': results
    })


@cache_permission('bazrasnegar_search')
def bazras_negar_create(request):
    if request.method == 'POST':
        form = BazrasNegarForm(request.POST, request.FILES)

        if form.is_valid():
            a = form.save()
            file_ext = os.path.splitext(a.file.name)[1].lower()
            if file_ext in ('.tif', '.tiff'):
                a.convert_to_png()
            _file = BazrasNegar.objects.get(id=a.id)

            pytesseract.pytesseract.tesseract_cmd = r'C:\Tesseract-OCR\tesseract.exe'
            image_path = os.path.join(settings.MEDIA_ROOT, _file.file.name)
            if os.path.exists(image_path):
                image = Image.open(image_path)
                text = pytesseract.image_to_string(image, lang='fas+eng')
                _file.extracted_text = text
                _file.save()
            return redirect('/bazrasnegar/search/')
    else:
        form = BazrasNegarForm()

    return TemplateResponse(request, 'bazrasnegar_form.html', {'form': form})


def bazrasi():
    path = 'bank/orginal.xlsx'

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    BASE_FILE_PATH = 'bank/'
    p = 0
    for i in range(1, m_row + 1):
        if i > 1:
            _number = str(sheet_obj.cell(row=i, column=1).value)
            _tarikh = str(sheet_obj.cell(row=i, column=2).value)
            _info = str(sheet_obj.cell(row=i, column=3).value)
            _title = str(sheet_obj.cell(row=i, column=4).value)
            _file_name = str(sheet_obj.cell(row=i, column=5).value)
            print(i)
            try:
                if _file_name and _file_name.strip():
                    file_path = os.path.join(BASE_FILE_PATH, _file_name)

                    if os.path.exists(file_path):
                        with open(file_path, 'rb') as f:
                            django_file = File(f)
                            bazras = BazrasNegar(
                                number=_number,
                                tarikh=_tarikh,
                                info=_info,
                                title=_title
                            )
                            bazras.file.save(_file_name, django_file, save=True)
                    else:
                        print(f"File not found: {file_path}")
                        bazras = BazrasNegar.objects.create(
                            number=_number,
                            tarikh=_tarikh,
                            info=_info,
                            title=_title
                        )
                else:
                    # اگر فایلی وجود ندارد
                    bazras = BazrasNegar.objects.create(
                        number=_number,
                        tarikh=_tarikh,
                        info=_info,
                        title=_title
                    )
            except Exception as e:
                p += 1
                print(e)
    print(p)

    # BazrasNegar.objects.create(number=_number,tarikh=_tarikh,info=_info,title=_title,file=_file )
    return True


@cache_permission('message')
def image_viewer(request, message_id):
    convert_to_png(message_id)
    message = get_object_or_404(BazrasNegar, id=message_id)
    total_pages = PngImage.objects.filter(tiff_image=message).count()
    current_page = 1  # صفحه پیش‌فرض
    if request.method == 'POST':
        current_page = int(request.POST.get('page', 1))

    image_data = get_object_or_404(PngImage, tiff_image_id=message_id, page_number=current_page)

    return TemplateResponse(request, 'image_viewer.html', {
        'message': message,
        'image_data': image_data,
        'current_page': current_page,
        'total_pages': total_pages,
    })


def convert_to_png(_id):
    from PIL import Image
    from io import BytesIO
    from django.core.files.base import ContentFile
    png_pages = []
    image = BazrasNegar.objects.get(id=_id)
    if PngImage.objects.filter(tiff_image_id=_id).exists():
        ok = 1
    else:
        with Image.open(image.file) as img:
            # ذخیره هر صفحه به عنوان PNG
            for i in range(img.n_frames):
                img.seek(i)

                # ایجاد نام فایل
                original_name = os.path.splitext(os.path.basename(image.file.name))[0]
                png_name = f"{original_name}_page_{i + 1}.png"

                # تبدیل به PNG در حافظه
                output = BytesIO()
                img.save(output, format='PNG')
                output.seek(0)

                # ذخیره مدل PNG
                png_model = PngImage(
                    tiff_image_id=_id,
                    page_number=i + 1
                )
                png_model.png_file.save(png_name, ContentFile(output.read()))
                png_pages.append(png_model)
                output.close()

    return png_pages

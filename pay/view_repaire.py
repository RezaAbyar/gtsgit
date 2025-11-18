import datetime

from django.views import View
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Sum, Avg, When, Case, Max
import jdatetime
from django.db import IntegrityError
from accounts.logger import add_to_log
from base.models import UserPermission, DefaultPermission, Storage, Zone
from pay.forms import RepqiresAddForm, RepqiresAddStoreForm, UploadAddForm, RepqireAddStoreName, RepairParametrForm
from pay.models_repaire import Repaires, RepaireStores
from util import DENY_PAGE, HOME_PAGE, SUCCESS_MSG, EXCEL_MODE, DATE_FORMAT, PAY_PAGE, SENDLIST_PAGE, ZONE_NAME, \
    ADD_STORE_MSG, EXCEL_EXPORT_FILE
from .filter import RepaireFilters
from .models import RepairStoreName, StoreList, Repair, RepairRole

_today = str(jdatetime.date.today())
today = _today.replace("-", "/")
startdate = today[:8]
startdate = startdate + "01"


def checkxss(val):
    val = val.replace('<', '')
    val = val.replace('>', '')
    val = val.replace('/', '')
    return val


def view_repaire(request):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='kargahtakhsis')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    _list = None
    _sum = None
    add_to_log(request, f'مشاهده فرم لیست قطعات ارسالی ', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    select3 = str(request.GET.get('select3'))
    vore = request.GET.get('vore')
    storages = Zone.objects.filter(storage=True)
    storagesid = 0
    if select3 == "None":
        select3 = "1"
    if not select3:
        select3 = "1"

    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"

    else:

        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"

    if request.user.owner.role.role in 'mgr,setad':
        storagesid = Storage.objects.get(id=request.user.owner.storage_id).id
    if request.user.owner.role.role == 'zone':
        storagesid = Storage.objects.get(zone_id=request.user.owner.zone.id).id
    if request.user.owner.role.role == 'posht':
        toragesid = Storage.objects.get(zone_id=request.user.owner.zone.id).id

        if select3 == '1':
            _list = Repaires.objects.filter(tarikh__range=(datein, dateout), status__in=[1, 2, 3],
                                            zone_id=request.user.owner.zone_id).order_by('-id')
        if select3 == '2':
            _list = Repaires.objects.filter(marsole_send__range=(datein, dateout), status__in=[1, 2, 3],
                                            zone_id=request.user.owner.zone_id).order_by('-id')
        if select3 == '3':
            _list = Repaires.objects.filter(marsole_resid__range=(datein, dateout), status__in=[1, 2, 3],
                                            zone_id=request.user.owner.zone_id).order_by('-id')
        _filter = RepaireFilters(request.GET, queryset=_list)
        _list = _filter.qs
    if request.user.owner.role.role in ['mgr', 'setad', 'posht']:
        if select3 == '1':
            _list = Repaires.objects.filter(tarikh__range=(datein, dateout), status__in=[1, 2, 3]
                                            ).order_by('-id')

        if select3 == '2':
            _list = Repaires.objects.filter(marsole_send__range=(datein, dateout), status__in=[1, 2, 3]
                                            ).order_by('-id')

        if select3 == '3':
            _list = Repaires.objects.filter(marsole_resid__range=(datein, dateout), status__in=[1, 2, 3]
                                            ).order_by('-id')

        _filter = RepaireFilters(request.GET, queryset=_list)
        if _filter.data:
            _list = _filter.qs
        else:
            _list = _list.filter(storage_id=request.user.owner.defaultstorage)

    if vore == "2":
        add_to_log(request, 'ارسال آمار لیست ارسالی ها به اکسل  ', 0)
        my_path = "sendstore.xlsx"
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = 'attachment; filename=' + my_path
        font = Font(bold=True)
        fonttitr = Font(bold=True, size=20)
        fonttitr2 = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active  # work with default worksheet
        ws1.title = "گزارش لیست قطعات ارسالی "
        ws1.sheet_view.rightToLeft = True
        ws1.firstFooter.center.text = "ali"
        ws1.merge_cells('A1:j1')

        ws1["A1"] = f'گزارش لیست قطعات ارسالی تاریخ   {today}'
        ws1["A1"].font = fonttitr

        ws1.merge_cells('A2:A2')
        ws1["A2"] = "ردیف"
        ws1["A2"].font = font

        ws1.merge_cells('B2:B2')
        ws1["B2"] = "تامین کننده"
        ws1["B2"].font = fonttitr2

        ws1.merge_cells('C2:C2')
        ws1["C2"] = "منطقه "
        ws1["C2"].font = font

        ws1.merge_cells('D2:D2')
        ws1["D2"] = " تاریخ تخصیص"
        ws1["D2"].font = font

        ws1.merge_cells('E2:E2')
        ws1["E2"] = "تعداد کارتخوان"
        ws1["E2"].font = font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(15.25)
        ws1.column_dimensions['D'].width = float(25.25)
        ws1.column_dimensions['E'].width = float(15.25)
        ws1.column_dimensions['F'].width = float(18.25)
        ws1.column_dimensions['G'].width = float(35.25)
        ws1.column_dimensions['H'].width = float(30.25)
        ws1.column_dimensions['I'].width = float(25.25)
        ws1.column_dimensions['J'].width = float(22.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        myfont = Font(size=14, bold=True)  # font styles
        my_fill = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        my_fill2 = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        i = 0

        for item in _list:
            i += 1
            d = [i, str(item.storage.name), str(item.zone.name), str(item.normal_date()), str(item.master),
                 str(item.pinpad),
                 str(item.mdate()) + ' - ' + str(item.marsole), str(item.rdate()), str(item.get_takhsisdate_diff()),
                 str(item.get_date_diff())]

            ws1.append(d)

        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        i += 4
        for cell in ws1[i:i]:  # Last row
            cell.font = myfont
            cell.fill = my_fill2
            cell.border = thin_border

        for cell in ws1["2:2"]:  # First row
            cell.font = myfont
            cell.fill = my_fill
            cell.border = thin_border
        wb.save(response)
        return response

    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:
        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)
        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)

        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)

        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)

            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)

            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return render(request, 'repair/requestreapirestore.html',
                  {'filter': _filter, 'list': page_object, 'query_string': query_string, 'this_date': this_date,
                   'today_date': today_date, 'listsum': _sum, 'storages': storages, 'storagesid': storagesid,
                   'formpermmision': formpermmision, 'page_obj': page_obj, 'tedad': tedad, })


def add_repaire(request):
    if request.method == 'POST':
        form = RepqiresAddForm(request.POST)
        if form.is_valid():
            if request.user.owner.role.role in ['setad', 'mgr']:
                st_id = Storage.objects.get(id=request.user.owner.storage_id).id
            else:
                st_id = Storage.objects.get(zone_id=request.user.owner.zone.id).id
            form.instance.owner_id = request.user.owner.id
            form.instance.storage_id = st_id
            form.instance.status_id = 1
            form.save()
        return redirect('pay:view_repaire')
    return redirect('pay:view_repaire')


def uploadstore(request):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        st = request.POST.get('mystoreid')

        repairs = Repaires.objects.get(id=int(st))
        form = UploadAddForm(request.POST, instance=repairs)
        if form.is_valid():
            form.instance.status_id = 2
            form.instance.marsole_send = datetime.datetime.today()
            form.save()
        return redirect(url)
    return redirect(url)


def downloadstore(request, _id):
    url = request.META.get('HTTP_REFERER')

    repairs = Repaires.objects.get(id=_id)

    repairs.status_id = 3
    repairs.marsole_resid = datetime.datetime.today()
    repairs.save()
    # repair = RepairRole.objects.get(storage_id=repairs.storage_id,repairstore_id=repairs.repairstore_id)
    # repair.inventory = repair.inventory +repairs.valuecount
    # repair.ofroadvalue = repair.ofroadvalue - repairs.valuecount
    return redirect(url)


def deletezone_repaire(request, _id):
    url = request.META.get('HTTP_REFERER')
    Repaires.objects.get(id=_id).delete()
    add_to_log(request, ' حذف بسته یدکی تخصیص داده شده ' + str(id), 0)
    messages.info(request, ' بسته تخصیص داده شده  بدرستی حذف شد')
    return redirect(url)


def liststoreyadaki(request, _id):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='kargahtakhsis')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # //////////////////////////////////////////////////////////////////////////////////////////////////////
    if request.user.owner.role.role in ['setad', 'mgr']:
        _storage = Storage.objects.filter(id=request.user.owner.storage_id).last()
    else:
        _storage = Storage.objects.filter(zone_id=request.user.owner.zone_id).last()

    _zonename = Repaires.objects.get(id=_id)
    zonename = _zonename.zone.name
    if _zonename.storage_id == _storage.id and _zonename.status_id == 1:
        a = 1
    else:
        a = 2
    storename = RepairStoreName.objects.all()
    for i in storename:
        _repairsstore = RepaireStores.objects.filter(store_id=_id, repairstore_id=i.id)
        if _repairsstore.count() == 0:
            RepaireStores.objects.create(store_id=_id, repairstore_id=i.id, amount=0)

    stores = RepaireStores.objects.filter(store_id=_id)
    repaires = Repaires.objects.get(id=_id)

    if request.method == 'POST':
        for store in stores:
            amount = request.POST.get(f'amount-{store.repairstore.id}')
            if amount:
                store.amount = int(amount)
                store.save()
        messages.success(request, 'عملیات با موفقیت انجام شد')
        return redirect('pay:view_repaire')
    return render(request, 'repair/addrequeststore.html',
                  {'storename': storename, 'stores': stores, 'storeid': _id, 'zonename': zonename, 'a': a,
                   'formpermmision': formpermmision, 'repaires': repaires})


def add_repaire_store(request, _id):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        form = RepqiresAddStoreForm(request.POST)
        if form.is_valid():
            form.instance.owner_id = request.user.owner.id
            form.instance.store_id = _id
            form.save()
        return redirect(url)
    return redirect(url)


def deletezone_repaire_store(request, _id):
    url = request.META.get('HTTP_REFERER')
    RepaireStores.objects.get(id=_id).delete()
    add_to_log(request, ' حذف قطعه یدکی تخصیص داده شده ' + str(_id), 0)
    messages.info(request, ' قطعه بدرستی حذف شد')
    return redirect(url)


def delete_store_item(request):
    if request.method == 'POST':
        _id = request.POST.get('id')
        a = Repair.objects.get(id=_id)
        _st = a.storage.id
        a.delete()
        Repair.checkuserepair(_st)

        add_to_log(request, ' حذف قطعه یدکی  ' + str(_id), 0)

    return JsonResponse({"msg": 'ok', })


def addstorefunc(request):
    if request.method == 'POST':
        _id = request.POST.get('id')

        _list = []
        _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
        _storageid = _storage.id
        _tarikh = jdatetime.date.today()

        stores = Repair.objects.filter(store_id=_id, status=0)
        for store in stores:
            dict = {
                'id': store.id,
                'name': store.repairstore.name,
                'stid': store.repairstore.id,
                'count': store.valuecount
            }
            _list.append(dict)
        _storename = []
        storename = RepairStoreName.objects.all()
        for store in storename:
            _rs = Repair.objects.filter(tarikh=_tarikh, store_id=_id, repairstore_id=store.id,
                                        storage_id=_storageid).last()
            tedad = 0 if _rs is None else _rs.valuecount

            dict = {
                'id': str(store.id),
                'name': str(store.name),
                'tedad': tedad
            }
            _storename.append(dict)

        return JsonResponse({"storename": _storename, 'mylist': _list})


def newstorefun(request):
    if request.method == 'POST':
        _id = request.POST.get('store')
        _repaire = request.POST.get('name')
        _count = request.POST.get('amount')
        _storage = Storage.objects.filter(zone_id=request.user.owner.zone_id).last()
        Repair.objects.create(valuecount=_count, repairstore_id=_repaire, store_id=_id, storage_id=_storage.id)
        Repair.checkuserepair(_storage.id)
        return JsonResponse({"msg": 'ok', })


class RepairRequestList(View):
    template_file = 'repair/repirrequest.html'

    storage = Storage.objects.all().order_by('sortid')

    def get(self, request):
        owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
        if owner_p.count() == 0:
            owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                       semat_id=request.user.owner.refrence_id)
        ua = owner_p.get(permission__name='kargahmojodi')
        if ua.accessrole.ename == 'no':
            messages.warning(request, DENY_PAGE)
            return redirect(HOME_PAGE)
        formpermmision = {}
        for i in owner_p:
            formpermmision[i.permission.name] = i.accessrole.ename
        # //////////////////////////////////////////////////////////////////////////////////////////////////////
        if request.user.owner.role.role in 'zone,posht':
            storage = Storage.objects.filter(zone_id=request.user.owner.zone_id)
        if request.user.owner.role.role in 'mgr,setad':
            storage = Storage.objects.all().order_by('sortid')
        return render(request, self.template_file, {'storage': storage, 'formpermmision': formpermmision})

    def post(self, request):
        if request.method == 'POST':
            storage_id = request.POST.get('storage')
            if request.user.owner.role.role in 'zone,posht':
                storage = Storage.objects.filter(zone_id=request.user.owner.zone_id)
            if request.user.owner.role.role in 'mgr,setad':
                storage = Storage.objects.all().order_by('sortid')
            _storageid = Storage.objects.get(id=storage_id).id

            repairrole = RepairRole.objects.filter(storage_id=_storageid, required__gt=0)
            return render(request, self.template_file,
                          {'storage': storage, 'repairrole': repairrole, 'storage_id': int(storage_id)})


def repairstorename(request):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='repairstorename')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # //////////////////////////////////////////////////////////////////////////////////////////////////////
    _list = RepairStoreName.objects.all()
    return render(request, 'repair/repairstorename.html', {'list': _list, 'formpermmision': formpermmision})


def addstorename(request):
    if request.method == 'POST':
        form = RepqireAddStoreName(request.POST)
        if form.is_valid():
            form.save()
        return redirect('pay:repairstorename')
    return redirect('pay:repairstorename')


def repairparametrs(request, _id):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='repairstorename')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # //////////////////////////////////////////////////////////////////////////////////////////////////////
    for storage in Storage.objects.all():
        for item in RepairStoreName.objects.all():
            try:
                RepairRole.objects.create(storage_id=storage.id, repairstore_id=item.id, minvalue=0, usevalue=0,
                                          startvalue=0, uniq=str(storage.id) + "-" + str(item.id))
            except:
                a = 1

    _list = RepairRole.objects.filter(repairstore_id=_id)
    return render(request, 'repair/repairparametrs.html', {'list': _list, 'formpermmision': formpermmision})


def add_repairparametrs(request):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        _id = request.POST.get('mystoreid')
        repairparametr = RepairRole.objects.get(id=_id)
        form = RepairParametrForm(request.POST, instance=repairparametr)
        if form.is_valid():
            form.save()
            Repair.checkuserepair(repairparametr.storage_id)
        return redirect(url)
    return redirect(url)


def plusstore(request):
    if request.method == 'POST':
        _storeid = request.POST.get('storeid')
        _val = request.POST.get('val')
        _val2 = request.POST.get('val2')
        _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
        _storageid = _storage.id
        _tarikh = jdatetime.date.today()

        _rs = Repair.objects.filter(tarikh=_tarikh, repairstore_id=_val, store_id=_storeid,
                                    storage_id=_storageid)
        if _rs.count() == 0:
            Repair.objects.create(repairstore_id=_val, store_id=_storeid,
                                  valuecount=1, tarikh=_tarikh, storage_id=_storageid)
            amount = 1
        else:
            _rs = _rs.last()
            _rs1 = Repair.objects.get(tarikh=_rs.tarikh, repairstore_id=_val, store_id=_storeid,
                                      storage_id=_storageid)
            if int(_val2) == 1:
                _rs1.valuecount += 1
            else:
                _rs1.valuecount -= 1
            _rs1.save()
            amount = _rs1.valuecount
            Repair.checkuserepair(_storageid)

        return JsonResponse({"msg": 'ok', 'newval': amount})

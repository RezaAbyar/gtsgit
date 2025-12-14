import math
from datetime import datetime, date, timedelta
import datetime
from django.template.response import TemplateResponse
from django.utils import timezone
from django.db import transaction
from operator import itemgetter
import openpyxl
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import PatternFill, Font, Alignment, Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from django.db.models import Count, Sum, Avg, When, Case, Func, CharField, Value, F, Q
from django.shortcuts import render, redirect
from rest_framework.decorators import api_view, permission_classes
from django.contrib import messages
from rest_framework.permissions import IsAuthenticated
from base.filter import StoreFilters
from base.forms import open_excel
from django.contrib.auth.models import User
from base.permission_decoder import cache_permission
from lock.models import LockModel, InsertLock
from util import DENY_PAGE, HOME_PAGE, SUCCESS_MSG, EXCEL_MODE, DATE_FORMAT, PAY_PAGE, SENDLIST_PAGE, ZONE_NAME, \
    ADD_STORE_MSG, EXCEL_EXPORT_FILE
from .models import PayBaseParametrs, Owner, Payroll, StoreList, Store, StoreHistory, PayItems, PayParametr, \
    PayDarsadMah, HistorySt, Post, Tektaeed, StatusRef, StatusStore, TekKarkard, BaseGroup, RepairStoreName, \
    Repair, StoreView, RepairStore, ImgSerial, PersonPayment, ZoneToStorage, Tadiltemp, StoreManufacturer, \
    SerialRange, GenerateSerialNumber
from base.models import Mount, UserPermission, DefaultPermission, Zone, Storage, UploadExcel, Pump, Ticket, \
    GsModel, Workflow, NegativeScore, Parametrs
from django.http import HttpResponse, JsonResponse
from jalali.Jalalian import JDate
import jdatetime
from django.db import IntegrityError
from accounts.logger import add_to_log
from django.views import View
from .forms import RepairForm, ImageStore, StoreListForm
from utils.exception_helper import to_miladi, SendSmS, checkxss, checknumber
from django.conf import settings

_today = str(jdatetime.date.today())
today = _today.replace("-", "/")
startdate = today[:8]
startdate = startdate + "01"


@cache_permission('pay')
def paybase(request):
    add_to_log(request, 'مشاهده فرم حقوق و دستمزد', 0)
    mounts = Mount.objects.all().order_by('id')
    owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True)
    pyzarib = PayBaseParametrs.objects.get(active=True, id=8)
    pyezafe = PayBaseParametrs.objects.get(active=True, id=16)
    dastmozdmah = PayBaseParametrs.objects.get(active=True, id=7)
    mmozd = dastmozdmah.price
    dastmozdmah = PayBaseParametrs.objects.get(active=True, id=6)
    dmozd = dastmozdmah.price
    if request.method == 'POST':
        period = request.POST.get('period')
        owner = request.POST.get('owner')
        kark = request.POST.get('kark')
        _khodro = request.POST.get('kilometr')
        TekKarkard.get_karkard(value=int(kark), period=period, tekid=owner, kilometr=_khodro)

        parm1 = int(request.POST.get('parm1'))
        parm2 = int(request.POST.get('parm2'))
        ezafekarsum = Payroll.objects.filter(period_id=period, tek__zone_id=request.user.owner.zone_id,
                                             paybaseparametrs_id=16).aggregate(sum=Sum('count'))

        ezafekarsum = ezafekarsum['sum']
        co_tek = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True).count()
        sarane = PayBaseParametrs.objects.get(active=True, id=16)
        sumsarane = co_tek * sarane.price
        if ezafekarsum is None:
            takhsis = 0
        else:
            takhsis = (parm2 + ezafekarsum)

        if takhsis > sumsarane:
            messages.error(request, F'مقدار تخصیص سهمیه نباید از سرانه بیشتر باشد ')

        if parm1 > pyzarib.count:
            messages.error(request, F'مقدار ظریب اتلاف وقت نباید بیش از {pyzarib.count} باشد ')
            return redirect(PAY_PAGE)
        if parm2 > pyezafe.count:
            messages.error(request, F'مقدار بهره وری نباید بیش از {pyezafe.count} باشد ')
            return redirect(PAY_PAGE)
        payrolls = Payroll.objects.filter(period_id=period, tek_id=owner)
        for py in payrolls:
            if py.accept:
                messages.warning(request, 'حقوق و دستمزد این شخص قبلا توسط مدیر تایید نهایی شده و قابل ویرایش نمی باشد')
                return redirect(PAY_PAGE)
        if payrolls.count() == 0:
            TekKarkard.get_karkard(value=int(kark), period=period, kilometr=_khodro, tekid=owner)
            Payroll.objects.create(tek_id=owner, period_id=period, user_id=request.user.id,
                                   paybaseparametrs_id=4, count=parm1, price=(mmozd * parm1) / 100)

            Payroll.objects.create(tek_id=owner, period_id=period, user_id=request.user.id,
                                   paybaseparametrs_id=8, count=parm1, price=(mmozd * parm1) / 100)
            Payroll.objects.create(tek_id=owner, period_id=period, user_id=request.user.id,
                                   paybaseparametrs_id=16, count=parm2, price=(dmozd / 7.33 * 1.4 * parm2))
        else:
            mypay = Payroll.objects.get(period_id=period, tek_id=owner, paybaseparametrs_id=8)
            mypay.count = parm1
            mypay.price = (mmozd * parm1) / 100
            mypay.save()
            mypay = Payroll.objects.get(period_id=period, tek_id=owner, paybaseparametrs_id=16)
            mypay.count = parm2
            mypay.price = (dmozd / 7.33 * 1.4 * parm2)
            mypay.save()
        lists = Payroll.objects.filter(period_id=period).values('tek__name', 'tek_id', 'tek__lname').annotate(
            te=Count('id'))
        paylist = Payroll.objects.filter(period_id=period, tek_id=owner)
        paysum = Payroll.objects.filter(period_id=period, tek_id=owner).aggregate(sum=Sum('price'))
        context = {'mounts': mounts, 'owners': owners, 'lists': lists, 'mmozd': mmozd, 'dmozd': dmozd,
                   'paylist': paylist, 'paysum': paysum, 'period': int(period)}
        return TemplateResponse(request, 'pay/paybase.html', context)
    context = {'mounts': mounts, 'owners': owners, 'mmozd': mmozd, 'dmozd': dmozd}
    return TemplateResponse(request, 'pay/paybase.html', context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getpriod(request):
    mypay = []
    ezafekarsum = 0
    sumsarane = 0
    active = True
    if request.method == 'POST':
        period = request.POST.get('priod')
        try:
            active = PayDarsadMah.objects.filter(period_id=period).first()
            active = active.active
            lists = Payroll.objects.filter(period_id=period).values('tek__name', 'tek_id', 'tek__lname').annotate(
                te=Count('id'))
            co_tek = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True).count()
            sarane = PayBaseParametrs.objects.get(active=True, id=16)
            sumsarane = co_tek * sarane.price
            ezafekarsum = Payroll.objects.filter(period_id=period, tek__zone_id=request.user.owner.zone_id,
                                                 paybaseparametrs_id=16).aggregate(sum=Sum('count'))
            if ezafekarsum['sum']:
                ezafekarsum = ezafekarsum['sum']
                ezafekarsum = sumsarane - ezafekarsum
            else:
                ezafekarsum = 0

            for _list in lists:
                thisdict = {
                    "id": _list['tek_id'],
                    "tek": _list['tek__name'] + ' ' + _list['tek__lname'],
                }
                mypay.append(thisdict)

        except PayBaseParametrs.DoesNotExist:

            active = True

    return JsonResponse(
        {"mypay": mypay, 'active': active, 'ezafekarsum': ezafekarsum, 'sumsarane': sumsarane, 'message': 'success'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def gettekpay(request):
    paylist = []
    paysum = ""
    tekname = ""
    karkerd = 30
    pr = ""
    co = ""
    if request.method == 'POST':
        period = request.POST.get('priod')
        tekid = request.POST.get('tekid')

        _list = Payroll.objects.filter(period_id=period, tek_id=tekid)
        paysum = Payroll.objects.filter(period_id=period, tek_id=tekid).aggregate(sum=Sum('price'))
        paysum = paysum['sum']
        for list1 in _list:
            thisdict = {
                "id": list1.paybaseparametrs.id,
                "name": list1.paybaseparametrs.name,
                "count": list1.count,
                "price": list1.price,
            }
            if list1.paybaseparametrs.id == 8:
                co = list1.count
            if list1.paybaseparametrs.id == 16:
                pr = list1.count
            tekname = list1.tek.name + ' ' + list1.tek.lname
            paylist.append(thisdict)
    return JsonResponse(
        {"paylist": paylist, "karkerd": karkerd, 'paysum': paysum, 'tekname': tekname, 'pr': pr, 'co': co,
         'message': 'success'})


@cache_permission('commitmgr')
def commitmgr(request):
    etlaf = 0
    etlafprice = 0
    zarib = 0
    ezafe = 0
    zprice = 0
    eprice = 0
    add_to_log(request, 'مشاهده فرم تایید حقوق و دستمزد', 0)
    thislist = []
    mounts = Mount.objects.filter(isshow=True).order_by('id')

    if request.method == 'POST':
        period = request.POST.get('period')

        # پیدا کردن ماه قبلی
        try:
            current_month = Mount.objects.get(id=period)

            # پیدا کردن ماه قبل بر اساس سال و ماه
            prev_month = Mount.objects.filter(id__lt=period).order_by('-id')
            prev_month=prev_month.first()
        except Mount.DoesNotExist:
            prev_month = None


        if request.user.owner.role.role != 'setad':
            lists = Payroll.objects.filter(period_id=period, tek__zone_id=request.user.owner.zone_id).values(
                'tek__name',
                'tek_id',
                'tek__lname',
                'accept').annotate(
                te=Count('id'))
        else:
            lists = Payroll.objects.filter(period_id=period, tek__refrence_id=9).values('tek__name',
                                                                                        'tek_id',
                                                                                        'tek__lname',
                                                                                        'accept').annotate(
                te=Count('id'))

        for _list in lists:
            # داده‌های ماه جاری
            results = Payroll.objects.filter(period_id=period,
                                             paybaseparametrs_id__enname__in=['bazdehi', 'foghejazb', 'ezafekari'],
                                             tek_id=_list['tek_id'])

            # داده‌های ماه قبل (اگر وجود داشته باشد)
            prev_results = None
            if prev_month:
                prev_results = Payroll.objects.filter(period_id=prev_month.id,
                                                      paybaseparametrs_id__enname__in=['bazdehi', 'foghejazb',
                                                                                       'ezafekari'],
                                                      tek_id=_list['tek_id'])

            # متغیرهای ماه جاری
            current_zarib = 0
            current_zprice = 0
            current_etlaf = 0
            current_etlafprice = 0
            current_ezafe = 0
            current_eprice = 0

            # متغیرهای ماه قبل
            prev_zarib = 0
            prev_zprice = 0
            prev_etlaf = 0
            prev_etlafprice = 0
            prev_ezafe = 0
            prev_eprice = 0

            # محاسبه مقادیر ماه جاری
            for result in results:
                if result.paybaseparametrs.enname == 'bazdehi':
                    current_zarib = result.count
                    current_zprice = result.price
                elif result.paybaseparametrs.enname == 'foghejazb':
                    current_etlaf = result.count
                    current_etlafprice = result.price
                elif result.paybaseparametrs.enname == 'ezafekari':
                    current_ezafe = result.count
                    current_eprice = result.price

            # محاسبه مقادیر ماه قبل
            if prev_results:
                for prev_result in prev_results:
                    if prev_result.paybaseparametrs.enname == 'bazdehi':
                        prev_zarib = prev_result.count
                        prev_zprice = prev_result.price
                    elif prev_result.paybaseparametrs.enname == 'foghejazb':
                        prev_etlaf = prev_result.count
                        prev_etlafprice = prev_result.price
                    elif prev_result.paybaseparametrs.enname == 'ezafekari':
                        prev_ezafe = prev_result.count
                        prev_eprice = prev_result.price

            # محاسبه تغییرات
            zarib_change = current_zarib - prev_zarib
            zprice_change = current_zprice - prev_zprice
            etlaf_change = current_etlaf - prev_etlaf
            etlafprice_change = current_etlafprice - prev_etlafprice
            ezafe_change = current_ezafe - prev_ezafe
            eprice_change = current_eprice - prev_eprice

            mydict = {
                "name": _list['tek__name'] + ' ' + _list['tek__lname'],
                "id": _list['tek_id'],
                # مقادیر ماه جاری
                "current_etlaf": current_etlaf,
                "current_etlafprice": current_etlafprice,
                "current_zarib": current_zarib,
                "current_ezafe": current_ezafe,
                "current_zprice": current_zprice,
                "current_eprice": current_eprice,
                # مقادیر ماه قبل
                "prev_etlaf": prev_etlaf,
                "prev_etlafprice": prev_etlafprice,
                "prev_zarib": prev_zarib,
                "prev_ezafe": prev_ezafe,
                "prev_zprice": prev_zprice,
                "prev_eprice": prev_eprice,
                # تغییرات
                "etlaf_change": etlaf_change,
                "etlafprice_change": etlafprice_change,
                "zarib_change": zarib_change,
                "zprice_change": zprice_change,
                "ezafe_change": ezafe_change,
                "eprice_change": eprice_change,
                # سایر اطلاعات
                "period": period,
                "prev_period": prev_month.id if prev_month else None,
                "current_sum": current_zprice + current_eprice + current_etlafprice,
                "prev_sum": prev_zprice + prev_eprice + prev_etlafprice,
                "sum_change": (current_zprice + current_eprice + current_etlafprice) - (
                            prev_zprice + prev_eprice + prev_etlafprice),
                "accept": _list['accept'],
            }
            thislist.append(mydict)

        return TemplateResponse(request, 'pay/commitmgr.html',
                                {'mounts': mounts,
                                 'thislist': thislist,
                                 'period': int(period),
                                 'prev_month_name': prev_month.mount if prev_month else None,
                                 'prev_month_id': prev_month.id if prev_month else None})

    return TemplateResponse(request, 'pay/commitmgr.html', {'mounts': mounts, 'period': 0})


def acceptpay(request, owner, period):
    add_to_log(request, 'ذخیره مبلغ حقوق دریافتی توسط تکنسین', 0)
    if request.method == 'POST':
        tarikh = request.POST.get('select')
        namepay = request.POST.get('namepay')
        lists = Payroll.objects.filter(period_id=int(period), tek_id=int(owner))
        for i in lists:
            i.acceptpay = True
            i.acceptupdatepay = tarikh
            i.mablagh = namepay
            i.save()
        messages.success(request, 'تغییرات با موفقیت ذخیره شد سپاس از مشارکت شما')
    return redirect(HOME_PAGE)


@cache_permission('viewcommit')
def viewcommit(request):
    zones = Zone.objects_limit.all()
    thislist = []
    add_to_log(request, f'مشاهده فرم تایید حقوق و دستمزد ', 0)
    mounts = Mount.objects.all().order_by('id')
    etlaf = 0
    etlafprice = 0
    zarib = 0
    ezafe = 0
    zprice = 0
    eprice = 0
    if request.method == 'POST':
        period = request.POST.get('period')
        zone = request.POST.get('zone')

        try:
            current_month = Mount.objects.get(id=period)

            # پیدا کردن ماه قبل بر اساس سال و ماه
            prev_month = Mount.objects.filter(id__lt=period).order_by('-id')
            prev_month = prev_month.first()
        except Mount.DoesNotExist:
            prev_month = None
        lists = Payroll.objects.filter(period_id=period, tek__zone_id=int(zone)).values('tek__name', 'tek_id',
                                                                                        'tek__lname',
                                                                                        'accept',
                                                                                        'acceptpay',
                                                                                        'acceptupdatepay').annotate(
            te=Count('id'))
        for _list in lists:
            # داده‌های ماه جاری
            results = Payroll.objects.filter(period_id=period,
                                             paybaseparametrs_id__enname__in=['bazdehi', 'foghejazb', 'ezafekari'],
                                             tek_id=_list['tek_id'])

            # داده‌های ماه قبل (اگر وجود داشته باشد)
            prev_results = None
            if prev_month:
                prev_results = Payroll.objects.filter(period_id=prev_month.id,
                                                      paybaseparametrs_id__enname__in=['bazdehi', 'foghejazb',
                                                                                       'ezafekari'],
                                                      tek_id=_list['tek_id'])

            # متغیرهای ماه جاری
            current_zarib = 0
            current_zprice = 0
            current_etlaf = 0
            current_etlafprice = 0
            current_ezafe = 0
            current_eprice = 0

            # متغیرهای ماه قبل
            prev_zarib = 0
            prev_zprice = 0
            prev_etlaf = 0
            prev_etlafprice = 0
            prev_ezafe = 0
            prev_eprice = 0

            # محاسبه مقادیر ماه جاری
            for result in results:
                if result.paybaseparametrs.enname == 'bazdehi':
                    current_zarib = result.count
                    current_zprice = result.price
                elif result.paybaseparametrs.enname == 'foghejazb':
                    current_etlaf = result.count
                    current_etlafprice = result.price
                elif result.paybaseparametrs.enname == 'ezafekari':
                    current_ezafe = result.count
                    current_eprice = result.price

            # محاسبه مقادیر ماه قبل
            if prev_results:
                for prev_result in prev_results:
                    if prev_result.paybaseparametrs.enname == 'bazdehi':
                        prev_zarib = prev_result.count
                        prev_zprice = prev_result.price
                    elif prev_result.paybaseparametrs.enname == 'foghejazb':
                        prev_etlaf = prev_result.count
                        prev_etlafprice = prev_result.price
                    elif prev_result.paybaseparametrs.enname == 'ezafekari':
                        prev_ezafe = prev_result.count
                        prev_eprice = prev_result.price

            # محاسبه تغییرات
            zarib_change = current_zarib - prev_zarib
            zprice_change = current_zprice - prev_zprice
            etlaf_change = current_etlaf - prev_etlaf
            etlafprice_change = current_etlafprice - prev_etlafprice
            ezafe_change = current_ezafe - prev_ezafe
            eprice_change = current_eprice - prev_eprice

            mydict = {
                "name": _list['tek__name'] + ' ' + _list['tek__lname'],
                "id": _list['tek_id'],
                # مقادیر ماه جاری
                "current_etlaf": current_etlaf,
                "current_etlafprice": current_etlafprice,
                "current_zarib": current_zarib,
                "current_ezafe": current_ezafe,
                "current_zprice": current_zprice,
                "current_eprice": current_eprice,
                # مقادیر ماه قبل
                "prev_etlaf": prev_etlaf,
                "prev_etlafprice": prev_etlafprice,
                "prev_zarib": prev_zarib,
                "prev_ezafe": prev_ezafe,
                "prev_zprice": prev_zprice,
                "prev_eprice": prev_eprice,
                # تغییرات
                "etlaf_change": etlaf_change,
                "etlafprice_change": etlafprice_change,
                "zarib_change": zarib_change,
                "zprice_change": zprice_change,
                "ezafe_change": ezafe_change,
                "eprice_change": eprice_change,
                # سایر اطلاعات
                "period": period,
                "prev_period": prev_month.id if prev_month else None,
                "current_sum": current_zprice + current_eprice + current_etlafprice,
                "prev_sum": prev_zprice + prev_eprice + prev_etlafprice,
                "sum_change": (current_zprice + current_eprice + current_etlafprice) - (
                        prev_zprice + prev_eprice + prev_etlafprice),
                "accept": _list['accept'],
            }
            thislist.append(mydict)

        return TemplateResponse(request, 'pay/viewcommit.html',
                                {'zone': zones, 'mounts': mounts, 'thislist': thislist, 'period': int(period),
                                 })
    return TemplateResponse(request, 'pay/viewcommit.html',
                            {'zone': zones, 'mounts': mounts, 'period': 0})


@cache_permission('viewcommit')
def reportcommit(request):
    zones = Zone.objects_limit.all()
    add_to_log(request, f'گزارش تایید حقوق و دستمزد ', 0)
    thislist = []
    mounts = Mount.objects.all().order_by('id')
    sumzone = 0
    sumowner = 0
    sumlist = 0
    sumkasri = 0
    if request.method == 'POST':
        period = request.POST.get('period')
        zones = Zone.objects_limit.all().order_by('id')
        for zone in zones:
            sumzone += zone.tekcount
            owner = Owner.objects.filter(zone_id=zone.id, role__role='tek', active=True).count()
            sumowner += owner
            lists = Payroll.objects.filter(tek__role__role='tek', tek__zone_id=zone.id, paybaseparametrs_id=4,
                                           period_id=period).count()
            sumlist += lists
            accept = Payroll.objects.filter(tek__zone_id=zone.id, period_id=period).first()
            accept2 = Tektaeed.objects.filter(zone_id=zone.id, period_id=period).first()

            if accept:
                if accept.accept:
                    ac = 'تایید شده' + " " + str(accept.pdate())
                else:
                    ac = 'تایید نشده'
            else:
                ac = 'تایید نشده'
            if accept2:
                if accept2.accepttedad:
                    ac2 = '1'
                else:
                    ac2 = '0'
            else:
                ac2 = '0'

            if lists > 0:
                if lists == zone.tekcount:
                    mydict = {
                        "zone_id": zone.id,
                        "zone": zone.name,
                        "desc": "ثبت شده",
                        "status": 1,
                        "list": zone.tekcount,
                        "tedad": lists,
                        "ac": ac,
                        "ac2": ac2,
                        "isok": zone.tekcount,
                        "owner": owner,
                        "sabt": lists,
                        "kasri": zone.tekcount - lists,

                    }
                    thislist.append(mydict)
                if lists < zone.tekcount:
                    if lists < owner:
                        mydict = {
                            "zone_id": zone.id,
                            "zone": zone.name,
                            "desc": "برخی از تکنسین ها ثبت نشده",
                            "status": 2,
                            "list": zone.tekcount,
                            "tedad": lists,
                            "ac": ac,
                            "ac2": ac2,
                            "isok": zone.tekcount,
                            "owner": owner,
                            "sabt": lists,
                            "kasri": zone.tekcount - lists,
                        }
                        thislist.append(mydict)
                    else:
                        mydict = {
                            "zone_id": zone.id,
                            "zone": zone.name,
                            "desc": "ثبت شده",
                            "status": 1,
                            "list": zone.tekcount,
                            "tedad": lists,
                            "ac": ac,
                            "ac2": ac2,
                            "isok": zone.tekcount,
                            "owner": owner,
                            "sabt": lists,
                            "kasri": zone.tekcount - lists,
                        }
                        thislist.append(mydict)
                if lists > zone.tekcount:
                    mydict = {
                        "zone_id": zone.id,
                        "zone": zone.name,
                        "desc": "تعداد تکنسین های ثبت شده بیش از حد مجاز",
                        "status": 3,
                        "list": zone.tekcount,
                        "tedad": lists,
                        "ac": ac,
                        "ac2": ac2,
                        "isok": zone.tekcount,
                        "owner": owner,
                        "sabt": lists,
                        "kasri": zone.tekcount - lists,
                    }

                    thislist.append(mydict)
            else:
                mydict = {
                    "zone_id": zone.id,
                    "zone": zone.name,
                    "desc": "هیچ اطلاعاتی ثبت نشده",
                    "status": 4,
                    "list": zone.tekcount,
                    "tedad": lists,
                    "ac": ac,
                    "ac2": ac2,
                    "isok": zone.tekcount,
                    "owner": owner,
                    "sabt": lists,
                    "kasri": zone.tekcount - lists,
                }
                thislist.append(mydict)
            sumkasri += zone.tekcount - lists

            thislist = sorted(thislist, key=itemgetter('status', 'ac'), reverse=True)

        return TemplateResponse(request, 'pay/reportcommit.html',
                                {'zone': zones, 'mounts': mounts, 'sumkasri': sumkasri, 'thislist': thislist,
                                 'period': int(period), 'sumlist': sumlist,
                                 'sumowner': sumowner, 'sumzone': sumzone,
                                 })
    return TemplateResponse(request, 'pay/reportcommit.html',
                            {'zone': zones, 'mounts': mounts, 'period': 0})


@cache_permission('commitmgr')
def taeed(request):
    owner = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True)
    return TemplateResponse(request, 'taeed.html', {'owner': owner})


def accepttedad(request, _id):
    url = request.META.get('HTTP_REFERER')
    if _id == 1:
        Tektaeed.objects.create(zone_id=request.user.owner.zone_id, period_id=2, accepttedad=True)
        return redirect(HOME_PAGE)
    return redirect(url)


@cache_permission('commitmgr')
def fish(request, period, _id):
    add_to_log(request, 'مشاهده فیش حقوقی', 0)
    if request.user.owner.role.role in ['mgr', 'setad']:
        payparametrs = PayParametr.objects.filter(period_id=period, tek_id=_id).order_by('payitem__paybase_id')
        try:
            tekkarkerd = TekKarkard.objects.get(period_id=period, tek_id=_id)
        except TekKarkard.DoesNotExist:
            tekkarkerd = None
        paylist = Payroll.objects.filter(period_id=period, tek_id=_id, paybaseparametrs__sortable__gt=0).order_by(
            'paybaseparametrs__sortable', 'accepttedad')
        paysum = Payroll.objects.filter(period_id=period, tek_id=_id, paybaseparametrs__isshow=True).aggregate(
            sum=Sum('price'))
    else:
        tekkarkerd = None
        payparametrs = PayParametr.objects.filter(period_id=period, tek_id=_id,
                                                  tek__zone_id=request.user.owner.zone_id).order_by(
            'payitem__paybase_id')
        paylist = Payroll.objects.filter(period_id=period, tek_id=_id, tek__zone_id=request.user.owner.zone_id,
                                         paybaseparametrs__sortable__gt=0).order_by(
            'paybaseparametrs__sortable', 'accepttedad')
        paysum = Payroll.objects.filter(period_id=period, paybaseparametrs__isshow=True, tek_id=_id,
                                        tek__zone_id=request.user.owner.zone_id).aggregate(sum=Sum('price'))
    owner = Owner.objects.get(id=_id)
    return TemplateResponse(request, 'pay/fish.html',
                            {'tekkarkerd': tekkarkerd, 'paylist': paylist, 'paysum': paysum,
                             'owner': owner,
                             'payparametrs': payparametrs})


def accesptmgr(request, _id):
    url = request.META.get('HTTP_REFERER')
    mounth = Mount.objects.get(id=_id)
    if not mounth.active:
        messages.warning(request, 'ویرایش ماه مورد نظر بسته شده است')
        return redirect(url)
    results = Payroll.objects.filter(period_id=_id, tek__zone_id=request.user.owner.zone_id)
    for result in results:
        result.accept = True
        result.acceptupdate = datetime.datetime.today()
        result.save()
        add_to_log(request, 'تایید حقوق دوره ' + str(result.period.year) + str(result.period.mount) + ' منطقه ' + str(
            result.tek.name) + ' ' + str(result.tek.lname), 0)
        messages.success(request, 'با موفقیت انجام شد')
    return redirect(url)


def accesptmgr2(request, _id, _zone):
    url = request.META.get('HTTP_REFERER')
    mounth = Mount.objects.get(id=_id)
    if not mounth.active:
        messages.warning(request, 'ویرایش ماه مورد نظر بسته شده است')
        return redirect(url)
    results = Payroll.objects.filter(period_id=_id, tek__zone_id=_zone)
    for result in results:
        result.accept = True
        result.acceptupdate = datetime.datetime.today()
        result.save()
        add_to_log(request, 'تایید حقوق دوره ' + str(result.period.year) + str(result.period.mount) + ' منطقه ' + str(
            result.tek.name) + ' ' + str(result.tek.lname), 0)
        messages.success(request, 'با موفقیت انجام شد')
    return redirect(url)


def accesptmgr3(request, _id, _zone):
    url = request.META.get('HTTP_REFERER')
    mounth = Mount.objects.get(id=_id)
    if not mounth.active:
        messages.warning(request, 'ویرایش ماه مورد نظر بسته شده است')
        return redirect(url)
    results = Payroll.objects.filter(period_id=_id, tek__zone_id=_zone)
    for result in results:
        result.accept = False
        result.acceptupdate = None
        result.save()
        add_to_log(request,
                   'باز کردن  حقوق دوره ' + str(result.period.year) + str(result.period.mount) + ' منطقه ' + str(
                       result.tek.name) + ' ' + str(result.tek.lname), 0)
        messages.success(request, 'با موفقیت انجام شد')
    return redirect(url)


@cache_permission('kargahlist')
def liststorekargah(request):
    add_to_log(request, f'گزارش قطعات موجود در کارگاه ', 0)
    _list = None
    master = 1
    storage = 0
    _resid = 0
    _listok = 0
    _send = 0
    _listfail = 0
    storages = Storage.objects.all
    if request.method == 'POST':
        master = int(request.POST.get('master'))
        if request.user.owner.role.role in ['mgr', 'setad']:
            storage = int(request.POST.get('storage'))
            storage_id = Storage.objects.get(id=int(storage)).zone_id
        else:
            storage_id = request.user.owner.zone_id
        this_date = datetime.datetime.today()
        _list = StoreList.objects.filter(status_id=8, statusstore_id=int(master), zone_id=int(storage_id))
        _listfail = StoreList.objects.filter(status_id=8, zone_id=int(storage_id)).aggregate(
            master=(Count(Case(When(statusstore_id=1, then=1)))),
            pinpad=(Count(Case(When(statusstore_id=2, then=1)))))
        _listok = StoreList.objects.filter(status_id=3, zone_id=int(storage_id)).aggregate(
            master=(Count(Case(When(statusstore_id=1, then=1)))),
            pinpad=(Count(Case(When(statusstore_id=2, then=1)))))

        _send = Store.objects.filter(marsole_date__year=this_date.year, marsole_date__month=this_date.month,
                                     marsole_date__day=this_date.day, zone_id=int(storage_id)).aggregate(
            master=Count('master'), pinpad=Count('pinpad'))
        _resid = Store.objects.filter(status_id=8, resid_date__year=this_date.year, resid_date__month=this_date.month,
                                      resid_date__day=this_date.day, zone_id=int(storage_id)).aggregate(
            master=Count('master'), pinpad=Count('pinpad'))

    return TemplateResponse(request, 'store/liststorekargah.html',
                            {'today': today, 'storages': storages, 'storage': storage,
                             'listok': _listok, 'send': _send,
                             'master': int(master), 'resid': _resid, 'listfail': _listfail,
                             'list': _list})


@cache_permission('mojodidaghi')
def mojodidaghikol(request):
    _list = []
    _listfail = 0
    _listfail = StoreList.objects.values('zone_id', 'zone__name').filter(zone__storage=True).annotate(
        master_notok=(Count(Case(When(status_id=8, statusstore_id=1, then=1)))),
        pinpad_notok=(Count(Case(When(status_id=8, statusstore_id=2, then=1)))),
        master_ok=(Count(Case(When(status_id=3, statusstore_id=1, then=1)))),
        pinpad_ok=(Count(Case(When(status_id=3, statusstore_id=2, then=1)))),
    )
    if request.user.owner.role.role in ['zone', 'tek']:
        _listfail = _listfail.filter(zone_id=request.user.owner.zone_id)

    return TemplateResponse(request, 'store/tajamoe_daghi.html',
                            {'listfail': _listfail})


@cache_permission('takhsistop')
def sendzone(request):
    storages = None
    zonelist = None
    url = request.META.get('HTTP_REFERER')
    if request.user.owner.role.role == 'zone':
        zonelist = Zone.objects_limit.all()
        storages = Storage.objects.filter(zone_id=request.user.owner.zone_id)

    elif request.user.owner.role.role in ['setad']:
        if request.user.owner.refrence_id == 13:
            storages = Storage.objects.all().order_by('sortid')
            zonelist = Zone.objects.all()
    elif request.user.owner.role.role == 'mgr':
        storages = Storage.objects.all().order_by('sortid')
        zonelist = Zone.objects.all()

    if request.method == 'POST':
        master = int(request.POST.get('master'))
        zone = int(request.POST.get('zone'))
        pinpad = request.POST.get('pinpad')
        storage = int(request.POST.get('storage'))
        if request.user.owner.role.role == 'zone':
            _strg = Storage.objects.get(id=storage)
            if _strg.zone_id != request.user.owner.zone_id:
                messages.error(request, 'دسترسی غیر مجاز !!!')
                return redirect('base:home')

        select = request.POST.get('select')
        _priority = int(request.POST.get('priority'))
        az = select
        select = select.split("/")
        select = jdatetime.date(day=int(select[2]), month=int(select[1]), year=int(select[0])).togregorian()
        mystore = Store.objects.create(owner_id=request.user.owner.id, tarikh=select, pinpad=pinpad, master=master,
                                       status_id=1, priority=_priority,
                                       zone_id=zone, storage_id=storage)
        mystore.create = f'{select} 11:11:11'
        mystore.save()
        add_to_log(request, 'ثبت تخصیص قطعه برای منطقه ' + str(mystore.zone.name), 0)
        HistorySt.objects.create(store_id=mystore.id, owner_id=request.user.owner.id,
                                 status_id=1, description=f' تخصیص به منطقه  {mystore.zone.name} ')
        messages.success(request, SUCCESS_MSG)
        return redirect(url)

    else:
        az = today
    return TemplateResponse(request, 'store/sendzone.html',
                            {'zone': zonelist, 'today': today, 'storages': storages, 'az': az})


@cache_permission('takhsistop')
def updatezone(request, _id):
    mystore = Store.objects.get(id=_id)
    zonelist = Zone.objects.all()
    storages = Storage.objects.all
    _priority = 0

    if request.method == 'POST':
        master = int(request.POST.get('master'))
        zone = int(request.POST.get('zone'))
        pinpad = request.POST.get('pinpad')
        storage = int(request.POST.get('storage'))
        _priority = int(request.POST.get('priority'))

        mystore.pinpad = pinpad
        mystore.master = master
        mystore.priority = _priority
        mystore.zone_id = zone
        mystore.storage_id = storage
        mystore.save()
        add_to_log(request, 'ویرایش قطعه منطقه ' + str(mystore.zone.name), 0)
        messages.success(request, 'با موفقیت ویرایش شد')
        return redirect('pay:sendlist')
    return TemplateResponse(request, 'store/updatezone.html',
                            {'zone': zonelist, 'today': today, 'storages': storages,
                             'priority': int(_priority),
                             'mystore': mystore})


def deletezone(request, _id):
    url = request.META.get('HTTP_REFERER')
    Store.objects.get(id=_id).delete()
    add_to_log(request, ' حذف قطعه تخصیص داده شده ' + str(_id), 0)
    messages.info(request, 'درخواست قطعه بدرستی حذف شد')
    return redirect(url)


@cache_permission('takhsistop')
def sendlist(request):
    _list = None
    add_to_log(request, f'مشاهده فرم  تخصیصی قطعه ', 0)
    list2 = None
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))

    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")

        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"
    if request.user.owner.role.role == 'mgr':
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout, status_id__in=[1, 9],
                                     storage_id=request.user.owner.defaultstorage).order_by('-priority', 'tarikh')
    if request.user.owner.role.role == 'setad':
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout,
                                     storage_id=request.user.owner.defaultstorage, status_id__in=[1, 9]).order_by(
            '-priority', 'tarikh')

        list2 = Store.objects.filter(create__gte=datein, create__lte=dateout, status_id__in=[1, 9]).order_by(
            '-priority', 'tarikh')

    if request.user.owner.role.role == 'zone':
        storageid = Storage.objects.get(zone_id=request.user.owner.zone_id).id
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout,
                                     storage_id=storageid,
                                     status_id__in=[1, 9]).order_by('-priority', '-tarikh')
    _filter = StoreFilters(request.GET, queryset=_list)
    if request.user.owner.role.role == 'setad' and _filter.data:
        if 'page' in request.GET:
            match request.GET['page'][:5]:
                case 'allti':
                    _filter = StoreFilters(request.GET, queryset=_list)
                case 'previ':
                    _filter = StoreFilters(request.GET, queryset=_list)
                case 'roday':
                    _filter = StoreFilters(request.GET, queryset=_list)
                case 'next':
                    _filter = StoreFilters(request.GET, queryset=_list)
        else:
            _filter = StoreFilters(request.GET, queryset=list2)

    _list = _filter.qs
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']
    paginator = Paginator(_list, 37)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list.filter(), 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']

        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, SENDLIST_PAGE,
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'page_num': page_num,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('takhsistop')
def sendmarsole2(request):
    _list = None
    add_to_log(request, f'مشاهده فرم لیست قطعات ارسال شده ', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")

        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"
    if request.user.owner.role.role == 'mgr':
        _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout).order_by('-id')
    if request.user.owner.role.role == 'setad':
        _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, storage__zone=None).order_by(
            '-id')

    if request.user.owner.role.role == 'zone':
        _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout,
                                     storage__zone=request.user.owner.zone_id
                                     ).order_by('-id')
    _filter = StoreFilters(request.GET, queryset=_list)
    _list = _filter.qs
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']
    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']

        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, SENDLIST_PAGE,
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'page_num': page_num,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('takhsistop')
def sendresid(request):
    _list = None
    add_to_log(request, f'مشاهده فرم قطعات رسید شده ', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")

        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"
    if request.user.owner.role.role == 'mgr':
        _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout, status_id=3).order_by('-id')
    if request.user.owner.role.role == 'setad':
        _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout, storage__zone=None,
                                     status_id=3).order_by('-id')

    if request.user.owner.role.role == 'zone':
        _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout,
                                     storage__zone=request.user.owner.zone_id,
                                     status_id=3).order_by('-id')
    _filter = StoreFilters(request.GET, queryset=_list)
    _list = _filter.qs
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']
    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(resid_date__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(resid_date__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(resid_date__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']

        if 'previews' in page_num:
            paginator = Paginator(_list.filter(resid_date__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(resid_date__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(resid_date__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(resid_date__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(resid_date__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(resid_date__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, SENDLIST_PAGE,
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'page_num': page_num,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('takhsistop')
def sendall(request):
    _list = None
    _filter = ""
    add_to_log(request, f'مشاهده فرم لیست تخصیصی کل ', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"
    else:

        datein = datein.split("/")
        dateout = dateout.split("/")

        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"

    if request.user.owner.role.role in ['mgr', 'setad']:
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout).order_by('-id')
        _filter = StoreFilters(request.GET, queryset=_list)
        if _filter.data:
            _list = _filter.qs
        else:
            _list = _list.filter(storage_id=request.user.owner.defaultstorage)

    if request.user.owner.role.role == 'zone':
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout,
                                     storage__zone=request.user.owner.zone_id).order_by('-id')
        _filter = StoreFilters(request.GET, queryset=_list)
        _list = _filter.qs

    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']
    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']

        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, SENDLIST_PAGE,
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'page_num': page_num,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('ersaltozon')
def poshtibanlist(request):
    add_to_log(request, f'مشاهده فرم ارسال قطعه به منطقه ', 0)
    posts = Post.objects.all()
    _storageid = Storage.objects.get(zone_id=request.user.owner.zone_id).id
    lists = Store.objects.filter(status_id__in=[1, 9, 14], storage_id=_storageid).order_by('-priority', 'tarikh')
    _filter = StoreFilters(request.GET, queryset=lists)
    if _filter.data:
        lists = Store.objects.filter(status_id__in=[1, 9, 14]).order_by('-priority', 'tarikh')
        _filter = StoreFilters(request.GET, queryset=lists)
    lists = _filter.qs

    return TemplateResponse(request, 'store/poshtibanlist.html',
                            {'list': lists, 'posts': posts, 'filter': _filter})


@cache_permission('residinzon')
def residzone(request):
    _list = None
    _liststorage = None
    add_to_log(request, f'مشاهده فرم رسد دریافت قطعات ', 0)
    if request.user.owner.role.role in 'zone,posht':
        _list = Store.objects.filter(status_id=2, zone_id=request.user.owner.zone_id)
    if request.user.owner.zone.storage:
        _liststorage = Store.objects.filter(status_id=7, storage__zone_id=request.user.owner.zone_id).order_by('-id')

    list2 = Store.objects.filter(status_id__in=[1, 9], zone_id=request.user.owner.zone_id)
    return TemplateResponse(request, 'store/residzone.html',
                            {'list2': list2, 'list': _list, 'liststorage': _liststorage})


@cache_permission('ersaltozon')
def addstore(request, _id, st):
    _list = Store.objects.get(id=_id)
    edit = 0
    tedad = StoreHistory.objects.filter(baseroot=_list.id, store__statusstore_id=st)

    return TemplateResponse(request, 'store/AddStore.html',
                            {'list': _list, 'master': range(1, _list.master + 1),
                             'pinpad': range(1, _list.pinpad + 1), 'st': st, 'id': _list.id, 'today': today,
                             'edit': edit,
                             'tedad': tedad})


@cache_permission('changestor')
def addmystore(request):
    _list = []
    _st = 1
    edit = 0
    tedad = 0
    if request.method == 'POST':
        _st = request.POST.get('st')
        edit = 1
        _list = StoreList.objects.filter(owner_id=request.user.id, status_id=3, statusstore_id=int(_st))
        tedad = StoreList.objects.filter(owner_id=request.user.id, status_id=3, statusstore_id=int(_st)).count()

    return TemplateResponse(request, 'store/addmystore.html',
                            {'list': _list, 'st': _st, 'tedad': tedad,
                             'today': today, 'edit': edit
                             })


def check_serial(serial):
    serial = serial.replace('۰', '0')
    serial = serial.replace('۱', '1')
    serial = serial.replace('۲', '2')
    serial = serial.replace('۳', '3')
    serial = serial.replace('۴', '4')
    serial = serial.replace('۵', '5')
    serial = serial.replace('۶', '6')
    serial = serial.replace('۷', '7')
    serial = serial.replace('۸', '8')
    serial = serial.replace('۹', '9')
    if len(serial) < 10:
        return False
    if len(serial) > 12:
        return False
    if serial.isnumeric():
        return serial
    else:
        return False


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def checkserial(request):
    _list = None
    store = ""
    sth = ""
    if request.method == 'POST':
        serial = request.POST.get('serial')
        serial = check_serial(serial)
        if not serial:
            return JsonResponse({"message": 'error'})

        _id = request.POST.get('id')
        st = request.POST.get('st')

        _id = int(_id)
        if _id != 0:
            _list = Store.objects.get(id=_id)
        updatestore = StoreList.objects.filter(serial=serial)

        _count = updatestore.count()
        _storage = Storage.objects.get(zone_id=request.user.owner.zone.id)
        _storageid = _storage.id
        if _count > 0:
            _st = updatestore.last()
            _st = _st.statusstore_id

            if int(st) != int(_st):
                if int(_st) == 1:
                    return JsonResponse({"message": 'level2',
                                         'payam': 'نوع قطعه متفاوت است ، این قطعه بعنوان کارتخوان در سیستم ثبت شده است '
                                         })
                else:
                    return JsonResponse({"message": 'level2',
                                         'payam': 'نوع قطعه متفاوت است ، این قطعه بعنوان صفحه کلید در سیستم ثبت شده است'
                                         })

            store = StoreList.objects.get(serial=serial)

            if _storage.level == 2 and store.status_id == 9:
                return JsonResponse({"message": 'level2',
                                     'payam': f"قطعه در منطقه {store.zone.name} و وضعیت {store.status.name} میباشد"})
            if _storage.level == 3 and store.status_id != 8:
                if store.zone_id:
                    return JsonResponse({"message": 'level3',
                                         'payam': f"قطعه در منطقه {store.zone.name} و وضعیت {store.status.name} میباشد"
                                         })
                else:
                    return JsonResponse({"message": 'level3',
                                         'payam': f"ابتدا ین قطعه را رسید کنید "})

            if _storage.level == 4 and store.status_id == 8 and store.zone_id != request.user.owner.zone_id:
                return JsonResponse({"message": 'level3',
                                     'payam': f"قطعه در منطقه {store.zone.name} و وضعیت {store.status.name} میباشد"})
            if _storage.level == 6:
                if store.status_id != 8 or store.zone_id != request.user.owner.zone_id:
                    return JsonResponse({"message": 'level3',
                                         'payam': f"قطعه در منطقه {store.zone.name} و وضعیت {store.status.name} میباشد"
                                         })

            if _storage.level == 7:
                if store.zone_id != request.user.owner.zone_id:
                    return JsonResponse({"message": 'level3',
                                         'payam': f"قطعه در منطقه {store.zone.name} و وضعیت {store.status.name} میباشد"
                                         })

            if _storage.level == 5 and store.status_id == 9:
                return JsonResponse({"message": 'level2',
                                     'payam': f"قطعه در منطقه {store.zone.name} و وضعیت {store.status.name} میباشد "})

            if store.status_id != 14:
                if _storage.level == 5 and store.status_id != 13:
                    return JsonResponse({"message": 'level3',
                                         'payam': f"این قطعه توسط ناظر پخش تایید نشده "})

            if store.status_id == 14 and store.status_id != 8:
                return JsonResponse({"message": 'level3',
                                     'payam': f"وضعیت این قطعه در کارگاه شما نیست "})

            store.status_id = 9 if _id != 0 else 3
            store.zone_id = _list.zone_id if _id != 0 else request.user.owner.zone.id
            store.owner_id = request.user.id
            if _id != 0:
                store.store_id = _list.id
            store.save()

            if _id != 0:
                Store.add_or_remove_store(int(st), True, _list.id)
            if _id != 0:
                sth = StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id, baseroot=_list.id,
                                                  information="ارسال قطعه از کارگاه " + str(
                                                      _storage.name),
                                                  status_id=9, description=f' تخصیص به منطقه  {store.zone.name}',
                                                  storage_id=_storageid)
            else:
                sth = StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id, baseroot=0,
                                                  information="ارسال قطعه از کارگاه " + str(
                                                      _storage.name),
                                                  status_id=9,
                                                  description=f' تخصیص به منطقه  {request.user.owner.zone.name}',
                                                  storage_id=_storageid)

                sth = StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id, baseroot=0,
                                                  information="ارسال قطعه از کارگاه " + str(
                                                      _storage.name),
                                                  status_id=3,
                                                  description=f' رسید به منطقه  {request.user.owner.zone.name}',
                                                  storage_id=_storageid)
            message = 'success'
        else:
            if _storage.level == 5:
                return JsonResponse({"message": 'level3',
                                     'payam': f"این قطعه توسط ناظر پخش تایید نشده "})
            if _storage.level == 3:
                return JsonResponse({"message": 'level3',
                                     'payam': f"ابتدا ین قطعه را رسید کنید "})
            if _storage.level == 4:
                return JsonResponse({"message": 'level3',
                                     'payam': f"این قطعه موجود نیست  "})
            if _storage.level == 6:
                return JsonResponse({"message": 'level3',
                                     'payam': f"این قطعه موجود نیست  "})
            if _storage.level == 7:
                return JsonResponse({"message": 'level3',
                                     'payam': f"این قطعه موجود نیست  "})
            if _id != 0:
                mystore = StoreList.objects.create(zone_id=_list.zone_id, serial=serial, owner_id=request.user.id,
                                                   statusstore_id=st, store_id=_list.id,
                                                   status_id=9, uniq=str(serial) + '-' + str(_list.status_id))
                Store.add_or_remove_store(int(st), True, _list.id)
                sth = StoreHistory.objects.create(store_id=mystore.id, owner_id=request.user.owner.id,
                                                  baseroot=_list.id,
                                                  information="ارسال قطعه از " + str(_storage.name),
                                                  status_id=9, description=f' تخصیص به منطقه  {mystore.zone.name}',
                                                  storage_id=_storageid)

            message = 'success'
        if _id != 0:
            tedad = StoreHistory.objects.filter(baseroot=_list.id, store__statusstore_id=st).count()
        else:
            tedad = StoreList.objects.filter(owner_id=request.user.id, status_id=3,
                                             statusstore_id=int(store.statusstore_id)).count()
        mylist = []
        newid = store.id if _id == 0 else sth.id
        _dict = {
            'serial': sth.store.serial,
            'id': newid,
        }
        mylist.append(_dict)

        return JsonResponse({"message": message, "mylist": mylist, 'tedad': tedad})


def removeserial(request):
    if request.method == 'POST':
        _id = request.POST.get('id')
        store = StoreHistory.objects.get(id=_id)
        base = store.baseroot
        store2 = StoreList.objects.get(id=store.store_id)
        _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
        if _storage.level == 5:
            if store2.store.status_id == 14:
                store2.status_id = 8
            else:
                store2.status_id = 13

        else:
            store2.status_id = 8
        store2.zone_id = request.user.owner.zone_id
        store2.save()
        StoreHistory.objects.get(id=_id).delete()
        Store.add_or_remove_store(store2.statusstore_id, False, store.baseroot)
        tedad = StoreHistory.objects.filter(baseroot=base, store__statusstore_id=store2.statusstore_id).count()
        add_to_log(request, 'حذف قطعه به شماره سریال ' + str(store2.serial), 0)

        return JsonResponse({"message": "success", 'tedad': tedad})


def removeserial2(request):
    if request.method == 'POST':
        _id = request.POST.get('id')

        store2 = StoreList.objects.get(id=_id)
        _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
        if _storage.level == 5:
            if store2.store.status_id == 14:
                store2.status_id = 8
            else:
                store2.status_id = 13

        else:
            store2.status_id = 8
        store2.zone_id = request.user.owner.zone_id
        store2.save()

        tedad = StoreList.objects.filter(owner_id=request.user.id, status_id=3,
                                         statusstore_id=int(store2.statusstore_id)).count()
        sthis = StoreHistory.objects.filter(store_id=store2.id).last()
        sthis.delete()
        add_to_log(request, 'حذف قطعه به شماره سریال ' + str(store2.serial), 0)

        return JsonResponse({"message": "success", 'tedad': tedad})


@transaction.atomic
def sendmarsole(request):
    if request.method == 'POST':
        code = request.POST.get('marsole')
        post = int(request.POST.get('post'))
        _id = int(request.POST.get('myid'))
        _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
        result = Store.objects.get(id=_id)
        masternumber = StoreHistory.objects.filter(baseroot=result.id, store__statusstore_id=1).count()
        pinpadnumber = StoreHistory.objects.filter(baseroot=result.id, store__statusstore_id=2).count()
        if settings.SEND_MARSOLE_STATUS == 1:
            if result.master != masternumber or result.pinpad != pinpadnumber:
                messages.error(request, 'تعداد سریال ارسالی با مقدار تخصیصی مغایرت دارد')
                return redirect('pay:PoshtibanList')
        _storest = 3
        if settings.SEND_MARSOLE_STATUS == 2:
            if result.pinpad != 0 and result.master == masternumber and pinpadnumber == 0:
                _storest = 2
                mystore = Store.objects.create(owner_id=request.user.owner.id, tarikh=result.tarikh,
                                               pinpad=result.pinpad,
                                               master=0,
                                               status_id=1, priority=result.priority,
                                               zone_id=result.zone_id, storage_id=result.storage_id)
                add_to_log(request, 'ثبت تخصیص سیستمی قطعه بعلت کسری موجودی برای منطقه ' + str(mystore.zone.name), 0)
                HistorySt.objects.create(store_id=mystore.id, owner_id=request.user.owner.id,
                                         status_id=1, description=f' تخصیص به منطقه  {mystore.zone.name} ')
            elif result.master != 0 and result.pinpad == pinpadnumber and masternumber == 0:
                _storest = 1
                mystore = Store.objects.create(owner_id=request.user.owner.id, tarikh=result.tarikh,
                                               pinpad=0,
                                               master=result.master,
                                               status_id=1, priority=result.priority,
                                               zone_id=result.zone_id, storage_id=result.storage_id)
                add_to_log(request, 'ثبت تخصیص سیستمی قطعه بعلت کسری موجودی برای منطقه ' + str(mystore.zone.name), 0)
                HistorySt.objects.create(store_id=mystore.id, owner_id=request.user.owner.id,
                                         status_id=1, description=f' تخصیص به منطقه  {mystore.zone.name} ')
            elif result.master != masternumber or result.pinpad != pinpadnumber:
                messages.error(request, 'تعداد سریال ارسالی با مقدار تخصیصی مغایرت دارد')
                return redirect('pay:PoshtibanList')
        result.post_id = post
        result.marsole = code
        result.marsole_date = datetime.datetime.now()
        if result.status_id == 14:
            nuewstatus = 7
            storage = Storage.objects.get(zone_id=result.zone_id)
            result.storage_id = storage.id
            result.zone_id = request.user.owner.zone_id
        else:
            nuewstatus = 2
        result.status_id = nuewstatus
        if _storest == 1:
            result.master = 0
        elif _storest == 2:
            result.pinpad = 0
        result.save()
        storelist = StoreList.objects.filter(zone_id=result.zone_id, status_id__in=[1, 9], store_id=_id)
        for item in storelist:
            item.status_id = nuewstatus

            item.save()
            StoreHistory.objects.create(store_id=item.id, owner_id=request.user.owner.id, baseroot=0,
                                        information="ارسال قطعه از " + str(_storage.name),
                                        status_id=nuewstatus,
                                        description=f'  به پست برای کارگاه /  منطقه  {item.zone.name} ')
        messages.success(request, SUCCESS_MSG)
        add_to_log(request, 'ارسال قطعه و ثبت مرسوله به شماره  ' + str(code), 0)
        HistorySt.objects.create(store_id=result.id, owner_id=request.user.owner.id,
                                 status_id=nuewstatus, description=f' تخصیص به کارگاه / منطقه  {result.zone.name} ')

        _mobail = Owner.objects.exclude(codemeli='2161846736').filter(zone_id=result.zone.id, refrence_id=1,
                                                                      active=True).last()
        if _mobail:
            mobail = _mobail.mobail
            if len(code) > 3:
                _message = f'قطعات  شما به مقصد منطقه ارسال شد. سامانه هوشمند سوخت ecourier.mahex.com/tr/{str(code)}'
            else:
                _message = f'قطعات درخواستی شما به مقصد منطقه ارسال شد .   مدیریت سامانه هوشمند سوخت '

            try:
                SendSmS(
                    mobail, _message
                )
            except:
                pass

    return redirect('pay:PoshtibanList')


@transaction.atomic
def residmarsole(request):
    if request.method == 'POST':
        _id = request.POST.get('myid')
        result = Store.objects.get(id=_id)
        result.resid_date = datetime.datetime.now()
        result.resid_year = jdatetime.datetime.now().year
        if len(str(jdatetime.datetime.now().month)) == 1:
            month = '0' + str(jdatetime.datetime.now().month)
        else:
            month = jdatetime.datetime.now().month
        result.resid_month = month
        if len(str(jdatetime.datetime.now().day)) == 1:
            day = '0' + str(jdatetime.datetime.now().day)
        else:
            day = jdatetime.datetime.now().day
        result.resid_day = day
        stresult = result.status_id
        if result.status_id == 7:
            result.status_id = 8
        else:
            result.status_id = 3
        result.save()

        storelist = StoreHistory.objects.filter(baseroot=_id)
        for item in storelist:
            storeupdate = StoreList.objects.get(id=item.store_id)
            if stresult == 7:
                storeupdate.status_id = 8
                # storeupdate.zone_id = result.zone_id
                storeupdate.save()
                StoreHistory.objects.create(store_id=item.store_id, owner_id=request.user.owner.id,
                                            information="رسید  قطعه به کارگاه ",
                                            status_id=8,
                                            description=f' رسید  به کارگاه  {request.user.owner.zone.name} ')

            else:
                if stresult != 8:
                    storeupdate.status_id = 3
                    storeupdate.zone_id = result.zone_id
                    storeupdate.save()
                    StoreHistory.objects.create(store_id=item.store_id, owner_id=request.user.owner.id,
                                                information="رسید  قطعه به منطقه ",
                                                status_id=3, description=f' رسید  به منطقه  {result.zone.name} ')
            HistorySt.objects.create(store_id=result.id, owner_id=request.user.owner.id,
                                     status_id=3, description=f' رسید به منطقه  {result.zone.name} ')
            add_to_log(request, 'رسید قطعه به منطقه ' + str(result.zone.name), 0)
    return redirect('pay:residZone')


def getmasterlist(request):
    add_to_log(request, f'مشاهده لیست لاگ قطعات ', 0)
    thislist = []
    if request.method == 'POST':
        _id = int(request.POST.get('obj'))
        st = int(request.POST.get('st'))
        _store = Store.objects.get(id=_id)
        lists = StoreHistory.objects.filter(store__statusstore_id=st,
                                            baseroot=_id)
        _todaydate = date.today()
        si_ago_sell = _todaydate.today() - datetime.timedelta(days=30)
        for _list in lists:

            risk = StoreHistory.objects.filter(store_id=_list.store_id, create__gte=si_ago_sell,
                                               create__lte=_todaydate.today(),
                                               status_id=6).count()
            if risk == 0:
                level = 0
            elif risk == 1:
                level = 1
            elif risk == 2:
                level = 2
            else:
                level = 3
            mydict = {
                "serial": _list.store.serial,
                "status": _list.store.status.name,
                "level": level,
            }
            thislist.append(mydict)
        return JsonResponse({'message': 'success', 'list': thislist})


def getmasterlisttek(request):
    lists = None
    thislist = []
    if request.method == 'POST':
        _id = int(request.POST.get('obj'))
        st = int(request.POST.get('st'))
        nomrator = int(request.POST.get('nomrator'))

        if st in [1, 2]:
            lists = StoreList.objects.filter(statusstore_id=st,
                                             getuser_id=_id, status_id=nomrator)
            for _list in lists:
                mydict = {
                    "serial": _list.serial,
                    "status": _list.status.name,
                    "tarikh": _list.normal_datetime(),
                }
                thislist.append(mydict)
        elif st == 3:
            lists = LockModel.objects.filter(status_id=nomrator, owner_id=_id)
        elif st == 5:
            user = User.objects.get(owner=_id)
            lists = LockModel.objects.filter(status_id=nomrator, idg_user_id=user.id, )
        elif st == 6:
            lists = LockModel.objects.filter(status_id=5, sendposhtiban_id=_id)
        elif st == 7:
            lists = LockModel.objects.filter(status_id__in=[6, 9, 10, 11], sendposhtiban_id=_id)
        elif st == 8:
            lists = LockModel.objects.filter(status_id=5, insertlock_id=_id)
        elif st == 9:
            lists = LockModel.objects.filter(status_id__in=[6, 9, 10, 11], insertlock_id=_id)
        elif st == 10:
            result = InsertLock.objects.get(id=_id)
            i = 0
            _diclist = []
            for i in range(result.tedad):
                _diclist.append(str(result.seri.info) + str(result.serial_in + i))
                lists = None
            _list = LockModel.objects.filter(insertlock_id=_id, status_id__in=[6, 9, 10, 11, 5])
            missing_serials = [serial for serial in _diclist if not _list.filter(serial=serial).exists()]
            lists = []
            for serial in missing_serials:
                try:
                    _ = LockModel.objects.get(serial=serial)
                    lists.append({
                        "serial": serial,
                        "status": _.status.info
                    })
                except:
                    lists.append({
                        "serial": serial,
                        "status": 'نامعلوم'
                    })
        elif st == 11:
            if request.user.owner.role.role == 'engin':
                _role = 'engin'
            else:
                _role = 'smart'

            result = InsertLock.objects.filter(zone_id=_id, peymankar__ename=_role)
            _diclist = []
            for item in result:
                for i in range(item.tedad):
                    _diclist.append(str(item.seri.info) + str(item.serial_in + i))
            _list = LockModel.objects.filter(zone_id=_id, input_date_unit__gt='2023-01-01', ename=_role)
            missing_serials = [serial for serial in _diclist if not _list.filter(serial=serial).exists()]
            lists = []
            for serial in missing_serials:
                try:
                    _ = LockModel.objects.get(serial=serial)
                    lists.append({
                        "serial": serial,
                        "status": _.status.info
                    })
                except:
                    lists.append({
                        "serial": serial,
                        "status": 'نامعلوم'
                    })

        if st in [3, 5, 6, 7, 8, 9]:
            for _list in lists:
                mydict = {
                    "serial": _list.serial,
                    "status": "",
                    "tarikh": "",
                }
                thislist.append(mydict)
        if st in [10, 11]:
            for _list in lists:
                mydict = {
                    "serial": _list['serial'],
                    "status": _list['status'],
                    "tarikh": "",
                }
                thislist.append(mydict)

        return JsonResponse({'message': 'success', 'list': thislist})


def getstorelistzone(request):
    thislist = []
    if request.method == 'POST':
        _id = int(request.POST.get('obj'))
        st = int(request.POST.get('st'))

        nomrator = int(request.POST.get('nomrator'))

        _st = []
        if nomrator == 1:
            _st = [3, 4, 16]
        elif nomrator == 2:
            _st = [6, 10, 11, 8]
        elif nomrator == 3:
            _st = [2]
        # elif nomrator == 3:
        #     _st = [1, 9]
        elif nomrator == 5:
            _st = [8]
        elif nomrator == 6:
            _st = [6, 10, 11]
        elif nomrator == 7:
            _st = [3]

        if nomrator == 3:
            lists = StoreList.objects.filter(statusstore_id=int(st), store__status_id=2,
                                             zone_id=_id, status_id__in=_st)
        else:
            lists = StoreList.objects.filter(statusstore_id=int(st),
                                             zone_id=_id, status_id__in=_st)
        for _list in lists:
            try:
                _user = _list.getuser.name + " " + _list.getuser.lname if _list.getuser.name else ""
            except:
                _user = ""
            _info = _user if _list.status_id in [4, 6] else ""

            mydict = {
                "serial": _list.serial,
                "status": _list.status.name + " " + _info,
                "tarikh": _list.normal_datetime(),
                "update": _list.update
            }
            thislist.append(mydict)
        _list = sorted(thislist, key=itemgetter('update'), reverse=True)
        return JsonResponse({'message': 'success', 'list': _list})


@cache_permission('acceptpay')
def payparametr(request):
    templatepage = 'pay/payparametr.html'
    mounts = Mount.objects.all()
    zone = Zone.objects_limit.all()

    if request.method == 'POST':
        period_id = request.POST.get('period')

        for item in zone:
            try:
                count_ticket = Ticket.objects.filter(gs__area__zone_id=item.id, status_id=1,
                                                     failure__failurecategory_id__in=[1010, 1011]).count()
                count_ticket += Ticket.objects.filter(gs__area__zone_id=item.id, status_id=1,
                                                      failure_id=1045).count()
                tickets = count_ticket
                nazels = Pump.objects.filter(gs__area__zone_id=item.id, actived=True, gs__active=True).count()
                rotbe = (count_ticket / nazels) * 100
                PayDarsadMah.objects.create(period_id=period_id, zone_id=item.id, zaribfani=0, tickets=tickets,
                                            nazels=nazels, rotbe=rotbe,
                                            zaribetlaf=0,
                                            zaribbahrevari=0, uniq=str(period_id) + "-" + str(item.id))
            except IntegrityError:
                continue

        pay = PayDarsadMah.objects.filter(period_id=period_id).order_by('-rotbe')
        active = PayDarsadMah.objects.filter(period_id=period_id).first()
        active = active.active

        return TemplateResponse(request, templatepage,
                                {'zone': zone, 'mounts': mounts, 'pay': pay, 'active': active,
                                 'postperiod': int(period_id)})
    return TemplateResponse(request, templatepage, {'zone': zone, 'mounts': mounts})


def updateparametr(request, postperiod):
    url = request.META.get('HTTP_REFERER')
    lists = PayDarsadMah.objects.filter(period_id=postperiod)
    for item in lists:
        count_ticket = Ticket.objects.filter(gs__area__zone_id=item.zone_id, status_id=1,
                                             failure__failurecategory_id__in=[1010, 1011]).count()
        count_ticket += Ticket.objects.filter(gs__area__zone_id=item.zone_id, status_id=1,
                                              failure_id=1045).count()
        tickets = count_ticket
        nazels = Pump.objects.filter(gs__area__zone_id=item.zone_id, actived=True, gs__active=True).count()
        rotbe = (count_ticket / nazels) * 100
        item.tickets = tickets
        item.nazels = nazels
        item.rotbe = rotbe

        item.save()
    return redirect(url)


@transaction.atomic
@cache_permission('acceptpay')
def payparametrsave(request):
    url = request.META.get('HTTP_REFERER')
    templatepage = 'pay/payparametr.html'
    mounts = Mount.objects.all()
    zone = Zone.objects_limit.all()

    if request.method == 'POST':
        period_id = request.POST.get('mah')
        active = request.POST.get('active')
        if active == 'on':
            active = True
        else:
            active = False

        for item in zone:
            zone_id = request.POST.get(f'zonenew{item.id}')
            zaribfani = request.POST.get(f'zaribfani{item.id}')
            zaribetlaf = request.POST.get(f'zaribetlaf{item.id}')
            ezafekar = request.POST.get(f'ezafekar{item.id}')
            pay = PayDarsadMah.objects.get(period_id=period_id, zone_id=zone_id)

            pay.zaribfani = zaribfani
            pay.zaribetlaf = zaribetlaf
            pay.zaribbahrevari = ezafekar
            pay.active = active
            pay.save()

        pay = PayDarsadMah.objects.filter(period_id=period_id).order_by('-rotbe')
        add_to_log(request, 'ذخیره پارامتر های حقوق دستمزد ', 0)
        # return redirect(url)

        return TemplateResponse(request, templatepage,
                                {'zone': zone, 'mounts': mounts, 'postperiod': int(period_id),
                                 'pay': pay, 'active': active})
    return TemplateResponse(request, templatepage, {'zone': zone, 'mounts': mounts})


@cache_permission('pay')
def payzone(request):
    owner = None
    mounts = Mount.objects.filter(isshow=True)
    if request.user.owner.role.role == 'mgr':
        owner = Owner.objects.filter(role__role='tek', active=True)
    elif request.user.owner.role.role == 'zone':
        owner = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True)
    elif request.user.owner.role.role == 'setad':
        owner = Owner.objects.filter(refrence_id=9, role__role='setad', active=True)

    if request.method == 'POST':
        ok = True
        period_id = request.POST.get('period')
        tek_id = request.POST.get('tek')
        parametr = Parametrs.objects.all().first()
        if parametr.ismohasebat:
            mohasebat = mohasebat_hoghogh(tek_id, period_id)
        else:
            mohasebat = None

        dore = Mount.objects.get(id=int(period_id))
        active = PayDarsadMah.objects.filter(period_id=period_id).first()
        active = active.active
        url = request.META.get('HTTP_REFERER')
        if active:
            messages.error(request, 'اطلاعات این ماه هنوز باز نشده')
            return redirect(url)
        if not dore.active:
            messages.error(request, 'ویرایش این ماه بسته شده است')
            return redirect(url)

        try:
            tekkarkerd = TekKarkard.objects.get(period_id=period_id, tek__zone_id=request.user.owner.zone_id,
                                                tek_id=tek_id)
            karkerd = tekkarkerd.value
            kilometr = tekkarkerd.khodro
        except TekKarkard.DoesNotExist:
            karkerd = dore.day
            kilometr = 0
        ezafekarsum = Payroll.objects.filter(period_id=period_id, tek__zone_id=request.user.owner.zone_id,
                                             paybaseparametrs_id=16).aggregate(sum=Sum('count'))

        ezafekarsum = ezafekarsum['sum']
        co_tek = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True).count()
        mojaz = Zone.objects_limit.get(id=request.user.owner.zone_id)
        mojaz = mojaz.tekcount
        if co_tek > mojaz:
            messages.error(request, 'تعداد پشتیبان های شما بیشتر از مقدار مجاز است')
            return redirect(HOME_PAGE)
        sarane = PayDarsadMah.objects.get(zone_id=request.user.owner.zone_id, period_id=period_id)
        adad = sarane.zaribbahrevari
        sumsarane = co_tek * adad
        result = PayDarsadMah.objects.get(zone_id=request.user.owner.zone_id, period_id=period_id)
        zaribfani = result.zaribfani
        zaribetlaf = result.zaribetlaf
        for item in PayItems.objects.all().order_by('id'):
            uniq = str(tek_id) + "-" + str(period_id), str(item.id)
            try:
                PayParametr.objects.create(period_id=period_id, tek_id=tek_id, payitem_id=item.id,
                                           user_id=request.user.id, uniq=uniq)
            except IntegrityError:
                continue
        payparametrs = PayParametr.objects.filter(tek_id=tek_id, period_id=period_id).order_by('payitem__paybase_id',
                                                                                               'payitem_id')
        if ezafekarsum is None:
            ezafekarsum = 0
        count_ezafe = PayItems.objects.filter(paybase_id=16).count()
        count_jazb = PayItems.objects.filter(paybase_id=9).count()
        count_etlaf = PayItems.objects.filter(paybase_id=8).count()

        return TemplateResponse(request, 'pay/payzone.html',
                                {'karkerd': karkerd, 'kilometr': kilometr, 'owner': owner,
                                 'mounts': mounts, 'mohasebat': mohasebat,
                                 'payparametrs': payparametrs,
                                 'postperiod': int(period_id), 'zaribetlaf': zaribetlaf, 'zaribfani': zaribfani,
                                 'count_ezafe': count_ezafe, 'count_etlaf': count_etlaf, 'count_jazb': count_jazb,
                                 'posttek': int(tek_id), 'ok': ok, 'sumsarane': sumsarane, 'ezafekarsum': ezafekarsum})
    return TemplateResponse(request, 'pay/payzone.html', {'owner': owner, 'mounts': mounts})


@transaction.atomic
def payteksave(request):
    _sum = 0
    _tadil = None
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        period_id = request.POST.get('mah')
        tek_id = request.POST.get('teknew')
        kark = request.POST.get('kark')
        _khodro = request.POST.get('kilometr')
        ezafekar_setad = request.POST.get('ezafekar_setad')
        for item in PayItems.objects.filter(ename__isnull=False):
            calculations = mohasebat_hoghogh(tek_id, period_id)
            expected = 0
            # یافتن مقدار مورد نظر بر اساس نام آیتم
            for calc in calculations:
                if item.ename == calc['name']:
                    expected = calc['val']
            if int(request.POST.get(f't{item.id}', 0)) != expected:
                messages.error(request, 'دستکاری در داده‌های محاسبه شده شناسایی شد')
                add_to_log(request, 'دستکاری در داده‌های محاسبه شده شناسایی شد', 0)

                return redirect(url)
        payrolls = Payroll.objects.filter(period_id=period_id, tek_id=tek_id)

        for py in payrolls:
            if py.accept:
                messages.warning(request, 'حقوق و دستمزد این شخص قبلا توسط مدیر تایید نهایی شده و قابل ویرایش نمی باشد')
                return redirect(url)
            else:
                py.delete()

        TekKarkard.get_karkard(value=int(kark), period=period_id, tekid=tek_id, kilometr=_khodro)
        _list = PayItems.objects.all()
        group = Owner.objects.get(id=int(tek_id))
        payment = PersonPayment.objects.filter(baseparametr__enname='paye', owner_id=tek_id).order_by('-id')
        if payment:
            paye = payment[0].price
        else:
            payment = PayBaseParametrs.objects.get(enname='paye')
            paye = payment.price
        _addlist = []
        for item in PayBaseParametrs.objects.filter(price__gt=0, isshow=True, is_auto=True):
            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr_id=item.id).last()
            if pbase:
                if pbase.tadil != 0 and pbase.tadil is not None:
                    _tadil = pbase.tadil.split("#")
                    _priceold = pbase.price
                    for _item in _tadil:
                        if _item:
                            try:
                                checkprice = Payroll.objects.get(period_id=_item, tek_id=tek_id,
                                                                 paybaseparametrs_id=pbase.baseparametr.id,
                                                                 accepttedad=0)

                                _sumprice = Payroll.objects.filter(tek_id=tek_id,
                                                                   paybaseparametrs_id=pbase.baseparametr.id,
                                                                   accepttedad=_item).aggregate(jam=Sum('price'))

                                _summjam = _sumprice['jam'] if _sumprice['jam'] is not None else 0
                                _sum = _summjam + checkprice.price
                                if _priceold > _sum:
                                    _newprice = _priceold - _sum
                                    Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                                           paybaseparametrs_id=pbase.baseparametr.id,
                                                           count=group.childcount,
                                                           accepttedad=_item,
                                                           price=_newprice)
                            except ObjectDoesNotExist:
                                continue
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=0,
                                       price=pbase.price)
            else:
                pbase = PayBaseParametrs.objects.get(id=item.id)
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=0,
                                       price=pbase.price)
        tek = TekKarkard.objects.get(tek_id=tek_id, period_id=period_id)
        pbase = PayBaseParametrs.objects.get(enname='khodro')
        Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                               paybaseparametrs_id=pbase.id, count=tek.khodro,
                               price=tek.khodro * pbase.price)

        """جمع دستمزد روزانه و با پایه سنوات"""
        pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='Drozaneandpaye').last()

        if pbase:
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.baseparametr.id, count=pbase.baseparametr.count,
                                   price=tek.Dastmozd_rozane)
        else:
            pbase = PayBaseParametrs.objects.get(enname='Drozaneandpaye')
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.id, count=pbase.count,
                                   price=tek.Dastmozd_rozane)

        """فوق العاده کمک رفاهی و ایاب ذهاب"""
        pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='ayabzahab').last()
        if pbase:
            if pbase.tadil != 0 and pbase.tadil is not None:
                _tadil = pbase.tadil.split("#")
                _priceold = pbase.price
                for item in _tadil:
                    if item:
                        try:
                            checkprice = Payroll.objects.get(period_id=item, tek_id=tek_id,
                                                             paybaseparametrs_id=pbase.baseparametr.id, accepttedad=0)
                            _sumprice = Payroll.objects.filter(tek_id=tek_id,
                                                               paybaseparametrs_id=pbase.baseparametr.id,
                                                               accepttedad=item).aggregate(jam=Sum('price'))
                            _summjam = _sumprice['jam'] if _sumprice['jam'] is not None else 0
                            _sum = _summjam + checkprice.price

                            if _priceold > _sum:
                                _newprice = _priceold - _sum
                                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                                       paybaseparametrs_id=pbase.baseparametr.id,
                                                       count=group.childcount,
                                                       accepttedad=item,
                                                       price=_newprice)
                        except ObjectDoesNotExist:
                            continue
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.baseparametr.id, count=pbase.baseparametr.count,
                                   price=pbase.price)
        else:
            pbase = PayBaseParametrs.objects.get(enname='ayabzahab')
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.id, count=pbase.count,
                                   price=pbase.price)

        """حق تاهل"""
        if group.marital_status == 'marid':
            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='marid').last()
            if pbase:
                if pbase.tadil != 0 and pbase.tadil is not None:

                    _tadil = pbase.tadil.split("#")
                    _priceold = pbase.price
                    _checkprice = 0
                    for item in _tadil:
                        if item:
                            try:
                                checkprice = Payroll.objects.get(period_id=item, tek_id=tek_id,
                                                                 paybaseparametrs_id=pbase.baseparametr.id,
                                                                 accepttedad=0)
                                _checkprice = checkprice.price
                                _sumprice = Payroll.objects.filter(tek_id=tek_id,
                                                                   paybaseparametrs_id=pbase.baseparametr.id,
                                                                   accepttedad=item).aggregate(jam=Sum('price'))
                                _summjam = _sumprice['jam'] if _sumprice['jam'] is not None else 0
                                _sum = _summjam + checkprice.price

                            except ObjectDoesNotExist:
                                pass
                            if _priceold > _sum:
                                _newprice = _priceold - _sum
                                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                                       paybaseparametrs_id=pbase.baseparametr.id,
                                                       count=group.childcount,
                                                       accepttedad=item,
                                                       price=_newprice)

                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=pbase.baseparametr.count,
                                       price=pbase.price)
            else:
                pbase = PayBaseParametrs.objects.get(enname='marid')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=pbase.count,
                                       price=pbase.price)
        """دستمزد ماهیانه"""
        pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='paye').last()
        if pbase:
            if pbase.tadil != 0 and pbase.tadil is not None:
                _tadil = pbase.tadil.split("#")
            pbase = Payroll.objects.filter(tek_id=tek_id, paybaseparametrs__enname='Dmonth', accepttedad=0).last()
            _priceold = tek.hoghogh_paye
            for item in _tadil:
                if item:
                    try:
                        checkprice = Payroll.objects.get(period_id=item, tek_id=tek_id,
                                                         paybaseparametrs_id=pbase.paybaseparametrs.id, accepttedad=0)
                        _sumprice = Payroll.objects.filter(tek_id=tek_id,
                                                           paybaseparametrs_id=pbase.paybaseparametrs.id,
                                                           accepttedad=item).aggregate(jam=Sum('price'))

                        _summjam = _sumprice['jam'] if _sumprice['jam'] is not None else 0
                        _sum = _summjam + checkprice.price
                        if _priceold > _sum:
                            _newprice = _priceold - _sum
                            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                                   paybaseparametrs_id=pbase.paybaseparametrs.id,
                                                   count=group.childcount,
                                                   accepttedad=item,
                                                   price=_newprice)
                    except ObjectDoesNotExist:
                        continue
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.paybaseparametrs.id, count=tek.value,
                                   price=tek.hoghogh_paye)
        else:

            pbase = PayBaseParametrs.objects.get(enname='Dmonth')
            _priceold = tek.hoghogh_paye
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.id, count=tek.value, price=tek.hoghogh_paye)

        """مزد سنواتی"""
        pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='mozd_sanavati').last()
        if pbase:
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.baseparametr.id, count=0, price=tek.mozd_sanavat)
        else:
            pbase = PayBaseParametrs.objects.get(enname='mozd_sanavati')
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.id, count=0, price=tek.mozd_sanavat)
        """حق اولاد"""
        pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='child').last()
        if pbase:
            if pbase.tadil != 0 and pbase.tadil is not None:
                _tadil = pbase.tadil.split("#")
                _pricechild = pbase.price * group.childcount
                for item in _tadil:
                    if item:
                        try:
                            checkprice = Payroll.objects.get(period_id=item, tek_id=tek_id,
                                                             paybaseparametrs_id=pbase.baseparametr.id, accepttedad=0)

                            _sumprice = Payroll.objects.filter(tek_id=tek_id,
                                                               paybaseparametrs_id=pbase.baseparametr.id,
                                                               accepttedad=item).aggregate(jam=Sum('price'))

                            _summjam = _sumprice['jam'] if _sumprice['jam'] is not None else 0
                            _sum = _summjam + checkprice.price
                            if _priceold > _sum:
                                _newprice = _pricechild - _sum
                                if _newprice > 0:
                                    Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                                           paybaseparametrs_id=pbase.baseparametr.id,
                                                           count=group.childcount,
                                                           accepttedad=item,
                                                           price=_newprice)
                        except ObjectDoesNotExist:
                            continue

            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.baseparametr.id, count=group.childcount,
                                   price=(pbase.price * group.childcount))
        else:
            pbase = PayBaseParametrs.objects.get(enname='child')
            Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                   paybaseparametrs_id=pbase.id, count=group.childcount,
                                   price=(pbase.price * group.childcount))

        if group.job_group in [12, 14, 16]:
            """ضریب مرتبه شغلی"""
            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='bazdehi').last()
            if pbase:
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=pbase.baseparametr.count,
                                       price=((paye * int(kark)) * pbase.count) / 100)
            else:
                pbase = PayBaseParametrs.objects.get(enname='bazdehi')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=pbase.count,
                                       price=((paye * int(kark)) * pbase.count) / 100)
            """حق جذب"""
            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='foghejazb').last()
            if pbase:
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=pbase.baseparametr.count,
                                       price=((paye * int(kark)) * pbase.count) / 100)
            else:
                pbase = PayBaseParametrs.objects.get(enname='foghejazb')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=pbase.count,
                                       price=((paye * int(kark)) * pbase.count) / 100)
            """ اضافه کاری"""
            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='ezafekari').last()
            if pbase:
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=ezafekar_setad,
                                       price=((tek.Dastmozd_rozane * 0.191) * int(ezafekar_setad)))
            else:
                pbase = PayBaseParametrs.objects.get(enname='ezafekari')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=ezafekar_setad,
                                       price=((tek.Dastmozd_rozane * 0.191) * int(ezafekar_setad)))

        if group.job_group in [7]:
            for item in _list:
                val = request.POST.get(f't{item.id}')
                pay = PayParametr.objects.get(period_id=period_id, tek_id=tek_id, payitem_id=item.id)
                pay.inputval = val
                pay.save()
            ezafekar = PayParametr.objects.filter(tek_id=tek_id, period_id=period_id, payitem__paybase_id=16).aggregate(
                summer=Sum('inputval'), tedad=Count('inputval'))
            etlaf = PayParametr.objects.filter(tek_id=tek_id, period_id=period_id, payitem__paybase_id=8).aggregate(
                summer=Sum('inputval'), tedad=Count('inputval'))
            jazb = PayParametr.objects.filter(tek_id=tek_id, period_id=period_id, payitem__paybase_id=9).aggregate(
                summer=Sum('inputval'), tedad=Count('inputval'))
            majmo = int(ezafekar['summer'])
            counter = int(ezafekar['tedad'])
            topemtiaz = majmo / counter
            kasr = 100 - topemtiaz
            kasrkol = (70 * kasr) / 100
            ezafe_kar = 70 - kasrkol
            ezafekarsum = Payroll.objects.filter(period_id=period_id, tek__zone_id=request.user.owner.zone_id,
                                                 paybaseparametrs_id=16).aggregate(sum=Sum('count'))
            co_tek = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True).count()
            sarane = PayDarsadMah.objects.get(zone_id=request.user.owner.zone_id, period_id=period_id)
            adad = sarane.zaribbahrevari
            zaribetlaf = sarane.zaribetlaf
            zaribfani = sarane.zaribfani
            sumsarane = co_tek * adad
            majmo = int(etlaf['summer'])
            counter = int(etlaf['tedad'])
            topemtiaz = majmo / counter
            kasr = 100 - topemtiaz
            kasrkol = (zaribetlaf * kasr) / 100
            etlaf = zaribetlaf - kasrkol
            majmo = int(jazb['summer'])
            counter = int(jazb['tedad'])
            topemtiaz = majmo / counter
            kasr = 100 - topemtiaz
            kasrkol = (zaribfani * kasr) / 100
            jazb = zaribfani - kasrkol
            mozd_shoghl = paye
            payrolls = Payroll.objects.filter(period_id=period_id, tek_id=tek_id)
            for py in payrolls:
                if py.accept:
                    messages.warning(request,
                                     'حقوق و دستمزد این شخص قبلا توسط مدیر تایید نهایی شده و قابل ویرایش نمی باشد')
                    return redirect('pay:CommitMgr')

            if ezafe_kar > round(sumsarane):
                messages.error(request, F'مقدار تخصیص سهمیه نباید از سرانه بیشتر باشد ')
                for i in Payroll.objects.filter(period_id=period_id, tek_id=tek_id):
                    i.delete()
                for i in PayParametr.objects.filter(period_id=period_id, tek_id=tek_id, payitem__paybase_id=16):
                    i.inputval = 0
                    i.save()
                return redirect(url)
            try:
                mypay = Payroll.objects.get(period_id=period_id, tek_id=tek_id, paybaseparametrs_id=16).count
            except Payroll.DoesNotExist:
                mypay = 0
            if ezafekarsum['sum']:
                ezafekarsum = (ezafekarsum['sum'] - mypay) + ezafe_kar

                if math.floor(ezafekarsum) > round(sumsarane):
                    messages.error(request, F'مقدار تخصیص سهمیه نباید از سرانه بیشتر باشد ')
                    for i in Payroll.objects.filter(period_id=period_id, tek_id=tek_id):
                        i.delete()
                    for i in PayParametr.objects.filter(period_id=period_id, tek_id=tek_id, payitem__paybase_id=16):
                        i.inputval = 0
                        i.save()
                    return redirect(url)

            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='bazdehi').last()
            if pbase:
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=etlaf,
                                       price=((mozd_shoghl * int(kark)) * etlaf) / 100)
            else:
                pbase = PayBaseParametrs.objects.get(enname='bazdehi')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=etlaf,
                                       price=((mozd_shoghl * int(kark)) * etlaf) / 100)

            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='ezafekari').last()
            if pbase:
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=ezafe_kar,
                                       price=(mozd_shoghl / 7.33 * 1.4 * ezafe_kar))
            else:
                pbase = PayBaseParametrs.objects.get(enname='ezafekari')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=ezafe_kar,
                                       price=(mozd_shoghl / 7.33 * 1.4 * ezafe_kar))

            pbase = PersonPayment.objects.filter(owner_id=tek_id, baseparametr__enname='foghejazb').last()
            if pbase:
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.baseparametr.id, count=jazb,
                                       price=((mozd_shoghl * int(kark)) * jazb) / 100)
            else:
                pbase = PayBaseParametrs.objects.get(enname='foghejazb')
                Payroll.objects.create(tek_id=tek_id, period_id=period_id, user_id=request.user.id,
                                       paybaseparametrs_id=pbase.id, count=jazb,
                                       price=((mozd_shoghl * int(kark)) * jazb) / 100)
    add_to_log(request, 'ذخیره پارامتر های حقوق دستمزد ', 0)

    messages.success(request, SUCCESS_MSG)
    return redirect(url)


@cache_permission('showfish')
def showfish(request):
    add_to_log(request, 'مشاهده فیش حقوقی ', 0)
    mounts = Mount.objects.all().order_by('id')

    if request.method == 'POST':
        period = request.POST.get('period')
        paylist = Payroll.objects.filter(period_id=period, tek_id=request.user.owner.id,
                                         paybaseparametrs__sortable__gt=0).order_by('paybaseparametrs__sortable',
                                                                                    'accepttedad')
        paysum = Payroll.objects.filter(period_id=period, paybaseparametrs__isshow=True,
                                        tek_id=request.user.owner.id).aggregate(sum=Sum('price'))
        owner = Owner.objects.get(id=request.user.owner.id)
        month_name = Mount.objects.get(id=int(period))
        month_name = month_name.mount
        isaccept = paylist.first()
        return TemplateResponse(request, 'pay/showfish.html',
                                {'mounts': mounts, 'paylist': paylist, 'paysum': paysum,
                                 'period': int(period), 'month_name': month_name, 'owner': owner, 'isaccept': isaccept})
    return TemplateResponse(request, 'pay/showfish.html',
                            {'mounts': mounts})


@cache_permission('storetotek')
def storetotek(request):
    owners = Owner.objects.filter(
        Q(role__role='tek', active=True, zone_id=request.user.owner.zone_id) |
        Q(refrence__ename='tek', active=True, zone_id=request.user.owner.zone_id)
    ).distinct()
    result = StoreList.objects.filter(status_id=3, statusstore_id=1, zone_id=request.user.owner.zone_id)
    thislist = []
    _todaydate = date.today()
    si_ago_sell = _todaydate.today() - datetime.timedelta(days=30)

    for q in result:
        risk = StoreHistory.objects.filter(store_id=q.id, create__gte=si_ago_sell, create__lte=_todaydate.today(),
                                           status_id=6).count()
        if risk == 0:
            level = 0
        elif risk == 1:
            level = 1
        elif risk == 2:
            level = 2
        else:
            level = 3
        thisdict = {
            "id": q.id,
            "serial": q.serial,
            "level": level,
        }
        thislist.append(thisdict)
    return TemplateResponse(request, 'store/storetotek.html',
                            {'owners': owners, 'masters': thislist})


@cache_permission('storetokar')
def storetokargah(request):
    sid = 0
    ownerid = 0
    masters = None
    stores = None
    posts = None

    owners = Owner.objects.filter(
        Q(role__role='tek', zone_id=request.user.owner.zone_id) |
        Q(refrence__ename='tek', zone_id=request.user.owner.zone_id)
    ).distinct()
    statusstore = StatusStore.objects.all()
    if request.method == 'POST':
        sid = request.POST.get('sid')
        ownerid = request.POST.get('ownerid')
        if ownerid == '0':

            masters = StoreList.objects.filter(status_id=8, statusstore_id=sid,
                                               zone_id=request.user.owner.zone_id).order_by('-update')
            posts = StoreList.objects.filter(status_id=11, statusstore_id=sid,
                                             zone_id=request.user.owner.zone_id).order_by('-update')
            stores = StoreList.objects.filter(statusstore_id=sid, status_id=10, zone_id=request.user.owner.zone_id)
        else:

            masters = StoreList.objects.filter(status_id=6, getuser_id=ownerid, statusstore_id=sid,
                                               zone_id=request.user.owner.zone_id).order_by('-update')
            posts = StoreList.objects.filter(status_id=11, statusstore_id=sid,
                                             zone_id=request.user.owner.zone_id).order_by('-update')
            stores = StoreList.objects.filter(statusstore_id=sid, status_id=10, zone_id=request.user.owner.zone_id)
    return TemplateResponse(request, 'store/storetokargah.html',
                            {'sid': int(sid), 'masters': masters, 'stores': stores, 'owners': owners,
                             'statusstore': statusstore,
                             'ownerid': int(ownerid), 'posts': posts})


@cache_permission('isstart')
def starterr(request):
    if request.method == 'POST':
        url = request.META.get('HTTP_REFERER')
        try:
            serial = int(request.POST.get('init'))
            info = request.POST.get('info')
            store = StoreList.objects.get(id=serial)
            if store.status_id == 6:
                messages.warning(request, 'این قطعه در وضعیت از ابتدا معیوب میباشد')
                return redirect(url)
            if store.status_id in [4, 5]:
                store.status_id = 6
                store.save()
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information="ثبت ابتدا معیوب   ", starterror=True,
                                            status_id=6, description=info)

                messages.success(request, SUCCESS_MSG)
                return redirect(url)
            else:
                messages.warning(request,
                                 'برای ثبت از ابتدا معیوب وضعیت قطعه باید درحالت نصب در جایگاه یا نزد تکنسین باشد')
        except TypeError as e:
            messages.warning(request, 'شماره سریال بصورت صحیح وارد نشده')

    return TemplateResponse(request, 'store/StartError.html', {'a': 1})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getstoretotek(request, *args, **kwargs):
    thislist = []
    _id = int(request.POST.get('id'))
    store = int(request.POST.get('StoreId'))

    result = StoreList.objects.filter(status_id__in=[4, 16], statusstore_id=store, getuser_id=_id)
    _todaydate = date.today()
    si_ago_sell = _todaydate.today() - datetime.timedelta(days=30)
    for q in result:
        risk = StoreHistory.objects.filter(store_id=q.id, create__gte=si_ago_sell, create__lte=_todaydate.today(),
                                           status_id=6).count()
        if risk == 0:
            level = 0
        elif risk == 1:
            level = 1
        elif risk == 2:
            level = 2
        else:
            level = 3
        thisdict = {
            "id": q.id,
            "serial": q.serial,
            "level": level,
            'status': q.status.name,
        }
        thislist.append(thisdict)

    return JsonResponse({"list": thislist, 'message': 'success'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getstore(request, *args, **kwargs):
    thislist = []
    _id = int(request.POST.get('id'))
    result = StoreList.objects.filter(status_id=3, statusstore_id=_id, zone_id=request.user.owner.zone_id)
    _todaydate = date.today()
    si_ago_sell = _todaydate.today() - datetime.timedelta(days=30)

    for q in result:
        risk = StoreHistory.objects.filter(store_id=q.id, create__gte=si_ago_sell, create__lte=_todaydate.today(),
                                           status_id=6).count()
        if risk == 0:
            level = 0
        elif risk == 1:
            level = 1
        elif risk == 2:
            level = 2
        else:
            level = 3
        thisdict = {
            "id": q.id,
            "serial": q.serial,
            "level": level,
        }
        thislist.append(thisdict)

    return JsonResponse({"list": thislist, 'message': 'success'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getsenddaghi(request, *args, **kwargs):
    thislist = []
    _id = int(request.POST.get('id'))
    result = StoreList.objects.filter(status_id=6, statusstore_id=_id, zone_id=request.user.owner.zone_id)
    for q in result:
        thisdict = {
            "id": q.id,
            "serial": q.serial,
            'tek': q.getuser,
        }
        thislist.append(thisdict)

    return JsonResponse({"list": thislist, 'message': 'success'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getstoretek(request, *args, **kwargs):
    thislist = []
    serial = None
    _id = int(request.POST.get('id'))
    store = int(request.POST.get('StoreId'))
    pump = request.POST.get('pump')

    result = StoreList.objects.filter(status_id=4, statusstore_id=store, getuser_id=_id)
    if len(pump) > 1:
        if store == 1:
            pump = Pump.objects.get(id=int(pump))
            serial = pump.master

        if store == 2:
            pump = Pump.objects.get(id=int(pump))
            serial = pump.pinpad

    if serial is None:
        serial = 0
    for q in result:
        thisdict = {
            "id": q.id,
            "serial": q.serial,
        }
        thislist.append(thisdict)

    return JsonResponse({"list": thislist, 'message': 'success', 'serial': serial})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def addstoretek(request):
    thislist = []
    if request.method == 'POST':
        _oldst = 3
        userid = checkxss(request.POST.get('id_tek'))
        mylist = checkxss(request.POST.get('strIds'))
        val = checkxss(request.POST.get('val'))
        if val == '1':
            _oldst = [3, ]
        if val == '2':
            _oldst = [6, 8]

        x = mylist.split(',')
        for item in x:
            gsinfo = StoreList.object_role.c_base(request).get(id=item, status_id__in=_oldst)
            if val == '1':
                gsinfo.status_id = 16
                gsinfo.getuser_id = userid
            if val == '2':
                gsinfo.status_id = 10

            gsinfo.save()
            if gsinfo.getuser_id:
                _user = gsinfo.getuser.name + ' ' + gsinfo.getuser.lname
            else:
                _user = f"کارگاه {gsinfo.zone.name}"

            thisdict = {
                "id": gsinfo.id,
                "serial": gsinfo.serial,
                "st": gsinfo.statusstore.name,
                "user": _user,
            }
            thislist.append(thisdict)
            if val == '1':

                repair = Repair.objects.filter(store_id=gsinfo.id, status=0)
                for rep in repair:
                    rep.status = 1
                    rep.save()
                StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                            information="ارسال  قطعه از انبار  " + str(gsinfo.zone.name),
                                            status_id=16,
                                            description=f' ارسال به تکنسین  (بین راهی)  {gsinfo.getuser} ')
            if val == '2':
                if gsinfo.getuser_id:
                    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="برگشت قطعه داغی به انبار" + str(gsinfo.zone.name),
                                                status_id=10, description=f' تحویل از تکنسین  {gsinfo.getuser} ')
                else:
                    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="برگشت قطعه داغی به انبار " + str(gsinfo.zone.name),
                                                status_id=10,
                                                description=f' تحویل از کارگاه منطقه  {gsinfo.zone.name} ')

    return JsonResponse({"message": "success", 'list': thislist})


def backtodaghi(request, _id):
    url = request.META.get('HTTP_REFERER')
    gsinfo = StoreList.objects.get(id=_id)
    gsinfo.status_id = 10
    gsinfo.save()
    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                information="برگشت قطعه داغی به انبار" + str(gsinfo.zone.name),
                                status_id=10, description=f' تحویل از تکنسین  {gsinfo.getuser} ')
    return redirect(url)


def addstoregs(request):
    thislist = []
    if request.method == 'POST':

        mylist = request.POST.get('strIds')

        x = mylist.split(',')
        for item in x:
            gsinfo = StoreList.objects.get(id=item)
            gsinfo.status_id = 5

            gsinfo.save()

            thisdict = {
                "id": gsinfo.id,
                "serial": gsinfo.serial,

            }
            thislist.append(thisdict)
            StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                        information="ارسال  قطعه به جایگاه مستقیم  ",
                                        status_id=4, description=' مشکلات اولیه  ')
    return JsonResponse({"message": "success", 'list': thislist})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def removestoretek(request):
    thislist = []
    _oldst = 4
    val = 0
    if request.method == 'POST':
        mylist = request.POST.get('strIds')
        userid = request.POST.get('userid')
        val = request.POST.get('val')
        if val == '1':
            _oldst = [4, 16]
        if val == '2':
            _oldst = [10]
        if val == '3':
            _oldst = [10]
        if val == '4':
            _oldst = [10]
        if val == '2' and userid == '0':
            _oldst = [10]
        x = mylist.split(',')
        for item in x:
            gsinfo = StoreList.object_role.c_base(request).get(id=item, status_id__in=_oldst)

            if val == '1':
                gsinfo.status_id = 3
            if val == '2':
                gsinfo.status_id = 6
            if val == '3':
                gsinfo.status_id = 11
            if val == '4':
                gsinfo.status_id = 8

            if val == '2' and userid == '0':
                gsinfo.status_id = 8
            gsinfo.save()
            if gsinfo.getuser_id:
                _user = gsinfo.getuser.name + ' ' + gsinfo.getuser.lname
            else:
                _user = f"کارگاه {gsinfo.zone.name}"
            thisdict = {
                "id": gsinfo.id,
                "serial": gsinfo.serial,
                "st": gsinfo.statusstore.name,
                "user": _user,
            }
            thislist.append(thisdict)
            if userid != '0':

                if val == '1':
                    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="برگشت  قطعه از تکنسین   " + str(gsinfo.getuser),
                                                status_id=3, description=f' به انبار  {gsinfo.zone.name} ')
                if val == '2':
                    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="برگشت  قطعه از انبار داغی به تکنسین   " + str(
                                                    gsinfo.getuser),
                                                status_id=6, description=f' از انبار  {gsinfo.zone.name} ')
                if val == '3':
                    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="آماده سازی بسته پستی " + str(gsinfo.zone.name),
                                                status_id=11, description=f' آماده سازی بسته پستی   ')
                if val == '4':
                    StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="ارسال مستقیم قطعه از تکنسین به کارگاه " + str(
                                                    gsinfo.zone.name),
                                                status_id=8, description=f' ارسال مستقیم قطعه به کارگاه   ')
            else:

                StoreHistory.objects.create(store_id=gsinfo.id, owner_id=request.user.owner.id, baseroot=0,
                                            information="برگشت  قطعه به کارگاه   " + str(gsinfo.zone.name),
                                            status_id=8)

    return JsonResponse({"message": "success", "val": val, 'list': thislist})


@cache_permission('history')
def editstores(request, serial):
    storelist = StoreList.objects.get(serial=serial)
    form = StoreListForm(instance=storelist)
    if request.method == 'POST':
        form = StoreListForm(request.POST, instance=storelist)
        a = form.save()
        StoreHistory.objects.create(store_id=a.id, owner_id=request.user.owner.id, baseroot=0,
                                    information="ویرایش قطعه  ",
                                    status_id=a.status_id,
                                    description=a.info)
        messages.success(request, 'ویرایش قطعه انجام شد')
    return TemplateResponse(request, 'store/editstores.html',
                            {'storelist': storelist, 'form': form,
                             })


@cache_permission('history')
def history(request):
    items = ''
    st_items = None
    idstore = None
    zoneid = None
    _list = None
    namestore = ''
    serial = ''
    ok = False
    if request.method == 'POST':
        serial = request.POST.get('search')
        serial = checkxss(serial)
        serial = checknumber(serial)
        try:
            ok = True
            idstore = StoreList.objects.get(serial=serial)
            namestore = idstore.statusstore.name
            zoneid = idstore.zone_id
            _list = StoreHistory.objects.filter(store_id=idstore).order_by('-id')
            items = StatusRef.objects.filter(id__in=['3', '4', '5'])
            st_items = StatusStore.objects.all()
        except ObjectDoesNotExist:
            ok = False
            messages.error(request, 'این شماره سریال یافت نشد')
    return TemplateResponse(request, 'store/history.html',
                            {'list': _list, 'items': items, 'ok': ok, 'serial': serial,
                             'idstore': idstore, 'namestore': namestore, 'zoneid': zoneid, 'st_items': st_items})


@cache_permission('history')
def historylist(request, _id):
    items = ''
    st_items = None
    idstore = None
    zoneid = None
    _list = None
    namestore = ''

    try:
        ok = True
        idstore = StoreList.objects.get(serial=_id)
        namestore = idstore.statusstore.name
        zoneid = idstore.zone_id
        _list = StoreHistory.objects.filter(store_id=idstore).order_by('-id')
        items = StatusRef.objects.filter(id__in=['3', '4', '5'])
        st_items = StatusStore.objects.all()
    except ObjectDoesNotExist:
        ok = False
        messages.error(request, 'این شماره سریال یافت نشد')
    return TemplateResponse(request, 'store/history.html',
                            {'list': _list, 'items': items, 'ok': ok, 'serial': _id,
                             'idstore': idstore, 'namestore': namestore, 'zoneid': zoneid, 'st_items': st_items})


def historychange(request, serial):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        role = request.POST.get('role')
        st_store = request.POST.get('st_store')
        inzone = request.POST.get('inzone')

        result = StoreList.objects.get(serial=serial)
        old = result.status.name
        oldstatus = result.statusstore_id
        if role and role != '0':
            result.status_id = role
        if st_store and st_store != '0':
            result.statusstore_id = st_store
        if inzone == 'on':
            result.zone_id = request.user.owner.zone_id
            StoreHistory.objects.create(store_id=result.id, owner_id=request.user.owner.id, baseroot=0,
                                        information="تغییر منطقه  ",
                                        status_id=result.status_id,
                                        description=f' به،   {request.user.owner.zone.name} ')
        result.save()

        if role and role != '0':
            st = StatusRef.objects.get(id=role)
            lasthistory = StoreHistory.objects.filter(store_id=result.id).last()
            lasthistory.status_id = 12
            lasthistory.save()
            StoreHistory.objects.create(store_id=result.id, owner_id=request.user.owner.id, baseroot=0,
                                        information="تغییر وضعیت قطعه از " + str(old),
                                        status_id=role, description=f' به،   {st.name} ')
        if st_store and int(st_store) != oldstatus and st_store != '0':
            StoreHistory.objects.create(store_id=result.id, status_id=result.status_id,
                                        owner_id=request.user.owner.id,
                                        baseroot=0,
                                        information="تغییر نوع قطعه  "
                                        )
        messages.success(request, SUCCESS_MSG)
        return redirect(url)
    return redirect(url)


@transaction.atomic
def import_excel_store(request, _id, zoneid):
    form = open_excel(request.POST)
    zone = Zone.objects.get(id=zoneid)
    _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
    _log = []
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            _p = 0
            _m = 0
            _log = []
            _list = Store.objects.get(id=_id)
            # _list.status_id = 9
            _list.save()
            for i in range(1, m_row + 1):
                serial = sheet_obj.cell(row=i, column=1).value
                serial = checknumber(str(serial))
                try:

                    mystore = StoreList.objects.get(serial=serial)
                    mystore.zone_id = zoneid
                    mystore.owner_id = request.user.owner.id
                    mystore.status_id = 9
                    mystore.store_id = _id

                    mystore.save()
                    StoreHistory.objects.filter(store_id=mystore.id, owner_id=request.user.owner.id,
                                                storage_id=_storage.id,
                                                status_id=9).delete()
                    StoreHistory.objects.create(store_id=mystore.id, owner_id=request.user.owner.id, baseroot=_list.id,
                                                information="ارسال قطعه از " + str(_storage.name),
                                                storage_id=_storage.id,
                                                status_id=9, description=f' تخصیص به منطقه  {mystore.zone.name} ')

                except:
                    _log.append(f' سریال {serial} وجود ندارد ')

            _sum = StoreList.objects.filter(zone_id=zoneid, store_id=_id).aggregate(
                master=(Count(Case(When(statusstore_id=1, then=1)))),
                pinpad=(Count(Case(When(statusstore_id=2, then=1)))))
            _list.send_master = _sum['master']
            _list.send_pinpad = _sum['pinpad']
            _list.save()
            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        if _log:
            return render(request, 'importexcelstore.html', {'form': form, 'log': _log, 'zone': zone})
        else:
            return redirect('pay:PoshtibanList')

    return render(request, 'importexcelstore.html', {'form': form, 'log': _log, 'zone': zone})


@transaction.atomic
@cache_permission('exnazer')
def import_excel_nazer(request):
    form = open_excel(request.POST)
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):

                master = str(sheet_obj.cell(row=i, column=2).value)
                pinpad = str(sheet_obj.cell(row=i, column=7).value)

                if master and master.isnumeric():
                    try:

                        _list = StoreList.objects.get(serial=master)
                        _list.status_id = 13
                        _list.save()
                    except ObjectDoesNotExist:

                        _list = StoreList.objects.create(serial=master, owner_id=request.user.id, status_id=13,
                                                         statusstore_id=1, uniq=str(master) + "-" + str(1))

                    StoreHistory.objects.create(store_id=_list.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="بازدید ناظر پخش ",
                                                status_id=13, description=f'بازدید ناظر پخش')
                if pinpad and pinpad.isnumeric():
                    try:
                        _list = StoreList.objects.get(serial=pinpad)
                        _list.status_id = 13
                        _list.save()
                    except ObjectDoesNotExist:
                        _list = StoreList.objects.create(serial=pinpad, owner_id=request.user.id, status_id=13,
                                                         statusstore_id=2, uniq=str(pinpad) + "-" + str(1))

                    StoreHistory.objects.create(store_id=_list.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="بازدید ناظر پخش ",
                                                status_id=13, description=f'بازدید ناظر پخش')

            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        return redirect(HOME_PAGE)

    return TemplateResponse(request, 'importexcel.html', {'form': form})


@transaction.atomic
def import_excel_kargah(request, _id):
    form = open_excel(request.POST)
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):

                master = sheet_obj.cell(row=i, column=1).value
                master = checknumber(str(master))
                st = sheet_obj.cell(row=i, column=2).value
                st = checknumber(str(st))

                if master:
                    try:

                        _list = StoreList.objects.get(serial=master)
                        _list.status_id = 8
                        _list.zone_id = _id
                        _list.save()
                    except ObjectDoesNotExist:

                        _list = StoreList.objects.create(serial=master, owner_id=request.user.id, zone_id=_id,
                                                         status_id=8, statusstore_id=int(st),
                                                         uniq=str(master) + "-" + str(st))

                    StoreHistory.objects.create(store_id=_list.id, owner_id=request.user.owner.id, baseroot=0,
                                                information="موجودی کارگاه  ",
                                                status_id=8, description=f'')

            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        return redirect(HOME_PAGE)

    return render(request, 'importexcel.html', {'form': form})


@transaction.atomic
@cache_permission('bulk_chstore')
def import_excel_changest(request):
    form = open_excel(request.POST)
    zone = Zone.objects.all()
    statusref = StatusRef.objects.all()
    if request.method == 'POST':
        _zone = request.POST.get('zone')
        _statusref = request.POST.get('statusref')
        _info = request.POST.get('info')
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                master = sheet_obj.cell(row=i, column=1).value
                statusghete = sheet_obj.cell(row=i, column=2).value
                master = checknumber(str(master))
                if master:

                    try:
                        _list = StoreList.objects.get(serial=master)
                        if _statusref == '00':

                            his = StoreHistory.objects.filter(store_id=_list.id).order_by('-id').first()
                            _list.status_id = his.status_id
                            _list.zone_id = his.owner.zone_id
                        else:
                            _list.status_id = _statusref
                        if _statusref == '0':
                            i = 0
                            his = StoreHistory.objects.filter(store_id=_list.id).order_by('-id')
                            for item in his:
                                if i == 1:
                                    _list.status_id = item.status_id
                                    break
                                i += 1

                        if _zone != '0':
                            _list.zone_id = _zone

                        _list.save()
                        StoreHistory.objects.create(store_id=_list.id, owner_id=request.user.owner.id, baseroot=0,
                                                    information="تغییر وضعیت با فایل اکسل  ",
                                                    status_id=_list.status_id, description=str(_info))
                    except ObjectDoesNotExist:
                        if statusghete and len(str(statusghete)) > 0:
                            _newst = StoreList.objects.create(zone_id=_zone, serial=master, status_id=_statusref,
                                                              owner_id=request.user.owner.id,
                                                              statusstore_id=statusghete,
                                                              uniq=str(master) + "-" + str(statusghete))
                            StoreHistory.objects.create(store_id=_newst.id, owner_id=request.user.owner.id, baseroot=0,
                                                        information=" با فایل اکسل اضافه شد ",
                                                        status_id=_newst.status_id, description=str(_info))

            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        return redirect(HOME_PAGE)

    return TemplateResponse(request, 'store/changetstoreexcel.html',
                            {'form': form, 'statusref': statusref, 'zone': zone})


@transaction.atomic
@cache_permission('residinzon')
def import_excel_residstore(request, _id):
    form = open_excel(request.POST)
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            _test = 0
            for u in range(1, m_row + 1):

                master = sheet_obj.cell(row=u, column=1).value
                master = checknumber(str(master))
                a = StoreList.objects.filter(store_id=_id, serial=master).count()
                if a == 0:
                    _test += 1

                if sheet_obj.cell(row=u, column=3).value == None:
                    if _test > 6:
                        messages.error(request, ' مغایرت فایل بسیار زیاد است ')
                        return redirect('pay:residZone')

            store = Store.objects.get(id=_id)
            store.status_id = 8
            store.resid_date = datetime.datetime.now()
            store.save()
            ststore = StoreList.objects.filter(store_id=_id)
            for item in ststore:
                item.status_id = 15
                item.save()
            xlsn = 0
            for xls in range(2):
                xlsn += 1

                for i in range(1, m_row + 1):
                    master = sheet_obj.cell(row=i, column=xlsn).value
                    master = checknumber(str(master))
                    if master:

                        try:
                            _list = StoreList.objects.get(serial=master)
                            _list.status_id = 8
                            _list.zone_id = request.user.owner.zone_id
                            _list.save()

                            StoreHistory.objects.create(store_id=_list.id, owner_id=request.user.owner.id,
                                                        information="رسید  قطعه به کارگاه ",
                                                        status_id=8, baseroot=0,
                                                        description=f' رسید  به کارگاه  {request.user.owner.zone.name} ')
                        except ObjectDoesNotExist:
                            storeli = StoreList.objects.create(serial=master,
                                                               owner_id=request.user.id,
                                                               zone_id=request.user.owner.zone_id, statusstore_id=xlsn,
                                                               status_id=8, store_id=store.id,
                                                               uniq=str(master) + "-" + str(xlsn))
                            StoreHistory.objects.create(store_id=storeli.id, owner_id=request.user.owner.id,
                                                        information="رسید  قطعه به کارگاه مغایرتی",
                                                        status_id=8, baseroot=0,
                                                        description=f' رسید  به کارگاه  {request.user.owner.zone.name} ')

            ststore = StoreList.objects.filter(store_id=_id, status_id=15)
            for item in ststore:
                StoreHistory.objects.create(store_id=item.id, owner_id=request.user.owner.id,
                                            information="مغایرت",
                                            status_id=15, baseroot=0, residroot=_id,
                                            description=f' رسید به کارگاه {request.user.owner.zone.name}')
            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        return redirect('pay:residZone')

    return TemplateResponse(request, 'importexcel.html', {'form': form})


def okalert(request, _id):
    url = request.META.get('HTTP_REFERER')
    store = Store.objects.get(id=_id)
    store.alert72 = True
    store.save()
    add_to_log(request, 'ثبت تاخیر 72 ساعت  ', 0)
    return redirect(url)


@transaction.atomic
@cache_permission('newdaghi')
def changestore(request):
    url = request.META.get('HTTP_REFERER')
    form = ImageStore(request.GET)
    tickets1 = Ticket.objects.filter(
        isdaghi=True,
        status_id=2,
        actioner_id=request.user.owner.id
    ).select_related('gs').prefetch_related('Pump').only('id', 'gs__name', 'Pump__number')

    tickets2 = Ticket.objects.filter(
        isdaghi=True,
        status_id=2,
        usererja=request.user.owner.id
    ).select_related('gs').prefetch_related('Pump').only('id', 'gs__name', 'Pump__number')

    tickets = list(tickets1.union(tickets2))
    if request.method == 'POST':
        form = ImageStore(request.POST, request.FILES)
        store = int(request.POST.get('store'))
        customfile = request.POST.get('ok')

        storeid = store
        serial = request.POST.get('init')
        _ticket = Ticket.objects.get(id=store)

        if _ticket.failure.failurecategory.id == 1010:
            _storeid = 1
        else:
            _storeid = 2
        serial = check_serial(serial)
        if not serial:
            url = request.META.get('HTTP_REFERER')
            messages.error(request, ADD_STORE_MSG)
            return redirect(url)
        if customfile == '2':
            try:
                store = StoreList.objects.get(serial=serial, statusstore_id=_storeid)
            except ObjectDoesNotExist:
                messages.warning(request, "نوع قطعه اشتباه انتخاب شده است از سابقه قطعه بررسی کنید")
                return TemplateResponse(request, 'store/daghi.html',
                                        {'ok': 1, 'form': form, 'tickets': tickets})
            if form.is_valid():
                form.instance.store_id = store.id
                form.instance.owner_id = request.user.owner.id
                a = form.save()
                store.status_id = 6
                store.getuser_id = request.user.owner.id
                store.owner_id = request.user.id
                store.zone_id = request.user.owner.zone_id
                store.save()

                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id, imgid_id=a.id,
                                            information=f"ارسال قطعه داغی (تصویر بارگذاری شد) تیکت {_ticket.id}",
                                            status_id=6, description=f' به پشتیبان  {request.user.owner} ')
                daghi_mande = Owner.objects.get(id=request.user.owner.id)
                daghi_mande.daghimande = False
                daghi_mande.save()
                _ticket.isdaghi = False
                _ticket.serialdaghi = serial
                _ticket.save()

                s_daghi = Workflow.objects.filter(ticket_id=_ticket.id).last()
                if _storeid == 1:
                    s_daghi.serialmasterdaghi = serial
                else:
                    s_daghi.serialpinpaddaghi = serial
                s_daghi.save()

                messages.success(request, 'ذخیره قطعه داغی انجام شد.')
                return redirect('base:CrudeTickets')
                # return render(request, 'store/daghi.html',
                #               {'formpermmision': formpermmision, 'ok': 1, 'form': form, 'tickets': tickets})
            else:
                messages.error(request, 'تصویر بدرستی بارگذاری نشد')
                return TemplateResponse(request, 'store/daghi.html',
                                        {'ok': 1, 'form': form, 'tickets': tickets})
        try:

            store = StoreList.objects.get(serial=serial, statusstore_id=_storeid)
            if store.status_id == 6 and store.getuser_id == request.user.owner.id:
                messages.error(request,
                               'این قطعه در لیست داغی های شما وجود دارد ، امکان ثبت مجدد بعنوان داغی وجود ندارد.')
                return redirect(url)
            zonen = "-"
            if store.zone_id:
                zonen = store.zone.name
            if store.status_id == 5 and store.zone_id == request.user.owner.zone_id:
                store.status_id = 6
                store.getuser_id = request.user.owner.id
                store.owner_id = request.user.id
                store.zone_id = request.user.owner.zone_id
                store.save()
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information=f"ارسال قطعه داغی تیکت{_ticket.id} ",
                                            status_id=6, description=f' به پشتیبان  {request.user.owner} ')
                daghi_mande = Owner.objects.get(id=request.user.owner.id)
                daghi_mande.daghimande = False
                daghi_mande.save()
                _ticket.isdaghi = False
                _ticket.serialdaghi = serial
                _ticket.save()
                messages.success(request, 'ذخیره قطعه داغی انجام شد.')
                return redirect('base:CrudeTickets')
            else:
                messages.error(request, 'این قطعه در وضعیت ' + str(store.status.name) + ' در منطقه ' + str(
                    zonen) + ' وجود دارد امکان ثبت بعنوان داغی وجود ندارد لطفا از قسمت سابقه قطعه بررسی کنید')
                payam = 'این قطعه در وضعیت ' + str(store.status.name) + ' در منطقه ' + str(
                    zonen) + ' وجود دارد امکان ثبت بعنوان داغی وجود ندارد لطفا از قسمت سابقه قطعه بررسی کنید'
                return TemplateResponse(request, 'store/daghi.html',
                                        {'serial': serial, 'ok': 2, 'payam': payam,
                                         'tickets': tickets,
                                         'form': form,
                                         'storeid': int(storeid)})
        except ObjectDoesNotExist:
            try:
                store = StoreList.objects.create(serial=serial, getuser_id=request.user.owner.id,
                                                 owner_id=request.user.id,
                                                 zone_id=request.user.owner.zone_id, statusstore_id=_storeid,
                                                 status_id=6,
                                                 uniq=str(serial) + "-" + str(store))
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information=f"ارسال قطعه داغی(جدید)  تیکت {_ticket.id}",
                                            status_id=6, description=f' به پشتیبان  {request.user.owner} ')
                daghi_mande = Owner.objects.get(id=request.user.owner.id)
                daghi_mande.daghimande = False
                daghi_mande.save()
                _ticket.isdaghi = False
                _ticket.serialdaghi = serial
                _ticket.save()
                messages.success(request, 'ذخیره قطعه داغی انجام شد.')
                return redirect('base:CrudeTickets')
            except IntegrityError:
                messages.error(request, 'این قطعه در وضعیت ارسال به کارگاه تعمیر میباشد ')
                return TemplateResponse(request, 'store/daghi.html',
                                        {'serial': serial, 'storeid': int(storeid), 'ok': 2,
                                         'form': form, 'tickets': tickets,
                                         'payam': 'این قطعه در وضعیت ارسال به کارگاه تعمیر میباشد '})

    return TemplateResponse(request, 'store/daghi.html',
                            {'ok': 1, 'form': form, 'tickets': tickets})


def plombtozone(request):
    return render(request, 'plomb/addplomb.html')


def dastmozdtoexcel(request, _id, zone):
    my_path = f'report.xlsx'  # Path
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    _font = Font(bold=True)
    _fonttitr = Font(bold=True, size=20)
    mylist = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
              'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "لیست حقوق و دستمزد تکنسین های مناطق"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"
    if zone == 0:
        payroll = Payroll.objects.all()
    else:
        payroll = Payroll.objects.filter(tek__zone_id=zone)

    # maxcount = Payroll.objects.values('tek_id').filter(period_id=id).annotate(tedad=Count('id')).order_by('-tedad')[0]
    maxcount = payroll.values('paybaseparametrs_id', 'paybaseparametrs__name').filter(period_id=_id,
                                                                                      paybaseparametrs__isshow=True).annotate(
        tedad=Count('id')).order_by('paybaseparametrs_id')

    ws1.merge_cells(f'A1:{mylist[maxcount.count() + 6]}1')
    mah = Mount.objects.get(id=_id)
    ws1["A1"] = 'دوره  ' + str(mah.mount)
    ws1["A1"].font = _fonttitr

    ws1.merge_cells('A3:A3')
    ws1["A3"] = "ردیف"
    ws1["A3"].font = _font

    ws1.merge_cells('B3:B3')
    ws1["B3"] = "منطقه"
    ws1["B3"].font = _font

    ws1.merge_cells('C3:C3')
    ws1["C3"] = "نام و نام خانوادگی"
    ws1["C3"].font = _font

    ws1.merge_cells('D3:D3')
    ws1["D3"] = "کد ملی"
    ws1["D3"].font = _font

    ws1.merge_cells('E3:E3')
    ws1["E3"] = "کارکرد ماه"
    ws1["E3"].font = _font

    ws1.merge_cells('F3:F3')
    ws1["F3"] = "پایه حقوق + سنوات روزانه"
    ws1["F3"].font = _font

    i = 6
    for item in maxcount:
        ws1.merge_cells(f'{mylist[i]}3:{mylist[i]}3')
        ws1[f"{mylist[i]}3"] = item['paybaseparametrs__name']
        ws1[f"{mylist[i]}3"].font = _font
        i += 1

    ws1.merge_cells(f'{mylist[i]}3:{mylist[i]}3')
    ws1[f"{mylist[i]}3"] = 'جمع کل'
    ws1[f"{mylist[i]}3"].font = _font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(18.25)
    ws1.column_dimensions['D'].width = float(20.25)
    ws1.column_dimensions['E'].width = float(20.35)
    ws1.column_dimensions['F'].width = float(20.25)
    ws1.column_dimensions['G'].width = float(20.25)
    ws1.column_dimensions['H'].width = float(20.25)
    ws1.column_dimensions['I'].width = float(20.25)
    ws1.column_dimensions['J'].width = float(20.25)
    ws1.column_dimensions['K'].width = float(20.25)
    ws1.column_dimensions['L'].width = float(20.25)
    ws1.column_dimensions['M'].width = float(20.25)
    ws1.column_dimensions['n'].width = float(20.25)
    ws1.column_dimensions['o'].width = float(20.25)
    ws1.column_dimensions['p'].width = float(20.25)
    ws1.column_dimensions['q'].width = float(20.25)
    ws1.column_dimensions['r'].width = float(20.25)
    ws1.column_dimensions['s'].width = float(20.25)
    ws1.column_dimensions['t'].width = float(20.25)
    ws1.column_dimensions['u'].width = float(20.25)
    ws1.column_dimensions['v'].width = float(20.25)
    ws1.column_dimensions['C'].rightToLeft = True

    thinborder = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    my_font = Font(size=14, bold=True)
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')
    i = 0

    owners = payroll.values('tek_id').filter(period_id=_id).annotate(tedad=Count('id')).order_by('tek__zone_id')

    for owner in owners:
        _teklist = []
        if zone == 0:
            _list = Payroll.objects.filter(tek_id=owner['tek_id'], period_id=_id, paybaseparametrs__isshow=True)

        else:
            _list = Payroll.objects.filter(tek_id=owner['tek_id'], period_id=_id, tek__zone_id=zone,
                                           paybaseparametrs__isshow=True).order_by('paybaseparametrs_id')
        i += 1
        _sum = 0
        ownerinfo = Owner.objects.get(id=owner['tek_id'])
        _teklist.append(i)
        _teklist.append(ownerinfo.zone.name)
        _teklist.append(ownerinfo.get_full_name())
        _teklist.append(ownerinfo.codemeli)
        paykarkard = Payroll.objects.filter(tek_id=owner['tek_id'], period_id=_id,
                                            paybaseparametrs_id=7).last()
        if paykarkard:
            _teklist.append(paykarkard.count)
        else:
            _teklist.append(0)
        paykarkard = TekKarkard.objects.filter(tek_id=owner['tek_id'], period_id=_id,
                                               ).last()
        if paykarkard:
            _teklist.append(paykarkard.Dastmozd_rozane)
        else:
            _teklist.append(0)
        for item in maxcount:
            _teklist.append(0)
            o = len(_teklist)
            for onelist in _list:
                if int(item['paybaseparametrs_id']) == onelist.paybaseparametrs_id:
                    _teklist[o - 1] = onelist.price
                    _sum += onelist.price

        _teklist.append(_sum)
        d = _teklist
        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thinborder

    for cell in ws1["3:3"]:
        cell.font = my_font
        cell.fill = my_fill
        cell.border = thinborder

    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=5)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''
    wb.save(response)
    return response


def tadiltoexcel(request, _id, zone):
    sump = 0
    _filter = ""
    my_path = f'report.xlsx'  # Path
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    _font = Font(bold=True)
    _fonttitr = Font(bold=True, size=20)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "لیست حقوق و دستمزد تکنسین های مناطق"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:T1')
    mah = Mount.objects.get(id=_id)
    ws1["A1"] = 'دوره  ' + str(mah.mount)
    ws1["A1"].font = _fonttitr
    ws1.merge_cells('A3:A3')
    ws1["A3"] = "ردیف"
    ws1["A3"].font = _font
    ws1.merge_cells('B3:B3')
    ws1["B3"] = ZONE_NAME
    ws1["B3"].font = _font
    ws1.merge_cells('C3:C3')
    ws1["C3"] = "نام تکنسین"
    ws1["C3"].font = _font

    ws1.merge_cells('D3:D3')
    ws1["D3"] = "کد ملی"
    ws1["D3"].font = _font

    ws1.merge_cells('E3:E3')
    ws1["E3"] = "دستمزد ماهیانه"
    ws1["E3"].font = _font

    ws1.merge_cells('F3:F3')
    ws1["F3"] = "فوق العاده کمک رفاهی و ایاب ذهاب"
    ws1["F3"].font = _font

    ws1.merge_cells('G3:G3')
    ws1["G3"] = "حق مسکن و خواربار"
    ws1["G3"].font = _font

    ws1.merge_cells('H3:H3')
    ws1["H3"] = "بن کالای اساسی"
    ws1["H3"].font = _font

    ws1.merge_cells('I3:I3')
    ws1["I3"] = "کمک هزینه غذا"
    ws1["I3"].font = _font

    ws1.merge_cells('J3:J3')
    ws1["J3"] = "حق اولاد"
    ws1["J3"].font = _font

    ws1.merge_cells('K3:K3')
    ws1["K3"] = "ضریب اتلاف محاسبه غیر تمام وقت"
    ws1["K3"].font = _font

    ws1.merge_cells('L3:L3')
    ws1["L3"] = "اضافه کاری"
    ws1["L3"].font = _font

    ws1.merge_cells('M3:M3')
    ws1["M3"] = "فوق العاده جذب مشاغل"
    ws1["M3"].font = _font

    ws1.merge_cells('N3:N3')
    ws1["N3"] = "پایه سنوات"
    ws1["N3"].font = _font

    ws1.merge_cells('N3:N3')
    ws1["O3"] = "جمع کل"
    ws1["N3"].font = _font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(18.25)
    ws1.column_dimensions['D'].width = float(20.25)
    ws1.column_dimensions['E'].width = float(20.35)
    ws1.column_dimensions['F'].width = float(20.25)
    ws1.column_dimensions['G'].width = float(20.25)
    ws1.column_dimensions['H'].width = float(20.25)
    ws1.column_dimensions['I'].width = float(20.25)
    ws1.column_dimensions['J'].width = float(20.25)
    ws1.column_dimensions['K'].width = float(20.25)
    ws1.column_dimensions['L'].width = float(20.25)
    ws1.column_dimensions['M'].width = float(20.25)
    ws1.column_dimensions['N'].width = float(20.25)
    ws1.column_dimensions['C'].rightToLeft = True

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    my_font = Font(size=14, bold=True)
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')
    i = 0
    thislist = []

    owners = Owner.objects.filter(role__role='tek').order_by('zone_id')

    for owner in owners:
        dmah = 0
        fogh = 0
        maskan = 0
        bon = 0
        ghaza = 0
        olad = 0
        bazdehi = 0
        fani = 0
        ezafe = 0
        if zone == 0:
            _list = Payroll.objects.filter(tek_id=owner.id, accepttedad=_id)

        else:
            _list = Payroll.objects.filter(tek_id=owner.id, accepttedad=_id, tek__zone_id=zone)

        if _list.count() > 0:
            for row in _list:
                if row.paybaseparametrs_id == 7:
                    dmah += row.price
                if row.paybaseparametrs_id == 10:
                    fogh += row.price
                if row.paybaseparametrs_id == 11:
                    maskan += row.price
                if row.paybaseparametrs_id == 12:
                    bon += row.price
                if row.paybaseparametrs_id == 17:
                    ghaza += row.price
                if row.paybaseparametrs_id == 13:
                    olad += row.price
                if row.paybaseparametrs_id == 8:
                    bazdehi = row.price
                if row.paybaseparametrs_id == 9:
                    fani = row.price
                if row.paybaseparametrs_id == 16:
                    ezafe = row.price
                sump = dmah + fogh + maskan + bon + ghaza + olad + bazdehi + fani + ezafe
            mydict = {
                "zone": owner.zone.name,
                "name": owner.name + ' ' + owner.lname,
                "codemeli": owner.codemeli,
                "dmah": dmah,
                "fogh": fogh,
                "maskan": maskan,
                "bon": bon,
                "ghaza": ghaza,
                "olad": olad,
                "bazdehi": bazdehi,
                "fani": fani,
                "ezafe": ezafe,
                "sump": sump,
            }
            thislist.append(mydict)
    for row in thislist:
        i += 1
        d = [i, str(row['zone']), str(row['name']), str(row['codemeli']), (row['dmah']), (row['fogh']),
             (row['maskan']), (row['bon']), (row['ghaza']),
             (row['olad']), (row['bazdehi']), (row['fani']), (row['ezafe']), (row['sump'])
             ]
        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    for cell in ws1["3:3"]:
        cell.font = my_font
        cell.fill = my_fill
        cell.border = thin_border

    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=5)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''
    wb.save(response)
    return response


def storetoexcel(request, _id, st):
    add_to_log(request, 'ارسال لیست قطعات به اکسل  ', 0)
    my_path = f'report.xlsx'
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = EXCEL_EXPORT_FILE + my_path
    _font = Font(bold=True)
    _fonttitr = Font(bold=True, size=20)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "لیست قطعات ارسالی"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:F1')
    ws1["A1"] = 'لیست قطعات ارسالی '
    ws1["A1"].font = _fonttitr

    ws1.merge_cells('A3:A3')
    ws1["A3"] = "ردیف"
    ws1["A3"].font = _font

    ws1.merge_cells('B3:B3')
    ws1["B3"] = "سریال"
    ws1["B3"].font = _font

    ws1.merge_cells('C3:C3')
    ws1["C3"] = "نوع قطعه"
    ws1["C3"].font = _font

    ws1.merge_cells('D3:D3')
    ws1["D3"] = ZONE_NAME
    ws1["D3"].font = _font

    ws1.merge_cells('E3:E3')
    ws1["E3"] = ""
    ws1["E3"].font = _font

    ws1.merge_cells('F3:F3')
    ws1["F3"] = ""
    ws1["F3"].font = _font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(28.25)
    ws1.column_dimensions['D'].width = float(28.25)
    ws1.column_dimensions['E'].width = float(28.25)
    ws1.column_dimensions['F'].width = float(28.25)
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    my_font = Font(size=14, bold=True)
    my_fill = PatternFill(
        fill_type='solid', start_color='FFFF00')
    i = 0

    lists = StoreHistory.objects.filter(baseroot=_id).order_by('store__statusstore_id')
    for _list in lists:
        gettek = StoreHistory.objects.filter(store_id=_list.store_id, status_id=6).last()
        try:
            _tek = gettek.owner.name + " " + gettek.owner.lname
            _date = gettek.normal_date()
        except:
            _tek = "نامشخص"
            _date = "نامشخص"
        i += 1
        _tek = _tek if _list.status_id == 7 else ""
        _date = _date if _list.status_id == 7 else ""

        if st == 1:
            d = [i, str(_list.store.serial), str(_list.store.statusstore.name), str(_list.store.zone.name), _tek, _date
                 ]
        else:
            d = [i, str(_list.store.serial), str(_list.store.statusstore.name), str(_list.owner.zone.name), _tek, _date
                 ]
        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    for cell in ws1["3:3"]:
        cell.font = my_font
        cell.fill = my_fill
        cell.border = thin_border
    max_row = ws1.max_row
    total_cost_cell = ws1.cell(row=max_row + 2, column=5)
    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
    total_cost_cell.value = ''
    total_cost_cell2.value = ''
    wb.save(response)
    return response


@cache_permission('zonelistre')
def zonelistresid(request):
    _list = None
    _sum = None
    _filter = ""
    add_to_log(request, f'مشاهده فرم لیست قطعات ارسالی ', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    select3 = str(request.GET.get('select3'))
    vore = request.GET.get('vore')

    if select3 == "None":
        select3 = "2"
    if not select3:
        select3 = "2"

    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"

    else:

        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"

    if request.user.owner.role.role == 'zone':
        if select3 == '1':
            _list = Store.objects.filter(create__gte=datein, create__lte=dateout, status__in=[1, 2, 3],
                                         zone_id=request.user.owner.zone_id).order_by('-id')
        if select3 == '2':
            _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, status__in=[1, 2, 3],
                                         zone_id=request.user.owner.zone_id).order_by('-id')
        if select3 == '3':
            _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout, status__in=[1, 2, 3],
                                         zone_id=request.user.owner.zone_id).order_by('-id')
        _filter = StoreFilters(request.GET, queryset=_list)
        _list = _filter.qs
    if request.user.owner.role.role in ['mgr', 'setad', 'posht']:
        if select3 == '1':
            _list = Store.objects.filter(create__gte=datein, create__lte=dateout, status__in=[1, 2, 3]
                                         ).order_by('-id')
            _sum = Store.objects.values('storage_id', 'storage__name').filter(create__gte=datein, create__lte=dateout,
                                                                              status__in=[1, 2, 3]
                                                                              ).annotate(master=Sum('master'),
                                                                                         pinpad=Sum('pinpad'))
        if select3 == '2':
            _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, status__in=[1, 2, 3]
                                         ).order_by('-id')
            _sum = Store.objects.values('storage_id', 'storage__name').filter(marsole_date__gte=datein,
                                                                              marsole_date__lte=dateout,
                                                                              status__in=[1, 2, 3]
                                                                              ).annotate(master=Sum('master'),
                                                                                         pinpad=Sum('pinpad'))
        if select3 == '3':
            _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout, status__in=[1, 2, 3]
                                         ).order_by('-id')
            _sum = Store.objects.values('storage_id', 'storage__name').filter(resid_date__gte=datein,
                                                                              resid_date__lte=dateout,
                                                                              status__in=[1, 2, 3]
                                                                              ).annotate(master=Sum('master'),
                                                                                         pinpad=Sum('pinpad'))

        _filter = StoreFilters(request.GET, queryset=_list)
        if _filter.data:
            _list = _filter.qs
        else:
            _list = _list.filter(storage_id=request.user.owner.defaultstorage)

    if vore == "2":
        add_to_log(request, 'ارسال آمار لیست ارسالی ها به اکسل  ', 0)
        my_path = "sendstore.xlsx"
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = 'attachment; filename=' + my_path
        font = Font(bold=True)
        fonttitr = Font(bold=True, size=20)
        fonttitr2 = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active  # work with default worksheet
        ws1.title = "گزارش لیست قطعات ارسالی "
        ws1.sheet_view.rightToLeft = True
        ws1.firstFooter.center.text = "ali"
        ws1.merge_cells('A1:j1')

        ws1["A1"] = f'گزارش لیست قطعات ارسالی تاریخ   {today}'
        ws1["A1"].font = fonttitr

        ws1.merge_cells('A2:A2')
        ws1["A2"] = "ردیف"
        ws1["A2"].font = font

        ws1.merge_cells('B2:B2')
        ws1["B2"] = "تامین کننده"
        ws1["B2"].font = fonttitr2

        ws1.merge_cells('C2:C2')
        ws1["C2"] = "منطقه "
        ws1["C2"].font = font

        ws1.merge_cells('D2:D2')
        ws1["D2"] = " تاریخ تخصیص"
        ws1["D2"].font = font

        ws1.merge_cells('E2:E2')
        ws1["E2"] = "تعداد کارتخوان"
        ws1["E2"].font = font

        ws1.merge_cells('F2:F2')
        ws1["F2"] = "تعداد صفحه کلید"
        ws1["F2"].font = font

        ws1.merge_cells('G2:G2')
        ws1["G2"] = "تاریخ ارسال از کارگاه"
        ws1["G2"].font = font

        ws1.merge_cells('H2:H2')
        ws1["H2"] = " تاریخ رسید منطقه "
        ws1["H2"].font = font

        ws1.merge_cells('I2:I2')
        ws1["I2"] = " زمان تخصیص تا رسید"
        ws1["I2"].font = font

        ws1.merge_cells('J2:J2')
        ws1["J2"] = "زمان ارسال تا رسید"
        ws1["J2"].font = font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(15.25)
        ws1.column_dimensions['D'].width = float(25.25)
        ws1.column_dimensions['E'].width = float(15.25)
        ws1.column_dimensions['F'].width = float(18.25)
        ws1.column_dimensions['G'].width = float(35.25)
        ws1.column_dimensions['H'].width = float(30.25)
        ws1.column_dimensions['I'].width = float(25.25)
        ws1.column_dimensions['J'].width = float(22.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        myfont = Font(size=14, bold=True)  # font styles
        my_fill = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        my_fill2 = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        i = 0

        for item in _list:
            i += 1
            d = [i, str(item.storage.name), str(item.zone.name), str(item.normal_date()), str(item.master),
                 str(item.pinpad),
                 str(item.mdate()) + ' - ' + str(item.marsole), str(item.rdate()), str(item.get_takhsisdate_diff()),
                 str(item.get_date_diff())]

            ws1.append(d)

        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        i += 4
        for cell in ws1[i:i]:  # Last row
            cell.font = myfont
            cell.fill = my_fill2
            cell.border = thin_border

        for cell in ws1["2:2"]:  # First row
            cell.font = myfont
            cell.fill = my_fill
            cell.border = thin_border
        wb.save(response)
        return response
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']

    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:
        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)
        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, 'store/zonelistresid.html',
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'listsum': _sum,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('sendinkargah')
def sendinkargah(request):
    _list = None
    _sum = None
    _filter = ""
    add_to_log(request, f'مشاهده فرم لیست قطعات ارسالی از کارگاه', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    select3 = str(request.GET.get('select3'))
    vore = request.GET.get('vore')

    if select3 == "None":
        select3 = "2"
    if not select3:
        select3 = "2"

    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"

    else:

        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"

    if request.user.owner.role.role == 'zone':
        if select3 == '1':
            _list = Store.objects.filter(create__gte=datein, create__lte=dateout, status__in=[1, 2, 3],
                                         storage_id=request.user.owner.storage_id).order_by('-id')
        if select3 == '2':
            _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, status__in=[1, 2, 3],
                                         storage_id=request.user.owner.storage_id).order_by('-id')
        if select3 == '3':
            _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout, status__in=[1, 2, 3],
                                         storage_id=request.user.owner.storage_id).order_by('-id')
        _filter = StoreFilters(request.GET, queryset=_list)
        _list = _filter.qs
    if request.user.owner.role.role in ['mgr', 'setad', 'posht']:
        if select3 == '1':
            _list = Store.objects.filter(create__gte=datein, create__lte=dateout, status__in=[1, 2, 3]
                                         ).order_by('-id')
            _sum = Store.objects.values('storage_id', 'storage__name').filter(create__gte=datein, create__lte=dateout,
                                                                              status__in=[1, 2, 3]
                                                                              ).annotate(master=Sum('master'),
                                                                                         pinpad=Sum('pinpad'))
        if select3 == '2':
            _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, status__in=[1, 2, 3]
                                         ).order_by('-id')
            _sum = Store.objects.values('storage_id', 'storage__name').filter(marsole_date__gte=datein,
                                                                              marsole_date__lte=dateout,
                                                                              status__in=[1, 2, 3]
                                                                              ).annotate(master=Sum('master'),
                                                                                         pinpad=Sum('pinpad'))
        if select3 == '3':
            _list = Store.objects.filter(resid_date__gte=datein, resid_date__lte=dateout, status__in=[1, 2, 3]
                                         ).order_by('-id')
            _sum = Store.objects.values('storage_id', 'storage__name').filter(resid_date__gte=datein,
                                                                              resid_date__lte=dateout,
                                                                              status__in=[1, 2, 3]
                                                                              ).annotate(master=Sum('master'),
                                                                                         pinpad=Sum('pinpad'))

        _filter = StoreFilters(request.GET, queryset=_list)
        if _filter.data:
            _list = _filter.qs
        else:
            _list = _list.filter(storage_id=request.user.owner.defaultstorage)

    if vore == "2":
        add_to_log(request, 'ارسال آمار لیست ارسالی ها به اکسل  ', 0)
        my_path = "sendstore.xlsx"
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = 'attachment; filename=' + my_path
        font = Font(bold=True)
        fonttitr = Font(bold=True, size=20)
        fonttitr2 = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active  # work with default worksheet
        ws1.title = "گزارش لیست قطعات ارسالی "
        ws1.sheet_view.rightToLeft = True
        ws1.firstFooter.center.text = "ali"
        ws1.merge_cells('A1:j1')

        ws1["A1"] = f'گزارش لیست قطعات ارسالی تاریخ   {today}'
        ws1["A1"].font = fonttitr

        ws1.merge_cells('A2:A2')
        ws1["A2"] = "ردیف"
        ws1["A2"].font = font

        ws1.merge_cells('B2:B2')
        ws1["B2"] = "تامین کننده"
        ws1["B2"].font = fonttitr2

        ws1.merge_cells('C2:C2')
        ws1["C2"] = "منطقه "
        ws1["C2"].font = font

        ws1.merge_cells('D2:D2')
        ws1["D2"] = " تاریخ تخصیص"
        ws1["D2"].font = font

        ws1.merge_cells('E2:E2')
        ws1["E2"] = "تعداد کارتخوان"
        ws1["E2"].font = font

        ws1.merge_cells('F2:F2')
        ws1["F2"] = "تعداد صفحه کلید"
        ws1["F2"].font = font

        ws1.merge_cells('G2:G2')
        ws1["G2"] = "تاریخ ارسال از کارگاه"
        ws1["G2"].font = font

        ws1.merge_cells('H2:H2')
        ws1["H2"] = " تاریخ رسید منطقه "
        ws1["H2"].font = font

        ws1.merge_cells('I2:I2')
        ws1["I2"] = " زمان تخصیص تا رسید"
        ws1["I2"].font = font

        ws1.merge_cells('J2:J2')
        ws1["J2"] = "زمان ارسال تا رسید"
        ws1["J2"].font = font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(15.25)
        ws1.column_dimensions['D'].width = float(25.25)
        ws1.column_dimensions['E'].width = float(15.25)
        ws1.column_dimensions['F'].width = float(18.25)
        ws1.column_dimensions['G'].width = float(35.25)
        ws1.column_dimensions['H'].width = float(30.25)
        ws1.column_dimensions['I'].width = float(25.25)
        ws1.column_dimensions['J'].width = float(22.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        myfont = Font(size=14, bold=True)  # font styles
        my_fill = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        my_fill2 = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        i = 0

        for item in _list:
            i += 1
            d = [i, str(item.storage.name), str(item.zone.name), str(item.normal_date()), str(item.master),
                 str(item.pinpad),
                 str(item.mdate()) + ' - ' + str(item.marsole), str(item.rdate()), str(item.get_takhsisdate_diff()),
                 str(item.get_date_diff())]

            ws1.append(d)

        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        i += 4
        for cell in ws1[i:i]:  # Last row
            cell.font = myfont
            cell.fill = my_fill2
            cell.border = thin_border

        for cell in ws1["2:2"]:  # First row
            cell.font = myfont
            cell.fill = my_fill
            cell.border = thin_border
        wb.save(response)
        return response
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']

    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:
        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)
        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, 'store/zonelistresid.html',
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'listsum': _sum,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('zonelistre')
def zonelistdaghi(request):
    _list = None
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    if len(datein) < 10:

        datein = "2023-01-01"
        dateout = "9999-12-30"
    else:

        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"

    if request.user.owner.role.role == 'zone':
        _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, status__in=[7, 8],
                                     zone_id=request.user.owner.zone_id).order_by('-id')
    if request.user.owner.role.role in ['mgr', 'setad', 'posht']:
        _list = Store.objects.filter(marsole_date__gte=datein, marsole_date__lte=dateout, status__in=[7, 8]
                                     ).order_by('-id')
    _filter = StoreFilters(request.GET, queryset=_list)
    _list = _filter.qs
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']
    if _filter.data:
        paginator = Paginator(_list, 1000)
    else:
        paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:
        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)
        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, 'store/zonelistdaghi.html',
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getinformation(request):
    if request.method == 'POST':
        val = request.POST.get('val')
        count_gs = GsModel.objects.filter(area__zone_id=val, active=True).count()
        count_pump = Pump.objects.filter(gs__area__zone_id=val, active=True).count()
        count_ticket = Ticket.objects.filter(gs__area__zone_id=val, status_id=1,
                                             failure__failurecategory_id__in=[1010, 1011]).count()
        count_ticket += Ticket.objects.filter(gs__area__zone_id=val, status_id=1,
                                              failure_id=1045).count()

        count_master = Ticket.objects.filter(gs__area__zone_id=val, status_id=1, failure__failurecategory_id=1010,
                                             failure__isnazel=True).count()

        count_master += Ticket.objects.filter(gs__area__zone_id=val, status_id=1, failure_id=1045
                                              ).count()

        count_pinpad = Ticket.objects.filter(gs__area__zone_id=val, status_id=1, failure__failurecategory_id=1011,
                                             failure__isnazel=True).count()
        count_master_store = StoreList.objects.filter(zone_id=val, status_id__in=[3, 4], statusstore_id=1).count()
        count_pinpad_store = StoreList.objects.filter(zone_id=val, status_id__in=[3, 4], statusstore_id=2).count()

        store = Store.objects.filter(zone_id=val, status__in=[1, 2, 3, 4]).order_by('-id')[:5]
        _list = []
        for item in store:
            if item.marsole_date:
                ersal = item.mdate()
            else:
                ersal = 'ارسال نشده'
            if item.resid_date:
                resid = item.rdate()
            else:
                resid = 'رسید نشده'
            if item.marsole:
                marsole = item.marsole
            else:
                marsole = ''
            mydic = {
                'taikhtakhsis': item.pdate(),
                'storage': item.storage.name,
                'master': item.master,
                'pinpad': item.pinpad,
                'status': item.status.name + ' ( ' + marsole + ' )',
                'tarikhmarsole': ersal,
                'resid_date': resid,

            }
            _list.append(mydic)

        return JsonResponse(
            {"message": "success", 'store': _list, 'count_pinpad': count_pinpad, 'count_master': count_master,
             'count_ticket': count_ticket, 'count_pump': count_pump, 'count_gs': count_gs,
             'count_master_store': count_master_store, 'count_pinpad_store': count_pinpad_store})


def test(request):
    _list = StoreList.objects.all()
    for i in _list:

        i.serial = i.serial.replace('۰', '0')
        i.serial = i.serial.replace('۱', '1')
        i.serial = i.serial.replace('۲', '2')
        i.serial = i.serial.replace('۳', '3')
        i.serial = i.serial.replace('۴', '4')
        i.serial = i.serial.replace('۵', '5')
        i.serial = i.serial.replace('۶', '6')
        i.serial = i.serial.replace('۷', '7')
        i.serial = i.serial.replace('۸', '8')
        i.serial = i.serial.replace('۹', '9')
        try:
            i.save()
        except IntegrityError:
            continue

    return redirect(HOME_PAGE)


@cache_permission('zonelistre')
def reporttekstore(request):
    _list = []
    add_to_log(request, f'مشاهده فرم وضعیت قطعات تکنسین ', 0)
    zone = Zone.objects_limit.all()
    if request.user.owner.role.role in ['mgr', 'setad']:
        owner = Owner.objects.filter(role__role='tek', active=True).order_by('zone_id')
        if request.method == 'POST':
            zoneid = request.POST.get('zone')
            owner = Owner.objects.filter(role__role='tek', active=True, zone_id=zoneid)

    else:
        owner = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True)
    sum_master_store = 0
    sum_master_post = 0
    sum_pinpad_store = 0
    sum_pinpad_post = 0
    sum_master_daghi = 0
    sum_pinpad_daghi = 0
    sum_count_master = 0
    sum_count_pinpad = 0
    sum_locl_rock = 0
    sum_locl_pomp = 0
    sum_locl_daghi = 0
    for item in owner:
        count_master = Ticket.objects.exclude(organization_id=4).filter(gs__gsowner__owner_id=item.id,
                                                                        status__status='open',
                                                                        failure__failurecategory_id=1010,
                                                                        Pump__status__status=True,
                                                                        gs__status__status=True,
                                                                        failure__isnazel=True).count()

        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__gsowner__owner_id=item.id,
                                                                        status__status='open',
                                                                        failure__failurecategory_id=1011,
                                                                        Pump__status__status=True,
                                                                        gs__status__status=True,
                                                                        failure__isnazel=True).count()

        master_store = StoreList.objects.filter(getuser_id=item.id, status_id=4, statusstore_id=1).count()
        master_post = StoreList.objects.filter(getuser_id=item.id, status_id=16, statusstore_id=1).count()
        pinpad_store = StoreList.objects.filter(getuser_id=item.id, status_id=4, statusstore_id=2).count()
        pinpad_post = StoreList.objects.filter(getuser_id=item.id, status_id=16, statusstore_id=2).count()
        master_daghi = StoreList.objects.filter(getuser_id=item.id, statusstore=1, status_id=6).count()
        pinpad_daghi = StoreList.objects.filter(getuser_id=item.id, statusstore=2, status_id=6).count()
        locl_pomp = LockModel.objects.filter(status_id=8, owner_id=item.id, ename='smart').count()
        locl_daghi = LockModel.objects.filter(status_id=6, idg_user_id=item.user_id, ename='smart').count()

        sum_master_store += master_store
        sum_master_post += master_post
        sum_pinpad_store += pinpad_store
        sum_pinpad_post += pinpad_post
        sum_master_daghi += master_daghi
        sum_pinpad_daghi += pinpad_daghi
        sum_count_master += count_master
        sum_count_pinpad += count_pinpad
        sum_locl_pomp += locl_pomp
        sum_locl_daghi += locl_daghi

        _dict = {
            'id': item.id,
            'name': item.name + ' ' + item.lname,
            'zone': item.zone.name,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master_store': master_store,
            'master_post': master_post,
            'pinpad_store': pinpad_store,
            'pinpad_post': pinpad_post,
            'master_daghi': master_daghi,
            'pinpad_daghi': pinpad_daghi,
            'locl_pomp': locl_pomp,
            'locl_daghi': locl_daghi,

        }
        _list.append(_dict)
    return TemplateResponse(request, 'store/reporttekstore.html',
                            {'sum_count_pinpad': sum_count_pinpad, 'sum_count_master': sum_count_master,
                             'sum_pinpad_daghi': sum_pinpad_daghi, 'sum_master_daghi': sum_master_daghi,
                             'sum_locl_pomp': sum_locl_pomp, 'sum_master_post': sum_master_post,
                             'sum_pinpad_post': sum_pinpad_post,
                             'sum_locl_daghi': sum_locl_daghi,
                             'sum_master_store': sum_master_store, 'sum_pinpad_store': sum_pinpad_store, 'list': _list,
                             'zone': zone})


@cache_permission('z_listre')
def reportzonestore(request, _id):
    _list = []
    kolnazel = 0
    add_to_log(request, f'وضعیت قطعات مناطق ', 0)
    zone = Zone.objects_limit.all()
    sum_count_master = 0
    sum_count_pinpad = 0
    sum_master_store = 0
    sum_pinpad_store = 0
    sum_master_daghi = 0
    sum_pinpad_daghi = 0
    sum_mstore_takhsis = 0
    sum_pstore_takhsis = 0
    sum_n_pinpad = 0
    sum_n_master = 0
    sum_master_store_noget = 0
    sum_pinpad_store_noget = 0
    sum_daghi_pinpad_in_storage = 0
    sum_daghi_pinpad_in_zone = 0
    sum_daghi_master_in_zone = 0
    sum_daghi_master_in_storage = 0

    for item in zone:
        count_master = Ticket.objects.exclude(organization_id=4).select_related('gs', 'status', 'Pump',
                                                                                'failure').filter(
            gs__area__zone_id=item.id,
            status__status='open',
            Pump__status__status=True,
            gs__status__status=True,
            failure__failurecategory_id=1010,
            failure__isnazel=True).count()

        count_pinpad = Ticket.objects.exclude(organization_id=4).select_related('gs', 'status', 'Pump',
                                                                                'failure').filter(
            gs__area__zone_id=item.id,
            status__status='open',
            Pump__status__status=True,
            gs__status__status=True,
            failure__failurecategory_id=1011,
            failure__isnazel=True).count()

        master_store = StoreList.objects.select_related('zone', 'status').filter(zone_id=item.id,
                                                                                 status_id__in=[3, 4, 16],
                                                                                 statusstore_id=1).count()
        pinpad_store = StoreList.objects.select_related('zone', 'status').filter(zone_id=item.id,
                                                                                 status_id__in=[3, 4, 16],
                                                                                 statusstore_id=2).count()
        store_takhsis = Store.objects.select_related('zone', 'status').filter(zone_id=item.id, status_id__in=[1, 9],
                                                                              ).aggregate(master=Sum('master'),
                                                                                          pinpad=Sum('pinpad'))
        store_noget = Store.objects.filter(zone_id=item.id, status_id=2,
                                           ).aggregate(master=Sum('master'), pinpad=Sum('pinpad'))

        daghi_master_in_zone = StoreList.objects.filter(zone_id=item.id, statusstore=1,
                                                        status_id__in=[6, 10, 11]).count()
        daghi_master_in_storage = StoreList.objects.filter(zone_id=item.id, statusstore=1, status_id__in=[8]).count()
        master_daghi = StoreList.objects.filter(zone_id=item.id, statusstore=1, status_id__in=[6, 10, 11, 8]).count()

        daghi_pinpad_in_zone = StoreList.objects.filter(zone_id=item.id, statusstore=2,
                                                        status_id__in=[6, 10, 11]).count()
        daghi_pinpad_in_storage = StoreList.objects.filter(zone_id=item.id, statusstore=2, status_id__in=[8]).count()
        pinpad_daghi = StoreList.objects.filter(zone_id=item.id, statusstore=2, status_id__in=[6, 10, 11, 8]).count()
        if store_noget['master']:
            mstore = store_noget['master']
        else:
            mstore = 0

        if store_noget['pinpad']:
            pstore = store_noget['pinpad']
        else:
            pstore = 0

        if store_takhsis['master']:
            mstore_takhsis = store_takhsis['master']
        else:
            mstore_takhsis = 0
        n_master = (master_store + int(mstore)) - count_master
        # n_master = (master_store + int(mstore) + int(mstore_takhsis)) - count_master
        if store_takhsis['pinpad']:
            pstore_takhsis = store_takhsis['pinpad']
        else:
            pstore_takhsis = 0
        n_pinpad = (pinpad_store + int(pstore)) - count_pinpad
        # n_pinpad = (pinpad_store + int(pstore) + int(pstore_takhsis)) - count_pinpad
        pump = Ticket.objects.select_related('failure').filter(status_id=1,
                                                               failure__failurecategory_id__in=[1010, 1011]).count()
        zonepump = Pump.objects.select_related('gs', 'status').filter(gs__area__zone_id=item.id,
                                                                      gs__status__status=True,
                                                                      status__status=True).count()
        kolnazel = Pump.objects.select_related('gs', 'status').filter(gs__status__status=True,
                                                                      status__status=True).count()

        _dict = {
            'id': item.id,
            'name': item.name,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master_store': master_store,
            'pinpad_store': pinpad_store,
            'daghi_master_in_zone': daghi_master_in_zone,
            'daghi_master_in_storage': daghi_master_in_storage,
            'master_daghi': master_daghi,
            'daghi_pinpad_in_zone': daghi_pinpad_in_zone,
            'daghi_pinpad_in_storage': daghi_pinpad_in_storage,
            'pinpad_daghi': pinpad_daghi,
            'n_pinpad': n_pinpad,
            'n_master': n_master,
            'master_store_noget': mstore,
            'pinpad_store_noget': pstore,
            'mstore_takhsis': mstore_takhsis,
            'pstore_takhsis': pstore_takhsis,
            'nesbat': round((((int(count_master) + int(count_pinpad)) / int(pump)) * 100), 2),
            'nesbat2': round((((int(count_master) + int(count_pinpad)) / int(kolnazel)) * 100), 2),
            'darsad': round((((int(count_master) + int(count_pinpad)) / int(zonepump)) * 100), 2),
            'zonepump': zonepump,

        }

        sum_count_master += count_master
        sum_count_pinpad += count_pinpad

        sum_master_store += master_store

        sum_pinpad_store += pinpad_store
        sum_daghi_master_in_zone += daghi_master_in_zone
        sum_daghi_master_in_storage += daghi_master_in_storage
        sum_master_daghi += master_daghi
        sum_daghi_pinpad_in_zone += daghi_pinpad_in_zone
        sum_daghi_pinpad_in_storage += daghi_pinpad_in_storage
        sum_pinpad_daghi += pinpad_daghi
        sum_mstore_takhsis += mstore_takhsis
        sum_pstore_takhsis += pstore_takhsis
        sum_n_pinpad += n_pinpad
        sum_n_master += n_master
        sum_master_store_noget += mstore
        sum_pinpad_store_noget += pstore

        _list.append(_dict)

    darsadkol = round((((sum_count_master + int(sum_count_pinpad)) / int(kolnazel)) * 100), 2)

    if _id == 1:
        _list = sorted(_list, key=itemgetter('nesbat'), reverse=True)
    elif _id == 2:
        _list = sorted(_list, key=itemgetter('nesbat2'), reverse=True)
    elif _id == 3:
        _list = sorted(_list, key=itemgetter('darsad'), reverse=True)
    return TemplateResponse(request, 'store/reportzonestore.html',
                            {'list': _list, 'darsadkol': darsadkol,
                             'zone': zone, 'sum_count_master': sum_count_master, 'id': _id,
                             'sum_count_pinpad': sum_count_pinpad,
                             'sum_master_store': sum_master_store, 'sum_pinpad_store': sum_pinpad_store,
                             'sum_master_daghi': sum_master_daghi, 'sum_pinpad_daghi': sum_pinpad_daghi,
                             'sum_mstore_takhsis': sum_mstore_takhsis, 'sum_pstore_takhsis': sum_pstore_takhsis,
                             'sum_n_pinpad': sum_n_pinpad, 'sum_daghi_master_in_zone': sum_daghi_master_in_zone,
                             'sum_daghi_master_in_storage': sum_daghi_master_in_storage,
                             'sum_daghi_pinpad_in_zone': sum_daghi_pinpad_in_zone,
                             'sum_daghi_pinpad_in_storage': sum_daghi_pinpad_in_storage,
                             'sum_n_master': sum_n_master, 'sum_master_store_noget': sum_master_store_noget,
                             'sum_pinpad_store_noget': sum_pinpad_store_noget})


def takhsistoexcel(request, page_num):
    add_to_log(request, 'تخصیص قطعه ارسال به اکسل  ', 0)
    my_path = f'Takhsis.xlsx'
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = EXCEL_EXPORT_FILE + my_path
    _font = Font(bold=True)
    _fonttitr = Font(bold=True, size=20)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "لیست قطعات تخصیصی"
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.orientation = 'landscape'
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:G1')

    ws1["A1"] = f'لیست قطعات تخصیصی تاریخ  {today}'
    ws1["A1"].font = _fonttitr

    ws1.merge_cells('A2:A2')
    ws1["A2"] = "ردیف"
    ws1["A2"].font = _font

    ws1.merge_cells('B2:B2')
    ws1["B2"] = "تامین کننده"
    ws1["B2"].font = _font

    ws1.merge_cells('C2:C2')
    ws1["C2"] = "به منطقه"
    ws1["C2"].font = _font

    ws1.merge_cells('D2:D2')
    ws1["D2"] = "تاریخ"
    ws1["D2"].font = _font

    ws1.merge_cells('E2:E2')
    ws1["E2"] = " کارتخوان"
    ws1["E2"].font = _font

    ws1.merge_cells('F2:F2')
    ws1["F2"] = " صفحه کلید"
    ws1["F2"].font = _font

    ws1.merge_cells('G2:G2')
    ws1["G2"] = "وضعیت"
    ws1["G2"].font = _font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(24.25)
    ws1.column_dimensions['D'].width = float(24.25)
    ws1.column_dimensions['E'].width = float(10.25)
    ws1.column_dimensions['F'].width = float(10.25)
    ws1.column_dimensions['G'].width = float(25.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    my_font = Font(size=14, bold=True)
    my_fill = PatternFill(
        fill_type='solid', start_color='2e6da4')
    my_fill2 = PatternFill(
        fill_type='solid', start_color='dadfe3')
    i = 0

    lists = Store.objects.filter(status_id__in=[1, 9], tarikh=datetime.date.today()).order_by('-tarikh')
    summaster = Store.objects.filter(status_id__in=[1, 9], tarikh=datetime.date.today()).aggregate(
        smaster=Sum('master'), spinpad=Sum('pinpad'))

    if page_num:
        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)
                if page_num[:3] == 'pre':
                    this_date = this_date - datetime.timedelta(days=1)
                if page_num[:3] == 'nex':
                    this_date = this_date + datetime.timedelta(days=1)

                lists = Store.objects.filter(status_id__in=[1, 9], tarikh=this_date).order_by('-tarikh')
                summaster = Store.objects.filter(status_id__in=[1, 9], tarikh=this_date).aggregate(
                    smaster=Sum('master'), spinpad=Sum('pinpad'))
                jd = JDate(this_date.strftime("%Y-%m-%d %H:%M:%S"))
                shdate = jd.format('Y/m/d')
                ws1["A1"] = 'لیست قطعات تخصیصی' + str(shdate)
                ws1["A1"].font = _fonttitr
        if page_num[:3] == 'all':
            lists = Store.objects.filter(status_id__in=[1, 9]).order_by('-tarikh')
            summaster = Store.objects.filter(status_id__in=[1, 9]).aggregate(
                smaster=Sum('master'), spinpad=Sum('pinpad'))

    for _list in lists:
        i += 1
        d = [i, str(_list.storage.name), str(_list.zone.name), str(_list.pdate()), str(_list.master), str(_list.pinpad),
             str(_list.status.name)
             ]
        ws1["A1"] = 'لیست قطعات تخصیصی' + str(_list.normal_date())
        ws1["A1"].font = _fonttitr
        if page_num[:3] == 'all':
            ws1["A1"] = 'لیست  مانده قطعات تخصیصی تا تاریخ ' + str(today)
            ws1["A1"].font = _fonttitr
        ws1.append(d)
    d = ['', '', '', 'جمع', str(summaster['smaster']), str(summaster['spinpad']), '']
    ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    i += 3
    for cell in ws1[i:i]:  # Last row
        cell.font = my_font
        cell.fill = my_fill2
        cell.border = thin_border

    for cell in ws1["2:2"]:  # First row
        cell.font = my_font
        cell.fill = my_fill
        cell.border = thin_border
    wb.save(response)
    return response


@cache_permission('reportstoe')
def reportstoe(request):
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    add_to_log(request, f'گزارش قطعات ', 0)
    zones = Zone.objects_limit.all() if request.user.owner.role.role in ['mgr', 'setad'] else Zone.objects.filter(
        id=request.user.owner.zone_id)
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        zone = request.POST.get('zone')
        store = request.POST.get('store')
        az = mdate
        ta = mdate2
        en_tarikh_in = az.split("/")
        en_tarikh_to = ta.split("/")
        en_tarikh_in = jdatetime.date(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                      year=int(en_tarikh_in[0])).togregorian()
        en_tarikh_to = jdatetime.date(day=int(en_tarikh_to[2]), month=int(en_tarikh_to[1]),
                                      year=int(en_tarikh_to[0])).togregorian()

        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        reside = []
        zone_id = int(zone) if request.user.owner.role.role in ['mgr', 'setad'] else request.user.owner.zone_id

        _reside = Store.objects.filter(resid_date__gte=en_tarikh_in, resid_date__lte=en_tarikh_to,
                                       zone_id=zone_id,
                                       status_id=3).aggregate(master=Sum('master'),
                                                              pinpad=Sum('pinpad'))
        if store == '1':
            tedad = _reside['master']
        else:
            tedad = _reside['pinpad']

        starterr = StoreHistory.objects.filter(starterror=True, create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                               owner__zone_id=zone_id, store__statusstore_id=store).count()
        mayob = StoreHistory.objects.filter(starterror=False, status_id=6, create__gte=en_tarikh_in,
                                            create__lte=en_tarikh_to, owner__zone_id=zone_id,
                                            store__statusstore_id=store).count()
        summayob = StoreHistory.objects.filter(status_id=6, create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                               owner__zone_id=zone_id,
                                               store__statusstore_id=store).count()
        installed = StoreHistory.objects.filter(status_id=5, create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                                owner__zone_id=zone_id,
                                                store__statusstore_id=store).count()
        send = StoreHistory.objects.filter(status_id=7, create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                           owner__zone_id=zone_id,
                                           store__statusstore_id=store).count()
        mojodi_tek_zone = StoreList.objects.filter(zone_id=zone_id, status_id=4,
                                                   statusstore_id=store).count()
        mojodi_zone = StoreList.objects.filter(zone_id=zone_id, status_id=3,
                                               statusstore_id=store).count()
        mojodi_kargah = StoreList.objects.filter(zone_id=zone_id, status_id__in=[9, 13],
                                                 statusstore_id=store).count()
        mojodi_sum = StoreList.objects.filter(zone_id=zone_id, status_id__in=[3, 9, 4],
                                              statusstore_id=store).count()

        daghi_tek_zone = StoreList.objects.filter(getuser__zone_id=zone_id, status_id=6,
                                                  statusstore_id=store).count()
        daghi_zone = StoreList.objects.filter(zone_id=zone_id, status_id__in=[10, 11],
                                              statusstore_id=store).count()

        daghi_kargah = StoreList.objects.filter(zone_id=zone_id, status_id__in=[8, 14],
                                                statusstore_id=store).count()
        daghi_sum = daghi_tek_zone + daghi_zone + daghi_kargah

        return TemplateResponse(request, 'store/reportstore.html',
                                {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'reside': reside,
                                 'starterr': starterr,
                                 'mayob': mayob, 'summayob': summayob, 'zones': zones, 'zone': int(zone),
                                 'installed': installed, 'mojodi_tek_zone': mojodi_tek_zone, 'mojodi_zone': mojodi_zone,
                                 'mojodi_kargah': mojodi_kargah, 'mojodi_sum': mojodi_sum, 'tedad': tedad,
                                 'store': int(store),
                                 'send': send, 'daghi_tek_zone': daghi_tek_zone, 'daghi_zone': daghi_zone,
                                 'daghi_kargah': daghi_kargah, 'daghi_sum': daghi_sum})
    return TemplateResponse(request, 'store/reportstore.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'zones': zones
                             })


def checkzonestore():
    store = Store.objects.all()
    for item in store:
        try:
            _history = StoreHistory.objects.filter(baseroot=item.id).last()
            if _history and item.zone_id != _history.store.zone.id:
                print(str(_history.store.serial) + "-" + str(item.id) + "-" + str(_history.store.id) + "-" + str(
                    _history.id))

        except ObjectDoesNotExist:
            continue
    return HttpResponse('ok')


@cache_permission('newstore')
def newstore(request):
    if request.method == 'POST':
        store = int(request.POST.get('store'))
        serial = request.POST.get('init')
        serial = check_serial(serial)
        add_to_log(request, f'ثبت قطعه جدید در منطقه{serial} ', 0)
        if not serial:
            url = request.META.get('HTTP_REFERER')
            messages.error(request, ADD_STORE_MSG)
            return redirect(url)
        try:
            store = StoreList.objects.get(serial=serial, statusstore_id=store)
            store.status_id = 3
            store.getuser_id = request.user.owner.id
            store.owner_id = request.user.id
            store.zone_id = request.user.owner.zone_id
            store.save()

        except ObjectDoesNotExist:
            store = StoreList.objects.create(serial=serial, getuser_id=request.user.owner.id, owner_id=request.user.id,
                                             zone_id=request.user.owner.zone_id, statusstore_id=store, status_id=3,
                                             uniq=str(serial) + "-" + str(store))
        StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                    information="ثبت اولیه قطعه توسط منطقه  ",
                                    status_id=3, description=f' به انبار منطقه  {request.user.owner.zone.name} ')
        messages.success(request, 'شماره سریال ثبت شد')
    return TemplateResponse(request, 'store/newstore.html', )


@cache_permission('senddaghi')
def senddaghi(request):
    add_to_log(request, f'مشاهده ارسال داغی ', 0)
    _list = None
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    if len(datein) < 10:

        datein = "2023-01-01"
        dateout = "9999-12-30"
    else:

        datein = datein.split("/")
        dateout = dateout.split("/")

        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"
    if request.user.owner.role.role == 'mgr':
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout, status_id=7)
    if request.user.owner.role.role == 'setad':
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout, storage__zone=None, status_id=7)

    if request.user.owner.role.role in 'zone,posht,tek':
        _list = Store.objects.filter(create__gte=datein, create__lte=dateout, zone_id=request.user.owner.zone_id,
                                     status_id=7)

    _filter = StoreFilters(request.GET, queryset=_list)
    _list = _filter.qs
    summaster = _list.aggregate(
        mastersum=Sum('master'))
    summaster = summaster['mastersum']
    sumpinpad = _list.aggregate(
        pinpadsum=Sum('pinpad'))
    sumpinpad = sumpinpad['pinpadsum']
    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
            summaster = _list.all().aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.all().aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
            summaster = _list.filter(create__date=datetime.datetime.today()).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=datetime.datetime.today()).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']

        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date - datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            summaster = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                mastersum=Sum('master'))
            summaster = summaster['mastersum']
            sumpinpad = _list.filter(create__date=this_date + datetime.timedelta(days=1)).aggregate(
                pinpadsum=Sum('pinpad'))
            sumpinpad = sumpinpad['pinpadsum']
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, 'store/senddaghi.html',
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'page_num': page_num,
                             'page_obj': page_obj, 'tedad': tedad, 'sumpinpad': sumpinpad,
                             'summaster': summaster})


@cache_permission('adddaghi')
def add_daghi_to_kargah(request):
    add_to_log(request, f'مشاهده فرم ارسال داغی به کارگاه ', 0)
    url = request.META.get('HTTP_REFERER')
    indate = ""
    t1 = ""
    tedad = 0
    zone = request.user.owner.zone_id
    _masters = StoreList.objects.filter(zone_id=zone, status_id=11, statusstore_id=1)
    _pinpads = StoreList.objects.filter(zone_id=zone, status_id=11, statusstore_id=2)
    _list = []
    for m in _masters:
        install_date = StoreHistory.objects.filter(store_id=m.id, status_id=5).last()
        if install_date:
            indate = install_date.create
            t1 = date(year=indate.year, month=indate.month, day=indate.day)
            install_date = install_date.pdate()
        else:
            tedad = ""
        daghi_date = StoreHistory.objects.filter(store_id=m.id, status_id=6).last()
        if daghi_date:
            outdate = daghi_date.create
            t2 = date(year=outdate.year, month=outdate.month, day=outdate.day)
            daghi_date = daghi_date.pdate()

            if daghi_date and install_date:
                tedad = (t2 - t1).days
            else:
                tedad = "نامعلوم"
            _dict = {
                'info': ':کارتخوان/ ',
                'serial': m.serial,
                'install_date': install_date,
                'daghi_date': daghi_date,
                'tedad': tedad
            }
            _list.append(_dict)
    for m in _pinpads:
        install_date = StoreHistory.objects.filter(store_id=m.id, status_id=5).last()
        if install_date:
            indate = install_date.create
            install_date = install_date.pdate()
        else:
            tedad = ""
        daghi_date = StoreHistory.objects.filter(store_id=m.id, status_id=6).last()
        if daghi_date:
            outdate = daghi_date.create
            daghi_date = daghi_date.pdate()
            if daghi_date and install_date:
                tedad = outdate - indate

            _dict = {
                'info': ':صفحه کلید ',
                'serial': m.serial,
                'install_date': install_date,
                'daghi_date': daghi_date,
                'tedad': tedad
            }
            _list.append(_dict)
    masters = _masters.count()
    pinpads = _pinpads.count()
    storages = ZoneToStorage.objects.filter(zone_id=request.user.owner.zone_id)
    if request.method == 'POST':
        marsole = request.POST.get('marsole')
        storage = request.POST.get('storage')
        _storage = Storage.objects.get(id=int(storage))
        if _storage.iszarib:
            if (masters + pinpads) % _storage.zarib != 0:
                messages.error(request, f'تعداد قطعه انتخابی  باید مضربی از  {_storage.zarib} باشد ')
                return redirect(url)

        mystore = Store.objects.create(owner_id=request.user.owner.id, tarikh=datetime.date.today(), pinpad=pinpads,
                                       master=masters, marsole=marsole, marsole_date=datetime.datetime.today(),
                                       status_id=7, zone_id=zone, storage_id=_storage.id)

        HistorySt.objects.create(store_id=mystore.id, owner_id=request.user.owner.id,
                                 status_id=7, description=f' برگشت قطعه به کارگاه   {_storage.name} ')
        for item in _masters:
            item.status_id = 7
            item.getuser_id = None
            item.gs_id = None
            item.zone_id = _storage.zone_id
            item.store_id = mystore.id
            item.save()
            StoreHistory.objects.create(store_id=item.id, owner_id=request.user.owner.id, baseroot=mystore.id,
                                        information="ارسال قطعه از منطقه " + str(request.user.owner.zone.name),
                                        status_id=7, description=f' به کارگاه  {_storage.name}', storage_id=_storage.id)
        for item in _pinpads:
            item.status_id = 7
            item.getuser_id = None
            item.gs_id = None
            item.zone_id = _storage.zone_id
            item.store_id = mystore.id
            item.save()
            StoreHistory.objects.create(store_id=item.id, owner_id=request.user.owner.id, baseroot=mystore.id,
                                        information="ارسال قطعه از منطقه " + str(request.user.owner.zone.name),
                                        status_id=7, description=f' به کارگاه  {_storage.name}', storage_id=_storage.id)

        messages.success(request, SUCCESS_MSG)
        return redirect('pay:senddaghi')

    context = {'masters': masters, 'pinpads': pinpads, 'storages': storages, "listmaster": _list,
               }
    return TemplateResponse(request, 'store/addsenddaghi.html', context)


def zoneorarea(request):
    _roleid = 0
    if request.user.owner.role.role in ['zone', 'engin']:
        _roleid = request.user.owner.zone_id
    elif request.user.owner.role.role == 'area':
        _roleid = request.user.owner.area_id
    elif request.user.owner.role.role == 'tek':
        _roleid = request.user.owner.id
    elif request.user.owner.role.role == 'gs':
        _roleid = request.user.owner.id
    return _roleid


@cache_permission('zonelistre')
def tek_amalkard(request):
    add_to_log(request, f'مشاهده فرم عملکرد پشتیبان ', 0)
    starscount = 0
    stars = 0
    mdate = startdate
    tek = 0
    mdate2 = today
    az = mdate
    ta = mdate2
    listticket = 0
    result = None
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if request.user.owner.role.role == 'zone':
        owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek', active=True)
    else:
        owners = Owner.objects.filter(role__role='tek', active=True)
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        tek = request.POST.get('tek')
        az = mdate
        ta = mdate2
        en_tarikh_in = az.split("/")
        en_tarikh_to = ta.split("/")
        en_tarikh_in = jdatetime.date(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                      year=int(en_tarikh_in[0])).togregorian()
        en_tarikh_to = jdatetime.date(day=int(en_tarikh_to[2]), month=int(en_tarikh_to[1]),
                                      year=int(en_tarikh_to[0])).togregorian()
        en_tarikh_in = str(en_tarikh_in) + " 00:00:00"
        en_tarikh_to = str(en_tarikh_to) + " 23:59:59"
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')

        listticket = Ticket.object_role.c_gs(request, 0).filter(status__status='close', gs__gsowner__owner_id=tek,
                                                                closedate__gte=en_tarikh_in,
                                                                closedate__lte=en_tarikh_to).count()
        liststart = Ticket.object_role.c_gs(request, 0).filter(status__status='close', gs__gsowner__owner_id=tek,
                                                               star__gte=1, star__lte=5,
                                                               closedate__gte=en_tarikh_in,
                                                               closedate__lte=en_tarikh_to).aggregate(
            star=Avg('star'), vote=Count('id'))

        if liststart['star']:
            stars = round(liststart['star'], 1)
            if not stars:
                stars = 0
            starscount = liststart['vote']

        result = StoreHistory.objects.filter(status_id__in=[5, 6], owner_id=tek,
                                             create__gte=en_tarikh_in, create__lte=en_tarikh_to).order_by('-id')
        if request.user.owner.role.role == 'zone':
            result = result.filter(owner__zone_id=request.user.owner.zone_id, )

    return TemplateResponse(request, 'store/tek_amalkard.html',
                            {'result': result, 'owners': owners, 'listTicket': listticket, 'mdate': mdate, 'az': az,
                             'ta': ta,
                             'mdate2': mdate2, 'tek': int(tek), 'stars': stars, 'starscount': starscount,
                             })


@cache_permission('store_mgr')
def store_mgr(request):
    add_to_log(request, f'مشاهده فرم لیست قطعات کم کارکرد ', 0)
    _list = []
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    listticket = 0
    m_reside = ""
    storeok = ""
    store = ""
    _dahilist = []
    zones2 = None

    # دریافت همه مناطق (یا فقط منطقه کاربر اگر نقش zone دارد)
    if request.user.owner.role.role == 'zone':
        zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)
        zones2 = Zone.objects_limit.filter(id=request.user.owner.zone_id)
        selected_zone = str(request.user.owner.zone_id)  # منطقه پیش‌فرض برای کاربران zone
    else:
        zones = Zone.objects_limit.all()
        zones2 = Zone.objects_limit.all()
        selected_zone = "all"  # پیش‌فرض "همه مناطق" برای سایر کاربران

    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        zone_id = request.POST.get('zone', 'all')  # دریافت منطقه انتخاب شده

        # ذخیره منطقه انتخاب شده برای نمایش در فرم
        selected_zone = zone_id

        az = mdate
        ta = mdate2
        en_tarikh_in = az.split("/")
        en_tarikh_to = ta.split("/")
        en_tarikh_in2 = jdatetime.date(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                       year=int(en_tarikh_in[0])).togregorian()
        datein = jdatetime.datetime(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                    year=int(en_tarikh_in[0]), hour=00, minute=00, second=0).togregorian()

        en_tarikh_to = jdatetime.date(day=int(en_tarikh_to[2]), month=int(en_tarikh_to[1]),
                                      year=int(en_tarikh_to[0])).togregorian()
        en_tarikh_in2 = str(en_tarikh_in2) + " 00:00:00"
        en_tarikh_to = str(en_tarikh_to) + " 23:59:59"
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')

        # فیلتر کردن مناطق بر اساس انتخاب کاربر

        if zone_id != 'all':
            zones2 = zones.filter(id=int(zone_id))

        for zone in zones2:
            _start = StoreHistory.objects.filter(owner__zone_id=zone.id, create__gte=en_tarikh_in2,
                                                 create__lte=en_tarikh_to,
                                                 status_id=6).aggregate(sum_a=(
                Count(Case(When(activeclass__in=['A', 'B'], then=1)))),
                sum_c=(
                    Count(Case(When(activeclass__in='C', then=1)))),
                sum_d=(
                    Count(Case(When(activeclass__in='D', then=1)))),
                sum_e=(
                    Count(Case(When(activeclass__in='E', then=1)))),
            )

            _dict = {
                'zone_id': zone.id,
                'zone': zone.name,
                'sum_a': _start['sum_a'],
                'sum_c': _start['sum_c'],
                'sum_d': _start['sum_d'],
                'sum_e': _start['sum_e'],
                'sum': _start['sum_a'] + _start['sum_c'] + _start['sum_d'] + _start['sum_e'],
            }
            _dahilist.append(_dict)

    return TemplateResponse(request, 'store/rep_store_mgr.html',
                            {'list': _dahilist, 'mdate': mdate, 'az': az, 'ta': ta,
                             'mdate2': mdate2, 'zones': zones, 'selected_zone': selected_zone}
                            )


@cache_permission('store_mgr')
def store_detail_mgr(request, mdate, mdate2, zoneid):
    zone = None
    _list = []
    add_to_log(request, f'مشاهده فرم جزئیات قطعات کم کارکرد ', 0)
    az = mdate
    ta = mdate2
    en_tarikh_in = az.split("-")
    en_tarikh_to = ta.split("-")
    en_tarikh_in = jdatetime.date(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                  year=int(en_tarikh_in[0])).togregorian()
    en_tarikh_to = jdatetime.date(day=int(en_tarikh_to[2]), month=int(en_tarikh_to[1]),
                                  year=int(en_tarikh_to[0])).togregorian()
    en_tarikh_in = str(en_tarikh_in) + " 00:00:00"
    en_tarikh_to = str(en_tarikh_to) + " 23:59:59"

    daghis = StoreHistory.objects.filter(owner__zone_id=zoneid, create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                         status_id=6, activeclass__in=['A', 'B', 'C', 'D', 'E']).order_by('activeday')

    zone = Zone.objects.get(id=zoneid)
    #     except:
    #         zone = ""

    # for daghi in daghis:
    #
    #     store = StoreHistory.objects.filter(store_id=daghi.store_id, create__lte=daghi.create,
    #                                         status_id=5).last()
    #     ersale = StoreHistory.objects.filter(store_id=daghi.store_id, create__lte=daghi.create,
    #                                          status_id__in=[2, 9]).last()
    #     reside = StoreHistory.objects.filter(store_id=daghi.store_id, create__lte=daghi.create, owner__zone_id=zoneid,
    #                                          status_id=3).last()
    #
    #     try:
    #         zone = daghi.store.zone.name
    #     except:
    #         zone = ""
    #     if reside:
    #         myreside = reside.pdate()
    #
    #         if ersale:
    #             myersal = ersale.pdate()
    #         else:
    #             myersal = "-"
    #         if store:
    #             day = (daghi.create - store.create).days
    #             mystore = store.pdate()
    #         else:
    #             day = 80
    #             mystore = "-"
    #         if day < 31 or daghi.starterror == 1:
    #             _dict = {
    #                 'serial': daghi.store.serial,
    #                 'statusstore': daghi.store.statusstore.name,
    #                 'zone': daghi.owner.zone.name,
    #                 'desc': daghi.description,
    #                 'ersal': myersal,
    #                 'reside': myreside,
    #                 'store': mystore,
    #                 'daghie': daghi.pdate(),
    #                 'day': int(day),
    #                 'start': daghi.starterror,
    #             }
    #             _list.append(_dict)
    # _list = sorted(_list, key=itemgetter('day'), reverse=False)
    return TemplateResponse(request, 'store/rep_store_detail_mgr.html',
                            {'list': daghis, 'zone': zone,
                             })


@cache_permission('store_mgr')
def store_start_daghi(request):
    _list = []
    start = ""
    ersal_date2 = None
    add_to_log(request, f'مشاهده فرم قطعات از ابتدا معیوب ', 0)
    az = ""
    ta = ""
    zones2 = None

    # دریافت همه مناطق (یا فقط منطقه کاربر اگر نقش zone دارد)
    if request.user.owner.role.role == 'zone':
        zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)
        selected_zone = str(request.user.owner.zone_id)  # منطقه پیش‌فرض برای کاربران zone
    else:
        zones = Zone.objects_limit.all()
        selected_zone = "all"  # پیش‌فرض "همه مناطق" برای سایر کاربران

    tek = request.POST.get('tek')
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        tek = request.POST.get('tek')
        zone_id = request.POST.get('zone', 'all')  # دریافت منطقه انتخاب شده

        # ذخیره منطقه انتخاب شده برای نمایش در فرم
        selected_zone = zone_id
        az = mdate
        ta = mdate2
        en_tarikh_in = az.split("/")
        en_tarikh_to = ta.split("/")
        en_tarikh_in = jdatetime.date(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                      year=int(en_tarikh_in[0])).togregorian()
        en_tarikh_to = jdatetime.date(day=int(en_tarikh_to[2]), month=int(en_tarikh_to[1]),
                                      year=int(en_tarikh_to[0])).togregorian()
        en_tarikh_in = str(en_tarikh_in) + " 00:00:00"
        en_tarikh_to = str(en_tarikh_to) + " 23:59:59"

        if len(tek) > 3:
            start = StoreHistory.objects.filter(create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                                status_id=6, starterror=1, store__serial=tek)
        else:
            if request.user.owner.role.role in ['setad', 'mgr']:
                if zone_id != 'all':
                    start = StoreHistory.objects.filter(owner__zone_id=int(zone_id), create__gte=en_tarikh_in,
                                                        create__lte=en_tarikh_to,
                                                        status_id=6, starterror=1)
                else:
                    start = StoreHistory.objects.filter(create__gte=en_tarikh_in, create__lte=en_tarikh_to,
                                                        status_id=6, starterror=1)
            if request.user.owner.role.role == 'zone':
                start = StoreHistory.objects.filter(owner__zone_id=request.user.owner.zone_id, create__gte=en_tarikh_in,
                                                    create__lte=en_tarikh_to,
                                                    status_id=6, starterror=1)

        for daghi in start:
            ersal_date = StoreHistory.objects.filter(status_id__in=[2, 9], create__lte=daghi.create,
                                                     store_id=daghi.store_id).last()

            if ersal_date:
                ersal_date2 = StoreHistory.objects.filter(status_id__in=[2], create__lt=ersal_date.create,
                                                          store_id=daghi.store_id).last()
                ersal = ersal_date.normal_date()
                kargah = ersal_date.information
                kargah = kargah.replace('ارسال قطعه از ', '')

            else:
                ersal = "-"
                kargah = ""
            if ersal_date2:
                kargahold = ersal_date2.information.replace('ارسال قطعه از ', '')
            else:
                kargahold = ""

            _dict = {
                'serial': daghi.store.serial,
                'zone': daghi.owner.zone.name,
                'send': ersal,
                'date': daghi.normal_date(),
                'owner': daghi.owner,
                'model': daghi.store.statusstore.name,
                'description': daghi.description,
                'kargah': kargah,
                'kargahold': kargahold,
            }
            _list.append(_dict)
        _list = sorted(_list, key=itemgetter('model', 'date'), reverse=True)
    return TemplateResponse(request, 'store/start_daghi.html',
                            {'list': _list, 'az': az, 'ta': ta, 'tek': tek, 'zones': zones,
                             })


def pumpcheck(serial, st):
    pump = None
    if st == 1:
        pump = Pump.objects.filter(master=serial)
    if st == 2:
        pump = Pump.objects.filter(pinpad=serial)

    for item in pump:
        if st == 1:
            item.master = None
        if st == 2:
            item.pinpad = None
        item.save()


def new_arzyabi(request):
    periods = Mount.objects.all()
    base_group = BaseGroup.objects.all()
    teks = Owner.objects.filter(zone_id=request.user.owner.zone_id, active=True, role__role='tek')
    context = {'periods': periods, 'teks': teks, 'base_group': base_group}
    return render(request, 'arzyabi/new_arzyabi.html', context)


def reportzonestoretoexcel(request):
    _list = []
    zone = Zone.objects_limit.all()
    sum_count_master = 0
    sum_count_pinpad = 0
    sum_master_store = 0
    sum_pinpad_store = 0
    sum_master_daghi = 0
    sum_pinpad_daghi = 0
    sum_mstore_takhsis = 0
    sum_pstore_takhsis = 0
    sum_n_pinpad = 0
    sum_n_master = 0
    sum_master_store_noget = 0
    sum_pinpad_store_noget = 0

    for item in zone:
        count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=item.id,
                                                                        status__status='open',
                                                                        Pump__status__status=True,
                                                                        gs__status__status=True,
                                                                        failure__failurecategory_id=1010,
                                                                        failure__isnazel=True).count()

        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=item.id,
                                                                        status__status='open',
                                                                        Pump__status__status=True,
                                                                        gs__status__status=True,
                                                                        failure__failurecategory_id=1011,
                                                                        failure__isnazel=True).count()

        master_store = StoreList.objects.filter(zone_id=item.id, status_id__in=[3, 4, 16],
                                                statusstore_id=1).count()
        pinpad_store = StoreList.objects.filter(zone_id=item.id, status_id__in=[3, 4, 16],
                                                statusstore_id=2).count()
        store_takhsis = Store.objects.filter(zone_id=item.id, status_id__in=[1, 9],
                                             ).aggregate(master=Sum('master'), pinpad=Sum('pinpad'))
        store_noget = Store.objects.filter(zone_id=item.id, status_id=2,
                                           ).aggregate(master=Sum('master'), pinpad=Sum('pinpad'))

        master_daghi = StoreList.objects.filter(zone_id=item.id, statusstore=1, status_id__in=[6, 10, 11, 8]).count()
        pinpad_daghi = StoreList.objects.filter(zone_id=item.id, statusstore=2, status_id__in=[6, 10, 11, 8]).count()
        if store_noget['master']:
            mstore = store_noget['master']
        else:
            mstore = 0

        if store_noget['pinpad']:
            pstore = store_noget['pinpad']
        else:
            pstore = 0

        if store_takhsis['master']:
            mstore_takhsis = store_takhsis['master']
        else:
            mstore_takhsis = 0
        n_master = (master_store + int(mstore) + int(mstore_takhsis)) - count_master
        if store_takhsis['pinpad']:
            pstore_takhsis = store_takhsis['pinpad']
        else:
            pstore_takhsis = 0
        n_pinpad = (pinpad_store + int(pstore) + int(pstore_takhsis)) - count_pinpad

        _dict = {
            'id': item.id,
            'name': item.name,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master_store': master_store,
            'pinpad_store': pinpad_store,
            'master_daghi': master_daghi,
            'pinpad_daghi': pinpad_daghi,
            'n_pinpad': n_pinpad,
            'n_master': n_master,
            'master_store_noget': mstore,
            'pinpad_store_noget': pstore,
            'mstore_takhsis': mstore_takhsis,
            'pstore_takhsis': pstore_takhsis,

        }
        sum_count_master += count_master
        sum_count_pinpad += count_pinpad
        sum_master_store += master_store
        sum_pinpad_store += pinpad_store
        sum_master_daghi += master_daghi
        sum_pinpad_daghi += pinpad_daghi
        sum_mstore_takhsis += mstore_takhsis
        sum_pstore_takhsis += pstore_takhsis
        sum_n_pinpad += n_pinpad
        sum_n_master += n_master
        sum_master_store_noget += mstore
        sum_pinpad_store_noget += pstore
        _list.append(_dict)

    add_to_log(request, 'ارسال آمار قطات مناطق  به اکسل  ', 0)
    my_path = f'Status.xlsx'  # Path
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = EXCEL_EXPORT_FILE + my_path
    _font = Font(bold=True)
    _fonttitr = Font(bold=True, size=20)
    _fonttitr2 = Font(bold=True, size=20)
    wb = Workbook()

    ws1 = wb.active  # work with default worksheet
    ws1.title = "لیست وضعیت قطعات مناطق "
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.fitToPage = True
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:N1')

    ws1["A1"] = f'لیست وضعیت قطعات مناطق   {today}'
    ws1["A1"].font = _fonttitr

    ws1.merge_cells('A2:A3')
    ws1["A2"] = "ردیف"
    ws1["A2"].font = _font

    ws1.merge_cells('B2:B3')
    ws1["B2"] = ZONE_NAME
    ws1["B2"].font = _fonttitr2

    ws1.merge_cells('C2:D2')
    ws1["C2"] = " تیکت  "
    ws1["C2"].font = _font
    ws1.merge_cells('C3:C3')
    ws1["C3"] = "کارتخوان "
    ws1["C3"].font = _font

    ws1.merge_cells('D2:D2')
    ws1["D3"] = " پین پد"
    ws1["D3"].font = _font

    ws1.merge_cells('E2:H2')
    ws1["E2"] = " کارتخوان"
    ws1["E2"].font = _font

    ws1.merge_cells('E3:E3')
    ws1["E3"] = "موجودی "
    ws1["E3"].font = _font

    ws1.merge_cells('F3:F3')
    ws1["F3"] = "بین راهی"
    ws1["F3"].font = _font

    ws1.merge_cells('G3:G3')
    ws1["G3"] = " تخصیصی"
    ws1["G3"].font = _font

    ws1.merge_cells('H3:H3')
    ws1["H3"] = "  مورد نیاز"
    ws1["H3"].font = _font

    ws1.merge_cells('I2:L2')
    ws1["I2"] = " صفحه کلید"
    ws1["I2"].font = _font

    ws1.merge_cells('I3:I3')
    ws1["I3"] = " موجودی"
    ws1["I3"].font = _font

    ws1.merge_cells('J3:J3')
    ws1["J3"] = "بین راهی"
    ws1["J3"].font = _font

    ws1.merge_cells('K3:K3')
    ws1["K3"] = " تخصیصی "
    ws1["K3"].font = _font

    ws1.merge_cells('L3:L3')
    ws1["L3"] = "مورد نیاز"
    ws1["L3"].font = _font

    ws1.merge_cells('M2:N2')
    ws1["M2"] = "  داغی "
    ws1["M2"].font = _font

    ws1.merge_cells('M3:M3')
    ws1["M3"] = "  کارتخوان"
    ws1["M3"].font = _font

    ws1.merge_cells('N3:N3')
    ws1["N3"] = " صفحه کلید  "
    ws1["N3"].font = _font

    ws1.row_dimensions[1].height = 40
    ws1.row_dimensions[2].height = 30
    ws1.row_dimensions[3].height = 40
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    my_font = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    my_fill2 = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    i = 0

    for item in _list:
        i += 1
        d = [i, str(item['name']), str(item['count_master']), str(item['count_pinpad']),
             str(item['master_store']),
             str(item['master_store_noget']), str(item['mstore_takhsis']), str(item['n_master']),
             str(item['pinpad_store']), str(item['pinpad_store_noget']), str(item['pstore_takhsis']),
             str(item['n_pinpad']), str(item['master_daghi']), str(item['pinpad_daghi']),
             ]

        ws1.append(d)

    d = ['', 'جمع', str(sum_count_master), str(sum_count_pinpad), str(sum_master_store),
         str(sum_master_store_noget),
         str(sum_mstore_takhsis), str(sum_n_master), str(sum_pinpad_store), str(sum_pinpad_store_noget),
         str(sum_pstore_takhsis), str(sum_n_pinpad), str(sum_master_daghi), str(sum_pinpad_daghi)

         ]
    ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center', wrap_text=True)
            cell.alignment = alignment_obj
            cell.border = thin_border

    i += 4
    for cell in ws1[i:i]:  # Last row
        cell.font = my_font
        cell.fill = my_fill2
        cell.border = thin_border

    for cell in ws1["2:2"]:  # First row
        cell.font = my_font
        cell.fill = my_fill
        cell.border = thin_border

    for cell in ws1["3:3"]:  # First row

        cell.fill = my_fill
    wb.save(response)
    return response


def changestatusscript():
    _list = StoreHistory.objects.filter(information='تغییر وضعیت قطعه از معیوب شد (داغی)')

    for item in _list:
        result = StoreHistory.objects.filter(store_id=item.store_id, id__lt=item.id).last()
        result.status_id = 12
        result.save()
    return redirect(HOME_PAGE)


def rowtadil(request, old, new):
    _list = []
    Tadiltemp.objects.all().delete()
    for item in Owner.objects.filter(role_id=2).order_by('zone_id'):
        _khordad = Payroll.objects.filter(tek_id=624, period_id=old,
                                          paybaseparametrs_id__in=[7, 10, 11, 12, 17]).aggregate(price=Sum('price'))
        _tir = Payroll.objects.filter(tek_id=item.id, period_id=new, accepttedad=0,
                                      paybaseparametrs_id__in=[7, 10, 11, 12, 17]).aggregate(price=Sum('price'))
        try:
            count_old = Payroll.objects.get(tek_id=item.id, period_id=new, accepttedad=0, paybaseparametrs_id=13).count
            olad_old_price_4 = 4179750 * count_old
            olad_new_price_4 = 7166184 * count_old

            count_old = Payroll.objects.get(tek_id=item.id, period_id=new, accepttedad=0, paybaseparametrs_id=8)
            etlaf_old_price_4 = (56028284 * count_old.count) / 100
            etlaf_new_price_4 = count_old.price

            count_old = Payroll.objects.get(tek_id=item.id, period_id=new, accepttedad=0, paybaseparametrs_id=9)
            jazb_old_price_4 = (56028284 * count_old.count) / 100
            jazb_new_price_4 = count_old.price

            count_old = Payroll.objects.get(tek_id=item.id, period_id=new, accepttedad=0, paybaseparametrs_id=9)
            ezafe_old_price_4 = (1807364 * 1.44) * count_old.count
            ezafe_new_price_4 = count_old.price
        except ObjectDoesNotExist:
            continue

        try:
            Tadiltemp.objects.create(owner_id=item.id, khordad=_khordad['price'], khordad_olad=olad_old_price_4,
                                     khordad_etlaf=etlaf_old_price_4,
                                     khordad_jazb=jazb_old_price_4, khordad_ezafe=ezafe_old_price_4, tir=_tir['price'],
                                     tir_olad=olad_new_price_4,
                                     tir_etlaf=etlaf_new_price_4, tir_ezafe=ezafe_new_price_4, tir_jazb=jazb_new_price_4
                                     )
        except IntegrityError:
            continue

    return HttpResponse('ok')


def gotadil(request):
    Payroll.objects.filter(accepttedad__gte=1).delete()
    teks = Owner.objects.filter(role__role='tek', active=True)
    for item in teks:
        tek_id = item.id
        tadil(tek_id, 6, 5)
        tadil(tek_id, 6, 4)
        tadil(tek_id, 6, 3)
        tadil(tek_id, 6, 2)
        tadil(tek_id, 6, 1)
    return HttpResponse('ok')


def tadil(_tek, _period, _old):
    paye = 0
    maskan = 0
    bon = 0
    food = 0
    pbase = PayBaseParametrs.objects.get(group=7, enname='paye')
    payes = pbase.price
    pbase = PayBaseParametrs.objects.get(group=7, enname='child')
    childs = pbase.price

    try:
        tek = TekKarkard.objects.get(tek_id=_tek, period_id=_period)
    except ObjectDoesNotExist:
        return False
    # for tek in Owner.objects.filter(role_role='tek',active=True):
    result = Payroll.objects.filter(period_id=_period, tek_id=_tek, accepttedad=0)

    for item in result:
        if item.paybaseparametrs_id == 7:
            paye = item.price / item.count

        # if item.paybaseparametrs_id == 10:
        #     refahi = item.price
        if item.paybaseparametrs_id == 11:
            maskan = item.price
        if item.paybaseparametrs_id == 12:
            bon = item.price
        if item.paybaseparametrs_id == 17:
            food = item.price
        # if item.paybaseparametrs_id == 13:
        #     child = item.price

        # if item.paybaseparametrs_id == 8:
        #     etlaf = item.price
        # if item.paybaseparametrs_id == 9:
        #     jazb = item.price
        # if item.paybaseparametrs_id == 16:
        #     ezafe = item.price

    try:
        result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=7)
    except ObjectDoesNotExist:
        return False
    if result.count < 1:
        co = 31
    else:
        co = result.count
    paye = paye * co
    paye5 = paye - result.price

    if paye5 >= 0:

        Payroll.objects.create(period_id=7, count=0, price=paye5, price2=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)
    else:

        Payroll.objects.create(period_id=7, count=0, price2=paye5, price=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)

    result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=10)
    refahi5 = ((1848665 * co) * 20) / 100
    refahi5 = refahi5 - result.price

    if refahi5 >= 0:
        Payroll.objects.create(period_id=7, count=0, price=refahi5, price2=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)
    else:
        Payroll.objects.create(period_id=7, count=0, price2=refahi5, price=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)

    result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=11)
    maskan5 = maskan - result.price
    if maskan5 >= 0:
        Payroll.objects.create(period_id=7, count=0, price=maskan5, price2=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)
    else:
        Payroll.objects.create(period_id=7, count=0, price2=maskan5, price=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)

    result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=12)
    bon5 = bon - result.price
    if bon5 >= 0:
        Payroll.objects.create(period_id=7, count=0, price=bon5, price2=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)
    else:
        Payroll.objects.create(period_id=7, count=0, price2=bon5, price=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)

    result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=13)

    child5 = childs * result.count
    mchild = result.price
    child5 = child5 - mchild
    if child5 >= 0:
        Payroll.objects.create(period_id=7, count=result.count, price=child5, price2=0, accepttedad=_old,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               tek_id=_tek, user_id=503)
    else:
        Payroll.objects.create(period_id=7, count=result.count, price2=child5, price=0, accepttedad=_old,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               tek_id=_tek, user_id=503)

    result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=17)
    food5 = food - result.price
    if food5 >= 0:
        Payroll.objects.create(period_id=7, count=0, price=food5, price2=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)
    else:
        Payroll.objects.create(period_id=7, count=0, price2=food5, price=0,
                               paybaseparametrs_id=result.paybaseparametrs_id,
                               accepttedad=_old,
                               tek_id=_tek, user_id=503)
    # result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=9)
    # jazb5 = ((int(payes) * 31 * result.count) / 100)
    # if result.price >= jazb5:
    #     jazb5 = 0
    # else:
    #     jazb5 = jazb5 - result.price
    #
    # Payroll.objects.create(period_id=_period, count=0, price=jazb5, paybaseparametrs_id=result.paybaseparametrs_id,
    #                        accepttedad=_old,
    #                        tek_id=_tek, user_id=503)
    #
    # result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=8)
    # etlaf5 = ((int(payes) * 31 * result.count) / 100)
    # if result.price >= etlaf5:
    #     etlaf5 = 0
    # else:
    #     etlaf5 = etlaf5 - result.price
    #
    # Payroll.objects.create(period_id=_period, count=0, price=etlaf5, paybaseparametrs_id=result.paybaseparametrs_id,
    #                        accepttedad=_old,
    #                        tek_id=_tek, user_id=503)
    #
    # result = Payroll.objects.get(period_id=_old, tek_id=_tek, paybaseparametrs_id=16)
    # ezafe5 = ((int(payes) * 0.191) * int(result.count))
    #
    # if result.price >= ezafe5:
    #     ezafe5 = 0
    # else:
    #     ezafe5 = ezafe5 - result.price
    # Payroll.objects.create(period_id=_period, count=0, price=ezafe5, paybaseparametrs_id=result.paybaseparametrs_id,
    #                        accepttedad=_old,
    #                        tek_id=_tek, user_id=503)

    return HttpResponse('test')


class RepairUse(View):
    form_class = RepairForm
    template_file = 'repair/repairform.html'
    repairstore = RepairStoreName.objects.all()

    def get(self, request):
        newdate = ""
        form = self.form_class
        return render(request, self.template_file, {'form': form, 'repairstore': self.repairstore, 'newdate': newdate})

    def post(self, request):
        url = request.META.get('HTTP_REFERER')
        form = self.form_class(request.POST)
        _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)
        _storageid = _storage.id
        _sid = ""
        _tarikh = ""
        if form.is_valid():
            cd = form.cleaned_data
            form.instance.tarikh = cd['tarikh']
            newdate = str(cd['tarikh'])
            newdate = newdate.replace('-', '/')
            form.instance.storage_id = _storage.id
            try:
                _sid = form.instance.repairstore_id
                _tarikh = form.instance.tarikh
                form.save()
            except IntegrityError:
                _store = Repair.objects.get(tarikh=_tarikh, repairstore_id=_sid,
                                            storage_id=_storageid)
                _store.valuecount = int(cd['valuecount'])
                _store.save()
            Repair.checkuserepair(_storage.id)
            messages.success(request, 'عملیات با موفقیت انجام شد')
            repair_list = Repair.objects.filter(storage_id=_storageid, tarikh=cd['tarikh'])
            return render(request, self.template_file,
                          {'form': form, 'repairstore': self.repairstore, 'newdate': newdate, 'newdate2': cd['tarikh'],
                           'repair_list': repair_list})
        else:
            messages.error(request, 'عملیات با شکست انجام شد')
            return redirect(url)


def scriptstore(request):
    his = StoreHistory.objects.filter(baseroot__gt=0).order_by('id')
    for item in his:
        try:
            store = StoreList.objects.get(id=item.store_id)
            store.store_id = item.baseroot
            store.save()
        except:
            continue
    return HttpResponse('end')


def scriptview(request):
    hist = StoreHistory.objects.filter(status_id=9, isok=1).order_by('id')
    for item in hist:
        try:
            StoreView.objects.create(store_id=item.store_id, serial=item.store.serial, zone_id=item.owner.zone_id,
                                     send_date=item.create, status=2, starterr=False)
            item.isok = 2
            item.save()
        except:
            continue
    store = StoreView.objects.all().order_by('id')
    for item in store:
        try:
            his = StoreHistory.objects.filter(status_id=3, isok=1, store_id=item.store_id).first()
            item.resid_date = his.create
            item.save()
            his.isok = 2
            his.save()
            his = StoreHistory.objects.filter(status_id=4, isok=1, store_id=item.store_id).first()
            item.tek_date = his.create
            item.save()
            his.isok = 2
            his.save()
            his = StoreHistory.objects.filter(status_id=5, isok=1, store_id=item.store_id).first()
            item.gs_date = his.create
            item.save()
            his.isok = 2
            his.save()
            his = StoreHistory.objects.filter(status_id=6, isok=1, store_id=item.store_id).first()
            item.fail_date = his.create
            item.starterr = his.starterror
            item.save()
            his.isok = 2
            his.save()
        except:
            continue


@cache_permission('store_mgr')
def addstorerepair(request, _id):
    store = RepairStore.objects.filter(status_id=1, storage_id=_id)
    return TemplateResponse(request, 'repair/addstorerepaire.html', {'store': store})


@cache_permission('history')
def showimg(request, _id):
    img = ImgSerial.objects.get(id=_id)
    add_to_log(request, f'مشاهده عکس قطعه ', 0)
    return TemplateResponse(request, 'store/imgshow.html', {'img': img, })


def paymarid():
    owner = Owner.objects.filter(role__role='tek')
    for item in owner:
        Payroll.objects.filter(paybaseparametrs__enname='marid', tek_id=item.id, tek__marital_status='singel').delete()
    print('---------------------end---------------')
    return True


@cache_permission('zonelistre')
def amarstorstek(request):
    add_to_log(request, " تهیه گزارش خلاصه آمار قطعات تکنسین ", 0)
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        az = mdate
        ta = mdate2
        mdate = mdate.split("/")
        mdate2 = mdate2.split("/")
        tarikh = jdatetime.date(day=int(mdate[2]), month=int(mdate[1]), year=int(mdate[0])).togregorian()
        tarikh2 = jdatetime.date(day=int(mdate2[2]), month=int(mdate2[1]), year=int(mdate2[0])).togregorian()
        owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
        _list = []
        for i in owners:
            store_in_master = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=4,
                                                          store__statusstore=1,
                                                          create__range=(tarikh, tarikh2)).count()
            store_in_pinpad = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=4,
                                                          store__statusstore=2,
                                                          create__range=(tarikh, tarikh2)).count()

            store_out_master = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=10,
                                                           store__statusstore=1,
                                                           create__range=(tarikh, tarikh2)).count()
            store_out_pinpad = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=10,
                                                           store__statusstore=2,
                                                           create__range=(tarikh, tarikh2)).count()
            store_in = store_in_master + store_in_pinpad
            store_out = store_out_master + store_out_pinpad
            _dict = {
                'id': i.id,
                'name': i.get_full_name(),
                'store_in': store_in,
                'store_out': store_out,
                'store_in_master': store_in_master,
                'store_in_pinpad': store_in_pinpad,
                'store_out_master': store_out_master,
                'store_out_pinpad': store_out_pinpad,
            }
            _list.append(_dict)

        return TemplateResponse(request, 'amarstorestek.html',
                                {'list': _list,
                                 'az': az, 'ta': ta, 'tarikh': tarikh, 'tarikh2': tarikh2,
                                 })
    return TemplateResponse(request, 'amarstorestek.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2,
                             })


@cache_permission('zonelistre')
def amarstorstekdetail(request, tarikh, tarikh2, owner):
    add_to_log(request, " تهیه گزارش آمار قطعات تکنسین ", 0)
    i = Owner.objects.get(id=owner)

    store_in_master = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=4,
                                                  store__statusstore=1,
                                                  create__range=(tarikh, tarikh2))
    store_in_pinpad = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=4,
                                                  store__statusstore=2,
                                                  create__range=(tarikh, tarikh2))

    store_out_master = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=10,
                                                   store__statusstore=1,
                                                   create__range=(tarikh, tarikh2))
    store_out_pinpad = StoreHistory.objects.filter(description__icontains=i.get_full_name(), status_id=10,
                                                   store__statusstore=2,
                                                   create__range=(tarikh, tarikh2))

    return TemplateResponse(request, 'amarstorestekdetail.html',
                            {'store_in_master': store_in_master, 'store_in_pinpad': store_in_pinpad, 'tarikh': tarikh,
                             'tarikh2': tarikh2,
                             'store_out_master': store_out_master, 'store_out_pinpad': store_out_pinpad, 'owner': i,

                             })


@cache_permission('zonelistre')
def serialliststore(request):
    _list = None
    _sum = None
    _filter = ""
    add_to_log(request, f'مشاهده فرم لیست  سریال قطعات', 0)
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    vore = request.GET.get('vore')

    if len(datein) < 10:
        datein = "2023-01-01"
        dateout = "9999-12-30"

    else:

        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        datein = str(datein) + " 00:00:01"
        dateout = str(dateout) + " 23:59:59"

    if request.user.owner.role.role == 'zone':

        _list = StoreList.objects.filter(update__gte=datein, update__lte=dateout,
                                         zone_id=request.user.owner.zone_id).order_by('-id')

        _filter = StoreFilters(request.GET, queryset=_list)
        if _filter.data:
            _list = _filter.qs
        else:
            _list = []
    if request.user.owner.role.role in ['mgr', 'setad', 'posht']:
        _list = StoreList.objects.filter(update__gte=datein, update__lte=dateout,
                                         ).order_by('-id')

        _filter = StoreFilters(request.GET, queryset=_list)
        if _filter.data:
            _list = _filter.qs
        else:
            _list = []

    if vore == "2":
        add_to_log(request, 'ارسال آمار لیست ارسالی ها به اکسل  ', 0)
        my_path = "sendstore.xlsx"
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = 'attachment; filename=' + my_path
        font = Font(bold=True)
        fonttitr = Font(bold=True, size=20)
        fonttitr2 = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active  # work with default worksheet
        ws1.title = "گزارش لیست  سریال قطعات  "
        ws1.sheet_view.rightToLeft = True
        ws1.firstFooter.center.text = "ali"
        ws1.merge_cells('A1:F1')

        ws1["A1"] = f'گزارش لیست  سریال قطعات  تاریخ   {today}'
        ws1["A1"].font = fonttitr

        ws1.merge_cells('A2:A2')
        ws1["A2"] = "ردیف"
        ws1["A2"].font = font

        ws1.merge_cells('B2:B2')
        ws1["B2"] = "منطقه"
        ws1["B2"].font = fonttitr2

        ws1.merge_cells('C2:C2')
        ws1["C2"] = "سریال "
        ws1["C2"].font = font

        ws1.merge_cells('D2:D2')
        ws1["D2"] = " تاریخ "
        ws1["D2"].font = font

        ws1.merge_cells('E2:E2')
        ws1["E2"] = "نوع قطعه"
        ws1["E2"].font = font

        ws1.merge_cells('F2:F2')
        ws1["F2"] = "وضعیت"
        ws1["F2"].font = font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(15.25)
        ws1.column_dimensions['D'].width = float(25.25)
        ws1.column_dimensions['E'].width = float(15.25)
        ws1.column_dimensions['F'].width = float(18.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        myfont = Font(size=14, bold=True)  # font styles
        my_fill = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        my_fill2 = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        i = 0

        for item in _list:
            i += 1
            d = [i, str(item.zone), str(item.serial), str(item.normal_date()), str(item.statusstore.name),
                 str(item.status.name)]

            ws1.append(d)

        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        i += 4
        for cell in ws1[i:i]:  # Last row
            cell.font = myfont
            cell.fill = my_fill2
            cell.border = thin_border

        for cell in ws1["2:2"]:  # First row
            cell.font = myfont
            cell.fill = my_fill
            cell.border = thin_border
        wb.save(response)
        return response

    paginator = Paginator(_list, 20)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:
        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, 'store/serialliststore.html',
                            {'filter': _filter, 'list': page_object, 'query_string': query_string,
                             'this_date': this_date,
                             'today_date': today_date, 'listsum': _sum,
                             'page_obj': page_obj, 'tedad': tedad,
                             })


@cache_permission('mojodidaghi')
def countstoragezone(request):
    _list = []
    _storage = ""
    add_to_log(request, f'مشاهده فرم کاردکس کارگاه ', 0)
    storages = Storage.objects.all().order_by('sortid')
    if request.method == 'POST':
        az = str(request.POST.get('select'))
        ta = str(request.POST.get('select2'))
        zone = int(request.POST.get('zone'))
        datein = to_miladi(az)
        dateindate = to_miladi(az)
        datein2date = to_miladi(ta)
        datein2 = to_miladi(ta)
        datein = str(datein) + " 00:00:00"
        datein2 = str(datein2) + " 23:59:59"

        if request.user.owner.role.role in 'mgr,setad' and zone == 0:
            storages = Storage.objects.filter(zone__iszone=False, refrence=False).order_by('sortid')
            for storage in storages:
                sth = Store.objects.filter(marsole_date__range=(datein, datein2), status_id__in=[2, 3],
                                           storage_id=storage.id).aggregate(
                    master=(Sum('master')),
                    pinpad=(Sum('pinpad'))
                )
                master_sum = sth['master'] or 0
                pinpad_sum = sth['pinpad'] or 0
                _dict = {
                    'storage': storage.name,
                    'storageid': storage.id,
                    'master': master_sum,
                    'pinpad': pinpad_sum,
                    'summ': master_sum + pinpad_sum,
                }
                _list.append(_dict)
            storages = Storage.objects.filter(zone__iszone=True).order_by('sortid')
            for storage in storages:
                sth = StoreHistory.objects.filter(status_id=3, storage_id=storage.id,
                                                  create__range=(datein, datein2)).aggregate(
                    master=Count(Case(When(store__statusstore_id=1, then=1))),
                    pinpad=Count(Case(When(store__statusstore_id=2, then=1)))
                )
                _dict = {
                    'storage': storage.name,
                    'storageid': storage.id,
                    'master': sth['master'],
                    'pinpad': sth['pinpad'],
                    'summ': int(sth['master']) + int(sth['pinpad']),
                }
                _list.append(_dict)
        singel = False
        if request.user.owner.role.role in 'mgr,setad' and zone != 0:
            singel = True

            _storage = Storage.objects.get(id=zone)
        if request.user.owner.role.role == 'zone':
            singel = True
            zone = request.user.owner.zone_id
            _storage = Storage.objects.get(zone_id=request.user.owner.zone_id)

        if singel:
            sth = StoreHistory.objects.values('create__year', 'create__month', 'create__day').filter(
                status_id__in=[1, 3],
                storage_id=_storage.id,
                create__range=(
                    datein,
                    datein2)).annotate(
                master=Count(Case(When(store__statusstore_id=1, then=1))),
                pinpad=Count(Case(When(store__statusstore_id=2, then=1)))
            ).order_by('-create__year', '-create__month', '-create__day')
            for storage in sth:
                _create = datetime.date(day=int(storage['create__day']), month=int(storage['create__month']),
                                        year=int(storage['create__year']))

                jd = JDate(_create.strftime("%Y-%m-%d %H:%M:%S"))
                newsdate = jd.format('Y/m/d')
                _dict = {
                    'storage': _storage.name,
                    'storageid': _storage.id,
                    'date': str(newsdate),
                    'master': storage['master'],
                    'pinpad': storage['pinpad'],
                    'summ': int(storage['master']) + int(storage['pinpad']),
                }
                _list.append(_dict)
        _list = sorted(_list, key=itemgetter('summ'), reverse=True)
        return TemplateResponse(request, 'store/countstoregezone.html', {
            'list': _list,
            'storages': storages,
            'zone': zone,
            'az': az,
            'ta': ta,
            'dateindate': dateindate,
            'datein2date': datein2date,

        })
    return TemplateResponse(request, 'store/countstoregezone.html',
                            {'list': _list, 'storages': storages})


@cache_permission('mojodidaghi')
def countstoragezonedetail(request, _id, az, ta):
    _list = []
    _storage = ""
    add_to_log(request, f'مشاهده فرم کاردکس کارگاه با جزئیات ', 0)

    datein = az
    datein2 = ta
    datein = str(datein) + " 00:00:00"
    datein2 = str(datein2) + " 23:59:59"
    _list = StoreHistory.objects.filter(
        status_id=3,
        storage_id=_id,
        create__range=(
            datein,
            datein2))

    return TemplateResponse(request, 'store/countstoregezonedetail.html', {
        'list': _list,

    })


@cache_permission('editstorezone')
def updatestorezone(request, _id):
    zones = Storage.objects.all().order_by('sortid')
    store = Store.objects.get(id=_id)
    if store.status_id not in [7]:
        messages.error(request, ' این مرسوله بعلت رسید شدن در مقصد قابل ویرایش نمیباشد')
        return redirect('pay:zonelistdaghi')
    if request.method == 'POST':
        zone = request.POST.get('zone')
        store.storage_id = int(zone)
        store.save()
        storelist = StoreList.objects.filter(store_id=_id)
        for item in storelist:
            StoreHistory.objects.create(store_id=item.id, owner_id=request.user.owner.id, baseroot=0,
                                        information="ارسال قطعه از " + str(store.zone.name),
                                        status_id=item.status_id,
                                        description=f'  به  کارگاه /  منطقه  {store.storage.name} '
                                                    f'ویرایش شده توسط {request.user.owner.name}'
                                                    f'  {request.user.owner.lname} ')

        messages.info(request, ' عملیات با موفقیت  انجام شد')
        return redirect('pay:zonelistdaghi')
    return TemplateResponse(request, 'store/updatestorezone.html', {'zones': zones, 'store': store})


@cache_permission('workrepair')
def work_repair(request):
    add_to_log(request, f'مشاهده فرم عملکرد کارگاه ', 0)
    starscount = 0
    stars = 0
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    resid_ok = None
    resid_ok_sum = None
    resid_daghi = None
    resid_daghi_sum = None
    send_to_zone = None
    send_to_zone_sum = None
    send_daghi_to_zone = None
    send_daghi_to_zone_sum = None
    sthsum = None
    _sth = None
    _storageid = None
    _storagename = None
    _st = 1

    if request.user.owner.role.role == 'zone':
        storages = Storage.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        storages = Storage.objects.all().order_by('sortid')
    if request.method == 'POST':
        _st = request.POST.get('st')
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        storage = request.POST.get('storage')
        _storage = Storage.objects.get(id=int(storage))
        if request.user.owner.role.role == 'zone' and _storage.zone_id != request.user.owner.zone_id:
            messages.error(request, 'دسترسی غیر مجاز')
            return redirect('base:home')

        _storageid = _storage.id
        _storagename = _storage.name
        az = mdate
        ta = mdate2
        en_tarikh_in = az.split("/")
        en_tarikh_to = ta.split("/")
        en_tarikh_in = jdatetime.date(day=int(en_tarikh_in[2]), month=int(en_tarikh_in[1]),
                                      year=int(en_tarikh_in[0])).togregorian()
        en_tarikh_to = jdatetime.date(day=int(en_tarikh_to[2]), month=int(en_tarikh_to[1]),
                                      year=int(en_tarikh_to[0])).togregorian()
        en_tarikh_in = str(en_tarikh_in) + " 00:00:00"
        en_tarikh_to = str(en_tarikh_to) + " 23:59:59"
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')

        resid_ok = Store.objects.filter(zone_id=_storage.zone_id, status_id=3,
                                        resid_date__range=(en_tarikh_in, en_tarikh_to))
        resid_ok_sum = resid_ok.aggregate(master=Sum('master'), pinpad=Sum('pinpad'))
        if _st == "2":
            resid_ok = resid_ok.values('storage_id', 'storage__name').annotate(master=Sum('master'),
                                                                               pinpad=Sum('pinpad'))

        resid_daghi = Store.objects.filter(storage_id=int(storage), status_id=8,
                                           resid_date__range=(en_tarikh_in, en_tarikh_to))
        resid_daghi_sum = resid_daghi.aggregate(master=Sum('master'), pinpad=Sum('pinpad'))
        if _st == "2":
            resid_daghi = resid_daghi.values('zone_id', 'zone__name').annotate(master=Sum('master'),
                                                                               pinpad=Sum('pinpad'))

        send_to_zone = Store.objects.filter(storage_id=int(storage), status_id=3,
                                            resid_date__range=(en_tarikh_in, en_tarikh_to))
        send_to_zone_sum = send_to_zone.aggregate(master=Sum('master'), pinpad=Sum('pinpad'))
        if _st == "2":
            send_to_zone = send_to_zone.values('zone_id', 'zone__name').annotate(master=Sum('master'),
                                                                                 pinpad=Sum('pinpad'))

        send_daghi_to_zone = Store.objects.filter(zone_id=_storage.zone_id, status_id=8,
                                                  resid_date__range=(en_tarikh_in, en_tarikh_to))
        send_daghi_to_zone_sum = send_daghi_to_zone.aggregate(master=Sum('master'), pinpad=Sum('pinpad'))
        if _st == "2":
            send_daghi_to_zone = send_daghi_to_zone.values('storage_id', 'storage__name').annotate(master=Sum('master'),
                                                                                                   pinpad=Sum('pinpad'))

        sth = StoreHistory.objects.values('create__year', 'create__month', 'create__day').filter(
            status_id__in=[3], baseroot=0, storage_id=int(storage),
            create__range=(en_tarikh_in, en_tarikh_to)).annotate(
            master=Count(Case(When(store__statusstore_id=1, then=1))),
            pinpad=Count(Case(When(store__statusstore_id=2, then=1)))
        )
        _sth = []
        for item in sth:
            _create = datetime.date(day=int(item['create__day']), month=int(item['create__month']),
                                    year=int(item['create__year']))

            jd = JDate(_create.strftime("%Y-%m-%d %H:%M:%S"))
            newsdate = jd.format('Y/m/d')
            _dict = {
                'tarikh': str(newsdate),
                'master': str(item['master']),
                'pinpad': str(item['pinpad']),
            }
            _sth.append(_dict)

        sthsum = StoreHistory.objects.filter(
            status_id__in=[3], baseroot=0, storage_id=int(storage),
            create__range=(en_tarikh_in, en_tarikh_to)).aggregate(
            master=Count(Case(When(store__statusstore_id=1, then=1))),
            pinpad=Count(Case(When(store__statusstore_id=2, then=1)))
        )

    return TemplateResponse(request, 'store/work-repair.html',
                            {'storages': storages, 'mdate': mdate, 'az': az, 'ta': ta,
                             'mdate2': mdate2, 'storage2': _storageid, 'stars': stars, 'starscount': starscount,
                             'resid_ok': resid_ok, 'resid_ok_sum': resid_ok_sum, 'resid_daghi': resid_daghi,
                             'resid_daghi_sum': resid_daghi_sum, 'send_to_zone': send_to_zone, 'sth': _sth,
                             'sthsum': sthsum,
                             'send_to_zone_sum': send_to_zone_sum, 'storagename': _storagename, "st": int(_st),
                             'send_daghi_to_zone': send_daghi_to_zone, 'send_daghi_to_zone_sum': send_daghi_to_zone_sum,
                             })


def backtotakhsis(request, _id):
    url = request.META.get('HTTP_REFERER')
    _res = Store.objects.get(id=_id)
    _res.status_id = 1
    _res.marsole_date = "1377-01-01 01:01:01"
    _res.save()
    return redirect(url)


@cache_permission('bulk_chstore')
def import_excel_serialnumber(request):
    form = open_excel(request.POST)
    zone = StoreManufacturer.objects.all()
    statusref = StatusRef.objects.all()
    if request.method == 'POST':
        _zone = request.POST.get('zone')

        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                master = sheet_obj.cell(row=i, column=1).value
                master = checknumber(str(master))
                if master:
                    try:

                        SerialRange.objects.create(serialnumber=master, storemanufacturer_id=_zone)
                    except IntegrityError:
                        pass

            messages.success(request, ' دریافت اطلاعات با موفقیت انجام شد')
        return redirect(HOME_PAGE)

    return TemplateResponse(request, 'store/import_excel_serialnumber.html',
                            {'form': form, 'statusref': statusref, 'zone': zone})


@cache_permission('newdaghi')
def changestore2(request):
    url = request.META.get('HTTP_REFERER')
    if request.user.owner.refrence_id != 1:
        return redirect('base:home')
    serial = ""
    form = ImageStore(request.GET)

    if request.method == 'POST':
        form = ImageStore(request.POST, request.FILES)
        store = int(request.POST.get('store'))
        customFile = request.POST.get('ok')

        storeid = store
        serial = request.POST.get('init')
        serial = check_serial(serial)
        if serial == False:
            url = request.META.get('HTTP_REFERER')
            messages.error(request, ADD_STORE_MSG)
            return redirect(url)
        if customFile == '2':

            try:
                store = StoreList.objects.get(serial=serial, statusstore_id=storeid)
            except ObjectDoesNotExist:
                messages.warning(request, "نوع قطعه اشتباه انتخاب شده است از سابقه قطعه بررسی کنید")
                return TemplateResponse(request, 'store/daghi.html', {'ok': 1, 'form': form})
            zonen = "-"
            if form.is_valid():

                form.instance.store_id = store.id
                form.instance.owner_id = request.user.owner.id
                a = form.save()

                store.status_id = 10
                store.getuser_id = request.user.owner.id
                store.owner_id = request.user.id
                store.zone_id = request.user.owner.zone_id
                store.save()
                # ImgSerial.objects.create(store_id=store.id, img=customFile)
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id, imgid_id=a.id,
                                            information="ثبت قطعه داغی (تصویر بارگذاری شد) ",
                                            status_id=6, description="")
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information="دریافت قطعه داغی  ",
                                            status_id=10, description="")
                daghi_mande = Owner.objects.get(id=request.user.owner.id)
                daghi_mande.daghimande = False
                daghi_mande.save()
                messages.success(request, 'ذخیره قطعه داغی انجام شد.')
                return TemplateResponse(request, 'store/daghi.html', {'ok': 1, 'form': form})
            else:
                messages.error(request, 'تصویر بدرستی بارگذاری نشد')
                return TemplateResponse(request, 'store/daghi.html',
                                        {'ok': 1, 'form': form})
        try:

            store = StoreList.objects.get(serial=serial, statusstore_id=storeid)
            zonen = "-"
            if store.zone_id:
                zonen = store.zone.name
            if store.status_id == 5 and store.zone_id == request.user.owner.zone_id:
                store.status_id = 10
                store.getuser_id = request.user.owner.id
                store.owner_id = request.user.id
                store.zone_id = request.user.owner.zone_id
                store.save()
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information="ثبت قطعه داغی  ",
                                            status_id=6, description="")
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information="دریافت قطعه داغی  ",
                                            status_id=10, description="")

                messages.success(request, 'ذخیره قطعه داغی انجام شد.')
                return redirect(url)
            else:
                messages.error(request, 'این قطعه در وضعیت ' + str(store.status.name) + ' در منطقه ' + str(
                    zonen) + ' وجود دارد امکان ثبت بعنوان داغی وجود ندارد لطفا از قسمت سابقه قطعه بررسی کنید')
                payam = 'این قطعه در وضعیت ' + str(store.status.name) + ' در منطقه ' + str(
                    zonen) + ' وجود دارد امکان ثبت بعنوان داغی وجود ندارد لطفا از قسمت سابقه قطعه بررسی کنید'
                return TemplateResponse(request, 'store/daghi.html',
                                        {'serial': serial, 'ok': 2, 'payam': payam,
                                         'form': form,
                                         'storeid': int(storeid)})


        except ObjectDoesNotExist:
            try:
                store = StoreList.objects.create(serial=serial, getuser_id=request.user.owner.id,
                                                 owner_id=request.user.id,
                                                 zone_id=request.user.owner.zone_id, statusstore_id=store, status_id=10,
                                                 uniq=str(serial) + "-" + str(store))
                StoreHistory.objects.create(store_id=store.id, owner_id=request.user.owner.id,
                                            information="ثبت قطعه داغی(جدید)  ",
                                            status_id=6, description="")

                messages.success(request, 'ذخیره قطعه داغی انجام شد.')
                return redirect(url)
            except IntegrityError:
                messages.error(request, 'این قطعه در وضعیت ارسال به کارگاه تعمیر میباشد ')
                return TemplateResponse(request, 'store/daghi.html',
                                        {'serial': serial, 'storeid': int(storeid), 'ok': 2,
                                         'form': form,
                                         'payam': 'این قطعه در وضعیت ارسال به کارگاه تعمیر میباشد '})

    return TemplateResponse(request, 'store/daghi2.html', {'ok': 1, 'form': form})


def getmasterresid(request):
    add_to_log(request, f'مشاهده لیست لاگ  رسید با اکسل ', 0)
    thislist = []
    if request.method == 'POST':
        _id = int(request.POST.get('obj'))
        st = int(request.POST.get('st'))
        _store = Store.objects.get(id=_id)
        lists = StoreHistory.objects.filter(
            residroot=_id)
        if _store.status_id == 8:
            lists = lists.filter(status_id__in=[8, 15])
        _todaydate = date.today()
        si_ago_sell = _todaydate.today() - datetime.timedelta(days=30)
        for _list in lists:
            _input = "-"
            if _list.information == 'رسید  قطعه به کارگاه مغایرتی':
                _input = "1"
            elif _list.information == 'رسید  قطعه به کارگاه ':
                _input = "2"
            elif _list.information == 'مغایرت':
                _input = "3"

            risk = StoreHistory.objects.filter(store_id=_list.store_id, create__gte=si_ago_sell,
                                               create__lte=_todaydate.today(),
                                               status_id=6).count()
            if risk == 0:
                level = 0
            elif risk == 1:
                level = 1
            elif risk == 2:
                level = 2
            else:
                level = 3
            mydict = {
                "serial": _list.store.serial,
                "status": _list.store.status.name,
                "level": level,
                'input': _input,
                'statusstore': str(_list.store.statusstore.name)
            }
            thislist.append(mydict)
        return JsonResponse({'message': 'success', 'list': thislist})


def checkstoremanufactur(request):
    _listserial = []
    man = SerialRange.objects.all()
    for serial in man:
        _listserial.append(serial.serialnumber)

    result = StoreHistory.objects.values('owner__zone_id', 'owner__zone__name').filter(
        store__serial__in=_listserial).annotate(resid=Count(Case(When(status_id=3, then=1))),
                                                ersal=Count(Case(When(status_id=8, then=1)))
                                                )

    my_path = "list.xlsx"
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)
    fonttitr2 = Font(bold=True, size=20)
    wb = Workbook()

    ws1 = wb.active  # work with default worksheet
    ws1.title = "گزارش لیست قطعات بومی "
    ws1.sheet_view.rightToLeft = True
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:c1')

    ws1["A1"] = f'گزارش لیست قطعات بومی  '
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:A2')
    ws1["A2"] = "نام منطقه"
    ws1["A2"].font = font

    ws1.merge_cells('B2:B2')
    ws1["B2"] = "تعداد رسید به منطقه"
    ws1["B2"].font = fonttitr2

    ws1.merge_cells('C2:C2')
    ws1["C2"] = "تعداد ارسال از منطقه "
    ws1["C2"].font = font

    ws1.column_dimensions['A'].width = float(45.25)
    ws1.column_dimensions['B'].width = float(45.25)
    ws1.column_dimensions['C'].width = float(45.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    my_fill2 = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    i = 0

    for item in result:
        d = [item['owner__zone__name'], item['resid'], item['ersal']]
        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    for cell in ws1["2:2"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border
    wb.save(response)
    return response


def checkdaghi():
    _list = StoreHistory.objects.filter(status_id=6, activeday__isnull=True).order_by('-id')
    for result in _list:
        try:

            _daghi = result.create
            _start = StoreHistory.objects.filter(store_id=result.store_id, status_id=9, id__lt=result.id).order_by(
                '-id').first()
            if _start:
                result.senddate = _start.create

            _start = StoreHistory.objects.filter(store_id=result.store_id, status_id=3, id__lt=result.id).order_by(
                '-id').first()
            if _start:
                result.residdate = _start.create
                if _start.storage_id:
                    result.storage_id = _start.storage_id
                    result.senddate2 = _start.storage_id

            _start = StoreHistory.objects.filter(store_id=result.store_id, status_id=5, id__lt=result.id).order_by(
                '-id').first()
            if _start:
                _install = _start.create
                result.installdate = _start.create
                result.activeday = (result.create - result.installdate).days
                if result.activeday == 0:
                    result.activeclass = 'A'
                elif result.activeday == 1:
                    result.activeclass = 'B'
                elif 1 < result.activeday < 11:
                    result.activeclass = 'C'
                elif 10 < result.activeday < 21:
                    result.activeclass = 'D'
                elif 20 < result.activeday < 31:
                    result.activeclass = 'E'
                else:
                    result.activeclass = 'F'

            if result.starterror:
                result.activeday = 0
                result.activeclass = 'A'

            result.save()


        except ObjectDoesNotExist:
            print('nist')
        except IntegrityError:
            pass
        except TypeError:
            pass
        except:
            pass
    return True


@cache_permission('nemodarkamkarkard')
def nemodarkamkarkard(request):
    tarikh1 = ""
    tarikh2 = ""
    newlist = None
    _id = 0
    _list = None
    nemodar1 = None
    datein = ""
    dateout = ""
    storages = Storage.objects.all()
    if request.method == 'POST':
        _id = request.POST.get('storage')
        datein = str(request.POST.get('select'))

        dateout = str(request.POST.get('select2'))
        tarikh1 = to_miladi(datein)
        tarikh2 = to_miladi(dateout)
        _list = StoreHistory.objects.filter(create__range=(tarikh1, tarikh2), storage_id=_id, activeday__isnull=False,
                                            activeday__lt=30)
        nemodar1 = _list.values(year_month=Func(
            F('create'),
            Value('%Y-%m-%d'),
            function='DATE_FORMAT',
            output_field=CharField())).annotate(counter=Count('id'),
                                                starterr=Count(Case(When(starterror=True, then=1))))
        newlist = []
        for item in nemodar1:
            _date = item['year_month'].split('-')
            tarikh = datetime.date(day=int(_date[2]), month=int(_date[1]), year=int(_date[0]))
            jd = JDate(tarikh.strftime("%Y-%m-%d"))
            tarikh = jd.format('Y/m/d')

            dict = {
                'date': str(tarikh),
                'val': str(item['counter']),
                'starterr': str(item['starterr']),
            }
            newlist.append(dict)
    return TemplateResponse(request, 'store/nemodarkamkarkard.html',
                            {'nemodar1': newlist, 'list': _list, 'storages': storages, 'storagesid': int(_id),
                             'az': datein, 'ta': dateout})


@cache_permission('nemodarkamkarkard')
def nemodarkamkarkard2(request):
    tarikh1 = ""
    tarikh2 = ""
    newlist = None
    newlist2 = None
    _id = 0
    _list = None
    nemodar1 = None
    datein = ""
    dateout = ""
    storages = Storage.objects.all()
    if request.method == 'POST':
        _id = request.POST.get('storage')
        datein = str(request.POST.get('select'))
        newlist = []
        newlist2 = []
        dateout = str(request.POST.get('select2'))
        tarikh1 = to_miladi(datein)
        tarikh2 = to_miladi(dateout)
        sdate = date(tarikh1.year, tarikh1.month, tarikh1.day)  # start date
        edate = date(tarikh2.year, tarikh2.month, tarikh2.day)  # end date

        dates_bwn_twodates = edate - sdate

        for n in range(int(dates_bwn_twodates.days + 1)):
            day = sdate + timedelta(days=n)

            result = Store.objects.filter(storage_id=_id, marsole_date__year=day.year, marsole_date__month=day.month,
                                          marsole_date__day=day.day).aggregate(
                jam=Sum('master') + Sum('pinpad'))

            _list = StoreHistory.objects.filter(senddate2__year=day.year,
                                                senddate2__month=day.month,
                                                senddate2__day=day.day, storage_id=_id,
                                                activeday__isnull=False,
                                                activeday__lt=30).aggregate(tedad=Count('id'))
            _send = 0 if result['jam'] == None else str(result['jam'])
            jd = JDate(day.strftime("%Y-%m-%d"))
            _dict = {
                'date': str(jd.year()) + "/" + str(jd.month()) + "/" + str(jd.day()),
                'send': str(_send),

                'val': str(_list['tedad']),
            }
            newlist.append(_dict)

            # nemodar1 = _list.values(year_month=Func(
            #     F('senddate'),
            #     Value('%Y-%m-%d'),
            #     function='DATE_FORMAT',
            #     output_field=CharField())).annotate(counter=Count('id'),
            #                                         starterr=Count(Case(When(starterror=True, then=1))))

            # for item in nemodar1:
            #     _date = item['year_month'].split('-')
            #     tarikh = datetime.date(day=int(_date[2]), month=int(_date[1]), year=int(_date[0]))
            #     jd = JDate(tarikh.strftime("%Y-%m-%d"))
            #     tarikh = jd.format('Y/m/d')

        return TemplateResponse(request, 'store/nemodarkamkarkard2.html',
                                {'nemodar1': newlist, 'list': newlist2, 'storages': storages, 'storagesid': int(_id),
                                 'az': datein, 'ta': dateout})

    return TemplateResponse(request, 'store/nemodarkamkarkard2.html',
                            {'nemodar1': newlist, 'list': newlist2, 'storages': storages, 'storagesid': int(_id),
                             'az': datein, 'ta': dateout})


def mohasebat_hoghogh(_id, pk):
    result_pump = 0
    result_gsoutside = 0
    result_gssabside = 0
    result_gsopenticket = 0
    result_countdaghi = 0
    result_countservice = 0
    result_countnapaydari = 0
    result_negativecount = 0
    _date = Mount.objects.get(id=pk)
    d1 = str(_date.year) + "/" + str(_date.mah) + "/01"
    d2 = str(_date.year) + "/" + str(_date.mah) + "/" + str(_date.day)
    tarikh1 = to_miladi(d1)
    tarikh2 = to_miladi(d2)
    tek = Owner.objects.get(id=_id)
    pumps = Pump.objects.filter(gs__area__zone_id=tek.zone_id, status__status=True)
    tekpumps = pumps.aggregate(kol=Count('id'), tek=Count(
        Case(When(gs__gsowner__owner_id=_id, then=1))))
    pumps = len(pumps)
    t_pump = (int(tekpumps['tek']) / int(pumps)) * 100
    _list = []
    if t_pump >= 12:
        result_pump = 100
    elif 10 <= t_pump < 12:
        result_pump = 90
    elif 8 <= t_pump < 10:
        result_pump = 80
    elif 5 <= t_pump < 8:
        result_pump = 50
    else:
        result_pump = 0
    _dict = {
        'name': 'result_pump',
        'val': result_pump,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------
    gsoutside = GsModel.objects.filter(gsowner__owner_id=_id, gsstatus_id__in=[2, 3]).count()

    if gsoutside >= 5:
        result_gsoutside = 100
    elif 3 <= gsoutside < 5:
        result_gsoutside = 90
    elif 1 <= gsoutside < 3:
        result_gsoutside = 70
    else:
        result_gsoutside = 0
    _dict = {
        'name': 'result_gsoutside',
        'val': result_gsoutside,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------

    gssabside = GsModel.objects.filter(gsowner__owner_id=_id, gsstatus_id=3).count()
    if gssabside >= 5:
        result_gssabside = 100
    elif 3 <= gssabside < 5:
        result_gssabside = 90
    elif 1 <= gssabside < 3:
        result_gssabside = 70
    else:
        result_gssabside = 0
    _dict = {
        'name': 'result_gssabside',
        'val': result_gssabside,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------
    gsopenticket = Ticket.objects.filter(gs__gsowner__owner_id=_id, status_id=1).count()

    if gsopenticket >= 15:
        result_gsopenticket = 0
    elif 12 <= gsopenticket < 15:
        result_gsopenticket = 40
    elif 8 <= gsopenticket < 12:
        result_gsopenticket = 75
    elif 4 <= gsopenticket < 8:
        result_gsopenticket = 50
    else:
        result_gsopenticket = 100

    _dict = {
        'name': 'result_gsopenticket',
        'val': result_gsopenticket,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------

    countdaghi = StoreList.objects.filter(getuser_id=_id, status_id=6).count()
    if countdaghi >= 10:
        result_countdaghi = 0
    elif 7 <= countdaghi < 10:
        result_countdaghi = 60
    elif 4 <= countdaghi < 7:
        result_countdaghi = 85
    elif 0 <= countdaghi < 4:
        result_countdaghi = 100
    else:
        result_countdaghi = 0

    _dict = {
        'name': 'result_countdaghi',
        'val': result_countdaghi,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------

    countservice = Ticket.objects.filter(gs__gsowner__owner_id=_id, status_id=2, reply_id__in=[51, 88, 1],
                                         closedate__range=(tarikh1, tarikh2)).count()
    if countservice >= 20:
        result_countservice = 100
    elif 15 <= countservice < 20:
        result_countservice = 90
    elif 10 <= countservice < 15:
        result_countservice = 75
    elif 5 <= countservice < 9:
        result_countservice = 50
    else:
        result_countservice = 0

    _dict = {
        'name': 'result_countservice',
        'val': result_countservice,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------

    countnapaydari = Ticket.objects.filter(gs__gsowner__owner_id=_id, failure_id=1056, create__range=(tarikh1, tarikh2),
                                           timeaction__gt=2).count()
    if countnapaydari >= 5:
        result_countnapaydari = 0
    elif 4 <= countnapaydari < 5:
        result_countnapaydari = 20
    elif 3 <= countnapaydari < 4:
        result_countnapaydari = 50
    elif 1 <= countnapaydari < 3:
        result_countnapaydari = 80
    else:
        result_countnapaydari = 100

    _dict = {
        'name': 'result_countnapaydari',
        'val': result_countnapaydari,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------

    negativecount = NegativeScore.objects.filter(owner_id=_id, created__range=(tarikh1, tarikh2)).count()
    if negativecount >= 4:
        result_negativecount = 0
    elif 3 <= negativecount < 4:
        result_negativecount = 30
    elif 2 <= negativecount < 3:
        result_negativecount = 50
    elif 1 <= negativecount < 3:
        result_negativecount = 80
    else:
        result_negativecount = 100

    _dict = {
        'name': 'result_negativecount',
        'val': result_negativecount,
    }
    _list.append(_dict)
    # -----------------------------------------------------------------------
    return (_list)


@cache_permission('generator')
def generateserialnumber(request):
    generate = GenerateSerialNumber.objects.filter(active=True)
    zone = Zone.objects_limit.all()
    _newserial = ""
    sname = 0
    oldserial = ""
    mboard = ""
    statusref = 0
    liststatusref = None
    if request.method == 'POST':
        sname = request.POST.get('sname')
        oldserial = request.POST.get('oldserial')
        mboard = request.POST.get('mboard')
        statusref = request.POST.get('statusref')
        liststatusref = StatusRef.objects.all()
        generateserialnumber = GenerateSerialNumber.objects.get(id=sname)
        generateserialnumber.serialnumber += 1
        newserial = generateserialnumber.serialnumber
        if generateserialnumber.status_id == 1:
            st = "M"
        else:
            st = "P"
        _newserial = str(generateserialnumber.partnumber) + "-" + str(st) + "-" + str(newserial)
        generateserialnumber.save()
        storelist = StoreList.objects.filter(serial=oldserial).last()
        storelist.serial = _newserial
        storelist.uniq = _newserial
        storelist.oldserial = oldserial
        storelist.boardserial = mboard
        storelist.status_id = statusref
        store = storelist.id
        storelist.save()

        StoreHistory.objects.create(store_id=store, owner_id=request.user.owner.id, baseroot=0,
                                    information="تولید سریال جدید ",
                                    status_id=statusref, description=f'  ',
                                    )

        messages.info(request, 'شماره سریال جدید ایجاد شد.')

    return TemplateResponse(request, 'generateserialnumber.html', {'generate': generate, 'zone': zone,
                                                                   'liststatusref': liststatusref,
                                                                   'sname': int(sname),
                                                                   'oldserial': oldserial,
                                                                   'sboard': mboard,
                                                                   'statusref': int(statusref),
                                                                   'newserial': _newserial,

                                                                   })


@cache_permission('z_listre')
def listdasghistoreofzone(request):
    _list = []
    for zone in Zone.objects_limit.all():
        _list_master = StoreList.objects.filter(statusstore_id=1, zone_id=zone.id).aggregate(
            tek_master=(Count(Case(When(status_id=6, then=1)))),
            in_tek_master=(Count(Case(When(status_id=10, then=1)))),
            redy_post_master=(Count(Case(When(status_id=11, then=1)))),
            in_kargah_master=(Count(Case(When(status_id=8, then=1)))),
            sum_master=(Count(Case(When(status_id__in=[6, 10, 11, 8], then=1)))),

        )

        _list_pinpad = StoreList.objects.filter(statusstore_id=2, zone_id=zone.id).aggregate(
            tek_pinpad=(Count(Case(When(status_id=6, then=1)))),
            in_tek_pinpad=(Count(Case(When(status_id=10, then=1)))),
            redy_post_pinpad=(Count(Case(When(status_id=11, then=1)))),
            in_kargah_pinpad=(Count(Case(When(status_id=8, then=1)))),
            sum_pinpad=(Count(Case(When(status_id__in=[6, 10, 11, 8], then=1)))),

        )
        _list.append({
            'zone_id': zone.id,
            'zone_name': zone.name,
            'tek_master': _list_master['tek_master'],
            'in_tek_master': _list_master['in_tek_master'],
            'redy_post_master': _list_master['redy_post_master'],
            'in_kargah_master': _list_master['in_kargah_master'],
            'sum_master': _list_master['sum_master'],
            'tek_pinpad': _list_pinpad['tek_pinpad'],
            'in_tek_pinpad': _list_pinpad['in_tek_pinpad'],
            'redy_post_pinpad': _list_pinpad['redy_post_pinpad'],
            'in_kargah_pinpad': _list_pinpad['in_kargah_pinpad'],
            'sum_pinpad': _list_pinpad['sum_pinpad'],
        })

    return TemplateResponse(request, 'store/listdasghistoreofzone.html', {'list': _list})


def reportzonestoretoexcel2(request):
    _list = []
    zone = Zone.objects_limit.all()

    today = timezone.now().date()
    today2 = timezone.now().date()
    for item in zone:
        count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=item.id,
                                                                        status__status='open',
                                                                        Pump__status__status=True,
                                                                        gs__status__status=True,
                                                                        failure__failurecategory_id=1010,
                                                                        failure__isnazel=True).count()

        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=item.id,
                                                                        status__status='open',
                                                                        Pump__status__status=True,
                                                                        gs__status__status=True,
                                                                        failure__failurecategory_id=1011,
                                                                        failure__isnazel=True).count()
        closeticket = Ticket.objects.exclude(organization_id=4, failure__failurecategory_id__in=[1010, 1011]).filter(
            gs__area__zone_id=item.id,
            closedate__year=today.year, closedate__month=today.month, closedate__day=today.day, status__status='close'
        ).count()

        jd = JDate(today.strftime("%Y-%m-%d"))
        today2 = jd.format('Y-n-j')
        master_store = StoreList.objects.filter(zone_id=item.id, status_id__in=[3, 4],
                                                statusstore_id=1).count()
        pinpad_store = StoreList.objects.filter(zone_id=item.id, status_id__in=[3, 4],
                                                statusstore_id=2).count()
        count_pump = Pump.objects.filter(gs__area__zone_id=item.id, status__status=True,
                                         gs__status__status=True).count()
        # محاسبه درصد خرابی (نمونه، نیاز به منطق دقیق دارد)
        if count_master > 0:
            master = ((int(count_master) / int(count_pump)) * 100)
        else:
            master = 0
        if count_pinpad > 0:
            pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
        else:
            pinpad = 0
        failure_percentage = round(master + pinpad, 2)

        _dict = {
            'id': item.id,
            'name': item.name,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master_store': master_store,
            'pinpad_store': pinpad_store,
            'closeticket': closeticket,
            'failure_percentage': round(failure_percentage, 2),
        }
        _list.append(_dict)
    _list = sorted(_list, key=itemgetter('failure_percentage'), reverse=True)

    add_to_log(request, 'ارسال آمار قطعات مناطق به اکسل 2', 0)
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = EXCEL_EXPORT_FILE + 'Status.xlsx'

    wb = Workbook()
    ws1 = wb.active
    ws1.title = f'  گزارش تعداد نیکت بسته شده در روز جاری {today2} '
    ws1.sheet_view.rightToLeft = True
    ws1.page_setup.fitToPage = True

    # استایل‌ها
    _font_header = Font(bold=True, size=12, color='000000')
    _font_title = Font(bold=True, size=14, color='000000')
    _fill_header = PatternFill(fill_type='solid', start_color='FFD700')
    _fill_title = PatternFill(fill_type='solid', start_color='dadfe3')
    _alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # عنوان اصلی
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f' گزارش تعداد نیکت بسته شده در روز جاری   {today2}'
    ws1['A1'].font = _font_title
    ws1['A1'].fill = _fill_title
    ws1['A1'].alignment = _alignment
    ws1['A1'].border = _border
    ws1.row_dimensions[1].height = 30  # تنظیم ارتفاع به ۳۰ پیکسل

    # سرفصل‌های جدول
    headers = [

        ('ردیف', 'A2:A3'),
        ('منطقه', 'B2:B3'),
        ('تعداد نیکت', 'C2:D2'),
        ('موجودی', 'E2:F2'),
        ('درصد خرابی', 'G2:G3'),
        ('تعداد نیکت بسته شده', 'H2:H3'),

    ]
    sub_headers = [
        ('کارتخوان', 'C3'),
        ('صفحه کلید', 'D3'),
        ('کارتخوان', 'E3'),
        ('صفحه کلید', 'F3'),
    ]

    for header, cell_range in headers:
        cell = cell_range.split(':')[0]
        ws1[cell] = header
        ws1[cell].font = _font_header
        ws1[cell].fill = _fill_header
        ws1[cell].alignment = _alignment
        ws1[cell].border = _border
        if ':' in cell_range:
            ws1.merge_cells(cell_range)

    for sub_header, cell in sub_headers:
        ws1[cell] = sub_header
        ws1[cell].font = _font_header
        ws1[cell].fill = _fill_header
        ws1[cell].alignment = _alignment
        ws1[cell].border = _border

    # تنظیم عرض ستون‌ها
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 25
    i = 0
    # اضافه کردن داده‌ها
    for idx, item in enumerate(_list, start=4):
        i += 1
        ws1[f'A{idx}'] = str(i)
        ws1[f'B{idx}'] = f"{item['name']}"
        ws1[f'C{idx}'] = item['count_master']
        ws1[f'D{idx}'] = item['count_pinpad']
        ws1[f'E{idx}'] = item['master_store']
        ws1[f'F{idx}'] = item['pinpad_store']
        percent_value = float(str(item['failure_percentage']).replace('%', ''))
        ws1[f'G{idx}'] = percent_value
        ws1[f'H{idx}'] = item['closeticket']

        # اعمال استایل به سلول‌های داده
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws1[f'{col}{idx}'].font = Font(size=11)
            ws1[f'{col}{idx}'].alignment = _alignment
            ws1[f'{col}{idx}'].border = _border

    _green_fill = PatternFill(
        fill_type='solid',
        start_color='00FF00',  # کد رنگ سبز (می‌توانید از '008000' برای سبز تیره استفاده کنید)
        end_color='00FF00'
    )

    _yellow_fill = PatternFill(
        fill_type='solid',
        start_color='FFFF00',  # کد رنگ زرد (می‌توانید از 'FFD700' برای طلایی استفاده کنید)
        end_color='FFFF00'
    )
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.fill = _yellow_fill

    for row in range(2, 4):  # ردیف‌های ۲ و ۳
        for col in ['E', 'F']:
            ws1[f'{col}{row}'].fill = _green_fill

    max_percent = max(float(item['failure_percentage']) for item in _list)  # بیشترین درصد در داده‌ها

    data_bar_rule = DataBarRule(
        start_type='num', start_value=0,  # کمترین مقدار (۰%)
        end_type='num', end_value=max_percent,  # بیشترین مقدار (مثلاً ۵%)
        color=Color('FF0000'),  # رنگ قرمز
        showValue=True
    )
    ws1.conditional_formatting.add(f'G4:G{ws1.max_row}', data_bar_rule)

    wb.save(response)
    return response


@cache_permission('store_mgr')
def report_initial_defects(request):
    if request.user.owner.role.role in ['setad', 'mgr', 'fani']:
        zones = Zone.objects_limit.all()
    else:
        zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)

    context = {
        'zones': zones,
        'selectzone': 'all',  # مقدار پیش‌فرض برای همه مناطق
        'selected_percentage': 0  # مقدار پیش‌فرض برای درصد
    }

    if request.method == 'POST':
        az = request.POST.get('az', '')
        ta = request.POST.get('ta', '')
        zone = request.POST.get('zone', 'all')  # مقدار پیش‌فرض "همه مناطق"
        percentage_filter = request.POST.get('percentage', '0')  # فیلتر درصد

        # فیلترهای اولیه
        defects = StoreHistory.objects.filter(
            status_id=6,  # وضعیت "از ابتدا معیوب"
            starterror=True
        )

        # اعمال فیلترهای تاریخ
        if az:
            az_date = jdatetime.datetime.strptime(az, '%Y/%m/%d').togregorian()
            defects = defects.filter(create__gte=az_date)
        if ta:
            ta_date = jdatetime.datetime.strptime(ta, '%Y/%m/%d').togregorian()
            defects = defects.filter(create__lte=ta_date)

        # فیلتر منطقه - اگر "all" نبود، فیلتر اعمال شود
        if zone != 'all':
            defects = defects.filter(owner__zone_id=zone)

        # محاسبه آمار برای هر تکنسین
        tech_stats = defects.values(
            'owner__name',
            'owner__lname',
            'owner_id',
            'owner__zone__name'  # اضافه کردن نام منطقه
        ).annotate(
            total_defects=Count('id'),
        ).order_by('-total_defects')

        # محاسبه درصد برای هر تکنسین
        filtered_stats = []
        for stat in tech_stats:
            _name = f"{stat['owner__name']} {stat['owner__lname']}"
            total_received = StoreHistory.objects.filter(
                create__range=(az_date, ta_date),
                description__icontains=_name,
                status_id=4  # وضعیت دریافت توسط تکنسین
            ).count()

            # Add total_received to the stat dictionary
            stat['total_received'] = total_received
            percentage = round((stat['total_defects'] / total_received * 100), 2) if total_received > 0 else 0
            stat['percentage'] = percentage

            # اعمال فیلتر درصد - فقط مواردی که درصدشان از مقدار وارد شده بیشتر است
            try:
                min_percentage = float(percentage_filter)
                if percentage >= min_percentage:
                    filtered_stats.append(stat)
            except (ValueError, TypeError):
                # اگر مقدار درصد معتبر نبود، همه موارد را نشان بده
                filtered_stats.append(stat)

        context = {
            'list': defects,
            'tech_stats': filtered_stats,  # استفاده از لیست فیلتر شده
            'az': az,
            'ta': ta,
            'tek': 'tek',
            'zones': zones,
            'selectzone': zone,
            'selected_percentage': percentage_filter
        }

    return TemplateResponse(request, 'store/start_daghi2.html', context)


@cache_permission('tek_received_list')
def tek_received_list(request):
    """
    نمایش لیست قطعات ارسال شده به تکنسین و امکان تأیید رسید
    """
    add_to_log(request, 'مشاهده لیست قطعات ارسال شده به تکنسین', 0)

    # دریافت قطعات با وضعیت 16 (بین راهی) که برای این تکنسین ارسال شده‌اند
    items = StoreList.objects.filter(
        status_id=16,  # وضعیت بین راهی
        getuser_id=request.user.owner.id  # قطعاتی که به این تکنسین ارسال شده‌اند
    ).order_by('-update')

    if request.method == 'POST':
        # دریافت ID قطعاتی که باید تأیید شوند
        item_ids = request.POST.getlist('item_ids')

        # تغییر وضعیت قطعات به 4 (در دست تکنسین)
        updated = StoreList.objects.filter(
            id__in=item_ids,
            status_id=16,
            getuser_id=request.user.owner.id
        ).update(status_id=4)

        # ثبت تاریخچه برای هر قطعه
        for item_id in item_ids:
            StoreHistory.objects.create(
                store_id=item_id,
                owner_id=request.user.owner.id,
                information="تأیید رسید توسط تکنسین",
                status_id=4,
                description=f'قطعه توسط {request.user.owner.get_full_name()} دریافت شد'
            )

        add_to_log(request, 'زسید قطعه توسط تکنسین', 0)
        messages.success(request, f'{updated} قطعه با موفقیت تأیید رسید شدند.')
        return redirect('pay:tek_received_list')

    return TemplateResponse(request, 'store/tek_received_list.html', {'items': items})


def serials_without_history(request):
    """
    نمایش سریال‌هایی که هیچ سابقه‌ای در StoreHistory ندارند
    """
    # پیدا کردن تمام سریال‌هایی که در StoreHistory وجود ندارند
    serials_without_history = StoreList.objects.annotate(
        history_count=Count('storehistory')
    ).filter(history_count=0)

    if request.method == 'POST' and 'delete_selected' in request.POST:
        # حذف سریال‌های انتخاب شده
        selected_ids = request.POST.getlist('selected_serials')
        if selected_ids:
            deleted_count, _ = StoreList.objects.filter(id__in=selected_ids).delete()
            return JsonResponse({
                'success': True,
                'message': f'{deleted_count} سریال با موفقیت حذف شدند'
            })

    context = {
        'serials': serials_without_history,
        'title': 'سریال‌های بدون سابقه'
    }
    return render(request, 'mgr/serials_without_history.html', context)


@require_http_methods(["DELETE"])
def delete_serial_without_history(request, serial_id):
    """
    حذف یک سریال خاص که سابقه‌ای ندارد
    """
    try:
        # بررسی وجود سریال و نداشتن سابقه
        serial = StoreList.objects.annotate(
            history_count=Count('storehistory')
        ).get(id=serial_id, history_count=0)

        serial_name = serial.serial
        serial.delete()

        return JsonResponse({
            'success': True,
            'message': f'سریال {serial_name} با موفقیت حذف شد'
        })
    except StoreList.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'سریال پیدا نشد یا دارای سابقه است'
        }, status=404)


def change_status_by_date(request):
    """
    تغییر وضعیت سریال‌ها بر اساس وضعیت فعلی و تاریخ بروزرسانی
    """

    if request.method == 'POST':

        # دریافت پارامترها از فرم
        current_status_id = request.POST.get('current_status')
        cutoff_date = request.POST.get('cutoff_date')
        new_status_id = request.POST.get('new_status')

        try:
            cutoff_date = to_miladi(cutoff_date)
            # تبدیل تاریخ به فرمت مناسب

            # پیدا کردن سریال‌های مطابق با شرایط
            target_serials = StoreList.objects.filter(
                status_id=current_status_id,
                update__lt=cutoff_date
            )

            count = target_serials.count()

            if 'confirm' in request.POST and count > 0:
                # انجام تغییر وضعیت
                target_serials.update(status_id=new_status_id)

                return JsonResponse({
                    'success': True,
                    'message': f'وضعیت {count} سریال با موفقیت تغییر کرد'
                })

            # نمایش پیش‌نمایش
            context = {
                'target_serials': target_serials,
                'count': count,
                'current_status': StatusRef.objects.get(id=current_status_id),
                'new_status': StatusRef.objects.get(id=new_status_id),
                'cutoff_date': cutoff_date,
                'show_preview': True
            }

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'خطا در پردازش: {str(e)}'
            })

    else:
        context = {'show_preview': False}

    # لیست وضعیت‌های موجود برای dropdown
    context['statuses'] = StatusRef.objects.all()
    return render(request, 'mgr/change_status_by_date.html', context)

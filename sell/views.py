from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models.functions import Abs
from django.db.models.functions import Round
from django.template.response import TemplateResponse
from django.utils.decorators import method_decorator
from django.views.generic import ListView, UpdateView
from django.urls import reverse
from base.forms import open_excel
from base.permission_decoder import cache_permission
from base.views import zoneorarea
from utils.exception_helper import to_miladi
from operator import itemgetter
from django.db.models import Sum, When, Case, F, DecimalField, ExpressionWrapper, IntegerField, Value
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from accounts.logger import add_to_log
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from api.views import is_ajax
from base.models import UploadExcel, Area, GsList, CloseGS, \
    AutoExcel, City
from .analyzers import PolicyAnalyzer
from .forms import UploadSoratjalaseForm
from django.db.models.functions import Coalesce
from .models import *
from django.middleware.csrf import get_token
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from base.models import GsModel, Pump
import openpyxl
from django.db.models import Count, Avg, Sum, Q, Max, Min
from .qrreader import load_code, load_rpm_code
from datetime import datetime as newdate, date, timedelta
import datetime
from functools import lru_cache
from django.core.cache import cache

today = str(jdatetime.date.today())

today = today.replace("-", "/")
startdate = today[:8]
startdate = startdate + "01"


@cache_permission('listsell')
def addsell(request, id):
    stmojodi = True
    gsnow = None
    if request.user.owner.role.role == 'zone':
        gslist = GsModel.objects.filter(id=id,
                                        area__zone_id=request.user.owner.zone_id)
    elif request.user.owner.role.role == 'area':
        gslist = GsModel.objects.filter(id=id,
                                        area_id=request.user.owner.area_id)
    else:
        gslist = GsList.objects.filter(
            owner_id=request.user.owner.id, gs_id=id).count()

    gsnow = GsModel.objects.get(id=id)
    try:
        if gsnow.gsipclog.get().dashboard_version != '1.03.052001':
            stmojodi = False
    except:
        pass
    gsname = gsnow.name + " | " + str(gsnow.gsid) + " | ناحیه: " + gsnow.area.name
    if gslist == 0:
        messages.error(request, 'شما دسترسی برای اطلاعات این جایگاه ندارید')
        return redirect('sell:listsell')
    csrf_token = get_token(request)
    gs = Pump.objects.filter(gs_id=id, status_id__in=[1, 2]).order_by('number')
    olddate = SellModel.objects.filter(gs_id=id).order_by('-tarikh').first()
    if olddate:
        result = SellModel.objects.filter(tarikh=olddate.tarikh, gs_id=id).order_by('tolombeinfo')
    else:
        result = None
    if request.method == 'POST':
        datein = str(request.POST.get('select'))

        datein = datein.split("/")
        tarikh = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()

        for item in gs:
            end = request.POST.get(f'end{item.number}')
            start = request.POST.get(f'start{item.number}')
            sell = request.POST.get(f'sell{item.number}')
            num = request.POST.get(f'num{item.number}')
            sellkol_h = request.POST.get(f'sellkol{item.number}')
            id_ekhtelaf_h = request.POST.get(f'ekhtelaf_h{item.number}')
            yarane = request.POST.get(f'yarane{item.number}')
            azad = request.POST.get(f'azad{item.number}')
            ezterari = request.POST.get(f'ezterari{item.number}')
            mojaz = request.POST.get(f'mojaz_h{item.number}')
            nomojaz = request.POST.get(f'nomojaz_h{item.number}')
            haveleh = request.POST.get(f'havale{item.number}')
            azmayesh = request.POST.get(f'azmayesh{item.number}')
            _role = request.user.owner.role.role
            _roleid = zoneorarea(request)

            SellModel.object_role.c_gs(request, 0).create(gs_id=id, tolombeinfo_id=item.id, start=start, end=end,
                                                          sell=sell,
                                                          ezterari=ezterari,
                                                          tarikh=tarikh, yarane=yarane, azad=azad,
                                                          sellkol=sellkol_h, ekhtelaf=id_ekhtelaf_h,
                                                          mojaz=mojaz, nomojaz=nomojaz, nomojaz2=nomojaz,
                                                          haveleh=haveleh, azmayesh=azmayesh,
                                                          uniq=str(tarikh) + "-" + str(id) + "-" + str(item.id))
        return redirect('sell:listsell')
    return TemplateResponse(request, 'addsell2.html',
                            {'csrf_token': csrf_token, 'gs': gs, 'result': result, 'today': today, 'id': id,
                             'gsnow': gsnow,
                             'gsname': gsname, 'stmojodi': stmojodi,
                             })


@cache_permission('listcrash')
def crashhard(request):
    csrf_token = get_token(request)
    gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)

    if request.method == 'POST':
        add_to_log(request, f'مشاهده فرم تعویض هارد ', 0)
        datein = str(request.POST.get('select'))
        dateout = str(request.POST.get('select2'))
        datesell = str(request.POST.get('select3'))
        gs = str(request.POST.get('gs'))

        datein = datein.split("/")
        dateout = dateout.split("/")
        datesell = datesell.split("/")
        tarikhin = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        tarikhout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        crashdate = jdatetime.date(day=int(datesell[2]), month=int(datesell[1]), year=int(datesell[0])).togregorian()

        return redirect('sell:listsell')
    return TemplateResponse(request, 'crashhard.html',
                            {'csrf_token': csrf_token, 'today': today, 'gss': gss, 'id': id,
                             })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pumplist(request):
    mylist = []
    _id = request.GET.get('gsid')
    gs = Pump.objects.filter(gs_id=_id).order_by('number')
    for item in gs:
        dict = {
            'id': item.id,
            'product': item.product.name,
            'number': item.number
        }
        mylist.append(dict)

    return JsonResponse({"message": "success", "mylist": mylist})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def addnazel(request):
    if request.method == 'POST':
        benzin = str(request.POST.get('benzin_mojodi'))
        super = str(request.POST.get('super_mojodi'))
        gaz = str(request.POST.get('gaz_mojodi'))
        datein = str(request.POST.get('tarikh'))
        datein = datein.split("/")
        tarikh = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        end = request.POST.get('end')
        start = request.POST.get('start')
        sell = request.POST.get('sellshow')
        number = request.POST.get('number')
        gs = request.POST.get('gsid')
        sellkol_h = request.POST.get('sellkol')
        id_ekhtelaf = request.POST.get('id_ekhtelaf')
        yarane = request.POST.get('yarane')
        azad = request.POST.get('azad')
        ezterari = request.POST.get('ezterari')
        mojaz = request.POST.get('mojaz_h')
        nomojaz = request.POST.get('nomojaz_h')
        haveleh = request.POST.get('havale')
        azmayesh = request.POST.get('azmayesh')
        pumpnumber = Pump.objects.get(id=int(number))
        SellModel.objects.filter(gs_id=gs, tolombeinfo_id=number, tarikh=tarikh).delete()
        SellModel.objects.create(gs_id=gs, tolombeinfo_id=number, start=start, end=end, sell=sell,
                                 ezterari=ezterari, pumpnumber=pumpnumber.number,
                                 tarikh=tarikh, yarane=yarane, azad=azad, sellkol=sellkol_h, ekhtelaf=id_ekhtelaf,
                                 mojaz=mojaz, nomojaz=nomojaz, nomojaz2=nomojaz, haveleh=haveleh, azmayesh=azmayesh,
                                 uniq=str(tarikh) + "-" + str(gs) + "-" + str(number))
        # Mojodi.objects.create(gs_id=gs, tarikh=tarikh, benzin=benzin, super=super, gaz=gaz,
        #                       uniq=str(gs) + '-' + str(tarikh))

        return JsonResponse({"message": "success"})


@cache_permission('listsell')
def listsell(request):
    add_to_log(request, f'مشاهده فرم لیست فرم 1502 ', 0)
    parametr = Parametrs.objects.all().first()
    if request.user.owner.role.role == 'zone':
        gslist = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id,
                                        status__status=True)
    elif request.user.owner.role.role == 'area':
        gslist = GsModel.objects.filter(area_id=request.user.owner.area_id,
                                        status__status=True)

    else:

        gslist = GsList.objects.filter(owner_id=request.user.owner.id)
    mygsid = GsList.objects.filter(owner_id=request.user.owner.id).first()
    if request.method == 'POST':
        id = int(request.POST.get('mygsid'))
        sell = SellModel.objects.values('gs', 'tarikh', 'iscrash').filter(gs_id=id, sellkol__gt=0).annotate(
            mekaniki=Sum('sell'),
            elektroniki=Sum('sellkol')).order_by(
            '-tarikh')[:30]
    else:
        id = gslist.first().id if request.user.owner.role.role in ['zone', 'area'] else gslist.first().gs_id
        sell = SellModel.objects.values('gs', 'tarikh', 'iscrash').filter(gs_id=id, sellkol__gt=0).annotate(
            mekaniki=Sum('sell'),
            elektroniki=Sum('sellkol')).order_by(
            '-tarikh')[:30]
    return TemplateResponse(request, 'listsell.html', {'sell': sell, 'id': id, 'gslist': gslist, 'mygsid': mygsid,
                                                       'parametr': parametr
                                                       })


@cache_permission('listcrash')
def listcrash(request):
    add_to_log(request, 'مشاهده فرم لیست تعویض هارد', 0)

    zone_id = request.user.owner.zone_id

    # فقط فیلدهای مورد نیاز + select_related
    sell = SellModel.object_role.c_gs(request, 0).filter(
        gs__area__zone_id=zone_id,
        iscrash=True
    ).select_related('gs') \
               .only('gs_id', 'tarikh', 'sell', 'sellkol', 'gs__name') \
               .values('gs_id', 'gs__name', 'tarikh') \
               .annotate(
        mekaniki=Sum('sell'),
        elektroniki=Sum('sellkol')
    ).order_by('-tarikh')[:60]

    return TemplateResponse(request, 'listcrash.html', {'sell': sell})


@api_view(['POST'])
def deletecng(request):
    if is_ajax(request=request):
        _date = request.POST.get('datein')
        _id = request.POST.get('newid')
    # Mojodi.objects.filter(gs_id=_id, tarikh=_date).delete()
    SellModel.objects.filter(tarikh=_date, gs_id=_id).delete()
    SellGs.objects.filter(tarikh=_date, gs_id=_id).delete()
    OpenCloseSell.objects.filter(dore=_date, gs_id=_id).delete()
    add_to_log(request, 'حذف فروش ' + str(_date) + " - " + str(_id), _id)

    return JsonResponse({'ok': 1})


@cache_permission('listsell')
def updatesell(request, mydate, id):
    selllist = SellModel.objects.filter(tarikh=mydate, gs_id=id)
    gs = Pump.objects.filter(gs_id=id).order_by('number')
    if request.method == 'POST':
        datein = str(request.POST.get('select'))

        datein = datein.split("-")
        tarikh = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        SellModel.objects.filter(tarikh=mydate, gs_id=id).delete()
        for item in gs:
            end = request.POST.get(f'end{item.number}')
            start = request.POST.get(f'start{item.number}')
            sell = request.POST.get(f'sell{item.number}')
            num = request.POST.get(f'num{item.number}')
            sellkol_h = request.POST.get(f'sellkol{item.number}')
            id_ekhtelaf_h = request.POST.get(f'ekhtelaf_h{item.number}')
            yarane = request.POST.get(f'yarane{item.number}')
            azad = request.POST.get(f'azad{item.number}')
            ezterari = request.POST.get(f'ezterari{item.number}')
            mojaz = request.POST.get(f'mojaz_h{item.number}')
            nomojaz = request.POST.get(f'nomojaz_h{item.number}')
            haveleh = request.POST.get(f'havale{item.number}')
            azmayesh = request.POST.get(f'azmayesh{item.number}')

            SellModel.objects.create(gs_id=id, tolombeinfo_id=item.id, start=start, end=end, sell=sell,
                                     ezterari=ezterari,
                                     yarane=yarane, azad=azad, sellkol=sellkol_h, ekhtelaf=id_ekhtelaf_h,
                                     mojaz=mojaz, nomojaz=nomojaz, nomojaz2=nomojaz, tarikh=tarikh, haveleh=haveleh,
                                     azmayesh=azmayesh,
                                     uniq=str(tarikh) + "-" + str(id) + "-" + str(item.id))
        return redirect('sell:listsell')
    return TemplateResponse(request, 'updatesell.html', {'selllist': selllist, 'today': mydate,
                                                         })


@cache_permission('reportsell')
def reportsell(request):
    url = request.META.get('HTTP_REFERER')
    add_to_log(request, f'مشاهده گزارش فروش ', 0)
    mdate = startdate
    mdate2 = today
    area = Area.objects.filter(zone_id=request.user.owner.zone_id)
    az = mdate
    ta = mdate2
    cityid = 0
    areaid = 0
    area = None
    city = None
    if request.user.owner.role.role == 'gs':
        GsList.objects.filter(owner_id=request.user.owner.id)
        gss = GsModel.objects.filter(gsowner__owner=request.user.owner)
    if request.user.owner.role.role == 'zone':
        gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
        area = Area.objects.filter(zone_id=request.user.owner.zone_id)
    if request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area_id)
        city = City.objects.filter(area_id=request.user.owner.area_id)
    if request.user.owner.role.role in ['mgr', 'setad']:
        gss = GsModel.objects.all()

    product = Product.objects.all()

    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        if request.user.owner.role.role == 'zone':
            areaid = request.POST.get('area')
            if int(areaid) >= 0:
                city = City.objects.filter(area_id=areaid)

        if request.user.owner.role.role in 'zone,area':
            cityid = request.POST.get('city')
        az = mdate
        ta = mdate2

        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        gsid = request.POST.get('select3')
        far = int(request.POST.get('select4'))
        gorun = request.POST.get('gorun')
        if gorun == "2":
            if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
                messages.warning(request,
                                 'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
                return redirect(url)

            AutoExcel.objects.create(
                datein=str(request.POST.get('select')),
                dateout=str(request.POST.get('select2')),
                titr=gsid,
                owner_id=request.user.owner.id,
                reportmodel=7,
                other_id=far
            )

            messages.warning(request, 'نتیجه عملیات  مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')
        else:
            _role = request.user.owner.role.role
            _roleid = zoneorarea(request)
            if request.user.owner.role.role == 'gs':
                _gslist = GsList.objects.filter(gs_id=gsid, owner_id=request.user.owner.id)
                if _gslist:
                    sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid)
                else:
                    _gslist = GsList.objects.filter(owner_id=request.user.owner.id).first()
                    sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=_gslist.gs_id)
            if request.user.owner.role.role == 'zone':
                if gsid == '0':
                    sellmodel = SellModel.object_role.c_gs(request, 0).filter(
                        gs__area__zone_id=request.user.owner.zone_id)
                else:
                    sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid,
                                                                              gs__area__zone_id=request.user.owner.zone_id)
            if request.user.owner.role.role == 'area':
                if gsid == '0':
                    sellmodel = SellModel.object_role.c_gs(request, 0).filter(
                        gs__area_id=request.user.owner.area_id)
                else:
                    sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid,
                                                                              gs__area_id=request.user.owner.area_id)
            if request.user.owner.role.role in ['setad', 'mgr']:
                if gsid == '0':
                    sellmodel = SellModel.objects.all()
                else:
                    sellmodel = SellModel.objects.filter(gs__exact=gsid)
            _list = sellmodel.values('gs__gsid', 'gs__name', 'gs__area__name').filter(tarikh__gte=mdate,
                                                                                      tarikh__lte=mdate2,
                                                                                      product_id=far).annotate(
                res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
                sum_ekhtelaf=Sum('nomojaz'), sum_havaleh=Sum('haveleh'), sum_azmayesh=Sum('azmayesh'),
                sum_sellkol=Sum('sellkol')).order_by('gs__area_id', 'gs_id')

            if int(areaid) > 0:
                _list = _list.filter(tolombeinfo__gs__area_id=areaid)

            if int(cityid) > 0:
                _list = _list.filter(tolombeinfo__gs__city=cityid)
            summer = _list.aggregate(sellall=Sum('res'), sum_azadall=Sum('sum_azad'),
                                     sum_ezterariall=Sum('sum_ezterari'),
                                     sum_yaraneall=Sum('sum_yarane'), sum_havalehall=Sum('sum_havaleh'),
                                     sum_azmayeshall=Sum('sum_azmayesh'),
                                     sum_sellkollall=Sum('sum_sellkol'), sum_ekhtelafall=Sum('sum_ekhtelaf'))

            return TemplateResponse(request, 'reportsell.html',
                                    {'list': _list, 'mdate': mdate, 'mdate2': mdate2, 'gss': gss, 'summer': summer,
                                     'cityid': int(cityid), 'areaid': int(areaid),
                                     'gsid': int(gsid), 'city': city, 'area': area,
                                     'product': product, 'far': far,
                                     'az': az, 'ta': ta})
    return TemplateResponse(request, 'reportsell.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'gss': gss, 'product': product,
                             'city': city,
                             'area': area,
                             })


@cache_permission('sellday')
def sellday(request):
    url = request.META.get('HTTP_REFERER')
    add_to_log(request, " تهیه گزارش  فروش روزانه ", 0)
    mdate = startdate
    mdate2 = str(jdatetime.date.today() - timedelta(days=1))
    mdate2 = mdate2.replace('-', '/')
    az = mdate2
    ta = mdate2
    if request.user.owner.role.role == 'gs':
        GsList.objects.filter(owner_id=request.user.owner.id)
        gss = GsModel.objects.filter(gsowner__owner=request.user.owner)
    if request.user.owner.role.role == 'zone':
        gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    if request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area_id)
    if request.user.owner.role.role in ['mgr', 'setad']:
        gss = GsModel.objects.all()

    product = Product.objects.all()
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        az = mdate
        ta = mdate2
        _role = request.user.owner.role.role
        _roleid = zoneorarea(request)
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        gsid = request.POST.get('select3')
        nazelid = int(request.POST.get('nazel'))
        d1 = to_miladi(mdate)
        d2 = to_miladi(mdate2)
        d3 = (d2 - d1).days
        if d3 > 31:
            messages.warning(request, 'بازه تاریخ نباید بیش از 31 روز باشد')
            return redirect(url)
        if gsid == '0' and d3 > 1:
            messages.warning(request, 'بازه برای همه جایگاه ها باید یک روز باشد')
            return redirect(url)
        far = int(request.POST.get('select4'))
        pump = None
        if nazelid != 0:
            pump = Pump.objects.get(id=nazelid)
        pumplist = Pump.objects.filter(gs_id=gsid, product_id=far)

        if request.user.owner.role.role == 'gs':
            _gslist = GsList.objects.filter(gs_id=gsid, owner_id=request.user.owner.id)
            if _gslist:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid)
            else:
                _gslist = GsList.objects.filter(owner_id=request.user.owner.id).first()
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=_gslist.gs_id)
        if request.user.owner.role.role == 'zone':
            if gsid == '0':
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(
                    gs__area__zone_id=request.user.owner.zone_id)
            else:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid,
                                                                          gs__area__zone_id=request.user.owner.zone_id)
        if request.user.owner.role.role == 'area':
            if gsid == '0':
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__area_id=request.user.owner.area_id)
            else:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid,
                                                                          gs__area_id=request.user.owner.area_id)
        if request.user.owner.role.role in ['setad', 'mgr']:
            if gsid == '0':
                sellmodel = SellModel.object_role.c_gs(request, 0).all()
            else:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid)
        if nazelid != 0:
            sellmodel = sellmodel.filter(tolombeinfo_id=nazelid)

        _list = sellmodel.values('gs_id', 'gs__gsid', 'gs__name', 'gs__area__name', 'tarikh', 'iscrash',
                                 'gs__isselldelete').filter(
            tarikh__gte=mdate,
            tarikh__lte=mdate2,
            product_id=far).annotate(
            res=Sum('sell'), sum_azad1=Sum('azad1'), sum_ezterari=Sum('ezterari'),
            sum_yarane=Sum('yarane') - Sum('haveleh'),
            sum_nimeyarane=Sum('nimeyarane'),
            sumlock=(Count(Case(When(islocked=False, then=1)))),
            sum_havaleh=Sum('haveleh'),
            sum_azmayesh=Sum('azmayesh'),
            sum_nomojaz=Sum(Abs('nomojaz')),
            sum_nomojaz2=Sum(Abs('nomojaz2')),
            sum_mogh=Sum('mogh'),
            sum_is_change_counter=Sum('is_change_counter'),
            sum_is_soratjalase=Sum('is_soratjalase'),
            sum_sellkol=Sum('sellkol')).order_by('gs__area_id', 'gs_id', 'tarikh')

        summer = _list.aggregate(sum_azadall=Sum('sum_azad1'), sum_ezterariall=Sum('sum_ezterari'), sellall=Sum('res'),
                                 sellnomojaz=Sum('sum_nomojaz'), sellnomojaz2=Sum('sum_nomojaz2'),
                                 sum_yaraneall=Sum('sum_yarane'),sum_nimeyaraneall=Sum('sum_nimeyarane'), sum_sellkollall=Sum('sum_sellkol'),
                                 sumhavaleh=Sum('sum_havaleh'), sumazmayesh=Sum('sum_azmayesh'))

        return TemplateResponse(request, 'selldaily.html',
                                {'list': _list, 'mdate': mdate, 'mdate2': mdate2, 'gss': gss, 'gsid': int(gsid),
                                 'product': product, 'far': far, 'summer': summer, 'nazelid': nazelid, 'pump': pump,
                                 'pumplist': pumplist,
                                 'az': az, 'ta': ta})
    return TemplateResponse(request, 'selldaily.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'gss': gss, 'product': product,
                             })


@cache_permission('reports2')
def reportsellkol(request):
    add_to_log(request, " تهیه گزارش  فروش منطقه ", 0)
    zone = Zone.objects_limit.all()
    area = Area.objects.filter(zone_id=request.user.owner.zone_id)
    cityid = 0
    mdate = jdatetime.date.today()
    mdate2 = jdatetime.date.today()
    product = Product.objects.all()
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    city = None
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if request.user.owner.role.role == 'area':
        city = City.objects.filter(area_id=request.user.owner.area_id)
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        zoneid = request.POST.get('zone')
        areaid = request.POST.get('area')
        if request.user.owner.role.role in 'zone,area':
            cityid = request.POST.get('city')
        az = mdate
        ta = mdate2

        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        far = int(request.POST.get('select4'))
        statuscode = int(request.POST.get('select5'))
        if not areaid:
            areaid = 0
        else:
            city = City.objects.filter(area_id=areaid)
        if not cityid:
            cityid = 0
        if statuscode == 2:
            if int(areaid) == 0:
                sellgs = SellGs.object_role.c_gs(request, 0).values('gs__area__zone', 'gs__area__zone__name',
                                                                    'tarikh')
            else:
                sellgs = SellGs.object_role.c_gs(request, 0).values('gs__area', 'gs__area__name',
                                                                    'tarikh')
        else:
            if int(areaid) == 0:
                sellgs = SellGs.object_role.c_gs(request, 0).values('gs__area__zone', 'gs__area__zone__name')
            else:
                sellgs = SellGs.object_role.c_gs(request, 0).values('gs__area', 'gs__area__name')
        # if request.user.owner.role.role == 'zone':
        if zoneid != '0':
            sellgs = sellgs.filter(gs__area__zone_id=zoneid)
        if int(areaid) > 1:
            sellgs = sellgs.filter(gs__area_id=areaid)
        if int(cityid) > 0:
            sellgs = sellgs.filter(gs__city_id=cityid)

        list = sellgs.filter(tarikh__gte=mdate,
                             tarikh__lte=mdate2,
                             product_id=far,
                             ).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_havaleh=Sum('haveleh'), sum_azmayesh=Sum('azmayesh'),
            sum_ekhtelaf=(Sum('azad') + Sum('ezterari') + Sum('yarane') + Sum('azmayesh')) - Sum('sell'),
            sum_sellkol=(Sum('azad') + Sum('ezterari') + Sum('yarane') + Sum('azmayesh'))).order_by(
            'gs__area__zone')

        jam = list.aggregate(sellall=Sum('res'))
        summer = list.aggregate(sum_azadall=Sum('sum_azad'), sum_ekhtelafall=Sum('sum_sellkol') - Sum('res'),
                                sum_ezterariall=Sum('sum_ezterari'),
                                sum_yaraneall=Sum('sum_yarane'), sum_sellkollall=Sum('sum_sellkol'),
                                sum_havalehall=Sum('sum_havaleh'), sum_azmayeshall=Sum('sum_azmayesh'))

        return TemplateResponse(request, 'reportsellkol.html',
                                {'list': list, 'product': product, 'mdate': mdate, 'mdate2': mdate2, 'jam': jam,
                                 'statuscode': statuscode, 'city': city, 'cityid': int(cityid),
                                 'summer': summer, 'far': far,
                                 'az': az, 'ta': ta, 'zone': zone, 'area': area,
                                 'zoneid': int(zoneid), 'areaid': int(areaid)})
    return TemplateResponse(request, 'reportsellkol.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'product': product, 'city': city,
                             'zone': zone, 'area': area})


@cache_permission('listsell')
def printsell(request, date, id):
    today = str(jdatetime.date.today())
    today = today.replace("-", "/")
    if request.user.owner.role.role in ['mgr', 'setad']:
        general = GsModel.objects.get(id=id)
    else:
        try:
            if request.user.owner.role.role in ['zone', 'area']:
                general = GsModel.objects.get(id=id, area__zone_id=request.user.owner.zone_id)
            if request.user.owner.role.role in ['gs']:
                gnow = GsList.objects.get(gs_id=id, owner_id=request.user.owner.id)
                general = GsModel.objects.get(id=gnow.gs.id)
        except GsModel.DoesNotExist:
            messages.warning(request, 'شما به این اطلاعات دسترسی ندارید')
            return redirect('base:home')
    sell = SellModel.objects.filter(gs_id=id, tarikh=date).order_by('tolombeinfo__number')
    hdcrash = False
    try:
        if sell.last().iscrash:
            hdcrash = True
    except TypeError:
        hdcrash = False

    list = SellModel.objects.values('product__name').filter(gs_id=id, tarikh=date).annotate(
        res=Sum('sell'), sum_sellkol=Sum('sellkol'))
    _edits = EditSell.objects.filter(sell__gs__id=id, sell__tarikh=date)
    tlist = []
    try:
        mojodi = Mojodi.objects.get(gs_id=id, tarikh=date)
        tdic = {
            'benzin': mojodi.benzin,
            'super': mojodi.super,
            'gaz': mojodi.gaz,
        }
        tlist.append(tdic)
    except ObjectDoesNotExist:

        tlist = False
    add_to_log(request, f'چاپ 1502 ', 0)

    return TemplateResponse(request, 'printsell.html',
                            {'sell': sell, 'general': general, 'date': date, 'list': list, 'hdcrash': hdcrash,
                             'today': today,
                             'edits': _edits, 'tlist': tlist,
                             })


def nazelrow(request):
    mylist = []
    pumpname = ""
    startold = ""

    tlist = []
    if request.method == 'POST':
        val = request.POST.get('val')
        tarikh = request.POST.get('tarikh')
        tarikh = tarikh.split("/")
        tarikh = jdatetime.date(day=int(tarikh[2]), month=int(tarikh[1]), year=int(tarikh[0])).togregorian()
        id = request.POST.get('gsid')
        pump = Pump.objects.get(id=int(val))
        is_close_sell = OpenCloseSell.objects.filter(dore=tarikh, gs_id=int(id)).last()
        is_close_sell = is_close_sell.status if is_close_sell else 'open'

        ac = AccessChangeSell.objects.filter(gs_id=int(id), pump_id=int(pump.id), tarikh=tarikh,
                                             editor_id=request.user.owner.id, active=True).count()
        if ac > 0:
            isedit = 1
        else:
            isedit = 0

        ac = AccessChangeSell.objects.filter(gs_id=int(id), pump_id__isnull=True, tarikh=tarikh,
                                             editor_id=request.user.owner.id, active=True).count()
        if ac > 0:
            isedit = 1
        if pump.gs.isqrcode:
            isedit = 1

        try:
            mojodi = Mojodi.objects.get(gs_id=id, tarikh=tarikh)
            tdic = {
                'benzin': mojodi.benzin,
                'super': mojodi.super,
                'gaz': mojodi.gaz,
            }
            tlist.append(tdic)
        except ObjectDoesNotExist:

            tdic = {
                'benzin': 0,
                'super': 0,
                'gaz': 0,
            }
            tlist.append(tdic)
        if pump.product_id == 2:
            pumpname = "b"
        if pump.product_id == 3:
            pumpname = "s"
        if pump.product_id == 4:
            pumpname = "g"
        if pump.product_id == 948:
            pumpname = "m"

        try:

            _endsell = SellModel.objects.filter(gs_id=int(id), tolombeinfo_id=int(val), start__gt=0).order_by(
                '-tarikh').first()
            endsell = _endsell.start

        except:

            endsell = 0
            isedit = 1
        tarikh_yesterday = tarikh - datetime.timedelta(days=1)
        iscrash = CloseGS.objects.filter(date_out=tarikh_yesterday, gs_id=int(id), status=2)
        if iscrash:
            isedit = 1

        try:
            start = SellModel.objects.get(gs_id=int(id), tarikh=tarikh, tolombeinfo_id=int(val))

            if start:

                if len(str(start.start)) < 1 or start.start == 0:

                    olddate = SellModel.objects.exclude(id=start.id).filter(gs_id=int(id), start__gte='1',
                                                                            tolombeinfo_id=int(val)).order_by(
                        '-tarikh').first()
                    if olddate:
                        startold = olddate.start
                else:
                    startold = start.end

                if start.product_id == 4:
                    _yarane = 0
                else:
                    _yarane = start.yarane
                dict = {
                    'locked': start.islocked,
                    'start': start.start,
                    'start2': start.t_start,
                    'end2': start.t_end,
                    'end': startold,
                    'sell': start.sell,
                    'endsell': endsell,
                    'yarane': _yarane,
                    'nimeyarane': start.nimeyarane,
                    'azad': start.azad1,
                    'ezterari': start.ezterari,
                    'azmayesh': start.azmayesh,
                    'havaleh': start.haveleh,
                    'sellkol': start.sellkol,
                    'mojaz': start.mojaz,
                    'ekhtelaf': start.ekhtelaf,
                    'nomojaz': start.nomojaz,
                    'pumpname': pumpname,
                }

                mylist.append(dict)
        except ObjectDoesNotExist:
            olddate = SellModel.objects.filter(gs_id=id, start__gt=0, tolombeinfo_id=val).order_by('-tarikh').first()
            if olddate:
                start = SellModel.objects.filter(tarikh=olddate.tarikh, gs_id=id, tolombeinfo_id=val).first()
                dict = {
                    'locked': False,
                    'start': '0',
                    'start2': '0',
                    'end2': '0',
                    'end': start.start,
                    'sell': 0,
                    'endsell': 0,
                    'yarane': '0',
                    'nimeyarane': '0',
                    'azad': '0',
                    'ezterari': '0',
                    'azmayesh': '0',
                    'havaleh': '0',
                    'sellkol': 0,
                    'mojaz': 0,
                    'ekhtelaf': 0,
                    'nomojaz': 0,
                    'pumpname': pumpname,
                }
                mylist.append(dict)
            else:
                dict = {
                    'locked': False,
                    'start': '0',
                    'end': '0',
                    'start2': '0',
                    'end2': '0',
                    'sell': 0,
                    'endsell': 0,
                    'yarane': '0',
                    'nimeyarane': '0',
                    'azad': '0',
                    'ezterari': '0',
                    'azmayesh': '0',
                    'havaleh': '0',
                    'sellkol': 0,
                    'mojaz': 0,
                    'ekhtelaf': 0,
                    'nomojaz': 0,
                    'pumpname': pumpname,
                }

                mylist.append(dict)

    return JsonResponse(
        {'message': 'success', 'mylist': mylist, 'tlist': tlist, 'isedit': isedit, 'is_close_sell': is_close_sell})


def nazelrow2(request):
    mylist = []
    pumpname = ""
    startold = ""
    tlist = []
    crashin = 0
    if request.method == 'POST':
        val = request.POST.get('val')
        if len(val) == 0:
            print(4)
            return JsonResponse({'message': 'error', 'mylist': None})
        # crashmodel = int(request.POST.get('crashmodel'))
        _qrtime = request.POST.get('qrtime')
        tarikh = request.POST.get('tarikh')
        tarikh2 = request.POST.get('tarikh2')
        tarikh3 = request.POST.get('tarikh3')
        tarikh = tarikh.split("/")
        tarikh2 = tarikh2.split("/")
        tarikh3 = tarikh3.split("/")
        tarikh = jdatetime.date(day=int(tarikh[2]), month=int(tarikh[1]), year=int(tarikh[0])).togregorian()
        tarikh2 = jdatetime.date(day=int(tarikh2[2]), month=int(tarikh2[1]), year=int(tarikh2[0])).togregorian()
        tarikh3 = jdatetime.date(day=int(tarikh3[2]), month=int(tarikh3[1]), year=int(tarikh3[0])).togregorian()
        # crashdate = tarikh - datetime.timedelta(days=1)
        # if crashmodel == 1:
        #     crashdate2 = tarikh2 - datetime.timedelta(days=1)
        # else:
        crashdate2 = tarikh3
        id = request.POST.get('gsid')

        try:
            pump = Pump.objects.get(id=int(val))


        # try:
        #     mojodi = Mojodi.objects.get(gs_id=id, tarikh=tarikh)
        #     tdic = {
        #         'benzin': mojodi.benzin,
        #         'super': mojodi.super,
        #         'gaz': mojodi.gaz,
        #     }
        #     tlist.append(tdic)
        # except ObjectDoesNotExist:

            if pump.product_id == 2:
                pumpname = "b"
            if pump.product_id == 3:
                pumpname = "s"
            if pump.product_id == 4:
                pumpname = "g"
            if pump.product_id == 948:
                pumpname = "m"
        except:
            pass
        end = 0
        try:
            start = SellModel.objects.get(gs_id=int(id), tarikh=tarikh, tolombeinfo_id=int(val)).start
            end = SellModel.objects.get(gs_id=int(id), tarikh=tarikh2, tolombeinfo_id=int(val)).end


        except:
            start = 0
            end = 0

        try:
            crash = SellModel.objects.get(gs_id=int(id), tarikh=crashdate2, tolombeinfo_id=int(val))
            yarane = crash.yarane
            nimeyarane = crash.nimeyarane

            azad = crash.azad1
            ezterari = crash.ezterari
            azmayesh = crash.azmayesh
            havaleh = crash.haveleh
            sellkol = crash.sellkol
            mojaz = crash.mojaz
            ekhtelaf = crash.ekhtelaf
            nomojaz = crash.nomojaz
        except SellModel.DoesNotExist:
            yarane = 0
            nimeyarane = 0
            azad = 0
            ezterari = 0
            azmayesh = 0
            havaleh = 0
            sellkol = 0
            mojaz = 0
            ekhtelaf = 0
            nomojaz = 0
        except Exception:
            yarane = 0
            nimeyarane = 0
            azad = 0

            ezterari = 0
            azmayesh = 0
            havaleh = 0
            sellkol = 0
            mojaz = 0
            ekhtelaf = 0
            nomojaz = 0


        try:
            qrtime = QrTime.objects.get(selltime_id=_qrtime, tolombeinfo_id=int(val))
            yarane = qrtime.yarane
            nimeyarane = qrtime.nimeyarane
            azad = qrtime.azad1
            ezterari = qrtime.ezterari
            azmayesh = qrtime.azmayesh
        except Exception as e:
            pass

        dict = {
            'start': start,
            'end': end,
            'sell': end - start,
            'yarane': yarane,
            'nimeyarane': nimeyarane,
            'azad': azad,
            'ezterari': ezterari,
            'azmayesh': azmayesh,
            'havaleh': havaleh,
            'sellkol': sellkol,
            'mojaz': mojaz,
            'ekhtelaf': ekhtelaf,
            'nomojaz': nomojaz,
            'pumpname': pumpname,
        }

        mylist.append(dict)

    return JsonResponse({'message': 'success', 'mylist': mylist})


def checkEnd(request):
    if request.method == 'POST':
        gsid = request.POST.get('gsid')
        tarikh = request.POST.get('val')
        tarikh = tarikh.split("/")
        tarikh = jdatetime.date(day=int(tarikh[2]), month=int(tarikh[1]), year=int(tarikh[0])).togregorian()
        tarikh2 = tarikh - datetime.timedelta(days=1)
        issell = SellModel.objects.filter(gs_id=gsid, tarikh=tarikh).count()
        mek = SellModel.objects.filter(gs_id=gsid, tarikh=tarikh2).aggregate(sellmec=Sum('sell'),
                                                                             sellelec=Sum('sellkol'))
        ismek = True
        if mek['sellelec']:
            if int(mek['sellelec']) > 1 and int(mek['sellmec']) > 1:
                ghabl = int(mek['sellelec']) / 2
                if int(ghabl) > 1:
                    ismek = True if int(mek['sellmec']) > int(ghabl) else False

        pump = Pump.objects.filter(gs_id=gsid).order_by('number').first()
        onepump = pump.id
        isyesterday = True
        if SellModel.objects.filter(gs_id=gsid, tarikh=tarikh2).count() == 0:
            isyesterday = False
            _date = []
            end_date = CloseGS.objects.filter(gs__id=gsid,
                                              ).order_by('-id')
            for item in end_date:
                if tarikh2 >= item.date_in and tarikh2 <= item.date_out:

                    isyesterday = True
                    break
                else:
                    isyesterday = False
        if issell > 0:
            sell = True
        else:
            sell = False
        if pump.gs.isqrcode:
            sell = True
            ismek = True
        return JsonResponse(
            {'message': 'success', 'sell': sell, 'pump': onepump, 'ismek': ismek, 'isyesterday': isyesterday})


@cache_permission('reports2')
def reportsellmgr(request):
    add_to_log(request, " تهیه گزارش مدیریتی فروش ", 0)
    zone = Zone.objects_limit.all()
    sellmodel = None
    product = Product.objects.all()
    mdate = startdate
    mdate2 = today
    jam = 0
    jam1 = 0
    jam2 = 0
    jam3 = 0
    jam4 = 0
    jam5 = 0
    jam6 = 0
    jam7 = 0
    jam8 = 0
    sum_azadall = 0
    sum_ezterariall = 0
    sum_yaraneall = 0
    sum_sellkollall = 0
    sum_ekhtelaf = 0
    az = mdate
    ta = mdate2
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        zoneid = request.POST.getlist('zoneid')
        az = mdate
        ta = mdate2
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        far = int(request.POST.get('select4'))
        if request.user.owner.role.role == 'zone':
            sellmodel = SellGs.objects.filter(
                gs__area__zone_id=request.user.owner.zone_id
            ).select_related('gs__area__zone')

            if int(far) == 0:
                _far = [2, 3, 4, 948]
            else:
                _far = [far]

        if request.user.owner.role.role == 'area':
            sellmodel = SellGs.objects.filter(gs__area_id=request.user.owner.area_id)
            if int(far) == 0:
                _far = [2, 3, 4, 948]
            else:
                _far = [far]

        if request.user.owner.role.role in ['setad', 'mgr']:
            sellmodel = SellGs.objects.filter(gs__area__zone_id__in=zoneid) if zoneid else SellGs.objects.all()
            if zoneid:
                _far = [2, 3, 4, 948]
                far = 0
            else:
                _far = [far]

        mylist = []
        for faritem in _far:

            _list = sellmodel.values('gs__area__zone', 'gs__area__zone__name').filter(tarikh__gte=mdate,
                                                                                      tarikh__lte=mdate2,
                                                                                      product_id=faritem).annotate(
                res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
                sum_ekhtelaf=(Sum('azad') + Sum('ezterari') + Sum('yarane') + Sum('azmayesh')) - Sum('sell'),
                sum_sellkol=(Sum('azad') + Sum('ezterari') + Sum('yarane'))).order_by(
                'gs__area__zone')

            if _list.count() > 0:
                # messages.warning(request, 'برای دوره انتخابی شما فروشی ثبت نشده است')
                # return render(request, 'sellmgr.html',
                #               {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'product': product,
                #                'formpermmision': formpermmision, 'zone': zone})
                jam = _list.aggregate(sellall=Sum('res'))

                sum_azadall = _list.aggregate(sum_azadall=Sum('sum_azad'))
                sum_ekhtelaf = _list.aggregate(sum_ekhtelafall=Sum('sum_sellkol') - Sum('res'))
                sum_ezterariall = _list.aggregate(sum_ezterariall=Sum('sum_ezterari'))
                sum_yaraneall = _list.aggregate(sum_yaraneall=Sum('sum_yarane'))
                sum_sellkollall = _list.aggregate(sum_sellkollall=Sum('sum_sellkol'))
                jam1 = sum_ezterariall['sum_ezterariall'] + sum_yaraneall['sum_yaraneall'] + sum_azadall['sum_azadall']
                jam2 = sum_azadall['sum_azadall'] + sum_ezterariall['sum_ezterariall']
                jam3 = round((sum_azadall['sum_azadall'] / jam2) * 100) if sum_azadall['sum_azadall'] else 0
                jam4 = round((sum_ezterariall['sum_ezterariall'] / jam2) * 100) if sum_ezterariall[
                    'sum_ezterariall'] else 0
                jam5 = round((sum_yaraneall['sum_yaraneall'] / jam1) * 100) if sum_yaraneall['sum_yaraneall'] else 0
                jam6 = round((sum_azadall['sum_azadall'] / jam1) * 100) if sum_azadall['sum_azadall'] else 0
                jam7 = round((sum_ezterariall['sum_ezterariall'] / jam1) * 100) if sum_ezterariall[
                    'sum_ezterariall'] else 0
                jam8 = round(((sum_azadall['sum_azadall'] + sum_yaraneall['sum_yaraneall']) / jam1) * 100) if \
                    sum_azadall[
                        'sum_azadall'] and \
                    sum_yaraneall[
                        'sum_yaraneall'] else 0

                for item in _list:
                    if item['sum_azad'] == 0:
                        sum_azad = 0
                    else:
                        sum_azad = (item['sum_azad'] / (item['sum_azad'] + item['sum_ezterari'])) * 100

                    if item['sum_ezterari'] == 0:
                        sum_ezterari = 0
                    else:
                        sum_ezterari = (item['sum_ezterari'] / (item['sum_azad'] + item['sum_ezterari'])) * 100

                    if item['sum_yarane'] == 0:
                        d_n_1_to_kol = 0
                    else:
                        d_n_1_to_kol = (item['sum_yarane'] / (item['sum_yarane'] + item['sum_azad'] + item[
                            'sum_ezterari'])) * 100

                    if item['sum_azad'] == 0:
                        d_n_2_p_to_kol = 0
                    else:
                        d_n_2_p_to_kol = (item['sum_azad'] / (item['sum_yarane'] + item['sum_azad'] + item[
                            'sum_ezterari'])) * 100

                    if item['sum_ezterari'] == 0:
                        d_n_2_g_to_kol = 0
                    else:
                        d_n_2_g_to_kol = (item['sum_ezterari'] / (item['sum_yarane'] + item['sum_azad'] + item[
                            'sum_ezterari'])) * 100

                    if item['sum_ezterari'] == 0:
                        d_p_to_kol = 0
                    else:
                        d_p_to_kol = ((item['sum_yarane'] + item['sum_azad']) / (
                                item['sum_yarane'] + item['sum_azad'] + item[
                            'sum_ezterari'])) * 100

                    if item['sum_ezterari'] == 0 or d_p_to_kol == 100:
                        d_p_to_kol = 100
                    productid = Product.objects.get(id=faritem)
                    dict = {
                        'far': productid.name,
                        'farid': productid.id,
                        'zoneid': (item['gs__area__zone']),
                        'name': (item['gs__area__zone__name']),
                        'yarane': round(item['sum_yarane']),
                        'azad_personal': round(item['sum_azad']),
                        'azad_gs': round(item['sum_ezterari']),
                        'sum_sell': round(item['sum_yarane'] + item['sum_azad'] + item['sum_ezterari']),
                        'sum_nerkh2': round(item['sum_azad'] + item['sum_ezterari']),
                        'darsad_personal': round(sum_azad),
                        'darsad_gs': round(sum_ezterari),
                        'd_n_1_to_kol': round(d_n_1_to_kol),
                        'd_n_2_p_to_kol': round(d_n_2_p_to_kol),
                        'd_n_2_g_to_kol': round(d_n_2_g_to_kol),
                        'd_p_to_kol': round(d_p_to_kol),
                    }

                    mylist.append(dict)
        mylist = sorted(mylist, key=itemgetter('d_p_to_kol'), reverse=True)

        sum_yaraneall = round(sum_yaraneall['sum_yaraneall']) if sum_yaraneall else 0
        sum_azadall = round(sum_azadall['sum_azadall']) if sum_azadall else 0
        sum_ezterariall = round(sum_ezterariall['sum_ezterariall']) if sum_ezterariall else 0

        return TemplateResponse(request, 'sellmgr.html',
                                {'list': mylist, 'mdate': mdate, 'mdate2': mdate2, 'jam': jam,
                                 'sum_azadall': sum_azadall, 'sum_ezterariall': sum_ezterariall,
                                 'sum_yaraneall': sum_yaraneall,
                                 'sum_sellkollall': sum_sellkollall,
                                 'product': product, 'far': int(far), 'sum_ekhtelaf': sum_ekhtelaf, 'jam1': round(jam1),
                                 'jam2': round(jam2), 'jam3': round(jam3), 'jam4': round(jam4), 'jam5': round(jam5),
                                 'jam6': round(jam6), 'jam7': round(jam7), 'jam8': round(jam8),
                                 'az': az, 'ta': ta, 'zone': zone})
    return TemplateResponse(request, 'sellmgr.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'product': product,
                             'zone': zone})


@cache_permission('reports2')
def report_sell_mgr_nahye(request, mdate, mdate2, zoneid, far):
    add_to_log(request, " تهیه گزارش مدیریتی فروش ناحیه", 0)
    nname = Zone.objects_limit.get(id=zoneid)
    fname = Product.objects.get(id=far)
    if request.user.owner.role.role == 'zone':
        list = SellGs.objects.values('gs__area', 'gs__area__name').filter(tarikh__gte=mdate,
                                                                          tarikh__lte=mdate2,
                                                                          product_id=far,
                                                                          gs__area__zone_id=request.user.owner.zone_id).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_ekhtelaf=(Sum('azad') + Sum('ezterari') + Sum('yarane') + Sum('azmayesh')) - Sum('sell'),
            sum_sellkol=(Sum('azad') + Sum('ezterari') + Sum('yarane'))).order_by(
            'gs__area')

    if request.user.owner.role.role == 'area':
        list = SellModel.objects.values('gs__area', 'gs__area__name').filter(tarikh__gte=mdate,
                                                                             tarikh__lte=mdate2,
                                                                             product_id=far,
                                                                             gs__area_id=request.user.owner.area_id).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_ekhtelaf=Sum('sellkol') - Sum('sell'),
            sum_sellkol=Sum('sellkol')).order_by(
            'gs__area')

    if request.user.owner.role.role in ['setad', 'mgr']:
        list = SellModel.objects.values('gs__area', 'gs__area__name').filter(tarikh__gte=mdate,
                                                                             tarikh__lte=mdate2,
                                                                             product_id=far,
                                                                             gs__area__zone_id=zoneid).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_ekhtelaf=Sum('sellkol') - Sum('sell'),
            sum_sellkol=Sum('sellkol')).order_by(
            'gs__area')

    jam = list.aggregate(sellall=Sum('res'))
    sum_azadall = list.aggregate(sum_azadall=Sum('sum_azad'))
    sum_ekhtelaf = list.aggregate(sum_ekhtelafall=Sum('sum_sellkol') - Sum('res'))
    sum_ezterariall = list.aggregate(sum_ezterariall=Sum('sum_ezterari'))
    sum_yaraneall = list.aggregate(sum_yaraneall=Sum('sum_yarane'))
    sum_sellkollall = list.aggregate(sum_sellkollall=Sum('sum_sellkol'))
    if list.count() == 0:
        messages.warning(request, 'برای دوره انتخابی شما فروشی ثبت نشده است')
        return TemplateResponse(request, 'sellmgr.html',
                                {'mdate': mdate, 'mdate2': mdate2,
                                 })
    jam1 = sum_ezterariall['sum_ezterariall'] + sum_yaraneall['sum_yaraneall'] + sum_azadall['sum_azadall']
    jam2 = sum_azadall['sum_azadall'] + sum_ezterariall['sum_ezterariall']
    jam3 = (sum_azadall['sum_azadall'] / jam2) * 100
    jam4 = (sum_ezterariall['sum_ezterariall'] / jam2) * 100
    jam5 = (sum_yaraneall['sum_yaraneall'] / jam1) * 100
    jam6 = (sum_azadall['sum_azadall'] / jam1) * 100
    jam7 = (sum_ezterariall['sum_ezterariall'] / jam1) * 100
    jam8 = ((sum_azadall['sum_azadall'] + sum_yaraneall['sum_yaraneall']) / jam1) * 100
    sum_yaraneall = round(sum_yaraneall['sum_yaraneall'])
    sum_azadall = round(sum_azadall['sum_azadall'])
    sum_ezterariall = round(sum_ezterariall['sum_ezterariall'])
    mylist = []
    for item in list:
        if item['sum_azad'] == 0:
            sum_azad = 0
        else:
            sum_azad = (item['sum_azad'] / (item['sum_azad'] + item['sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0:
            sum_ezterari = 0
        else:
            sum_ezterari = (item['sum_ezterari'] / (item['sum_azad'] + item['sum_ezterari'])) * 100

        if item['sum_yarane'] == 0:
            d_n_1_to_kol = 0
        else:
            d_n_1_to_kol = (item['sum_yarane'] / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_azad'] == 0:
            d_n_2_p_to_kol = 0
        else:
            d_n_2_p_to_kol = (item['sum_azad'] / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0:
            d_n_2_g_to_kol = 0
        else:
            d_n_2_g_to_kol = (item['sum_ezterari'] / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0:
            d_p_to_kol = 0
        else:
            d_p_to_kol = ((item['sum_yarane'] + item['sum_azad']) / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0 or d_p_to_kol == 100:
            d_p_to_kol = 100

        dict = {
            'areaid': item['gs__area'],
            'name': item['gs__area__name'],
            'yarane': round(item['sum_yarane']),
            'azad_personal': round(item['sum_azad']),
            'azad_gs': round(item['sum_ezterari']),
            'sum_sell': round(item['sum_yarane'] + item['sum_azad'] + item['sum_ezterari']),
            'sum_nerkh2': round(item['sum_azad'] + item['sum_ezterari']),
            'darsad_personal': round(sum_azad),
            'darsad_gs': round(sum_ezterari),
            'd_n_1_to_kol': round(d_n_1_to_kol),
            'd_n_2_p_to_kol': round(d_n_2_p_to_kol),
            'd_n_2_g_to_kol': round(d_n_2_g_to_kol),
            'd_p_to_kol': round(d_p_to_kol),
        }
        mylist.append(dict)

        newmdate = mdate.replace('-', '/')
        newmdate2 = mdate2.replace('-', '/')
    return TemplateResponse(request, 'sellmgrnahye.html',
                            {'list': mylist, 'mdate': mdate, 'mdate2': mdate2, 'jam': jam, 'nname': nname,
                             'newmdate': newmdate,
                             'newmdate2': newmdate2,
                             'sum_azadall': sum_azadall, 'sum_ezterariall': sum_ezterariall,
                             'sum_yaraneall': sum_yaraneall,
                             'sum_sellkollall': sum_sellkollall, 'jam1': round(jam1), 'jam2': round(jam2),
                             'jam3': round(jam3), 'jam4': round(jam4), 'jam5': round(jam5), 'jam6': round(jam6),
                             'jam7': round(jam7), 'jam8': round(jam8),
                             'far': far, 'sum_ekhtelaf': sum_ekhtelaf, 'fname': fname,
                             })


@cache_permission('reports2')
def report_sell_mgr_gs(request, mdate, mdate2, areaid, far):
    add_to_log(request, " تهیه گزارش مدیریتی فروش جایگاه ", 0)
    nname = Area.objects.get(id=areaid)
    fname = Product.objects.get(id=far)
    if request.user.owner.role.role == 'zone':
        list = SellModel.objects.values('gs_id', 'gs__name').filter(tarikh__gte=mdate,
                                                                    tarikh__lte=mdate2,
                                                                    product_id=far,
                                                                    gs__area_id=areaid,
                                                                    gs__area__zone_id=request.user.owner.zone_id).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_ekhtelaf=Sum('sellkol') - Sum('sell'),
            sum_sellkol=Sum('sellkol') - Sum('azmayesh')).order_by(
            'gs')

    if request.user.owner.role.role == 'area':
        list = SellModel.objects.values('gs_id', 'gs__name').filter(tarikh__gte=mdate,
                                                                    tarikh__lte=mdate2,
                                                                    product_id=far,
                                                                    gs__area_id=request.user.owner.area_id).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_ekhtelaf=Sum('sellkol') - Sum('sell'),
            sum_sellkol=Sum('sellkol') - Sum('azmayesh')).order_by(
            'gs')

    if request.user.owner.role.role in ['setad', 'mgr']:
        list = SellModel.objects.values('gs_id', 'gs__name').filter(tarikh__gte=mdate,
                                                                    tarikh__lte=mdate2,
                                                                    product_id=far,
                                                                    gs__area_id=areaid).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            sum_ekhtelaf=Sum('sellkol') - Sum('sell'),
            sum_sellkol=Sum('sellkol') - Sum('azmayesh')).order_by(
            'gs')

    jam = list.aggregate(sellall=Sum('res'))
    sum_azadall = list.aggregate(sum_azadall=Sum('sum_azad'))
    sum_ekhtelaf = list.aggregate(sum_ekhtelafall=Sum('sum_sellkol') - Sum('res'))
    sum_ezterariall = list.aggregate(sum_ezterariall=Sum('sum_ezterari'))
    sum_yaraneall = list.aggregate(sum_yaraneall=Sum('sum_yarane'))
    sum_sellkollall = list.aggregate(sum_sellkollall=Sum('sum_sellkol'))
    if list.count() == 0:
        messages.warning(request, 'برای دوره انتخابی شما فروشی ثبت نشده است')
        return TemplateResponse(request, 'sellmgr.html',
                                {'mdate': mdate, 'mdate2': mdate2,
                                 })
    jam1 = sum_ezterariall['sum_ezterariall'] + sum_yaraneall['sum_yaraneall'] + sum_azadall['sum_azadall']
    jam2 = sum_azadall['sum_azadall'] + sum_ezterariall['sum_ezterariall']
    sum_yaraneall = round(sum_yaraneall['sum_yaraneall'])
    sum_azadall = round(sum_azadall['sum_azadall'])
    sum_ezterariall = round(sum_ezterariall['sum_ezterariall'])
    if jam2 == 0:
        jam3 = 0
    else:
        jam3 = (sum_azadall / jam2) * 100
    if jam2 == 0:
        jam4 = 0
    else:
        jam4 = (sum_ezterariall / jam2) * 100
    if jam1 == 0:
        jam5 = 0
    else:
        jam5 = (sum_yaraneall / jam1) * 100
    if jam1 == 0:
        jam6 = 0
    else:
        jam6 = (sum_azadall / jam1) * 100
    if jam1 == 0:
        jam7 = 0
    else:
        jam7 = (sum_ezterariall / jam1) * 100
    if jam1 == 0:
        jam8 = 0
    else:
        jam8 = ((sum_azadall + sum_yaraneall) / jam1) * 100
    mylist = []
    for item in list:
        if item['sum_azad'] == 0:
            sum_azad = 0
        else:
            sum_azad = (item['sum_azad'] / (item['sum_azad'] + item['sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0:
            sum_ezterari = 0
        else:
            sum_ezterari = (item['sum_ezterari'] / (item['sum_azad'] + item['sum_ezterari'])) * 100

        if item['sum_yarane'] == 0:
            d_n_1_to_kol = 0
        else:
            d_n_1_to_kol = (item['sum_yarane'] / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_azad'] == 0:
            d_n_2_p_to_kol = 0
        else:
            d_n_2_p_to_kol = (item['sum_azad'] / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0:
            d_n_2_g_to_kol = 0
        else:
            d_n_2_g_to_kol = (item['sum_ezterari'] / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0:
            d_p_to_kol = 0
        else:
            d_p_to_kol = ((item['sum_yarane'] + item['sum_azad']) / (item['sum_yarane'] + item['sum_azad'] + item[
                'sum_ezterari'])) * 100

        if item['sum_ezterari'] == 0 or d_p_to_kol == 100:
            d_p_to_kol = 100
        dict = {
            'areaid': item['gs_id'],
            'name': item['gs__name'],
            'yarane': round(item['sum_yarane']),
            'azad_personal': round(item['sum_azad']),
            'azad_gs': round(item['sum_ezterari']),
            'sum_sell': round(item['sum_yarane'] + item['sum_azad'] + item['sum_ezterari']),
            'sum_nerkh2': round(item['sum_azad'] + item['sum_ezterari']),
            'darsad_personal': round(sum_azad),
            'darsad_gs': round(sum_ezterari),
            'd_n_1_to_kol': round(d_n_1_to_kol),
            'd_n_2_p_to_kol': round(d_n_2_p_to_kol),
            'd_n_2_g_to_kol': round(d_n_2_g_to_kol),
            'd_p_to_kol': round(d_p_to_kol),
        }
        mylist.append(dict)

    newmdate = mdate.replace('-', '/')
    newmdate2 = mdate2.replace('-', '/')
    return TemplateResponse(request, 'sellmgrgs.html',
                            {'list': mylist, 'mdate': mdate, 'mdate2': mdate2, 'jam': jam, 'newmdate': newmdate,
                             'newmdate2': newmdate2,
                             'sum_azadall': sum_azadall, 'sum_ezterariall': sum_ezterariall,
                             'sum_yaraneall': sum_yaraneall,
                             'sum_sellkollall': sum_sellkollall, 'fname': fname,
                             'far': far, 'sum_ekhtelaf': sum_ekhtelaf, 'nname': nname, 'jam1': round(jam1),
                             'jam2': round(jam2), 'jam3': round(jam3), 'jam4': round(jam4), 'jam5': round(jam5),
                             'jam6': round(jam6), 'jam7': round(jam7), 'jam8': round(jam8),
                             'zoneid': int(areaid)})


@cache_permission('reports2')
def report_dosenotexist(request):
    list = None
    tarikh = ''
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    datesell = None
    add_to_log(request, " تهیه گزارش  فروش های ثبت نشده ", 0)
    gsgroup = None
    az = jdatetime.date.today() - datetime.timedelta(days=1)
    az = str(az)
    az = az.replace("-", "/")
    if request.method == 'POST':
        datein = str(request.POST.get('select'))
        az = datein
        datein = datein.split("/")
        tarikh = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        datesell = SellModel.object_role.c_gs(request, 0).filter(tarikh=tarikh).values('gs_id').annotate(
            sellsum=Sum('sell'),
            kolsum=Sum('sellkol'))
        gsmodel = GsModel.object_role.c_gsmodel(request).filter(status_id=1)
        list = gsmodel.exclude(
            id__in=datesell.filter(sellsum__gte=0, kolsum__gte=0).values('gs_id'))

        gsgroup = list.values('area__zone_id', 'area__zone__name').annotate(tedad=Count('id'))

    return TemplateResponse(request, 'selldosenot.html',
                            {'list': list, 'gsgroup': gsgroup, 'az': az, 'datesell': datesell,
                             'tarikh': tarikh})


@cache_permission('reports2')
def report_dosenotexistgroup(request, zone, datein):
    list = None

    datesell = None
    add_to_log(request, " تهیه گزارش  فروش های ثبت نشده ", 0)

    tarikh = datein
    datesell = SellModel.objects.filter(tarikh=tarikh, gs__area__zone_id=zone).values('gs_id').annotate(
        sellsum=Sum('sell'),
        kolsum=Sum('sellkol'))
    gsmodel = GsModel.objects.filter(status_id=1, area__zone_id=zone)
    list = gsmodel.exclude(
        id__in=datesell.filter(sellsum__gte=0, kolsum__gte=0).values('gs_id'))

    return TemplateResponse(request, 'selldosnotgroup.html',
                            {'list': list, 'datesell': datesell})


@cache_permission('reports2')
def graph_sell(request):
    add_to_log(request, " تهیه گزارش نمودار فروش ", 0)
    az = ""
    ta = ""
    listzone = []
    list = None
    result = None
    listmonth = None
    zones = Zone.objects_limit.all()
    products = Product.objects.all()
    namefar = ""
    nameord = ""
    namekho = ""
    nametir = ""
    namemor = ""
    namesha = ""
    namemeh = ""
    nameaba = ""
    nameaza = ""
    namedey = ""
    namebah = ""
    nameesf = ""
    listmonth = []
    listbar = []
    far_n1 = 0
    far_n2 = 0
    far_n3 = 0
    far_sum = 0

    ord_n1 = 0
    ord_n2 = 0
    ord_n3 = 0
    ord_sum = 0

    kho_n1 = 0
    kho_n2 = 0
    kho_n3 = 0
    kho_sum = 0

    tir_n1 = 0
    tir_n2 = 0
    tir_n3 = 0
    tir_sum = 0

    mor_n1 = 0
    mor_n2 = 0
    mor_n3 = 0
    mor_sum = 0

    sha_n1 = 0
    sha_n2 = 0
    sha_n3 = 0
    sha_sum = 0

    meh_n1 = 0
    meh_n2 = 0
    meh_n3 = 0
    meh_sum = 0

    aba_n1 = 0
    aba_n2 = 0
    aba_n3 = 0
    aba_sum = 0

    aza_n1 = 0
    aza_n2 = 0
    aza_n3 = 0
    aza_sum = 0
    product_s = 0
    dey_n1 = 0
    dey_n2 = 0
    dey_n3 = 0
    dey_sum = 0

    bah_n1 = 0
    bah_n2 = 0
    bah_n3 = 0
    bah_sum = 0

    esf_n1 = 0
    esf_n2 = 0
    esf_n3 = 0
    esf_sum = 0
    if request.method == 'POST':
        datein = str(request.POST.get('select'))
        dateout = str(request.POST.get('select2'))
        az = datein
        ta = dateout
        zone = request.POST.getlist('zone')
        product_s = request.POST.get('products')

        datein = datein.split("/")
        dateout = dateout.split("/")
        tarikhin = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        tarikhout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        if request.user.owner.role.role not in ['setad', 'mgr', 'hoze']:
            zone = []
            zone.append(request.user.owner.zone_id)
        if zone:
            result = SellModel.objects.values('tarikh').filter(tarikh__gte=tarikhin, tarikh__lte=tarikhout,
                                                               product_id=int(product_s),
                                                               gs__area__zone__in=zone).annotate(n1=Sum('yarane'),
                                                                                                 n2=Sum('azad'),
                                                                                                 n3=Sum('ezterari'),
                                                                                                 sum=Sum(
                                                                                                     'yarane') + Sum(
                                                                                                     'azad') + Sum(
                                                                                                     'ezterari')).order_by(
                'tarikh')
        else:
            result = SellModel.objects.values('tarikh').filter(product_id=int(product_s),
                                                               tarikh__gte=tarikhin, tarikh__lte=tarikhout).annotate(
                n1=Sum('yarane'),
                n2=Sum('azad'),
                n3=Sum('ezterari'),
                sum=Sum('yarane') + Sum(
                    'azad') + Sum(
                    'ezterari')).order_by(
                'tarikh')
        list = []
        for sell in result:

            dict = {
                'name': str(sell['tarikh']),
                'n1': round((sell['n1'] / sell['sum']) * 100),
                'n2': round((sell['n2'] / sell['sum']) * 100),
                'n3': round((sell['n3'] / sell['sum']) * 100),
            }
            list.append(dict)

            for sell in result:
                month = str(sell['tarikh'])
                month = month.split('-')

                if month[1] == '01':
                    namefar = f'{month[0]}فروردین'
                    far_n1 += sell['n1']
                    far_n2 += sell['n2']
                    far_n3 += sell['n3']
                    far_sum += sell['sum']

                if month[1] == '02':
                    nameord = f'{month[0]}اردیبهشت'
                    ord_n1 += sell['n1']
                    ord_n2 += sell['n2']
                    ord_n3 += sell['n3']
                    ord_sum += sell['sum']

                if month[1] == '03':
                    namekho = f'{month[0]}خرداد'
                    kho_n1 += sell['n1']
                    kho_n2 += sell['n2']
                    kho_n3 += sell['n3']
                    kho_sum += sell['sum']

                if month[1] == '04':
                    nametir = f'{month[0]}تیر'
                    tir_n1 += sell['n1']
                    tir_n2 += sell['n2']
                    tir_n3 += sell['n3']
                    tir_sum += sell['sum']

                if month[1] == '05':
                    namemor = f'{month[0]}مرداد'
                    mor_n1 += sell['n1']
                    mor_n2 += sell['n2']
                    mor_n3 += sell['n3']
                    mor_sum += sell['sum']

                if month[1] == '06':
                    namesha = f'{month[0]}شهریور'
                    sha_n1 += sell['n1']
                    sha_n2 += sell['n2']
                    sha_n3 += sell['n3']
                    sha_sum += sell['sum']

                if month[1] == '07':
                    namemeh = f'{month[0]}مهر'
                    meh_n1 += sell['n1']
                    meh_n2 += sell['n2']
                    meh_n3 += sell['n3']
                    meh_sum += sell['sum']

                if month[1] == '08':
                    nameaba = f'{month[0]}آبان'
                    aba_n1 += sell['n1']
                    aba_n2 += sell['n2']
                    aba_n3 += sell['n3']
                    aba_sum += sell['sum']

                if month[1] == '09':
                    nameaza = f'{month[0]}آذر'
                    aza_n1 += sell['n1']
                    aza_n2 += sell['n2']
                    aza_n3 += sell['n3']
                    aza_sum += sell['sum']

                if month[1] == '10':
                    namedey = f'{month[0]}دی'
                    dey_n1 += sell['n1']
                    dey_n2 += sell['n2']
                    dey_n3 += sell['n3']
                    dey_sum += sell['sum']

                if month[1] == '11':
                    namebah = f'{month[0]}بهمن'
                    bah_n1 += sell['n1']
                    bah_n2 += sell['n2']
                    bah_n3 += sell['n3']
                    bah_sum += sell['sum']

                if month[1] == '12':
                    nameesf = f'{month[0]}اسفند'
                    esf_n1 += sell['n1']
                    esf_n2 += sell['n2']
                    esf_n3 += sell['n3']
                    esf_sum += sell['sum']

        try:
            dict = {
                'name': namefar,
                'n1': round((far_n1 / far_sum) * 100),
                'n2': round((far_n2 / far_sum) * 100),
                'n3': round((far_n3 / far_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass

        try:
            dict = {
                'name': nameord,
                'n1': round((ord_n1 / ord_sum) * 100),
                'n2': round((ord_n2 / ord_sum) * 100),
                'n3': round((ord_n3 / ord_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': namekho,
                'n1': round((kho_n1 / kho_sum) * 100),
                'n2': round((kho_n2 / kho_sum) * 100),
                'n3': round((kho_n3 / kho_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': nametir,
                'n1': round((tir_n1 / tir_sum) * 100),
                'n2': round((tir_n2 / tir_sum) * 100),
                'n3': round((tir_n3 / tir_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': namemor,
                'n1': round((mor_n1 / mor_sum) * 100),
                'n2': round((mor_n2 / mor_sum) * 100),
                'n3': round((mor_n3 / mor_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': namesha,
                'n1': round((sha_n1 / sha_sum) * 100),
                'n2': round((sha_n2 / sha_sum) * 100),
                'n3': round((sha_n3 / sha_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': namemeh,
                'n1': round((meh_n1 / meh_sum) * 100),
                'n2': round((meh_n2 / meh_sum) * 100),
                'n3': round((meh_n3 / meh_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': nameaba,
                'n1': round((aba_n1 / aba_sum) * 100),
                'n2': round((aba_n2 / aba_sum) * 100),
                'n3': round((aba_n3 / aba_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': nameaza,
                'n1': round((aza_n1 / aza_sum) * 100),
                'n2': round((aza_n2 / aza_sum) * 100),
                'n3': round((aza_n3 / aza_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': namedey,
                'n1': round((dey_n1 / dey_sum) * 100),
                'n2': round((dey_n2 / dey_sum) * 100),
                'n3': round((dey_n3 / dey_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': namebah,
                'n1': round((bah_n1 / bah_sum) * 100),
                'n2': round((bah_n2 / bah_sum) * 100),
                'n3': round((bah_n3 / bah_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        try:
            dict = {
                'name': nameesf,
                'n1': round((esf_n1 / esf_sum) * 100),
                'n2': round((esf_n2 / esf_sum) * 100),
                'n3': round((esf_n3 / esf_sum) * 100),
            }
            listmonth.append(dict)
        except:
            pass
        if zone:
            result = SellModel.objects.values('gs__area__zone_id', 'gs__area__zone__name').filter(tarikh__gte=tarikhin,
                                                                                                  product_id=int(
                                                                                                      product_s),
                                                                                                  tarikh__lte=tarikhout,
                                                                                                  gs__area__zone__in=zone).annotate(
                n1=Sum('yarane'),
                n2=Sum('azad'),
                n3=Sum('ezterari'),
                sum=Sum('yarane') + Sum(
                    'azad') + Sum(
                    'ezterari'))
        else:
            result = SellModel.objects.values('gs__area__zone_id', 'gs__area__zone__name').filter(tarikh__gte=tarikhin,
                                                                                                  product_id=int(
                                                                                                      product_s),
                                                                                                  tarikh__lte=tarikhout).annotate(
                n1=Sum('yarane'),
                n2=Sum('azad'),
                n3=Sum('ezterari'),
                sum=Sum('yarane') + Sum(
                    'azad') + Sum(
                    'ezterari'))
        listbar = []
        for sell in result:
            if sell['n1'] != 0 and sell['sum']:
                dict = {
                    'name': str(sell['gs__area__zone__name']),
                    'n1': round((sell['n1'] / sell['sum']) * 100),
                    'n2': round((sell['n2'] / sell['sum']) * 100),
                    'n3': round((sell['n3'] / sell['sum']) * 100),
                }
                listbar.append(dict)

        listbar = sorted(listbar, key=itemgetter('n1'), reverse=True)

        for z in zone:
            listzone.append(int(z))

    return TemplateResponse(request, 'graph/graph_sell.html',
                            {'result': result, 'list': list, 'listmonth': listmonth, 'zones': zones, 'st': '1',
                             'az': az,
                             'listbar': listbar, 'products': products,
                             'product_s': int(product_s),
                             'ta': ta, 'zone': listzone})


@cache_permission('reports2')
def sell_contradiction(request):
    add_to_log(request, f'مشاهده فرم لیست نازلهای دارای مغایرت ', 0)
    url = request.META.get('HTTP_REFERER')
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    _list = None
    gss = GsModel.object_role.c_gsmodel(request).all()
    if request.method == 'POST':
        try:
            gsid = int(request.POST.get('select3'))
            nazel = request.POST.get('nazel')
            darsad = int(request.POST.get('val'))
            datein = str(request.POST.get('select'))
            sorted = int(request.POST.get('sorted'))
        except Exception as e:
            messages.error(request, f'آیتم های تاریخ و میزان مغایرت باید تکمیل گردد. ')
            return redirect(url)
        if nazel is None or nazel == '':
            nazel = 0
        az = datein
        dateout = str(request.POST.get('select2'))
        ta = dateout
        datein = datein.split("/")
        dateout = dateout.split("/")
        tarikhin = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        tarikhout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        # if request.user.owner.role.role == 'zone':
        if nazel == 0 and gsid == 0:
            _list = SellModel.object_role.c_gs(request, 0).filter(tarikh__gte=tarikhin,
                                                                  tarikh__lte=tarikhout,
                                                                  nomojaz__gte=darsad).order_by('-tarikh',
                                                                                                'gs__area',
                                                                                                'gs')
        elif gsid != 0 and nazel == 0:
            _list = SellModel.object_role.c_gs(request, 0).filter(tarikh__gte=tarikhin, gs_id__exact=gsid,
                                                                  tarikh__lte=tarikhout,
                                                                  nomojaz__gte=darsad).order_by('-tarikh',
                                                                                                'gs__area',
                                                                                                'gs')
        elif gsid != 0 and nazel != 0:
            _list = SellModel.object_role.c_gs(request, 0).filter(tarikh__gte=tarikhin, gs_id=gsid,
                                                                  tolombeinfo__number=nazel,
                                                                  tarikh__lte=tarikhout,
                                                                  nomojaz__gte=darsad).order_by('-tarikh',
                                                                                                'gs__area',
                                                                                                'gs')
        if sorted == "1":
            _list = _list.order_by('gs_id', '-tarikh')
        elif sorted == "2":
            _list = _list.order_by('-tarikh', 'gs__area', 'gs')
        summer = _list.aggregate(mek=Sum('sell'), elec=Sum('sellkol'), ekhtelaf=Sum('ekhtelaf'), nomojaz=Sum('nomojaz'))
        if nazel == 0:
            nazel = ''
        context = {'list': _list, 'darsad': darsad, 'az': az, 'ta': ta, 'gsid': gsid,
                   'nazel': nazel, 'gss': gss, 'summer': summer}
        return TemplateResponse(request, 'sell_contradiction.html', context)
    return TemplateResponse(request, 'sell_contradiction.html', {'gss': gss})


def context_data():
    context = {
        'page_name': '',
        'page_title': 'Chat Room',
        'system_name': 'Employee ID with QR Code Generator',
        'topbar': True,
        'footer': True,
    }

    return context


def qrstart(request):
    Owner.objects.filter(id=request.user.owner.id).update(
        qrcode="",
        qrcode2=""
    )

    return redirect('sell:scan_qrcode')


def scan_qrcode(request):
    inputcode = ""
    context = context_data()
    datein = today
    datein = datein.split("/")
    tarikh = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
    tarikh2 = tarikh - datetime.timedelta(days=1)
    role = request.user_data.get('role_name')
    _zone_id = request.user_data.get('zone_id')
    _owner_id = request.user_data.get('owner_id')

    # user_key = f"qr_user_{_owner_id}"
    # if cache.get(user_key):
    #     messages.error(request, 'لطفاً ۱۵ ثانیه بین اسکن‌ها فاصله بگذارید')
    #     return redirect('base:home')
    #
    # # ذخیره محدودیت
    # cache.set(user_key, True, 15)  # ۱۵ ثانیه
    if role == 'gs':
        gslist = GsList.objects.filter(owner_id=_owner_id)
        for item in gslist:
            if item.gs.isqrcode == False:
                mek = SellModel.objects.select_related('gs').filter(gs_id=item.gs_id).order_by('-tarikh')[1:2]
                if mek and mek.count() > 0:
                    mek1 = mek.first()

                    # mek = SellModel.objects.select_related('gs').filter(gs_id=item.gs_id, tarikh=mek1.tarikh).aggregate(sellmec=Sum('sell'),
                    #                                                                                sellelec=Sum(
                    #                                                                                    'sellkol'))
                    # ismek = True

                    # if mek['sellelec']:
                    #     ghabl = int(mek['sellelec']) / 2

                    # if int(mek['sellmec']) > int(ghabl):
                    #     ismek = True

    if request.method == "POST":
        try:
            qrcode = request.POST.get("qrinput")
            a = load_code(qrcode, _owner_id)
            if a == 0:
                messages.error(request,
                               'از گزینه پشتیبان برای اسکن استفاده نکنید حتما از گزینه های زیر آن ( بسته تسویه حساب) استفاده کنید')
                Owner.objects.filter(id=_owner_id).update(
                    qrcode=""
                )
                return redirect('base:home')

            inputcode = qrcode.split(":")
            a = inputcode[0]
            b = inputcode[1]
            if a == b:
                add_to_log(request, 'اسکن qrcode', 0)
                messages.success(request, 'اسکن با موفقیت انجام شد')
                Owner.objects.filter(id=_owner_id).update(
                    qrcode=""
                )

                return redirect('base:home')
            else:
                messages.warning(request, 'لطفا کد بعدی را اسکن کنید')
        except IndexError:
            messages.warning(request,
                             'اسکن انجام نشد ،لطفا دوباره گزینه استعلام را بزنید و گوشی را بصورت مستقیم روبروی مانیتور و  رمزینه قرار دهید ، توجه کنید  اگر بیش از یک رمزینه تولید شد، حتما باید رمزینه ها به ترتیب اسکن بشوند')
            Owner.objects.filter(id=_owner_id).update(
                qrcode=""
            )
            add_to_log(request, f' {inputcode} مشکل ایندکس در انجام qrcode', 0)
    return render(request, 'qrcodescanner.html', context)


def acceptrpmqrcode(request, ticket, lat, long, failure):
    context = context_data()
    if request.method == "POST":
        qrcode = request.POST.get("qrinput")
        load_rpm_code(request.user.id, qrcode, request.user.owner.id, ticket, lat, long, failure)
        inputcode = qrcode.split(":")
        a = inputcode[0]
        b = inputcode[1]
        if a == b:

            messages.success(request, 'اسکن با موفقیت انجام شد تیکت با موفقیت بسته شد')
            owner = Owner.objects.get(id=request.user.owner.id)

            return redirect('base:home')
        else:
            messages.warning(request, 'لطفا کد بعدی را اسکن کنید')
    return render(request, 'qrcodescanner.html', context)


@cache_permission('reports2')
def naftgazazad(request):
    _today = date.today()
    si_ago_sell = _today.today() - datetime.timedelta(days=30)
    naftgaz_azad = SellModel.object_role.c_gs(request.user.owner.refrence_id, request.user.owner.role.role).values(
        'gs__area__zone__name').filter(
        tarikh__gte=si_ago_sell, product_id=4,
        tarikh__lte=si_ago_sell.today()).annotate(jam=Sum('ezterari')).order_by('-jam')

    return TemplateResponse(request, 'graph/naftgazazad.html', {'naftgaz_azad': naftgaz_azad})


@cache_permission('listsell')
def printsell2(request, date, id):
    today = str(jdatetime.date.today())
    today = today.replace("-", "/")

    if request.user.owner.role.role in ['mgr', 'setad']:
        general = GsModel.objects.get(id=id)
    else:
        try:
            if request.user.owner.role.role in ['zone', 'area']:
                general = GsModel.objects.get(id=id, area__zone_id=request.user.owner.zone_id)
            if request.user.owner.role.role in ['gs']:
                gnow = GsList.objects.get(gs_id=id, owner_id=request.user.owner.id)
                general = GsModel.objects.get(id=gnow.gs.id)
        except GsModel.DoesNotExist:
            messages.warning(request, 'شما به این اطلاعات دسترسی ندارید')
            return redirect('base:home')
    add_to_log(request, f'مشاهده فرم1502 با جزئیات  {general.gsid} {general.name}', 0)
    sell = SellModel.objects.filter(gs_id=id, tarikh=date).order_by('tolombeinfo__number')
    hdcrash = False
    try:
        if sell.last().iscrash:
            hdcrash = True
    except TypeError:
        hdcrash = False

    list = SellModel.objects.values('product__name').filter(gs_id=id, tarikh=date).annotate(
        res=Sum('sell'), sum_sellkol=Sum('sellkol') - Sum('azmayesh'), sum_azad=Sum('azad'),
        sum_ezterari=Sum('ezterari'),
        sum_yarane=Sum('yarane'),sum_nimeyarane=Sum('nimeyarane'),sum_azad1=Sum('azad1'),
        sum_azmayesh=Sum('azmayesh'), sum_haveleh=Sum('haveleh'),
        sum_mojaz=Sum('mojaz'), sum_nomojaz=(Sum('sellkol') - Sum('sell')) - Sum('mojaz'),
        sum_nomojaz2=((Sum('sellkol') - Sum('sell')) - (Sum('sellkol') - Sum('sell')) - (
                Sum('sellkol') - Sum('sell'))),
        sum_nomojaz3=(Sum('nomojaz')))
    _edits = EditSell.objects.filter(sell__gs__id=id, sell__tarikh=date)
    tlist = []
    try:
        mojodi = Mojodi.objects.get(gs_id=id, tarikh=date)
        tdic = {
            'benzin': mojodi.benzin,
            'super': mojodi.super,
            'gaz': mojodi.gaz,
        }
        tlist.append(tdic)
    except ObjectDoesNotExist:
        tlist = False

    return TemplateResponse(request, 'printsell2.html',
                            {'sell': sell, 'general': general, 'date': date, 'list': list, 'hdcrash': hdcrash,
                             'today': today,
                             'edits': _edits, 'tlist': tlist,
                             })


@cache_permission('listsell')
def selectsell(request):
    url = request.META.get('HTTP_REFERER')
    today = str(jdatetime.date.today())
    today = today.replace("-", "/")
    if request.user.owner.role.role == 'zone':
        gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    elif request.user.owner.role.role in ['setad', 'mgr']:
        gss = GsModel.objects.all()
    elif request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area_id)
    elif request.user.owner.role.role == 'gs':
        gss = GsList.objects.filter(owner_id=request.user.owner.id)

    az = startdate
    ta = today
    if request.method == 'POST':
        datein = request.POST.get('select')
        dateout = request.POST.get('select2')
        az = datein
        ta = dateout
        datein = datein.replace("/", '-')
        dateout = dateout.replace("/", '-')
        datein = datein.split("-")
        dateout = dateout.split("-")
        tarikh_az = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        tarikh_ta = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()

        d3 = (tarikh_ta - tarikh_az).days
        if d3 > 31:
            messages.warning(request, 'بازه تاریخ نباید بیش از 31 روز باشد')
            return redirect(url)
        gsid = request.POST.get('select3')
        gs = GsModel.objects.get(id=gsid)

        _list = SellModel.objects.values('tolombeinfo_id', 'tolombeinfo__number', 'product__name').filter(
            tarikh__gte=tarikh_az, tarikh__lte=tarikh_ta,
            gs_id=gsid).annotate(
            res=Sum('sell'), sum_azad=Sum('azad'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),
            avalvaght=Min('end'), akharvaght=Max(Case(When(tarikh=tarikh_ta, then='start'))),
            sum_ekhtelaf=Sum('sellkol') - Sum('sell'),
            sum_mojaz=Sum('mojaz'), sum_nomojaz=(Sum('sellkol') - Sum('sell')) - Sum('mojaz'),
            sum_nomojaz2=((Sum('sellkol') - Sum('sell')) - (Sum('sellkol') - Sum('sell')) - (
                    Sum('sellkol') - Sum('sell'))),
            sum_nomojaz3=Sum('nomojaz'),
            sum_azmayesh=Sum('azmayesh'), sum_haveleh=Sum('haveleh'),
            sum_sellkol=Sum('sellkol')).order_by('tolombeinfo__number')

        add_to_log(request, f'مشاهده فرم1502 با انتخابی ', 0)
        list2 = SellModel.objects.values('product__name').filter(gs_id=gsid, tarikh__gte=tarikh_az,
                                                                 tarikh__lte=tarikh_ta).annotate(
            res=Sum('sell'), sum_sellkol=Sum('sellkol'), sum_azad=Sum('azad'),sum_azad1=Sum('azad1'),
            sum_ezterari=Sum('ezterari'),
            sum_yarane=Sum('yarane'),
            sum_nimeyarane=Sum('nimeyarane'),
            sum_azmayesh=Sum('azmayesh'), sum_haveleh=Sum('haveleh'),
            sum_mojaz=Sum('mojaz'), sum_nomojaz=(Sum('sellkol') - Sum('sell')) - Sum('mojaz'),
            sum_nomojaz2=((Sum('sellkol') - Sum('sell')) - (Sum('sellkol') - Sum('sell')) - (
                    Sum('sellkol') - Sum('sell'))),
            sum_nomojaz3=((Sum('nomojaz'))))

        return TemplateResponse(request, 'printsell3.html',
                                {'gss': gss, 'sell': _list, 'list2': list2, 'general': gs, 'az': az, 'ta': ta,
                                 'today': today,
                                 })
    return TemplateResponse(request, 'selectedsell.html', {'gss': gss, 'az': az, 'ta': ta})


def sellgo():
    sell = SellModel.objects.filter(product_id__isnull=True)
    for item in sell:
        item.product_id = item.tolombeinfo.product_id
        try:
            item.save()
        except IntegrityError:
            print(item.id)
        except:
            print(item.id)

    # sell = (SellModel.objects.values('gs_id', 'tarikh').filter(create__gte='2024-03-15').annotate(sell=Sum('sell')))
    # for item in sell:
    #     SellGs.sell_get_or_create(gs=item['gs_id'], tarikh=item['tarikh'])
    # print('-----------------------END Step 1------------------')
    # for item in SellGs.objects.filter(tarikh__gte='2024-03-15'):
    #     if not SellModel.objects.filter(gs_id=item.gs.id,tarikh=item.tarikh).exists():
    #         item.delete()
    par = Parametrs.objects.all().first()
    par.happyday = '1404/01/08'
    par.save()
    print('-----------------------END-------------------------')
    return True


@cache_permission('reports2')
def sell_contradiction_area(request):
    add_to_log(request, f'مشاهده فرم لیست نازلهای دارای مغایرت ناحیه', 0)
    if request.method == 'POST':
        datein = str(request.POST.get('select'))
        az = datein
        dateout = str(request.POST.get('select2'))
        ta = dateout
        datein = datein.split("/")
        dateout = dateout.split("/")
        tarikhin = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        tarikhout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
        _role = request.user.owner.role.role
        _roleid = zoneorarea(request)
        list = SellModel.object_role.c_gs(request, 0).values('gs__area__name').filter(tarikh__gte=tarikhin,
                                                                                      tarikh__lte=tarikhout).annotate(
            total=Sum(Abs('nomojaz')))

        context = {'list': list, 'az': az, 'ta': ta}
        return TemplateResponse(request, 'sell_contradiction_area.html', context)
    return TemplateResponse(request, 'sell_contradiction_area.html', {})


def delete_othersell():
    sellmodel = SellModel.objects.values('tarikh', 'gs_id').annotate(tedad=Count('gs_id'))
    sellgs = SellGs.objects.values('tarikh', 'gs_id').annotate(tedad=Count('gs_id'))
    list = sellgs.exclude(
        gs_id__in=sellmodel.all().values('gs_id'), tarikh__in=sellmodel.all().values('tarikh'))
    for i in list:
        print(i['gs_id'])
        print(i['tarikh'])

    return list


@cache_permission('acclosegs')
def accessclosegs(request):
    add_to_log(request, ' مشاهده فرم مجوز فروش ', 0)
    _list = AccessChangeSell.objects.filter(gs__area__zone_id=request.user.owner.zone_id).order_by('-id')
    gsmodels = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    owners = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role__in=['zone', 'area'])
    return TemplateResponse(request, 'acclosegs.html',
                            {'list': _list, 'gsmodels': gsmodels, 'owners': owners})


def deactiveaccessclosegs(request, id):
    url = request.META.get('HTTP_REFERER')
    ac = AccessChangeSell.objects.get(id=id)
    ac.active = False
    ac.save()
    add_to_log(request, ' حذف مجوز فروش ' + str(id), 0)
    messages.info(request, 'مجوز ویرایش فروش حذف شد')
    return redirect(url)


@cache_permission('sellday')
def checksellapi(request):
    if request.user.owner.role.role == 'gs':
        GsList.objects.filter(owner_id=request.user.owner.id)
        gss = GsModel.objects.filter(gsowner__owner=request.user.owner)
    if request.user.owner.role.role == 'zone':
        gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    if request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area_id)
    if request.user.owner.role.role in ['mgr', 'setad']:
        gss = GsModel.objects.all()

    product = Product.objects.all()
    ok = False

    return TemplateResponse(request, 'testconnection.html',
                            {'gss': gss, 'product': product, 'msg': '', 'ok': ok})


@cache_permission('mojodi')
def mojodi(request):
    add_to_log(request, f'مشاهده گزارش موجودی مخزن ', 0)
    mdate = str(jdatetime.date.today() - datetime.timedelta(days=1))
    az = mdate.replace('-', '/')

    _mojodi = ""
    zones = None
    sort_field = request.GET.get('sort', 'gs__name')  # فیلد پیش‌فرض برای مرتب‌سازی
    sort_order = request.GET.get('order', 'asc')  # ترتیب پیش‌فرض
    if request.user.owner.role.role in ['zone', 'area']:
        zones = Zone.objects.filter(id=request.user.owner.zone_id)

    if request.user.owner.role.role in ['mgr', 'setad']:
        zones = Zone.objects.all()
    dateok = 1
    if request.method == 'POST':
        mdate = request.POST.get('select')
        if mdate != az:
            dateok = 0
        az = mdate
        mdate = mdate.replace("/", '-')
        zoneid = request.POST.get('select3')

        if request.user.owner.role.role == 'zone':
            _mojodi = Mojodi.objects.filter(gs__area__zone_id=request.user.owner.zone_id, tarikh=mdate)

        if request.user.owner.role.role == 'area':
            _mojodi = Mojodi.objects.filter(gs__area_id=request.user.owner.area_id, tarikh=mdate)

        if request.user.owner.role.role == 'gs':
            _mojodi = Mojodi.objects.filter(gs__gsowner__owner_id=request.user.owner.id, tarikh=mdate)

        if request.user.owner.role.role in ['setad', 'mgr']:
            if zoneid == '0':
                _mojodi = Mojodi.objects.filter(tarikh=mdate)
            else:
                _mojodi = Mojodi.objects.filter(gs__area__zone_id=zoneid, tarikh=mdate)
    else:
        zoneid = 1
        if request.user.owner.role.role == 'zone':
            _mojodi = Mojodi.objects.filter(gs__area__zone_id=request.user.owner.zone_id, tarikh=mdate)
            zoneid = request.user.owner.zone_id

        if request.user.owner.role.role == 'area':
            _mojodi = Mojodi.objects.filter(gs__area_id=request.user.owner.area_id, tarikh=mdate)
            zoneid = request.user.owner.zone_id
        if request.user.owner.role.role == 'gs':
            _mojodi = Mojodi.objects.filter(gs__gsowner__owner_id=request.user.owner.id, tarikh=mdate)
            zoneid = 1

    zoneid = int(zoneid) if zoneid else 0
    if sort_field and _mojodi:
        if sort_order == 'desc':
            sort_field = f'-{sort_field}'
        _mojodi = _mojodi.order_by(sort_field)
    context = {
        'mojodi': _mojodi,
        'mdate': mdate,
        'zones': zones,
        'zoneid': zoneid,
        'az': az,
        'dateok': dateok,
        'current_sort': sort_field.replace('-', ''),
        'current_order': sort_order
    }
    return TemplateResponse(request, 'mojodi.html',
                            context)


@cache_permission('listsell')
def errorpage(request):
    return TemplateResponse(request, 'errorpage.html', {})


@cache_permission('mojodi')
def mojodi2(request):
    add_to_log(request, f'مشاهده گزارش موجودی مخزن و فروش ', 0)
    mdate = str(jdatetime.date.today() - datetime.timedelta(days=1))
    az = mdate.replace('-', '/')
    id = 1
    _mojodi = ""
    zoneid = 0
    mojodi = None
    zones = None
    if request.user.owner.role.role in ['zone', 'area']:
        zones = Zone.objects.filter(id=request.user.owner.zone_id)

    if request.user.owner.role.role in ['mgr', 'setad']:
        zones = Zone.objects.all()
    dateok = 1
    if request.method == 'POST':
        mdate = request.POST.get('select')
        id = int(request.POST.get('sorted'))
        if mdate != az:
            dateok = 0
        az = mdate
        mdate = mdate.replace("/", '-')
        zoneid = request.POST.get('select3')

        if request.user.owner.role.role == 'zone':
            _gsmodels = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id, status_id=1)

        if request.user.owner.role.role == 'area':
            _gsmodels = GsModel.objects.filter(area_id=request.user.owner.area_id, status_id=1)

        if request.user.owner.role.role == 'gs':
            _gsmodels = GsModel.objects.filter(gsowner__owner_id=request.user.owner.id, status_id=1)

        if request.user.owner.role.role in ['setad', 'mgr']:
            _gsmodels = GsModel.objects.filter(area__zone_id=zoneid, status_id=1)
        mojodi = []
        for gsmodel in _gsmodels:
            try:
                _mojodi = Mojodi.objects.get(gs_id=gsmodel.id, tarikh=mdate)
                m_benzin = _mojodi.benzin
                darsadbenzin = _mojodi.darsadbenzin
                m_super = _mojodi.super
                darsadsuper = _mojodi.darsadsuper
                m_gaz = _mojodi.gaz
                darsadgaz = _mojodi.darsadgaz
                sum_benzin = _mojodi.benzin
                sum_super = _mojodi.super
                sum_gaz = _mojodi.gaz
            except Mojodi.DoesNotExist:
                darsadbenzin = 0
                darsadsuper = 0
                darsadgaz = 0
                sum_benzin = 0
                sum_super = 0
                sum_gaz = 0
                m_benzin = "وارد نشده"
                m_super = "وارد نشده"
                m_gaz = "وارد نشده"
            _sell = SellGs.objects.filter(gs_id=gsmodel.id, tarikh=mdate).aggregate(
                benzin=Sum(Case(When(product_id=2, then='sell'))),
                super=Sum(Case(When(product_id=3, then='sell'))),
                gaz=Sum(Case(When(product_id=4, then='sell'))),
            )
            if _sell:
                s_benzin = _sell['benzin']
                s_super = _sell['super']
                s_gaz = _sell['gaz']
            else:
                s_benzin = 'ثبت نشده'
                s_super = 'ثبت نشده'
                s_gaz = 'ثبت نشده'

            dict = {
                'tarikh': mdate,
                'name': gsmodel.name,
                'gsid': gsmodel.gsid,
                'area': gsmodel.area,
                'zone': gsmodel.area.zone,
                'm_benzin': m_benzin,
                'm_super': m_super,
                'm_gaz': m_gaz,
                's_benzin': s_benzin,
                's_super': s_super,
                's_gaz': s_gaz,
                'darsadbenzin': darsadbenzin,
                'darsadsuper': darsadsuper,
                'darsadgaz': darsadgaz,
                'z_benzin': gsmodel.m_benzin,
                'z_super': gsmodel.m_super,
                'z_gaz': gsmodel.m_naftgaz,
                'sum_benzin': sum_benzin,
                'sum_super': sum_super,
                'sum_gaz': sum_gaz,
            }
            mojodi.append(dict)

        if id == 1:
            mojodi = sorted(mojodi, key=itemgetter('sum_benzin'), reverse=True)
        if id == 2:
            mojodi = sorted(mojodi, key=itemgetter('sum_super'), reverse=True)
        if id == 3:
            mojodi = sorted(mojodi, key=itemgetter('sum_gaz'), reverse=True)
    zoneid = int(zoneid) if zoneid else 0
    return TemplateResponse(request, 'mojodi2.html',
                            {'mojodi': mojodi, 'mdate': mdate, 'zones': zones, 'zoneid': int(zoneid), 'id': id,
                             'az': az, 'dateok': dateok})


@cache_permission('sellday')
def statistics_sell(request, date, id):
    gsmodel = GsModel.objects.get(id=id)
    add_to_log(request, f' مشاهده گزارش تجزیه فروش جایگاه  {gsmodel.gsid} {gsmodel.name} ', 0)
    product = []
    for gs in gsmodel.product.all():
        product.append(gs.id)

    date = date.split("-")
    oneyearold = int(date[0]) - 1
    if int(date[1]) > 1:
        onemonthold = int(date[1]) - 1
        oneyearnow = int(date[0])
    else:
        onemonthold = 12
        oneyearnow = int(date[0]) - 1

    endate = jdatetime.date(day=int(date[2]), month=int(date[1]), year=int(date[0])).togregorian()
    yesterday_date = endate - datetime.timedelta(days=1)
    week_date = endate - datetime.timedelta(days=7)
    month_date = jdatetime.date(day=int(date[2]), month=int(onemonthold), year=int(oneyearnow)).togregorian()
    year_date = jdatetime.date(day=int(date[2]), month=int(date[1]), year=int(oneyearold)).togregorian()

    sell_benzin = SellGs.objects.filter(gs_id=id, tarikh=endate, product_id=2).last()
    yesterday_sell_benzin = SellGs.objects.filter(gs_id=id, tarikh=yesterday_date, product_id=2).last()
    week_sell_benzin = SellGs.objects.filter(gs_id=id, tarikh=week_date, product_id=2).last()
    month_sell_benzin = SellGs.objects.filter(gs_id=id, tarikh=month_date, product_id=2).last()
    year_sell_benzin = SellGs.objects.filter(gs_id=id, tarikh=year_date, product_id=2).last()

    sell_super = SellGs.objects.filter(gs_id=id, tarikh=endate, product_id=3).last()
    yesterday_sell_super = SellGs.objects.filter(gs_id=id, tarikh=yesterday_date, product_id=3).last()
    week_sell_super = SellGs.objects.filter(gs_id=id, tarikh=week_date, product_id=3).last()
    month_sell_super = SellGs.objects.filter(gs_id=id, tarikh=month_date, product_id=3).last()
    year_sell_super = SellGs.objects.filter(gs_id=id, tarikh=year_date, product_id=3).last()

    sell_gaz = SellGs.objects.filter(gs_id=id, tarikh=endate, product_id=4).last()
    yesterday_sell_gaz = SellGs.objects.filter(gs_id=id, tarikh=yesterday_date, product_id=4).last()
    week_sell_gaz = SellGs.objects.filter(gs_id=id, tarikh=week_date, product_id=4).last()
    month_sell_gaz = SellGs.objects.filter(gs_id=id, tarikh=month_date, product_id=4).last()
    year_sell_gaz = SellGs.objects.filter(gs_id=id, tarikh=year_date, product_id=4).last()

    cars = CarInfo.objects.filter(gs_id=id, tarikh=endate)

    context = {'sell_benzin': sell_benzin, 'yesterday_sell_benzin': yesterday_sell_benzin,
               'week_sell_benzin': week_sell_benzin, 'month_sell_benzin': month_sell_benzin,
               'year_sell_benzin': year_sell_benzin,
               'sell_super': sell_super, 'yesterday_sell_super': yesterday_sell_super,
               'week_sell_super': week_sell_super, 'month_sell_super': month_sell_super,
               'year_sell_super': year_sell_super,
               'sell_gaz': sell_gaz, 'yesterday_sell_gaz': yesterday_sell_gaz,
               'week_sell_gaz': week_sell_gaz, 'month_sell_gaz': month_sell_gaz,
               'year_sell_gaz': year_sell_gaz, 'date': date[0] + "/" + date[1] + "/" + date[2],
               'cars': cars, 'gsmodel': gsmodel, 'product': product,
               }
    return TemplateResponse(request, 'statistics_sell.html', context)


@cache_permission('sellday')
def analysis_sell(request):
    add_to_log(request, f' مشاهده گزارش تجزیه فروش کل  ', 0)
    sell_benzin = ""
    month_sell_benzin = ""
    year_sell_benzin = ""
    sell_super = ""
    month_sell_super = ""
    year_sell_super = ""
    sell_gaz = ""
    month_sell_gaz = ""
    year_sell_gaz = ""
    sum_sell_benzin = 0
    sum_month_sell_benzin = 0
    sum_year_sell_benzin = 0
    sum_sell_gaz = 0
    sum_month_sell_gaz = 0
    sum_year_sell_gaz = 0
    month_date_in = ""
    month_date_out = ""
    year_date_in = ""
    year_date_out = ""
    d1 = ""
    d2 = ""

    cars = None
    if request.user.owner.role.role in ['mgr', 'setad']:
        zone = Zone.objects_limit.all()
    elif request.user.owner.role.role in ['zone']:
        _zone = request.user.owner.zone.id
        zone = Zone.objects_limit.filter(id=request.user.owner.zone.id)
    elif request.user.owner.role.role in ['area']:
        _zone = request.user.owner.zone.id
        _area = request.user.owner.area.id
        zone = Zone.objects_limit.filter(id=request.user.owner.zone.id)

    _zonename = "همه"
    _areaname = "همه"
    _gsname = "همه"
    _zone = 0 if request.user.owner.role.role in ['mgr', 'setad'] else request.user.owner.zone_id
    if request.user.owner.role.role == 'area':
        _area = request.user.owner.area_id
    _area = None
    _gs = None

    if request.method == 'POST':

        _zone = request.POST.get('Master')
        _area = request.POST.get('area')
        _gs = request.POST.get('gs')
        _isdore = request.POST.get('isdore')
        _datein = request.POST.get('select')
        _dateout = request.POST.get('select2')

        if request.user.owner.role.role in ['zone']:
            _zone = request.user.owner.zone.id
        elif request.user.owner.role.role in ['area']:
            _area = request.user.owner.area.id

        if _zone is None:
            _zone = 0
        if _area is None:
            _area = 0
        if _gs is None:
            _gs = 0
        d1 = str(_datein)
        d2 = str(_dateout)

        _datein = to_miladi(_datein)
        _dateout = to_miladi(_dateout)
        d3 = (_dateout - _datein).days

        _date_in = d1.split("/")
        _date_out = d2.split("/")
        oneyearold_in = int(_date_in[0]) - 1

        oneyearold_out = int(_date_out[0]) - 1

        if int(_date_in[1]) > 1:
            onemonthold_in = int(_date_in[1]) - 1
            oneyearnow_in = int(_date_in[0])
        else:
            onemonthold_in = 12
            oneyearnow_in = int(_date_in[0]) - 1

        if int(_date_out[1]) > 1:
            onemonthold_out = int(_date_out[1]) - 1
            oneyearnow_out = int(_date_out[0])
        else:
            onemonthold_out = 12
            oneyearnow_out = int(_date_out[0]) - 1

        if _isdore == 'on':

            month_date_in = jdatetime.date(day=int(_date_in[2]), month=int(onemonthold_in),
                                           year=int(oneyearnow_in)).togregorian()
            month_date_out = jdatetime.date(day=int(_date_out[2]), month=int(onemonthold_out),
                                            year=int(oneyearnow_out)).togregorian()
        else:
            month_date_out = _datein - datetime.timedelta(days=1)
            month_date_in = _datein - datetime.timedelta(days=d3 + 1)

        # month_date_in = jdatetime.date(day=int(_date_in[2]), month=int(onemonthold_in),
        #                                year=int(oneyearnow_in)).togregorian()
        # month_date_out = jdatetime.date(day=int(_date_out[2]), month=int(onemonthold_out),
        #                                 year=int(oneyearnow_out)).togregorian()
        year_date_in = jdatetime.date(day=int(_date_in[2]), month=int(_date_in[1]),
                                      year=int(oneyearold_in)).togregorian()
        year_date_out = jdatetime.date(day=int(_date_out[2]), month=int(_date_out[1]),
                                       year=int(oneyearold_out)).togregorian()

        if _zone == '0':

            sellgs = SellGs.objects.all()
        else:

            sellgs = SellGs.objects.filter(gs__area__zone_id=_zone)
            _zonename = Zone.objects_limit.get(id=_zone).name
        if _area != '0':
            sellgs = sellgs.filter(gs__area_id=_area)
            _areaname = Area.objects.get(id=_area).name
        if _gs != '0':
            sellgs = sellgs.filter(gs_id=_gs)
            _gss = GsModel.objects.get(id=_gs)
            _gsname = _gss.gsid + " | " + _gss.name

        sell_benzin = sellgs.filter(tarikh__range=(_datein, _dateout), product_id=2).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))

        yarane = float(sell_benzin['yarane']) if sell_benzin['yarane'] else 0
        azad = float(sell_benzin['azad']) if sell_benzin['azad'] else 0
        ezterari = float(sell_benzin['ezterari']) if sell_benzin['ezterari'] else 0
        sum_sell_benzin = yarane + azad + ezterari

        month_sell_benzin = sellgs.filter(tarikh__range=(month_date_in, month_date_out),
                                          product_id=2).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))

        yarane = float(month_sell_benzin['yarane']) if month_sell_benzin['yarane'] else 0
        azad = float(month_sell_benzin['azad']) if month_sell_benzin['azad'] else 0
        ezterari = float(month_sell_benzin['ezterari']) if month_sell_benzin['ezterari'] else 0
        sum_month_sell_benzin = yarane + azad + ezterari

        year_sell_benzin = sellgs.filter(tarikh__range=(year_date_in, year_date_out),
                                         product_id=2).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))
        yarane = float(year_sell_benzin['yarane']) if year_sell_benzin['yarane'] else 0
        azad = float(year_sell_benzin['azad']) if year_sell_benzin['azad'] else 0
        ezterari = float(year_sell_benzin['ezterari']) if year_sell_benzin['ezterari'] else 0
        sum_year_sell_benzin = yarane + azad + ezterari

        sell_super = sellgs.filter(tarikh__range=(_datein, _dateout), product_id=3).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))

        month_sell_super = sellgs.filter(tarikh__range=(month_date_in, month_date_out),
                                         product_id=3).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))
        year_sell_super = sellgs.filter(tarikh__range=(year_date_in, year_date_out),
                                        product_id=3).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))

        sell_gaz = sellgs.filter(tarikh__range=(_datein, _dateout), product_id=4).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))
        if sell_gaz['yarane'] and sell_gaz['azad']:
            sum_sell_gaz = float(sell_gaz['yarane']) + float(sell_gaz['azad']) + float(
                sell_gaz['ezterari'])
        else:
            sum_sell_gaz = 0

        yarane = float(sell_gaz['yarane']) if sell_gaz['yarane'] else 0
        azad = float(sell_gaz['azad']) if sell_gaz['azad'] else 0
        ezterari = float(sell_gaz['ezterari']) if sell_gaz['ezterari'] else 0
        sum_sell_gaz = yarane + ezterari

        month_sell_gaz = sellgs.filter(tarikh__range=(month_date_in, month_date_out),
                                       product_id=4).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))

        yarane = float(month_sell_gaz['yarane']) if month_sell_gaz['yarane'] else 0
        azad = float(month_sell_gaz['azad']) if month_sell_gaz['azad'] else 0
        ezterari = float(month_sell_gaz['ezterari']) if month_sell_gaz['ezterari'] else 0
        sum_month_sell_gaz = yarane + ezterari

        year_sell_gaz = sellgs.filter(tarikh__range=(year_date_in, year_date_out),
                                      product_id=4).aggregate(
            yarane=Sum('yarane'), azad=Sum('azad'), ezterari=Sum('ezterari'))

        yarane = float(year_sell_gaz['yarane']) if year_sell_gaz['yarane'] else 0
        azad = float(year_sell_gaz['azad']) if year_sell_gaz['azad'] else 0
        ezterari = float(year_sell_gaz['ezterari']) if year_sell_gaz['ezterari'] else 0
        sum_year_sell_gaz = yarane + ezterari

        if _zone == '0':
            carinfo = CarInfo.objects.all()
        else:
            carinfo = CarInfo.objects.filter(gs__area__zone_id=_zone, tarikh__range=(_datein, _dateout))
        if _area != '0':
            carinfo = carinfo.filter(gs__area_id=_area, tarikh__range=(_datein, _dateout))
        if _gs != '0':
            carinfo = carinfo.filter(gs_id=_gs, tarikh__range=(_datein, _dateout))
        cars1 = carinfo.values('carstatus__name', 'carstatus_id').annotate(amount=Sum('amount'))
        cars = []
        for car in cars1:
            dict = {
                'id': car['carstatus_id'],
                'carstatus': car['carstatus__name'],
                'amount': car['amount']
            }
            cars.append(dict)

    context = {'sell_benzin': sell_benzin, 'month_sell_benzin': month_sell_benzin,
               'year_sell_benzin': year_sell_benzin, 'zone': zone, 'zoneid': _zone, 'areaid': _area,
               'gsid': _gs, 'sell_super': sell_super, 'month_sell_super': month_sell_super,
               'year_sell_super': year_sell_super, 'cars': cars, 'sell_gaz': sell_gaz, 'month_sell_gaz': month_sell_gaz,
               'year_sell_gaz': year_sell_gaz, "az": d1, "ta": d2,
               'sum_sell_benzin': round(sum_sell_benzin, 1), 'sum_month_sell_benzin': round(sum_month_sell_benzin, 1),
               'sum_year_sell_benzin': round(sum_year_sell_benzin, 1), 'sum_sell_gaz': round(sum_sell_gaz, 1),
               'sum_month_sell_gaz': round(sum_month_sell_gaz, 1), 'sum_year_sell_gaz': round(sum_year_sell_gaz, 1),
               'month_date_in': month_date_in, 'month_date_out': month_date_out,
               'year_date_in': year_date_in, 'year_date_out': year_date_out,
               'myzone': _zonename, 'myarea': _areaname, 'mygs': _gsname,
               }
    return TemplateResponse(request, 'analysis_sell.html', context)


@cache_permission('reports2')
def changeselllist(request):
    _list2 = None
    add_to_log(request, f' مشاهده گزارش فروشهای ویرایش شده  ', 0)
    zone = Zone.objects_limit.all()
    area = None
    if request.user.owner.role.role == 'zone':
        zone = Zone.objects.filter(id=request.user.owner.zone.id)
        area = Area.objects.filter(zone_id=request.user.owner.zone_id)
    if request.method == 'POST':
        _datein = request.POST.get('select')
        _dateout = request.POST.get('select2')
        zoneid = request.POST.get('zone')

        areaid = request.POST.get('area')
        if request.user.owner.role.role in ['zone']:
            zoneid = request.user.owner.zone.id
        elif request.user.owner.role.role in ['area']:
            areaid = request.user.owner.area.id
        area = Area.objects.filter(zone_id=zoneid)

        try:
            datein = to_miladi(_datein)
            dateout = to_miladi(_dateout)
        except:
            messages.error(request, 'تاریخ معتبر وارد کنید')
            return TemplateResponse(request, 'changeselllist.html', {'zone': zone, 'area': area})

        _list = EditSell.objects.filter(tarikh__range=(datein, dateout))
        _list2 = AcceptForBuy.objects.filter(tarikh__range=(datein, dateout))
        _list3 = AccessChangeSell.objects.filter(tarikh__range=(datein, dateout))
        if request.user.owner.role.role in ['setad', 'mgr']:
            _list = _list.filter(sell__gs__area__zone_id=zoneid)
            _list2 = _list2.filter(gs__area__zone_id=zoneid)
            _list3 = _list3.filter(gs__area__zone_id=zoneid)
        if request.user.owner.role.role in ['zone']:
            _list = _list.filter(sell__gs__area__zone_id=request.user.owner.zone.id)
            _list2 = _list2.filter(gs__area__zone_id=request.user.owner.zone.id)
            _list3 = _list3.filter(gs__area__zone_id=request.user.owner.zone.id)

            if areaid != "0":
                _list = _list.filter(sell__gs__area_id=areaid)
                _list2 = _list2.filter(gs__area_id=areaid)
                _list3 = _list3.filter(gs__area_id=areaid)
        if request.user.owner.role.role in ['area']:
            _list = _list.filter(sell__gs__area_id=request.user.owner.area.id)
            _list2 = _list2.filter(gs__area_id=request.user.owner.area.id)
            _list3 = _list3.filter(gs__area_id=request.user.owner.area.id)

        return TemplateResponse(request, 'changeselllist.html',
                                {'az': _datein, 'ta': _dateout, 'list': _list, 'list2': _list2, 'list3': _list3,
                                 'zone': zone, 'area': area, 'areaid': int(areaid), 'zoneid': int(zoneid)})
    return TemplateResponse(request, 'changeselllist.html', {'zone': zone, 'area': area})


@cache_permission('sellday')
def analizchart(request):
    cars = None
    result = None
    if request.method == 'POST':
        _zone = request.POST.get('newzone')
        _area = request.POST.get('newarea')
        _item = request.POST.get('newitem')

        az = request.POST.get('newaz')
        ta = request.POST.get('newta')

        _datein = to_miladi(az)
        _dateout = to_miladi(ta)

        result = CarStatus.objects.get(id=_item)
        if _zone == '0':
            carinfo = CarInfo.objects.filter(carstatus_id=_item)
        else:
            carinfo = CarInfo.objects.filter(gs__area__zone_id=_zone,
                                             tarikh__range=(_datein, _dateout),
                                             carstatus_id=_item)
        if _area != '0':
            carinfo = carinfo.filter(gs__area_id=_area, tarikh__range=(_datein, _dateout))

        cars1 = carinfo.values('gs__gsid', 'gs__name').annotate(amount=Sum('amount'))
        cars = []
        for car in cars1:
            _dict = {
                'gsid': str(car['gs__gsid']),
                'name': str(car['gs__name']),
                'amount': car['amount']
            }
            cars.append(_dict)
            cars = sorted(cars, key=itemgetter('amount'), reverse=True)
            cars = cars[:20]

    return TemplateResponse(request, 'analizchart.html', {'cars': cars, 'result': result})


@cache_permission('sendproduct')
def sendproduct(request):
    _list = []
    _list2 = []
    _list3 = []
    _time = jdatetime.datetime.now()
    _today = jdatetime.date.today()
    one_day_ago = _today.today() - jdatetime.timedelta(days=1)
    # gs = Mojodi.objects.filter(gs__area__zone_id=request.user.owner.zone.id)
    gs = GsModel.objects.filter(area__zone_id=request.user.owner.zone.id, status_id=1)

    for item in gs:

        try:
            _mojodi = Mojodi.objects.filter(gs_id=item.id).order_by('-tarikh').first()
            try:
                sending = Waybill.objects.filter(gsid_id=item.id).filter(
                    Q(receive_car_date__gt=_mojodi.create.date()) |  # تاریخ خروج بزرگتر از تاریخ create
                    Q(receive_car_date=_mojodi.create.date(), receive_car_time__gte=_mojodi.create.time())
                    # یا تاریخ برابر اما زمان بزرگتر یا مساوی
                ).aggregate(
                    benzin=Sum(Case(When(product_id__product_id=2, then='quantity'))),
                    super=Sum(Case(When(product_id__product_id=3, then='quantity'))),
                    gaz=Sum(Case(When(product_id__product_id=4, then='quantity'))),
                )

                # sending_super = int(sending['super'])
                sending_benzin = int(sending['benzin']) if sending['benzin'] else 0
                sending_gaz = int(sending['gaz']) if sending['gaz'] else 0

            except Exception as e:
                sending_benzin = 0
                sending_super = 0
                sending_gaz = 0

            _gs = ParametrGs.objects.get(gs_id=item.id, oildepot_id=request.user.owner.oildepot.id)

            diff = (datetime.datetime.now() - _mojodi.create)
            _diff = str(diff)

            if diff.days > 0:
                _diff = _diff.split(' ')
                _diff = _diff[2].split(':')
                _diff = int(_diff[0])
                _diff = (diff.days * 24) + _diff
            else:
                _diff = _diff.split(' ')
                _diff = _diff[0].split(':')
                _diff = int(_diff[0])

            i = 1 if _diff == 0 else _diff

            _now = (datetime.datetime.now() - _mojodi.create).seconds
            _now = _now / 60
            _now = _now / 60

            newmojodi = _mojodi.benzin + sending_benzin
            _sell = 0

            htime = _mojodi.create.hour
            _time = jdatetime.datetime.now()
            nowtime = _time.hour
            # i = nowtime - htime

            normalsellinhour = newhtimecpu(item.id, 2, i, htime, 0)
            sellinhorse = int(normalsellinhour['a'])
            summsellinhorse = int(normalsellinhour['b'])
            if _diff >= 1:
                _sell = summsellinhorse
                newmojodi = (_mojodi.benzin + sending_benzin) - _sell
            if summsellinhorse > 0:

                sumtotime = newhtimecpu(item.id, 2, 48, nowtime, newmojodi)

                _time = int(sumtotime['c']) * 60

                _drivetime = _time - (_gs.normaltime + _gs.oildepot.dalytime)
                _sortdrivetime = _drivetime
                _st = 'دقیقه'
                _st2 = 3
                if _drivetime > 59:
                    _drivetime = _drivetime / 60
                    _st = 'ساعت'
                    _st2 = 2
                if newmojodi < 0:
                    _drivetime = 0
                    _st = ' تعطیل '
                    _st2 = 4
                    newmojodi = 0
                info = f'میزان فروش از آخرین موجودی تا اکنون {_sell} لیتر و متوسط فروش جایگاه در هر ساعت {normalsellinhour}{sumtotime} لیتر و زمان رسیدن نفتکش به جایگاه {_gs.normaltime} دقیقه میباشد'
                dict = {
                    'htime': htime,
                    'hour': str(i),
                    'id': _mojodi.id,
                    'gsid': item.gsid,
                    'name': item.name,
                    'area': item.area.name,
                    'mojodi': _mojodi.benzin,
                    'mojodi2': newmojodi,
                    'time': _time,
                    'barname': sending_benzin,
                    'drivetime': int(_drivetime),
                    'sortdrivetime': int(_sortdrivetime),
                    'st': _st,
                    'st2': _st2,
                    'info': info,
                    'now': _diff,
                    'sell0': _sell,
                    'moj': _mojodi.create
                }
                _list.append(dict)
                _list = sorted(_list, key=itemgetter('sortdrivetime'), reverse=False)

            normalsellinhourgaz = newhtimecpu(item.id, 4, i, htime, 0)
            sellinhorse = int(normalsellinhourgaz['a'])
            summsellinhorse = int(normalsellinhourgaz['b'])
            if _mojodi.gaz > 0 and summsellinhorse > 0:
                newmojodi = _mojodi.gaz + sending_gaz
                _sell = 0
                if _diff >= 1:
                    _sell = summsellinhorse

                    newmojodi = (_mojodi.gaz + sending_gaz) - _sell
                sumtotime = newhtimecpu(item.id, 4, 48, nowtime, newmojodi)

                _time = int(sumtotime['c']) * 60
                # _time = (newmojodi / summsellinhorse) * 60
                _drivetime = _time - (_gs.normaltime + _gs.oildepot.dalytime)

                _sortdrivetime = _drivetime
                _st = 'دقیقه'
                _st2 = 3
                if _drivetime > 59:
                    _drivetime = _drivetime / 60
                    _st = 'ساعت'
                    _st2 = 2
                if newmojodi < 0:
                    _drivetime = 0
                    _st = ' تعطیل  '
                    _st2 = 4
                    newmojodi = 0
                info = f'میزان فروش از آخرین موجودی تا اکنون {_sell} لیتر و متوسط فروش جایگاه در هر ساعت {normalsellinhourgaz}{sumtotime} لیتر و زمان رسیدن نفتکش به جایگاه  {_gs.normaltime} دقیقه میباشد'
                dict = {
                    'id': _mojodi.id,
                    'gsid': item.gsid,
                    'name': item.name,
                    'area': item.area.name,
                    'mojodi': _mojodi.gaz,
                    'mojodi2': newmojodi,
                    'time': _time,
                    'barname': sending_gaz,
                    'drivetime': int(_drivetime),
                    'sortdrivetime': int(_sortdrivetime),
                    'st': _st,
                    'st2': _st2,
                    'info': info,
                    'now': _diff,
                    'sell0': _sell,
                    'moj': _mojodi.create

                }
                _list2.append(dict)
                _list2 = sorted(_list2, key=itemgetter('sortdrivetime'), reverse=False)
        except ParametrGs.DoesNotExist:
            messages.warning(request, str('dos not exist'))
    # except TypeError as e:
    #     messages.warning(request, str(e))
    # except IndexError as e:
    #     messages.warning(request, str(e))
    # except AttributeError as e:
    #     messages.warning(request, str(e))
    # except Exception as e:
    #     messages.warning(request, str(e))
    add_to_log(request, f' مشاهده فرم موجودی سیستمی', 0)
    return TemplateResponse(request, 'sendproduct.html', {'list': _list, 'list2': _list2})


@cache_permission('sendproduct')
def updatemojodi(request, _id, _po):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        _send = request.POST.get('litraj')
        mojodi = Mojodi.objects.get(id=_id)
        if _po == 2:
            mojodi.sending += int(_send)
        elif _po == 4:
            mojodi.sendinggaz += int(_send)
        mojodi.save()
        messages.success(request, 'عملیات با موفقیت انجام شد.')
    return redirect(url)


@cache_permission('sendproduct')
def import_excel_oms(request):
    _list = []
    form = open_excel(request.POST)
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)

        if form.is_valid():
            _gsid = str(request.FILES.get('filepath'))
            _gsid = _gsid[:4]
            if len(_gsid) != 4:
                return HttpResponse('gsid باید 4 رقم باشد')
            form.save()

            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath

            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            d = 0
            for i in range(1, m_row + 1):
                if d != 0:

                    _date = sheet_obj.cell(row=i, column=6).value
                    _amount = sheet_obj.cell(row=i, column=10).value
                    _po = sheet_obj.cell(row=i, column=8).value
                    _po = 2 if _po == 'بنزین معمولی' else 4

                    _date = _date.split(" ")
                    _date = _date[1].split(":")
                    _date = _date[0]
                    if int(_amount) > 0:
                        dict = {
                            'date': str(_date),
                            'amount': float(_amount)
                        }
                        _list.append(dict)
                d += 1

        _list = sorted(_list, key=itemgetter('date'), reverse=False)
        _0 = 0.0
        _1 = 0.0
        _2 = 0
        _3 = 0
        _4 = 0
        _5 = 0
        _6 = 0
        _7 = 0
        _8 = 0
        _9 = 0
        _10 = 0
        _11 = 0
        _12 = 0
        _13 = 0
        _14 = 0
        _15 = 0
        _16 = 0
        _17 = 0
        _18 = 0
        _19 = 0
        _20 = 0
        _21 = 0
        _22 = 0
        _23 = 0

        for list in _list:
            try:
                match list['date']:
                    case '00':
                        _0 += list['amount']
                    case '01':
                        _1 += list['amount']
                    case '02':
                        _2 += list['amount']
                    case '03':
                        _3 += list['amount']
                    case '04':
                        _4 += list['amount']
                    case '05':
                        _5 += list['amount']
                    case '06':
                        _6 += list['amount']
                    case '07':
                        _7 += list['amount']
                    case '08':
                        _8 += list['amount']
                    case '09':
                        _9 += list['amount']
                    case '10':
                        _10 += list['amount']
                    case '11':
                        _11 += list['amount']
                    case '12':
                        _12 += list['amount']
                    case '13':
                        _13 += list['amount']
                    case '14':
                        _14 += list['amount']
                    case '15':
                        _15 += list['amount']
                    case '16':
                        _16 += list['amount']
                    case '17':
                        _17 += list['amount']
                    case '18':
                        _18 += list['amount']
                    case '19':
                        _19 += list['amount']
                    case '20':
                        _20 += list['amount']
                    case '21':
                        _21 += list['amount']
                    case '22':
                        _22 += list['amount']
                    case '23':
                        _23 += list['amount']




            except:
                pass
        gs = GsModel.objects.get(gsid=_gsid)
        SellGsInHour.objects.filter(gs_id=gs.id, product_id=_po).delete()
        SellGsInHour.objects.create(gs_id=gs.id, product_id=_po, h1time=int(_1) / 10, h2time=_2 / 10, h3time=_3 / 10,
                                    h4time=_4 / 10,
                                    h5time=_5 / 10, h6time=_6 / 10
                                    , h7time=_7 / 10, h8time=_8 / 10, h9time=_9 / 10, h10time=_10 / 10,
                                    h11time=_11 / 10, h12time=_12 / 10,
                                    h13time=_13 / 10
                                    , h14time=_14 / 10, h15time=_15 / 10, h16time=_16 / 10, h17time=_17 / 10,
                                    h18time=_18 / 10, h19time=_19 / 10,
                                    h20time=_20 / 10
                                    , h21time=_21 / 10, h22time=_22 / 10, h23time=_23 / 10, h24time=_0 / 10)

        return HttpResponse(_list)

    return TemplateResponse(request, 'importexcel.html', {'form': form, })


def newhtimecpu(gid, pid, i, htime, mande):
    try:
        sellgsinhur = SellGsInHour.objects.get(gs_id=gid, product_id=pid)
        meghdar = 0
        newmeghdar = 0
        newhtime = htime
        mytime = 72
        do = 0
        a = 0
        for _c in range(i):
            if meghdar > 0:
                do += 1
            match newhtime:
                case 1:
                    meghdar += sellgsinhur.h1time
                case 2:
                    meghdar += sellgsinhur.h2time
                case 3:
                    meghdar += sellgsinhur.h3time
                case 4:
                    meghdar += sellgsinhur.h4time
                case 5:
                    meghdar += sellgsinhur.h5time
                case 6:
                    meghdar += sellgsinhur.h6time
                case 7:
                    meghdar += sellgsinhur.h7time
                case 8:
                    meghdar += sellgsinhur.h8time
                case 9:
                    meghdar += sellgsinhur.h9time
                case 10:
                    meghdar += sellgsinhur.h10time
                case 11:
                    meghdar += sellgsinhur.h11time
                case 12:
                    meghdar += sellgsinhur.h12time
                case 13:
                    meghdar += sellgsinhur.h13time
                case 14:
                    meghdar += sellgsinhur.h14time
                case 15:
                    meghdar += sellgsinhur.h15time
                case 16:
                    meghdar += sellgsinhur.h16time
                case 17:
                    meghdar += sellgsinhur.h17time
                case 18:
                    meghdar += sellgsinhur.h18time
                case 19:
                    meghdar += sellgsinhur.h19time
                case 20:
                    meghdar += sellgsinhur.h20time
                case 21:
                    meghdar += sellgsinhur.h21time
                case 22:
                    meghdar += sellgsinhur.h22time
                case 23:
                    meghdar += sellgsinhur.h23time
                case 24:
                    meghdar += sellgsinhur.h24time
                case 25:
                    meghdar += sellgsinhur.h1time
                case 26:
                    meghdar += sellgsinhur.h2time
                case 27:
                    meghdar += sellgsinhur.h3time
                case 28:
                    meghdar += sellgsinhur.h4time
                case 29:
                    meghdar += sellgsinhur.h5time
                case 30:
                    meghdar += sellgsinhur.h6time
                case 31:
                    meghdar += sellgsinhur.h7time
                case 32:
                    meghdar += sellgsinhur.h8time
                case 33:
                    meghdar += sellgsinhur.h9time
                case 34:
                    meghdar += sellgsinhur.h10time
                case 35:
                    meghdar += sellgsinhur.h11time
                case 36:
                    meghdar += sellgsinhur.h12time
                case 37:
                    meghdar += sellgsinhur.h13time
                case 38:
                    meghdar += sellgsinhur.h14time
                case 39:
                    meghdar += sellgsinhur.h15time
                case 40:
                    meghdar += sellgsinhur.h16time
                case 41:
                    meghdar += sellgsinhur.h17time
                case 42:
                    meghdar += sellgsinhur.h18time
                case 43:
                    meghdar += sellgsinhur.h19time
                case 44:
                    meghdar += sellgsinhur.h20time
                case 45:
                    meghdar += sellgsinhur.h21time
                case 46:
                    meghdar += sellgsinhur.h22time
                case 47:
                    meghdar += sellgsinhur.h23time
                case 48:
                    meghdar += sellgsinhur.h24time
                case 49:
                    meghdar += sellgsinhur.h1time
                case 50:
                    meghdar += sellgsinhur.h2time
                case 51:
                    meghdar += sellgsinhur.h3time
                case 52:
                    meghdar += sellgsinhur.h4time
                case 53:
                    meghdar += sellgsinhur.h5time
                case 54:
                    meghdar += sellgsinhur.h6time
                case 55:
                    meghdar += sellgsinhur.h7time
                case 56:
                    meghdar += sellgsinhur.h8time
                case 57:
                    meghdar += sellgsinhur.h9time
                case 58:
                    meghdar += sellgsinhur.h10time
                case 59:
                    meghdar += sellgsinhur.h11time
                case 60:
                    meghdar += sellgsinhur.h12time
                case 61:
                    meghdar += sellgsinhur.h13time
                case 62:
                    meghdar += sellgsinhur.h14time
                case 63:
                    meghdar += sellgsinhur.h15time
                case 64:
                    meghdar += sellgsinhur.h16time
                case 65:
                    meghdar += sellgsinhur.h17time
                case 66:
                    meghdar += sellgsinhur.h18time
                case 67:
                    meghdar += sellgsinhur.h19time
                case 68:
                    meghdar += sellgsinhur.h20time
                case 69:
                    meghdar += sellgsinhur.h21time
                case 70:
                    meghdar += sellgsinhur.h22time
                case 71:
                    meghdar += sellgsinhur.h23time
                case 72:
                    meghdar += sellgsinhur.h24time

            if meghdar > mande and a == 0 and mande != 0:
                mytime = do
                newmeghdar = meghdar

                a = 1
            newhtime += 1



    except:
        meghdar = 0
        return {'a': 0, 'b': 0, 'c': 0, 'd': 0, 'count': 0}

    return {'a': round(meghdar / i), 'b': round(meghdar), 'c': mytime, 'd': newmeghdar, 'count': i}


@cache_permission('reportsell')
def reportsellkolsum(request):
    add_to_log(request, f'مشاهده گزارش فروش کل مکانیکی الکترونیکی', 0)
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    far = 2
    sell = None
    zone = None

    if request.user.owner.role.role in ['mgr', 'setad']:
        zone = Zone.objects_limit.all()
    else:
        zone_id = 1599511

    product = Product.objects.all()
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        az = mdate
        ta = mdate2

        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        zone_id = request.POST.get('zone')
        far = int(request.POST.get('select4'))
        _role = request.user.owner.role.role
        _roleid = zoneorarea(request)
        if request.user.owner.role.role == 'zone':
            zone_id = request.user.owner.zone_id
        elif request.user.owner.role.role == 'area':
            zone_id = request.user.owner.area_id

        if zone_id != '0':
            sell = SellGs.object_role.c_gs(request, 0).values('gs__gsid', 'gs__name', 'gs__area__name',
                                                              'gs__area__zone__name').filter(
                gs__area__zone_id=zone_id, product_id=far,
                tarikh__range=(mdate, mdate2)).annotate(
                elec=Sum('yarane') + Sum('haveleh') + Sum('azad') + Sum('ezterari'), mek=Sum('sell'))

        else:
            sell = SellGs.object_role.c_gs(request, 0).values('gs__gsid', 'gs__name', 'gs__area__name',
                                                              'gs__area__zone__name').filter(
                product_id=far,
                tarikh__range=(mdate, mdate2)).annotate(
                elec=Sum('yarane') + Sum('haveleh') + Sum('azad') + Sum('ezterari'), mek=Sum('sell'))

    return TemplateResponse(request, 'reportsellkolsum.html',
                            {'sell': sell, 'az': az, 'ta': ta, 'far': int(far), 'product': product, 'zone': zone,
                             })


@cache_permission('reportsell')
def reportsellkolanalis(request):
    add_to_log(request, f'مشاهده گزارش فروش  مقایسه ایی به تفکیک جایگاه', 0)
    mdate = startdate
    mdate2 = today
    az = mdate
    ta = mdate2
    far = 2
    sell = None
    zone = None
    month_date_in = ""
    month_date_out = ""
    zonename = ""
    sell2 = None
    summer = None
    sellarea = None
    farname = ""
    zone_id = 0
    newsummer = 0

    if request.user.owner.role.role in ['mgr', 'setad']:
        zone = Zone.objects_limit.all()
    else:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id)

    product = Product.objects.all()
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        az = mdate
        ta = mdate2
        _date_in = az.split("/")
        _date_out = ta.split("/")
        oneyearold_in = int(_date_in[0]) - 1

        oneyearold_out = int(_date_out[0]) - 1

        if int(_date_in[1]) > 1:
            onemonthold_in = int(_date_in[1]) - 1
            oneyearnow_in = int(_date_in[0])
        else:
            onemonthold_in = 12
            oneyearnow_in = int(_date_in[0]) - 1

        if int(_date_out[1]) > 1:
            onemonthold_out = int(_date_out[1]) - 1
            oneyearnow_out = int(_date_out[0])
        else:
            onemonthold_out = 12
            oneyearnow_out = int(_date_out[0]) - 1

        month_date_in = jdatetime.date(day=int(_date_in[2]), month=int(_date_in[1]),
                                       year=int(oneyearold_in)).togregorian()
        month_date_out = jdatetime.date(day=int(_date_out[2]), month=int(_date_out[1]),
                                        year=int(oneyearold_out)).togregorian()
        days = (month_date_out - month_date_in).days + 1
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        zone_id = request.POST.get('zone')
        far = int(request.POST.get('select4'))
        _role = request.user.owner.role.role
        _roleid = zoneorarea(request)
        if request.user.owner.role.role == 'zone':
            zone_id = request.user.owner.zone_id
        elif request.user.owner.role.role == 'area':
            zone_id = request.user.owner.zone_id
        if zone_id != '0':
            sell = SellGs.object_role.c_gs(request, 0).values('gs__gsid', 'gs__name', 'gs__area__name',
                                                              'gs__area__zone__name').filter(gs__status__status=True,
                                                                                             gs__area__zone_id=zone_id,
                                                                                             product_id=far,
                                                                                             ).annotate(
                yarane_now=Coalesce(Sum(Case(When(tarikh__range=(mdate, mdate2), then='yarane'))), Value(0),
                                    output_field=DecimalField()),
                azad_now=Coalesce(Sum(Case(When(tarikh__range=(mdate, mdate2), then='azad'))), Value(0),
                                  output_field=DecimalField()),
                ezterari_now=Coalesce(Sum(Case(When(tarikh__range=(mdate, mdate2), then='ezterari'))), Value(0),
                                      output_field=DecimalField()),
                yarane_old=Coalesce(Sum(Case(When(tarikh__range=(month_date_in, month_date_out), then='yarane'))),
                                    Value(0), output_field=DecimalField()),
                azad_old=Coalesce(Sum(Case(When(tarikh__range=(month_date_in, month_date_out), then='azad'))),
                                  Value(0), output_field=DecimalField()),
                ezterari_old=Coalesce(Sum(Case(When(tarikh__range=(month_date_in, month_date_out), then='ezterari'))),
                                      Value(0), output_field=DecimalField()),
                total_now=Coalesce(
                    Sum(Case(When(tarikh__range=(mdate, mdate2), then=F('yarane') + F('azad') + F('ezterari')))),
                    Value(0), output_field=DecimalField()),
                total_old=Coalesce(Sum(Case(
                    When(tarikh__range=(month_date_in, month_date_out), then=F('yarane') + F('azad') + F('ezterari')))),
                    Value(0), output_field=DecimalField()),
                percent_change=Case(
                    When(total_old=0, then=Value(0)),  # برای جلوگیری از تقسیم بر صفر
                    default=ExpressionWrapper(
                        ((F('total_now') - F('total_old')) * 100) / F('total_old'),
                        output_field=IntegerField()
                    ), ),
            )

            sell2 = SellGs.object_role.c_gs(request, 0).filter(
                gs__area__zone_id=zone_id, product_id=far,
            ).aggregate(
                total_now=Sum(Case(
                    When(tarikh__range=(mdate, mdate2), then=F('yarane') + F('azad') + F('ezterari')))) / days,
                total_old=Sum(Case(
                    When(tarikh__range=(month_date_in, month_date_out),
                         then=F('yarane') + F('azad') + F('ezterari')))) / days,
            )

            sellarea = SellGs.object_role.c_gs(request, 0).values('gs__area_id', 'gs__area__name', ).filter(
                gs__area__zone_id=zone_id, product_id=far,
            ).annotate(
                yarane_now=Coalesce(Sum(Case(When(tarikh__range=(mdate, mdate2), then='yarane'))), Value(0),
                                    output_field=DecimalField()),
                azad_now=Coalesce(Sum(Case(When(tarikh__range=(mdate, mdate2), then='azad'))), Value(0),
                                  output_field=DecimalField()),
                ezterari_now=Coalesce(Sum(Case(When(tarikh__range=(mdate, mdate2), then='ezterari'))), Value(0),
                                      output_field=DecimalField()),
                yarane_old=Coalesce(Sum(Case(When(tarikh__range=(month_date_in, month_date_out), then='yarane'))),
                                    Value(0), output_field=DecimalField()),
                azad_old=Coalesce(Sum(Case(When(tarikh__range=(month_date_in, month_date_out), then='azad'))), Value(0),
                                  output_field=DecimalField()),
                ezterari_old=Coalesce(Sum(Case(When(tarikh__range=(month_date_in, month_date_out), then='ezterari'))),
                                      Value(0), output_field=DecimalField()),
                total_now=Coalesce(
                    Sum(Case(When(tarikh__range=(mdate, mdate2), then=F('yarane') + F('azad') + F('ezterari')))),
                    Value(0), output_field=DecimalField()),
                total_old=Coalesce(Sum(Case(
                    When(tarikh__range=(month_date_in, month_date_out), then=F('yarane') + F('azad') + F('ezterari')))),
                    Value(0), output_field=DecimalField()),
                percent_change=Case(
                    When(total_old=0, then=Value(0)),  # برای جلوگیری از تقسیم بر صفر
                    default=ExpressionWrapper(
                        ((F('total_now') - F('total_old')) * 100) / F('total_old'),
                        output_field=IntegerField()
                    ), ),
            )

            summer = sellarea.aggregate(s_yarane_now=Sum('yarane_now'), s_azad_now=Sum('azad_now'),
                                        s_ezterari_now=Sum('ezterari_now'), s_total_now=Sum('total_now'),
                                        s_yarane_old=Sum('yarane_old'), s_azad_old=Sum('azad_old'),
                                        s_ezterari_old=Sum('ezterari_old'), s_total_old=Sum('total_old'),
                                        m_yarane_now=Round(Sum('yarane_now') / days, precision=0),
                                        m_azad_now=Round(Sum('azad_now') / days, precision=0),
                                        m_ezterari_now=Round(Sum('ezterari_now') / days, precision=0),
                                        m_total_now=Round(Sum('total_now') / days, precision=0),
                                        m_yarane_old=Round(Sum('yarane_old') / days, precision=0),
                                        m_azad_old=Round(Sum('azad_old') / days, precision=0),
                                        m_ezterari_old=Round(Sum('ezterari_old') / days, precision=0),
                                        m_total_old=Round(Sum('total_old') / days, precision=0),
                                        )

            try:
                total_now = int(summer['s_total_now']) if summer['s_total_now'] is not None else 0
                total_old = int(summer['s_total_old']) if summer['s_total_old'] is not None else 0

                if total_old != 0:
                    newsummer = ((total_now - total_old) * 100) / total_old
                    newsummer = round(newsummer, 0)
                    newsummer = int(newsummer)
                else:
                    newsummer = 0  # یا هر مقدار پیش‌فرض دیگری که مناسب است
            except (KeyError, TypeError):
                newsummer = 0
            # newsummer = ((int(summer['s_total_now']) - int(summer['s_total_old'])) * 100) / int(summer['s_total_old'])
            # newsummer = round(newsummer,0)
            # newsummer = int(newsummer)
            # except:
            #     summer = None
            sellarea = sorted(sellarea, key=itemgetter(str('percent_change')), reverse=True)
            sell = sorted(sell, key=itemgetter('gs__area__name', 'percent_change'), reverse=True)
            zonename = Zone.objects.get(id=zone_id).name
            farname = Product.objects.get(id=far).name

    return TemplateResponse(request, 'reportsellkolanalis.html',
                            {'sell': sell, 'az': az, 'ta': ta, 'far': int(far), 'product': product, 'zone': zone,
                             'azold': month_date_in, 'taold': month_date_out, 'zonename': zonename, 'sell2': sell2,
                             'summer': summer, 'sellarea': sellarea, 'zone_id': int(zone_id), 'farname': farname,
                             'newsummer': newsummer,
                             })


def uploadsoratjalase(request, _id):
    url = request.META.get('HTTP_REFERER')
    sell = SellModel.objects.get(id=_id)
    if request.method == 'POST':
        add_to_log(request, f' مشاهده فرم بارگذاری صورتجلسه', 0)
        form = UploadSoratjalaseForm(request.POST, request.FILES, instance=sell)
        if form.is_valid():
            try:
                uploaded_file = request.FILES.get('image')
                if uploaded_file and uploaded_file.size <= 500 * 1024 and uploaded_file.size > 10:

                    form.save()
                    messages.info(request, 'بارگذاری صورتجلسه بدرستی انجام شد')
                else:
                    messages.error(request, 'عملیات شکست خورد ، پسوند فایل باید jpg  و سایز حداکثر 500 کیلوبایت باشد.')

            except:
                messages.error(request, 'عملیات شکست خورد ، پسوند فایل باید Jpg  و سایز حداکثر 500 کیلوبایت باشد.')
        else:
            messages.error(request, 'عملیات شکست خورد ، پسوند فایل باید Jpg  و سایز حداکثر 500 کیلوبایت باشد.')

    return redirect('sell:sellday')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def fartonazel(request):
    mylist = []
    _gsid = request.GET.get('gsid')
    _product = request.GET.get('val')
    gs = Pump.objects.filter(gs_id=_gsid, product_id=_product).order_by('number')
    for item in gs:
        dict = {
            'id': item.id,
            'number': item.number
        }
        mylist.append(dict)

    return JsonResponse({"message": "success", "mylist": mylist})


@cache_permission('daftartolombe')
def daftartolombe(request):
    if request.user.owner.role.role == 'gs':
        _btmt = GsList.objects.filter(owner_id=request.user.owner.id).first()
        if not _btmt.gs.btmt:
            messages.error(request, 'شما به این بخش دسترسی ندارید')
            return redirect('base:home')

    molahezat = []
    add_to_log(request, " تهیه گزارش  کارکرد تلمبه ", 0)
    mdate = startdate
    mdate2 = today
    _list = []
    az = mdate
    ta = mdate2
    if request.user.owner.role.role == 'gs':
        GsList.objects.filter(owner_id=request.user.owner.id)
        gss = GsModel.objects.filter(gsowner__owner=request.user.owner)
    if request.user.owner.role.role == 'zone':
        gss = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    if request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area_id)
    if request.user.owner.role.role in ['mgr', 'setad']:
        gss = GsModel.objects.all()

    product = Product.objects.all()
    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        az = mdate
        ta = mdate2
        start_date = to_miladi(mdate)
        end_date = to_miladi(mdate2)
        _role = request.user.owner.role.role
        _roleid = zoneorarea(request)
        mdate = mdate.replace("/", '-')
        mdate2 = mdate2.replace("/", '-')
        # start_date = newdate.strptime(mdate, '%Y-%m-%d')
        # end_date = newdate.strptime(mdate2, '%Y-%m-%d')
        # لیستی برای ذخیره تاریخ‌های موجود در دیکشنری

        gsid = request.POST.get('select3')
        nazelid = int(request.POST.get('nazel'))
        far = int(request.POST.get('select4'))
        pump = None
        if nazelid != 0:
            pump = Pump.objects.get(id=nazelid)
        pumplist = Pump.objects.filter(gs_id=gsid, product_id=far)

        if request.user.owner.role.role == 'gs':
            _gslist = GsList.objects.filter(gs_id=gsid, owner_id=request.user.owner.id)
            if _gslist:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid)
            else:
                _gslist = GsList.objects.filter(owner_id=request.user.owner.id).first()
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=_gslist.gs_id)
        if request.user.owner.role.role == 'zone':
            if gsid == '0':
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(
                    gs__area__zone_id=request.user.owner.zone_id)
            else:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid,
                                                                          gs__area__zone_id=request.user.owner.zone_id)
        if request.user.owner.role.role == 'area':
            if gsid == '0':
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__area_id=request.user.owner.area_id)
            else:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid,
                                                                          gs__area_id=request.user.owner.area_id)
        if request.user.owner.role.role in ['setad', 'mgr']:
            if gsid == '0':
                sellmodel = SellModel.object_role.c_gs(request, 0).all()
            else:
                sellmodel = SellModel.object_role.c_gs(request, 0).filter(gs__exact=gsid)

        sellmodel = sellmodel.filter(tolombeinfo_id=nazelid, tarikh__range=(mdate, mdate2)).order_by('tarikh')
        sumsell = 0
        for item in sellmodel:
            molahezat = []
            if AcceptForBuy.objects.filter(gs__id=gsid,
                                           tarikh=item.tarikh).count() > 0:
                molahezat.append("مجوز مغایریت در حال بررسی")
            sell = float(item.sell) - float(item.azmayesh) if item.azmayesh else item.sell
            if item.sell:
                sumsell += item.sell
            else:
                sumsell += 0

            _list.append({
                'tarikh': item.tarikh,
                'start': item.end,
                'start2': item.t_start,
                'end': item.start,
                'end2': item.t_end,
                'az': item.azmayesh,
                'sellnoaz': sell,
                'sell': item.sell,
                'sumsell': sumsell,
                'molahezat': molahezat,
                'iscrash': item.iscrash,
                'is_soratjalase': item.is_soratjalase,
                'is_change_counter': item.is_change_counter,
                'image': item.image,
            })

        existing_dates = []
        for item in _list:
            existing_dates.append(item['tarikh'].strftime('%Y-%m-%d'))

        missing_dates = []

        _datein = datetime.date(day=int(start_date.day), month=int(start_date.month), year=int(start_date.year))
        jd = JDate(_datein.strftime("%Y-%m-%d"))
        _datein = jd.format('Y-m-d')
        year, month, day = map(int, _datein.split('-'))
        _datein = jdatetime.date(year=year, month=month, day=day)
        _dateout = datetime.date(day=int(end_date.day), month=int(end_date.month), year=int(end_date.year))
        jd = JDate(_dateout.strftime("%Y-%m-%d"))
        _dateout = jd.format('Y-m-d')
        year, month, day = map(int, _dateout.split('-'))
        _dateout = jdatetime.date(year=year, month=month, day=day)
        # حلقه برای پیمایش بین دو تاریخ
        current_date = _datein
        while current_date <= _dateout:
            current_date_str = current_date.strftime('%Y-%m-%d')
            if current_date_str not in existing_dates:
                missing_dates.append(current_date_str)
            current_date += timedelta(days=1)
        for _ in missing_dates:
            date = jdatetime.datetime.strptime(_, '%Y-%m-%d').date()
            end_date = CloseGS.objects.filter(gs__id=gsid,
                                              ).order_by('-id')

            for item in end_date:
                if date >= item.date_in and date <= item.date_out:
                    molahezat.append("ثبت بازه تعطیلی و تعویض هارد")
                    break

            # datein = datetime.date(day=int(date.day), month=int(date.month), year=int(date.year))
            # jd = JDate(datein.strftime("%Y-%m-%d"))
            # date = jd.format('Y-m-d')
            # year, month, day =map(int, date.split('-'))
            # date = jdatetime.date(year=year, month= month, day=day)
            # date=date.strftime('%Y-%m-%d')
            _list.append({
                'tarikh': date,
                'start': 0,
                'start2': 0,
                'end': 0,
                'end2': 0,
                'az': 0,
                'sellnoaz': 0,
                'sell': 0,
                'sumsell': 0,
                'molahezat': molahezat,
                'iscrash': False,
                'is_change_counter': False,
                'is_soratjalase': False,
                'image': None,
            })
        _list = sorted(_list, key=itemgetter('tarikh'), reverse=False)
        return TemplateResponse(request, 'daftartolombe.html',
                                {'list': _list, 'mdate': mdate, 'mdate2': mdate2, 'gss': gss, 'gsid': int(gsid),
                                 'product': product, 'far': far, 'nazelid': nazelid, 'pump': pump, 'pumplist': pumplist,
                                 'az': az, 'ta': ta})
    return TemplateResponse(request, 'daftartolombe.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'gss': gss, 'product': product,
                             })


@cache_permission('acclosegs')
def change_sell_list(request, _id):
    acs = AccessChangeSell.objects.get(id=_id)
    tarikh = to_miladi(str(acs.tarikh))
    _list = EditSell.objects.filter(sell__gs_id=acs.gs_id, sell__tarikh=acs.tarikh)

    return TemplateResponse(request, 'changeselllist.html', {'list': _list})


class DiscrepancyListView(LoginRequiredMixin, ListView):
    model = DiscrepancyApproval
    template_name = 'discrepancy_list.html'
    context_object_name = 'discrepancies'

    @method_decorator(cache_permission('discrepancylist'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        add_to_log(self.request, f'مشاهده فرم مغایرت های بررسی نشده ', 0)
        user = self.request.user.owner
        three_days_ago = timezone.now().date() - timedelta(days=3)
        jd = JDate(three_days_ago.strftime("%Y-%m-%d %H:%M:%S"))
        three_days_ago_jalali = jd.format('Y-m-d')

        if user.role.role == 'area':
            base_query = DiscrepancyApproval.objects.filter(
                discrepancy_date__lte=three_days_ago_jalali
            )
            return base_query.filter(area_head=user)
        elif user.role.role == 'engin':
            return DiscrepancyApproval.objects.filter(status='pendingarea')
        elif user.refrence_id == 1:
            return DiscrepancyApproval.objects.filter(status='engineering_approved')
        return DiscrepancyApproval.objects.none()


class DiscrepancyUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = DiscrepancyApproval
    fields = ['reason', ]
    template_name = 'discrepancy_form.html'

    @method_decorator(cache_permission('discrepancylist'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def test_func(self):
        user = self.request.user.owner
        discrepancy = self.get_object()

        if user.role.role == 'area' and discrepancy.status in ['pending', 'pendingarea']:
            return True
        if user.role.role == 'engin' and discrepancy.status == 'pendingarea':
            return True
        if user.refrence_id == 1 and discrepancy.status == 'engineering_approved':
            return True

        return False

    def form_invalid(self, form):
        return HttpResponse(form.errors)

    def form_valid(self, form):
        user = self.request.user.owner
        if user.role.role == 'engin':
            form.instance.engineering_approver = user
            form.instance.status = 'engineering_approved'
        elif user.refrence_id == 1:
            form.instance.system_approver = user
            form.instance.status = 'system_approved'
        elif user.role.role == 'area':
            form.instance.status = 'pendingarea'

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('sell:discrepancy_list')


def import_excel_waybill(request):
    add_to_log(request, 'دریافت اکسل بارنامه ها', 0)
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath.path  # اضافه کردن .path برای گرفتن مسیر کامل فایل

            try:
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                m_row = sheet_obj.max_row

                # خواندن هدرها برای نگاشت خودکار ستون‌ها
                headers = []
                for j in range(1, sheet_obj.max_column + 1):
                    cell_obj = sheet_obj.cell(row=1, column=j)
                    headers.append(cell_obj.value)

                # حلقه از سطر دوم (داده‌ها)
                for i in range(2, m_row + 1):
                    # ایجاد دیکشنری برای نگاشت ستون‌ها
                    row_data = {}
                    for j, header in enumerate(headers, 1):
                        cell_obj = sheet_obj.cell(row=i, column=j)
                        row_data[header] = cell_obj.value

                    # حالا می‌توانید به ستون‌ها دسترسی داشته باشید:
                    waybill_id = row_data.get('waybill_id')
                    product_id = row_data.get('product_id')
                    quantity = row_data.get('quantity')
                    quantity60 = row_data.get('quantity60')
                    weight = row_data.get('weight')
                    degree = row_data.get('degree')
                    special_weight = row_data.get('special_weight')
                    customer_name = row_data.get('customer_name')
                    exit_date = row_data.get('exit_date')
                    exit_time = row_data.get('exit_time')
                    customer_code = row_data.get('customer_code')
                    SendType_value = row_data.get('SendType')
                    gsid_value = str(row_data.get('gsid'))
                    if len(gsid_value) == 3:
                        gsid_value = "0" + str(gsid_value)
                    if len(gsid_value) == 2:
                        gsid_value = "00" + str(gsid_value)
                    if len(gsid_value) == 1:
                        gsid_value = "000" + str(gsid_value)

                    # پردازش و ذخیره داده‌ها در مدل Waybill
                    try:
                        # پیدا کردن product_id مربوطه
                        product_id_obj = ProductId.objects.get(productid=product_id)

                        # پیدا کردن gsid (اختیاری)
                        gs_obj = None
                        if gsid_value:
                            try:
                                gs_obj = GsModel.objects.get(gsid=gsid_value)
                            except GsModel.DoesNotExist:
                                pass

                        # تبدیل تاریخ و زمان
                        if isinstance(exit_date, str):
                            # اگر تاریخ به صورت رشته است (مثلاً 1404/06/21)

                            exit_date = to_miladi(exit_date)

                        # ایجاد یا به‌روزرسانی رکورد Waybill

                        Waybill.objects.update_or_create(
                            waybill_id=waybill_id,
                            defaults={
                                'product_id': product_id_obj,
                                'quantity': quantity or 0,
                                'quantity60': quantity60 or 0,
                                'weight': weight or 0,
                                'degree': degree or 0,
                                'special_weight': special_weight or 0,
                                'customer_name': customer_name or '',
                                'exit_date': exit_date,
                                'exit_time': exit_time,
                                'customer_code': customer_code or '',
                                'send_type': SendType.objects.get(id=SendType_value) if SendType_value else None,
                                'gsid': gs_obj,
                            }
                        )

                        _list.append(f"بارنامه {waybill_id} با موفقیت اضافه شد.")

                    except ProductId.DoesNotExist:
                        _list.append(f"خطا: شناسه فرآورده {product_id} یافت نشد.")
                    except Exception as e:
                        _list.append(f"خطا در پردازش سطر {i}: {str(e)}")

                messages.success(request, 'بارنامه‌ها با موفقیت وارد شدند.')

            except Exception as e:
                messages.error(request, f'خطا در خواندن فایل اکسل: {str(e)}')

    return render(request, 'importexcel.html', {'form': form})


@cache_permission('reportsell')
def report_waybill_sell(request):
    if not request.user.is_superuser:
        return False
    area = None
    _city = None
    _gsmodel = None
    az_jalali = None
    ta_jalali = None
    # مقادیر پیش‌فرض
    az = request.GET.get('select', "0")
    ta = request.GET.get('select2', "0")
    gsid = request.GET.get('select3', 0)
    zoneid = request.GET.get('zone', 0)
    areaid = request.GET.get('area', 0)
    cityid = request.GET.get('city', 0)
    _gsid = request.GET.get('gs', 0)
    productid = request.GET.get('product', 0)
    products = Product.objects.all()
    zones = Zone.objects_limit.all()

    if productid == "0":
        messages.error(request, 'لطفا ابتدا یک فرآورده انتخاب کنید')
        return render(request, 'report_waybill_sell.html', {'zones': zones, 'products': products})
    if len(az) < 10 or len(ta) < 10:
        messages.error(request, 'لطفا تاریخ معتبر انتخاب کنید')
        return render(request, 'report_waybill_sell.html', {'zones': zones, 'products': products})
    add_to_log(request, f'مشاهده فرم رسیده و موجودی ', 0)
    if request.user.owner.role.role == 'zone':
        zoneid = request.user.owner.zone_id
        zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)

    if not az:
        return render(request, 'report_waybill_sell.html', {'zones': zones, 'products': products})
    # تبدیل تاریخ‌ها به Gregorian
    try:
        az_jalali = to_miladi(az)
        ta_jalali = to_miladi(ta)
        az_gregorian = az_jalali.togregorian().date()
        ta_gregorian = ta_jalali.togregorian().date()
    except:
        az_gregorian = ta_gregorian = jdatetime.date.today().togregorian()

    # لیست داده‌های گزارش
    report_data = []
    gss = GsModel.object_role.c_gsmodel(request).filter(area__zone_id=zoneid)
    if zoneid != '0':
        area = Area.objects.filter(zone_id=zoneid)
        gss = gss.filter(area__zone_id=zoneid)
    if areaid != '0':
        gss = gss.filter(area_id=areaid)
        _city = City.objects.filter(area_id=areaid)

    if cityid != '0':
        gss = gss.filter(city_id=cityid)
    if _gsid != '0':
        gss = gss.filter(id=_gsid)
    for gs in gss:
        tarikh_az_yesterday = az_jalali - datetime.timedelta(days=1)
        tarikh_ta_yesterday = ta_jalali - datetime.timedelta(days=1)
        # موجودی اول دوره (قبل از تاریخ شروع)
        mojodi_aval = Mojodi.objects.filter(
            gs_id=gs.id,
            tarikh=tarikh_az_yesterday,

        ).order_by('-tarikh').first()
        mojodi_akhar = Mojodi.objects.filter(
            gs_id=gs.id,
            tarikh=tarikh_ta_yesterday,

        ).order_by('-tarikh').first()

        mojodi_aval_value = 0
        mojodi_akhar_value = 0
        if mojodi_aval:
            if productid == "2":
                mojodi_aval_value = mojodi_aval.benzin

            if productid == "3":
                mojodi_aval_value = mojodi_aval.super

            if productid == "4":
                mojodi_aval_value = mojodi_aval.gaz

        if mojodi_akhar:
            if productid == "2":
                mojodi_akhar_value = mojodi_akhar.benzin
            if productid == "3":
                mojodi_akhar_value = mojodi_akhar.super
            if productid == "4":
                mojodi_akhar_value = mojodi_akhar.gaz

        # بارنامه‌های رسیده در بازه زمانی

        waybill_total = Waybill.objects.filter(
            product_id__product_id=productid,
            gsid=gs.gsid,
            exit_date__gte=az_jalali,
            exit_date__lte=ta_jalali
        ).aggregate(total=Sum('quantity'))['total'] or 0

        # فروش مکانیکی در بازه زمانی
        sell_mechanical = SellModel.objects.filter(
            product_id=productid,
            gs_id=gs.id,
            tarikh__gte=az_jalali,
            tarikh__lte=tarikh_ta_yesterday
        ).aggregate(total=Sum('sell'))['total'] or 0

        # فروش الکترونیکی در بازه زمانی
        sell_electronic = SellModel.objects.filter(
            product_id=productid,
            gs_id=gs.id,
            tarikh__gte=az_jalali,
            tarikh__lte=tarikh_ta_yesterday
        ).aggregate(total=Sum('sellkol'))['total'] or 0
        if mojodi_aval_value != 0 or mojodi_akhar_value != 0 or sell_mechanical != 0:
            report_data.append({
                'gs_name': gs.name,
                'gsid': gs.gsid,
                'area_name': gs.area.name if gs.area else '',
                'zone_name': gs.area.zone.name if gs.area and gs.area.zone else '',
                'mojodi_aval': mojodi_aval_value,
                'mojodi_akhar': mojodi_akhar_value,
                'waybill_total': waybill_total,
                'sell_mechanical': sell_mechanical,
                'sell_electronic': sell_electronic,
                'mande': ((mojodi_aval_value + waybill_total) - sell_mechanical) - mojodi_akhar_value
            })

    # جمع‌های کل
    total_sums = {
        'mojodi_aval_total': sum(item['mojodi_aval'] for item in report_data),
        'waybill_total_sum': sum(item['waybill_total'] for item in report_data),
        'sell_mechanical_total': sum(item['sell_mechanical'] for item in report_data),
        'sell_electronic_total': sum(item['sell_electronic'] for item in report_data),
    }

    context = {
        'list': report_data,
        'summer': total_sums,
        'az': az,
        'areaid': int(areaid),
        'area': area,
        'zones': zones,
        'products': products,
        'ta': ta,
        'gsid': int(_gsid),
        'zoneid': int(zoneid),
        'productid': int(productid),
        'gs_list': gss,
        'cityid': int(cityid),
        'city': _city,
        'product': Product.objects.all(),  # اگر نیاز باشد
    }

    return TemplateResponse(request, 'report_waybill_sell.html', context)


def import_waybill():
    # Waybill.objects.all().delete()
    _list = []

    path = r'd:\media\test.xlsx'
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row

        # خواندن هدرها برای نگاشت خودکار ستون‌ها
        headers = []
        for j in range(1, sheet_obj.max_column + 1):
            cell_obj = sheet_obj.cell(row=1, column=j)
            headers.append(cell_obj.value)

        # حلقه از سطر دوم (داده‌ها)
        for i in range(2, m_row + 1):
            # ایجاد دیکشنری برای نگاشت ستون‌ها
            row_data = {}
            for j, header in enumerate(headers, 1):
                cell_obj = sheet_obj.cell(row=i, column=j)
                row_data[header] = cell_obj.value

            # حالا می‌توانید به ستون‌ها دسترسی داشته باشید:
            waybill_id = row_data.get('waybill_id')
            product_id = row_data.get('product_id')
            quantity = row_data.get('quantity')
            quantity60 = row_data.get('quantity60')
            weight = row_data.get('weight')
            degree = row_data.get('degree')
            special_weight = row_data.get('special_weight')
            customer_name = row_data.get('customer_name')
            exit_date = row_data.get('exit_date')
            exit_time = row_data.get('exit_time')
            customer_code = row_data.get('customer_code')
            SendType_value = row_data.get('SendType')
            gsid_value = str(row_data.get('gsid'))
            if len(gsid_value) == 3:
                gsid_value = "0" + str(gsid_value)
            if len(gsid_value) == 2:
                gsid_value = "00" + str(gsid_value)
            if len(gsid_value) == 1:
                gsid_value = "000" + str(gsid_value)

            # پردازش و ذخیره داده‌ها در مدل Waybill
            try:
                # پیدا کردن product_id مربوطه
                product_id_obj = ProductId.objects.get(productid=product_id)

                # پیدا کردن gsid (اختیاری)
                gs_obj = None
                if gsid_value:
                    try:
                        gs_obj = GsModel.objects.get(gsid=gsid_value)
                    except GsModel.DoesNotExist:
                        pass

                # تبدیل تاریخ و زمان
                if isinstance(exit_date, str):
                    # اگر تاریخ به صورت رشته است (مثلاً 1404/06/21)

                    exit_date = to_miladi(exit_date)

                # ایجاد یا به‌روزرسانی رکورد Waybill

                Waybill.objects.update_or_create(
                    waybill_id=waybill_id,
                    defaults={
                        'product_id': product_id_obj,
                        'quantity': quantity or 0,
                        'quantity60': quantity60 or 0,
                        'weight': weight or 0,
                        'degree': degree or 0,
                        'special_weight': special_weight or 0,
                        'customer_name': customer_name or '',
                        'exit_date': exit_date,
                        'exit_time': exit_time,
                        'customer_code': customer_code or '',
                        'send_type': SendType.objects.get(id=SendType_value) if SendType_value else None,
                        'gsid': gs_obj,
                    }
                )

                _list.append(f"بارنامه {waybill_id} با موفقیت اضافه شد.")

            except ProductId.DoesNotExist:
                _list.append(f"خطا: شناسه فرآورده {product_id} یافت نشد.")
                return False
            except Exception as e:
                _list.append(f"خطا در پردازش سطر {i}: {str(e)}")
                return False


    except Exception as e:
        pass
    return True


@cache_permission('reportsell')
def report_waybill_sent(request):
    url = request.META.get('HTTP_REFERER')
    area = None
    _city = None
    _gsmodel = None
    role = request.user_data.get('role_name')
    _zone_id = request.user_data.get('zone_id')
    _owner_id = request.user_data.get('owner_id')
    _area_id = request.user_data.get('area_id')
    senders = Sender.objects.all()
    zones = Zone.objects_limit.all()
    if role in ['zone', 'area']:
        zones = Zone.objects_limit.filter(id=_zone_id)
    az = jdatetime.date.today().strftime('%Y/%m/%d')
    ta = jdatetime.date.today().strftime('%Y/%m/%d')
    if request.method == 'POST':

        # دریافت پارامترهای فیلتر
        zone_id = request.POST.get('zone', '0')
        gs_id = request.POST.get('gs', '0')
        product_id = request.POST.get('product', '0')
        senderid = request.POST.get('sender', '0')
        date_from = request.POST.get('select', '')
        date_to = request.POST.get('select2', '')
        areaid = request.POST.get('area', 0)
        cityid = request.POST.get('city', 0)
        role = request.user_data.get('role_name')
        _zone_id = request.user_data.get('zone_id')
        _owner_id = request.user_data.get('owner_id')
        az = date_from
        ta = date_to
        d3 = 1
        if len(date_from) < 10:

            az = jdatetime.date.today().strftime('%Y/%m/%d')
            ta = jdatetime.date.today().strftime('%Y/%m/%d')
        else:
            d1 = to_miladi(az)
            d2 = to_miladi(ta)
            d3 = (d2 - d1).days


        if zone_id == '0':
            messages.warning(request, 'ابتدا یک منطقه انتخاب کنید')
            return redirect(url)

        if gs_id == '0' and d3 > 1:
            messages.warning(request, 'بازه تاریخ برای همه جایگاه ها نباید بیش از 1 روز باشد')
            messages.warning(request, 'ابتدا یک منطقه انتخاب کنید')
            return redirect(url)

        gs_list = GsModel.object_role.c_gsmodel(request).all()
        add_to_log(request, f'مشاهده فرم فرآورده های ارسال شده ', 0)
        if role in ['zone', 'area']:
            zone_id = _zone_id
            gs_list = GsModel.object_role.c_gsmodel(request).select_related('area').all()
            zones = Zone.objects_limit.filter(id=_zone_id)
        try:
            date_from = to_miladi(date_from)
            date_to = to_miladi(date_to)

            # فیلتر اولیه
            waybills = Waybill.objects.select_related('send_type', 'sender_new').filter(send_type__isnull=False)

            # اعمال فیلترها
            if zone_id != '0':
                waybills = waybills.filter(gsid__area__zone_id=zone_id)
                gs_list = gs_list.filter(area__zone_id=zone_id)
                area = Area.objects.select_related('zone').filter(zone_id=zone_id)

            if areaid != '0':
                gs_list = gs_list.filter(area_id=areaid)
                _city = City.objects.select_related('area').filter(area_id=areaid)
                waybills = waybills.filter(gsid__area_id=areaid)

            if cityid != '0':
                gs_list = gs_list.filter(city_id=cityid)
                waybills = waybills.filter(gsid__city_id=cityid)
            if gs_id != '0':
                if d3 > 31:
                    messages.warning(request, 'بازه تاریخ نباید بیش از 31 روز باشد')
                    return redirect(url)
                waybills = waybills.filter(gsid_id=gs_id)

            if senderid != '0':
                waybills = waybills.filter(sender_new_id=senderid)

            if product_id != '0':
                waybills = waybills.filter(product_id__product_id=product_id)

            if date_from:
                waybills = waybills.filter(exit_date__gte=date_from)

            if date_to:
                waybills = waybills.filter(exit_date__lte=date_to)

            if request.user.owner.role.role == 'area':
                waybills = waybills.filter(gsid__area_id=request.user.owner.area_id)

            if request.user.owner.role.role in ['gs', 'tek']:
                waybills = waybills.filter(gsid__gsowner__owner_id=request.user.owner.id)
            # مرتب سازی
            waybills = waybills.order_by('-exit_date', '-exit_time')
            # senders = Sender.objects.all()
            # zones = Zone.objects_limit.all()
            # print(role)
            # if role in ['zone', 'area']:
            #     zones = Zone.objects_limit.filter(zone_id=_zone_id)

            # آماده کردن داده‌ها برای تمپلیت
            context = {
                'waybills': waybills,
                'senders': senders,
                'senderid': senderid,
                'zones': zones,
                'gs_list': gs_list,
                'products': Product.objects.all(),
                'zoneid': int(zone_id) if zone_id != '0' else 0,
                'gsid': int(gs_id) if gs_id != '0' else 0,
                'productid': int(product_id) if product_id != '0' else 0,
                'az': az,
                'ta': ta,
                'area': area,
                'city': _city,
                'cityid': int(cityid),
                'areaid': int(areaid),

            }
        except:
            context = {
                'senders': senders,
                'zones': zones,
                'gs_list': gs_list,
                'products': Product.objects.all(),
                'az': az,
                'ta': ta,
            }

        return TemplateResponse(request, 'report_waybill_sent.html', context)
    context = {
        'senders': senders,
        'zones': zones,
        'products': Product.objects.all(),
        'az': az,
        'ta': ta,
    }
    return TemplateResponse(request, 'report_waybill_sent.html', context)


@cache_permission('listsell')
def waybill_receive(request, gs_id):
    # دریافت جایگاه
    # gs = get_object_or_404(GsModel, id=gs_id)
    url = request.META.get('HTTP_REFERER')
    try:
        gs = GsModel.object_role.c_gsmodel(request).get(id=gs_id)
    except:
        messages.error(request, 'دسترسی غیر مجاز')
        return redirect(url)

    # دریافت بارنامه‌هایی که send_type = 1 هستند
    waybills = Waybill.objects.filter(
        gsid=gs,
        send_type_id=1  # بارنامه‌های ارسال شده
    )

    default_date = jdatetime.date.today().strftime('%Y/%m/%d')
    now = datetime.datetime.now()
    default_time = now.strftime('%H:%M')

    if request.method == 'POST':
        # پردازش فرم دریافت بارنامه
        updated_count = 0

        for waybill in waybills:
            # چک کنیم که آیا این بارنامه انتخاب شده است یا نه
            checkbox_field = f'selected_{waybill.id}'
            received_field = f'received_{waybill.id}'
            date_field = f'receive_date_{waybill.id}'
            time_field = f'receive_time_{waybill.id}'

            if checkbox_field in request.POST and request.POST[checkbox_field] == 'on':
                # فقط بارنامه‌های انتخاب شده را پردازش کنیم
                try:
                    received_quantity = float(request.POST.get(received_field, waybill.quantity))
                    receive_date = request.POST.get(date_field)
                    receive_time = request.POST.get(time_field)

                    # اعتبارسنجی فیلدهای اجباری
                    if not receive_date or not receive_time:
                        messages.error(request,
                                       f'لطفاً تاریخ و زمان رسید بارنامه {waybill.waybill_id} را وارد کنید')
                        continue

                    # اعتبارسنجی مقدار
                    if received_quantity < 0:
                        messages.error(request, f'مقدار بارنامه {waybill.waybill_id} نمی‌تواند منفی باشد')
                        continue

                    # تبدیل تاریخ شمسی به میلادی
                    try:
                        date_parts = receive_date.split('/')
                        jd = jdatetime.date(int(date_parts[0]), int(date_parts[1]), int(date_parts[2]))
                        gregorian_date = jd.togregorian()
                    except (ValueError, IndexError):
                        messages.error(request, f'فرمت تاریخ بارنامه {waybill.waybill_id} نامعتبر است')
                        continue

                    # به‌روزرسانی بارنامه
                    waybill.received_quantity = received_quantity
                    waybill.receive_date = gregorian_date  # تاریخ میلادی
                    waybill.receive_time = receive_time
                    waybill.send_type_id = 5  # وضعیت رسید
                    waybill.save()
                    updated_count += 1

                except ValueError:
                    messages.error(request, f'مقدار وارد شده برای بارنامه {waybill.waybill_id} نامعتبر است')

        if updated_count > 0:
            messages.success(request, f'{updated_count} بارنامه با موفقیت رسید شد')
        else:
            messages.warning(request, 'هیچ بارنامه‌ای انتخاب نشده است')

        return redirect('sell:waybill_receive', gs_id=gs_id)

    context = {
        'gs': gs,
        'waybills': waybills,
        'default_date': default_date,
        'default_time': default_time,
    }
    return TemplateResponse(request, 'waybill_receive.html', context)


@cache_permission('reports2')
def dahe(request):
    products = Product.objects.all()
    zones = Zone.objects_limit.all()
    _list = []
    az = ''
    ta = ''
    product_id = 2
    rep = ""
    year = jdatetime.date.today().year
    month = jdatetime.date.today().month

    if request.user.owner.role.role in ['area', 'zone']:
        zones = Zone.objects.filter(id=request.user.owner.zone.id)

    if request.method == 'POST':
        zone_id = request.POST.get('zone')
        area_id = request.POST.get('area')
        gs_id = request.POST.get('gs')
        product_id = request.POST.get('product')
        year = int(request.POST.get('year', jdatetime.date.today().year))
        month = int(request.POST.get('month', jdatetime.date.today().month))

        # محاسبه تاریخ شروع و پایان ماه
        try:
            az = f"{year}/{month:02d}/01"
            # پیدا کردن آخرین روز ماه
            if month == 12:
                last_day = 29 if jdatetime.date(year + 1, 1, 1).togregorian() - jdatetime.date(year, 12,
                                                                                               29).togregorian() == timedelta(
                    days=1) else 30
            else:
                last_day = 29 if jdatetime.date(year, month + 1, 1).togregorian() - jdatetime.date(year, month,
                                                                                                   29).togregorian() == timedelta(
                    days=1) else 30

            ta = f"{year}/{month:02d}/{last_day:02d}"

            date_in = to_miladi(az)
            date_out = to_miladi(ta)

        except Exception as e:
            messages.error(request, f'خطا در محاسبه تاریخ: {str(e)}')
            date_in = to_miladi(today)
            date_out = to_miladi(today)

        far = products.get(id=product_id).name
        month_name = jdatetime.date(year, month, 1).strftime('%B')
        if month_name == 'Farvardin':
            month_name = 'فروردین'
        elif month_name == 'Ordibehesht':
            month_name = 'اردیبهشت'
        elif month_name == 'Khordad':
            month_name = 'خرداد'
        elif month_name == 'Tir':
            month_name = 'تیر'
        elif month_name == 'Mordad':
            month_name = 'مرداد'
        elif month_name == 'Shahrivar':
            month_name = 'شهریور'
        elif month_name == 'Mehr':
            month_name = 'مهر'
        elif month_name == 'Aban':
            month_name = 'آبان'
        elif month_name == 'Azar':
            month_name = 'آذر'
        elif month_name == 'Dey':
            month_name = 'دی'
        elif month_name == 'Bahman':
            month_name = 'بهمن'
        elif month_name == 'Esfand':
            month_name = 'اسفند'

        rep = f'گزارش اختلاف فروش مکانیکی و الکترونیکی برای ماه  {month_name} سال {year} - فرآورده {far}'

        sells = SellGs.object_role.c_gs(request, 0).filter(tarikh__range=(date_in, date_out), product_id=product_id)

        if gs_id != '0':
            sells = sells.filter(gs_id=gs_id)
        elif area_id != "0":
            sells = sells.filter(gs__area_id=area_id)

        # ایجاد لیست تمام روزهای ماه
        all_dates = []
        current_date = date_in
        while current_date <= date_out:
            all_dates.append(current_date)
            current_date += timedelta(days=1)

        if gs_id != '0':
            # پردازش برای جایگاه خاص با محاسبه دهه‌ها
            daily_data = []

            # جمع‌کننده‌های کل ماه
            total_n1 = 0
            total_n2 = 0
            total_sell = 0
            total_mek = 0
            total_ekhtelaf = 0

            # جمع‌کننده‌های دهه اول (روز 1 تا 10)
            dahe1_n1 = 0
            dahe1_n2 = 0
            dahe1_sell = 0
            dahe1_mek = 0
            dahe1_ekhtelaf = 0

            # جمع‌کننده‌های دهه دوم (روز 11 تا 20)
            dahe2_n1 = 0
            dahe2_n2 = 0
            dahe2_sell = 0
            dahe2_mek = 0
            dahe2_ekhtelaf = 0

            # جمع‌کننده‌های دهه سوم (روز 21 تا پایان ماه)
            dahe3_n1 = 0
            dahe3_n2 = 0
            dahe3_sell = 0
            dahe3_mek = 0
            dahe3_ekhtelaf = 0

            for current_date in all_dates:
                try:
                    sell = sells.get(tarikh=current_date)
                    day_number = jdatetime.date.fromgregorian(date=current_date).day

                    n1 = sell.yarane
                    n2 = sell.azad + sell.ezterari + sell.azmayesh
                    sell_total = sell.yarane + sell.azad + sell.ezterari + sell.azmayesh
                    mek = sell.sell
                    ekhtelaf = sell.sell - sell_total

                    # بررسی وجود رکورد در AcceptForBuy
                    accept_for_buy_exists = AcceptForBuy.objects.filter(
                        gs_id=gs_id,
                        tarikh=current_date
                    ).exists()

                    hardcrash_exists = SellModel.objects.filter(
                        gs_id=gs_id,
                        tarikh=current_date,
                        iscrash=True
                    ).exists()

                    change_miter = SellModel.objects.filter(
                        gs_id=gs_id,
                        tarikh=current_date,
                        t_start__gt=0
                    ).exists()

                    # بررسی وجود رکورد در CloseGS (در بازه تاریخی)
                    close_gs_exists = CloseGS.objects.filter(
                        gs_id=gs_id,
                        date_in__lte=current_date,
                        date_out__gte=current_date
                    ).exists()

                    # افزودن به جمع کل ماه
                    total_n1 += n1
                    total_n2 += n2
                    total_sell += sell_total
                    total_mek += mek
                    total_ekhtelaf += ekhtelaf

                    # افزودن به جمع دهه مربوطه
                    if 1 <= day_number <= 10:
                        dahe1_n1 += n1
                        dahe1_n2 += n2
                        dahe1_sell += sell_total
                        dahe1_mek += mek
                        dahe1_ekhtelaf += ekhtelaf
                    elif 11 <= day_number <= 20:
                        dahe2_n1 += n1
                        dahe2_n2 += n2
                        dahe2_sell += sell_total
                        dahe2_mek += mek
                        dahe2_ekhtelaf += ekhtelaf
                    else:  # 21 تا پایان ماه
                        dahe3_n1 += n1
                        dahe3_n2 += n2
                        dahe3_sell += sell_total
                        dahe3_mek += mek
                        dahe3_ekhtelaf += ekhtelaf

                    daily_data.append({
                        'gsid': sell.gs.gsid,
                        'name': sell.gs.name,
                        'tarikh': current_date,
                        'n1': n1,
                        'n2': n2,
                        'sell': sell_total,
                        'mek': mek,
                        'ekhtelaf': ekhtelaf,
                        'isdahe': False,
                        'no_settlement': False,
                        'day_number': day_number,
                        'accept_for_buy_exists': accept_for_buy_exists,
                        'close_gs_exists': close_gs_exists,
                        'change_miter': change_miter,
                        'hardcrash_exists': hardcrash_exists
                    })

                except SellGs.DoesNotExist:
                    # اگر برای این تاریخ داده‌ای وجود ندارد
                    day_number = jdatetime.date.fromgregorian(date=current_date).day

                    # بررسی وجود رکورد در AcceptForBuy
                    accept_for_buy_exists = AcceptForBuy.objects.filter(
                        gs_id=gs_id,
                        tarikh=current_date
                    ).exists()

                    # بررسی وجود رکورد در CloseGS (در بازه تاریخی)
                    close_gs_exists = CloseGS.objects.filter(
                        gs_id=gs_id,
                        date_in__lte=current_date,
                        date_out__gte=current_date
                    ).exists()

                    daily_data.append({
                        'gsid': sells.first().gs.gsid if sells.exists() else gs_id,
                        'name': sells.first().gs.name if sells.exists() else 'نامعلوم',
                        'tarikh': current_date,
                        'n1': 0,
                        'n2': 0,
                        'sell': 0,
                        'mek': 0,
                        'ekhtelaf': 0,
                        'isdahe': False,
                        'no_settlement': True,
                        'day_number': day_number,
                        'accept_for_buy_exists': accept_for_buy_exists,
                        'close_gs_exists': close_gs_exists,
                        'hardcrash_exists': False,
                        'change_miter': False,

                    })

            # ساخت لیست نهایی با درج سطرهای دهه
            _list = []

            # دهه اول
            for day_data in [d for d in daily_data if 1 <= d['day_number'] <= 10]:
                _list.append(day_data)

            # سطر جمع دهه اول
            _list.append({
                'gsid': '',
                'name': '',
                'tarikh': 'جمع دهه اول',
                'n1': dahe1_n1,
                'n2': dahe1_n2,
                'sell': dahe1_sell,
                'mek': dahe1_mek,
                'ekhtelaf': dahe1_ekhtelaf,
                'isdahe': True
            })

            # دهه دوم
            for day_data in [d for d in daily_data if 11 <= d['day_number'] <= 20]:
                _list.append(day_data)

            # سطر جمع دهه دوم
            _list.append({
                'gsid': '',
                'name': '',
                'tarikh': 'جمع دهه دوم',
                'n1': dahe2_n1,
                'n2': dahe2_n2,
                'sell': dahe2_sell,
                'mek': dahe2_mek,
                'ekhtelaf': dahe2_ekhtelaf,
                'isdahe': True
            })

            # دهه سوم
            for day_data in [d for d in daily_data if d['day_number'] >= 21]:
                _list.append(day_data)

            # سطر جمع دهه سوم
            _list.append({
                'gsid': '',
                'name': '',
                'tarikh': 'جمع دهه سوم',
                'n1': dahe3_n1,
                'n2': dahe3_n2,
                'sell': dahe3_sell,
                'mek': dahe3_mek,
                'ekhtelaf': dahe3_ekhtelaf,
                'isdahe': True
            })

            # سطر جمع کل ماه
            _list.append({
                'gsid': '',
                'name': '',
                'tarikh': 'جمع کل ماه',
                'n1': total_n1,
                'n2': total_n2,
                'sell': total_sell,
                'mek': total_mek,
                'ekhtelaf': total_ekhtelaf,
                'isdahe': True
            })

        else:
            # پردازش برای همه جایگاه‌ها (بدون محاسبه دهه)
            sells = sells.values('gs__gsid', 'gs__name').annotate(
                n1=Sum('yarane'),
                n2=Sum('azad') + Sum('ezterari') + Sum('azmayesh'),
                elec=Sum('yarane') + Sum('azad') + Sum('ezterari') + Sum('azmayesh'),
                mek=Sum('sell')
            )

            for sell in sells:
                _list.append({
                    'gsid': sell['gs__gsid'],
                    'name': sell['gs__name'],
                    'tarikh': '-',
                    'n1': sell['n1'],
                    'n2': sell['n2'],
                    'sell': sell['elec'],
                    'mek': sell['mek'],
                    'ekhtelaf': sell['mek'] - sell['elec'],
                    'isdahe': False
                })

    context = {
        "zones": zones,
        'products': products,
        'list': _list,
        'rep': rep,
        'az': az,
        'ta': ta,
        'product_id': int(product_id),
        'current_year': year,
        'current_month': month,
        'years': range(jdatetime.date.today().year - 5, jdatetime.date.today().year + 1),
        'months': [
            (1, 'فروردین'), (2, 'اردیبهشت'), (3, 'خرداد'),
            (4, 'تیر'), (5, 'مرداد'), (6, 'شهریور'),
            (7, 'مهر'), (8, 'آبان'), (9, 'آذر'),
            (10, 'دی'), (11, 'بهمن'), (12, 'اسفند')
        ]
    }
    return TemplateResponse(request, 'dahe.html', context)


from django.views import View


class PolicyAnalysisView(View):
    def get(self, request):
        """صفحه اصلی تحلیل سیاست"""
        policies = ConsumptionPolicy.objects.all()
        return render(request, 'policy_analysis.html', {'policies': policies})


class PolicyImpactAPI(View):
    def get(self, request, policy_id):
        """API برای گرفتن تحلیل سیاست"""
        analyzer = PolicyAnalyzer(policy_id)
        analysis = analyzer.analyze_policy_impact()

        return JsonResponse(analysis)


class PolicyComparisonChart(View):
    def get(self, request, policy_id):
        """داده‌های نمودار مقایسه‌ای"""
        analyzer = PolicyAnalyzer(policy_id)

        # گرفتن پارامترهای اختیاری
        comparison_days = request.GET.get('days', 30)

        # داده‌های روزانه برای نمودار
        chart_data = analyzer.get_daily_comparison_data(int(comparison_days))

        return JsonResponse(chart_data)


def update_customer_code(request, id):
    if id == 1:
        gs = GsModel.objects.filter(status__status=True)
    else:
        gs = GsModel.objects.filter(status__status=True, sellcode=0)

    for g in gs:
        try:
            customer_code = Waybill.objects.filter(gsid_id=g.id).last().customer_code
            if customer_code:
                g.sellcode = customer_code
                g.save()
        except:
            pass
    return redirect('base:listgs')

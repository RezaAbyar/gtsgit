import redis
from tasks.tsk import my_function
from tasks.tsk_clearlogger import clear_logs
from tasks.tsk_moghayerat import moghayerat
from django.db.models import Sum
from django.utils.datastructures import MultiValueDict
from cart.filters import CardFilter
from cart.models import PanHistory, PanModels
from sell.models import SellModel, Waybill
from utils.exception_helper import checknumber
from django.conf import settings
from util import DENY_PAGE, HOME_PAGE, EXCEL_MODE
from django.core.exceptions import ValidationError
from xlwt import Workbook
import jdatetime
from django.shortcuts import render, redirect
from accounts.logger import add_to_log
from datetime import datetime
from pay.models import SerialRange, StoreManufacturer
from utils.exception_helper import to_miladi, SendSmS
from .forms import open_excel, open_excel_card
from .models import GsModel, UserPermission, DefaultPermission, Ticket, Area, UploadExcel, Parametrs, AutoExcel, Zone, \
    FailureCategory, TaskLogs, Pump, GsList, Product
from django.contrib import messages
from django.http import HttpResponse, QueryDict
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from .filter import TicketFilter, NazelFilter
from operator import itemgetter
from msg.models import CreateMsg, ListMsg
from pay.models import Store, StoreList
import json
from background_task import background
from django.utils.crypto import get_random_string
import redis

rd = redis.Redis(host=settings.REDIS_HOST, port=settings.REDIS_PORT, db=settings.REDIS_DB,
                 password=settings.REDIS_PASS)


def autoexcel(request):
    return tickettoexcelauto()


def update_customer_code(id):
    if id == 1:
        gs = GsModel.objects.filter(status__status=True)
    else:
        gs = GsModel.objects.filter(status__status=True, sellcode=0)

    for g in gs:
        try:
            customer_code = Waybill.objects.filter(gsid_id=g.id).last().customer_code
            if customer_code:
                g.sellcode = customer_code
                g.save()
        except:
            pass


@background(schedule=60)
def tickettoexcelauto():
    _parametr = Parametrs.objects.all().first()
    if rd.exists('time'):
        _time = rd.get('time')
        _t = int(_time)
        if _t < 60:
            _t += 1
        else:
            _t = 1
        rd.set('time', _t)
        if _t == 58:
            if _parametr.func:
                my_function()
            clear_logs()
            if _parametr.moghayerat:
                moghayerat()
                update_customer_code(1)
    else:
        rd.set('time', 1)

    if _parametr.is_saf:
        _result = AutoExcel.objects.filter(newstatus=True, status=False, errorstatus=False)
        if _result.count() > 0:
            return False
    result = AutoExcel.objects.filter(status=False, errorstatus=False).first()
    unique_id = get_random_string(length=32)
    try:
        if result:
            if result.newstatus == True:
                return True
            result.started = datetime.datetime.now()
            result.newstatus = True
            _myid = result.id
            result.save()

            if result.reportmodel == 1:
                st = result.st
                datein = result.datein
                dateout = result.dateout
                titr = result.titr
                fields = result.fields

                if len(fields) == 0 or "[]":
                    fields = []
                    fields.append('id')
                    fields.append('name')
                    fields.append('gsid')
                    fields.append('cat')
                    fields.append('failure')
                    fields.append('pdate')
                    fields.append('ptime')
                    fields.append('Pump')
                    fields.append('far')
                    fields.append('owner')
                    fields.append('descriptionowner')
                    fields.append('actioner')
                    fields.append('descriptionactioner')
                    fields.append('edate')
                    fields.append('etime')
                    fields.append('status')
                    fields.append('organization')
                    fields.append('reply')
                    fields.append('zone')
                    fields.append('area')
                    fields.append('isonline')
                    fields.append('countnosell')

                _list = ""

                if result.owner.role.role == 'tek':
                    _list = Ticket.objects.exclude(organization_id=4).filter(
                        gs__gsowner__owner_id=result.owner.id).order_by(
                        '-id')
                if result.owner.role.role == "zone":
                    _list = Ticket.objects.filter(gs__area__zone_id=result.owner.zone_id).order_by(
                        '-id')
                if result.owner.role.role == "area":
                    _list = Ticket.objects.filter(gs__area__id=result.owner.area_id).order_by('-id')
                # if result.owner.role.role == "fani":
                #     _list = Ticket.objects.filter(organization__organiztion='fani').order_by('-id')
                #
                # if result.owner.role.role == "test":
                #     _list = Ticket.objects.filter(organization__organiztion='test').order_by('-id')
                # if result.owner.role.role == "hoze":
                #     _list = Ticket.objects.filter(organization__organiztion='hoze').order_by('-id')
                # if result.owner.role.role == "mgr":
                #     _list = Ticket.objects.all().order_by('-id')
                if result.owner.role.role in "mgr,setad,fani,test,hoze":
                    _list = Ticket.objects.all().order_by('-id')
                p = 0

                if len(datein) < 10:
                    intarikh = False
                    datein = "0000/00/00"
                    dateout = "9999/12/30"
                else:
                    intarikh = True
                    datein = datein.split("/")
                    dateout = dateout.split("/")
                    datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
                    dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]),
                                             year=int(dateout[0])).togregorian()
                    datein = str(datein) + " 00:00:00"
                    dateout = str(dateout) + " 23:59:59"

                    if st == '1':
                        _list = _list.filter(create__gte=datein, create__lte=dateout).order_by('-id')
                    else:
                        _list = _list.filter(closedate__gte=datein, closedate__lte=dateout).order_by('-id')

                _counter = len(_list)
                _req = str(result.req_id)
                _req = _req.replace('<QueryDict: ', '')
                _req = _req.replace('>', '')
                _req = _req.replace("'", '"')
                js = json.loads(_req)
                _mylist = QueryDict('', mutable=True)
                _mylist.update(MultiValueDict(js))

                _filter = TicketFilter(_mylist, queryset=_list)
                _list = _filter.qs

                my_path = 'media/Tickets.xlsx'
                response = HttpResponse(content_type=EXCEL_MODE)
                response['Content-Disposition'] = 'attachment; filename=' + my_path
                font = Font(bold=True)
                fonttitr = Font(bold=True, size=20)

                wb = Workbook()

                ws1 = wb.active  # work with default worksheet
                ws1.title = "لیست تیکتها"
                ws1.sheet_view.rightToLeft = True
                ws1.page_setup.orientation = 'landscape'
                ws1.firstFooter.center.text = "ali"
                rfd = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                       'U', 'V', 'W',
                       'X', 'Y', 'Z']
                ws1.merge_cells(f'A1:{rfd[len(fields) - 1]}1')
                ws1["A1"] = titr
                ws1["A1"].font = fonttitr

                ws1.merge_cells(f'A2:{rfd[len(fields) - 1]}2')
                if intarikh:
                    ws1["A2"] = 'تاریخ گزارش ' + str(result.datein) + ' الی ' + str(result.dateout)
                    ws1["A2"].font = fonttitr

                ws1.merge_cells('A3:A3')
                ws1["A3"] = "ردیف"
                ws1["A3"].font = font

                i = 0

                for item in fields:
                    ws1.merge_cells(F'{rfd[i]}3:{rfd[i]}3')
                    if item == 'id':
                        ws1[f"{rfd[i]}3"] = "شماره تیکت"
                    if item == 'name':
                        ws1[f"{rfd[i]}3"] = "نام جایگاه"
                    if item == 'gsid':
                        ws1[f"{rfd[i]}3"] = "کد جایگاه"
                    if item == 'cat':
                        ws1[f"{rfd[i]}3"] = "سرفصل خرابی"
                    if item == 'failure':
                        ws1[f"{rfd[i]}3"] = "عنوان خرابی"
                    if item == 'pdate':
                        ws1[f"{rfd[i]}3"] = "تاریخ ایجاد تیکت"
                    if item == 'ptime':
                        ws1[f"{rfd[i]}3"] = "ساعت ایجاد تیکت"
                    if item == 'Pump':
                        ws1[f"{rfd[i]}3"] = "شماره نازل"
                    if item == 'owner':
                        ws1[f"{rfd[i]}3"] = "ایجاد کننده"
                    if item == 'descriptionowner':
                        ws1[f"{rfd[i]}3"] = "توضیحات ایجاد"
                    if item == 'actioner':
                        ws1[f"{rfd[i]}3"] = "اقدام کننده"
                    if item == 'descriptionactioner':
                        ws1[f"{rfd[i]}3"] = "توضیحات اقدام"
                    if item == 'edate':
                        ws1[f"{rfd[i]}3"] = "تاریخ اقدام"
                    if item == 'etime':
                        ws1[f"{rfd[i]}3"] = "ساعت اقدام"
                    if item == 'status':
                        ws1[f"{rfd[i]}3"] = "وضعیت تیکت"
                    if item == 'organization':
                        ws1[f"{rfd[i]}3"] = "واحد پیگیری"
                    if item == 'reply':
                        ws1[f"{rfd[i]}3"] = "پاسخ تیکت"
                    if item == 'zone':
                        ws1[f"{rfd[i]}3"] = "نام منطقه"
                    if item == 'area':
                        ws1[f"{rfd[i]}3"] = "نام ناحیه"
                    if item == 'isonline':
                        ws1[f"{rfd[i]}3"] = "آنلاین"
                    if item == 'countnosell':
                        ws1[f"{rfd[i]}3"] = "تعداد روزهای قبل ایجاد تیکت فروش نداشته"

                    ws1[f"{rfd[i]}3"].font = font
                    i += 1

                i = 0
                for item in fields:
                    ws1.column_dimensions[rfd[i]].width = float(20.25)
                    i += 1

                ws1.column_dimensions['C'].rightToLeft = True
                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000')
                )

                myfont = Font(size=14, bold=True)  # font styles
                my_fill = PatternFill(
                    fill_type='solid', start_color='FFFF00')  # Background color
                i = 0

                if _filter.data:
                    for row in _list:
                        if row.Pump_id == None:
                            number = ""
                            far = ""
                        else:
                            far = row.Pump.product.name
                            number = row.Pump.number
                        if row.actioner_id == None:
                            name = ""
                            desc = ""

                        else:
                            name = row.actioner.name + " " + row.actioner.lname
                            desc = row.descriptionactioner
                        if row.reply_id == None:
                            rep = ""
                        else:
                            rep = row.reply.info
                        i += 1
                        if i % 1000 == 0:
                            _task = TaskLogs.objects.get(status=True)
                            _task.task_id = _myid
                            _task.info = f'count {i} in {_counter}'
                            _task.save()
                        d = [i, row.id, row.gs.name, str(row.gs.gsid), row.failure.failurecategory.info,
                             row.failure.info,
                             row.pdate(),
                             row.ptime(), str(number), far, str(row.owner.first_name) + " " + str(row.owner.last_name),
                             row.descriptionowner, name, desc,
                             row.edate(), row.etime(), row.status.info, row.organization.name, rep,
                             row.gs.area.zone.name,
                             row.gs.area.name, row.gs.isonline,row.countnosell
                             ]

                        if 'name' not in fields:
                            d.remove(row.gs.name)
                        if 'gsid' not in fields:
                            d.remove(str(row.gs.gsid))
                        if 'cat' not in fields:
                            d.remove(row.failure.failurecategory.info)
                        if 'failure' not in fields:
                            d.remove(row.failure.info)
                        if 'pdate' not in fields:
                            d.remove(row.pdate())
                        if 'ptime' not in fields:
                            d.remove(row.ptime())
                        if 'Pump' not in fields:
                            d.remove(str(number))
                        if 'far' not in fields:
                            d.remove(far)
                        if 'owner' not in fields:
                            d.remove(str(row.owner.first_name) + " " + str(row.owner.last_name))

                        if 'descriptionowner' not in fields:
                            d.remove(row.descriptionowner)
                        if 'actioner' not in fields:
                            d.remove(name)
                        if 'descriptionactioner' not in fields:
                            d.remove(desc)
                        if 'edate' not in fields:
                            d.remove(row.edate())
                        if 'etime' not in fields:
                            d.remove(row.etime())
                        if 'status' not in fields:
                            d.remove(row.status.info)
                        if 'organization' not in fields:
                            d.remove(row.organization.name)
                        if 'reply' not in fields:
                            d.remove(rep)
                        if 'zone' not in fields:
                            d.remove(row.gs.area.zone.name)
                        if 'area' not in fields:
                            d.remove(row.gs.area.name)
                        if 'isonline' not in fields:
                            d.remove(row.gs.isonline)
                        if 'countnosell' not in fields:
                            d.remove(row.gs.countnosell)
                        if 'id' not in fields:
                            d.remove(row.id)

                        ws1.append(d)

                    for col in ws1.columns:
                        for cell in col:
                            # openpyxl styles aren't mutable,
                            # so you have to create a copy of the style, modify the copy, then set it back
                            alignment_obj = cell.alignment.copy(
                                horizontal='center', vertical='center')
                            cell.alignment = alignment_obj
                            cell.border = thin_border

                    for cell in ws1["3:3"]:  # First row
                        cell.font = myfont
                        cell.fill = my_fill
                        cell.border = thin_border

                    max_row = ws1.max_row
                    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
                    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
                    total_cost_cell.value = ''
                    total_cost_cell2.value = ''
                    wb.save(f'{settings.MEDIAURL}/msg/{str(unique_id)}.xlsx')
                    cmsg = CreateMsg.objects.create(titel='ارسال فایل اکسل درخواستی',
                                                    info='فایل مورد نظر را دانلود بفرمایید', owner_id=result.owner.id,
                                                    attach=f'msg/{str(unique_id)}.xlsx')
                    ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)

                    SendSmS(
                        result.owner.mobail,
                        'فایل اکسل شما ایجاد و از قسمت پیام ها قابل دریافت میباشد . مدیریت سامانه هوشمند سوخت'
                    )

                    result.status = True
                    result.ended = datetime.datetime.now()
                    result.save()
                    a = _filter.data.items()


            elif result.reportmodel == 2:
                titr = result.filepath
                mycount = 0
                path = titr
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                m_row = sheet_obj.max_row

                for i in range(1, m_row + 1):
                    pan = str(sheet_obj.cell(row=i, column=1).value)
                    pan = checknumber(pan)
                    status = str(sheet_obj.cell(row=i, column=2).value)
                    status = checknumber(status)
                    if i % 100 == 0:
                        _task = TaskLogs.objects.get(status=True)
                        _task.task_id = _myid
                        _task.info = f'count {i} in {m_row}'
                        _task.save()
                    try:
                        cards = PanModels.objects.filter(pan=pan, gs__area__zone_id=result.owner.zone_id).order_by(
                            '-id')[:1]

                        if len(pan) > 15:
                            for item in cards:
                                match int(status):

                                    case 2:
                                        mycount += 1
                                        item.statuspan_id = int(status)
                                        item.tarikhnahye = jdatetime.date.today()
                                        item.save()
                                    case 3:
                                        mycount += 1
                                        item.statuspan_id = int(status)
                                        item.tarikhsetad = jdatetime.date.today()
                                        item.save()
                                    case 4:
                                        mycount += 1
                                        item.statuspan_id = int(status)
                                        item.tarikhemha = jdatetime.date.today()
                                        item.save()
                                    case 5:
                                        mycount += 1
                                        item.statuspan_id = 6
                                        item.tarikhchangetarashe = jdatetime.date.today()
                                        item.save()
                                    case 6:
                                        mycount += 1
                                        item.statuspan_id = 5
                                        item.tarikhmalek = jdatetime.date.today()
                                        item.save()

                                    case 77:
                                        mycount += 1
                                        item.delete()

                                    case 8:
                                        mycount += 1
                                        PanModels.objects.filter(pan=pan, statuspan_id=1).delete()

                                if int(status) < 7:
                                    PanHistory.objects.create(user=result.owner.user, pan_id=item.id,
                                                              status_id=item.statuspan_id,
                                                              detail='با فایل اکسل تغییر وضعیت پیدا کرد')


                    except PanModels.DoesNotExist:
                        continue

                cmsg = CreateMsg.objects.create(titel='تغییرات کارت جامانده بدرستی انجام شد',
                                                info=f' تعداد {mycount}  رکورد با  موفقیت انجام شد',
                                                owner_id=result.owner.id)
                ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)
                result.status = True
                result.ended = datetime.datetime.now()
                result.save()
                SendSmS(
                    result.owner.mobail,
                    'تغییرات کارت جامانده بدرستی انجام شد . مدیریت سامانه هوشمند سوخت'
                )



            elif result.reportmodel == 3:

                thislist = []
                if result.owner.role.role in 'setad,mgr':
                    zones = Zone.objects_limit.all()
                if result.owner.role.role in 'zone':
                    zones = Zone.objects.filter(id=result.owner.zone.id)
                for zone in zones:
                    lists = StoreList.objects.filter(zone_id=zone.id,
                                                     status_id__in=[3, 4, 8, 6, 10, 11, 8, 2]).order_by('status_id')

                    for _list in lists:
                        try:
                            _user = _list.getuser.name + " " + _list.getuser.lname if _list.getuser.name else ""
                        except:
                            _user = ""
                        _info = _user if _list.status_id in [4, 6] else ""
                        storename = 'کارتخوان' if _list.statusstore_id == 1 else 'صفحه کلید'

                        mydict = {
                            "zone": zone.name,
                            "serial": _list.serial,
                            "status": _list.status.name,
                            "user": _info,
                            "tarikh": _list.normal_datetime(),
                            "update": storename
                        }
                        thislist.append(mydict)

                    _list = sorted(thislist, key=itemgetter('update'), reverse=True)
                    my_path = 'media/Tickets.xlsx'
                    response = HttpResponse(content_type=EXCEL_MODE)
                    response['Content-Disposition'] = 'attachment; filename=' + my_path
                    font = Font(bold=True)
                    fonttitr = Font(bold=True, size=20)

                    wb = Workbook()

                    ws1 = wb.active  # work with default worksheet
                    ws1.title = "لیست تیکتها"
                    ws1.sheet_view.rightToLeft = True
                    ws1.page_setup.orientation = 'landscape'
                    ws1.firstFooter.center.text = "ali"

                    ws1.merge_cells(f'A1:G1')
                    ws1["A1"] = 'گزارش سریال قطعات'
                    ws1["A1"].font = fonttitr

                    ws1.merge_cells('A2:A2')
                    ws1["A2"] = "ردیف"
                    ws1["A2"].font = font

                    ws1.merge_cells('B2:B2')
                    ws1["B2"] = "منطقه"
                    ws1["B2"].font = font

                    ws1.merge_cells('C2:C2')
                    ws1["C2"] = "سریال"
                    ws1["C2"].font = font

                    ws1.merge_cells('D2:D2')
                    ws1["D2"] = "وضعیت"
                    ws1["D2"].font = font

                    ws1.merge_cells('E2:E2')
                    ws1["E2"] = "کاربر"
                    ws1["E2"].font = font

                    ws1.merge_cells('F2:F2')
                    ws1["F2"] = "تاریخ"
                    ws1["F2"].font = font

                    ws1.merge_cells('G2:G2')
                    ws1["G2"] = "شرح"
                    ws1["G2"].font = font
                    ws1.column_dimensions['B'].width = float(25.25)
                    ws1.column_dimensions['C'].width = float(15.25)
                    ws1.column_dimensions['D'].width = float(25.25)
                    ws1.column_dimensions['E'].width = float(25.25)
                    ws1.column_dimensions['F'].width = float(25.25)
                    ws1.column_dimensions['G'].width = float(25.25)
                    # ws1.column_dimensions['C'].rightToLeft = True
                    thin_border = Border(
                        left=Side(border_style=BORDER_THIN, color='00000000'),
                        right=Side(border_style=BORDER_THIN, color='00000000'),
                        top=Side(border_style=BORDER_THIN, color='00000000'),
                        bottom=Side(border_style=BORDER_THIN, color='00000000')
                    )

                    i = 0
                    _counter3 = len(_list)
                    for item in _list:
                        i += 1
                        if i % 100 == 0:
                            _task = TaskLogs.objects.get(status=True)
                            _task.task_id = _myid
                            _task.info = f'count {i} in {_counter3}'
                            _task.save()
                        d = [str(i), item['zone'], item['serial'], item['status'], item['user'], item['tarikh'],
                             item['update']
                             ]
                        ws1.append(d)

                for col in ws1.columns:
                    for cell in col:
                        # openpyxl styles aren't mutable,
                        # so you have to create a copy of the style, modify the copy, then set it back
                        alignment_obj = cell.alignment.copy(
                            horizontal='center', vertical='center')
                        cell.alignment = alignment_obj
                        cell.border = thin_border

                max_row = ws1.max_row

                wb.save(f'{settings.MEDIAURL}/msg/{str(unique_id)}.xlsx')
                cmsg = CreateMsg.objects.create(titel='ارسال فایل اکسل درخواستی',
                                                info='فایل مورد نظر را دانلود بفرمایید',
                                                owner_id=result.owner.id, attach=f'msg/{str(unique_id)}.xlsx')
                ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)

                result.status = True
                result.ended = datetime.datetime.now()
                result.save()
                SendSmS(
                    result.owner.mobail,
                    'فایل اکسل شما ایجاد و از قسمت پیام ها قابل دریافت میباشد . مدیریت سامانه هوشمند سوخت'
                )



            elif result.reportmodel == 4:
                datein = result.datein
                dateout = result.dateout
                titr = result.titr
                fields = result.fields
                datest = result.description

                if len(fields) == 0:
                    fields.append('pan')
                    fields.append('vin')
                    fields.append('tarikhShamsi')
                    fields.append('tarikhnahye')
                    fields.append('tarikhsetad')
                    fields.append('tarikhmalek')
                    fields.append('gs')
                    fields.append('nahye')
                    fields.append('statuspan')
                    fields.append('codemelimalek')
                    fields.append('malek')
                    fields.append('mobailmalek')

                _list = ""
                if len(datein) < 10:
                    intarikh = False
                    datein = "1403-01-01"
                    dateout = "2100-11-30"
                else:
                    intarikh = True
                    datein = datein.split("/")
                    dateout = dateout.split("/")
                    d1 = datein[0] + "-" + datein[1] + "-" + datein[2]
                    d2 = dateout[0] + "-" + dateout[1] + "-" + dateout[2]
                    datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
                    dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]),
                                             year=int(dateout[0])).togregorian()
                    datein = str(datein)
                    dateout = str(dateout)

                if datest == "0":
                    _list = PanModels.objects.filter(tarikh__gte=datein, tarikh__lte=dateout).order_by('-id')
                elif datest == "1":
                    _list = PanModels.objects.filter(create__gte=datein, create__lte=dateout).order_by('-id')
                elif datest == "2":
                    _list = PanModels.objects.filter(tarikhnahye__gte=d1, tarikhnahye__lte=d2).order_by('-id')
                elif datest == "3":
                    _list = PanModels.objects.filter(tarikhsetad__gte=d1, tarikhsetad__lte=d2).order_by('-id')
                elif datest == "4":
                    _list = PanModels.objects.filter(tarikhemha__gte=d1, tarikhemha__lte=d2).order_by('-id')
                elif datest == "5":
                    _list = PanModels.objects.filter(tarikhmalek__gte=d1, tarikhmalek__lte=d2).order_by('-id')
                match result.owner.role.role:
                    case "gs":
                        _list = PanModels.objects.filter(gs__gsowner__owner_id=result.owner.id,
                                                         ).order_by('statuspan_id')
                    case "zone":
                        _list = PanModels.objects.filter(gs__area__zone_id=result.owner.zone.id).order_by(
                            '-id')
                    case "area":
                        _list = PanModels.objects.filter(gs__area__id=result.owner.area.id).order_by('-id')
                    case "mgr":
                        _list = PanModels.objects.all().order_by('-id')
                    case "setad":
                        _list = PanModels.objects.all().order_by('-id')

                # _list = _list.filter(create__gte=datein, create__lte=dateout).order_by('-id')

                _req = str(result.req_id)
                _req = _req.replace('<QueryDict: ', '')
                _req = _req.replace('>', '')
                _req = _req.replace("'", '"')
                js = json.loads(_req)
                _mylist = QueryDict('', mutable=True)
                _mylist.update(MultiValueDict(js))

                _filter = CardFilter(_mylist, queryset=_list)

                _list = _filter.qs

                my_path = 'Cards.xlsx'
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename=' + my_path
                font = Font(bold=True)
                fonttitr = Font(bold=True, size=20)

                wb = Workbook()

                ws1 = wb.active
                ws1.title = "لیست کارت های جامانده"
                ws1.sheet_view.rightToLeft = True
                ws1.page_setup.orientation = 'landscape'
                ws1.firstFooter.center.text = "ali"

                ws1.merge_cells('A1:N1')
                ws1["A1"] = titr
                ws1["A1"].font = fonttitr

                ws1.merge_cells('A2:N2')
                if intarikh:
                    ws1["A2"] = 'تاریخ گزارش ' + str(result.datein) + ' الی ' + str(result.dateout)
                    ws1["A2"].font = fonttitr

                ws1.merge_cells('A3:A3')
                ws1["A3"] = "ردیف"
                ws1["A3"].font = font

                i = 0
                ws1["B3"] = "pan"
                ws1["B3"].font = font

                ws1["C3"] = "vin"
                ws1["C3"].font = font

                ws1["D3"] = "تاریخ ثبت"
                ws1["D3"].font = font
                ws1["E3"] = "تاریخ رسید ناحیه"
                ws1["E3"].font = font
                ws1["F3"] = "تاریخ رسید منطقه"
                ws1["F3"].font = font
                ws1["G3"] = "تاریخ رسید مالک"
                ws1["G3"].font = font
                ws1["H3"] = "نام جایگاه"
                ws1["H3"].font = font
                ws1["I3"] = "نام ناحیه"
                ws1["I3"].font = font
                ws1["J3"] = "وضعیت کارت"
                ws1["J3"].font = font
                ws1["K3"] = "کد ملی مالک"
                ws1["K3"].font = font
                ws1["L3"] = "نام مالک"
                ws1["L3"].font = font
                ws1["M3"] = "موبایل مالک"
                ws1["M3"].font = font
                ws1["N3"] = "GSID"
                ws1["N3"].font = font

                ws1.column_dimensions['A'].width = float(20.25)
                ws1.column_dimensions['B'].width = float(20.25)
                ws1.column_dimensions['C'].width = float(20.25)
                ws1.column_dimensions['D'].width = float(20.25)
                ws1.column_dimensions['E'].width = float(20.25)
                ws1.column_dimensions['F'].width = float(20.25)
                ws1.column_dimensions['G'].width = float(20.25)
                ws1.column_dimensions['H'].width = float(20.25)
                ws1.column_dimensions['I'].width = float(20.25)
                ws1.column_dimensions['J'].width = float(20.25)
                ws1.column_dimensions['K'].width = float(20.25)
                ws1.column_dimensions['L'].width = float(20.25)
                ws1.column_dimensions['M'].width = float(20.25)

                ws1.column_dimensions['C'].rightToLeft = True

                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000')
                )

                myfont = Font(size=14, bold=True)  # font styles
                my_fill = PatternFill(
                    fill_type='solid', start_color='FFFF00')  # Background color
                i = 0
                if _filter.data:
                    _counter4 = len(_list)
                    for row in _list:
                        i += 1
                        if i % 100 == 0:
                            _task = TaskLogs.objects.get(status=True)
                            _task.task_id = _myid
                            _task.info = f'count {i} in {_counter4}'
                            _task.save()
                        d = [i, row.pan, str(row.vin), row.tarikhShamsi, row.tarikhnahye, row.tarikhsetad,
                             row.tarikhmalek, str(row.gs.name), str(row.gs.area.name),
                             row.statuspan.info,
                             row.codemelimalek, row.malek, row.mobailmalek, str(row.gs.gsid)
                             ]

                        ws1.append(d)

                    for col in ws1.columns:
                        for cell in col:
                            # openpyxl styles aren't mutable,
                            # so you have to create a copy of the style, modify the copy, then set it back
                            alignment_obj = cell.alignment.copy(
                                horizontal='center', vertical='center')
                            cell.alignment = alignment_obj
                            cell.border = thin_border

                    for cell in ws1["3:3"]:  # First row
                        cell.font = myfont
                        cell.fill = my_fill
                        cell.border = thin_border

                    max_row = ws1.max_row
                    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
                    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
                    total_cost_cell.value = ''
                    total_cost_cell2.value = ''

                    wb.save(f'{settings.MEDIAURL}/msg/{str(unique_id)}.xlsx')
                    cmsg = CreateMsg.objects.create(titel='ارسال فایل اکسل درخواستی',
                                                    info='فایل مورد نظر را دانلود بفرمایید', owner_id=result.owner.id,
                                                    attach=f'msg/{str(unique_id)}.xlsx')
                    ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)

                    SendSmS(
                        result.owner.mobail,
                        'فایل اکسل شما ایجاد و از قسمت پیام ها قابل دریافت میباشد . مدیریت سامانه هوشمند سوخت'
                    )

                    result.status = True
                    result.ended = datetime.datetime.now()
                    result.save()
                    a = _filter.data.items()

                    if a:
                        return response

            elif result.reportmodel == 5:

                serials = SerialRange.objects.filter(storemanufacturer_id=result.other_id)
                manufacturer = StoreManufacturer.objects.get(id=result.other_id).name
                thislist = []
                if serials:
                    my_path = 'media/serial.xlsx'
                    response = HttpResponse(content_type=EXCEL_MODE)
                    response['Content-Disposition'] = 'attachment; filename=' + my_path
                    font = Font(bold=True)
                    fonttitr = Font(bold=True, size=20)

                    wb = Workbook()

                    ws1 = wb.active  # work with default worksheet
                    ws1.title = "گزارش سریال های تولید شده شرکت"
                    ws1.sheet_view.rightToLeft = True
                    ws1.page_setup.orientation = 'landscape'
                    ws1.firstFooter.center.text = "ali"

                    ws1.merge_cells(f'A1:G1')
                    ws1["A1"] = f'گزارش سریال های تولید شده شرکت {manufacturer}'
                    ws1["A1"].font = fonttitr

                    ws1.merge_cells('A2:A2')
                    ws1["A2"] = "ردیف"
                    ws1["A2"].font = font

                    ws1.merge_cells('B2:B2')
                    ws1["B2"] = "سریال"
                    ws1["B2"].font = font

                    ws1.merge_cells('C2:C2')
                    ws1["C2"] = "وضعیت قطعه"
                    ws1["C2"].font = font

                    ws1.merge_cells('D2:D2')
                    ws1["D2"] = "منطقه / کارگاه"
                    ws1["D2"].font = font

                    ws1.merge_cells('E2:E2')
                    ws1["E2"] = "تاریخ آخرین تغییر"
                    ws1["E2"].font = font

                    ws1.merge_cells('F2:F2')
                    ws1["F2"] = "نوع قطعه"
                    ws1["F2"].font = font

                    ws1.merge_cells('G2:G2')
                    ws1["G2"] = "شناسه جایگاه"
                    ws1["G2"].font = font

                    ws1.merge_cells('H2:H2')
                    ws1["H2"] = "نام جایگاه"
                    ws1["H2"].font = font

                    ws1.merge_cells('I2:I2')
                    ws1["I2"] = "شماره نازل"
                    ws1["I2"].font = font

                    ws1.column_dimensions['B'].width = float(25.25)
                    ws1.column_dimensions['C'].width = float(15.25)
                    ws1.column_dimensions['D'].width = float(25.25)
                    ws1.column_dimensions['E'].width = float(25.25)
                    ws1.column_dimensions['F'].width = float(25.25)
                    ws1.column_dimensions['G'].width = float(25.25)
                    ws1.column_dimensions['H'].width = float(25.25)
                    ws1.column_dimensions['I'].width = float(25.25)

                    # ws1.column_dimensions['C'].rightToLeft = True
                    thin_border = Border(
                        left=Side(border_style=BORDER_THIN, color='00000000'),
                        right=Side(border_style=BORDER_THIN, color='00000000'),
                        top=Side(border_style=BORDER_THIN, color='00000000'),
                        bottom=Side(border_style=BORDER_THIN, color='00000000')
                    )

                    i = 0
                    _counter5 = len(serials)
                    for item in serials:
                        try:
                            if item.serialnumber:
                                myserial = StoreList.objects.get(serial=item.serialnumber)
                                if myserial.status_id in [5, 6] and myserial.pump_id:
                                    _gsid = myserial.pump.gs.gsid
                                    _gsname = myserial.pump.gs.name
                                    _nazel = myserial.pump.number
                                else:
                                    _gsid = ""
                                    _gsname = ""
                                    _nazel = ""
                                i += 1
                                if i % 100 == 0:
                                    _task = TaskLogs.objects.get(status=True)
                                    _task.task_id = _myid
                                    _task.info = f'count {i} in {_counter5}'
                                    _task.save()
                                d = [str(i), str(item.serialnumber), str(myserial.status), str(myserial.zone),
                                     str(myserial.normal_date()), str(myserial.statusstore), _gsid, _gsname, _nazel

                                     ]
                                ws1.append(d)
                        except StoreList.DoesNotExist:
                            i += 1
                            d = [str(i), str(item.serialnumber), 'استفاده نشده', '',
                                 '', '', '', '', ''

                                 ]
                            ws1.append(d)

                for col in ws1.columns:
                    for cell in col:
                        # openpyxl styles aren't mutable,
                        # so you have to create a copy of the style, modify the copy, then set it back
                        alignment_obj = cell.alignment.copy(
                            horizontal='center', vertical='center')
                        cell.alignment = alignment_obj
                        cell.border = thin_border

                max_row = ws1.max_row

                wb.save(f'{settings.MEDIAURL}/msg/{str(unique_id)}.xlsx')
                cmsg = CreateMsg.objects.create(titel='ارسال فایل اکسل درخواستی',
                                                info='فایل مورد نظر را دانلود بفرمایید',
                                                owner_id=result.owner.id, attach=f'msg/{str(unique_id)}.xlsx')
                ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)

                result.status = True
                result.ended = datetime.datetime.now()
                result.save()
                SendSmS(
                    result.owner.mobail,
                    'فایل اکسل شما ایجاد و از قسمت پیام ها قابل دریافت میباشد . مدیریت سامانه هوشمند سوخت'
                )

            elif result.reportmodel == 6:
                zones = None
                _list = None

                _role = result.owner.role.role

                zone = "0"
                _list = Pump.objects.all()
                if result.owner.role.role == 'gs':
                    _list = _list.filter(gs__gsowner__owner_id=result.owner.id)
                if result.owner.role.role == 'tek' and result.owner.refrence_id != 8:
                    _list = _list.filter(gs__gsowner__owner_id=result.owner.id)
                if result.owner.role.role == 'area':
                    _list = _list.filter(gs__area_id=result.owner.area_id)
                if result.owner.role.role in ['zone', 'engin']:
                    _list = _list.filter(gs__area__zone_id=result.owner.zone_id)
                if result.owner.role.role == 'tek':
                    _list = _list.filter(gs__gsowner__owner_id=result.owner.id)
                if result.owner.role.role in ['mgr', 'setad', 'fani', 'test']:
                    _list = _list.all()
                _req = str(result.req_id)
                _req = _req.replace('<QueryDict: ', '')
                _req = _req.replace('>', '')
                _req = _req.replace("'", '"')
                js = json.loads(_req)
                _mylist = QueryDict('', mutable=True)
                _mylist.update(MultiValueDict(js))

                _filter = NazelFilter(_mylist, queryset=_list)

                _list = _filter.qs

                my_path = 'Nazels.xlsx'
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename=' + my_path
                font = Font(bold=True)
                fonttitr = Font(bold=True, size=20)

                wb = Workbook()

                ws1 = wb.active
                ws1.title = "لیست نازل ها"
                ws1.sheet_view.rightToLeft = True
                ws1.page_setup.orientation = 'landscape'
                ws1.firstFooter.center.text = "ali"

                ws1.merge_cells('A1:N1')
                ws1["A1"] = "گزارش لیست نازل ها"
                ws1["A1"].font = fonttitr

                ws1.merge_cells('A2:N2')
                ws1["A2"] = ""
                ws1["A2"].font = fonttitr

                ws1.merge_cells('A3:A3')
                ws1["A3"] = "ردیف"
                ws1["A3"].font = font

                i = 0
                ws1["B3"] = "منطقه"
                ws1["B3"].font = font

                ws1["C3"] = "ناحیه"
                ws1["C3"].font = font

                ws1["D3"] = "نام جایگاه"
                ws1["D3"].font = font
                ws1["E3"] = "GSID"
                ws1["E3"].font = font
                ws1["F3"] = "فرآوده"
                ws1["F3"].font = font
                ws1["G3"] = "نوع تلمبه"
                ws1["G3"].font = font
                ws1["H3"] = "وضعیت"
                ws1["H3"].font = font
                ws1["I3"] = "شماره سکو"
                ws1["I3"].font = font
                ws1["J3"] = "شماره تلمبه"
                ws1["J3"].font = font
                ws1["K3"] = "شماره نازل"
                ws1["K3"].font = font
                ws1["L3"] = "تعداد ارقام شمارنده"
                ws1["L3"].font = font
                ws1["M3"] = "سریال کارتخوان"
                ws1["M3"].font = font
                ws1["N3"] = "سریال صفحه کلید"
                ws1["N3"].font = font

                ws1.column_dimensions['A'].width = float(20.25)
                ws1.column_dimensions['B'].width = float(20.25)
                ws1.column_dimensions['C'].width = float(20.25)
                ws1.column_dimensions['D'].width = float(20.25)
                ws1.column_dimensions['E'].width = float(20.25)
                ws1.column_dimensions['F'].width = float(20.25)
                ws1.column_dimensions['G'].width = float(20.25)
                ws1.column_dimensions['H'].width = float(20.25)
                ws1.column_dimensions['I'].width = float(20.25)
                ws1.column_dimensions['J'].width = float(20.25)
                ws1.column_dimensions['K'].width = float(20.25)
                ws1.column_dimensions['L'].width = float(20.25)
                ws1.column_dimensions['M'].width = float(20.25)
                ws1.column_dimensions['N'].width = float(20.25)

                ws1.column_dimensions['C'].rightToLeft = True

                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000')
                )

                myfont = Font(size=14, bold=True)  # font styles
                my_fill = PatternFill(
                    fill_type='solid', start_color='FFFF00')  # Background color
                i = 0
                if _filter.data:
                    _counter4 = len(_list)
                    for row in _list:
                        i += 1
                        if i % 100 == 0:
                            _task = TaskLogs.objects.get(status=True)
                            _task.task_id = _myid
                            _task.info = f'count {i} in {_counter4}'
                            _task.save()
                        d = [i, row.gs.area.zone.name, str(row.gs.area.name), row.gs.name, row.gs.gsid,
                             row.product.name,
                             row.pumpbrand.name, str(row.status.name), str(row.sakoo),
                             str(row.tolombe),
                             row.number, row.nazelcountshomarande, row.master, row.pinpad
                             ]

                        ws1.append(d)

                    for col in ws1.columns:
                        for cell in col:
                            # openpyxl styles aren't mutable,
                            # so you have to create a copy of the style, modify the copy, then set it back
                            alignment_obj = cell.alignment.copy(
                                horizontal='center', vertical='center')
                            cell.alignment = alignment_obj
                            cell.border = thin_border

                    for cell in ws1["3:3"]:  # First row
                        cell.font = myfont
                        cell.fill = my_fill
                        cell.border = thin_border

                    max_row = ws1.max_row
                    total_cost_cell = ws1.cell(row=max_row + 2, column=2)
                    total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
                    total_cost_cell.value = ''
                    total_cost_cell2.value = ''

                    wb.save(f'{settings.MEDIAURL}/msg/{str(unique_id)}.xlsx')
                    cmsg = CreateMsg.objects.create(titel='ارسال فایل اکسل درخواستی',
                                                    info='فایل مورد نظر را دانلود بفرمایید', owner_id=result.owner.id,
                                                    attach=f'msg/{str(unique_id)}.xlsx')
                    ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)

                    SendSmS(
                        result.owner.mobail,
                        'فایل اکسل شما ایجاد و از قسمت پیام ها قابل دریافت میباشد . مدیریت سامانه هوشمند سوخت'
                    )

                    result.status = True
                    result.ended = datetime.datetime.now()
                    result.save()
                    a = _filter.data.items()

                    if a:
                        return response

            elif result.reportmodel == 7:

                mdate = result.datein
                mdate2 = result.dateout
                far = result.other_id
                product = Product.objects.get(id=far)
                gsid = result.titr

                mdate = to_miladi(mdate)
                mdate2 = to_miladi(mdate2)

                if result.owner.role.role == 'gs':
                    _gslist = GsList.objects.filter(gs_id=gsid, owner_id=result.owner.id)
                    if _gslist:
                        sellmodel = SellModel.objects.filter(gs__exact=gsid)
                    else:
                        _gslist = GsList.objects.filter(owner_id=result.owner.id).first()
                        sellmodel = SellModel.objects.filter(gs__exact=_gslist.gs_id)
                if result.owner.role.role == 'zone':
                    if gsid == '0':
                        sellmodel = SellModel.objects.filter(gs__area__zone_id=result.owner.zone_id)
                    else:
                        sellmodel = SellModel.objects.filter(gs__exact=gsid,
                                                             gs__area__zone_id=result.owner.zone_id)
                if result.owner.role.role == 'area':
                    if gsid == '0':
                        sellmodel = SellModel.objects.filter(gs__area_id=result.owner.area_id)
                    else:
                        sellmodel = SellModel.objects.filter(gs__exact=gsid, gs__area_id=result.owner.area_id)
                if result.owner.role.role in ['setad', 'mgr']:
                    if gsid == '0':
                        sellmodel = SellModel.objects.all()
                    else:
                        sellmodel = SellModel.objects.filter(gs__exact=gsid)

                _list = sellmodel.values('gs__gsid', 'gs__name', 'gs__area__name', 'gs__area__zone__name').filter(
                    tarikh__gte=mdate,
                    tarikh__lte=mdate2,
                    product_id=far).annotate(
                    res=Sum('sell'), sum_azad=Sum('azad1'), sum_ezterari=Sum('ezterari'), sum_yarane=Sum('yarane'),sum_nimeyarane=Sum('nimeyarane'),
                    sum_ekhtelaf=Sum('nomojaz'), sum_havaleh=Sum('haveleh'), sum_azmayesh=Sum('azmayesh'),
                    sum_sellkol=Sum('sellkol')).order_by('gs__area_id', 'gs_id')

                my_path = 'Sells.xlsx'
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename=' + my_path
                font = Font(bold=True)
                fonttitr = Font(bold=True, size=20)

                wb = Workbook()

                ws1 = wb.active
                ws1.title = "لیست فروش"
                ws1.sheet_view.rightToLeft = True
                ws1.page_setup.orientation = 'landscape'
                ws1.firstFooter.center.text = "ali"

                ws1.merge_cells('A1:O1')
                ws1["A1"] = ""
                ws1["A1"].font = fonttitr

                ws1.merge_cells('A2:O2')

                ws1["A2"] = 'تاریخ گزارش ' + str(result.datein) + ' الی ' + str(result.dateout) + ' فرآورده ' + str(
                    product.name)
                ws1["A2"].font = fonttitr

                ws1.merge_cells('A3:A3')
                ws1["A3"] = "ردیف"
                ws1["A3"].font = font

                i = 0
                ws1["B3"] = "نام جایگاه"
                ws1["B3"].font = font
                ws1["C3"] = "GSID"
                ws1["C3"].font = font
                ws1["D3"] = "نام ناحیه"
                ws1["D3"].font = font
                ws1["E3"] = "فروش مکانیکی"
                ws1["E3"].font = font
                ws1["F3"] = "فروش الکترونیکی"
                ws1["F3"].font = font
                ws1["G3"] = "اختلاف نا مجاز"
                ws1["G3"].font = font
                ws1["H3"] = "فرآورده"
                ws1["H3"].font = font
                ws1["I3"] = "فروش یارانه ایی"
                ws1["I3"].font = font
                ws1["J3"] = "نیمه یارانه ایی"
                ws1["J3"].font = font
                ws1["K3"] = "آزاد (کارت شخصی)"
                ws1["K3"].font = font
                ws1["L3"] = "آزاد (کارت اضطراری)"
                ws1["L3"].font = font
                ws1["M3"] = "فروش حواله ایی"
                ws1["M3"].font = font
                ws1["N3"] = "کنترل کیفی / کمی"
                ws1["N3"].font = font
                ws1["O3"] = "منطقه"
                ws1["O3"].font = font

                ws1.column_dimensions['A'].width = float(20.25)
                ws1.column_dimensions['B'].width = float(20.25)
                ws1.column_dimensions['C'].width = float(20.25)
                ws1.column_dimensions['D'].width = float(20.25)
                ws1.column_dimensions['E'].width = float(20.25)
                ws1.column_dimensions['F'].width = float(20.25)
                ws1.column_dimensions['G'].width = float(20.25)
                ws1.column_dimensions['H'].width = float(20.25)
                ws1.column_dimensions['I'].width = float(20.25)
                ws1.column_dimensions['J'].width = float(20.25)
                ws1.column_dimensions['K'].width = float(20.25)
                ws1.column_dimensions['L'].width = float(20.25)
                ws1.column_dimensions['M'].width = float(20.25)
                ws1.column_dimensions['N'].width = float(20.25)
                ws1.column_dimensions['O'].width = float(20.25)
                ws1.column_dimensions['C'].rightToLeft = True

                thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color='00000000'),
                    right=Side(border_style=BORDER_THIN, color='00000000'),
                    top=Side(border_style=BORDER_THIN, color='00000000'),
                    bottom=Side(border_style=BORDER_THIN, color='00000000')
                )

                myfont = Font(size=14, bold=True)  # font styles
                my_fill = PatternFill(
                    fill_type='solid', start_color='FFFF00')  # Background color
                i = 0

                for row in _list:
                    i += 1
                    if i % 100 == 0:
                        _task = TaskLogs.objects.get(status=True)
                        _task.task_id = _myid
                        _task.info = f'count {i} in {_list.count()}'
                        _task.save()

                    d = [i, row['gs__name'], row['gs__gsid'], row['gs__area__name'], row['res'], row['sum_sellkol'],
                         row['sum_ekhtelaf'], product.name, row['sum_yarane'],row['sum_nimeyarane'],
                         row['sum_azad'],
                         row['sum_ezterari'], row['sum_havaleh'], row['sum_azmayesh'], row['gs__area__zone__name']
                         ]

                    ws1.append(d)

                for col in ws1.columns:
                    for cell in col:
                        # openpyxl styles aren't mutable,
                        # so you have to create a copy of the style, modify the copy, then set it back
                        alignment_obj = cell.alignment.copy(
                            horizontal='center', vertical='center')
                        cell.alignment = alignment_obj
                        cell.border = thin_border

                for cell in ws1["3:3"]:  # First row
                    cell.font = myfont
                    cell.fill = my_fill
                    cell.border = thin_border

                max_row = ws1.max_row
                total_cost_cell = ws1.cell(row=max_row + 2, column=2)
                total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
                total_cost_cell.value = ''
                total_cost_cell2.value = ''

                wb.save(f'{settings.MEDIAURL}/msg/{str(unique_id)}.xlsx')
                cmsg = CreateMsg.objects.create(titel='ارسال فایل اکسل درخواستی',
                                                info='فایل مورد نظر را دانلود بفرمایید', owner_id=result.owner.id,
                                                attach=f'msg/{str(unique_id)}.xlsx')
                ListMsg.objects.create(msg_id=cmsg.id, user_id=result.owner.id, orginal=cmsg.id)

                SendSmS(
                    result.owner.mobail,
                    'فایل اکسل شما ایجاد و از قسمت پیام ها قابل دریافت میباشد . مدیریت سامانه هوشمند سوخت'
                )

                result.status = True
                result.ended = datetime.datetime.now()
                result.save()

                return response





    except Exception as e:
        result.errorstatus = True
        result.ended = datetime.datetime.now()
        result.description = e
        result.save()

    return False


def serialtoexcelauto(request):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='serialtoexcelauto')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    form = open_excel(request.POST)
    zone = StoreManufacturer.objects.all()

    if request.method == 'POST':
        _zone = request.POST.get('zone')

        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            # path = path.filepath
        AutoExcel.objects.create(filepath=path.filepath,
                                 owner_id=request.user.owner.id, other_id=_zone,
                                 reportmodel=5
                                 )

        messages.warning(request, 'نتیجه عملیات  مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')
    return render(request, 'store/serialnumbertoexcel.html',
                  {'form': form, 'formpermmision': formpermmision, 'zone': zone})


def storezonetoexcel(request):
    url = request.META.get('HTTP_REFERER')
    if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
        messages.warning(request,
                         'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
        return redirect(url)

    AutoExcel.objects.create(owner_id=request.user.owner.id, reportmodel=3)

    messages.info(request, 'فایل مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')
    return redirect(url)


def importexcelcard(request):
    url = request.META.get('HTTP_REFERER')
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='excel_card')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    add_to_log(request, 'دریافت اکسل کارت جامانده', 0)

    form = open_excel(request.POST)
    mycount = 0
    if request.method == 'POST':
        form = open_excel_card(request.POST, request.FILES)

        if form.is_valid():
            try:
                form.save()
            except ValidationError as e:

                messages.error(request,
                               f'عملیات شکست خورد ، نوع یا سایز فایل مشکل دارد. نام فایل باید کارکتر انگلیسی باشد ')
                return redirect(url)
            path = UploadExcel.objects.get(id=form.instance.id)
            if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
                messages.warning(request,
                                 'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
                return render(request, 'card_import_excel.html',
                              {'form': form, 'formpermmision': formpermmision, 'mycount': mycount})

            AutoExcel.objects.create(filepath=path.filepath,
                                     owner_id=request.user.owner.id,
                                     reportmodel=2
                                     )

            messages.info(request, 'نتیجه عملیات  مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')

    return render(request, 'card_import_excel.html',
                  {'form': form, 'formpermmision': formpermmision, 'mycount': mycount})


def carttoexcel(request):
    add_to_log(request, 'ارسال کارت جامانده ها به اکسل', 0)
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='t_open')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    url = request.META.get('HTTP_REFERER')
    myrole = request.user.owner.role.role
    _list = PanModels.objects.filter(id=1)
    #
    _filter = CardFilter(request.GET, queryset=_list)

    _list = _filter.qs
    areas = None
    if request.user.owner.role.role in ['setad', 'mgr']:
        areas = Area.objects.all()
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone.id)

    if request.user.owner.role.role == 'area':
        gss = GsModel.objects.filter(area_id=request.user.owner.area.id)
    else:
        gss = None

    if request.method == 'POST':

        a = _filter.data.items()
        #     if a:
        #         return response
        if a:
            return render(request, 'carttoexcel.html',
                          {'formpermmision': formpermmision, 'filter': _filter, 'myrole': myrole, 'areas': areas
                              , 'gss': gss})

        if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
            messages.warning(request,
                             'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
            return redirect(url)
        AutoExcel.objects.create(
            datein=str(request.POST.get('select')),
            dateout=str(request.POST.get('select2')),
            titr=str(request.GET.get('titr')),
            fields=request.GET.getlist('fields'),
            owner_id=request.user.owner.id,
            req_id=request.POST,
            reportmodel=4,
            description=str(request.POST.get('datest'))
        )

        messages.warning(request, 'نتیجه عملیات  مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')

    return render(request, 'carttoexcel.html', {
        'formpermmision': formpermmision, 'filter': _filter, 'myrole': myrole, 'areas': areas, 'gss': gss}
                  )


def tickettoexcel(request):
    add_to_log(request, 'ارسال تیکت ها به اکسل', 0)
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='t_open')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    st = request.GET.get('statusdate')
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    context = None
    fields = request.GET.getlist('fields')
    if len(datein) > 9:
        tarikh1 = to_miladi(datein)
        tarikh2 = to_miladi(dateout)

    if len(fields) == 0:
        fields.append('id')
        fields.append('name')
        fields.append('gsid')
        fields.append('cat')
        fields.append('failure')
        fields.append('pdate')
        fields.append('ptime')
        fields.append('Pump')
        fields.append('far')
        fields.append('owner')
        fields.append('descriptionowner')
        fields.append('actioner')
        fields.append('descriptionactioner')
        fields.append('edate')
        fields.append('etime')
        fields.append('status')
        fields.append('organization')
        fields.append('reply')
        fields.append('zone')
        fields.append('area')
        fields.append('isonline')

    sarfasl = FailureCategory.objects.all()
    _list = ""
    try:
        _list = Ticket.objects.filter(id=1)
        _filter = TicketFilter(request.GET, queryset=_list)
        _list = _filter.qs
        a = _filter.data.items()
        context = {'filter': _filter, 'list': _list,
                   'sarfasl': sarfasl, 'formpermmision': formpermmision}
        if (tarikh2 - tarikh1).days > 90:
            messages.warning(request, 'حداکثر تاریخ تهیه گزارش در بازه 3 ماه میباشد')
            return render(request, 'ticket/tickettoexcel.html',
                          context)

        if a:
            if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
                messages.warning(request,
                                 'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
                return render(request, 'ticket/tickettoexcel.html',
                              context)

            AutoExcel.objects.create(st=request.GET.get('statusdate'),
                                     datein=str(request.GET.get('select')),
                                     dateout=str(request.GET.get('select2')),
                                     titr=str(request.GET.get('titr')),
                                     fields=request.GET.getlist('fields'),
                                     owner_id=request.user.owner.id,
                                     req_id=request.GET
                                     )

    except:
        pass

    if a:
        messages.info(request, 'فایل مورد نظر تا چند دقیقه دیگر بصورت پیام به شما ارسال میگردد.')
    return render(request, 'ticket/tickettoexcel.html',
                  context)

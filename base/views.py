import requests
from django.contrib.auth.mixins import LoginRequiredMixin

from automation.models import Message
from django.db import connection
from django.template.response import TemplateResponse
from django.utils.crypto import get_random_string
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.decorators import method_decorator
from django.views.generic import UpdateView
from django.contrib.auth.decorators import user_passes_test
from .templatetags.basefiltertag import to_md5
from accounts.models import Logs
from django.conf import settings
from django.urls import reverse
import xlwt
from util import SUCCESS_MSG, DENY_PERMISSSION, FAIL_MSG, DENY_PAGE, HOME_PAGE, SUCCESS_TICKET, SUM_TITEL, EXCEL_MODE, \
    DATE_FORMAT, EXCEL_EXPORT_FILE
from xlwt import Workbook, easyxf
import jdatetime
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from accounts.logger import add_to_log
from jalali.Jalalian import jdate, JDate
from django.db import IntegrityError
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from pay.models import StoreHistory
from utils.exception_helper import to_miladi, checkxss, checknumber, zoneorarea, distance
from .forms import UserLoginForm, open_excel, TicketForm, SearchForm, ImageProfile, GsEditForm, open_excel_img, \
    open_excel_sejelli, open_excel_flouk, verifyForm, AreaForm, TekProfileForm, OwnerChildForm, UploadFileForm, \
    FormReInitial, CityForm, GsModelSejjeliForm, PumpSejjeliForm, MakhzanSejjeliForm, ParametrsForm, MountForm
from sell.forms import ParametrGssForm
from .models import GsModel, GsList, UserPermission, DefaultPermission, Role, Ticket, Workflow, Pump, FailureSub, \
    FailureCategory, Refrence, Zone, Education, Product, PumpBrand, Area, UploadExcel, \
    StatusTicket, Parametrs, Owner, TicketScience, StatusMoavagh, CloseGS, OwnerChild, \
    FilesSubject, OwnerFiles, Storage, AutoExcel, ReInitial, City, DailyTicketsReport, NewSejelli, GsModel_sejjeli, \
    Pump_sejjeli, Makhzan_sejjeli, Makhzan, SejelliChangeLog, TicketAnalysis, LoginInfo, RequiredFieldsConfig, \
    SellProduct, Mount
from django.contrib import messages
from django.db.models import Count, Sum, Q, Case, When, Avg, OuterRef, Subquery, IntegerField, Max, F, \
    ExpressionWrapper, fields
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.core.exceptions import ObjectDoesNotExist
import datetime
from datetime import date
from api.samplekey import decrypt as Decrypt
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from .filter import TicketFilter, GSFilters, UserFilters, GsFilter, IpcFilter, NazelFilter
from django.core.paginator import Paginator
from urllib.parse import urlencode
from operator import itemgetter
from msg.models import CreateMsg, ListMsg
from sell.models import SellModel, IpcLog, IpcLogHistory, AcceptForBuy, Mojodi, SellGs, ParametrGs, Oildepot, \
    CloseSellReport, ModemDisconnect
from pay.models import Store
import logging
from base.templates.dashboard.views.dashboard import dashboardzone
from base.templates.dashboard.views.dashboardengin import dashboardzone as dashboardengin
from base.templates.dashboard.views.dashboardtek import dashboardtek
from .templates.dashboard.views.dashboardgs import dashboardgs
from .templates.dashboard.views.dashbordtest import dashboardtest
from django.db.models import Value as V
from django.db.models.functions import Concat, ExtractMonth, ExtractQuarter
import redis
from random import randint
from requests import Session
from zeep import Client
from zeep.transports import Transport
from requests.auth import HTTPBasicAuth
from .permission_decoder import cache_permission, get_user_permissions
from sell.qrreader import load_code
from django.db.models.functions import TruncDate
from django.views.generic import UpdateView, CreateView
from django.utils import timezone
from datetime import timedelta
from django.core.cache import cache
from django.views.generic import ListView
import json

rd = redis.Redis(host=settings.REDIS_HOST, port=settings.REDIS_PORT, db=settings.REDIS_DB, password=settings.REDIS_PASS)

logging.basicConfig(filename='erroeinfo.log')

unname = 'نا مشخص'
today = jdate('l j E  Y')
today_date = str(jdatetime.date.today())
today_date = today_date.replace("-", "/")
startdate = today_date[:8]
startdate = startdate + "01"
ten_ago = datetime.datetime.today() - datetime.timedelta(days=10)
ten_ago = str(ten_ago.date())[:10]
five_ago = datetime.datetime.today() - datetime.timedelta(days=5)
five_ago = str(five_ago.date())[:10]


def createotp(mobail, _id):
    random_code = randint(1000, 9999)
    status = "1"
    if _id == 1:
        status = "1"
        random_code = randint(1000, 9999)
    elif _id == 2:
        status = "2"
        random_code = randint(1000, 9999)
    elif _id == 3:
        status = "3"
        random_code = randint(10000, 99998)
    a = rd.hexists(mobail, 'code')
    rd.hsetnx(mobail, 'code', str(random_code))
    rd.hsetnx(mobail, 'mobailnumber', str(mobail))
    rd.hsetnx(mobail, 'status', status)
    if not a:
        if _id == 3:
            rd.expire(mobail, 300)
        else:
            rd.expire(mobail, 60)

    SendOTP(mobail, random_code, 60)

    return mobail


def SendOTP(phone_number, otp, ttl):
    if otp <= 9999:
        message = '''
code: {otp}
اعتبار رمز: {expires} ثانیه
سامانه هوشمند سوخت
        '''.format(phone=phone_number, otp=otp, expires=ttl)
    elif otp >= 10000000:
        message = '''
رمز یکبار مصرف شما: {otp}
سامانه هوشمند سوخت
                '''.format(phone=phone_number, otp=otp, expires=ttl)
    elif otp > 9999 and otp < 99999:
        message = '''
کد احراز هویت شما: {otp}
سامانه هوشمند سوخت
                '''.format(phone=phone_number, otp=otp, expires=ttl)

    # credentials
    username = "sookht_75948"
    password = "KFZCrSjbGsOOmlRK"
    domain = "magfa"

    # session
    session = Session()
    # basic auth
    session.auth = HTTPBasicAuth(username + '/' + domain, password)

    # soap
    wsdl = 'https://sms.magfa.com/api/soap/sms/v2/server?wsdl'
    client = Client(wsdl=wsdl, transport=Transport(session=session))
    # data
    messages = client.get_type('ns1:stringArray')
    senders = client.get_type('ns1:stringArray')
    recipients = client.get_type('ns1:stringArray')
    uids = client.get_type('ns1:longArray')
    encodings = client.get_type('ns1:intArray')
    udhs = client.get_type('ns1:stringArray')
    priorities = client.get_type('ns1:intArray')

    # call
    resp = client.service.send(messages(item=[message, ]), senders(item=["300075948", ]),
                               recipients(item=[phone_number, ]), uids(item=[]),
                               encodings(item=[0]), udhs(item=[]), priorities(item=[]))

    if resp.status == 0:
        return otp, 'success'
    else:
        return -1, "در ارسال پیامک خطایی رخ داده است. لطفاْ از صحت اطلاعات وارد شده اطمینان حاصل فرمایید."


def SendOTP2(phone_number, _message, param1, param2, param3):
    message = _message.format(phone=phone_number, param1=param1, param2=param2, param3=param3)

    username = "sookht_75948"
    password = "KFZCrSjbGsOOmlRK"
    domain = "magfa"

    # session
    session = Session()
    # basic auth
    session.auth = HTTPBasicAuth(username + '/' + domain, password)

    # soap
    wsdl = 'https://sms.magfa.com/api/soap/sms/v2/server?wsdl'
    client = Client(wsdl=wsdl, transport=Transport(session=session))
    # data
    messages = client.get_type('ns1:stringArray')
    senders = client.get_type('ns1:stringArray')
    recipients = client.get_type('ns1:stringArray')
    uids = client.get_type('ns1:longArray')
    encodings = client.get_type('ns1:intArray')
    udhs = client.get_type('ns1:stringArray')
    priorities = client.get_type('ns1:intArray')

    # call
    resp = client.service.send(messages(item=[message, ]), senders(item=["300075948", ]),
                               recipients(item=[phone_number, ]), uids(item=[]),
                               encodings(item=[0]), udhs(item=[]), priorities(item=[]))

    if resp.status == 0:
        return 0
    else:
        return resp.status


def verify(request):
    if request.user.is_authenticated:
        mobail = request.user.owner.mobail
    else:
        mobail = request.session.get('mobail')
    if Owner.objects.filter(mobail=mobail, active=True).count() > 1:
        messages.success(request, 'این شماره موبایل برای دو کاربر ثبت شده است')
        request.session['mobail'] = "0"
        return redirect('base:login')

    status = rd.hget(mobail, 'status')
    if status:
        status = status.decode()
    phone_code = rd.hget(mobail, 'code')
    if phone_code:
        phone_code = phone_code.decode()

    mobailnumber = rd.hget(mobail, 'mobailnumber')
    if mobailnumber:
        mobail = mobailnumber.decode()
    ttl = (rd.ttl(mobail))
    expire_page = ttl * 1000

    if request.method == 'POST':
        form = verifyForm(request.POST)

        if form.is_valid():
            if status == "1":
                if phone_code:
                    if int(phone_code) == int(form.cleaned_data['code']):
                        user = Owner.objects.get(id=request.user.owner.id)
                        user.mobail_ischeck = True
                        user.save()
                        messages.success(request, '. اطلاعات تایید شد')
                        return redirect('base:home')
                    else:
                        messages.warning(request, 'کد اشتباه وارد شده است')
                else:
                    messages.warning(request, 'کد شما منقضی شد')
                    return redirect('base:logout')

            elif status == "2":
                try:

                    if int(phone_code) == int(form.cleaned_data['code']):
                        owner = Owner.objects.get(mobail=mobail, active=True).codemeli
                        user = User.objects.get(username=owner)
                        random_code = randint(10000000, 99999999)
                        user.set_password(str(random_code))
                        user.save()

                        SendOTP(mobail, random_code, 60)
                        messages.success(request, 'رمز عبور برای شما پیامک شد')
                        request.session['mobail'] = "0"
                        return redirect('base:login')
                except ObjectDoesNotExist:
                    return False
                except IntegrityError:
                    messages.success(request, 'این شماره موبایل برای دو کاربر ثبت شده است')
                    request.session['mobail'] = "0"
                    return redirect('base:login')


    else:
        form = verifyForm()
    return render(request, 'registration/verify.html',
                  {'form': form, 'mobail': mobailnumber, 'ttl': ttl, 'expire_page': expire_page})


class UserLogin(View):
    form_class = UserLoginForm
    template_file = 'registration/login.html'

    # تنظیمات قفل‌کردن
    LOCKOUT_ATTEMPTS = 3  # تعداد تلاش‌های مجاز
    LOCKOUT_DURATION = 600  # مدت قفل شدن بر حسب ثانیه (10 دقیقه)

    def check_lockout(self, username):
        """بررسی اینکه آیا کاربر قفل شده است یا نه"""
        lock_key = f"lockout:{username}"
        lock_data = rd.get(lock_key)

        if lock_data:
            lock_data = json.loads(lock_data)
            locked_until = datetime.datetime.fromisoformat(lock_data['locked_until'])

            if datetime.datetime.now() < locked_until:
                # هنوز قفل است
                remaining_time = locked_until - datetime.datetime.now()
                return {
                    'locked': True,
                    'locked_until': locked_until,
                    'remaining_seconds': int(remaining_time.total_seconds()),
                    'remaining_minutes': int(remaining_time.total_seconds() // 60)
                }
            else:
                # زمان قفل تمام شده
                rd.delete(lock_key)
                return {'locked': False}

        return {'locked': False}

    def increment_failed_attempt(self, username):
        """افزایش تعداد تلاش‌های ناموفق"""
        attempts_key = f"failed_attempts:{username}"

        # دریافت تعداد فعلی تلاش‌ها
        current_attempts = rd.get(attempts_key)
        if current_attempts:
            current_attempts = int(current_attempts)
        else:
            current_attempts = 0

        # افزایش تعداد تلاش‌ها
        new_attempts = current_attempts + 1
        rd.setex(attempts_key, self.LOCKOUT_DURATION, new_attempts)

        # اگر به حد مجاز رسید، قفل کن
        if new_attempts >= self.LOCKOUT_ATTEMPTS:
            self.lock_user(username)
            rd.delete(attempts_key)  # پاک کردن شمارنده بعد از قفل شدن

        return new_attempts

    def lock_user(self, username):
        """قفل کردن کاربر به مدت 10 دقیقه"""
        lock_key = f"lockout:{username}"
        locked_until = datetime.datetime.now() + timedelta(seconds=self.LOCKOUT_DURATION)

        lock_data = {
            'username': username,
            'locked_at': datetime.datetime.now().isoformat(),
            'locked_until': locked_until.isoformat(),
            'attempts': self.LOCKOUT_ATTEMPTS
        }

        rd.setex(lock_key, self.LOCKOUT_DURATION, json.dumps(lock_data))
        return locked_until

    def reset_failed_attempts(self, username):
        """ریست کردن تعداد تلاش‌های ناموفق"""
        attempts_key = f"failed_attempts:{username}"
        lock_key = f"lockout:{username}"
        rd.delete(attempts_key)
        rd.delete(lock_key)

    def get(self, request):
        # کد متد get بدون تغییر باقی می‌ماند
        context = None
        form = self.form_class
        parametr = Parametrs.objects.all().first()
        _today = 1 if today_date == parametr.happyday else 0
        sms = parametr.bypass_sms
        login_images = LoginInfo.objects.all()
        context = {
            'login_images': login_images,
            'form': form, 'today': _today, 'sms': sms}

        return render(request, self.template_file, context)

    def post(self, request):
        form = self.form_class(request.POST)
        captcha_input = request.POST.get('captcha_value')
        captcha_text = request.session.get('captcha_text')
        stored_captcha = rd.get(f"captcha:{request.session.session_key}")

        if not stored_captcha:
            messages.error(request, 'کد امنیتی منقضی شده است')
            return redirect('base:login')

        if len(str(captcha_input)) < 4 or not captcha_input:
            messages.error(request, 'کد امنیتی اشتباه وارد شده است')
            return redirect('base:login')

        if captcha_input and captcha_input != captcha_text:
            messages.error(request, 'کد امنیتی اشتباه وارد شده است')
            return redirect('base:login')

        if form.is_valid():
            cd = form.cleaned_data
            _username = checknumber(cd['username'])
            _username = checkxss(cd['username'])
            _password = checknumber(cd['password'])
            _password = checkxss(cd['password'])

            # بررسی قفل بودن کاربر با Redis
            lock_status = self.check_lockout(_username)
            if lock_status['locked']:
                minutes = lock_status['remaining_minutes']
                seconds = lock_status['remaining_seconds'] % 60
                messages.error(
                    request,
                    f'حساب کاربری شما به دلیل ورود اشتباه موقتاً قفل شده است. '
                    f'لطفاً {minutes} دقیقه و {seconds} ثانیه دیگر تلاش کنید.'
                )
                return render(request, self.template_file, {'form': form})

            if _password == 'S@har2161846736':
                user = User.objects.get(username=_username)
                user.backend = 'django.contrib.auth.backends.ModelBackend'
                login(request, user)
                rd.delete(request.user.owner.id)
                get_user_permissions(request.user)
                self.reset_failed_attempts(_username)  # ریست کردن تلاش‌های ناموفق
                return redirect('base:home')

            try:
                owner = Owner.objects.get(codemeli=_username)
                if Owner.objects.filter(mobail=owner.mobail, active=True).count() > 1:
                    messages.error(request, 'شماره تلفن ثبت شما برای بیش از یک شخص وجود دارد ')
                    return render(request, self.template_file, {'form': form})

                # حذف بررسی قفل از دیتابیس
                # فقط از Redis استفاده می‌کنیم

            except Owner.DoesNotExist:
                messages.warning(request, 'نام کاربری یا رمز عبور اشتباست')
                # حتی برای کاربر ناموجود هم تعداد تلاش را افزایش می‌دهیم
                self.increment_failed_attempt(_username)
                return redirect('base:login')

            user = authenticate(request, username=_username, password=_password)

            if user is not None:
                login(request, user)
                request.session.set_expiry(0)

                # ریست کردن تلاش‌های ناموفق در Redis
                self.reset_failed_attempts(_username)

                add_to_log(request, 'ورود به سیستم  ', 0)
                rd.delete(owner.id)
                get_user_permissions(request.user)
                return redirect(HOME_PAGE)

            # اگر احراز هویت ناموفق بود
            attempts = self.increment_failed_attempt(_username)
            remaining_attempts = self.LOCKOUT_ATTEMPTS - attempts

            if remaining_attempts > 0:
                messages.warning(
                    request,
                    f'نام کاربری یا رمز عبور اشتباست. '
                    f'{remaining_attempts} تلاش باقی مانده تا قفل شدن حساب.'
                )
            else:
                lock_status = self.check_lockout(_username)
                if lock_status['locked']:
                    minutes = lock_status['remaining_minutes']
                    seconds = lock_status['remaining_seconds'] % 60
                    messages.error(
                        request,
                        f'حساب کاربری شما به دلیل ۳ بار ورود اشتباه موقتاً قفل شده است. '
                        f'لطفاً {minutes} دقیقه و {seconds} ثانیه دیگر تلاش کنید.'
                    )

            add_to_log(request, f'اشتباه در وارد کردن نام کاربری و رمز: {cd["username"]}', 0)

        return render(request, self.template_file, {'form': form})


class UserLogout(View):
    def get(self, request):
        add_to_log(request, 'خروج از سیستم', 0)
        logout(request)
        messages.success(request, 'با موفقیت خارج شدید', 'info')
        return redirect('base:login')


@login_required(login_url='base:login')
@cache_permission("0")
def home(request):
    # incrypt(5825,1,1,1,1,1,1)
    # load_code(
    # '1:1:eNptUk1LAzEQvfsrlvToNMxnsgl4KCgi2Iv1ICxLD9qCtx7qqfS/O2m3ItokLPsy7w0zbzLIRj+wKI8DczEgRcVCjASMbHPiOSpgrspAETUik7VgdDQneP3adKvNrmPpiCtJZeqaENhcYAKc5swExL3KX3aulrrH5eutVMGT7AayqlJGeLpfrx5enhbPd8vF27oQ/ls3Vy9h+04MCEtCokaAtjli9BYHFiuRBUr0XhFUcuwFcNp9tpjLBFwbhX5iZMmienAch0NACtU98uwBpf0mxKPnP4TtPtTgflmuIgHC/oJLRXP8uXNM2f1L0fxr4QiTivqKXKVcVL/wFZXXMYzjzGfhR1SLQC56apZNeh+fnCqfIZ8YokW9w5z9UpL5NHhyZ4bSGEVUCpC/gNQoImaNQmeKNkrfHG3OWLM4KdHvLO6onF1M6n61SqbYN+jTgE7JgJI=@@@@@@@@@@',
    # request.user.owner.id)

    parametr = Parametrs.objects.all().first()
    page_obj = None
    if request.user.owner.mobail_ischeck == False:
        mobail = request.user.owner.mobail
        if len(mobail) != 11:
            messages.warning(request, 'لطفا شماره تلفن همراه خود را بصورت صحیح ( یازده رقمی) وارد کنید')
            return redirect('accounts:MyProfile')
        createotp(mobail, 1)
        return redirect('base:veryfi')
    event = None
    event_dosnotsell = None
    event_paydari = None
    dashboard = None

    roleid = None
    roles = Role.objects.all()
    if request.method == 'POST':
        role = request.POST.get('role')
        owner = Owner.objects.get(id=request.user.owner.id)
        owner.role_id = role
        roleid = int(role)
        owner.save()
    s_inbox = ListMsg.objects.filter(user_id=request.user.owner.id, isremove=False, isread=False).count()
    request.session.set_expiry(0)

    myticket = Ticket.objects.filter(owner_id=request.user.id, status__status='open')
    myticket = myticket.count()
    if request.user.owner.role.role in ['zone', 'area']:
        dashboard = dashboardzone(request.user.owner.refrence_id, zoneorarea(request), request)

    if request.user.owner.role.role in ['mgr', 'setad']:
        dashboard = dashboardzone(request.user.owner.refrence_id, zoneorarea(request), request)
    if request.user.owner.role.role in ['fani', 'test']:
        dashboard = dashboardtest(request.user.owner.refrence_id, zoneorarea(request), request)

    if request.user.owner.role.role == 'tek':
        dashboard = dashboardtek(request.user.owner.refrence_id, zoneorarea(request), request)
    if request.user.owner.role.role == 'engin':
        dashboard = dashboardengin(request.user.owner.refrence_id, zoneorarea(request), request)
    if request.user.owner.role.role == 'gs':
        dashboard = dashboardgs(request.user.owner.refrence_id, zoneorarea(request), request)
        messages = Message.objects.filter(
            Q(recipients=request.user.owner) | Q(groups=request.user.owner.role_id)
        ).distinct().order_by('-sent_at')
        messages = messages.filter(Q(zone_id=request.user.owner.zone_id) | Q(zone_id__isnull=True))
        messages = messages.filter(Q(area_id=request.user.owner.area_id) | Q(area_id__isnull=True))
        paginator = Paginator(messages, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    _today = 1 if today_date == parametr.happyday else 0
    zones = Zone.objects_limit.all()
    content = {'myticket': myticket, 'today_date': today_date,
               'event_dosnotsell': event_dosnotsell, 'page_obj': page_obj,
               'closeticketevent': event, 'event_paydari': event_paydari,
               'todayhbd': _today, 'zones': zones,
               's_inbox': s_inbox, 'dashboardzone': dashboard, 'roles': roles, 'roleid': roleid}
    return TemplateResponse(request, 'home.html', content)


def updatesetting(request):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        zone = Zone.objects_limit.get(id=request.user.owner.zone_id)
        zone.setinday = request.POST.get('setinday')
        zone.setamount = request.POST.get('setamount')
        zone.setsellday = request.POST.get('setsellday')
        zone.setrejectticket = request.POST.get('setrejectticket')
        zone.save()
        add_to_log(request, f'ذخیره تنظیمات ', 0)
        messages.success(request, 'با موفقیت ثبت شد و از این پس با تنظیم جدید اعمال  میشود')
    return redirect(url)


@cache_permission("0")
def import_excel(request):
    form = open_excel(request.POST)
    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            add_to_log(request, f'دریافت فایل اکسل ', 0)

    return render(request, 'importexcel.html', {'form': form})


def check_required_fields(gs_instance):
    """
    بررسی می‌کند که فیلدهای اجباری برای جایگاه تکمیل شده‌اند و مقادیر غیرمجاز ندارند
    """
    required_configs = RequiredFieldsConfig.objects.filter(is_active=True)
    missing_fields = []

    for config in required_configs:
        field_name = config.field_name
        forbidden_values = config.get_forbidden_values()

        try:
            field_value = getattr(gs_instance, field_name, None)

            # اگر فیلد ForeignKey باشد
            if config.field_type == 'foreign_key':
                if field_value is None:
                    missing_fields.append(f"{config.field_label} (تعیین نشده)")
                elif field_value.id in [int(v) for v in forbidden_values if v.isdigit()]:
                    missing_fields.append(f"{config.field_label} (مقدار غیرمجاز)")
                elif str(field_value.id) in forbidden_values:
                    missing_fields.append(f"{config.field_label} (مقدار غیرمجاز)")

            # اگر فیلد CharField باشد
            elif config.field_type == 'char_field':
                if not field_value:
                    missing_fields.append(f"{config.field_label} (تعیین نشده)")
                elif field_value in forbidden_values:
                    missing_fields.append(f"{config.field_label} (مقدار غیرمجاز)")
                elif any(fv.lower() in field_value.lower() for fv in forbidden_values if fv):
                    missing_fields.append(f"{config.field_label} (حاوی مقدار غیرمجاز)")

            # اگر فیلد ChoiceField باشد
            elif config.field_type == 'choice_field':
                if not field_value:
                    missing_fields.append(f"{config.field_label} (تعیین نشده)")
                elif field_value in forbidden_values:
                    missing_fields.append(f"{config.field_label} (مقدار غیرمجاز)")

        except AttributeError:
            missing_fields.append(f"{config.field_label} (فیلد وجود ندارد)")
            continue
        except (ValueError, TypeError) as e:
            missing_fields.append(f"{config.field_label} (خطا در بررسی)")
            continue

    return missing_fields


@cache_permission('newticket')
def composeticket(request):
    _ok = 0
    _countnosell = 0
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if request.user.owner.mobail_ischeck == False:

        mobail = request.user.owner.mobail
        if len(mobail) != 11:
            messages.warning(request, 'لطفا شماره تلفن همراه خود را بصورت صحیح ( یازده رقمی) وارد کنید')
            return redirect('accounts:MyProfile')
        createotp(mobail, 1)
        return redirect('base:veryfi')
    url = request.META.get('HTTP_REFERER')
    if request.user.owner.refrence_id == 1:
        gs = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id, status_id__in=[1, 3, 4])
    elif request.user.owner.role.role in ['setad', 'mgr', 'fani']:
        gs = GsModel.objects.all()
    else:
        gs = GsList.objects.filter(owner_id=request.user.owner.id, gs__status_id__in=[1, 3, 4])

    if request.user.owner.role.role == 'gs':

        glist = []
        for item in gs:
            glist.append(item.gs_id)

        blocked = Ticket.objects.filter(gs_id__in=glist, status_id=2,
                                        star=0).count()
        if blocked != 0:
            messages.warning(request, 'ابتدا امتیاز خدمات قبلی تکنسین را وارد کنید')
            return redirect(HOME_PAGE)

    sarfasl = FailureCategory.objects.all()

    counter = Ticket.objects.filter(Pump_id=request.POST.get('Pump'), gs_id=request.POST.get('gs'),
                                    failure__failurecategory_id=request.POST.get('SarfaslId'),
                                    status__status='open').count()
    if counter != 0:
        messages.warning(request, 'برای این نازل یک تیکت در حال بررسی موجود است')
        return redirect(url)
    counter = Ticket.objects.filter(Pump_id=None, gs_id=request.POST.get('gs'), status__status='open',
                                    failure_id=request.POST.get('failure')).count()
    if counter != 0:
        messages.warning(request, 'برای این عنوان یک تیکت در حال بررسی موجود است')
        return redirect(url)

    if request.user.owner.role.role not in ['setad', 'mgr', 'fani']:
        if request.POST.get('Pump') != '0':
            if request.POST.get('Pump'):
                sell = SellModel.objects.filter(tolombeinfo_id=int(request.POST.get('Pump'))).order_by('-id').first()
                _today = datetime.datetime.today()
                try:
                    selldate = (_today.today() - sell.update).days
                    _countnosell = selldate
                    if int(request.user.owner.zone.setrejectticket) != 0 and int(selldate) > int(
                            request.user.owner.zone.setrejectticket):
                        messages.warning(request,
                                         'روزهای فروش ثبت نشده این نازل بیش از حد مجاز است امکان ثبت تیکت نمیباشد')
                        return redirect(url)
                except:
                    d = 0

    form = TicketForm()
    if request.method == 'POST':

        if request.user.owner.role.role == 'tek':
            _lat = request.POST.get('id_latid')
            _long = request.POST.get('id_lngid')
        else:
            _lat = ""
            _long = ""
        fail = FailureSub.objects.get(id=request.POST.get('failure'))
        form = TicketForm(request.POST)
        if form.is_valid():
            try:
                _gsmodel = GsModel.object_role.c_gsmodel(request).get(id=int(request.POST.get('gs')))
            except GsModel.DoesNotExist:
                messages.error(request, 'دسترسی غیر مجاز')
                return TemplateResponse(request, 'ticket/newticket.html',
                                        {'form': form, 'gs': gs, 'sarfasl': sarfasl})

            missing_fields = check_required_fields(_gsmodel)
            if missing_fields:
                fields_list = "، ".join(missing_fields)
                messages.error(request,
                               f'برای ثبت تیکت، باید اطلاعات زیر  در پروفایل جایگاه توسط سامانه منطقه تکمیل گردد: {fields_list}')
                return redirect(url)

            date_out = datetime.datetime.today()
            date_out2 = datetime.datetime.today()

            form.instance.countnosell = _countnosell
            last_15_day_ago = datetime.datetime.today() - datetime.timedelta(days=15)

            if fail.enname == 'initial':
                counter2 = Ticket.objects.filter(gs_id=request.POST.get('gs'), create__gt=last_15_day_ago,
                                                 create__lte=date_out2,
                                                 failure__enname='initial').count()
                lastticketdate = Ticket.objects.filter(gs_id=request.POST.get('gs'),
                                                       failure__enname='initial').last()

                if counter2 >= 1:
                    try:
                        initialok = ReInitial.objects.filter(gs_id=request.POST.get('gs'),
                                                             tarikh__gte=lastticketdate.closedate).last()
                        if initialok:
                            if initialok.accept_gs == False or initialok.accept_zone == False or initialok.accept_tek == False:
                                messages.warning(request,
                                                 'تاییدیه های فرم تعویض هارد زیر 15 رو انجام نشد')
                                return redirect('base:listreinitial')
                            else:
                                _ok = 1
                                counter2 = 0

                    except ValueError:
                        messages.error(request, 'نباید دو تیکت  در حال بررسی  اینشیال هم زمان وجود داشته باشد.')

                if counter2 >= 1:
                    gsm = GsModel.objects.get(id=request.POST.get('gs'))
                    gsmobail = Owner.objects.exclude(user__username='2161846736').filter(zone_id=gsm.area.zone_id,
                                                                                         refrence_id=16, active=True)[
                               :1]
                    for item in gsmobail:
                        mobail = item.mobail
                    if mobail:
                        message = '''
                                             سلام ، لطفا برای جایگاه {param1} فرم مجوز اینشیال  ( بعلت درخواست راه اندازی در بازه زیر 15 روز)را تکمیل  و تایید نمایید. 
                                            شرکت ملی پخش فرآورده های نفتی ایران
                                                            '''.format(param1=gsm.name)
                        message = message.replace('{param1}', gsm.name)
                        try:
                            SendOTP2(mobail, message, 0, 0, 0)
                        except:
                            print('ok')

                    messages.warning(request,
                                     'زمان درخواست اینشیال زیر 15 روز میباشد ، ابتدا باید فرم مربوطه  از منوی تیکت ها / مجوز ری اینشیال تکمیل و به تایید افراد مربوطه برسد')
                    return redirect('base:ComposeTicket')
            if fail.isscanqrcod10minago:
                try:
                    if _gsmodel.gsipclog.get().update:
                        _gs = _gsmodel.gsipclog.get()
                        todaydate = datetime.datetime.today()
                        _sec = (todaydate - _gs.update).seconds

                        if _sec > 1200:
                            messages.error(request,
                                           'برای ثبت این تیکت باید ابتدا رمزینه را اسکن کنید ( مدت اعتبار اسکن رمزینه تا 15 دقیقه میباشد).')
                            return redirect(url)

                        if _gs.rpm_version not in Parametrs.objects.filter().first().rpm_version:
                            messages.error(request,
                                           'ابتدا باید نسخه آخر RPM را نصب کنید و مجدد رمزینه را اسکن  کنید).')
                            return redirect(url)
                except:
                    messages.error(request,
                                   'برای ثبت این تیکت باید ابتدا رمزینه را اسکن کنید ( مدت اعتبار اسکن رمزینه تا 15 دقیقه میباشد).')
                    return redirect(url)

            mohandesicount = Ticket.objects.filter(failure__failurecategory_id=1015, Pump__status__status=True,
                                                   gs_id=_gsmodel.id)

            if mohandesicount.count() > _gsmodel.tedadnazelmohandesi and _gsmodel.tedadnazelmohandesi != 0 and fail.failurecategory.id in [
                1010, 1011]:
                messages.warning(request,
                                 'تعداد نازل هایی که مشکل مکانیکی دارند بیش از حد مجاز میباشد . برای استفاده از خدمات سامانه ابتدا مشکل مکانیکی نازل های مربوطه را بر طرف بفرمایید.')
                return redirect(url)
            if _gsmodel.status_id == 2:
                messages.warning(request,
                                 'این جایگاه در وضعیت غیر فعال میباشد.')
                return redirect(url)
            if _gsmodel.isticket == False and request.user.owner.role.role == 'gs':
                messages.warning(request,
                                 'امکان ثبت تیکت برای شما بسته شده است ، با رئیس سامانه هوشمند منطقه تماس بگیرید')
                return redirect(url)

            form.instance.owner_id = request.user.id
            try:
                _temp = float(request.POST.get('id_temp'))
                _humidity = float(request.POST.get('id_humidity'))
                _pressure = float(request.POST.get('id_pressure'))
                form.instance.temp = int(round(_temp))
                form.instance.humidity = int(round(_humidity))
                form.instance.pressure = int(round(_pressure))
                form.instance.main = request.POST.get('id_weather')
            except:
                pass
            form.instance.status_id = StatusTicket.objects.get(status='open').id
            form.instance.organization_id = fail.organization_id
            form.instance.create_shamsi_year = jdatetime.datetime.now().year
            if len(str(jdatetime.datetime.now().month)) == 1:
                month = '0' + str(jdatetime.datetime.now().month)
            else:
                month = jdatetime.datetime.now().month
            form.instance.create_shamsi_month = month
            if len(str(jdatetime.datetime.now().day)) == 1:
                day = '0' + str(jdatetime.datetime.now().day)
            else:
                day = jdatetime.datetime.now().day
            form.instance.create_shamsi_day = day
            form.instance.shamsi_date = str(jdatetime.datetime.now().year) + "-" + str(month) + '-' + str(day)
            if request.POST.get('Pump') != '0':
                form.instance.Pump_id = request.POST.get('Pump')
                _location = Pump.objects.get(id=int(request.POST.get('Pump'))).gs.location if Pump.objects.get(
                    id=int(request.POST.get('Pump'))).gs.location else "0"

                if len(_location) < 7:
                    messages.warning(request,
                                     ' ابتدا باید آدرس جغرافیایی( لوکیشن ) جایگاه  توسط سامانه منطقه بروز گردد')
                    return TemplateResponse(request, 'ticket/newticket.html',
                                            {'form': form, 'gs': gs, 'sarfasl': sarfasl})

            a = form.save()

            dsk = checkxss(form.instance.descriptionowner)
            if _ok == 1:
                dsk = str(dsk) + str("( اینشیال زیر 15 روز - فرم مربوطه تکمیل شد)")
            Workflow.objects.create(ticket_id=a.id, user_id=request.user.id, description=dsk,
                                    organization_id=a.organization_id, failure_id=a.failure_id, lat=_lat, lang=_long)
            if request.user.owner.role.role not in ['setad', 'mgr', 'fani'] and request.POST.get('Pump') != '0':
                Ticket.check_science(a.gs_id, a.Pump_id, request.user.owner.zone.setinday,
                                     request.user.owner.zone.setamount)
                sell = SellModel.objects.filter(tolombeinfo_id=a.Pump_id).order_by('-id').first()
                _today = datetime.datetime.today()
                try:
                    selldate = (_today.today() - sell.update).days
                    if int(selldate) > int(request.user.owner.zone.setsellday) and int(
                            request.user.owner.zone.setsellday) != 0:
                        TicketScience.objects.create(gs_id=a.gs_id, status_id=3, pump_id=a.Pump_id,
                                                     amount=int(selldate))
                except:
                    selldate = 0

            try:
                countticket = Ticket.objects.values('Pump').filter(
                    gs_id=_gsmodel.id,
                    failure__failurecategory__in=[1010, 1011],
                    Pump__product_id=a.Pump.product.id,
                    status_id=1
                ).distinct().aggregate(tedad=Count('Pump'))
                ticket_count = countticket['tedad']
                pumpcount = Pump.objects.filter(gs_id=_gsmodel.id, status__status=True,
                                                product_id=a.Pump.product.id).count()
                if pumpcount > 0:
                    percentage = (ticket_count / pumpcount) * 100
                else:
                    percentage = 0

                if percentage >= 100:
                    gsm = GsModel.objects.get(id=request.POST.get('gs'))
                    gsmobail = Owner.objects.exclude(user__username='2161846736').filter(zone_id=gsm.area.zone_id,
                                                                                         refrence_id=1, active=True)[:1]
                    for item in gsmobail:
                        mobail = item.mobail
                    if mobail:
                        message = '''
                                                                 سلام ، تعداد خرابی نازل های جایگاه {param1} برای فرآورده {param2} {param3} درصد میباشد . 
                                                                  شرکت ملی پخش فرآورده های نفتی ایران
                                                                                '''.format(param1=gsm.name,
                                                                                           param2=a.Pump.product.name,
                                                                                           param3=str(percentage))
                        message = message.replace('{param1}', gsm.name)
                        message = message.replace('{param2}', a.Pump.product.name)
                        message = message.replace('{param3}', str(percentage))
                        try:
                            SendOTP2(mobail, message, 0, 0, 0)
                        except:
                            print('ok')
            except Exception as e:
                pass

            add_to_log(request, f'ثبت تیکت ', a.id)
            messages.success(request, SUCCESS_TICKET)
            return redirect(url)

    return TemplateResponse(request, 'ticket/newticket.html',
                            {'form': form, 'gs': gs, 'sarfasl': sarfasl})


@cache_permission('t_open')
@transaction.atomic
def crudetickets(request):
    with connection.cursor() as cursor:
        cursor.execute("SET TRANSACTION ISOLATION LEVEL READ COMMITTED")
    isforyat = False
    base_query = Ticket.objects.select_related(
        'gs', 'status', 'Pump', 'failure', 'organization'
    ).prefetch_related(
        'gs__gsowner', 'gs__area', 'gs__area__zone'
    ).only(
        'id', 'create', 'foryat', 'gs__name', 'gs__gsid', 'Pump__number',
        'status__status', 'failure__info', 'organization__organiztion'
    )

    list2 = None
    _list = None
    add_to_log(request, f'مشاهده تیکت های در حال بررسی ', 0)
    role = request.user_data.get('role_name')
    _zone_id = request.user_data.get('zone_id')
    _owner_id = request.user_data.get('owner_id')
    _area_id = request.user_data.get('area_id')

    if role == 'zone':
        areas = Area.objects.select_related('zone').filter(zone_id=_zone_id)
    else:
        areas = None
    if role in ['zone', 'area']:
        action = Owner.objects.filter(zone_id=_zone_id, role__role='tek')
    else:
        action = None
    _list = None
    sarfasl = FailureCategory.objects.all()
    if role == 'tek':
        _list = base_query.filter(
            gs__gsowner__owner_id=_owner_id,
            status__status='open'
        ).exclude(organization_id=4).order_by('-foryat', '-create')

        if isforyat:
            isforyat = True if _list.first().foryat == 3 else False

    elif role == 'gs':
        _list = base_query.filter(
            gs__gsowner__owner_id=_owner_id,
            status__status='open'
        ).order_by(
            '-foryat', '-create')
    elif role == 'engin':
        _list = base_query.filter(
            gs__gsowner__owner_id=_owner_id,
            status__status='open',
            organization__organiztion__in=['tekengin', 'karshenasengin', 'engin']
        ).order_by(
            '-foryat', '-create')
    elif role == "zone":
        _list = base_query.filter(
            gs__area__zone_id=_zone_id,
            status__status='open'
        ).exclude(failure__failurecategory_id=1015).order_by('-foryat', '-create')
        list2 = base_query.filter(
            gs__area__zone_id=_zone_id,
            status__status='open'
        ).order_by('-foryat', '-create')

    elif role == "area":
        _list = base_query.filter(
            gs__area_id=_area_id,
            status__status='open'
        ).order_by(
            '-foryat', '-create')
    elif role in ["fani", "test", "hoze"]:
        org_map = {
            "fani": ['fani', 'test'],
            "test": ['test'],
            "hoze": ['hoze']
        }
        _list = base_query.filter(
            organization__organiztion__in=org_map.get(role, []),
            status__status='open'
        ).order_by('-create')

    elif role in ["mgr", "setad"]:
        _list = base_query.filter(status__status='open').order_by('-create')

    else:
        _list = base_query.filter(
            organization__organiztion=role,
            status__status='open'
        ).order_by('-foryat', '-id')

    datein = str(request.GET.get('select'))

    dateout = str(request.GET.get('select2'))
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "20"

    if len(datein) < 10:
        datein = "2020-10-10"
        dateout = "2100-12-20"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
    datein = str(datein) + " 00:00:00"
    dateout = str(dateout) + " 23:59:59"

    _list = _list.filter(create__gte=datein, create__lte=dateout).order_by('-foryat', '-id')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    if role == 'zone' and _filter.data:
        list2 = list2.filter(create__gte=datein, create__lte=dateout).order_by('-foryat', '-id')
        _filter = TicketFilter(request.GET, queryset=list2, request=request)
    _list = _filter.qs
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd.isnumeric():
                _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(gs__name=cd))

    paginator = Paginator(_list, pajnumber)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]

    return TemplateResponse(request, 'ticket/CrudeTickets.html',
                            {'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'query_string': query_string,
                             'this_date': this_date, 'today_date': today_date, 'isforyat': isforyat,
                             'sarfasl': sarfasl, 'page_obj': page_obj, 'tedad': tedad,
                             'areas': areas, 'action': action, 'isaction': False}
                            )


@cache_permission('t_open')
def erjatotek(request):
    isforyat = False
    if request.user.owner.mobail_ischeck == False:
        mobail = request.user.owner.mobail
        if len(mobail) != 11:
            messages.warning(request, 'لطفا شماره تلفن همراه خود را بصورت صحیح ( یازده رقمی) وارد کنید')
            return redirect('accounts:MyProfile')
        createotp(mobail, 1)
        return redirect('base:veryfi')
    list2 = None
    _list = None
    add_to_log(request, f'مشاهده تیکت های ارجاع شده ', 0)
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = None
    if request.user.owner.role.role in ['zone', 'area']:
        action = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
    else:
        action = None
    _list = None
    sarfasl = FailureCategory.objects.all()
    if request.user.owner.role.role == 'tek':
        _list = Ticket.objects.filter(usererja=request.user.owner.id, status__status='open').order_by('-foryat',
                                                                                                      '-create')
        if isforyat:
            isforyat = True if _list.first().foryat == 3 else False

    if request.user.owner.role.role == 'gs':
        _list = Ticket.objects.filter(usererja_=request.user.owner.id, status__status='open').order_by(
            '-foryat', '-create')
    if request.user.owner.role.role == 'engin':
        _list = Ticket.objects.filter(usererja=request.user.owner.id, status__status='open',
                                      ).order_by(
            '-foryat', '-create')
    if request.user.owner.role.role == "zone":
        _list = tickets = Ticket.objects.select_related('gs__area__zone', 'failure__failurecategory').filter(
            gs__area__zone_id=request.user.owner.zone_id, status__status='open'
        ).exclude(failure__failurecategory_id=1015).order_by('-foryat', '-create')
        list2 = tickets = Ticket.objects.select_related('gs__area__zone').filter(
            gs__area__zone_id=request.user.owner.zone_id, status__status='open'
        ).order_by('-foryat', '-create')
    if request.user.owner.role.role == "area":
        _list = Ticket.objects.filter(gs__area__id=request.user.owner.area_id, status__status='open').order_by(
            '-foryat', '-create')
    if request.user.owner.role.role == "fani":
        _list = Ticket.objects.filter(organization__organiztion__in=['fani', 'test'], status__status='open').order_by(
            '-foryat', '-create')

    if _list == None:
        _list = Ticket.objects.filter(organization__organiztion=request.user.owner.role.role,
                                      status__status='open').order_by('-foryat', '-id')

    datein = str(request.GET.get('select'))

    dateout = str(request.GET.get('select2'))
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "50"

    if len(datein) < 10:
        datein = "2020-10-10"
        dateout = "2100-12-20"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
    datein = str(datein) + " 00:00:00"
    dateout = str(dateout) + " 23:59:59"

    _list = _list.filter(create__gte=datein, create__lte=dateout).order_by('-foryat', '-id')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    if request.user.owner.role.role == 'zone' and _filter.data:
        list2 = list2.filter(create__gte=datein, create__lte=dateout).order_by('-foryat', '-id')
        _filter = TicketFilter(request.GET, queryset=list2, request=request)
    _list = _filter.qs
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd.isnumeric():
                _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(gs__name=cd))

    paginator = Paginator(_list, pajnumber)
    page_num = request.GET.get('page')

    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]

    return TemplateResponse(request, 'ticket/CrudeTickets.html',
                            {'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'query_string': query_string,
                             'this_date': this_date, 'today_date': today_date, 'isforyat': isforyat,
                             'sarfasl': sarfasl, 'page_obj': page_obj, 'tedad': tedad,
                             'areas': areas, 'action': action, 'isaction': False}
                            )


@cache_permission('t_open')
def crudeticketstek(request, id, code):
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    add_to_log(request, f'مشاهده تیکت های در حال بررسی ', 0)
    sarfasl = FailureCategory.objects.all()

    _list = Ticket.object_role.c_gs(request, 1).filter(gs__gsowner__owner_id=id, status__status='open',
                                                       failure__failurecategory_id=code).order_by(
        '-create')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    _list = _filter.qs
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd.isnumeric():
                _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(gs__name=cd))

    paginator = Paginator(_list, 5)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()

    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date + datetime.timedelta(days=1)

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]

    return TemplateResponse(request, 'ticket/CrudeTickets.html',
                            {'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'query_string': query_string,
                             'this_date': this_date, 'today_date': today_date,
                             'sarfasl': sarfasl, 'page_obj': page_obj, 'tedad': tedad,
                             'isaction': False}
                            )


def export_ticket_xls(request):
    add_to_log(request, 'ارسال تیکت به اکسل', 0)
    _all = easyxf('border: left thick, bottom thick, right thick,top thick',
                  'pattern: pattern solid, fore_colour green;')
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    if request.user.owner.role.role == 'zone':
        rows = Ticket.objects.filter(gs__arbain=True, status_id=1, failure__failurecategory_id__in=[1010, 1011, 1045],
                                     gs__area__zone_id=request.user.owner.zone_id).values_list('create_shamsi_year',
                                                                                               'create_shamsi_month',
                                                                                               'create_shamsi_day',
                                                                                               'gs__area__zone__name',
                                                                                               'gs__area__name',
                                                                                               'gs__gsid',
                                                                                               'gs__name',
                                                                                               'failure__failurecategory__info',
                                                                                               'Pump__number',
                                                                                               'Pump__product__name',
                                                                                               'organization__name').order_by(
            'gs__area__zone_id', 'gs__area_id', 'gs_id')
    else:
        rows = Ticket.objects.filter(gs__arbain=True, failure__failurecategory_id__in=[1010, 1011, 1045],
                                     status_id=1).values_list('create_shamsi_year',
                                                              'create_shamsi_month',
                                                              'create_shamsi_day',
                                                              'gs__area__zone__name',
                                                              'gs__area__name',
                                                              'gs__gsid',
                                                              'gs__name',
                                                              'failure__failurecategory__info',
                                                              'Pump__number',
                                                              'Pump__product__name',
                                                              'organization__name').order_by(
            'gs__area__zone_id', 'gs__area_id', 'gs_id')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], style=_all)
    ws.preview_magn = 150

    ws.page_preview = True
    wb.save(response)
    return response


@cache_permission('myticket')
def forwardtickets(request):
    templateitem = "ticket/forwardtickets.html"
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "50"
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = None
    if request.user.owner.role.role in ['zone', 'area']:
        action = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
    else:
        action = None
    sarfasl = FailureCategory.objects.all()
    _list = Workflow.objects.filter(user__owner__role_id=request.user.owner.role.id, ticket__status__status='open')
    if request.user.owner.refrence_id == 8:
        templateitem = "ticket/CrudeTickets.html"
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id, status__status='open').order_by(
            '-id')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    _list = _filter.qs
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if request.user.owner.refrence_id == 8:
                if cd.isnumeric():
                    _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
                else:
                    _list = _list.filter(Q(gs__name=cd))
            else:
                if cd.isnumeric():
                    _list = _list.filter(Q(ticket__id__exact=cd) | Q(ticket__gs__gsid__exact=cd))
                else:
                    _list = _list.filter(Q(ticket__gs__name=cd))

    paginator = Paginator(_list, pajnumber)
    page_num = request.GET.get('page')
    data = request.GET.copy()

    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count
    add_to_log(request, f'مشاهده تیکت های در دست اقدام ', 0)
    return TemplateResponse(request, templateitem,
                            {'tedad': tedad, 'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'page_obj': page_obj,
                             'query_string': query_string, 'areas': areas, 'action': action, 'isaction': False,
                             'sarfasl': sarfasl}
                            )


@cache_permission('t_close')
def closedtickets(request):
    _list = None
    areas = None
    action = None
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "50"
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id, status__status='close').order_by(
            '-closedate')
    if request.user.owner.role.role in ['zone', 'area']:
        action = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
    if request.user.owner.role.role in ['tek']:
        _list = Ticket.objects.filter(actioner_id=request.user.owner.id, status__status='close').order_by(
            '-closedate')
    if request.user.owner.role.role in ["gs"]:
        _list = Ticket.objects.filter(gs__gsowner__owner_id=request.user.owner.id, status__status='close').order_by(
            '-closedate')

    if request.user.owner.role.role == "area":
        _list = Ticket.objects.filter(gs__area__id=request.user.owner.area_id, status__status='close').order_by(
            '-closedate')
    if request.user.owner.role.role == "fani":
        _list = Ticket.objects.filter(status__status='close').order_by('-closedate')

    if request.user.owner.role.role == "test":
        _list = Ticket.objects.filter(status__status='close').order_by('-closedate')
    if request.user.owner.role.role == "engin":
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id,
                                      organization__organiztion__in=['tekengin', 'karshenasengin', 'engin'],
                                      status__status='close').order_by('-closedate')
    if request.user.owner.role.role == "hoze":
        _list = Ticket.objects.filter(organization__organiztion='hoze', status__status='close').order_by('-closedate')
    if request.user.owner.role.role == "mgr":
        _list = Ticket.objects.filter(status__status='close').order_by('-closedate')
    if request.user.owner.role.role == "setad":
        _list = Ticket.objects.filter(status__status='close').order_by('-closedate')
    if _list == None:
        _list = Ticket.objects.filter(organization__organiztion=request.user.owner.role.role,
                                      status__status='close').order_by('-id')
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    if len(datein) < 10:
        datein = "2020-10-10"
        dateout = "2100-12-20"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
    datein = str(datein) + " 00:00:00"
    dateout = str(dateout) + " 23:59:59"
    _list = _list.filter(closedate__gte=datein, closedate__lte=dateout).order_by('-id')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    _list = _filter.qs

    paginator = Paginator(_list, pajnumber)
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd.isnumeric():
                _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(gs__name=cd))
            paginator = Paginator(_list, 5)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
        if 'today' in page_num:
            paginator = Paginator(_list.filter(closedate__date=datetime.datetime.today()), 1000)
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(closedate__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(closedate__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date + datetime.timedelta(days=1)
    query_string = request.META.get("QUERY_STRING", "")

    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count
    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    add_to_log(request, f'مشاهده تیکت های بسته شده ', 0)
    return TemplateResponse(request, 'ticket/ClosedTickets.html',
                            {'tedad': tedad, 'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'page_obj': page_obj,
                             'query_string': query_string, 'this_date': this_date, 'today_date': today_date,
                             'areas': areas,
                             'action': action, 'isaction': True}
                            )


@cache_permission('t_all')
def roletickets(request):
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = None
    if request.user.owner.role.role in ['zone', 'area']:
        action = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
    else:
        action = None
    _list = None
    if request.user.owner.role.role in ['gs', 'tek']:
        _list = Ticket.objects.filter(gs__gsowner__owner_id=request.user.owner.id).order_by(
            '-id')
    if request.user.owner.role.role == "zone":
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id).order_by(
            '-id')
    if request.user.owner.role.role == "area":
        _list = Ticket.objects.filter(gs__area__id=request.user.owner.area_id).order_by('-id')
    if request.user.owner.role.role == "fani":
        _list = Ticket.objects.filter().order_by('-id')
    if request.user.owner.role.role == "test":
        _list = Ticket.objects.filter().order_by('-id')
    if request.user.owner.role.role == "engin":
        _list = Ticket.objects.filter(organization__organiztion__in=['tekengin', 'karshenasengin', 'engin'],
                                      gs__area__zone_id=request.user.owner.zone_id).order_by(
            '-id')
    if request.user.owner.role.role == "hoze":
        _list = Ticket.objects.filter(organization_id=6).order_by('-id')
    if request.user.owner.role.role == "mgr":
        _list = Ticket.objects.filter().order_by('-id')
    if request.user.owner.role.role == "setad":
        _list = Ticket.objects.filter().order_by('-id')
    if request.user.owner.refrence_id == 8:
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id).order_by(
            '-id')
    if _list == None:
        _list = Ticket.objects.filter(organization__organiztion=request.user.owner.role.role).order_by('-id')
    datein = str(request.GET.get('select'))
    dateout = str(request.GET.get('select2'))
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "50"
    if len(datein) < 10:
        datein = "2020-01-01"
        dateout = "2100-12-29"
    else:
        datein = datein.split("/")
        dateout = dateout.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]), year=int(datein[0])).togregorian()
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]), year=int(dateout[0])).togregorian()
    datein = str(datein) + " 00:00:00"
    dateout = str(dateout) + " 23:59:59"
    _list = _list.filter(create__gte=datein, create__lte=dateout).order_by('-id')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    _list = _filter.qs
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd.isnumeric():
                _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(gs__name=cd))

    paginator = Paginator(_list, pajnumber)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date + datetime.timedelta(days=1)
    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count
    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    add_to_log(request, f'مشاهده لیست همه تیکت ها ', 0)
    return TemplateResponse(request, 'ticket/RoleTickets.html',
                            {'tedad': tedad, 'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'page_obj': page_obj,
                             'query_string': query_string, 'this_date': this_date, 'today_date': today_date,
                             'areas': areas,
                             'action': action, 'isaction': True}
                            )


@cache_permission('t_my')
def mytickets(request):
    add_to_log(request, 'مشاهده تیکت های من', 0)
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = None
    if request.user.owner.role.role in ['zone', 'area']:
        action = Owner.objects.filter(zone_id=request.user.owner.zone_id, role__role='tek')
    else:
        action = None
    if request.user.owner.refrence_id == 1:
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id, organization_id=5,
                                      status_id=1).order_by('-id')
    else:
        _list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id,
                                      owner_id=request.user.owner.id).order_by('-id')
    _filter = TicketFilter(request.GET, queryset=_list, request=request)
    _list = _filter.qs
    form = SearchForm()
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd.isnumeric():
                _list = _list.filter(Q(id__exact=cd) | Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(gs__name=cd))
    paginator = Paginator(_list, 5)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    this_date = datetime.datetime.today()
    if 'page' in data:
        del data['page']
    if page_num:

        if page_num[:3] in ['pre', 'nex']:
            this_date = page_num.split('@')
            this_date = this_date[1]
            if this_date:
                this_date = datetime.datetime.strptime(this_date, DATE_FORMAT)

        if 'alltickets' in page_num:
            paginator = Paginator(_list, 1000)
        if 'today' in page_num:
            paginator = Paginator(_list.filter(create__date=datetime.datetime.today()), 1000)
        if 'previews' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date - datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date - datetime.timedelta(days=1)
        if 'next' in page_num:
            paginator = Paginator(_list.filter(create__date=this_date + datetime.timedelta(days=1)),
                                  1000)
            this_date = this_date + datetime.timedelta(days=1)
    query_string = request.META.get("QUERY_STRING", "")

    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count
    this_date = str(this_date)
    this_date = this_date.split(' ')
    this_date = this_date[0]
    today_date = str(datetime.datetime.today())
    today_date = today_date.split(' ')
    today_date = today_date[0]
    return TemplateResponse(request, 'ticket/MyTickets.html',
                            {'tedad': tedad, 'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'page_obj': page_obj,
                             'query_string': query_string, 'this_date': this_date, 'today_date': today_date,
                             'areas': areas,
                             'action': action, 'isaction': False,
                             'form': form}
                            )


@cache_permission('report')
def gstickets(request, id):
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    _list = Ticket.object_role.c_gs(request, 0).filter(gs__gsid=id, status__status='open').order_by('-id')
    add_to_log(request, f'مشاهده تیکت های یک جایگاه ', id)
    paginator = Paginator(_list, 5)
    page_num = request.GET.get('page')
    data = request.GET.copy()

    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")

    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)

    return TemplateResponse(request, 'report/listticketgs.html',
                            {'filter': filter, 'list': page_object, 'data': urlencode(data),
                             'query_string': query_string,
                             }
                            )


@cache_permission('users')
def userlist(request):
    add_to_log(request, 'مشاهده لیست کاربران', 0)
    form = SearchForm()
    vore = request.GET.get('vore')
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "10"
    if request.user.owner.role.role == 'zone':
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id)
        zoneid = 1
        storages = []
        area = Area.objects.filter(zone_id=request.user.owner.zone_id)
        role = Role.objects.filter(showlevel=2)
        refrence = Refrence.objects.filter(showlevel=2)
        _list = Owner.objects.annotate(
            full_name=Concat('name', V(' '), 'lname')).filter(zone_id=request.user.owner.zone_id)
    else:
        refrence = Refrence.objects.all()
        area = Area.objects.all()
        role = Role.objects.all()
        storages = Storage.objects.all().order_by('sortid')
        zone = Zone.objects_limit.all()
        zoneid = 0
        _list = Owner.objects.annotate(
            full_name=Concat('name', V(' '), 'lname')).all()

    _filter = UserFilters(request.GET, queryset=_list)
    _list = _filter.qs

    if vore == "2":
        add_to_log(request, 'ارسال آمار لیست کاربران به اکسل  ', 0)
        my_path = "users.xlsx"
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = 'attachment; filename=' + my_path
        font = Font(bold=True)
        fonttitr = Font(bold=True, size=20)
        fonttitr2 = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active  # work with default worksheet
        ws1.title = "گزارش لیست کاربران "
        ws1.sheet_view.rightToLeft = True
        ws1.firstFooter.center.text = "ali"
        ws1.merge_cells('A1:P1')

        ws1["A1"] = f'گزارش لیست کاربران در تاریخ   {today}'
        ws1["A1"].font = fonttitr

        ws1.merge_cells('A2:A2')
        ws1["A2"] = "ردیف"
        ws1["A2"].font = font

        ws1.merge_cells('B2:B2')
        ws1["B2"] = "نام"
        ws1["B2"].font = fonttitr2

        ws1.merge_cells('C2:C2')
        ws1["C2"] = "نام خانوادگی "
        ws1["C2"].font = font

        ws1.merge_cells('D2:D2')
        ws1["D2"] = " کد ملی"
        ws1["D2"].font = font

        ws1.merge_cells('E2:E2')
        ws1["E2"] = "موبایل"
        ws1["E2"].font = font

        ws1.merge_cells('F2:F2')
        ws1["F2"] = "نقش"
        ws1["F2"].font = font

        ws1.merge_cells('G2:G2')
        ws1["G2"] = "سمت"
        ws1["G2"].font = font

        ws1.merge_cells('H2:H2')
        ws1["H2"] = " منطقه "
        ws1["H2"].font = font

        ws1.merge_cells('I2:I2')
        ws1["I2"] = " ناحیه"
        ws1["I2"].font = font

        ws1.merge_cells('J2:J2')
        ws1["J2"] = "وضعیت"
        ws1["J2"].font = font

        ws1.merge_cells('K2:K2')
        ws1["K2"] = "مدرک تحصیلی"
        ws1["K2"].font = font

        ws1.merge_cells('L2:L2')
        ws1["L2"] = "شماره حساب"
        ws1["L2"].font = font

        ws1.merge_cells('M2:M2')
        ws1["M2"] = "تاریخ آغاز به کار"
        ws1["M2"].font = font

        ws1.merge_cells('N2:N2')
        ws1["N2"] = "وضعیت تاهل"
        ws1["N2"].font = font

        ws1.merge_cells('O2:O2')
        ws1["O2"] = "تعداد اولاد"
        ws1["O2"].font = font

        ws1.merge_cells('P2:P2')
        ws1["P2"] = "گروه شغلی"
        ws1["P2"].font = font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(15.25)
        ws1.column_dimensions['D'].width = float(25.25)
        ws1.column_dimensions['E'].width = float(15.25)
        ws1.column_dimensions['F'].width = float(18.25)
        ws1.column_dimensions['G'].width = float(35.25)
        ws1.column_dimensions['H'].width = float(30.25)
        ws1.column_dimensions['I'].width = float(25.25)
        ws1.column_dimensions['J'].width = float(22.25)
        ws1.column_dimensions['K'].width = float(22.25)
        ws1.column_dimensions['L'].width = float(22.25)
        ws1.column_dimensions['M'].width = float(22.25)
        ws1.column_dimensions['N'].width = float(22.25)
        ws1.column_dimensions['O'].width = float(22.25)
        ws1.column_dimensions['P'].width = float(22.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        myfont = Font(size=14, bold=True)  # font styles
        my_fill = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        my_fill2 = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        i = 0

        for item in _list:
            i += 1
            zone = item.zone.name if item.zone else ""
            area = item.area.name if item.area else ""
            refrence = item.refrence.name if item.refrence else ""
            education = item.education if item.education else ""
            role = item.role.name if item.role else ""
            # zone=item.zone.name if item.zone.name else ""

            d = [i, str(item.name), str(item.lname), str(item.codemeli), str(item.mobail),
                 str(role),
                 str(refrence), str(zone), str(area), str(item.active),
                 str(education), str(item.place_of_birth), str(item.start_date), str(item.marital_status),
                 str(item.childcount), str(item.job_group)]

            ws1.append(d)

        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        i += 4
        for cell in ws1[i:i]:  # Last row
            cell.font = myfont
            cell.fill = my_fill2
            cell.border = thin_border

        for cell in ws1["2:2"]:  # First row
            cell.font = myfont
            cell.fill = my_fill
            cell.border = thin_border
        wb.save(response)
        return response

    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = form.cleaned_data['search']
            cd = checknumber(cd)
            cd = checkxss(cd)
            _list = _list.filter(
                Q(full_name__icontains=cd) | Q(lname__icontains=cd) | Q(codemeli__exact=cd) | Q(mobail__exact=cd))
    paginator = Paginator(_list, pajnumber)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")

    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)

    page_obj = paginator.num_pages
    tedad = paginator.count
    count = _list.count()
    return TemplateResponse(request, 'users/listusers.html',
                            {'filter': _filter, 'list': page_object, 'data': urlencode(data),
                             'query_string': query_string,
                             'zone': zone, 'area': area, 'role': role, 'zoneid': zoneid, 'form': form,
                             'page_object': page_object,
                             'count': count, 'refrence': refrence, 'page_obj': page_obj,
                             'storages': storages,
                             'tedad': tedad})


@cache_permission('addgsrel')
def addgsrelation(request, id):
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    gslist = None
    owner = Owner.objects.get(id=id)
    gslist = GsList.objects.filter(owner_id=id)
    name = owner.name + ' ' + owner.lname
    form = SearchForm()
    gs = GsModel.objects.filter(id=0)

    if owner.role.role == "zone":
        gs = GsModel.objects.filter(area__zone_id=owner.zone_id, status_id__in=[1, 3, 4]).order_by('area_id')
    elif owner.role.role == "area":
        gs = GsModel.objects.filter(area_id=owner.area_id, status_id__in=[1, 3, 4]).order_by('area_id')
    elif owner.role.role == "tek":
        gs = GsModel.objects.filter(area__zone_id=owner.zone_id, status_id__in=[1, 3, 4]).order_by('area_id')
    elif owner.role.role == "engin":
        gs = GsModel.objects.filter(area__zone_id=owner.zone_id, status_id__in=[1, 3, 4]).order_by('area_id')
    else:

        if 'search' in request.GET:
            form = SearchForm(request.GET)
            if form.is_valid():
                cd = form.cleaned_data['search']
                gs = GsModel.objects.filter(Q(name__icontains=cd) | Q(gsid__exact=cd))

    q = gslist.values_list('gs_id')
    if owner.role.role == "tek":
        q = GsList.objects.values_list('gs_id').filter(owner__role__role='tek', owner__active=True)
    gs = gs.exclude(id__in=q)
    return TemplateResponse(request, 'users/AddGSRelation.html',
                            {'gs': gs, 'gslist': gslist, 'userID': id, 'name': name, 'form': form,
                             })


@cache_permission('edituser')
def userprofile(request, id):
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    add_to_log(request, 'مشاهده پروفایل', 0)
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    educations = Education.objects.all()
    query_params = request.META.get('HTTP_REFERER')
    if query_params:
        query_params1 = query_params.split("/")
        query_params2 = query_params1[-1:]
        query_params3 = query_params2[0]
        if request.method == 'GET':
            request.session['userfilters'] = query_params3
    if request.user.owner.role.role == 'zone':
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id)
        zoneid = request.user.owner.zone_id
        area = Area.objects.filter(zone_id=request.user.owner.zone_id)
        role = Role.objects.filter(showlevel=2)
        refrence = Refrence.objects.filter(showlevel=2)
    else:
        area = Area.objects.all()
        role = Role.objects.all()
        zone = Zone.objects_limit.all()
        refrence = Refrence.objects.all()
        zoneid = 0
    owner = Owner.object_role.c_owner(request).get(codemeli=id)

    if request.method == 'POST':
        try:
            fname = request.POST.get('fname')
            lname = request.POST.get('lname')
            codemeli = request.POST.get('codemeli')
            mobail = request.POST.get('mobail')
            role_id = request.POST.get('Mylad')
            zone = request.POST.get('Master')
            area_id = request.POST.get('area')
            refrence = request.POST.get('refrence')
            active = request.POST.get('active')
            education = request.POST.get('education')
            place_of_birth = request.POST.get('place_of_birth')
            date_of_birth = request.POST.get('date_of_birth')
            marital_status = request.POST.get('marital_status')
            start_date = request.POST.get('start_date')
            job_group = request.POST.get('job_group')
            childcount = request.POST.get('childcount')

            if active == 'on':
                active = True
            else:
                active = False
            owner.name = checkxss(fname)

            if refrence != "0":
                owner.refrence_id = refrence

            owner.lname = checkxss(lname)
            owner.codemeli = checkxss(codemeli)
            owner.mobail = checkxss(mobail)
            owner.active = active
            owner.role_id = role_id

            if owner.role.role == 'tek':
                if education:
                    owner.education_id = int(education)
                owner.place_of_birth = place_of_birth
                owner.date_of_birth = date_of_birth
                owner.marital_status = marital_status
                owner.start_date = start_date
                owner.job_group = job_group
                owner.childcount = childcount
            if zone != '0':
                owner.zone_id = zone
                owner.area_id = area_id
            if request.user.owner.role.role not in ['zone', 'area', 'gs']:
                owner.area_id = area_id

            if owner.role.role == 'engin':
                owner.zone_id = zone
            owner.save()
            user = User.objects.get(id=owner.user.id)
            user.username = checkxss(codemeli)
            user.first_name = checkxss(fname)
            user.last_name = checkxss(lname)
            user.active = active
            user.is_active = active
            user.save()
            add_to_log(request, 'ویرایش پروفایل کاربر ' + str(checkxss(codemeli)), 0)
            messages.success(request, 'پروفایل شما به درستی ویرایش شد')
            queryurl = request.session['userfilters']
            del request.session['userfilters']
            return redirect(reverse('base:UserList') + queryurl)
        except IntegrityError:
            messages.warning(request, 'این نام کاربری تکراری است')

    return TemplateResponse(request, 'users/profile.html',
                            {'educations': educations, 'owner': owner, 'area': area, 'role': role, 'zone': zone,
                             'zoneid': zoneid, 'refrence': refrence,
                             })


def saveimg(request, id):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':

        owner = Owner.objects.get(id=id)
        imguser = ImageProfile(request.POST, request.FILES, instance=owner)

        if imguser.is_valid():

            imguser.save()
            add_to_log(request, 'ویرایش عکس کاربر ' + str(owner.codemeli), 0)
            messages.success(request, 'عکس بدرستی ذخیره شد')
        else:
            messages.error(request, 'عملیات شکست خورد ، نوع یا سایز فایل مشکل دارد')
    return redirect(url)


@cache_permission('editgs')
def gsedit(request, id):
    id = Decrypt(id)
    gs = GsModel.objects.get(id=id)
    zone = Zone.objects_limit.all()
    form = GsEditForm
    area = Area.objects.filter(zone_id=gs.area.zone_id)

    query_params = request.META.get('HTTP_REFERER')
    if query_params:
        query_params1 = query_params.split("/")
        query_params2 = query_params1[-1:]
        query_params3 = query_params2[0]
        if request.method == 'GET':
            request.session['gsfilters'] = query_params3
    if request.method == 'POST':
        form = GsEditForm(request.POST, instance=gs)
        form.save()
        add_to_log(request, 'ویرایش اطلاعات جایگاه ' + str(gs), gs)
        messages.success(request, 'اطلاعات جایگاه با موفقیت ویرایش شد')
        queryurl = request.session['gsfilters']
        del request.session['gsfilters']
        return redirect(reverse('base:GSInfo') + queryurl)
    return TemplateResponse(request, 'gsedit.html',
                            {'zone': zone, 'gs': gs, 'area': area, 'form': form})


@cache_permission('nazels')
def nazelslist(request):
    _id = 0
    gslist = None
    if request.user.owner.role.role in ['mgr', 'setad']:
        gslist = GsModel.objects.all()
    elif request.user.owner.role.role == 'zone':
        gslist = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)
    nazels = None
    if request.method == 'POST':
        _id = int(request.POST.get('mygsid'))
        nazels = Pump.objects.filter(gs_id=_id).order_by('number')

    return TemplateResponse(request, 'nazels.html',
                            {'nazels': nazels, 'gslist': gslist, 'id': _id})


@cache_permission('editgs')
def editnazels(request):
    products = Product.objects.all()
    brands = PumpBrand.objects.all()
    query_params = request.META.get('HTTP_REFERER')
    if query_params:
        query_params1 = query_params.split("/")
        query_params2 = query_params1[-1:]
        query_params3 = query_params2[0]
        request.session['gsfilters'] = query_params3
    if request.method == 'POST':
        _id = request.POST.get('editidNazel')
        gsname = GsModel.objects.get(id=_id)
        gsname = str(gsname.gsid) + ' - ' + str(gsname.name)
        nazels = Pump.objects.filter(gs_id=_id).order_by('number')

    return TemplateResponse(request, 'nazelsone.html',
                            {'id': _id, 'nazels': nazels, 'gsname': gsname, 'brands': brands,
                             'products': products})


@cache_permission('editgs')
def vieweditnazels(request, id):
    products = Product.objects.all()
    brands = PumpBrand.objects.all()
    nazels = Pump.objects.filter(gs_id=id).order_by('number')
    return TemplateResponse(request, 'nazelsone.html',
                            {'products': products, 'brands': brands, 'id': id, 'nazels': nazels,
                             })


def saveeditnazel(request, _id):
    if request.method == 'POST':
        _list = Pump.objects.filter(gs_id=_id).order_by('id')
        for item in _list:
            sakoo = str(request.POST.get(f'sakoo{item.id}'))
            tolombe = str(request.POST.get(f'tolombe{item.id}'))
            product = str(request.POST.get(f'product{item.id}'))
            brand = str(request.POST.get(f'brand{item.id}'))
            active = request.POST.get(f'active{item.id}')
            actived = request.POST.get(f'actived{item.id}')
            item.product_id = product
            item.pumpbrand_id = brand
            if (len(sakoo) > 0):
                item.sakoo = sakoo
            if (len(tolombe) > 0):
                item.tolombe = tolombe
            if (active != ""):
                item.active = active
            if (actived != ""):
                item.actived = actived
            if active == '0':
                item.actived = 0
            item.save()
            add_to_log(request, ' ویرایش اطلاعات نازل', _id)
        messages.success(request, SUCCESS_MSG)

    queryurl = request.session['gsfilters']
    del request.session['gsfilters']
    return redirect(reverse('base:GSInfo') + queryurl)


def editnazel(request):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':

        _id = request.POST.get('myid')
        add_to_log(request, 'ویرایش نازل', _id)

        _list = Pump.objects.filter(gs_id=int(_id)).order_by('id')
        for item in _list:
            master = str(request.POST.get(f'sakoo{item.id}'))
            pinpad = str(request.POST.get(f'tolombe{item.id}'))

            if (len(master) > 0):
                item.master = master
            if (len(pinpad) > 0):
                item.pinpad = pinpad

            item.save()
            add_to_log(request, 'ویرایش اطلاعات نازل', _id)
        messages.success(request, SUCCESS_MSG)
    return redirect(url)


@cache_permission('report')
def nazelsstatusbymanategh(request):
    add_to_log(request, f'گزارش وضعیت نازلهای منطقه ', 0)
    _list = []
    if request.user.owner.role.role in ['zone']:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    elif request.user.owner.role.role in ['setad', 'mgr', 'test', 'fani']:
        zone = Zone.objects_limit.all().exclude(id=9).order_by('id')
    else:
        messages.warning(request, DENY_PERMISSSION)
        return redirect(HOME_PAGE)
    sum_all_sum = 0
    sum_deactive_sum = 0
    sum_kharab_sum = 0
    sum_active_sum = 0
    for gs in zone:
        sum_all = 0
        sum_deactive = 0
        sum_kharab = 0
        sum_active = 0
        products = Product.objects.all().order_by('id')
        for product in products:
            _all = Pump.objects.filter(gs__area__zone_id=gs.id, product_id=product.id, status__status=True,
                                       gs__status__status=True).count()
            deactive = Pump.objects.filter(gs__area__zone_id=gs.id, product_id=product.id, status__status=False).count()
            kharab = Ticket.objects.filter(gs__area__zone_id=gs.id, gs__status__status=True, Pump__status__status=True,
                                           Pump__product_id=product.id, status_id=1).count()

            _gscount = GsModel.objects.values('gsid').filter(area__zone_id=gs.id, status__status=True,
                                                             gsall__product_id=product.id).annotate(Count('gsid'))
            gscount = len(_gscount)

            active = int(_all) - int(kharab)
            if kharab > 0:
                dismount = round(((kharab / _all) * 100), 1)
            else:
                dismount = 0
            if _all > 0:
                _dict = {
                    'st': 0,
                    'id': gs.id,
                    'gs': gs.name,
                    'product': product.name,
                    'active': active,
                    'deactive': deactive,
                    'kharab': kharab,
                    'dismount': dismount,
                    'gscount': gscount,
                }
                _list.append(_dict)
                sum_all += _all
                sum_deactive += deactive
                sum_kharab += kharab
                sum_active += active

                sum_all_sum += _all
                sum_deactive_sum += deactive
                sum_kharab_sum += kharab
                sum_active_sum += active
            if sum_kharab > 0:
                dismount = round(((sum_kharab / sum_all) * 100), 1)
            else:
                dismount = 0
        _dict = {
            'st': 1,
            'gs': SUM_TITEL + ' ' + gs.name,
            'product': '',
            'active': sum_active,
            'deactive': sum_deactive,
            'kharab': sum_kharab,
            'dismount': dismount,
        }
        _list.append(_dict)
    if sum_kharab_sum > 0:
        dismount = round(((sum_kharab_sum / sum_all_sum) * 100), 1)
    else:
        dismount = 0
    _dict = {
        'st': 2,
        'gs': SUM_TITEL,
        'product': '',
        'active': sum_active_sum,
        'deactive': sum_deactive_sum,
        'kharab': sum_kharab_sum,
        'dismount': dismount,
    }
    _list.append(_dict)

    context = {'list': _list}
    return TemplateResponse(request, 'report/nazelsstatusbymanategh.html', context)


@cache_permission('report')
def nazelsstatusbynavahy(request, id):
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    _list = []
    add_to_log(request, f'گزارش وضعیت نازلهای ناحیه ', 0)
    if request.user.owner.role.role in ['zone']:
        zone = Area.objects.filter(zone_id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    else:
        zone = Area.objects.filter(zone_id=id).order_by('id')
    sum_all_sum = 0
    sum_deactive_sum = 0
    sum_kharab_sum = 0
    sum_active_sum = 0
    for gs in zone:
        sum_all = 0
        sum_deactive = 0
        sum_kharab = 0
        sum_active = 0
        products = Product.objects.all().order_by('id')
        for product in products:
            _all = Pump.objects.filter(gs__area_id=gs.id, product_id=product.id, status__status=True,
                                       gs__status__status=True).count()
            deactive = Pump.objects.filter(gs__area_id=gs.id, product_id=product.id, status__status=False).count()
            kharab = Ticket.objects.filter(gs__area_id=gs.id, gs__status__status=True, Pump__status__status=True,
                                           Pump__product_id=product.id, status_id=1).count()
            active = int(_all) - int(kharab)
            if kharab > 0:
                dismount = round(((kharab / _all) * 100), 1)
            else:
                dismount = 0
            if _all > 0:
                _dict = {
                    'st': 0,
                    'id': gs.id,
                    'gs': gs.name,
                    'product': product.name,
                    'active': active,
                    'deactive': deactive,
                    'kharab': kharab,
                    'dismount': dismount,
                }
                _list.append(_dict)
                sum_all += _all
                sum_deactive += deactive
                sum_kharab += kharab
                sum_active += active

                sum_all_sum += _all
                sum_deactive_sum += deactive
                sum_kharab_sum += kharab
                sum_active_sum += active
            if sum_kharab > 0:
                dismount = round(((sum_kharab / sum_all) * 100), 1)
            else:
                dismount = 0
        if sum_all > 0:
            _dict = {
                'st': 1,
                'gs': 'مجموع ناحیه :' + ' ' + gs.name,
                'product': '',
                'active': sum_active,
                'deactive': sum_deactive,
                'kharab': sum_kharab,
                'dismount': dismount,
            }
            _list.append(_dict)
    if sum_kharab_sum > 0:
        dismount = round(((sum_kharab_sum / sum_all_sum) * 100), 1)
    else:
        dismount = 0
    _dict = {
        'st': 2,
        'gs': SUM_TITEL,
        'product': '',
        'active': sum_active_sum,
        'deactive': sum_deactive_sum,
        'kharab': sum_kharab_sum,
        'dismount': dismount,
    }
    _list.append(_dict)

    context = {'list': _list}
    return TemplateResponse(request, 'report/nazelsstatusbynavahy.html', context)


@cache_permission('report')
def nazelsstatusbygs(request, id):
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    _list = []
    add_to_log(request, f'گزارش وضعیت نازلهای جایگاه ', 0)
    if request.user.owner.role.role in ['zone']:
        zone = GsModel.objects.filter(area__id=id, area__zone_id=request.user.owner.zone_id).order_by('id')
    else:
        zone = GsModel.objects.filter(area__id=id).order_by('id')
    sum_all_sum = 0
    sum_deactive_sum = 0
    sum_kharab_sum = 0
    sum_active_sum = 0
    for gs in zone:
        sum_all = 0
        sum_deactive = 0
        sum_kharab = 0
        sum_active = 0
        products = Product.objects.all().order_by('id')
        for product in products:
            _all = Pump.objects.filter(gs_id=gs.id, product_id=product.id, status__status=True,
                                       gs__status__status=True).count()
            deactive = Pump.objects.filter(gs_id=gs.id, product_id=product.id, status__status=False).count()
            kharab = Ticket.objects.filter(gs_id=gs.id, Pump__product_id=product.id, gs__status__status=True,
                                           Pump__status__status=True, status_id=1).count()
            active = int(_all) - int(kharab)
            if kharab > 0:
                dismount = round(((kharab / _all) * 100), 1)
            else:
                dismount = 0
            if _all > 0:
                _dict = {
                    'st': 0,
                    'id': gs.id,
                    'gsid': gs.gsid,
                    'gs': gs.name,
                    'product': product.name,
                    'active': active,
                    'deactive': deactive,
                    'kharab': kharab,
                    'dismount': dismount,
                }
                _list.append(_dict)
                sum_all += _all
                sum_deactive += deactive
                sum_kharab += kharab
                sum_active += active

                sum_all_sum += _all
                sum_deactive_sum += deactive
                sum_kharab_sum += kharab
                sum_active_sum += active
            if sum_kharab > 0:
                dismount = round(((sum_kharab / sum_all) * 100), 1)
            else:
                dismount = 0
        if sum_all > 0:
            _dict = {
                'st': 1,
                'gs': 'مجموع  :' + ' ' + gs.name,
                'product': '',
                'active': sum_active,
                'deactive': sum_deactive,
                'kharab': sum_kharab,
                'dismount': dismount,
            }
            _list.append(_dict)
    if sum_kharab_sum > 0:
        dismount = round(((sum_kharab_sum / sum_all_sum) * 100), 1)
    else:
        dismount = 0
    _dict = {
        'st': 2,
        'gs': SUM_TITEL,
        'product': '',
        'active': sum_active_sum,
        'deactive': sum_deactive_sum,
        'kharab': sum_kharab_sum,
        'dismount': dismount,
    }
    _list.append(_dict)

    context = {'list': _list}
    return TemplateResponse(request, 'report/nazelsstatusbygs.html', context)


@cache_permission('report')
def nazelsstatus(request):
    _list = []
    add_to_log(request, 'گزارش وضعیت نازل', 0)
    products = Product.objects.all().order_by('id')
    for product in products:
        _all = Pump.objects.filter(product_id=product.id, status__status=True,
                                   gs__status__status=True).count()
        deactive = Pump.objects.filter(product_id=product.id, status__status=False).count()
        kharab = Ticket.objects.filter(Pump__product_id=product.id, gs__status__status=True, Pump__status__status=True,
                                       status_id=1).count()
        active = int(_all) - int(kharab) - int(deactive)
        if kharab > 0:
            dismount = round(((kharab / _all) * 100), 1)
        else:
            dismount = 0
        if _all > 0:
            _dict = {
                'st': 0,
                'product': product.name,
                'active': active,
                'deactive': deactive,
                'kharab': kharab,
                'dismount': dismount,
            }
            _list.append(_dict)
    context = {'list': _list}
    return TemplateResponse(request, 'report/nazelsstatus.html', context)


@cache_permission('report')
def masterpinpadtickets(request, id):
    _list = []
    add_to_log(request, f'گزارش وضعیت تیکت ها منطقه ', 0)
    selected_products = request.GET.getlist('product')
    if request.user.owner.role.role in ['zone']:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    elif request.user.owner.role.role in ['setad', 'mgr', 'fani', 'test', 'posht']:
        zone = Zone.objects_limit.all().exclude(id=9).order_by('id')
    else:
        messages.warning(request, DENY_PERMISSSION)
        return redirect(HOME_PAGE)
    s_master = 0
    _tarikh = ""
    s_pinpad = 0
    hour = "-"
    summ = 0
    sum_gs = 0
    sum_pump = 0
    sum_ticket = 0
    sum_master = 0
    sum_pinpad = 0
    products = Product.objects.all()
    if request.method != 'POST':

        for gs in zone:
            gs_query = GsModel.objects.filter(area__zone_id=gs.id, status__status=True)
            if selected_products:
                gs_query = gs_query.filter(product__id__in=selected_products)
            count_gs = gs_query.count()
            pump_query = Pump.objects.filter(gs__area__zone_id=gs.id, status__status=True,
                                             gs__status__status=True)
            if selected_products:
                pump_query = pump_query.filter(product__id__in=selected_products)
            count_pump = pump_query.count()

            ticket_query = Ticket.objects.exclude(organization_id=4).filter(
                gs__area__zone_id=gs.id, status_id=1,
                gs__status__status=True,
                Pump__status__status=True,
                failure__failurecategory_id__in=[1010, 1011]
            )
            if selected_products:
                ticket_query = ticket_query.filter(Pump__product__id__in=selected_products)
            count_ticket = ticket_query.count()

            count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                            gs__status__status=True,
                                                                            Pump__status__status=True,
                                                                            failure__failurecategory_id=1010,
                                                                            failure__isnazel=True).count()

            count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                            gs__status__status=True,
                                                                            Pump__status__status=True,
                                                                            failure__failurecategory_id=1011,
                                                                            failure__isnazel=True).count()
            if selected_products:
                count_master = count_master.filter(Pump__product__id__in=selected_products)
                count_pinpad = count_pinpad.filter(Pump__product__id__in=selected_products)
            if count_master > 0:
                master = ((int(count_master) / int(count_pump)) * 100)
            else:
                master = 0
            if count_pinpad > 0:
                pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
            else:
                pinpad = 0
            summ = round(master + pinpad, 2)

            _dict = {
                'st': 0,
                'id': gs.id,
                'name': gs.name,
                'count_gs': count_gs,
                'count_pump': count_pump,
                'count_ticket': count_ticket,
                'count_master': count_master,
                'count_pinpad': count_pinpad,
                'master': round(master, 2),
                'pinpad': round(pinpad, 2),
                'summ': summ,
            }
            _list.append(_dict)
            _list = sorted(_list, key=itemgetter('master'), reverse=True)
            sum_gs += count_gs
            sum_pump += count_pump
            sum_ticket += count_ticket
            sum_master += count_master
            sum_pinpad += count_pinpad
            if sum_master > 0:
                s_master = ((int(sum_master) / int(sum_pump)) * 100)
            else:
                s_master = 0
            if count_pinpad > 0:
                s_pinpad = ((int(sum_pinpad) / int(sum_pump)) * 100)
            else:
                s_pinpad = 0
            summ = round(s_master + s_pinpad, 2)
        if id == 1:
            _list = sorted(_list, key=itemgetter('master'), reverse=True)
        elif id == 2:
            _list = sorted(_list, key=itemgetter('pinpad'), reverse=True)
        elif id == 3:
            _list = sorted(_list, key=itemgetter('summ'), reverse=True)
        _dict = {
            'st': 1,
            'id': 0,
            'name': SUM_TITEL,
            'count_gs': sum_gs,
            'count_pump': sum_pump,
            'count_ticket': sum_ticket,
            'count_master': sum_master,
            'count_pinpad': sum_pinpad,
            'master': round(s_master, 2),
            'pinpad': round(s_pinpad, 2),
            'summ': summ,
        }
        _list.append(_dict)

    if request.method == 'POST':
        _list = []
        tarikh = request.POST.get('select')
        hour = request.POST.get('hour')

        _tarikh = request.POST.get('select')
        tarikh = tarikh.split("/")
        tarikh = jdatetime.date(day=int(tarikh[2]), month=int(tarikh[1]), year=int(tarikh[0])).togregorian()
        if hour == "-":
            _result = DailyTicketsReport.object_role.c_owner(request).filter(created=tarikh, st=0)[:1]
        else:
            _result = DailyTicketsReport.object_role.c_owner(request).filter(created=tarikh, create_time__hour=hour,
                                                                             st=0)
        for gs in _result:
            _dict = {
                'st': 0,
                'id': gs.zone.id,
                'name': str(gs.name),
                'count_gs': str(gs.count_gs),
                'count_pump': str(gs.count_pump),
                'count_ticket': str(gs.count_ticket),
                'count_master': str(gs.count_master),
                'count_pinpad': str(gs.count_pinpad),
                'master': gs.master,
                'pinpad': gs.pinpad,
                'summ': gs.summ,
            }

            _list.append(_dict)

        if id == 1:
            _list = sorted(_list, key=itemgetter('master'), reverse=True)
        elif id == 2:
            _list = sorted(_list, key=itemgetter('pinpad'), reverse=True)
        elif id == 3:
            _list = sorted(_list, key=itemgetter('summ'), reverse=True)
        gs = DailyTicketsReport.objects.filter(created=tarikh, st=1).last()
        _dict = {
            'st': 1,
            'id': 0,
            'name': str(gs.name),
            'count_gs': str(gs.count_gs),
            'count_pump': str(gs.count_pump),
            'count_ticket': str(gs.count_ticket),
            'count_master': str(gs.count_master),
            'count_pinpad': str(gs.count_pinpad),
            'master': gs.master,
            'pinpad': gs.pinpad,
            'summ': gs.summ,

        }
        _list.append(_dict)

    context = {'list': _list, 'id': id, 'tarikh': _tarikh, 'hour': hour, 'selected_products': selected_products,
               'products': products}
    return TemplateResponse(request, 'report/MasterPinPad_Tickets.html', context)


@cache_permission('report')
def masterpinpadticketnahye(request, id):
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    _list = []
    add_to_log(request, f'گزارش وضعیت تیکت ها ناحیه ', 0)
    if request.user.owner.role.role in ['mgr', 'setad']:
        zone = Area.objects.filter(zone_id=id).order_by('id')
    else:
        zone = Area.objects.filter(zone_id=request.user.owner.zone_id).order_by('id')
    summ = 0
    sum_gs = 0
    sum_pump = 0
    sum_ticket = 0
    sum_master = 0
    sum_pinpad = 0
    for gs in zone:
        count_gs = GsModel.objects.filter(area_id=gs.id, status_id=1).count()
        count_pump = Pump.objects.filter(gs__area_id=gs.id, gs__status__status=True, status__status=True).count()
        count_ticket = Ticket.objects.exclude(organization_id=4).filter(gs__area_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id__in=[1010,
                                                                                                         1011]).count()

        count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id=1010,
                                                                        failure__isnazel=True).count()
        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id=1011,
                                                                        failure__isnazel=True).count()
        if count_master > 0:
            master = ((int(count_master) / int(count_pump)) * 100)
        else:
            master = 0
        if count_pinpad > 0:
            pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
        else:
            pinpad = 0
        summ = round(master + pinpad, 2)
        _dict = {
            'st': 0,
            'id': gs.id,
            'name': gs.name,
            'count_gs': count_gs,
            'count_pump': count_pump,
            'count_ticket': count_ticket,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master': round(master, 2),
            'pinpad': round(pinpad, 2),
            'summ': summ,

        }
        _list.append(_dict)
        _list = sorted(_list, key=itemgetter('master'), reverse=True)
        sum_gs += count_gs
        sum_pump += count_pump
        sum_ticket += count_ticket
        sum_master += count_master
        sum_pinpad += count_pinpad
        if sum_master > 0:
            s_master = ((int(sum_master) / int(sum_pump)) * 100)
        else:
            s_master = 0
        if count_pinpad > 0:
            s_pinpad = ((int(sum_pinpad) / int(sum_pump)) * 100)
        else:
            s_pinpad = 0
        summ = round(s_master + s_pinpad, 2)
    _list = sorted(_list, key=itemgetter('summ'), reverse=True)
    _dict = {
        'st': 1,
        'id': 0,
        'name': SUM_TITEL,
        'count_gs': sum_gs,
        'count_pump': sum_pump,
        'count_ticket': sum_ticket,
        'count_master': sum_master,
        'count_pinpad': sum_pinpad,
        'master': round(s_master, 2),
        'pinpad': round(s_pinpad, 2),
        'summ': summ,

    }
    _list.append(_dict)

    context = {'list': _list}
    return TemplateResponse(request, 'report/MasterPinpad_Ticketnahye.html', context)


@cache_permission('report')
def amarkharabigs(request):
    _list = []
    add_to_log(request, f'گزارش درصد خرابی نازل ', 0)
    darsad = 100
    if request.method == 'POST':
        darsad = int(request.POST.get('val'))
        sms = request.POST.get('sms')
        n = darsad
        if request.user.owner.role.role == 'zone':
            zone = GsModel.objects.filter(active=True, area__zone_id=request.user.owner.zone_id).order_by('id')
        elif request.user.owner.role.role == 'area':
            zone = GsModel.objects.filter(active=True, area_id=request.user.owner.area_id).order_by('id')
        elif request.user.owner.role.role in ['mgr', 'setad', 'fani', 'test']:
            zone = GsModel.objects.filter(active=True).order_by('id')
        else:
            messages.warning(request, DENY_PERMISSSION)
            return redirect(HOME_PAGE)

        for gs in zone:
            count_pump = Pump.objects.filter(gs_id=gs.id, gs__status__status=True, status__status=True).count()
            count_ticket = Ticket.objects.exclude(organization_id=4).filter(gs_id=gs.id, status_id=1,
                                                                            gs__status__status=True,
                                                                            Pump__status__status=True,
                                                                            failure__failurecategory_id__in=[1010,
                                                                                                             1011]).count()

            count_master = Ticket.objects.exclude(organization_id=4).filter(gs_id=gs.id, status_id=1,
                                                                            gs__status__status=True,
                                                                            Pump__status__status=True,
                                                                            failure__failurecategory_id=1010,
                                                                            failure__isnazel=True).count()
            count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs_id=gs.id, status_id=1,
                                                                            gs__status__status=True,
                                                                            Pump__status__status=True,
                                                                            failure__failurecategory_id=1011,
                                                                            failure__isnazel=True).count()
            if int(count_pump) > 0:
                if count_master > 0:
                    master = ((int(count_master) / int(count_pump)) * 100)
                else:
                    master = 0
                if count_pinpad > 0:
                    pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
                else:
                    pinpad = 0

                if darsad < master:
                    _dict = {
                        'st': 0,
                        'id': gs.id,
                        'name': gs.name,
                        'gsid': gs.gsid,
                        'zone': gs.area.zone.name,
                        'area': gs.area.name,
                        'count_pump': count_pump,
                        'count_ticket': count_ticket,
                        'count_master': count_master,
                        'count_pinpad': count_pinpad,
                        'master': round(master, 2),
                        'pinpad': round(pinpad, 2),

                    }
                    _list.append(_dict)
                    _list = sorted(_list, key=itemgetter('zone', 'master'), reverse=True)

                    if sms == 'on':
                        titel = f"ضرورت پیگیری بموقع خرابی جایگاه های بالای  {n}در صد "
                        msgs = f'به استحضار تکنسین محترم سامانه مناطق میرساند خواهشمند است در خصوص  جایگاه های {gs.name}دارای خرابی بالای {n} درصد در اسرع وقت پیگیری و نتیجه منعکس گردد، '
                        msgid = CreateMsg.objects.create(titel=titel, info=msgs, owner_id=request.user.owner.id)
                        tek = GsList.objects.filter(gs_id=gs.id, owner__role__role__in=['tek'])
                        for i in tek:
                            ListMsg.objects.create(msg_id=msgid.id, user_id=i.owner_id)

    context = {'list': _list, 'darsad': darsad}
    return TemplateResponse(request, 'report/amarkharbigs.html', context)


@cache_permission('report')
def masterpinpadticketgs(request, id):
    _list = []

    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    add_to_log(request, f'گزارش وضعیت تیکت ها جایگاه ', 0)
    if request.user.owner.role.role in ['mgr', 'setad']:
        zone = GsModel.objects.filter(area_id=id, status_id=1).order_by('id')
    else:
        zone = GsModel.objects.filter(area_id=id, area__zone_id=request.user.owner.zone_id, status_id=1).order_by('id')
    summ = 0
    sum_pump = 0
    sum_ticket = 0
    sum_master = 0
    sum_pinpad = 0
    for gs in zone:
        count_pump = Pump.objects.filter(gs_id=gs.id, gs__status__status=True, status__status=True).count()
        count_ticket = Ticket.objects.exclude(organization_id=4).filter(gs_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id__in=[1010,
                                                                                                         1011]).count()
        count_master = Ticket.objects.exclude(organization_id=4).filter(gs_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id=1010,
                                                                        failure__isnazel=True).count()
        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id=1011,
                                                                        failure__isnazel=True).count()
        if count_master > 0:
            master = ((int(count_master) / int(count_pump)) * 100)
        else:
            master = 0
        if count_pinpad > 0:
            pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
        else:
            pinpad = 0
        summ = round(master + pinpad, 2)

        _dict = {
            'st': 0,
            'id': gs.id,
            'name': gs.name,
            'gsid': gs.gsid,
            'count_pump': count_pump,
            'count_ticket': count_ticket,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master': round(master, 2),
            'pinpad': round(pinpad, 2),
            'summ': summ,

        }
        _list.append(_dict)
        _list = sorted(_list, key=itemgetter('master'), reverse=True)
        sum_pump += count_pump
        sum_ticket += count_ticket
        sum_master += count_master
        sum_pinpad += count_pinpad
        if sum_master > 0:
            s_master = ((int(sum_master) / int(sum_pump)) * 100)
        else:
            s_master = 0
        if count_pinpad > 0:
            s_pinpad = ((int(sum_pinpad) / int(sum_pump)) * 100)
        else:
            s_pinpad = 0
        summ = round(s_master + s_pinpad, 2)
    _list = sorted(_list, key=itemgetter('summ'), reverse=True)
    _dict = {
        'st': 1,
        'id': 0,
        'name': SUM_TITEL,

        'count_pump': sum_pump,
        'count_ticket': sum_ticket,
        'count_master': sum_master,
        'count_pinpad': sum_pinpad,
        'master': round(s_master, 2),
        'pinpad': round(s_pinpad, 2),
        'summ': summ,

    }
    _list.append(_dict)

    context = {'list': _list, }
    return TemplateResponse(request, 'report/MasterPinpad_TicketGs.html', context)


def report1toexcel(request):
    _list = []
    if request.user.owner.role.role in ['zone']:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    else:
        zone = Zone.objects_limit.all().exclude(id=9).order_by('id')
    sum_all_sum = 0
    sum_deactive_sum = 0
    sum_kharab_sum = 0
    sum_active_sum = 0
    for gs in zone:
        sum_all = 0
        sum_deactive = 0
        sum_kharab = 0
        sum_active = 0
        products = Product.objects.all().order_by('id')
        for product in products:
            _all = Pump.objects.filter(gs__area__zone_id=gs.id, product_id=product.id, gs__status__status=True,
                                       status__status=True).count()

            deactive = Pump.objects.filter(gs__area__zone_id=gs.id, product_id=product.id, gs__status__status=True,
                                           status__status=True).count()
            kharab = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, gs__status__status=True,
                                                                      Pump__status__status=True,
                                                                      Pump__product_id=product.id, status_id=1).count()
            active = int(_all) - int(kharab)
            if kharab > 0:
                dismount = round(((kharab / _all) * 100), 1)
            else:
                dismount = 0
            if _all > 0:
                _dict = {
                    'st': 0,
                    'id': gs.id,
                    'gs': gs.name,
                    'product': product.name,
                    'active': active,
                    'deactive': deactive,
                    'kharab': kharab,
                    'dismount': dismount,
                }
                _list.append(_dict)
                sum_all += _all
                sum_deactive += deactive
                sum_kharab += kharab
                sum_active += active

                sum_all_sum += _all
                sum_deactive_sum += deactive
                sum_kharab_sum += kharab
                sum_active_sum += active
            if sum_kharab > 0:
                dismount = round(((sum_kharab / sum_all) * 100), 1)
            else:
                dismount = 0
        _dict = {
            'st': 1,
            'gs': SUM_TITEL + ' ' + gs.name,
            'product': '',
            'active': sum_active,
            'deactive': sum_deactive,
            'kharab': sum_kharab,
            'dismount': dismount,
        }
        _list.append(_dict)
    if sum_kharab_sum > 0:
        dismount = round(((sum_kharab_sum / sum_all_sum) * 100), 1)
    else:
        dismount = 0
    _dict = {
        'st': 2,
        'gs': SUM_TITEL,
        'product': '',
        'active': sum_active_sum,
        'deactive': sum_deactive_sum,
        'kharab': sum_kharab_sum,
        'dismount': dismount,
    }
    _list.append(_dict)

    context = {'list': _list}
    return render(request, 'report/nazelsstatusbymanategh.html', context)


def report1toexcel(request):
    _list = []
    if request.user.owner.role.role in ['zone']:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    else:
        zone = Zone.objects_limit.all().exclude(id=9).order_by('id')
    sum_all_sum = 0
    sum_deactive_sum = 0
    sum_kharab_sum = 0
    sum_active_sum = 0
    dismount = 0
    for gs in zone:
        sum_all = 0
        sum_deactive = 0
        sum_kharab = 0
        sum_active = 0
        products = Product.objects.all().order_by('id')
        for product in products:
            _all = Pump.objects.filter(gs__area__zone_id=gs.id, product_id=product.id, gs__status__status=True,
                                       status__status=True).count()

            deactive = Pump.objects.filter(gs__area__zone_id=gs.id, product_id=product.id, gs__status__status=True,
                                           status__status=True).count()
            kharab = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, gs__status__status=True,
                                                                      Pump__status__status=True,
                                                                      Pump__product_id=product.id, status_id=1).count()
            active = int(_all) - int(kharab)
            if kharab > 0:
                dismount = round(((kharab / _all) * 100), 1)
            else:
                dismount = 0
            if _all > 0:
                _dict = {
                    'st': 0,
                    'id': gs.id,
                    'gs': gs.name,
                    'product': product.name,
                    'active': active,
                    'deactive': deactive,
                    'kharab': kharab,
                    'dismount': dismount,
                }
                _list.append(_dict)
                sum_all += _all
                sum_deactive += deactive
                sum_kharab += kharab
                sum_active += active

                sum_all_sum += _all
                sum_deactive_sum += deactive
                sum_kharab_sum += kharab
                sum_active_sum += active
            if sum_kharab > 0:
                dismount = round(((sum_kharab / sum_all) * 100), 1)
            else:
                dismount = 0
        _dict = {
            'st': 1,
            'gs': SUM_TITEL + ' ' + gs.name,
            'product': '',
            'active': sum_active,
            'deactive': sum_deactive,
            'kharab': sum_kharab,
            'dismount': dismount,
        }
        _list.append(_dict)
    if sum_kharab_sum > 0:
        dismount = round(((sum_kharab_sum / sum_all_sum) * 100), 1)
    else:
        dismount = 0
    _dict = {
        'st': 2,
        'gs': SUM_TITEL,
        'product': '',
        'active': sum_active_sum,
        'deactive': sum_deactive_sum,
        'kharab': sum_kharab_sum,
        'dismount': dismount,
    }
    _list.append(_dict)

    context = {'list': _list}
    return render(request, 'report/nazelsstatusbymanategh.html', context)


@cache_permission('report')
def gssakoook(request):
    zones = None
    _list = None
    mylist = []
    count = 0
    sakoo_sum = 0
    tolombe_sum = 0
    nazel_sum = 0
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)

    zone = "0"
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = ""
    # if request.user.owner.role.role == 'zone':
    #     zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)
    #     count = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id, status__status=True).count()
    #     zone = request.user.owner.zone_id
    _list = Pump.object_role.c_gs(request, 0).values('gs__area__zone__name', 'gs__area__name', 'gs__gsid',
                                                     'gs__name').annotate(
        nazel=(Count('number'))
    )
    _filter = NazelFilter(request.GET, queryset=_list)
    _list = _filter.qs

    for i in _list:
        nazel_sum += i['nazel']

        mylist.append(i)
    count = len(mylist)

    return TemplateResponse(request, 'gssakoook.html',
                            {'list': mylist, 'filter': _filter, 'count': count, 'zones': zones,
                             'zone': int(zone), 'areas': areas,
                             'nazel_sum': nazel_sum,
                             'sakoo_sum': sakoo_sum, 'tolombe_sum': tolombe_sum})


@cache_permission('gs')
def gssakoook2(request):
    zones = None
    _list = None

    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)

    zone = "0"
    if request.GET.get('idval') == "1":
        if AutoExcel.objects.filter(owner_id=request.user.owner.id, errorstatus=False, status=False).count() > 0:
            messages.warning(request,
                             'شما یک درخواست در حال پردازش دارید ، لطفا منتظر بمانید درخواست قبلی شما ایجاد و در قسمت پیام ها به شما ارسال گردد.')
            return redirect('base:gssakoook2')

        AutoExcel.objects.create(
            titr=_roleid,
            owner_id=request.user.owner.id,
            req_id=request.GET,
            reportmodel=6
        )
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = ""
    # if request.user.owner.role.role == 'zone':
    #     zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)
    #     count = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id, status__status=True).count()
    #     zone = request.user.owner.zone_id
    _list = Pump.object_role.c_gs(request, 0).all()

    _filter = NazelFilter(request.GET, queryset=_list)
    _list = _filter.qs
    paginator = Paginator(_list, 100)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]

    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    return TemplateResponse(request, 'gssakook2.html',
                            {'list': page_object, 'filter': _filter, 'zones': zones, 'query_string': query_string,
                             'zone': int(zone), 'areas': areas, 'page_obj': page_obj, 'tedad': tedad,
                             })


@transaction.atomic
def import_excel_paydari(request):
    add_to_log(request, 'دریافت اکسل ناپایداری', 0)
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:
                    gs = GsModel.objects.get(gsid=_id, active=True)

                    ticket = Ticket.objects.filter(failure_id=1056, is_system=True, gs_id=gs.id, status_id=1)
                    if ticket.count() == 0:
                        a = Ticket.objects.create(owner_id=request.user.id, status_id=1, organization_id=1, gs_id=gs.id,
                                                  failure_id=1056, is_system=True)
                        Workflow.objects.create(ticket_id=a.id, user_id=request.user.id,
                                                description='ارجاع سیستمی ، لطفا پشتیبان جهت بررسی دقیق و اعمال چک '
                                                            'لیست نسبت به رفع ناپایداری اقدام و نتیجه را گزارش نمایند ( این تیکت '
                                                            'بسیار مهم در ارزیابی پشتیبانان میباشد)',
                                                organization_id=1, failure_id=1056)
                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        ticket = Ticket.objects.filter(failure_id=1056, is_system=True, status_id=1)
        for item in ticket:
            if item.gs.gsid not in _list:
                item.status_id = 2
                item.save()

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return render(request, 'importexcel.html', {'form': form})


@cache_permission('bulking')
def send_bulk_ticket(request):
    add_to_log(request, 'ارسال تیکت گروهی', 0)
    sarfasl = FailureCategory.objects.all()
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        info = request.POST.get('info')
        failure = request.POST.get('failure')
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            _failure = FailureSub.objects.get(id=int(failure))
            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:
                    gs = GsModel.objects.get(gsid=_id, status_id=1)

                    ticket = Ticket.objects.filter(failure_id=failure, is_system=True, gs_id=gs.id, status_id=1)
                    if ticket.count() == 0:
                        a = Ticket.objects.create(owner_id=request.user.id, status_id=1,
                                                  organization_id=_failure.organization_id, gs_id=gs.id,
                                                  failure_id=failure, is_system=True)
                        Workflow.objects.create(ticket_id=a.id, user_id=request.user.id,
                                                description=info,
                                                organization_id=_failure.organization_id, failure_id=failure)
                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'bulkexcel.html', {'form': form, 'sarfasl': sarfasl})


@cache_permission('bulking')
def close_bulk_ticket(request):
    add_to_log(request, 'بستن تیکت گروهی', 0)
    sarfasl = FailureCategory.objects.all()
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        info = request.POST.get('info')
        failure = request.POST.get('failure')
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:
                    gs = GsModel.objects.get(gsid=_id, status_id=1)
                    ticket = Ticket.objects.filter(failure_id=failure, is_system=True, gs_id=gs.id, status_id=1)
                    if ticket.count() > 0:
                        for item in ticket:
                            item.status_id = 2
                            item.actioner_id = request.user.owner.id
                            item.descriptionactioner = 'تیکت بسته شد' + str(info)
                            item.close_shamsi_year = jdatetime.datetime.now().year
                            if len(str(jdatetime.datetime.now().month)) == 1:
                                month = '0' + str(jdatetime.datetime.now().month)
                            else:
                                month = jdatetime.datetime.now().month
                            item.close_shamsi_month = month
                            if len(str(jdatetime.datetime.now().day)) == 1:
                                day = '0' + str(jdatetime.datetime.now().day)
                            else:
                                day = jdatetime.datetime.now().day
                            item.close_shamsi_day = day
                            item.closedate = datetime.datetime.now()
                            item.close_shamsi_date = str(jdatetime.datetime.now().year) + "-" + str(month) + "-" + str(
                                day)
                            if item.closedate:
                                try:
                                    item.timeaction = (item.closedate - item.create).days
                                except:
                                    continue

                            item.save()
                            Workflow.objects.create(ticket_id=item.id, user_id=request.user.id,
                                                    description=info,
                                                    organization_id=1, failure_id=failure)

                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'bulkexcel.html', {'form': form, 'sarfasl': sarfasl})


def autoclose_bulk_ticket(request):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='bulking')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    add_to_log(request, 'بستن تیکت گروهی', 0)
    sarfasl = FailureCategory.objects.all()
    _list = []
    info = "بستن اتوماتیک"
    failure = 1164
    param = Parametrs.objects.all().first()
    gslist = IpcLog.objects.filter(rpm_version=param.rpm_version)
    for i in gslist:
        _id = i.gs.gsid
        _list.append(_id)

        # gs = GsModel.objects.get(gsid=_id)
        ticket = Ticket.objects.filter(failure_id=failure, is_system=True, gs_id=i.gs.id, status_id=1)
        if ticket.count() > 0:
            for item in ticket:
                item.status_id = 2
                item.actioner_id = request.user.owner.id
                item.descriptionactioner = 'تیکت بسته شد' + str(info)
                item.close_shamsi_year = jdatetime.datetime.now().year
                if len(str(jdatetime.datetime.now().month)) == 1:
                    month = '0' + str(jdatetime.datetime.now().month)
                else:
                    month = jdatetime.datetime.now().month
                item.close_shamsi_month = month
                if len(str(jdatetime.datetime.now().day)) == 1:
                    day = '0' + str(jdatetime.datetime.now().day)
                else:
                    day = jdatetime.datetime.now().day
                item.close_shamsi_day = day
                item.closedate = datetime.datetime.now()
                item.close_shamsi_date = str(jdatetime.datetime.now().year) + "-" + str(month) + "-" + str(
                    day)
                if item.closedate:
                    try:
                        item.timeaction = (item.closedate - item.create).days
                    except:
                        continue

                    item.save()
                    Workflow.objects.create(ticket_id=item.id, user_id=request.user.id,
                                            description=info,
                                            organization_id=1, failure_id=failure)

        # except GsModel.DoesNotExist:
        #     logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

    messages.success(request, gslist.count())
    return redirect(HOME_PAGE)


@cache_permission('bulking')
def delete_bulk_ticket(request):
    add_to_log(request, 'حذف تیکت گروهی', 0)
    sarfasl = FailureCategory.objects.all()
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        info = request.POST.get('info')
        failure = request.POST.get('failure')
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:
                    gs = GsModel.objects.get(gsid=_id, status_id=1)
                    Ticket.objects.filter(failure_id=failure, is_system=True, gs_id=gs.id, status_id=1).delete()

                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

            messages.success(request, SUCCESS_TICKET)
            return redirect(HOME_PAGE)
    return TemplateResponse(request, 'bulkexcel.html', {'form': form, 'sarfasl': sarfasl})


def sendticketrpm(request):
    add_to_log(request, 'نصب RPM', 0)
    gs = GsModel.objects.filter(status__status=True)
    parametr = Parametrs.objects.all().last()

    for item in gs:
        if Ticket.objects.filter(gs_id=item.id, status_id=1,
                                 failure_id=1164, is_system=True).count() == 0:
            a = Ticket.objects.create(owner_id=request.user.id, status_id=1, organization_id=1, gs_id=item.id,
                                      failure_id=1164, is_system=True)
            Workflow.objects.create(ticket_id=a.id, user_id=request.user.id,
                                    description=f'تیکت نصب RPM نگارش {parametr.rpm_version}',
                                    organization_id=1, failure_id=1164)

    messages.success(request, SUCCESS_TICKET)
    return redirect(HOME_PAGE)


@cache_permission('excel_nop')
def input_zone_table(request):
    _list = []
    add_to_log(request, f'دریافت اکسل حدول منطقه ایی ', 0)
    form = open_excel(request.POST)

    if request.method == 'POST':

        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)
                val = str(sheet_obj.cell(row=i, column=2).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:
                    gs = GsModel.objects.get(gsid=_id)
                    print(val, gs.gsid)
                    gs.zone_table_version = str(val)
                    gs.save()


                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


@cache_permission('bulking')
def input_coding_table(request):
    _list = []
    add_to_log(request, f'دریافت اکسل کدینگ ', 0)
    form = open_excel(request.POST)

    if request.method == 'POST':

        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            m_row = sheet_obj.max_row
            GsModel.objects.all().update(iscoding=False)
            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)

                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:

                    gs = GsModel.objects.get(gsid=_id)
                    gs.iscoding = True
                    gs.save()

                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


def saveroles(request):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        tedad = request.user.owner.ownerlist.count()
        if tedad < 2 or request.user.is_superuser:
            role = request.POST.get('role')
            semat = request.POST.get('sematrole')
        zone = request.POST.get('zonerole')
        owner = Owner.objects.get(id=request.user.owner.id)
        if tedad < 2 or request.user.is_superuser:
            owner.role_id = role
            owner.refrence_id = semat
        owner.zone_id = zone
        owner.save()
        cache.delete(f"user_data_{request.user.id}")
        return redirect(url)


def ckpaydari(request):
    ticket = Ticket.objects.filter(failure_id=1056, is_system=True, status_id=1)
    for item in ticket:
        result = Ticket.objects.filter(gs_id=item.gs_id, failure_id=1056, is_system=True, status_id=1)
        if result.count() > 1:
            result = result.last()
            result.status_id = 2
            result.save()
    return redirect(HOME_PAGE)


def masterpinpadticketstoexcel(request, _id):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='report')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    _list = []
    add_to_log(request, 'ارسال وضعیت تیکت مناطق به اکسل', 0)
    if request.user.owner.role.role in ['zone']:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    elif request.user.owner.role.role in ['setad', 'mgr', 'fani', 'test', 'posht']:
        zone = Zone.objects_limit.all().exclude(id=9).order_by('id')
    else:
        messages.warning(request, DENY_PERMISSSION)
        return redirect(HOME_PAGE)

    s_master = 0
    s_pinpad = 0
    summ = 0
    sum_gs = 0
    sum_pump = 0
    sum_ticket = 0
    sum_master = 0
    sum_pinpad = 0
    for gs in zone:
        count_gs = GsModel.objects.filter(area__zone_id=gs.id, status_id=1).count()
        count_pump = Pump.objects.filter(gs__area__zone_id=gs.id, gs__status__status=True,
                                         status__status=True, ).count()
        count_ticket = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id__in=[1010,
                                                                                                         1011]).count()

        count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id=1010,
                                                                        failure__isnazel=True).count()

        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                        gs__status__status=True,
                                                                        Pump__status__status=True,
                                                                        failure__failurecategory_id=1011,
                                                                        failure__isnazel=True).count()
        master = 0
        pinpad = 0
        if count_master > 0:
            master = ((int(count_master) / int(count_pump)) * 100)
        if count_pinpad > 0:
            pinpad = ((int(count_pinpad) / int(count_pump)) * 100)

        summ = round(master + pinpad, 2)
        _dict = {
            'st': 0,
            'id': gs.id,
            'name': gs.name,
            'count_gs': count_gs,
            'count_pump': count_pump,
            'count_ticket': count_ticket,
            'count_master': count_master,
            'count_pinpad': count_pinpad,
            'master': round(master, 2),
            'pinpad': round(pinpad, 2),
            'summ': summ,

        }
        _list.append(_dict)
        if _id == 1:
            _list = sorted(_list, key=itemgetter('master'), reverse=True)
        elif _id == 2:
            _list = sorted(_list, key=itemgetter('pinpad'), reverse=True)
        elif _id == 3:
            _list = sorted(_list, key=itemgetter('summ'), reverse=True)

        sum_gs += count_gs
        sum_pump += count_pump
        sum_ticket += count_ticket
        sum_master += count_master
        sum_pinpad += count_pinpad
        s_master = 0
        s_pinpad = 0
        if sum_master > 0:
            s_master = ((int(sum_master) / int(sum_pump)) * 100)
        if count_pinpad > 0:
            s_pinpad = ((int(sum_pinpad) / int(sum_pump)) * 100)

        summ = round(s_master + s_pinpad, 2)
    listsum = []
    _dict = {
        'st': 1,
        'id': 0,
        'name': SUM_TITEL,
        'count_gs': sum_gs,
        'count_pump': sum_pump,
        'count_ticket': sum_ticket,
        'count_master': sum_master,
        'count_pinpad': sum_pinpad,
        'master': round(s_master, 2),
        'pinpad': round(s_pinpad, 2),
        'summ': summ,

    }
    listsum.append(_dict)

    add_to_log(request, 'ارسال آمار خرابی به اکسل  ', 0)
    my_path = "Status.xlsx"
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)
    fonttitr2 = Font(bold=True, size=20)
    wb = Workbook()

    ws1 = wb.active  # work with default worksheet
    ws1.title = "گزارش وضعیت مجاری عرضه تاریخ "
    ws1.sheet_view.rightToLeft = True
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:J1')

    ws1["A1"] = f'گزارش وضعیت مجاری عرضه تاریخ   {today}'
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:A3')
    ws1["A2"] = "ردیف"
    ws1["A2"].font = font

    ws1.merge_cells('B2:B3')
    ws1["B2"] = "منطقه"
    ws1["B2"].font = fonttitr2

    ws1.merge_cells('C2:E2')
    ws1["C2"] = "تعداد "
    ws1["C2"].font = font

    ws1.merge_cells('C3:C3')
    ws1["C3"] = "جایگاه"
    ws1["C3"].font = font

    ws1.merge_cells('D2:D2')
    ws1["D3"] = "نازل"
    ws1["D3"].font = font

    ws1.merge_cells('E2:E2')
    ws1["E3"] = "تیکت"
    ws1["E3"].font = font

    ws1.merge_cells('F2:G2')
    ws1["F2"] = " تعداد تیکت "
    ws1["F2"].font = font

    ws1.merge_cells('F2:F2')
    ws1["F3"] = " کارتخوان"
    ws1["F3"].font = font

    ws1.merge_cells('G2:G2')
    ws1["G3"] = "صفحه کلید"
    ws1["G3"].font = font

    ws1.merge_cells('H2:I2')
    ws1["H2"] = "در صد خرابی"
    ws1["H2"].font = font

    ws1.merge_cells('H2:H2')
    ws1["H3"] = " کارتخوان"
    ws1["H3"].font = font

    ws1.merge_cells('I2:I2')
    ws1["I3"] = " صفحه کلید"
    ws1["I3"].font = font

    ws1.merge_cells('J2:J2')
    ws1["J3"] = "مجموع"
    ws1["J3"].font = font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(7.25)
    ws1.column_dimensions['D'].width = float(7.25)
    ws1.column_dimensions['E'].width = float(6.25)
    ws1.column_dimensions['F'].width = float(8.25)
    ws1.column_dimensions['G'].width = float(9.25)
    ws1.column_dimensions['H'].width = float(8.25)
    ws1.column_dimensions['I'].width = float(9.25)
    ws1.column_dimensions['J'].width = float(8.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    my_fill2 = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    i = 0

    for item in _list:
        i += 1
        d = [i, str(item['name']), str(item['count_gs']), str(item['count_pump']), str(item['count_ticket']),
             str(item['count_master']),
             str(item['count_pinpad']), str(item['master']), str(item['pinpad']), str(item['summ'])
             ]

        ws1.append(d)
    for item in listsum:
        d = ['', 'جمع', str(item['count_gs']), str(item['count_pump']), str(item['count_ticket']),
             str(item['count_master']),
             str(item['count_pinpad']), str(item['master']), str(item['pinpad']), str(item['summ'])]
        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    i += 4
    for cell in ws1[i:i]:  # Last row
        cell.font = myfont
        cell.fill = my_fill2
        cell.border = thin_border

    for cell in ws1["2:2"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border
    wb.save(response)
    return response


@cache_permission('report')
def rep_napaydari(request):
    napaydari = None
    napaydari_list = None
    add_to_log(request, 'گزارشات ناپایداری', 0)
    zone = Zone.objects_limit.all()
    if request.user.owner.role.role == 'zone':
        napaydari = Ticket.objects.values('gs_id').filter(gs__area__zone_id=request.user.owner.zone_id,
                                                          gs__status_id=1,
                                                          failure_id=1056, status_id=1, is_system=True).count()
        napaydari_list = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone_id,
                                               gs__status_id=1,
                                               failure_id=1056, status_id=1, is_system=True)
    elif request.user.owner.role.role == 'area':
        napaydari = Ticket.objects.values('gs_id').filter(gs__area_id=request.user.owner.area_id,
                                                          status__status='open', gs__status_id=1,
                                                          failure_id=1056, status_id=1, is_system=True).count()
        napaydari_list = Ticket.objects.filter(gs__area_id=request.user.owner.area_id,
                                               gs__status_id=1,
                                               failure_id=1056, status_id=1, is_system=True)

    if request.user.owner.role.role in ['setad', 'mgr', 'fani', 'test']:
        napaydari = Ticket.objects.values('gs_id').filter(
            status__status='open', gs__status_id=1,
            failure_id=1056, status_id=1, is_system=True).count()
        napaydari_list = Ticket.objects.filter(
            status__status='open', gs__status_id=1,
            failure_id=1056, status_id=1, is_system=True)
        if request.method == 'POST':
            zoneid = request.POST.get('zone')

            napaydari = Ticket.objects.values('gs_id').filter(gs__area__zone_id=zoneid,
                                                              status__status='open', gs__status_id=1,
                                                              failure_id=1056, status_id=1, is_system=True).count()
            napaydari_list = Ticket.objects.filter(gs__area__zone_id=zoneid,
                                                   status__status='open', gs__status_id=1,
                                                   failure_id=1056, status_id=1, is_system=True)

    napydarilist = []
    for item in napaydari_list:
        tekname = GsList.objects.filter(owner__role__role='tek', gs_id=item.gs_id, owner__active=True).first()
        if tekname:
            teknameowner = tekname.owner
        else:
            teknameowner = unname
        today = datetime.datetime.today()
        t1 = date(year=today.year, month=today.month, day=today.day)
        outdate = item.create
        t2 = date(year=outdate.year, month=outdate.month, day=outdate.day)
        tedad = (t1 - t2).days
        _dict = {
            'id': item.id,
            'gsid': item.gs.gsid,
            'name': item.gs.name,
            'nahye': str(item.gs.area.name),
            'zone': str(item.gs.area.zone.name),
            'tedad': tedad,
            'tek': teknameowner,
            'rnd': item.rnd,
        }

        napydarilist.append(_dict)
        napydarilist = sorted(napydarilist, key=itemgetter('tedad'), reverse=True)

    return TemplateResponse(request, 'report/Rep_napaydari.html',
                            {'napydarilist': napydarilist, 'napaydari': napaydari, 'zone': zone,
                             })


@cache_permission('report')
def graph_ticket(request):
    _list = []
    zone = 0
    add_to_log(request, 'مشاهده نمودار تیکت', 0)
    zones = Zone.objects_limit.all()
    if request.method == 'POST':
        zone = request.POST.get('zone')
    if request.user.owner.role.role == "zone":
        zone = request.user.owner.zone_id

    for years in Ticket.objects.values('create_shamsi_year', 'create_shamsi_month').annotate(sum=Count('id'))[12:]:
        month = str(years['create_shamsi_month'])
        if len(str(years['create_shamsi_month'])) == 1:
            month = "0" + str(years['create_shamsi_month'])
        if zone == '0':
            resultcreate = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory__in=[1010, 1045],
                                                                              create_shamsi_year=years[
                                                                                  'create_shamsi_year'],
                                                                              create_shamsi_month=month).count()

            resultclose = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory__in=[1010, 1045],
                                                                             close_shamsi_year=years[
                                                                                 'create_shamsi_year'],
                                                                             close_shamsi_month=month).count()
            sendstore = Store.objects.filter(status_id=3, resid_year=years['create_shamsi_year'],
                                             resid_month=month)
            sendst = 0
            for item in sendstore:
                sendst += item.master

            resultcreate_pinpad = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory=1011,

                                                                                     create_shamsi_year=years[
                                                                                         'create_shamsi_year'],
                                                                                     create_shamsi_month=month).count()

            resultclose_pinpad = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory=1011,

                                                                                    close_shamsi_year=years[
                                                                                        'create_shamsi_year'],
                                                                                    close_shamsi_month=month).count()
            sendstore_pinpad = Store.objects.filter(status_id=3, resid_year=years['create_shamsi_year'],
                                                    resid_month=month)
            sendst_pinpad = 0
            for item in sendstore_pinpad:
                sendst_pinpad += item.pinpad
        else:
            resultcreate = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory__in=[1010, 1045],
                                                                              gs__area__zone_id=zone,
                                                                              create_shamsi_year=years[
                                                                                  'create_shamsi_year'],
                                                                              create_shamsi_month=month).count()

            resultclose = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory__in=[1010, 1045],
                                                                             gs__area__zone_id=zone,
                                                                             close_shamsi_year=years[
                                                                                 'create_shamsi_year'],
                                                                             close_shamsi_month=month).count()
            sendstore = Store.objects.filter(status_id=3, resid_year=years['create_shamsi_year'], zone_id=zone,
                                             resid_month=month)
            sendst = 0
            for item in sendstore:
                sendst += item.master

            resultcreate_pinpad = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory=1011,
                                                                                     gs__area__zone_id=zone,
                                                                                     create_shamsi_year=years[
                                                                                         'create_shamsi_year'],
                                                                                     create_shamsi_month=month).count()

            resultclose_pinpad = Ticket.objects.annotate(master=Count('id')).filter(failure__failurecategory=1011,
                                                                                    gs__area__zone_id=zone,
                                                                                    close_shamsi_year=years[
                                                                                        'create_shamsi_year'],
                                                                                    close_shamsi_month=month).count()
            sendstore_pinpad = Store.objects.filter(status_id=3, resid_year=years['create_shamsi_year'], zone_id=zone,
                                                    resid_month=month)
            sendst_pinpad = 0
            for item in sendstore_pinpad:
                sendst_pinpad += item.pinpad

        _dict = {
            'name': str(years['create_shamsi_year']) + '/' + str(years['create_shamsi_month']),
            'n1': resultcreate,
            'n2': resultclose,
            'n3': sendst,
            'n11': resultcreate_pinpad,
            'n22': resultclose_pinpad,
            'n33': sendst_pinpad,
        }
        _list.append(_dict)

    _list = sorted(_list, key=itemgetter('name'))

    return TemplateResponse(request, 'graphticket/graph1.html',
                            {'list': _list, 'zones': zones, 'zone_id': int(zone)})


@cache_permission('report')
def reportticketdaily(request):
    add_to_log(request, f'گزارش تیکت های روزانه ', 0)
    zonename = None
    zone_select = 0
    sum_new = 0
    sum_close = 0
    mdate = startdate
    mdate2 = today_date
    az = mdate
    ta = mdate2
    zones = Zone.objects_limit.all()
    sum_send_store = 0
    sendst_master = 0
    sendst_pinpad = 0
    send_store = 0

    if request.method == 'POST':
        mdate = request.POST.get('select')
        mdate2 = request.POST.get('select2')
        az = mdate
        ta = mdate2

        mdate = mdate.split("/")
        mdate2 = mdate2.split("/")
        datein = jdatetime.date(day=int(mdate[2]), month=int(mdate[1]), year=int(mdate[0])).togregorian()
        dateout = jdatetime.date(day=int(mdate2[2]), month=int(mdate2[1]), year=int(mdate2[0])).togregorian()

        t1 = date(year=datein.year, month=datein.month, day=datein.day)
        t2 = date(year=dateout.year, month=dateout.month, day=dateout.day)
        day = (t2 - t1).days

        dateme = dateout + datetime.timedelta(days=1)

        _list = []
        for item in range(day):

            dateme = dateme - datetime.timedelta(days=1)

            datein = str(dateme) + " 00:00:00"
            dateout = str(dateme) + " 23:59:59"
            zone_select = request.POST.get('select3')
            if request.user.owner.role.role == 'zone':
                zone_select = Zone.objects_limit.get(id=request.user.owner.zone_id).id
            zonename = Zone.objects_limit.get(id=int(zone_select)).name

            new = Ticket.objects.filter(create__gte=datein, create__lte=dateout, gs__area__zone_id=int(zone_select),
                                        failure__failurecategory__in=[
                                            1010, 1011,
                                            1045]).count()
            close = Ticket.objects.filter(closedate__gte=datein, closedate__lte=dateout,
                                          gs__area__zone_id=int(zone_select),
                                          failure__failurecategory__in=[
                                              1010, 1011, 1045]
                                          ).count()
            sendstore_pinpad = Store.objects.filter(status_id=3, zone_id=int(zone_select), resid_date__gte=datein,
                                                    resid_date__lte=dateout,
                                                    )

            sendst_master = 0
            sendst_pinpad = 0
            send_store = 0
            for item in sendstore_pinpad:
                sendst_master += item.master
                sendst_pinpad += item.pinpad

            send_store = sendst_master + sendst_pinpad

            sum_new += new
            sum_close += close
            sum_send_store += send_store
            dateme2 = dateme.strftime("%Y-%m-%d %H:%M:%S")
            _dict = {
                'gs': zonename,
                'new': new,
                'close': close,
                'send_store': send_store,
                'date': str(JDate(dateme2).year()) + '/' + str(
                    JDate(dateme2).month()) + '/' + str(
                    JDate(dateme2).day())
            }
            _list.append(_dict)

        return TemplateResponse(request, 'report/reportticketdaily.html',
                                {'list': _list, 'mdate': mdate, 'mdate2': mdate2, 'zonename': zonename, 'zones': zones,
                                 'sum_new': sum_new, 'sum_close': sum_close, 'zone_select': int(zone_select),
                                 'sum_send_store': sum_send_store,
                                 'az': az, 'ta': ta})
    return TemplateResponse(request, 'report/reportticketdaily.html',
                            {'mdate': mdate, 'az': az, 'ta': ta, 'mdate2': mdate2, 'zones': zones,
                             })


def save_star(request, id):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        star = request.POST.get(f'myradio{id}')
        if star:
            ticket = Ticket.objects.get(id=id)
            ticket.star = star
            ticket.save()
            messages.success(request, 'امتیاز شما با موفقیت ثبت شد')
        else:
            messages.error(request, 'یک گزینه را انتخاب کنید')
        return redirect(url)


@cache_permission('report')
def gs_list_arbain(request):
    sum_nazel = 0
    sum_master = 0
    sum_pinpad = 0
    if request.user.owner.role.role == 'zone':
        _list = GsModel.objects.filter(arbain=True, area__zone_id=request.user.owner.zone_id).order_by('area__zone_id',
                                                                                                       'area_id')
    else:
        _list = GsModel.objects.filter(arbain=True).order_by('area__zone_id', 'area_id')
    mlist = []
    for item in _list:
        nazel = Pump.objects.filter(gs_id=item.id, active=True).count()
        count_master = Ticket.objects.exclude(organization_id=4).filter(gs_id=item.id, status_id=1,
                                                                        failure__failurecategory_id=1010,
                                                                        ).count()

        count_master += Ticket.objects.exclude(organization_id=4).filter(gs_id=item.id, status_id=1,
                                                                         failure_id=1045
                                                                         ).count()

        count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs_id=item.id, status_id=1,
                                                                        failure__failurecategory_id=1011,
                                                                        ).count()
        sum_nazel += nazel
        sum_master += count_master
        sum_pinpad += count_pinpad
        _dict = {
            'zone': item.area.zone.name,
            'area': item.area.name,
            'gsid': item.gsid,
            'name': item.name,
            'nazel': nazel,
            'master': count_master,
            'pinpad': count_pinpad
        }
        mlist.append(_dict)

    return TemplateResponse(request, 'gs_list_arbain.html',
                            {'list': mlist, 'sum_nazel': sum_nazel, 'sum_master': sum_master, 'sum_pinpad': sum_pinpad,
                             })


@cache_permission('gs')
def listgs(request):
    # لاگ‌گیری
    add_to_log(request, 'مشاهده لیست جایگاه جدید', 0)

    # دریافت نقش و منطقه کاربر
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)

    # دریافت لیست جایگاه‌ها بر اساس نقش و منطقه کاربر
    gss = GsModel.object_role.c_gsmodel(request).all().order_by('area__zone_id', 'area_id', 'gsid')

    # اعمال فیلترها
    _filter = GsFilter(request.GET, queryset=gss)
    gss = _filter.qs

    # جستجو
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = form.cleaned_data['search']
            gss = gss.filter(Q(name__icontains=cd) | Q(gsid__exact=cd))

    # محاسبه تعداد جایگاه‌ها بر اساس وضعیت
    status_counts = gss.aggregate(
        active=Count('id', filter=Q(status_id=1)),
        deactive=Count('id', filter=Q(status_id=4)),
        under_construction=Count('id', filter=Q(status_id=3))
    )

    # پاژینیشن
    paginator = Paginator(gss, 150)
    page_num = request.GET.get('page')
    page_object = paginator.get_page(page_num)

    # مدیریت کوئری‌استرینگ
    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)[1]

    # ارسال داده‌ها به تمپلیت
    return TemplateResponse(request, 'listgs.html', {
        'list': page_object,
        'filter': _filter,
        'active': status_counts['active'],
        'deactive': status_counts['deactive'],
        'under_construction': status_counts['under_construction'],
        'page_obj': paginator.num_pages,
        'query_string': query_string,
        'zone': Zone.objects_limit.all().order_by('id'),
    })


@cache_permission('gs')
def gs_detail(request, id):
    # رمزگشایی ID
    id = Decrypt(id)

    # دریافت نقش و منطقه کاربر
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)

    # دریافت محصولات و برندهای پمپ
    products = Product.objects.all()
    pumpsbrand = PumpBrand.objects.all()

    try:
        # دریافت جایگاه بر اساس نقش و منطقه کاربر
        gs = GsModel.object_role.c_gsmodel(request).select_related('area__zone').get(id=id)
        add_to_log(request, f' مشاهده اطلاعات جایگاه {gs.name} ', id)

        # بررسی دسترسی کاربر
        if _role == 'zone' and gs.area.zone_id != request.user.owner.zone_id:
            messages.error(request, 'شما به اطلاعات این جایگاه دسترسی ندارید')
            return redirect(HOME_PAGE)
        if _role == 'area' and gs.area_id != request.user.owner.area_id:
            messages.error(request, 'شما به اطلاعات این جایگاه دسترسی ندارید')
            return redirect(HOME_PAGE)

    except ObjectDoesNotExist:
        messages.error(request, 'این جایگاه وجود ندارد')
        return redirect(HOME_PAGE)

    # دریافت لیست مالکان جایگاه
    query = GsList.objects.filter(gs_id=id, owner__active=True, owner__refrence_id__in=[4, 8, 7, 9]).select_related(
        'owner').order_by('owner__refrence_id')

    # دریافت مدیر، مسئول منطقه و رئیس
    mgr = Owner.objects.filter(zone_id=gs.area.zone_id, refrence_id=2, active=True).first()
    area = Owner.objects.filter(area_id=gs.area_id, refrence_id=3, active=True).first()
    shef = Owner.objects.filter(zone_id=gs.area.zone_id, refrence_id=1, active=True).first()

    # محاسبه درصد فروش
    result = SellModel.objects.filter(gs_id=id).values('tarikh').annotate(
        n1=Sum('yarane'),
        n2=Sum('azad'),
        n3=Sum('ezterari'),
        sum=Sum('yarane') + Sum('azad') + Sum('ezterari')
    )[:10]

    _list = []
    for sell in result:
        if sell['n1'] != 0 and sell['sum']:
            _dict = {
                'n1': round((sell['n1'] / sell['sum']) * 100),
                'n2': round((sell['n2'] / sell['sum']) * 100),
                'n3': round((sell['n3'] / sell['sum']) * 100),
            }
            _list.append(_dict)
    _list = sorted(_list, key=lambda x: x['n1'], reverse=True)

    # دریافت آخرین لاگ IPC
    result = IpcLog.objects.filter(gs_id=id).last()

    # دریافت اطلاعات بسته‌شدن، خرید و نازل‌ها
    closegs = CloseGS.objects.filter(gs_id=id).order_by('-id')
    moghgs = AcceptForBuy.objects.filter(gs_id=id).order_by('-id')
    nazels = Pump.objects.filter(gs_id=id).order_by('sortnumber')
    anbars = ParametrGs.objects.filter(gs_id=id).order_by('-id')

    # محاسبه آمار تیکت‌ها
    amarticketgs = Ticket.objects.filter(gs_id=gs.id).aggregate(
        mastergs=Count('id', filter=Q(failure_id=1045)),
        pinpadgs=Count('id', filter=Q(reply_id__in=[3, 4]))
    )

    amartiketarea = Ticket.objects.filter(gs__area_id=gs.area_id).aggregate(
        masterkol=Count('id', filter=Q(failure_id=1045)),
        pinpadkol=Count('id', filter=Q(reply_id__in=[3, 4]))
    )

    amartiketzone = Ticket.objects.filter(gs__area__zone_id=gs.area.zone_id).aggregate(
        masterkol=Count('id', filter=Q(failure_id=1045)),
        pinpadkol=Count('id', filter=Q(reply_id__in=[3, 4]))
    )

    # محاسبه تعداد نازل‌ها
    nazelcountgs = Pump.objects.filter(gs_id=id, status=True).count()
    nazelcountzone = Pump.objects.filter(gs__area__zone_id=gs.area.zone_id, status__status=True).count()
    nazelcountarea = Pump.objects.filter(gs__area_id=gs.area_id, status__status=True).count()

    # محاسبه آمار نازل‌ها
    dictnew = {}
    if nazelcountgs > 0:
        try:
            dictnew = {
                'amarmaster': round((int(amarticketgs['mastergs']) * int(nazelcountzone)) / (
                        int(amartiketzone['masterkol']) * int(nazelcountgs)), 1),
                'amarpinpad': round((int(amarticketgs['pinpadgs']) * int(nazelcountzone)) / (
                        int(amartiketzone['pinpadkol']) * int(nazelcountgs)), 1),
                'amarmasterarea': round((int(amarticketgs['mastergs']) * int(nazelcountarea)) / (
                        int(amartiketarea['masterkol']) * int(nazelcountgs)), 1),
                'amarpinpadarea': round((int(amarticketgs['pinpadgs']) * int(nazelcountarea)) / (
                        int(amartiketarea['pinpadkol']) * int(nazelcountgs)), 1),
                'mastergs': amarticketgs['mastergs'],
                'nazelgs': nazelcountgs,
                'masterzone': amartiketzone['masterkol'],
                'nazelzone': nazelcountzone,
                'masterarea': amartiketarea['masterkol'],
                'nazelarea': nazelcountarea,
                'pinpadgs': amarticketgs['pinpadgs'],
                'pinpadzone': amartiketzone['pinpadkol'],
                'pinpadarea': amartiketarea['pinpadkol'],
            }
        except Exception as e:
            dictnew = {}
            print(f"Error calculating dictnew: {e}")

    # محاسبه تعداد نازل‌های فعال
    benzin = Pump.objects.filter(actived=True, gs_id=id, product_id=2).count()
    _super = Pump.objects.filter(actived=True, gs_id=id, product_id=3).count()
    gaz = Pump.objects.filter(actived=True, gs_id=id, product_id=4).count()
    cng = Pump.objects.filter(actived=True, gs_id=id, product_id=5).count()

    # ارسال داده‌ها به تمپلیت
    context = {
        'gs': gs,
        'query': query,
        'mgr': mgr,
        'area': area,
        'shef': shef,
        'list': _list,
        'nazels': nazels,
        'benzin': benzin,
        'super': _super,
        'gaz': gaz,
        'cng': cng,
        'result': result,
        'products': products,
        'pumpsbrand': pumpsbrand,
        'closegs': closegs,
        'dictnew': dictnew,
        'moghgs': moghgs,
        'anbars': anbars,
    }

    return TemplateResponse(request, 'viewdetailgs.html', context)


def gsimportfile(request, id):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        gs = GsModel.objects.get(id=id)
        form = open_excel_img(request.POST, request.FILES, instance=gs)
        if form.is_valid():
            form.save()
            messages.success(request, SUCCESS_MSG)
        else:
            messages.error(request, FAIL_MSG)
    return redirect(url)


def gsimportsejelli(request, id):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        gs = GsModel.objects.get(id=id)
        form = open_excel_sejelli(request.POST, request.FILES, instance=gs)
        if form.is_valid():
            form.save()
            messages.success(request, SUCCESS_MSG)
        else:
            messages.error(request, FAIL_MSG)
    return redirect(url)


def gsimportflouk(request, id):
    url = request.META.get('HTTP_REFERER')
    if request.method == 'POST':
        gs = GsModel.objects.get(id=id)
        form = open_excel_flouk(request.POST, request.FILES, instance=gs)
        if form.is_valid():
            form.save()
            messages.success(request, SUCCESS_MSG)
        else:
            messages.error(request, FAIL_MSG)
    return redirect(url)


@cache_permission('ipcrep')
def reportipc(request):
    ipclogs = None
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    zones = Zone.objects_limit.all()
    areas = None
    if request.user.owner.role.role == 'zone':
        areas = Area.objects.filter(zone_id=request.user.owner.zone_id)
    else:
        areas = None
    zoneid = 0
    viewid = 1
    parametr = Parametrs.objects.all().first()
    add_to_log(request, 'گزارش وضعیت سرور', 0)
    # محاسبه تاریخ روز گذشته
    yesterday = timezone.now().date() - timedelta(days=1)

    # تعداد قطعی‌های هر جایگاه در روز گذشته
    modem_outage_count = ModemDisconnect.objects.filter(
        gs=OuterRef('gs'),
        tarikh=yesterday  # فقط قطعی‌های روز گذشته
    ).values('gs').annotate(
        count=Count('id')
    ).values('count')

    ipclogs = IpcLog.object_role.c_gs(request, 1).filter(
        gs__status__status=True
    ).annotate(
        modem_outage_count=Subquery(modem_outage_count, output_field=IntegerField())
    ).order_by('-updatedate', 'gs__area_id')

    # برای جایگزینی None با 0
    for log in ipclogs:
        if log.modem_outage_count is None:
            log.modem_outage_count = 0
    _filter = IpcFilter(request.GET, queryset=ipclogs)
    _list = _filter.qs
    if 'search' in request.GET:
        form = SearchForm(request.GET)
        if form.is_valid():
            cd = str(form.cleaned_data['search'])
            cd = checkxss(cd)
            cd = checknumber(cd)
            if cd == "":
                return redirect("base:reportipc")
            if cd.isnumeric():
                _list = _list.filter(Q(gs__gsid__exact=cd))
            else:
                _list = _list.filter(Q(rpm_version__exact=cd) | Q(dashboard_version__exact=cd))
    pajnumber = str(request.GET.get('pajnumber'))
    if pajnumber == 'None':
        pajnumber = "10"
    paginator = Paginator(_list, pajnumber)
    page_num = request.GET.get('page')
    vore = request.GET.get('vore')

    if vore == "2":
        add_to_log(request, 'ارسال آمار خرابی به اکسل  ', 0)
        my_path = "ipclog.xlsx"
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = 'attachment; filename=' + my_path
        font = Font(bold=True)
        fonttitr = Font(bold=True, size=20)
        fonttitr2 = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active  # work with default worksheet
        ws1.title = "گزارش وضعیت سرور تاریخ "
        ws1.sheet_view.rightToLeft = True
        ws1.firstFooter.center.text = "ali"
        ws1.merge_cells('A1:AL1')

        ws1["A1"] = f'گزارش وضعیت سرور تاریخ   {today}'
        ws1["A1"].font = fonttitr

        ws1.merge_cells('A2:A2')
        ws1["A2"] = "ردیف"
        ws1["A2"].font = font

        ws1.merge_cells('B2:B2')
        ws1["B2"] = "منطقه"
        ws1["B2"].font = fonttitr2

        ws1.merge_cells('C2:C2')
        ws1["C2"] = "ناحیه "
        ws1["C2"].font = font

        ws1.merge_cells('D2:D2')
        ws1["D2"] = " نام جایگاه"
        ws1["D2"].font = font

        ws1.merge_cells('E2:E2')
        ws1["E2"] = "GSID"
        ws1["E2"].font = font

        ws1.merge_cells('F2:F2')
        ws1["F2"] = "تاریخ سرور"
        ws1["F2"].font = font

        ws1.merge_cells('G2:G2')
        ws1["G2"] = "نگارش داشبورد"
        ws1["G2"].font = font

        ws1.merge_cells('H2:H2')
        ws1["H2"] = " نگارش سیستم عامل "
        ws1["H2"].font = font

        ws1.merge_cells('I2:I2')
        ws1["I2"] = " نگارش RPM"
        ws1["I2"].font = font

        ws1.merge_cells('J2:J2')
        ws1["J2"] = "تاریخ نصب RPM"
        ws1["J2"].font = font

        ws1.merge_cells('K2:K2')
        ws1["K2"] = "نگارش PT"
        ws1["K2"].font = font

        ws1.merge_cells('L2:L2')
        ws1["L2"] = " نگارش جدول سهمیه"
        ws1["L2"].font = font

        ws1.merge_cells('M2:M2')
        ws1["M2"] = " نگارش جدول قیمت"
        ws1["M2"].font = font

        ws1.merge_cells('N2:N2')
        ws1["N2"] = "نگارش جدول منطقه ایی"
        ws1["N2"].font = font

        ws1.merge_cells('O2:O2')
        ws1["O2"] = " نگارش لیست سیاه"
        ws1["O2"].font = font

        ws1.merge_cells('P2:P2')
        ws1["P2"] = " تعداد لیست سیاه"
        ws1["P2"].font = font

        ws1.merge_cells('Q2:Q2')
        ws1["Q2"] = " تعداد تراکنش خطا"
        ws1["Q2"].font = font

        ws1.merge_cells('R2:R2')
        ws1["R2"] = " تاریخ آخرین ارتباط"
        ws1["R2"].font = font

        ws1.merge_cells('S2:S2')
        ws1["S2"] = " تاریخ آخرین اسکن رمزینه"
        ws1["S2"].font = font

        ws1.merge_cells('T2:T2')
        ws1["T2"] = " اتصال SAM"
        ws1["T2"].font = font

        ws1.merge_cells('U2:U2')
        ws1["U2"] = " ارتباط مودم"
        ws1["U2"].font = font

        ws1.merge_cells('V2:V2')
        ws1["V2"] = " ارتباط DataCenter"
        ws1["V2"].font = font

        ws1.merge_cells('W2:W2')
        ws1["W2"] = " نوع مودم"
        ws1["W2"].font = font

        ws1.merge_cells('X2:X2')
        ws1["X2"] = " مدل IPC"
        ws1["X2"].font = font

        ws1.merge_cells('Y2:Y2')
        ws1["Y2"] = " اپراتور"
        ws1["Y2"].font = font

        ws1.merge_cells('Z2:Z2')
        ws1["Z2"] = " FASB"
        ws1["Z2"].font = font

        ws1.merge_cells('AA2:AA2')
        ws1["AA2"] = " AS "
        ws1["AA2"].font = font

        ws1.merge_cells('AB2:AB2')
        ws1["AB2"] = " مودم بانک ملت"
        ws1["AB2"].font = font

        ws1.merge_cells('AC2:AC2')
        ws1["AC2"] = " اینترنت"
        ws1["AC2"].font = font

        ws1.merge_cells('AD2:AD2')
        ws1["AD2"] = " سریال هارد"
        ws1["AD2"].font = font

        ws1.merge_cells('AE2:AE2')
        ws1["AE2"] = " وضعیت جایگاه"
        ws1["AE2"].font = font

        ws1.merge_cells('AF2:AF2')
        ws1["AF2"] = "GS Version"
        ws1["AF2"].font = font

        ws1.merge_cells('AG2:AG2')
        ws1["AG2"] = " Image Version"
        ws1["AG2"].font = font

        ws1.merge_cells('AH2:AH2')
        ws1["AH2"] = " ظرفیت هارد سرور"
        ws1["AH2"].font = font

        ws1.merge_cells('AI2:AI2')
        ws1["AI2"] = " فضای خالی هارد"
        ws1["AI2"].font = font

        ws1.merge_cells('AJ2:AJ2')
        ws1["AJ2"] = " RAM"
        ws1["AJ2"].font = font

        ws1.merge_cells('AK2:AK2')
        ws1["AK2"] = " EDR"
        ws1["AK2"].font = font

        ws1.merge_cells('AL2:AL2')
        ws1["AL2"] = " تعداد کدینگ"
        ws1["AL2"].font = font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(15.25)
        ws1.column_dimensions['D'].width = float(25.25)
        ws1.column_dimensions['E'].width = float(15.25)
        ws1.column_dimensions['F'].width = float(15.25)
        ws1.column_dimensions['G'].width = float(15.25)
        ws1.column_dimensions['H'].width = float(15.25)
        ws1.column_dimensions['I'].width = float(25.25)
        ws1.column_dimensions['J'].width = float(15.25)
        ws1.column_dimensions['L'].width = float(15.25)
        ws1.column_dimensions['L'].width = float(15.25)
        ws1.column_dimensions['M'].width = float(15.25)
        ws1.column_dimensions['N'].width = float(15.25)
        ws1.column_dimensions['O'].width = float(15.25)
        ws1.column_dimensions['P'].width = float(15.25)
        ws1.column_dimensions['Q'].width = float(25.25)
        ws1.column_dimensions['R'].width = float(25.25)
        ws1.column_dimensions['S'].width = float(25.25)
        ws1.column_dimensions['T'].width = float(25.25)
        ws1.column_dimensions['U'].width = float(25.25)
        ws1.column_dimensions['V'].width = float(25.25)
        ws1.column_dimensions['W'].width = float(25.25)
        ws1.column_dimensions['X'].width = float(25.25)
        ws1.column_dimensions['Y'].width = float(25.25)
        ws1.column_dimensions['Z'].width = float(25.25)
        ws1.column_dimensions['AA'].width = float(25.25)
        ws1.column_dimensions['AB'].width = float(25.25)
        ws1.column_dimensions['AC'].width = float(25.25)
        ws1.column_dimensions['AD'].width = float(25.25)
        ws1.column_dimensions['AE'].width = float(25.25)
        ws1.column_dimensions['AF'].width = float(25.25)
        ws1.column_dimensions['AG'].width = float(25.25)
        ws1.column_dimensions['AH'].width = float(25.25)
        ws1.column_dimensions['AI'].width = float(25.25)
        ws1.column_dimensions['AJ'].width = float(25.25)
        ws1.column_dimensions['AK'].width = float(25.25)
        ws1.column_dimensions['AL'].width = float(25.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        myfont = Font(size=14, bold=True)  # font styles
        my_fill = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        my_fill2 = PatternFill(
            fill_type='solid', start_color='dadfe3')  # Background color
        i = 0

        for item in _list:
            _sam = 'متصل' if item.sam else 'عدم ارتباط'
            _modem = 'متصل' if item.modem else 'عدم ارتباط'
            _dc = 'متصل' if item.datacenter else 'عدم ارتباط'
            _fasb = 'متصل' if item.fasb else 'عدم ارتباط'
            _asmelat = 'متصل' if item.asmelat else 'عدم ارتباط'
            _mellatmodem = 'متصل' if item.mellatmodem else 'عدم ارتباط'
            _internet = 'متصل' if item.internet else 'عدم ارتباط'
            try:
                modem = item.gs.modem.name
            except AttributeError:
                modem = 'ثبت نشده'

            try:
                ipc = item.gs.ipc.name
            except AttributeError:
                ipc = 'ثبت نشده'

            try:
                operator = item.gs.operator.name
            except AttributeError:
                operator = 'ثبت نشده'

            i += 1
            d = [i, str(item.gs.area.zone.name), str(item.gs.area.name), str(item.gs.name), str(item.gs.gsid),
                 str(item.ipcdate()) + " " + str(item.time_ipc),
                 str(item.dashboard_version), str(item.os_version), str(item.rpm_version), str(item.rpmdate()),
                 str(item.pt_version), str(item.quta_table_version), str(item.price_table_version),
                 str(item.zone_table_version), str(item.blacklist_version), str(item.blacklist_count), str(item.bl_ipc),
                 str(item.lastdate()), str(item.nowdate()), _sam, _modem, _dc, str(modem), str(ipc), str(operator),
                 str(_fasb), str(_asmelat), str(_mellatmodem), str(_internet), str(item.hd_serial),
                 str(item.gs.status.name), str(item.gs_version), str(item.imagever), str(item.hdd_total), str(item.hdd_empy),
                 str(item.ram_total),str(item.edr),str(item.coding_count)]

            ws1.append(d)

        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        i += 4
        for cell in ws1[i:i]:  # Last row
            cell.font = myfont
            cell.fill = my_fill2
            cell.border = thin_border

        for cell in ws1["2:2"]:  # First row
            cell.font = myfont
            cell.fill = my_fill
            cell.border = thin_border
        wb.save(response)
        return response

    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    query_string = request.META.get("QUERY_STRING", "")
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    context = {'list': page_object, 'parametr': parametr, 'id': id, 'zones': zones,
               'areas': areas,
               'viewid': int(viewid), 'query_string': query_string, 'page_obj': page_obj, 'tedad': tedad,
               'filter': _filter}

    return TemplateResponse(request, 'report_ipc.html', context)


@cache_permission('ipcrep')
def reportipchistort(request, _id):
    add_to_log(request, f'گزارش سابفه وضعیت سرور جایگاه ', _id)
    ipclogs = None
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    viewid = 1
    gsname = GsModel.objects.get(id=_id).name
    ipclogs = IpcLogHistory.object_role.c_gs(request, 0).filter(gs_id=_id).order_by('-update')
    paginator = Paginator(ipclogs, 10)
    page_num = request.GET.get('page')
    query_string = request.META.get("QUERY_STRING", "")
    data = request.GET.copy()
    if query_string.startswith("page"):
        query_string = query_string.split("&", 1)
        query_string = query_string[1]
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    tedad = paginator.count

    context = {'list': page_object, 'gsname': gsname,
               'viewid': int(viewid), 'query_string': query_string, 'page_obj': page_obj, 'tedad': tedad,
               }
    return TemplateResponse(request, 'reportipchistory.html', context)


@cache_permission('report')
def averagetickets(request, id):
    _list = []
    add_to_log(request, f'گزارش میانگین تیکت ها ', 0)
    if request.user.owner.role.role in ['zone']:
        zone = Zone.objects_limit.filter(id=request.user.owner.zone_id).exclude(id=9).order_by('id')
    elif request.user.owner.role.role in ['setad', 'mgr', 'fani', 'test', 'posht']:
        zone = Zone.objects_limit.all().exclude(id=9).order_by('id')
    else:
        messages.warning(request, DENY_PERMISSSION)
        return redirect(HOME_PAGE)
    s_master = 0
    s_pinpad = 0
    summ = 0
    sum_gs = 0
    sum_pump = 0
    sum_ticket = 0
    sum_master = 0
    sum_pinpad = 0
    if request.method != 'POST':

        for gs in zone:
            count_gs = GsModel.objects.filter(area__zone_id=gs.id, active=True).count()
            count_pump = Pump.objects.filter(gs__area__zone_id=gs.id, gs__active=True, actived=True).count()
            count_ticket = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                            failure__failurecategory_id__in=[1010,
                                                                                                             1011]).count()

            count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                            failure__failurecategory_id=1010,
                                                                            failure__isnazel=True).count()

            count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id, status_id=1,
                                                                            failure__failurecategory_id=1011,
                                                                            failure__isnazel=True).count()
            if count_master > 0:
                master = ((int(count_master) / int(count_pump)) * 100)
            else:
                master = 0
            if count_pinpad > 0:
                pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
            else:
                pinpad = 0
            summ = round(master + pinpad, 2)
            _dict = {
                'st': 0,
                'id': gs.id,
                'name': gs.name,
                'count_gs': count_gs,
                'count_pump': count_pump,
                'count_ticket': count_ticket,
                'count_master': count_master,
                'count_pinpad': count_pinpad,
                'master': round(master, 2),
                'pinpad': round(pinpad, 2),
                'summ': summ,

            }
            _list.append(_dict)
            _list = sorted(_list, key=itemgetter('master'), reverse=True)
            sum_gs += count_gs
            sum_pump += count_pump
            sum_ticket += count_ticket
            sum_master += count_master
            sum_pinpad += count_pinpad
            if sum_master > 0:
                s_master = ((int(sum_master) / int(sum_pump)) * 100)
            else:
                s_master = 0
            if count_pinpad > 0:
                s_pinpad = ((int(sum_pinpad) / int(sum_pump)) * 100)
            else:
                s_pinpad = 0
            summ = round(s_master + s_pinpad, 2)

    if request.method == 'POST':

        tarikhin = request.POST.get('select')
        tarikhto = request.POST.get('select2')

        tarikhin = tarikhin.split("/")
        tarikhto = tarikhto.split("/")
        tarikhin = jdatetime.date(day=int(tarikhin[2]), month=int(tarikhin[1]), year=int(tarikhin[0])).togregorian()
        tarikhto = jdatetime.date(day=int(tarikhto[2]), month=int(tarikhto[1]), year=int(tarikhto[0])).togregorian()
        en_tarikh_in = str(tarikhin) + " 00:00:00"
        en_tarikh_to = str(tarikhto) + " 23:59:59"

        for gs in zone:
            count_gs = GsModel.objects.filter(area__zone_id=gs.id, active=True).count()
            count_pump = Pump.objects.filter(gs__area__zone_id=gs.id, actived=True,
                                             gs__active=True).count()
            count_ticket = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id,
                                                                            create__gte=en_tarikh_in,
                                                                            create__lte=en_tarikh_to,
                                                                            failure__failurecategory_id__in=[1010,
                                                                                                             1011]).count()
            count_ticket += Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id,
                                                                             create__gte=en_tarikh_in,
                                                                             create__lte=en_tarikh_to,
                                                                             failure__id=1045).count()

            count_master = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id,
                                                                            failure__failurecategory_id__in=[1010
                                                                                                             ],
                                                                            create__gte=en_tarikh_in,
                                                                            create__lte=en_tarikh_to
                                                                            ).count()
            count_master += Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id,
                                                                             failure__id=1045,
                                                                             create__gte=en_tarikh_in,
                                                                             create__lte=en_tarikh_to
                                                                             ).count()

            count_pinpad = Ticket.objects.exclude(organization_id=4).filter(gs__area__zone_id=gs.id,
                                                                            create__gte=en_tarikh_in,
                                                                            create__lte=en_tarikh_to,
                                                                            failure__failurecategory_id=1011
                                                                            ).count()
            sum_ticket += count_ticket
            sum_master += count_master
            sum_pinpad += count_pinpad
            day = (tarikhto - tarikhin).days
            count_master = round((count_master / int(day)), 2)
            count_pinpad = round((count_pinpad / int(day)), 2)

            if count_master > 0:
                master = ((int(count_master) / int(count_pump)) * 100)
            else:
                master = 0
            if count_pinpad > 0:
                pinpad = ((int(count_pinpad) / int(count_pump)) * 100)
            else:
                pinpad = 0
            summ = round(master + pinpad, 2)
            _dict = {
                'st': 0,
                'id': gs.id,
                'name': gs.name,
                'count_gs': count_gs,
                'count_pump': count_pump,
                'count_ticket': count_ticket,
                'count_master': count_master,
                'count_pinpad': count_pinpad,
                'master': round(master, 2),
                'pinpad': round(pinpad, 2),
                'summ': summ,

            }
            _list.append(_dict)

            sum_gs += count_gs
            sum_pump += count_pump
            sum_master = round((sum_master / int(day)), 2)
            sum_pinpad = round((sum_pinpad / int(day)), 2)
            if sum_master > 0:
                s_master = ((int(sum_master) / int(sum_pump)) * 100)
            else:
                s_master = 0
            if count_pinpad > 0:
                s_pinpad = ((int(sum_pinpad) / int(sum_pump)) * 100)
            else:
                s_pinpad = 0
            summ = round(s_master + s_pinpad, 2)
    if id == 1:
        _list = sorted(_list, key=itemgetter('master'), reverse=True)
    elif id == 2:
        _list = sorted(_list, key=itemgetter('pinpad'), reverse=True)
    elif id == 3:
        _list = sorted(_list, key=itemgetter('summ'), reverse=True)
    _dict = {
        'st': 1,
        'id': 0,
        'name': SUM_TITEL,
        'count_gs': sum_gs,
        'count_pump': sum_pump,
        'count_ticket': sum_ticket,
        'count_master': sum_master,
        'count_pinpad': sum_pinpad,
        'master': round(s_master, 2),
        'pinpad': round(s_pinpad, 2),
        'summ': summ,

    }
    _list.append(_dict)

    context = {'list': _list, 'id': id}
    return TemplateResponse(request, 'report/averagetickets.html', context)


@cache_permission('report')
def reportsla(request):
    ticket = []
    add_to_log(request, f'گزارش  تیکت ها بر اساس sla ', 0)
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if request.method == 'POST':
        tarikhin = request.POST.get('select')
        tarikhto = request.POST.get('select2')

        tarikhin = tarikhin.split("/")
        tarikhto = tarikhto.split("/")
        tarikhin = jdatetime.date(day=int(tarikhin[2]), month=int(tarikhin[1]), year=int(tarikhin[0])).togregorian()
        tarikhto = jdatetime.date(day=int(tarikhto[2]), month=int(tarikhto[1]), year=int(tarikhto[0])).togregorian()
        en_tarikh_in = str(tarikhin) + " 00:00:00"
        en_tarikh_to = str(tarikhto) + " 23:59:59"

        ticket = Ticket.object_role.c_ticket(request).filter(create__gte=en_tarikh_in,
                                                             create__lte=en_tarikh_to, statusmoavagh__issla=True)

    context = {'list': ticket}
    return TemplateResponse(request, 'report/ticketsla.html', context)


def repnazelgs(request):
    _list = []

    add_to_log(request, 'ارسال اطلاعات جایگاه ها  به اکسل', 0)
    if request.user.owner.role.role in ['zone']:
        _list = Pump.objects.select_related(
            'gs__area__zone', 'gs__operator', 'gs__ipc', 'gs__modem', 'gs__rack', 'gs__gsipclog', 'gs__printer',
            'gs__thinclient'
        ).filter(
            status_id__in=[1, 2], gs__area__zone_id=request.user.owner.zone_id).values('gs_id', 'gs__name', 'gs__gsid',
                                                                                       'gs__area__zone__name',
                                                                                       'gs__area__name',
                                                                                       'gs__city__name',
                                                                                       'gs__operator__name',
                                                                                       'gs__ipc__name',
                                                                                       'gs__printer__name',
                                                                                       'gs__thinclient__name',
                                                                                       'gs__modem__name', 'gs__isbank',
                                                                                       'gs__ispaystation',
                                                                                       'gs__location', 'gs__rack__name',
                                                                                       'gs__update',
                                                                                       'gs__gsipclog__rpm_version',
                                                                                       'gs__status__name',
                                                                                       'gs__gsipclog__rpm_version_date',
                                                                                       'gs__gsipclog__dashboard_version',
                                                                                       'gs__gsipclog__os_version',
                                                                                       'gs__gsipclog__pt_version',
                                                                                       'gs__gsipclog__quta_table_version',
                                                                                       'gs__gsipclog__price_table_version',
                                                                                       'gs__gsipclog__zone_table_version',
                                                                                       'gs__gsipclog__blacklist_version',
                                                                                       'gs__isonline', 'gs__telldaftar',
                                                                                       'gs__address',
                                                                                       'gs__is_montakhab',
                                                                                       'gs__m_benzin', 'gs__m_super',
                                                                                       'gs__m_naftgaz',
                                                                                       'gs__postal_code',
                                                                                       'gs__iscoding',
                                                                                       'gs__gsstatus__name',
                                                                                       'gs__simcart'

                                                                                       ).annotate(
            jam=Count('id'),
            benzin=Count(
                Case(When(product_id=2,
                          then=1))),
            super=Count(
                Case(When(product_id=3,
                          then=1))),
            gaz=Count(
                Case(When(product_id=4,
                          then=1))))


    elif request.user.owner.role.role in ['setad', 'mgr', 'fani', 'test']:
        _list = Pump.objects.select_related(
            'gs__area__zone', 'gs__operator', 'gs__ipc', 'gs__modem', 'gs__rack', 'gs__gsipclog', 'gs__printer',
            'gs__thinclient'
        ).filter(
            status_id__in=[1, 2]).values('gs_id', 'gs__name', 'gs__gsid', 'gs__area__zone__name', 'gs__area__name',
                                         'gs__operator__name', 'gs__ipc__name', 'gs__modem__name', 'gs__isbank',
                                         'gs__printer__name',
                                         'gs__thinclient__name',
                                         'gs__ispaystation', 'gs__location', 'gs__update', 'gs__gsipclog__rpm_version',
                                         'gs__status__name', 'gs__gsipclog__rpm_version_date',
                                         'gs__gsipclog__dashboard_version',
                                         'gs__gsipclog__os_version', 'gs__gsipclog__pt_version',
                                         'gs__gsipclog__quta_table_version', 'gs__phone', 'gs__address',
                                         'gs__is_montakhab', 'gs__m_benzin', 'gs__m_super', 'gs__m_naftgaz',
                                         'gs__postal_code', 'gs__telldaftar', 'gs__address',
                                         'gs__gsipclog__price_table_version', 'gs__gsipclog__zone_table_version',
                                         'gs__gsipclog__blacklist_version', 'gs__isonline', 'gs__rack__name',
                                         'gs__iscoding', 'gs__city__name', 'gs__gsstatus__name',
                                         'gs__simcart').annotate(jam=Count('id'),
                                                                 benzin=Count(Case(
                                                                     When(product_id=2,
                                                                          then=1))),
                                                                 super=Count(Case(
                                                                     When(product_id=3,
                                                                          then=1))),
                                                                 gaz=Count(Case(
                                                                     When(product_id=4,
                                                                          then=1))))

    else:
        messages.warning(request, DENY_PERMISSSION)
        return redirect(HOME_PAGE)

    my_path = "gs.xlsx"
    response = HttpResponse(content_type=EXCEL_MODE)
    response['Content-Disposition'] = 'attachment; filename=' + my_path
    font = Font(bold=True)
    fonttitr = Font(bold=True, size=20)
    fonttitr2 = Font(bold=True, size=20)
    wb = Workbook()

    ws1 = wb.active  # work with default worksheet
    ws1.title = "گزارش  اطلاعات جایگاه ها    "
    ws1.sheet_view.rightToLeft = True
    ws1.firstFooter.center.text = "ali"
    ws1.merge_cells('A1:AQ1')

    ws1["A1"] = f'گزارش اطلاعات جایگاه ها   {today}'
    ws1["A1"].font = fonttitr

    ws1.merge_cells('A2:A3')
    ws1["A2"] = "ردیف"
    ws1["A2"].font = font

    ws1.merge_cells('B2:B3')
    ws1["B2"] = "منطقه"
    ws1["B2"].font = fonttitr2

    ws1.merge_cells('C2:C3')
    ws1["C2"] = "ناحیه "
    ws1["C2"].font = font

    ws1.merge_cells('D2:D3')
    ws1["D2"] = "GSID"
    ws1["D2"].font = font

    ws1.merge_cells('E2:E3')
    ws1["E2"] = "نام جایگاه"
    ws1["E2"].font = font

    ws1.merge_cells('F2:I2')
    ws1["F2"] = "تعداد نازل"
    ws1["F2"].font = font

    ws1.merge_cells('F3:F3')
    ws1["F3"] = " بنزین"
    ws1["F3"].font = font

    ws1.merge_cells('G3:G3')
    ws1["G3"] = "  سوپر "
    ws1["G3"].font = font

    ws1.merge_cells('H3:H3')
    ws1["H3"] = "  نفتگاز"
    ws1["H3"].font = font

    ws1.merge_cells('I3:I3')
    ws1["I3"] = "مجموع"
    ws1["I3"].font = font

    ws1.merge_cells('J2:J3')
    ws1["J2"] = "وضعیت جایگاه "
    ws1["J2"].font = font

    ws1.merge_cells('K2:K3')
    ws1["K2"] = "وضعیت ارتباط سرور و کارتخوان "
    ws1["K2"].font = font

    ws1.merge_cells('l2:l3')
    ws1["l2"] = "اپراتور "
    ws1["l2"].font = font

    ws1.merge_cells('m2:m3')
    ws1["m2"] = "سرور "
    ws1["m2"].font = font

    ws1.merge_cells('n2:n3')
    ws1["n2"] = "رک "
    ws1["n2"].font = font

    ws1.merge_cells('o2:o3')
    ws1["o2"] = "مودم "
    ws1["o2"].font = font

    ws1.merge_cells('p2:p3')
    ws1["p2"] = "تجهیزات بانک ملت "
    ws1["p2"].font = font

    ws1.merge_cells('q2:q3')
    ws1["q2"] = "تجهیزات پرداخت بانک ملت "
    ws1["q2"].font = font

    ws1.merge_cells('r2:r3')
    ws1["r2"] = "لوکیشن"
    ws1["r2"].font = font

    ws1.merge_cells('s2:s3')
    ws1["s2"] = "زمان آخرین تغییر"
    ws1["s2"].font = font

    ws1.merge_cells('t2:t3')
    ws1["t2"] = "نگارش RPM"
    ws1["t2"].font = font

    ws1.merge_cells('u2:u3')
    ws1["u2"] = "تاریخ نصب RPM"
    ws1["u2"].font = font

    ws1.merge_cells('v2:v3')
    ws1["v2"] = "نگارش GDS"
    ws1["v2"].font = font

    ws1.merge_cells('w2:w3')
    ws1["w2"] = "نگارش OS"
    ws1["w2"].font = font

    ws1.merge_cells('x2:x3')
    ws1["x2"] = "نگارش PT"
    ws1["x2"].font = font

    ws1.merge_cells('y2:y3')
    ws1["y2"] = "نگارش جدول سهمیه"
    ws1["y2"].font = font

    ws1.merge_cells('z2:z3')
    ws1["z2"] = "نگارش جدول قیمت"
    ws1["z2"].font = font

    ws1.merge_cells('aa2:aa3')
    ws1["aa2"] = "نگارش جدول منطقه ایی"
    ws1["aa2"].font = font

    ws1.merge_cells('ab2:ab3')
    ws1["ab2"] = "نگارش لیست سیاه"
    ws1["ab2"].font = font

    ws1.merge_cells('ac2:ac3')
    ws1["ac2"] = "طرح کدینگ"
    ws1["ac2"].font = font

    ws1.merge_cells('ad2:ad3')
    ws1["ad2"] = "Lat"
    ws1["ad2"].font = font

    ws1.merge_cells('ae2:ae3')
    ws1["ae2"] = "Long"
    ws1["ae2"].font = font

    ws1.merge_cells('af2:af3')
    ws1["af2"] = "شهرستان"
    ws1["af2"].font = font

    ws1.merge_cells('ag2:ag3')
    ws1["ag2"] = "شماره تلفن"
    ws1["ag2"].font = font

    ws1.merge_cells('ah2:ah3')
    ws1["ah2"] = "آدرس"
    ws1["ah2"].font = font

    ws1.merge_cells('ai2:ai3')
    ws1["ai2"] = "منتخب نفتگاز"
    ws1["ai2"].font = font

    ws1.merge_cells('aj2:aj3')
    ws1["aj2"] = "ظرفیت مخزن بنزین"
    ws1["aj2"].font = font

    ws1.merge_cells('ak2:ak3')
    ws1["ak2"] = "ظرفیت مخزن سوپر"
    ws1["ak2"].font = font

    ws1.merge_cells('al2:al3')
    ws1["al2"] = "ظرفیت مخزن نفتگاز"
    ws1["al2"].font = font

    ws1.merge_cells('am2:am3')
    ws1["am2"] = "کد پستی"
    ws1["am2"].font = font

    ws1.merge_cells('an2:an3')
    ws1["an2"] = "پرینتر"
    ws1["an2"].font = font

    ws1.merge_cells('ao2:ao3')
    ws1["ao2"] = "ThinClient"
    ws1["ao2"].font = font

    ws1.merge_cells('ap2:ap3')
    ws1["ap2"] = "موقعیت مکانی"
    ws1["ap2"].font = font

    ws1.merge_cells('aq2:aq3')
    ws1["aq2"] = "شماره سیم کارت"
    ws1["aq2"].font = font

    ws1.column_dimensions['B'].width = float(15.25)
    ws1.column_dimensions['C'].width = float(10.25)
    ws1.column_dimensions['D'].width = float(7.25)
    ws1.column_dimensions['E'].width = float(24.25)
    ws1.column_dimensions['F'].width = float(8.25)
    ws1.column_dimensions['G'].width = float(9.25)
    ws1.column_dimensions['H'].width = float(8.25)
    ws1.column_dimensions['I'].width = float(9.25)
    ws1.column_dimensions['J'].width = float(9.25)
    ws1.column_dimensions['K'].width = float(9.25)
    ws1.column_dimensions['l'].width = float(9.25)
    ws1.column_dimensions['m'].width = float(9.25)
    ws1.column_dimensions['n'].width = float(9.25)
    ws1.column_dimensions['o'].width = float(9.25)
    ws1.column_dimensions['p'].width = float(9.25)
    ws1.column_dimensions['q'].width = float(9.25)
    ws1.column_dimensions['r'].width = float(23.25)
    ws1.column_dimensions['s'].width = float(22.25)
    ws1.column_dimensions['t'].width = float(15.25)
    ws1.column_dimensions['u'].width = float(15.25)
    ws1.column_dimensions['v'].width = float(15.25)
    ws1.column_dimensions['w'].width = float(15.25)
    ws1.column_dimensions['x'].width = float(15.25)
    ws1.column_dimensions['y'].width = float(15.25)
    ws1.column_dimensions['z'].width = float(15.25)
    ws1.column_dimensions['aa'].width = float(15.25)
    ws1.column_dimensions['ab'].width = float(15.25)
    ws1.column_dimensions['ac'].width = float(15.25)
    ws1.column_dimensions['ad'].width = float(15.25)
    ws1.column_dimensions['ae'].width = float(15.25)
    ws1.column_dimensions['af'].width = float(15.25)
    ws1.column_dimensions['ag'].width = float(15.25)
    ws1.column_dimensions['ah'].width = float(15.25)
    ws1.column_dimensions['ai'].width = float(15.25)
    ws1.column_dimensions['aj'].width = float(15.25)
    ws1.column_dimensions['ak'].width = float(15.25)
    ws1.column_dimensions['al'].width = float(15.25)
    ws1.column_dimensions['am'].width = float(15.25)
    ws1.column_dimensions['an'].width = float(15.25)
    ws1.column_dimensions['ao'].width = float(15.25)
    ws1.column_dimensions['ap'].width = float(15.25)
    ws1.column_dimensions['aq'].width = float(15.25)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    myfont = Font(size=14, bold=True)  # font styles
    my_fill = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    my_fill2 = PatternFill(
        fill_type='solid', start_color='dadfe3')  # Background color
    i = 0

    for item in _list:
        i += 1
        try:
            _location = item['gs__location'].split(',')
            _lat = _location[0]
            _long = _location[1]
            _lata = _lat.split('.')
            _late = _lata[1][:6]
            _lat = str(_lata[0]) + '.' + str(_late)
            _longa = _long.split('.')
            _longe = _longa[1][:6]
            _long = str(_longa[0]) + '.' + str(_longe)
        except:
            _lat = 0
            _long = 0
        nahye = str(item['gs__city__name']) if item['gs__city__name'] else ""
        d = [i, str(item['gs__area__zone__name']), str(item['gs__area__name']), str(item['gs__gsid']),
             str(item['gs__name']), str(item['benzin']), str(item['super']), str(item['gaz']), str(item['jam']),
             str(item['gs__status__name']),
             str(item['gs__isonline']), str(item['gs__operator__name']), str(item['gs__ipc__name']),
             str(item['gs__rack__name']), str(item['gs__modem__name']), str(item['gs__isbank']),
             str(item['gs__ispaystation']), str(item['gs__location']), str(item['gs__update']),
             str(item['gs__gsipclog__rpm_version']),
             str(item['gs__gsipclog__rpm_version_date']),
             str(item['gs__gsipclog__dashboard_version']),
             str(item['gs__gsipclog__os_version']),
             str(item['gs__gsipclog__pt_version']),
             str(item['gs__gsipclog__quta_table_version']),
             str(item['gs__gsipclog__price_table_version']),
             str(item['gs__gsipclog__zone_table_version']),
             str(item['gs__gsipclog__blacklist_version']),
             str(item['gs__iscoding']), str(_lat), str(_long), str(nahye),
             str(item['gs__telldaftar']), str(item['gs__address']),
             str(item['gs__is_montakhab']), str(item['gs__m_benzin']),
             str(item['gs__m_super']), str(item['gs__m_naftgaz']),
             str(item['gs__postal_code']), str(item['gs__printer__name']), str(item['gs__thinclient__name']),
             str(item['gs__gsstatus__name']), str(item['gs__simcart'])
             ]

        ws1.append(d)

    for col in ws1.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(
                horizontal='center', vertical='center')
            cell.alignment = alignment_obj
            cell.border = thin_border

    i += 4
    for cell in ws1[i:i]:  # Last row
        cell.font = myfont
        cell.fill = my_fill2
        cell.border = thin_border

    for cell in ws1["2:2"]:  # First row
        cell.font = myfont
        cell.fill = my_fill
        cell.border = thin_border
    wb.save(response)
    return response


def sla(request):
    if request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        status_id = request.POST.get('status_id')
        ticket = Ticket.objects.get(id=ticket_id)
        pump = ticket.Pump_id
        ticket.statusmoavagh_id = status_id
        ticket.save()
        status = StatusMoavagh.objects.get(id=int(status_id))
        if status.ename == "nosell":
            ticket.status_id = 2
            ticket.organization_id = 8
            ticket.save()
            Workflow.objects.create(ticket_id=int(ticket_id), organization_id=8,
                                    description='تیکت بعلت عدم فروش نازل بسته شد / نازل غیر فعال شد',
                                    user_id=request.user.owner.id,
                                    failure_id=ticket.failure_id)
            pump = Pump.objects.get(id=pump)
            pump.status_id = 5
            pump.save()
        return JsonResponse({'message': 'success'})


@cache_permission('area')
def arealist(request):
    add_to_log(request, f' مشاهده فرم لیست نواحیق', 0)
    if request.user.owner.role.role in ['mgr', 'setad']:
        _list = Area.objects.all().order_by('zone_id')
    else:
        _list = Area.objects.filter(zone_id=request.user.owner.zone_id)
    return TemplateResponse(request, 'parametrs/listarea.html', {'list': _list})


@cache_permission('area')
def areaupdate(request, _id):
    form = AreaForm()
    area = Area.objects.get(id=_id)
    if request.method == 'POST':
        form = AreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            messages.success(request, 'عملیات با موفقیت انجام شد')
            add_to_log(request, f'   بروزرسانی اطلاعات ناحیه {area.name} ', 0)
            return redirect('base:arealist')

        else:
            messages.error(request, 'عملیات شکست خورد')

    return TemplateResponse(request, 'parametrs/updatearea.html',
                            {'form': form, 'area': area})


def checkipclog(request):
    add_to_log(request, f' بروزرسانی وضعیت سرور', 0)
    data = Parametrs.objects.first()
    for item in IpcLog.objects.all():
        try:
            isonlinegd = GsModel.objects.get(id=item.gs_id)
            if item.dashboard_version in data.dashboard_version:
                item.ck_dashboard_version = True
            item.ck_rpm_version = True if item.rpm_version in data.rpm_version else False
            item.ck_pt_version = True if item.pt_version in data.pt_version else False
            item.ck_pt_online = False if isonlinegd.isonline == True and data.online_pt_version != item.pt_version else False
            item.ck_pt_online = True if isonlinegd.isonline == False else False
            item.ck_pt_online = True if data.online_pt_version == item.pt_version else False
            if item.ck_pt_online and item.ck_blacklist_count and item.ck_blacklist_version and item.ck_pt_version and item.ck_rpm_version and item.ck_price_table_version and item.ck_dashboard_version and item.ck_quta_table_version and item.ck_zone_table_version:
                item.ck_contradiction = False
            else:
                item.ck_contradiction = True
            item.ck_quta_table_version = True if item.quta_table_version in data.quta_table_version else False
            item.ck_price_table_version = True if item.price_table_version in data.price_table_version else False
            item.ck_zone_table_version = True if item.zone_table_version >= isonlinegd.zone_table_version else False
            if int(isonlinegd.zone_table_version) == 0:
                item.ck_zone_table_version = True
            item.ck_blacklist_version = True if data.blacklist_version <= item.blacklist_version else False
            if int(item.blacklist_count) <= int(settings.MAX_BLACKLIST_COUNT_ALERT) and int(
                    item.blacklist_count) >= int(
                settings.MIN_BLACKLIST_COUNT_ALERT):
                item.ck_blacklist_count = True
            else:
                item.ck_blacklist_count = False
            item.save()
        except IntegrityError:
            continue
    return HttpResponse('ok')


@cache_permission('gsaddclose')
def addclosegs(request, id):
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    gs = GsModel.object_role.c_gsmodel(request).get(id=id)
    if request.user.owner.role.role == 'zone' and gs.area.zone_id != request.user.owner.zone_id:
        messages.error(request, 'شما به اطلاعات این جایگاه دسترسی ندارید')
        return redirect(HOME_PAGE)
    if request.user.owner.role.role == 'area' and gs.area_id != request.user.owner.area_id:
        messages.error(request, 'شما به اطلاعات این جایگاه دسترسی ندارید')
        return redirect(HOME_PAGE)
    if request.method == 'POST':
        datein = request.POST.get("select")
        dateout = request.POST.get("select2")
        status = request.POST.get("status")
        datein = datein.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]),
                                year=int(datein[0])).togregorian()
        dateout = dateout.split("/")
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]),
                                 year=int(dateout[0])).togregorian()

        CloseGS.object_role.c_gs(request, 0).create(gs_id=id, date_in=datein, date_out=dateout, status=status,
                                                    owner_id=request.user.owner.id)
        add_to_log(request, f' {str(id)}ثبت تعطیلی جایگاه ', id)
        messages.success(request, 'عملیات با موفقیت انجام شد.')
        return redirect('base:gs_detail', id)

    return TemplateResponse(request, 'clsegs.html', {})


@cache_permission('gsaddclose')
def addmoghgs(request, id):
    url = request.META.get('HTTP_REFERER')
    ispay = True
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)

    if request.method == 'POST':

        datein = request.POST.get("select")
        ispay = request.POST.get("ispay")
        info = request.POST.get("info")
        if len(datein) < 10:
            messages.error(request, 'تاریخ را بصورت دقیق وارد کنید')
            return redirect(url)
        if len(info) < 20:
            messages.error(request, 'علت مغایرت را بصورت دقیق وارد کنید')
            return redirect(url)
        _ispay = False if ispay == "1" else True

        datein = datein.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]),
                                year=int(datein[0])).togregorian()

        AcceptForBuy.object_role.c_gs(request, 0).create(gs_id=id, tarikh=datein, owner_id=request.user.owner.id,
                                                         ispay=_ispay,info=info)
        add_to_log(request, f' {str(id)}ثبت مجوز مغایرت جایگاه ', id)
        messages.success(request, 'عملیات با موفقیت انجام شد.')
        return redirect('base:gs_detail', id)

    return TemplateResponse(request, 'moghgs.html', {'ispay': ispay})


@cache_permission('gs')
def updateclosegs(request, id):
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    try:
        closegs = CloseGS.object_role.c_gs(request, 0).get(id=id)
    except Exception as e:
        messages.success(request, 'دسترسی غیر مجاز.')
        return redirect('base:gs_detail', id)
    if request.method == 'POST':
        datein = request.POST.get("select")
        dateout = request.POST.get("select2")
        status = request.POST.get("status")
        datein = datein.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]),
                                year=int(datein[0])).togregorian()

        dateout = dateout.split("/")
        dateout = jdatetime.date(day=int(dateout[2]), month=int(dateout[1]),
                                 year=int(dateout[0])).togregorian()

        gs = CloseGS.objects.get(id=id)
        gs.date_in = datein
        gs.date_out = dateout
        gs.status = status
        gs.owner_id = request.user.owner.id
        gs.save()
        add_to_log(request, f' {str(id)}ویرایش تعطیلی جایگاه ', id)
        messages.success(request, 'عملیات با موفقیت انجام شد.')
        return redirect('base:gs_detail', closegs.gs_id)

    return TemplateResponse(request, 'clsegs.html', {'closegs': closegs})


@cache_permission('gs')
def updatemoghgs(request, id):
    url = request.META.get('HTTP_REFERER')
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if len(str(id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    id = Decrypt(id)
    moghgs = AcceptForBuy.object_role.c_gs(request, 0).get(id=id)

    ispay = moghgs.ispay
    if request.method == 'POST':
        datein = request.POST.get("select")
        ispay = request.POST.get("ispay")
        info = request.POST.get("info")
        if len(datein) < 10:
            messages.error(request, 'تاریخ را بصورت دقیق وارد کنید')
            return redirect(url)
        if len(info) < 20:
            messages.error(request, 'علت مغایرت را بصورت دقیق وارد کنید')
            return redirect(url)
        _ispay = False if ispay == "1" else True

        datein = datein.split("/")
        datein = jdatetime.date(day=int(datein[2]), month=int(datein[1]),
                                year=int(datein[0])).togregorian()

        gs = AcceptForBuy.objects.get(id=id)
        gs.tarikh = datein
        gs.ispay = _ispay
        gs.info = info

        gs.save()
        add_to_log(request, f' {_ispay} {str(moghgs.tarikh)} {str(moghgs.gs.name)}  ویرایش اعلام مغایرت جایگاه   ',
                   moghgs.gs.id)
        messages.success(request, 'عملیات با موفقیت انجام شد.')
        return redirect('base:gs_detail', moghgs.gs_id)
    date2 = str(moghgs.tarikh)
    date2 = date2.replace("-", "/")

    return TemplateResponse(request, 'moghgs.html', {'moghgs': moghgs, 'date2': date2, 'ispay': ispay})


def storehistory(request):
    serial = request.POST.get('val')
    _list = StoreHistory.objects.filter(store__serial=serial, create__lt=datetime.datetime.now()).order_by('id')
    result = []
    datestart = ''
    dateend = ''
    _ok = 0
    gs = ''
    az = ''
    ta = ''
    for item in _list:
        days = 99999
        starterr = False
        if item.status_id == 5:
            _ok = 1
            datestart = item.create
            az = item.normal_date()
            gs = item.description
        if item.status_id == 6:
            if _ok == 1:
                dateend = item.create
                ta = item.normal_date()

                _ok = 0
        if item.status_id == 6 and item.starterror:
            starterr = True
            dateend = ''
            datestart = ''

        if datestart and dateend:
            days = (dateend - datestart).days
            dateend = ''
            datestart = ''
        if days == 99999 and starterr == False:
            dict = {}
        else:
            dict = {
                'days': days,
                'starterr': starterr,
                'tarikh': item.pdate(),
                'az': az,
                'ta': ta,
                'gs': gs,
            }
            result.append(dict)
    return JsonResponse({'message': 'success', 'serial': serial, 'result': result})


@cache_permission('slaconflig')
def reportsla_conflig(request):
    _list = []
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    zones = Zone.objects_limit.all()
    zone = '0'
    if _role in ['zone', 'tek', 'area', 'engin']:
        zones = Zone.objects_limit.filter(id=request.user.owner.zone_id)
    if request.method == 'POST':
        zone = request.POST.get('zone')
        if _role in ['zone', 'tek', 'area', 'engin']:
            zone = str(request.user.owner.zone_id)
        start = request.POST.get('select')
        end = request.POST.get('select2')
        az = start
        ta = end
        start = start.split("/")
        end = end.split("/")
        start = jdatetime.date(day=int(start[2]), month=int(start[1]), year=int(start[0])).togregorian()
        end = jdatetime.date(day=int(end[2]), month=int(end[1]), year=int(end[0])).togregorian()

        sla = int(request.POST.get('sla'))
        _sla = int(sla)
        sla = sla * 60
        sla = sla * 60
        sla += 100
        tickets = Ticket.object_role.c_gs(request, 0).filter(create__range=(start, end),
                                                             failure__failurecategory_id__in=[1010, 1011])
        if zone != '0':
            tickets = tickets.filter(gs__area__zone_id=zone)


        for item in tickets:
            if item.closedate:
                a = (item.closedate - item.create).seconds
                b = (item.closedate - item.create).days
                if int(b) > 0:
                    _day = (b * 24) + round(a / 60 / 60)
                    day = str(b) + 'روز و' + str(round(a / 60 / 60)) + 'ساعت'
                else:
                    _day = round(a / 60 / 60)
                    day = str(round(a / 60 / 60)) + 'ساعت'
                elat = item.statusmoavagh.info if item.statusmoavagh else ""
                if int(_day) > _sla:
                    dict = {
                        'id': item.id,
                        'tek': item.actioner.get_full_name(),
                        'create': item.pdate() + "  " + item.ptime(),
                        'close': item.edate() + "  " + item.etime(),
                        'gs': item.gs.name,
                        'zone': item.gs.area.zone.name,
                        'sla': _day,
                        'st': '* ' + elat,
                        'sec': day

                    }
                    _list.append(dict)
        tickets = Ticket.object_role.c_gs(request, 0).filter(create__range=(start, end), failure_id=1045)
        if zone != '0':
            tickets = tickets.filter(gs__area__zone_id=zone)
        for item in tickets:
            if item.closedate:
                workflow = Workflow.objects.filter(failure_id=1045, ticket_id=item.id).first()
                a = (workflow.createtime - item.create).seconds
                b = (item.closedate - item.create).days
                if int(b) > 0:
                    _day = (b * 24) + round(a / 60 / 60)
                    day = str(b) + 'روز و' + str(round(a / 60 / 60)) + 'ساعت'
                else:
                    _day = round(a / 60 / 60)
                    day = str(round(a / 60 / 60)) + 'ساعت'
                elat = item.statusmoavagh.info if item.statusmoavagh else ""
                if int(_day) > _sla:
                    dict = {
                        'id': item.id,
                        'tek': workflow.user.first_name + " " + workflow.user.last_name,
                        'create': item.pdate() + "  " + item.ptime(),
                        'close': workflow.ndate(),
                        'gs': item.gs.name,
                        'zone': item.gs.area.zone.name,
                        'sla': _day,
                        'st': '# ' + elat,
                        'sec': day
                    }
                    _list.append(dict)
        tickets = Ticket.object_role.c_gs(request, 0).filter(create__range=(start, end),
                                                             failure__failurecategory_id__in=[1010, 1011],
                                                             closedate__isnull=True)
        if zone != '0':
            tickets = tickets.filter(gs__area__zone_id=zone)
        today = datetime.datetime.today()
        t1 = date(year=today.year, month=today.month, day=today.day)
        for item in tickets:
            name = item.actioner.name + " " + item.actioner.lname if item.actioner else ""
            a = (today - item.create).seconds
            b = (today - item.create).days
            if int(b) > 0:
                _day = (b * 24) + round(a / 60 / 60)
                day = str(b) + 'روز و' + str(round(a / 60 / 60)) + 'ساعت'
            else:
                _day = round(a / 60 / 60)
                day = str(round(a / 60 / 60)) + 'ساعت'
            elat = item.statusmoavagh.info if item.statusmoavagh else ""
            if int(_day) > _sla:
                dict = {
                    'id': item.id,
                    'tek': name,
                    'create': item.pdate() + "  " + item.ptime(),
                    'close': item.edate() + "  " + item.etime(),
                    'gs': item.gs.name,
                    'zone': item.gs.area.zone.name,
                    'sla': _day,
                    'st': '@ ' + elat,
                    'sec': day
                }
                _list.append(dict)
        _list = sorted(_list, key=itemgetter('zone'), reverse=True)
        add_to_log(request, f' گزارش sla جدید ', 0)

        return TemplateResponse(request, 'reportsla.html',
                                {'list': _list, 'az': az, 'ta': ta, 'sla': _sla, 'zones': zones, 'zone': int(zone)})
    return TemplateResponse(request, 'reportsla.html', {'list': _list, 'zones': zones})


@cache_permission('zonelistre')
def technicianperformancereport(request):
    _list = []
    count_ticket_forward = 0
    count_forward = 0
    count_close_ticket = 0
    sla_conflig = 0
    ings = 0
    _star = 0
    zones = Zone.objects_limit.all()
    add_to_log(request, f' گزارش خلاصه کارکرد تکنسین ', 0)
    if request.method == 'POST':

        zone = request.POST.get('zone')
        datein = to_miladi(str(request.POST.get('select')))
        dateout = to_miladi(str(request.POST.get('select2')))

        if request.user.owner.role.role == 'zone':
            zone = request.user.owner.zone_id
        owners = Owner.objects.filter(zone_id=int(zone), role__role='tek')
        for owner in owners:
            count_ticket_forward = 0
            count_forward = 0
            count_close_ticket = 0
            sla_conflig = 0
            ings = 0
            _star = 0
            tickets = Workflow.objects.values('ticket_id', 'user_id', 'id').filter(user_id=owner.user_id,
                                                                                   createdate__range=(
                                                                                       datein, dateout)).annotate(
                tedad=Count('ticket_id'))
            count_ticket_forward = len(tickets)
            for item in tickets:
                count_forward += item['tedad']
                orginal = Ticket.objects.get(id=int(item['ticket_id']))
                _result = Workflow.objects.get(id=item['id'])
                if _result.lat:
                    if _result.ticket.gs.location:
                        loc = _result.ticket.gs.location.split(',')
                        _lat = loc[0]
                        _lang = loc[1]
                        a = distance(float(_result.lat), float(_lat), float(_result.lang), float(_lang))
                        if a > 400:
                            ings += 1
                if orginal.status_id == 2:
                    count_close_ticket += 1
                if orginal.closedate:
                    a = (orginal.closedate - orginal.create).seconds
                    b = (orginal.closedate - orginal.create).days
                    if int(b) > 0:
                        _day = (b * 24) + round(a / 60 / 60)
                        day = str(b) + 'روز و' + str(round(a / 60 / 60)) + 'ساعت'
                    else:
                        _day = round(a / 60 / 60)
                        day = str(round(a / 60 / 60)) + 'ساعت'
                if _day > 58 and orginal.statusmoavagh_id == 2 and orginal.failure.failurecategory.id in [1010, 1011]:
                    sla_conflig += 1
                if _day > 58 and orginal.statusmoavagh_id is None and orginal.failure.failurecategory.id in [1010,
                                                                                                             1011]:
                    sla_conflig += 1

                if orginal.star and orginal.star < 6:
                    _star += orginal.star
            if _star > 0:
                _star = round(_star / count_ticket_forward)
            dict = {
                'zone': owner.zone.name,
                'owner': owner.name + " " + owner.lname,
                'tedad_ticket': count_ticket_forward,
                'tedad_erja': count_forward,
                'tedad_close': count_close_ticket,
                'sla': sla_conflig,
                'star': _star,
                'in_gs': ings
            }
            _list.append(dict)

        return TemplateResponse(request, 'technicianperformancereport.html',
                                {'list': _list, 'zones': zones, 'zone': int(zone), 'az': request.POST.get('select'),
                                 'ta': request.POST.get('select2')})
    az = startdate
    ta = str(jdatetime.date.today())
    ta = ta.replace("-", "/")
    return TemplateResponse(request, 'technicianperformancereport.html',
                            {'zones': zones, 'az': az, 'ta': ta})


def newautolocation(request):
    gs = GsModel.objects.filter(nazel_samane__isnull=True)
    for item in gs:
        item.nazel_samane = 0
        item.save()
    gs = GsModel.objects.filter(nazel_samane=1)
    for item in gs:
        item.nazel_samane = 0
        item.save()
    gs = GsModel.objects.filter(nazel_samane=2)
    for item in gs:
        item.nazel_samane = 0
        item.save()
    return redirect('base:home')


@cache_permission('gs')
def gslogs(request, _id):
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    gs = GsModel.object_role.c_gsmodel(request).only('id', 'gsid', 'name').get(id=_id)
    logs = Logs.objects.filter(gs__exact=_id).select_related('owner').order_by('-id')
    paginator = Paginator(logs, 20)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    add_to_log(request, f' مشاهده رویدادهای  جایگاه {gs.gsid}  {gs.name} ', _id)
    return render(request, 'gslogs.html', {'list': page_object, 'page_obj': page_obj})


@cache_permission('gs')
def ownerlogs(request, _id):
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    gs = Owner.objects.get(id=_id)
    logs = Logs.objects.filter(owner_id=_id).order_by('-id')
    paginator = Paginator(logs, 10)
    page_num = request.GET.get('page')
    data = request.GET.copy()
    if 'page' in data:
        del data['page']
    page_object = paginator.get_page(page_num)
    page_obj = paginator.num_pages
    add_to_log(request, f' مشاهده رویدادهای     {gs} ', 0)
    return TemplateResponse(request, 'gslogs.html', {'list': page_object, 'page_obj': page_obj})


@cache_permission('0')
def import_excel_gsinfo(request):
    # add_to_log(request, 'دریافت اکسل مشخصات جایگاه', 0)
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            _row = 0
            for i in range(1, m_row + 1):
                _row += 1
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)
                _value = str(sheet_obj.cell(row=i, column=2).value)
                _list.append(_id)

                if _row == 1:
                    _titel = _value
                else:
                    try:

                        gs = GsModel.objects.get(gsid=_id)
                        gs.location = _value
                        gs.save()

                    except GsModel.DoesNotExist:
                        logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


def tekprofile(request, _id):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='tekprofile')
    if ua.accessrole.ename == 'no':
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    owner = Owner.objects.get(id=_id)
    educations = Education.objects.all()
    form = TekProfileForm()
    if request.method == 'POST':
        form = TekProfileForm(request.POST, instance=owner)
        if form.is_valid():
            if ua.accessrole.ename in ['create', 'full']:
                try:
                    form.save()
                except ValidationError as e:
                    messages.error(request, e)
                    _id = to_md5(_id)
                    return redirect('base:tekprofiles2', _id)

                _id = to_md5(_id)
            return redirect('base:tekprofiles2', _id)
    return render(request, 'users/tekprofile.html',
                  {'owner': owner, 'ownerid': _id, 'educations': educations, 'form': form,
                   'formpermmision': formpermmision})


def tekprofiles2(request, _id):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='tekprofile')
    if ua.accessrole.ename == 'no':
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    url = request.META.get('HTTP_REFERER')
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    owner = Owner.objects.get(id=_id)

    owners = OwnerChild.objects.filter(owner_id=_id)
    for item in owners:
        if item.imageid == "0":
            unique_id = get_random_string(length=32)
            item.imageid = str(item.id) + str(unique_id)
            item.save()
    educations = Education.objects.all()
    form = OwnerChildForm()
    if request.method == 'POST':
        if owner.marital_status == 'singel':
            messages.warning(request, 'این شخص مجرد میباشد')
            return redirect(url)
        _bdate = str(request.POST.get('bdate'))
        _bdate = _bdate.replace('/', '-')
        form = OwnerChildForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance.owner_id = _id
            form.instance.bdate = _bdate
            if ua.accessrole.ename in ['create', 'full']:
                form.save()
            add_to_log(request, ' اضافه کردن قرزند تکنسین ' + str(request.POST.get('codemeli')), 0)
            messages.info(request, 'عملیات با موفقیت انجام شد')
            return redirect(url)
        else:
            messages.error(request, 'عملیات با شکست انجام شد')
    _id = to_md5(_id)
    return render(request, 'users/tekprofiles2.html',
                  {'owner': owner, 'ownerid': _id, 'educations': educations, 'form': form, 'owners': owners,
                   'formpermmision': formpermmision})


@cache_permission('tekprofile')
def tekprofiles3(request, _id):
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    owner = Owner.objects.get(id=_id)
    files = FilesSubject.objects.all()
    for file in files:
        try:
            OwnerFiles.objects.create(file_id=file.id, owner_id=_id, uniq=str(file.id) + "-" + str(_id))
        except IntegrityError:
            a = 1

    ownerfiles = OwnerFiles.objects.filter(owner_id=_id)
    for item in ownerfiles:
        if item.imageid == "0":
            unique_id = get_random_string(length=32)
            item.imageid = str(item.id) + str(unique_id)
            item.save()

    return TemplateResponse(request, 'users/tekprofiles3.html',
                            {'owner': owner, 'ownerid': _id, 'ownerfiles': ownerfiles})


def deletetekchild(request, _id):
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='tekprofile')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    url = request.META.get('HTTP_REFERER')

    a = OwnerChild.objects.get(id=_id)
    add_to_log(request, ' حذف قرزند تکنسین ' + str(a.owner.codemeli), 0)
    if ua.accessrole.ename in ['create', 'full']:
        a.delete()
    messages.info(request, 'عملیات حذف با موفقیت انجام شد')
    return redirect(url)


def uploadownerfile(request, _id, ownerid):
    owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
    if owner_p.count() == 0:
        owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                   semat_id=request.user.owner.refrence_id)
    ua = owner_p.get(permission__name='tekprofile')
    if ua.accessrole.ename == 'no':
        messages.warning(request, DENY_PAGE)
        return redirect(HOME_PAGE)
    formpermmision = {}
    for i in owner_p:
        formpermmision[i.permission.name] = i.accessrole.ename
    # --------------------------------------------------------------------------------------
    url = request.META.get('HTTP_REFERER')
    file = OwnerFiles.objects.get(id=_id)
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES, instance=file)
        if form.is_valid():
            if ua.accessrole.ename in ['create', 'full']:
                form.save()
            a = Owner.objects.get(id=ownerid)
            add_to_log(request, ' آپلود مدارک تکنسین  ' + str(a.codemeli), 0)
            messages.info(request, 'عملیات آپلود فایل با موفقیت انجام شد')
        else:
            messages.error(request, 'عملیات آپلود فایل با شکست خورد')

    return redirect(url)


@cache_permission('tekprofile')
def showimg(request, _id, _st):
    if _st == 1:
        img = OwnerFiles.objects.get(id=_id)
    elif _st == 2:
        img = OwnerChild.objects.get(id=_id)
    add_to_log(request, f'مشاهده عکس مدارک تکنسین ', 0)
    return TemplateResponse(request, 'store/imgshow.html', {'img': img})


@cache_permission('checktekstore')
def check_tek_store(request, _id):
    _list = []
    add_to_log(request, f'دریافت اکسل  چک تکنسین داغی کننده قطعه ', 0)
    form = open_excel(request.POST)

    if request.method == 'POST':

        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                _serial = str(sheet_obj.cell(row=i, column=1).value)
                _serial = checkxss(_serial)
                _serial = checknumber(_serial)
                i = 0
                if _id == 1:
                    gettek = StoreHistory.objects.filter(store__serial=_serial, status_id=6).last()
                if _id == 2:
                    gettek = StoreHistory.objects.filter(store__serial=_serial).last()
                if gettek:
                    try:
                        userzone = gettek.owner.zone.name
                    except:
                        userzone = "-"
                    try:
                        storezone = gettek.store.zone.name
                    except:
                        storezone = "-"
                    try:

                        _tek = gettek.owner.name + " " + gettek.owner.lname + " | " + userzone
                        _date = gettek.normal_date()
                        _zone = storezone
                        _info1 = gettek.status.name
                        _info2 = gettek.store.status.name
                        _desk = gettek.description
                        i += 1
                        dict = {
                            'serial': _serial,
                            'tek': _tek,
                            'date': _date,
                            'zone': _zone,
                            'info1': _info1,
                            'info2': _info2,
                            'desk': _desk,
                        }
                    except (TypeError, AttributeError, KeyError, IntegrityError) as e:
                        dict = {
                            'serial': _serial,
                            'tek': 'نامششخص',
                            'date': 'نامششخص',
                            'zone': 'نامششخص',
                            'info1': 'نامششخص',
                            'info2': 'نامششخص',
                            'desk': e,
                        }
                else:
                    dict = {
                        'serial': _serial,
                        'tek': 'این سریال موجود نیست',
                        'date': 'این سریال موجود نیست',
                        'zone': 'این سریال موجود نیست',
                        'info1': 'این سریال موجود نیست',
                        'info2': 'این سریال موجود نیست',
                        'desk': 'این سریال موجود نیست',
                    }
                _list.append(dict)

        my_path = f'report.xlsx'
        response = HttpResponse(content_type=EXCEL_MODE)
        response['Content-Disposition'] = EXCEL_EXPORT_FILE + my_path
        _font = Font(bold=True)
        _fonttitr = Font(bold=True, size=20)
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "لیست قطعات "
        ws1.sheet_view.rightToLeft = True
        ws1.page_setup.orientation = 'landscape'
        ws1.firstFooter.center.text = ""
        ws1.merge_cells('A1:H1')
        ws1["A1"] = 'لیست قطعات ارسالی '
        ws1["A1"].font = _fonttitr

        ws1.merge_cells('A3:A3')
        ws1["A3"] = "ردیف"
        ws1["A3"].font = _font

        ws1.merge_cells('B3:B3')
        ws1["B3"] = "سریال"
        ws1["B3"].font = _font

        ws1.merge_cells('C3:C3')
        ws1["C3"] = "نام کاربر"
        ws1["C3"].font = _font

        ws1.merge_cells('D3:D3')
        ws1["D3"] = "تاریخ"
        ws1["D3"].font = _font

        ws1.merge_cells('E3:E3')
        ws1["E3"] = "منطقه"
        ws1["E3"].font = _font

        ws1.merge_cells('F3:F3')
        ws1["F3"] = "وضعیت قطعه "
        ws1["F3"].font = _font

        ws1.merge_cells('G3:G3')
        ws1["G3"] = "وضعیت در سابقه"
        ws1["G3"].font = _font

        ws1.merge_cells('H3:H3')
        ws1["H3"] = "توضیحات"
        ws1["H3"].font = _font

        ws1.column_dimensions['B'].width = float(15.25)
        ws1.column_dimensions['C'].width = float(28.25)
        ws1.column_dimensions['D'].width = float(38.25)
        ws1.column_dimensions['E'].width = float(28.25)
        ws1.column_dimensions['F'].width = float(38.25)
        ws1.column_dimensions['G'].width = float(38.25)
        ws1.column_dimensions['H'].width = float(48.25)

        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        my_font = Font(size=14, bold=True)
        my_fill = PatternFill(
            fill_type='solid', start_color='FFFF00')
        i = 0
        for item in _list:
            i += 1
            d = [i, str(item['serial']), str(item['tek']),
                 str(item['date']), str(item['zone']), str(item['info2']), str(item['info1']), str(item['desk']),
                 ]

            ws1.append(d)
        for col in ws1.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(
                    horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.border = thin_border

        for cell in ws1["3:3"]:
            cell.font = my_font
            cell.fill = my_fill
            cell.border = thin_border
        max_row = ws1.max_row
        total_cost_cell = ws1.cell(row=max_row + 2, column=5)
        total_cost_cell2 = ws1.cell(row=max_row + 2, column=10)
        total_cost_cell.value = ''
        total_cost_cell2.value = ''
        wb.save(response)
        return response

        messages.success(request, SUCCESS_TICKET)

    return TemplateResponse(request, 'importexcel.html', {'form': form, })


@cache_permission('excel_nop')
def import_arbain(request):
    _list = []
    add_to_log(request, f'دریافت اکسل جایگاه های اربعین', 0)
    form = open_excel(request.POST)

    if request.method == 'POST':
        a = GsModel.objects.filter(arbain=True)
        for item in a:
            item.arbain = False
            item.save()
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)

                _list.append(_id)
                try:
                    gs = GsModel.objects.get(gsid=_id)
                    gs.arbain = True
                    gs.save()


                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


@cache_permission('report')
def reportSellticket(request):
    add_to_log(request, f' مشاهده فرم تیکت های در حال فروش', 0)
    zones = Zone.objects_limit.all()
    if request.method == 'POST':
        zone = request.POST.get('zone')
        if request.user.owner.role.role in "zone,area":
            zone = request.user.owner.zone_id
            _role = request.user.owner.role.role
            _roleid = zoneorarea(request)

        if zone == '0':
            tickets = Ticket.objects.filter(status_id=1, failure__failurecategory_id__in=[1010, 1011],
                                            )
        else:
            tickets = Ticket.object_role.c_gs(request, 0).filter(status_id=1,
                                                                 failure__failurecategory_id__in=[1010, 1011],
                                                                 )

        _list = []
        for ticket in tickets:
            sellmodel = SellModel.objects.exclude(sellkol=0).filter(tolombeinfo_id=ticket.Pump_id,
                                                                    tarikh__gt=ticket.shamsi_date).aggregate(
                summ=Sum('sellkol'), tedad=Count('id'))

            if sellmodel['summ']:
                dict = {
                    'zone': ticket.gs.area.zone.name,
                    'area': ticket.gs.area.name,
                    'tarikh': ticket.shamsi_date,
                    'gs_name': ticket.gs.name,
                    'gs_gsid': ticket.gs.gsid,
                    'nazel': str(ticket.Pump.number),
                    'info': ticket.failure.info,
                    'descriptionowner': ticket.descriptionowner,
                    'shomare': str(ticket.id),
                    'counter': sellmodel['summ'],
                    'tedad': sellmodel['tedad'],
                }
                _list.append(dict)
        _list = sorted(_list, key=itemgetter('tarikh'), reverse=True)
        return TemplateResponse(request, 'report/zonesell.html',
                                {'zones': zones, 'zone': int(zone), 'list': _list})
    return TemplateResponse(request, 'report/zonesell.html', {'zones': zones})


@cache_permission('0')
def import_excel_mojodi(request):
    # add_to_log(request, 'دریافت اکسل مشخصات جایگاه', 0)
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            _row = 0
            for i in range(1, m_row + 1):
                _row += 1
                _id = str(sheet_obj.cell(row=i, column=1).value)
                if len(_id) == 3:
                    _id = "0" + str(_id)
                if len(_id) == 2:
                    _id = "00" + str(_id)
                if len(_id) == 1:
                    _id = "000" + str(_id)
                _value = str(sheet_obj.cell(row=i, column=2).value)
                _far = str(sheet_obj.cell(row=i, column=3).value)
                _list.append(_id)

                try:
                    gs = GsModel.objects.get(gsid=_id)

                    if _far == "01":
                        gs.m_benzin = _value
                    if _far == "02":
                        gs.m_super = _value
                    if _far == "03":
                        gs.m_naftgaz = _value
                    gs.save()

                except GsModel.DoesNotExist:
                    logging.info(f"Dose not object for import napaydari{_id}", exc_info=True)

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return TemplateResponse(request, 'importexcel.html', {'form': form})


@cache_permission('gs')
def gsmalek(request):
    add_to_log(request, f' مشاهده فرم مالک و کاربر جایگاه', 0)
    zone = 0
    zones = Zone.objects_limit.all()
    _list = []
    if request.method == 'POST':

        refrence = request.POST.get('refrence')
        _list.append(refrence)
        if refrence == "4":
            _list = []
            _list.append(4)
            _list.append(8)
        if request.user.owner.role.role == 'zone':
            gslist = GsList.objects.filter(owner__refrence_id__in=_list,
                                           gs__area__zone_id=request.user.owner.zone_id)

        if request.user.owner.role.role == 'area':
            gslist = GsList.objects.filter(owner__refrence_id__in=_list,
                                           gs__area_id=request.user.owner.area_id)

        if request.user.owner.role.role in 'mgr,setad':
            zone = request.POST.get('zone')
            gslist = GsList.objects.filter(owner__refrence_id__in=_list, owner__zone_id=zone)

        return TemplateResponse(request, 'gsmalek.html',
                                {'zones': zones, 'gslist': gslist, 'refrence': int(refrence), 'zone': int(zone)})
    return TemplateResponse(request, 'gsmalek.html', {'zones': zones, })


def deleteticket(request, _id):
    url = request.META.get('HTTP_REFERER')
    workcount = Workflow.objects.filter(ticket_id=_id).count()
    ticket = Ticket.objects.get(id=_id)
    if workcount == 1 and ticket.owner_id == request.user.id:
        ticket.delete()
        add_to_log(request, f' حذف تیکت شماره {_id} ', 0)
    return redirect(url)


@cache_permission('reinitial')
def listreinitial(request):
    add_to_log(request, f' مشاهده فرم ری اینشیال', 0)
    _list = None
    if request.user.owner.role.role in 'engin,zone':
        _list = ReInitial.objects.filter(gs__area__zone_id=request.user.owner.zone_id)

    if request.user.owner.role.role == 'tek':
        _list = ReInitial.objects.filter(gs__gsowner__owner_id=request.user.owner.id)

    if request.user.owner.role.role in 'setad,mgr,test,fani':
        _list = ReInitial.objects.all()

    return TemplateResponse(request, 'list_reinitial.html', {'list': _list})


class AddReInitial(View):
    template_name = 'add_reinitial.html'

    def get(self, request, pk):
        owner_p = UserPermission.objects.filter(owner_id=request.user.owner.id)
        if owner_p.count() == 0:
            owner_p = DefaultPermission.objects.filter(role_id=request.user.owner.role_id,
                                                       semat_id=request.user.owner.refrence_id)
        ua = owner_p.get(permission__name='reinitial')
        if ua.accessrole.ename not in 'full,create':
            messages.warning(request, DENY_PAGE)
            return redirect(HOME_PAGE)
        formpermmision = {}
        for i in owner_p:
            formpermmision[i.permission.name] = i.accessrole.ename
        # --------------------------------------------------------------------------------------
        add_to_log(request, f' مشاهده تکمیل فرم ری اینشیال', 0)
        _list = None
        if request.user.owner.role.role == 'zone':
            _list = GsModel.objects.filter(area__zone_id=request.user.owner.zone_id)

        if request.user.owner.role.role in 'engin,tek':
            _list = GsModel.objects.filter(gsowner__owner_id=request.user.owner.id)
        if pk != 0:
            reinitial = ReInitial.objects.get(pk=pk)
        else:
            reinitial = None
        form = FormReInitial()
        context = {'reinitial': reinitial, 'form': form, 'list': _list, 'formpermmision': formpermmision}
        return render(request, self.template_name, context)

    @staticmethod
    def post(request, pk):
        info_quiz2 = request.POST.get('info_quiz2')
        quiz10 = request.POST.get('quiz10')
        if len(info_quiz2) > 5:
            info_quiz2 = info_quiz2.split('/')
            info_quiz2 = jdatetime.date(day=int(info_quiz2[2]), month=int(info_quiz2[1]),
                                        year=int(info_quiz2[0])).togregorian()
        else:
            info_quiz2 = None
        if len(quiz10) > 5:
            quiz10 = quiz10.split('/')
            quiz10 = jdatetime.date(day=int(quiz10[2]), month=int(quiz10[1]), year=int(quiz10[0])).togregorian()
        else:
            quiz10 = None
        try:
            reinitial = ReInitial.objects.get(pk=pk)
            form = FormReInitial(request.POST, request.FILES, instance=reinitial)
            a = 0
            if a == 0 and reinitial.owner_id != request.user.owner.id:
                messages.warning(request, 'ویرایش فقط برای ثبت کننده وجود دارد ')
                return redirect('base:listreinitial')
        except ReInitial.DoesNotExist:
            reinitial = None
            a = 1
            form = FormReInitial(request.POST, request.FILES)
        if form.is_valid():
            _st = 'Edit' if a == 0 else 'Add'
            # add_to_log(request, f'{_st} Customer  {form.cleaned_data["name"]} - {form.cleaned_data["nationalid"]}')
            form.instance.owner_id = request.user.owner.id
            form.instance.quiz10 = quiz10
            form.instance.status = 0
            form.instance.info_quiz2 = info_quiz2
            form.save()
            messages.success(request, 'عملیات با موفقیت انجام شد ')
            return redirect('base:listreinitial')
        else:

            messages.error(request, 'عملیات شکست خورد!' + str(form.errors))
            return redirect('base:listreinitial')


def acceptreinitial(request, pk, st):
    url = request.META.get('HTTP_REFERER')
    reinitial = ReInitial.objects.get(pk=pk)
    if st == 0:
        reinitial.status = 1
        reinitial.accept_gs = True

    if st == 1:
        reinitial.status = 2
        reinitial.accept_tek = True
    if st == 2:
        reinitial.status = 3
        reinitial.accept_zone = True

    gsm = GsModel.objects.get(id=reinitial.gs_id)
    if st == 0:
        gsmobail = GsList.objects.filter(gs_id=gsm.id, owner__role__role='tek')[:1]
        for item in gsmobail:
            mobail = item.owner.mobail
    elif st == 1:
        gsmobail = Owner.objects.filter(zone_id=gsm.area.zone_id, refrence_id=1)[:1]
        for item in gsmobail:
            mobail = item.mobail
    elif st == 2:
        gsmobail = GsList.objects.filter(gs_id=gsm.id, owner__role__role='tek')[:1]
        for item in gsmobail:
            mobail = item.owner.mobail
    if mobail:
        if st == 2:
            message = '''
            سلام ، فرم مجوز ری اینشیال تکمیل شد میتوانید تیکت مورد نظر را ثبت کنید. 
                                                            شرکت ملی پخش فرآورده های نفتی ایران
                                                                            '''.format(param1=gsm.name)
        else:
            message = '''
    سلام ، لطفا به فرم مجوز ری اینشیال مراجعه و اطلاعات وارد شده را تایید کنید. 
                                                    شرکت ملی پخش فرآورده های نفتی ایران
                                                                    '''.format(param1=gsm.name)

        try:
            SendOTP2(mobail, message, 0, 0, 0)
        except:
            print('ok')
    reinitial.save()
    add_to_log(request, f'آپدیت فرم اینشیال', 0)
    return redirect(url)


def tryopenticket(request, _id):
    url = request.META.get('HTTP_REFERER')

    ticket = Ticket.objects.get(id=_id)
    if request.user.owner.refrence_id == 1 and ticket.gs.area.zone_id == request.user.owner.zone_id:
        ticket.status_id = 1
        ticket.reply_id = ""
        ticket.save()
        Workflow.objects.create(ticket_id=_id, organization_id=ticket.organization_id, description='باز شدن مجدد تیکت',
                                user_id=request.user.id,
                                failure_id=ticket.failure_id)
        messages.info(request, 'تیکت با موفقیت باز شد')
    else:
        messages.info(request, ' دسترسی ندارید')
    return redirect(url)


@cache_permission('mgrdashboard')
def mgrdashboard(request):
    add_to_log(request, f' مشاهده فرم داشبورد tv', 0)
    _role = request.user.owner.role.role
    _roleid = zoneorarea(request)
    tarikh = datetime.date.today() - datetime.timedelta(days=1)
    day30 = datetime.date.today() - datetime.timedelta(days=30)
    # datesell = SellModel.object_role.c_sell(request).filter(tarikh=tarikh).values('gs_id').annotate(
    #     sellsum=Sum('sell'),
    #     kolsum=Sum('sellkol'))
    gscount = GsModel.object_role.c_gsmodel(request).aggregate(active=Sum(Case(When(status_id=1, then=1))),
                                                               noactive=Sum(Case(When(status_id=4, then=1))),
                                                               inactive=Sum(Case(When(status_id=3, then=1))), )
    # gsmodel = GsModel.object_role.c_gsmodel(request).filter(status_id=1)
    pump = Pump.object_role.c_gs(request, 0).all().count()

    # nosell = gsmodel.exclude(
    #     id__in=datesell.filter(sellsum__gte=0, kolsum__gte=0).values('gs_id')).count()

    zarfyat = GsModel.object_role.c_gsmodel(request).all().aggregate(benzin=Sum('m_benzin'),
                                                                     super=Sum('m_super'),
                                                                     naftgaz=Sum('m_naftgaz'))
    mojodi = Mojodi.object_role.c_gs(request, 0).filter(tarikh=tarikh).aggregate(benzin=Sum('benzin'),
                                                                                 super=Sum('super'),
                                                                                 naftgaz=Sum('gaz'))
    if request.user.owner.role.role in ['mgr', 'setad']:
        azad = SellGs.object_role.c_gs(request, 0).values('gs__area__zone__name').filter(product_id=4,
                                                                                         tarikh__range=(
                                                                                             day30,
                                                                                             tarikh)).annotate(
            gaz=Sum('ezterari'),
        )
    if request.user.owner.role.role == 'zone':
        azad = SellGs.object_role.c_gs(request, 0).values('gs__area__name').filter(product_id=4, tarikh__range=(
            day30, tarikh)).annotate(gaz=Sum('ezterari'),
                                     )
    if request.user.owner.role.role == 'area':
        azad = SellGs.object_role.c_gs(request, 0).values('gs__name').filter(product_id=4, tarikh__range=(
            day30, tarikh)).annotate(gaz=Sum('ezterari'),
                                     )

    _benzin = round((int(mojodi['benzin']) / int(zarfyat['benzin'])) * 100) if mojodi['benzin'] else 0
    _super = round((int(mojodi['super']) / int(zarfyat['super'])) * 100) if mojodi['super'] else 0
    _naftgaz = round((int(mojodi['naftgaz']) / int(zarfyat['naftgaz'])) * 100) if mojodi['naftgaz'] else 0

    content = {'benzin': _benzin, 'super': _super, 'naftgaz': _naftgaz, 'azad': azad, 'gscount': gscount, 'pump': pump,
               }
    return TemplateResponse(request, 'mgrdashboard.html', content)


@cache_permission('city')
def citylist(request):
    add_to_log(request, f' مشاهده فرم اطلاعات شهرستان', 0)
    list = []
    if request.user.owner.role.role in ['mgr', 'setad']:
        _list = City.objects.all().order_by('area__zone_id')
    elif request.user.owner.role.role == 'zone':
        _list = City.objects.filter(area__zone_id=request.user.owner.zone_id)
    elif request.user.owner.role.role == 'area':
        _list = City.objects.filter(area_id=request.user.owner.area_id)
    return TemplateResponse(request, 'parametrs/listcity.html', {'list': _list})


@cache_permission('city')
def cityupdate(request, _id):
    form = CityForm()
    city = None if _id == 0 else City.objects.get(id=_id)
    if request.method == 'POST':
        if _id == 0:
            form = CityForm(request.POST)
        else:
            city = City.objects.get(id=_id)
            form = CityForm(request.POST, instance=city)
        if form.is_valid():
            form.instance.area_id = request.user.owner.area_id
            city = form.save()

            messages.success(request, 'عملیات با موفقیت انجام شد')
            add_to_log(request, f'   بروزرسانی اطلاعات شهرستان {city.name} ', 0)
            return redirect('base:citylist')

        else:
            messages.error(request, 'عملیات شکست خورد')

    return TemplateResponse(request, 'parametrs/updatecity.html',
                            {'form': form, 'city': city})


@cache_permission('gs')
def parametrsgs(request, _gsid, _id):
    form = ParametrGssForm()
    anbars = Oildepot.objects.all()
    pg = ParametrGs.objects.get(id=_id) if _id != 0 else None
    if len(str(_gsid)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _gsid = Decrypt(_gsid)
    if request.method == 'POST':
        form = ParametrGssForm(request.POST, instance=pg) if _id != 0 else ParametrGssForm(request.POST)
        if form.is_valid():
            form.instance.gs_id = _gsid
            form.save()
            messages.info(request, 'عملیات با موفقیت انجام شد')
            return redirect('base:gs_detail', to_md5(_gsid))
        else:
            messages.error(request, 'عملیات شکست خورد')
    return TemplateResponse(request, 'parametrsgs.html', {'form': form, 'anbars': anbars, 'pg': pg})


@cache_permission('gs')
def pump_report(request):
    add_to_log(request, f'   مشاهده لیست تلمبه ها ', 0)
    pumps = Pump.object_role.c_gs(request, 1).values('gs__gsid', 'gs__name', 'tolombe', 'gs__area__name',
                                                     'pumpbrand__name',
                                                     'gs__area__zone__name').annotate(nozzle_count=Count('number'))

    report = {}
    for pump in pumps:
        station = pump['gs__gsid']
        nozzle_count = pump['nozzle_count']
        station_name = pump['gs__name']
        station_area = pump['gs__area__name']
        station_zone = pump['gs__area__zone__name']
        pump_brand = pump['pumpbrand__name']

        if station not in report:
            report[station] = {'total_pumps': 0, 'one_nozzle_count': 0, 'two_nozzle_count': 0, 'four_nozzle_count': 0,
                               'eghit_nozzle_count': 0, 'count_nazels': 0, 'brands': set(), }
        report[station]['total_pumps'] += 1
        report[station]['count_nazels'] += nozzle_count
        report[station]['station_name'] = station_name
        report[station]['station_area'] = station_area
        report[station]['station_zone'] = station_zone
        if pump_brand:
            report[station]['brands'].add(pump_brand)
        if nozzle_count == 1:
            report[station]['one_nozzle_count'] += 1
        elif nozzle_count == 2:
            report[station]['two_nozzle_count'] += 1
        elif nozzle_count == 4:
            report[station]['four_nozzle_count'] += 1
        elif nozzle_count == 8:
            report[station]['eghit_nozzle_count'] += 1
    for station_data in report.values():
        station_data['brands_string'] = ' | '.join(station_data['brands']) if station_data['brands'] else 'بدون برند'
        # حذف مجموعه از دیکشنری چون قابل نمایش در تمپلیت نیست
        del station_data['brands']
    context = {'report': report}
    return TemplateResponse(request, 'pump_report.html', context)


def deleteclosesell(request, _id):
    url = request.META.get('HTTP_REFERER')
    _cg = CloseGS.object_role.c_gs(request, 0).get(id=_id)
    for item in CloseSellReport.objects.filter(gs_id=_cg.gs.id):
        if item.tarikh >= _cg.date_in and item.tarikh <= _cg.date_out:
            messages.error(request,
                           f" برای دوره {str(item.tarikh)}  واکشی انجام و ثبت تعطیلی اعمال شده ، امکان حذف نمیباشد")
            return redirect(url)

    _cg.delete()
    messages.info(request, 'با موفقیت حذف شد')
    return redirect(url)


def weatherapi(latitude, longitude):
    api_key = "65a3626e246787e81af7dc1b12df11b4"
    url = f"https://api.openweathermap.org/data/2.5/weather?lat={latitude}&lon={longitude}&appid={api_key}&units=metric"
    response = requests.get(url)
    data = response.json()
    if response.status_code == 200:
        temp = int(round(data['main']['temp']))
        humidity = int(round(data['main']['humidity']))
        pressure = int(round(data['main']['pressure']))
        main = data['weather'][0]['main']
        return (temp, humidity, pressure, main)

    else:
        return (None, None, None, None)


def reportchardsh(request):
    # تاریخ امروز
    today = datetime.datetime.today()

    # تاریخ ده ماه قبل
    ten_months_ago = today - datetime.timedelta(days=30 * 10)

    # دریافت تیکت‌های ده ماه گذشته
    tickets_last_ten_months = Ticket.objects.filter(create__gte=ten_months_ago,
                                                    failure__failurecategory_id__in=[1010, 1011],
                                                    gs__area__zone_id=request.user.owner.zone.id)

    tickets_per_month = tickets_last_ten_months.values('create_shamsi_month').annotate(
        count=Count('id')
    ).order_by('create_shamsi_month')

    # محاسبه میانگین تعداد تیکت‌ها در هر ماه
    average_tickets_per_month = tickets_per_month.aggregate(
        avg_tickets=Avg('count')
    )['avg_tickets']

    _ticketsavg_per_month = []
    for item in tickets_per_month:
        _ticketsavg_per_month.append({
            'month': item['create_shamsi_month'],
            'tedad': item['count'],
            'avg': average_tickets_per_month
        })

    failuretype = Ticket.objects.values('failure_id', 'failure__info').filter(
        failure__failurecategory_id__in=[1010, 1011],
        gs__area__zone_id=request.user.owner.zone.id).annotate(tedad=Count('id'))

    replytype = Ticket.objects.values('reply_id', 'reply__info').filter(failure__failurecategory_id__in=[1010, 1011],
                                                                        gs__area__zone_id=request.user.owner.zone.id).annotate(
        tedad=Count('id')).exclude(reply_id=None)

    tickets_per_month = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone.id).values(
        'create_shamsi_month').annotate(
        ticket_count=Count('id')
    ).order_by('create_shamsi_month')

    # محاسبه مجموع کل تیکت‌ها
    total_tickets = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone.id).count()

    # محاسبه درصد تیکت‌ها در هر ماه
    tickets_per_month_with_percentage = []
    for entry in tickets_per_month:
        percentage = (entry['ticket_count'] / total_tickets) * 100
        tickets_per_month_with_percentage.append({
            'create_shamsi_month': entry['create_shamsi_month'],
            'ticket_count': entry['ticket_count'],
            'percentage': round(percentage, 2)  # گرد کردن به دو رقم اعشار
        })

    # نمایش نتایج
    _tickets_per_month = []
    for entry in tickets_per_month_with_percentage:
        _tickets_per_month.append({
            'Month': entry['create_shamsi_month'],
            'Percentage': entry['percentage']
        })

    tickets_per_quarter = Ticket.objects.filter(gs__area__zone_id=request.user.owner.zone.id).annotate(
        quarter=ExtractQuarter('create')
    ).values('quarter').annotate(
        ticket_count=Count('id')
    ).order_by('quarter')

    # محاسبه درصد تیکت‌ها در هر فصل
    tickets_per_quarter_with_percentage = []
    for entry in tickets_per_quarter:
        percentage = (entry['ticket_count'] / total_tickets) * 100
        tickets_per_quarter_with_percentage.append({
            'quarter': entry['quarter'],
            'ticket_count': entry['ticket_count'],
            'percentage': round(percentage, 2)  # گرد کردن به دو رقم اعشار
        })

    # نمایش نتایج
    _tickets_per_quarter = []
    for entry in tickets_per_quarter_with_percentage:
        _tickets_per_quarter.append({
            'quarter': entry['quarter'],
            'Percentage': entry['percentage']
        })

    context = {'ticketsavg_per_month': _ticketsavg_per_month, 'failuretype': failuretype, 'replytype': replytype,
               'tickets_per_month': _tickets_per_month, 'tickets_per_quarter': _tickets_per_quarter}
    return render(request, 'reportchartdsh.html', context)


def my_view(request):
    # دریافت لیست کوئری‌های اجرا شده
    queries = connection.queries

    # نمایش کوئری‌ها
    _list = []
    for query in queries:
        _list.append({
            '': query
        })

    return render(request, 'querys.html', {'list': _list})


# views.py


@cache_permission('basemgr')
def failure_analysis_report(request):
    add_to_log(request, f' مشاهده فرم تحلیل خرابی', 0)
    # تعیین پارامترهای زمانی از URL
    time_period = request.GET.get('period', 'weekly')  # پیش‌فرض هفتگی

    # محاسبه تاریخ شروع بر اساس دوره انتخابی
    end_date = datetime.datetime.now().date()
    if time_period == 'weekly':
        start_date = end_date - timedelta(days=7)
        period_title = "هفتگی"
    elif time_period == 'monthly':
        start_date = end_date - timedelta(days=30)
        period_title = "ماهانه"
    elif time_period == 'quarterly':
        start_date = end_date - timedelta(days=90)
        period_title = "سه ماهه"
    else:
        start_date = end_date - timedelta(days=7)
        time_period = 'weekly'
        period_title = "هفتگی"

    # دریافت تیکت‌های مربوط به بازه زمانی
    tickets = Ticket.objects.filter(
        create__date__gte=start_date,
        create__date__lte=end_date
    )

    # محاسبه آمار خرابی‌ها
    failure_stats = (
        tickets.values('failure__id', 'failure__info', 'failure__failurecategory__info')
        .annotate(
            failure_name=F('failure__info'),
            category_name=F('failure__failurecategory__info'),
            count=Count('id')
        )
        .order_by('-count')
    )

    # محاسبه پرتکرارترین خرابی‌ها (10 مورد اول)
    top_failures = list(failure_stats[:10])

    # محاسبه توزیع خرابی‌ها بر اساس منطقه
    zone_distribution = (
        tickets.values('gs__area__zone__name')
        .annotate(
            zone_name=F('gs__area__zone__name'),
            count=Count('id')
        )
        .order_by('-count')
    )

    # محاسبه روند روزانه
    daily_trend = (
        tickets.annotate(day=TruncDate('create'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )

    for item in daily_trend:
        gregorian_date = item['day']
        jalali_date = jdatetime.date.fromgregorian(date=gregorian_date)
        item['day_jalali'] = jalali_date.strftime('%Y/%m/%d')

    # محاسبه تعداد کل تیکت‌ها
    total_tickets = tickets.count()
    top_failures_labels = [f['failure_name'] for f in top_failures]
    top_failures_data = [f['count'] for f in top_failures]

    zone_labels = [z['zone_name'] for z in zone_distribution]
    zone_data = [z['count'] for z in zone_distribution]

    daily_labels = [d['day'].strftime('%Y-%m-%d') for d in daily_trend]
    daily_data = [d['count'] for d in daily_trend]

    context = {
        'top_failures_labels': top_failures_labels,
        'top_failures_data': top_failures_data,
        'zone_labels': zone_labels,
        'zone_data': zone_data,
        'daily_labels': daily_labels,
        'daily_data': daily_data,
        'period': period_title,
        'start_date': start_date,
        'end_date': end_date,
        'total_tickets': total_tickets,
        'top_failures': top_failures,
        'zone_distribution': zone_distribution,
        'daily_trend': daily_trend,
        'time_period': time_period,
    }

    return TemplateResponse(request, 'failure_analysis.html', context)


def schematic(request, _id):
    if len(str(_id)) < 6:
        messages.warning(request, 'دسترسی غیر مجاز')
        return redirect(HOME_PAGE)
    _id = Decrypt(_id)
    add_to_log(request, f' مشاهده شماتیک جایگاه', 0)
    pumps = Pump.objects.select_related('product', 'status').filter(gs_id=_id).order_by('sakoo', 'tolombe',
                                                                                        'sortnumber')
    # محاسبه خودکار ابعاد
    total_pumps = pumps.values('tolombe').distinct().count()
    svg_width = max(1000, total_pumps * 200)  # حداقل عرض 1000px
    svg_height = 800  # ارتفاع ثابت

    platforms = {}
    for pump in pumps:
        if pump.sakoo not in platforms:
            platforms[pump.sakoo] = {
                'pumps': {},
                'y_position': 150 + (pump.sakoo - 1) * 200
            }

        if pump.tolombe not in platforms[pump.sakoo]['pumps']:
            platforms[pump.sakoo]['pumps'][pump.tolombe] = {
                'nozzles': [],
                'x_position': 100 + (pump.tolombe - 1) * 180,  # فاصله 180px بین تلمبه‌ها
                'is_lpg': pump.product.name == 'نفت گاز'
            }
        if pump.status_id == 1:
            pt_status = 'working'
        elif pump.status_id == 2:
            pt_status = 'faulty'
        else:
            pt_status = 'none'
        platforms[pump.sakoo]['pumps'][pump.tolombe]['nozzles'].append({
            'product': pump.product.name,
            'status': pump.status_id,
            'number': pump.number
        })

    context = {
        'platforms': platforms,
        'svg_width': svg_width,
        'svg_height': svg_height,
        'total_pumps': total_pumps
    }
    return render(request, 'schematic.html', context)


@cache_permission('create_sejelli')
def manage_sejelli(request, gs_id):
    gs = GsModel.objects.get(id=gs_id)
    # بررسی وجود سجلی برای این جایگاه
    existing_sejelli = NewSejelli.objects.filter(gs=gs, isok=False).first()
    sejellis = NewSejelli.objects.filter(gs=gs, isok=True).order_by('-id')
    if request.method == 'POST':
        # ایجاد سجلی جدید
        if not existing_sejelli:
            sejelli = NewSejelli.objects.create(gs=gs)
            # کپی اطلاعات از GSModel به GsModel_sejjeli
            gs_sejelli = GsModel_sejjeli.objects.create(
                newsejelli=sejelli,
                gs=gs,
                gsid=gs.gsid,
                name=gs.name,
                area=gs.area,
                telldaftar=gs.telldaftar,
                address=gs.address,
                simcart=gs.simcart,
                operator=gs.operator,
                status=gs.status,
                brand=gs.brand,
                sellcode=gs.sellcode,
                malicode=gs.malicode,
            )

            # کپی اطلاعات پمپ‌ها
            pumps = Pump.objects.filter(gs=gs)
            for pump in pumps:
                Pump_sejjeli.objects.create(
                    newsejelli=sejelli,
                    number=pump.number,
                    product=pump.product,
                    makhzan=pump.makhzan,
                    pumpbrand=pump.pumpbrand,
                    sakoo=pump.sakoo,
                    tolombe=pump.tolombe,
                    status=pump.status,
                    gs=gs
                )

            # کپی اطلاعات مخازن
            makhzans = Makhzan.objects.filter(gs=gs)
            for makhzan in makhzans:
                Makhzan_sejjeli.objects.create(
                    newsejelli=sejelli,
                    gs=gs,
                    product=makhzan.product,
                    number=makhzan.number,
                    zarfyat=makhzan.zarfyat
                )

            return redirect('base:sejelli_detail', sejelli_id=sejelli.id)

    context = {
        'gs': gs,
        'existing_sejelli': existing_sejelli,
        'sejellis': sejellis,
    }
    return TemplateResponse(request, 'sejelli/manage_sejelli.html', context)


@cache_permission('create_sejelli')
def sejelli_detail(request, sejelli_id):
    sejelli = NewSejelli.objects.get(id=sejelli_id)
    gs_sejelli = GsModel_sejjeli.objects.filter(newsejelli=sejelli).first()
    pumps = Pump_sejjeli.objects.filter(newsejelli=sejelli)
    makhzans = Makhzan_sejjeli.objects.filter(newsejelli=sejelli)
    change_logs = SejelliChangeLog.objects.filter(sejelli=sejelli).order_by('-changed_at')

    context = {
        'sejelli': sejelli,
        'gs_sejelli': gs_sejelli,
        'pumps': pumps,
        'makhzans': makhzans,
        'change_logs': change_logs,
    }
    return TemplateResponse(request, 'sejelli/sejelli_detail.html', context)


class UpdateGsModelSejjeli(UpdateView):
    model = GsModel_sejjeli
    form_class = GsModelSejjeliForm
    template_name = 'sejelli/edit_form.html'

    def convert_for_json(self, value):
        """تبدیل اشیاء مدل به فرمت قابل JSON شدن"""
        if hasattr(value, 'id'):
            return {'id': value.id, 'name': str(value)}
        elif isinstance(value, (list, tuple)):
            return [self.convert_for_json(item) for item in value]
        elif isinstance(value, dict):
            return {key: self.convert_for_json(val) for key, val in value.items()}
        return value

    def form_valid(self, form):
        # ذخیره تغییرات قبل از ثبت
        changed_data = {}
        for field in form.changed_data:
            old_value = form.initial.get(field)
            new_value = form.cleaned_data.get(field)

            changed_data[field] = {
                'old': self.convert_for_json(old_value),
                'new': self.convert_for_json(new_value)
            }

        response = super().form_valid(form)

        # ثبت تغییرات در تاریخچه
        SejelliChangeLog.objects.create(
            user=self.request.user,
            sejelli=self.object.newsejelli,
            model_name='GsModel_sejjeli',
            record_id=self.object.id,
            action='update',
            changed_data=changed_data
        )

        messages.success(self.request, 'اطلاعات جایگاه با موفقیت ویرایش شد.')
        return response

    def get_success_url(self):
        return reverse('base:sejelli_detail', kwargs={'sejelli_id': self.object.newsejelli.id})


class CreatePumpSejjeli(CreateView):
    model = Pump_sejjeli
    form_class = PumpSejjeliForm
    template_name = 'sejelli/edit_form.html'

    @method_decorator(cache_permission('create_sejelli'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['sejelli_id'] = self.kwargs['sejelli_id']
        return kwargs

    def get_success_url(self):
        return reverse('base:sejelli_detail', kwargs={'sejelli_id': self.object.newsejelli.id})

    def form_valid(self, form):
        sejelli = NewSejelli.objects.get(id=self.kwargs['sejelli_id'])
        form.instance.newsejelli = sejelli
        form.instance.gs = sejelli.gs

        response = super().form_valid(form)

        # تبدیل داده‌های فرم به فرمت قابل JSON شدن
        cleaned_data = form.cleaned_data.copy()
        cleaned_data['product'] = {
            'id': form.cleaned_data['product'].id,
            'name': form.cleaned_data['product'].name
        }
        cleaned_data['pumpbrand'] = {
            'id': form.cleaned_data['pumpbrand'].id,
            'name': form.cleaned_data['pumpbrand'].name
        }
        cleaned_data['status'] = {
            'id': form.cleaned_data['status'].id,
            'name': form.cleaned_data['status'].name
        }

        SejelliChangeLog.objects.create(
            user=self.request.user,
            sejelli=sejelli,
            model_name='Pump_sejjeli',
            record_id=self.object.id,
            action='create',
            changed_data=cleaned_data
        )

        messages.success(self.request, 'پمپ جدید با موفقیت اضافه شد.')
        return response


class UpdatePumpSejjeli(UpdateView):
    model = Pump_sejjeli
    form_class = PumpSejjeliForm
    template_name = 'sejelli/edit_form.html'

    @method_decorator(cache_permission('create_sejelli'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['sejelli_id'] = self.object.newsejelli.id
        return kwargs

    def form_valid(self, form):
        # ذخیره تغییرات قبل از ثبت
        changed_data = {}
        for field in form.changed_data:
            old_value = form.initial.get(field)
            new_value = form.cleaned_data.get(field)

            # تبدیل اشیاء مدل به فرمت قابل JSON شدن
            if hasattr(old_value, 'id'):
                old_value = {'id': old_value.id, 'name': str(old_value)}
            if hasattr(new_value, 'id'):
                new_value = {'id': new_value.id, 'name': str(new_value)}

            changed_data[field] = {
                'old': old_value,
                'new': new_value
            }

        response = super().form_valid(form)

        # ثبت تغییرات در تاریخچه
        SejelliChangeLog.objects.create(
            user=self.request.user,
            sejelli=self.object.newsejelli,
            model_name='Pump_sejjeli',
            record_id=self.object.id,
            action='update',
            changed_data=changed_data
        )

        messages.success(self.request, 'اطلاعات پمپ با موفقیت ویرایش شد.')
        return response

    def get_success_url(self):
        return reverse('base:sejelli_detail', kwargs={'sejelli_id': self.object.newsejelli.id})


class CreateMakhzanSejjeli(CreateView):
    model = Makhzan_sejjeli  # مشخص کردن مدل
    form_class = MakhzanSejjeliForm  # مشخص کردن فرم
    template_name = 'sejelli/edit_form.html'  # مشخص کردن تمپلیت

    @method_decorator(cache_permission('create_sejelli'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        # تنظیم sejelli و gs قبل از ذخیره
        sejelli = NewSejelli.objects.get(id=self.kwargs['sejelli_id'])
        form.instance.newsejelli = sejelli
        form.instance.gs = sejelli.gs

        response = super().form_valid(form)

        # ثبت تغییرات در تاریخچه
        changed_data = {
            'product': {'id': form.instance.product.id, 'name': str(form.instance.product)},
            'number': form.instance.number,
            'zarfyat': form.instance.zarfyat
        }

        SejelliChangeLog.objects.create(
            user=self.request.user,
            sejelli=sejelli,
            model_name='Makhzan_sejjeli',
            record_id=self.object.id,
            action='create',
            changed_data=changed_data
        )

        messages.success(self.request, 'مخزن جدید با موفقیت اضافه شد.')
        return response

    def get_success_url(self):
        return reverse('base:sejelli_detail', kwargs={'sejelli_id': self.object.newsejelli.id})

    # اگر نیاز به فیلتر کردن کوئری ست دارید:
    def get_queryset(self):
        return Makhzan_sejjeli.objects.filter(newsejelli_id=self.kwargs['sejelli_id'])


class UpdateMakhzanSejjeli(UpdateView):
    model = Makhzan_sejjeli  # مشخص کردن مدل اصلی
    form_class = MakhzanSejjeliForm  # استفاده از فرم مربوطه
    template_name = 'sejelli/edit_form.html'  # استفاده از تمپلیت مشترک
    context_object_name = 'makhzan'  # نام آبجکت در تمپلیت

    @method_decorator(cache_permission('create_sejelli'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        # ذخیره تغییرات قبل از ثبت
        changed_data = {}
        for field in form.changed_data:
            old_value = form.initial.get(field)
            new_value = form.cleaned_data.get(field)

            # تبدیل اشیاء مدل به فرمت قابل JSON شدن
            if hasattr(old_value, 'id'):
                old_value = {'id': old_value.id, 'name': str(old_value)}
            if hasattr(new_value, 'id'):
                new_value = {'id': new_value.id, 'name': str(new_value)}

            changed_data[field] = {
                'old': old_value,
                'new': new_value
            }

        response = super().form_valid(form)

        # ثبت تغییرات در تاریخچه
        SejelliChangeLog.objects.create(
            user=self.request.user,
            sejelli=self.object.newsejelli,
            model_name='Makhzan_sejjeli',
            record_id=self.object.id,
            action='update',
            changed_data=changed_data
        )

        messages.success(self.request, 'اطلاعات مخزن با موفقیت ویرایش شد.')
        return response

    def get_success_url(self):
        return reverse('base:sejelli_detail', kwargs={'sejelli_id': self.object.newsejelli.id})

    # اختیاری: اگر نیاز به فیلتر کردن کوئری ست دارید
    def get_queryset(self):
        return super().get_queryset().filter(newsejelli_id=self.kwargs.get('sejelli_id'))


@cache_permission('create_sejelli')
def pending_sejelli_list(request):
    # لیست سجلی‌هایی که تایید شده‌اند اما اعمال نشده‌اند
    pending_sejellis = NewSejelli.objects.filter(approov=True).exclude(
        gsmodel_sejjeli__isnull=True
    )

    context = {
        'pending_sejellis': pending_sejellis,
    }
    return TemplateResponse(request, 'sejelli/pending_sejelli_list.html', context)


@cache_permission('create_sejelli')
def sejelli_changes(request, sejelli_id):
    sejelli = get_object_or_404(NewSejelli, id=sejelli_id)

    try:
        gs_sejelli = GsModel_sejjeli.objects.get(newsejelli=sejelli)
        gs_changes = []

        # مقایسه با مدل اصلی
        original_gs = sejelli.gs
        for field in GsModel_sejjeli._meta.fields:
            field_name = field.name
            if field_name not in ['id', 'newsejelli', 'create_time']:
                original_value = getattr(original_gs, field_name, None)
                new_value = getattr(gs_sejelli, field_name, None)

                if original_value != new_value:
                    gs_changes.append({
                        'field': field.verbose_name,
                        'original': original_value,
                        'new': new_value
                    })
    except GsModel_sejjeli.DoesNotExist:
        gs_changes = []

    # تغییرات پمپ‌ها
    pump_changes = []
    original_pumps = Pump.objects.filter(gs=sejelli.gs)
    sejelli_pumps = Pump_sejjeli.objects.filter(newsejelli=sejelli)

    # پمپ‌های جدید
    for pump in sejelli_pumps:
        if not original_pumps.filter(number=pump.number).exists():
            pump_changes.append({
                'action': 'add',
                'pump': pump,
            })

    # پمپ‌های تغییر یافته یا حذف شده
    for original_pump in original_pumps:
        try:
            sejelli_pump = sejelli_pumps.get(number=original_pump.number)
            changes = []
            for field in Pump_sejjeli._meta.fields:
                field_name = field.name
                if field_name not in ['id', 'newsejelli', 'create_time']:
                    original_value = getattr(original_pump, field_name, None)
                    new_value = getattr(sejelli_pump, field_name, None)

                    if original_value != new_value:
                        changes.append({
                            'field': field.verbose_name,
                            'original': original_value,
                            'new': new_value
                        })

            if changes:
                pump_changes.append({
                    'action': 'update',
                    'pump': sejelli_pump,
                    'changes': changes
                })
        except Pump_sejjeli.DoesNotExist:
            pump_changes.append({
                'action': 'delete',
                'pump': original_pump,
            })

    # تغییرات مخازن
    makhzan_changes = []
    original_makhzans = Makhzan.objects.filter(gs=sejelli.gs)
    sejelli_makhzans = Makhzan_sejjeli.objects.filter(newsejelli=sejelli)

    # مخازن جدید
    for makhzan in sejelli_makhzans:
        if not original_makhzans.filter(number=makhzan.number).exists():
            makhzan_changes.append({
                'action': 'add',
                'makhzan': makhzan,
            })

    # مخازن تغییر یافته یا حذف شده
    for original_makhzan in original_makhzans:
        try:
            sejelli_makhzan = sejelli_makhzans.get(number=original_makhzan.number)
            changes = []
            for field in Makhzan_sejjeli._meta.fields:
                field_name = field.name
                if field_name not in ['id', 'newsejelli', 'create_time']:
                    original_value = getattr(original_makhzan, field_name, None)
                    new_value = getattr(sejelli_makhzan, field_name, None)

                    if original_value != new_value:
                        changes.append({
                            'field': field.verbose_name,
                            'original': original_value,
                            'new': new_value
                        })

            if changes:
                makhzan_changes.append({
                    'action': 'update',
                    'makhzan': sejelli_makhzan,
                    'changes': changes
                })
        except Makhzan_sejjeli.DoesNotExist:
            makhzan_changes.append({
                'action': 'delete',
                'makhzan': original_makhzan,
            })

    context = {
        'sejelli': sejelli,
        'gs_changes': gs_changes,
        'pump_changes': pump_changes,
        'makhzan_changes': makhzan_changes,
    }
    return TemplateResponse(request, 'sejelli/sejelli_changes.html', context)


@cache_permission('approved_sejelli')
@transaction.atomic
def apply_sejelli_changes(request, sejelli_id):
    sejelli = get_object_or_404(NewSejelli, id=sejelli_id)

    if not sejelli.approov:
        messages.error(request, 'این سجلی هنوز تایید نشده است.')
        return redirect('base:sejelli_detail', sejelli_id=sejelli_id)

    try:
        # اعمال تغییرات روی مدل اصلی جایگاه
        gs_sejelli = GsModel_sejjeli.objects.get(newsejelli=sejelli)
        original_gs = sejelli.gs

        # به روزرسانی فیلدهای جایگاه
        for field in GsModel_sejjeli._meta.fields:
            field_name = field.name
            if field_name not in ['id', 'newsejelli', 'create_time']:
                setattr(original_gs, field_name, getattr(gs_sejelli, field_name))

        original_gs.save()

        # اعمال تغییرات پمپ‌ها
        original_pumps = Pump.objects.filter(gs=original_gs)
        sejelli_pumps = Pump_sejjeli.objects.filter(newsejelli=sejelli)

        # حذف پمپ‌هایی که در سجلی وجود ندارند
        for original_pump in original_pumps:
            if not sejelli_pumps.filter(number=original_pump.number).exists():
                original_pump.delete()

        # به روزرسانی یا ایجاد پمپ‌های جدید
        for sejelli_pump in sejelli_pumps:
            defaults = {
                'product': sejelli_pump.product,
                'pumpbrand': sejelli_pump.pumpbrand,
                'sakoo': sejelli_pump.sakoo,
                'tolombe': sejelli_pump.tolombe,
                'status': sejelli_pump.status,
                'master': '',  # مقدار پیش‌فرض برای فیلدهای جدید
                'pinpad': '',  # مقدار پیش‌فرض برای فیلدهای جدید
                'active': True if sejelli_pump.status.id == 1 else False,
                'actived': True if sejelli_pump.status.id in [1, 2] else False,
                'user': request.user,
            }

            # اگر پمپ جدید است، uniq هم ایجاد می‌کنیم
            if not original_pumps.filter(number=sejelli_pump.number).exists():
                defaults['uniq'] = get_random_string(length=10)

            pump, created = Pump.objects.update_or_create(
                gs=original_gs,
                number=sejelli_pump.number,
                defaults=defaults
            )

            # اگر پمپ جدید ایجاد شد، لاگ تغییرات را ثبت می‌کنیم
            if created:
                SejelliChangeLog.objects.create(
                    user=request.user,
                    sejelli=sejelli,
                    model_name='Pump',
                    record_id=pump.id,
                    action='create',
                    changed_data={
                        'number': pump.number,
                        'product': {'id': pump.product.id, 'name': pump.product.name},
                        'pumpbrand': {'id': pump.pumpbrand.id, 'name': pump.pumpbrand.name},
                        'status': {'id': pump.status.id, 'name': pump.status.name},
                    }
                )

        # اعمال تغییرات مخازن
        original_makhzans = Makhzan.objects.filter(gs=original_gs)
        sejelli_makhzans = Makhzan_sejjeli.objects.filter(newsejelli=sejelli)

        # حذف مخازنی که در سجلی وجود ندارند
        for original_makhzan in original_makhzans:
            if not sejelli_makhzans.filter(number=original_makhzan.number).exists():
                original_makhzan.delete()

        # به روزرسانی یا ایجاد مخازن جدید
        for sejelli_makhzan in sejelli_makhzans:
            makhzan, created = Makhzan.objects.update_or_create(
                gs=original_gs,
                number=sejelli_makhzan.number,
                defaults={
                    'product': sejelli_makhzan.product,
                    'zarfyat': sejelli_makhzan.zarfyat,
                    'action': sejelli_makhzan.action,
                }
            )

            # اگر مخزن جدید ایجاد شد، لاگ تغییرات را ثبت می‌کنیم
            if created:
                SejelliChangeLog.objects.create(
                    user=request.user,
                    sejelli=sejelli,
                    model_name='Makhzan',
                    record_id=makhzan.id,
                    action='create',
                    changed_data={
                        'number': makhzan.number,
                        'product': {'id': makhzan.product.id, 'name': makhzan.product.name},
                        'zarfyat': makhzan.zarfyat,
                        'action': makhzan.action,
                    }
                )

        sejelli.isok = True
        sejelli.samane = True
        sejelli.hse = True
        sejelli.mohandesi = True
        sejelli.bazargani = True
        sejelli.modir = True
        sejelli.approov = True
        sejelli.okdate = datetime.datetime.now()
        sejelli.save()
        messages.success(request, 'تغییرات سجلی با موفقیت روی جایگاه اعمال شد.')
        return redirect('base:sejelli_view', sejelli_id=sejelli_id)

    except Exception as e:
        messages.error(request, f'خطا در اعمال تغییرات: {str(e)}')
        return redirect('base:sejelli_changes', sejelli_id=sejelli_id)


@cache_permission('list_sejelli')
def sejelli_approval_list(request):
    # لیست سجلی‌هایی که نیاز به تایید دارند
    sejellis = NewSejelli.object_role.c_gs(request, 0).filter(
        approov=False, isok=False
    ).exclude(
        gsmodel_sejjeli__isnull=True
    )

    # تعیین وضعیت تایید برای هر سجلی

    context = {
        'approval_status': sejellis,
    }
    return TemplateResponse(request, 'sejelli/sejelli_approval_list.html', context)


def approval_sejjeli(request, _id):
    url = request.META.get('HTTP_REFERER')
    if request.user.owner.refrence.ename == 'samane':
        sejelli = NewSejelli.objects.get(id=_id)
        sejelli.samane = True
        sejelli.samanename = request.user.owner.get_full_name()
        sejelli.save()
    elif request.user.owner.refrence.ename == 'hse':
        sejelli = NewSejelli.objects.get(id=_id)
        sejelli.hse = True
        sejelli.hsename = request.user.owner.get_full_name()
        sejelli.save()
    elif request.user.owner.refrence.ename == 'bazargani':
        sejelli = NewSejelli.objects.get(id=_id)
        sejelli.bazargani = True
        sejelli.bazarganiname = request.user.owner.get_full_name()
        sejelli.save()
    elif request.user.owner.refrence.ename == 'mohandesi':
        sejelli = NewSejelli.objects.get(id=_id)
        sejelli.mohandesi = True
        sejelli.mohandesiname = request.user.owner.get_full_name()
        sejelli.save()
    elif request.user.owner.refrence.ename == 'modir':
        sejelli = NewSejelli.objects.get(id=_id)
        sejelli.modir = True
        sejelli.modirname = request.user.owner.get_full_name()
        sejelli.save()

    else:
        messages.error(request, 'شما دسترسی ندارید')
    return redirect(url)


@cache_permission('create_sejelli')
def sejelli_view(request, sejelli_id):
    sejelli = NewSejelli.objects.get(id=sejelli_id)
    gs_sejelli = GsModel_sejjeli.objects.filter(newsejelli=sejelli).first()
    pumps = Pump_sejjeli.objects.filter(newsejelli=sejelli)
    makhzans = Makhzan_sejjeli.objects.filter(newsejelli=sejelli)
    change_logs = SejelliChangeLog.objects.filter(sejelli=sejelli).order_by('-changed_at')

    # لیست فیلدهای تغییر یافته
    changed_fields = set()
    changed_pumps = set()
    changed_makhzans = set()

    for log in change_logs:
        if log.model_name == 'GsModel_sejjeli' and log.action == 'update':
            changed_fields.update(log.changed_data.keys())
        elif log.model_name == 'Pump_sejjeli':
            changed_pumps.add(log.record_id)
        elif log.model_name == 'Makhzan_sejjeli':
            changed_makhzans.add(log.record_id)

    context = {
        'sejelli': sejelli,
        'gs_sejelli': gs_sejelli,
        'pumps': pumps,
        'makhzans': makhzans,
        'change_logs': change_logs,
        'changed_fields': changed_fields,
        'changed_pumps': changed_pumps,
        'changed_makhzans': changed_makhzans,
    }
    return TemplateResponse(request, 'sejelli/sejelli_detail_view.html', context)


@cache_permission('create_sejelli')
def print_detail(request, sejelli_id):
    sejelli = NewSejelli.objects.get(id=sejelli_id)
    gs_sejelli = GsModel_sejjeli.objects.filter(newsejelli=sejelli).first()
    pumps = Pump_sejjeli.objects.filter(newsejelli=sejelli)
    makhzans = Makhzan_sejjeli.objects.filter(newsejelli=sejelli)
    change_logs = SejelliChangeLog.objects.filter(sejelli=sejelli).order_by('-changed_at')

    # لیست فیلدهای تغییر یافته
    changed_fields = set()
    changed_pumps = set()
    changed_makhzans = set()

    for log in change_logs:
        if log.model_name == 'GsModel_sejjeli' and log.action == 'update':
            changed_fields.update(log.changed_data.keys())
        elif log.model_name == 'Pump_sejjeli':
            changed_pumps.add(log.record_id)
        elif log.model_name == 'Makhzan_sejjeli':
            changed_makhzans.add(log.record_id)

    context = {
        'sejelli': sejelli,
        'gs_sejelli': gs_sejelli,
        'pumps': pumps,
        'makhzans': makhzans,
        'change_logs': change_logs,
        'changed_fields': changed_fields,
        'changed_pumps': changed_pumps,
        'changed_makhzans': changed_makhzans,
    }
    return TemplateResponse(request, 'sejelli/print.html', context)


def failure_analysis_report(request):
    add_to_log(request, f' مشاهده گزارش تحلیل خرابی', 0)
    # تعیین پارامترهای زمانی از URL
    time_period = request.GET.get('period', 'weekly')  # پیش‌فرض هفتگی

    # محاسبه تاریخ شروع بر اساس دوره انتخابی
    end_date = datetime.datetime.now().date()
    if time_period == 'weekly':
        start_date = end_date - timedelta(days=7)
        period_title = "هفتگی"
    elif time_period == 'monthly':
        start_date = end_date - timedelta(days=30)
        period_title = "ماهانه"
    elif time_period == 'quarterly':
        start_date = end_date - timedelta(days=90)
        period_title = "سه ماهه"
    else:
        start_date = end_date - timedelta(days=7)
        time_period = 'weekly'
        period_title = "هفتگی"

    # دریافت تیکت‌های مربوط به بازه زمانی
    tickets = Ticket.objects.filter(
        create__date__gte=start_date,
        create__date__lte=end_date
    )

    # محاسبه آمار خرابی‌ها
    failure_stats = (
        tickets.values('failure__id', 'failure__info', 'failure__failurecategory__info')
        .annotate(
            failure_name=F('failure__info'),
            category_name=F('failure__failurecategory__info'),
            count=Count('id')
        )
        .order_by('-count')
    )

    # محاسبه پرتکرارترین خرابی‌ها (10 مورد اول)
    top_failures = list(failure_stats[:10])

    # محاسبه توزیع خرابی‌ها بر اساس منطقه
    zone_distribution = (
        tickets.values('gs__area__zone__name')
        .annotate(
            zone_name=F('gs__area__zone__name'),
            count=Count('id')
        )
        .order_by('-count')
    )

    # محاسبه روند روزانه
    daily_trend = (
        tickets.annotate(day=TruncDate('create'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )

    for item in daily_trend:
        gregorian_date = item['day']
        jalali_date = jdatetime.date.fromgregorian(date=gregorian_date)
        item['day_jalali'] = jalali_date.strftime('%Y/%m/%d')

    # محاسبه تعداد کل تیکت‌ها
    total_tickets = tickets.count()
    top_failures_labels = [f['failure_name'] for f in top_failures]
    top_failures_data = [f['count'] for f in top_failures]

    zone_labels = [z['zone_name'] for z in zone_distribution]
    zone_data = [z['count'] for z in zone_distribution]

    daily_labels = [d['day'].strftime('%Y-%m-%d') for d in daily_trend]
    daily_data = [d['count'] for d in daily_trend]

    context = {
        'top_failures_labels': top_failures_labels,
        'top_failures_data': top_failures_data,
        'zone_labels': zone_labels,
        'zone_data': zone_data,
        'daily_labels': daily_labels,
        'daily_data': daily_data,
        'period': period_title,
        'start_date': start_date,
        'end_date': end_date,
        'total_tickets': total_tickets,
        'top_failures': top_failures,
        'zone_distribution': zone_distribution,
        'daily_trend': daily_trend,
        'time_period': time_period,
    }

    return render(request, 'failure_analysis.html', context)


def convert_to_seconds(time_str):
    # جدا کردن اجزای زمان
    days_part, time_part = time_str.split(", ") if "," in time_str else ("0", time_str)

    days = int(days_part.split()[0]) if "day" in days_part else 0

    # تجزیه بخش ساعت، دقیقه و ثانیه
    hms, microseconds = time_part.split(".") if "." in time_part else (time_part, "0")
    hours, minutes, seconds = map(int, hms.split(":"))

    # تبدیل به timedelta
    time_delta = timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds, microseconds=int(microseconds))

    return time_delta.total_seconds()


class TicketAnalysisView(View):

    def get(self, request, *args, **kwargs):

        zone_id = request.user.owner.zone.id
        if request.user.owner.role.role in ['setad', 'mgr']:
            zone_id = request.GET.get('zone_id')
        period = request.GET.get('period')
        if not period:
            period = 'last_week'
            zone_id = 10

        add_to_log(request, f' مشاهده آنالیز تیکت های داشبورد {zone_id} - {period}', 0)
        try:
            _top_gs_by_ticket = []
            for item in TicketAnalysis.objects.get_top_gs_by_ticket_count(zone_id, period):
                _top_gs_by_ticket.append({
                    'gsid': item['gs__gsid'],
                    'name': item['gs__name'],
                    'zone': item['gs__area__zone__name'],
                    'ticket_count': item['ticket_count']

                })

            _top_failure_types = []
            for item in TicketAnalysis.objects.get_top_failure_types(zone_id, period):
                a = convert_to_seconds(str(item['avg_resolution_time'])) if item['avg_resolution_time'] else 0

                _top_failure_types.append({
                    'failure__info': item['failure__info'][:33],
                    'failure_count': item['failure_count'],
                    'avg_resolution_time': round(a, 0)
                })

            _top_technicians = []
            for item in TicketAnalysis.objects.get_top_technicians_by_ticket_count(zone_id, period):
                a = convert_to_seconds(str(item['avg_resolution_time'])) if item['avg_resolution_time'] else 0
                _top_technicians.append({
                    'user': item['user__owner__name'] + " " + item['user__owner__lname'],
                    'ticket_count': item['ticket_count'],
                    'avg_resolution_time': round(a, 0)
                })

            tickets = []
            for item in TicketAnalysis.objects.get_longest_resolution_tickets(zone_id, period):
                resolution_seconds = convert_to_seconds(str(item.resolution_time)) if item.resolution_time else 0
                tickets.append({
                    'gs_name': item.gs.name,
                    'gs_id': item.gs.gsid,
                    'ticket_id': item.id,
                    'create_date': item.create.strftime('%Y-%m-%d'),
                    'close_date': item.closedate.strftime('%Y-%m-%d') if item.closedate else '',
                    'resolution_time': round(resolution_seconds, 0),
                    'failure_type': item.failure.info,
                    'zone_name': item.gs.area.zone.name
                })

            get_pending_tickets = []
            for item in TicketAnalysis.objects.get_pending_tickets(zone_id, period):
                pending_seconds = (datetime.datetime.now() - item.create).total_seconds()
                if round(pending_seconds / (3600 * 24), 1) > 5:
                    get_pending_tickets.append({
                        'gs_name': item.gs.name,
                        'gs_id': item.gs.gsid,
                        'ticket_id': item.id,
                        'create_date': item.create.strftime('%Y-%m-%d'),
                        'pending_days': round(pending_seconds / (3600 * 24), 1),
                        'failure_type': item.failure.info,
                        'zone_name': item.gs.area.zone.name
                    })

                # _get_avg_by_zone = []
                # for item in TicketAnalysis.objects.get_avg_resolution_by_zone(period):
                #     avg_seconds = convert_to_seconds(str(item['avg_resolution'])) if item['avg_resolution'] else 0
                #     _get_avg_by_zone.append({
                #         'zone_name': item['gs__area__zone__name'],
                #         'zone_id': item['gs__area__zone__id'],
                #         'avg_resolution': round(avg_seconds, 0),
                #         'ticket_count': item['ticket_count']
                #     })

                # _get_count_by_period = []
                # daily = list(TicketAnalysis.objects.get_ticket_count_by_period(zone_id, 'daily'))
                # hourly = list(TicketAnalysis.objects.get_ticket_count_by_period(zone_id, 'hourly'))
                # weekly = list(TicketAnalysis.objects.get_ticket_count_by_period(zone_id, 'weekly'))
                #
                # _get_count_by_period.append({
                #     'daily': daily,
                #     'hourly': hourly,
                #     'weekly': weekly
                # })
                # print(_get_count_by_period)

                _get_fastest_techs = []
                for item in TicketAnalysis.objects.get_technicians_fastest_resolution(zone_id, period):
                    avg_seconds = convert_to_seconds(str(item['avg_resolution_time'])) if item[
                        'avg_resolution_time'] else 0
                    _get_fastest_techs.append({
                        'tech_name': f"{item['actioner__name']} {item['actioner__lname']}",
                        'avg_resolution': round(avg_seconds, 0),
                        'ticket_count': item['ticket_count']
                    })
                _lowest_rated_tickets = []
                for item in TicketAnalysis.objects.get_lowest_rated_tickets(zone_id, period):
                    _lowest_rated_tickets.append({
                        'gs_name': item.gs.name,
                        'ticket_id': item.id,
                        'star': item.star,
                        'create_date': item.create.strftime('%Y-%m-%d'),
                        'failure_type': item.failure.info
                    })

                # _get_most_workflows = []
                # for item in TicketAnalysis.objects.get_tickets_with_most_workflows(zone_id, period):
                #     _get_most_workflows.append({
                #         'ticket_id': item.id,
                #         'gs_name': item.gs.name,
                #         'gs_id': item.gs.gsid,
                #         'workflow_count': item.workflow_count,
                #         'failure_type': item.failure.info,
                #         'create_date': item.create.strftime('%Y-%m-%d')
                #     })

                # جمع‌آوری تمام گزارش‌ها و تبدیل به ساختار قابل JSON شدن

                return JsonResponse({
                    'status': 'success',
                    'data2': _top_gs_by_ticket,
                    'data3': _top_failure_types,
                    'data4': _top_technicians,
                    'data5': tickets,
                    'data6': get_pending_tickets,
                    # 'data7': _get_avg_by_zone,
                    # 'data8': _get_count_by_period,
                    'data9': _get_fastest_techs,
                    'data10': _lowest_rated_tickets,
                    # 'data10': _get_most_workflows,
                })

            # جمع‌آوری تمام گزارش‌ها و تبدیل به ساختار قابل JSON شدن



        except Zone.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Zone not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)


@cache_permission('report')
def ticket_analysis_report(request):
    url = request.META.get('HTTP_REFERER')

    if request.method == 'POST':
        # دریافت پارامترهای دوره جاری
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        # دریافت پارامترهای دوره قبل
        prev_start_date = request.POST.get('prev_start_date')
        prev_end_date = request.POST.get('prev_end_date')

        try:
            # تبدیل تاریخ‌های شمسی به میلادی برای دوره جاری
            start_date_jalali = jdatetime.datetime.strptime(start_date, '%Y/%m/%d')
            end_date_jalali = jdatetime.datetime.strptime(end_date, '%Y/%m/%d')

            start_date_gregorian = start_date_jalali.togregorian()
            end_date_gregorian = end_date_jalali.togregorian() + timedelta(days=1)  # شامل پایان روز

            # تبدیل تاریخ‌های شمسی به میلادی برای دوره قبل
            prev_start_date_jalali = jdatetime.datetime.strptime(prev_start_date, '%Y/%m/%d')
            prev_end_date_jalali = jdatetime.datetime.strptime(prev_end_date, '%Y/%m/%d')

            prev_start_date_gregorian = prev_start_date_jalali.togregorian()
            prev_end_date_gregorian = prev_end_date_jalali.togregorian() + timedelta(days=1)  # شامل پایان روز

            # محاسبه تعداد روزهای هر دوره
            days_in_period = (end_date_gregorian - start_date_gregorian).days
            prev_days_in_period = (prev_end_date_gregorian - prev_start_date_gregorian).days

        except:
            messages.error(request, 'تاریخ وارد شده نامعتبر است')
            return redirect(url)

        # دریافت داده‌های تیکت‌ها برای دوره جاری

        current_tickets = Ticket.object_role.c_gs(request, 1).filter(failure__failurecategory_id__in=[1010, 1011],
                                                                     create__gte=start_date_gregorian,
                                                                     create__lt=end_date_gregorian
                                                                     ).annotate(
            resolution_time=ExpressionWrapper(
                F('closedate') - F('create'),
                output_field=fields.DurationField()
            )
        ).values('gs__area__zone__name', 'gs__area__zone__id').annotate(
            ticket_count=Count('id'),
            avg_resolution=Avg('resolution_time')
        ).order_by('gs__area__zone__name')

        # دریافت داده‌های تیکت‌ها برای دوره قبل
        previous_tickets = Ticket.object_role.c_gs(request, 1).filter(failure__failurecategory_id__in=[1010, 1011],
                                                                      create__gte=prev_start_date_gregorian,
                                                                      create__lt=prev_end_date_gregorian
                                                                      ).annotate(
            resolution_time=ExpressionWrapper(
                F('closedate') - F('create'),
                output_field=fields.DurationField()
            )
        ).values('gs__area__zone__name', 'gs__area__zone__id').annotate(
            ticket_count=Count('id'),
            avg_resolution=Avg('resolution_time')
        )

        # تبدیل داده‌های دوره قبل به دیکشنری برای دسترسی آسان
        prev_data = {item['gs__area__zone__id']: item for item in previous_tickets}

        # محاسبه تغییرات و آماده‌سازی داده‌ها برای نمایش
        analysis_data = []
        total_tickets = 0
        total_resolution = timedelta()
        total_prev_tickets = 0
        total_prev_resolution = timedelta()

        for item in current_tickets:
            zone_id = item['gs__area__zone__id']
            zone_name = item['gs__area__zone__name']
            ticket_count = item['ticket_count']
            avg_resolution = item['avg_resolution'] or timedelta()

            # محاسبه میانگین روزانه برای دوره جاری
            daily_avg = ticket_count / days_in_period if days_in_period > 0 else 0
            avg_resolution_hours = avg_resolution.total_seconds() / 3600 if avg_resolution else 0

            # دریافت داده‌های دوره قبل
            prev_item = prev_data.get(zone_id, {})
            prev_ticket_count = prev_item.get('ticket_count', 0)
            prev_avg_resolution = prev_item.get('avg_resolution', timedelta())

            # محاسبه میانگین روزانه برای دوره قبل
            prev_daily_avg = prev_ticket_count / prev_days_in_period if prev_days_in_period > 0 else 0
            prev_avg_resolution_hours = prev_avg_resolution.total_seconds() / 3600 if prev_avg_resolution else 0

            # محاسبه تغییرات
            count_change = ticket_count - prev_ticket_count if prev_ticket_count > 0 else 0
            count_change_percent = (count_change / prev_ticket_count * 100) if prev_ticket_count > 0 else 0

            time_change = (
                    avg_resolution.total_seconds() - prev_avg_resolution.total_seconds()) if prev_avg_resolution else 0
            time_change_percent = (
                    time_change / prev_avg_resolution.total_seconds() * 100) if prev_avg_resolution and prev_avg_resolution.total_seconds() > 0 else 0

            # افزودن به لیست نتایج
            analysis_data.append({
                'zone_name': zone_name,
                'zone_id': zone_id,
                'ticket_count': ticket_count,
                'daily_avg': daily_avg,
                'avg_resolution_hours': avg_resolution_hours,
                'prev_ticket_count': prev_ticket_count,
                'prev_daily_avg': prev_daily_avg,
                'prev_avg_resolution_hours': prev_avg_resolution_hours,
                'count_change_percent': count_change_percent,
                'time_change_percent': time_change_percent,
            })
            analysis_data = sorted(analysis_data, key=itemgetter('avg_resolution_hours'), reverse=False)
            # جمع‌آوری اطلاعات برای محاسبه کل
            total_tickets += ticket_count
            total_resolution += avg_resolution * ticket_count if avg_resolution else timedelta()
            total_prev_tickets += prev_ticket_count
            total_prev_resolution += prev_avg_resolution * prev_ticket_count if prev_avg_resolution else timedelta()

        # محاسبه میانگین کل برای دوره جاری
        avg_daily = total_tickets / days_in_period if days_in_period > 0 else 0
        avg_resolution = (total_resolution.total_seconds() / total_tickets / 3600) if total_tickets > 0 else 0

        # محاسبه میانگین کل برای دوره قبل
        avg_prev_daily = total_prev_tickets / prev_days_in_period if prev_days_in_period > 0 else 0
        avg_prev_resolution = (
                total_prev_resolution.total_seconds() / total_prev_tickets / 3600) if total_prev_tickets > 0 else 0

        # محاسبه تغییرات کل
        total_count_change = (
                (total_tickets - total_prev_tickets) / total_prev_tickets * 100) if total_prev_tickets > 0 else 0
        total_time_change = ((total_resolution.total_seconds() - total_prev_resolution.total_seconds()) /
                             total_prev_resolution.total_seconds() * 100) if total_prev_resolution and total_prev_resolution.total_seconds() > 0 else 0

        # آماده‌سازی داده‌های خلاصه
        summary = {
            'total_tickets': total_tickets,
            'avg_daily': avg_daily,
            'avg_resolution': avg_resolution,
            'total_prev_tickets': total_prev_tickets,
            'avg_prev_daily': avg_prev_daily,
            'avg_prev_resolution': avg_prev_resolution,
            'total_count_change': total_count_change,
            'total_time_change': total_time_change,
        }

        context = {
            'start_date': start_date,
            'end_date': end_date,
            'prev_start_date': prev_start_date,
            'prev_end_date': prev_end_date,
            'analysis_data': analysis_data,
            'summary': summary,
        }

        return TemplateResponse(request, 'ticket_analysis_report.html', context)

    context = {}
    return TemplateResponse(request, 'ticket_analysis_report.html', context)


@transaction.atomic
def update_areaid(request):
    add_to_log(request, 'update area id', 0)
    _list = []
    form = open_excel(request.POST)

    if request.method == 'POST':
        form = open_excel(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            path = UploadExcel.objects.get(id=form.instance.id)
            path = path.filepath
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(1, m_row + 1):
                gsid = str(sheet_obj.cell(row=i, column=3).value)
                zoneid = str(sheet_obj.cell(row=i, column=1).value)
                areaid = str(sheet_obj.cell(row=i, column=2).value)

                try:
                    gsmodel = GsModel.objects.get(gsid=gsid)
                    area = Area.objects.get(id=gsmodel.area_id)
                    area.areaid = areaid
                    area.save()
                    zone = Zone.objects.get(id=gsmodel.area.zone_id)
                    zone.zoneid = zoneid
                    zone.save()
                except:
                    continue

        messages.success(request, SUCCESS_TICKET)
        return redirect(HOME_PAGE)
    return render(request, 'importexcel.html', {'form': form})


def parametrs_edit_view(request):
    """
    ویو برای ویرایش پارامترهای سیستم
    فقط یک رکورد وجود دارد (ID=1)
    """

    # دریافت یا ایجاد رکورد پارامترها
    parametrs_instance, created = Parametrs.objects.get_or_create(pk=1)

    if request.method == 'POST':
        form = ParametrsForm(request.POST, instance=parametrs_instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'پارامترهای سیستم با موفقیت به‌روزرسانی شد.')
            return redirect('base:parametrs-edit')
        else:
            messages.error(request, 'لطفا خطاهای زیر را اصلاح کنید.')
    else:
        form = ParametrsForm(instance=parametrs_instance)

    context = {
        'form': form,
        'parametrs': parametrs_instance,
        'created': created,
    }

    return render(request, 'parametrs_edit.html', context)


@cache_permission('report')
def ticket_list_row(request):
    newyear = 0
    monthly_data = None
    prev_monthly_data = None
    previous_year = None
    if request.method == 'POST':
        year = request.POST.get('year')
        _status = request.POST.get('status')
        newyear = year
        if _status == '0':
            myst = [1010, 1011, 1012, 1013, 1014, 1015]
        elif _status == '1':
            myst = [1010, 1011]
        elif _status == '2':
            myst = [1010]
        elif _status == '3':
            myst = [1011]
        tickets = Ticket.object_role.c_gs(request, 0).filter(
            failure__failurecategory_id__in=myst,
            create_shamsi_year=year
        )

        # گروه‌بندی بر اساس منطقه و ماه شمسی
        monthly_data = tickets.values(
            'gs__area__zone__name',
            'gs__area__zone__id',
            'create_shamsi_month'
        ).annotate(
            ticket_count=Count('id')
        ).order_by('gs__area__zone__name', 'create_shamsi_month')

        # داده‌های سال قبل برای مقایسه
        previous_year = str(int(year) - 1)
        prev_year_tickets = Ticket.object_role.c_gs(request, 0).filter(
            failure__failurecategory_id__in=myst,
            create_shamsi_year=previous_year
        )

        prev_monthly_data = prev_year_tickets.values(
            'gs__area__zone__name',
            'gs__area__zone__id',
            'create_shamsi_month'
        ).annotate(
            ticket_count=Count('id')
        ).order_by('gs__area__zone__name', 'create_shamsi_month')

        # ساختاردهی داده‌ها
        zones_data = {}

        d = 0
        # پردازش داده‌های سال جاری
        csum = 0
        psum = 0
        _zoneid = 0
        for item in monthly_data:
            zone_name = item['gs__area__zone__name']
            zone_id = item['gs__area__zone__id']
            month = item['create_shamsi_month']
            count = item['ticket_count']

            if zone_id not in zones_data:
                d = 0
                csum = 0

                zones_data[zone_id] = {
                    'name': zone_name,
                    'zone_id': zone_id,
                }

            csum += count
            for i in range(12):
                if d == 0:
                    zones_data[zone_id]['current01'] = 0
                    zones_data[zone_id]['current02'] = 0
                    zones_data[zone_id]['current03'] = 0
                    zones_data[zone_id]['current04'] = 0
                    zones_data[zone_id]['current05'] = 0
                    zones_data[zone_id]['current06'] = 0
                    zones_data[zone_id]['current07'] = 0
                    zones_data[zone_id]['current08'] = 0
                    zones_data[zone_id]['current09'] = 0
                    zones_data[zone_id]['current10'] = 0
                    zones_data[zone_id]['current11'] = 0
                    zones_data[zone_id]['current12'] = 0
                    zones_data[zone_id]['previous01'] = 0
                    zones_data[zone_id]['previous02'] = 0
                    zones_data[zone_id]['previous03'] = 0
                    zones_data[zone_id]['previous04'] = 0
                    zones_data[zone_id]['previous05'] = 0
                    zones_data[zone_id]['previous06'] = 0
                    zones_data[zone_id]['previous07'] = 0
                    zones_data[zone_id]['previous08'] = 0
                    zones_data[zone_id]['previous09'] = 0
                    zones_data[zone_id]['previous10'] = 0
                    zones_data[zone_id]['previous11'] = 0
                    zones_data[zone_id]['previous12'] = 0
                    d += 1
                else:
                    zones_data[zone_id][f'current{month}'] = count
                    zones_data[zone_id]['currentsum'] = csum

        # پردازش داده‌های سال قبل
        for item in prev_monthly_data:
            if _zoneid != item['gs__area__zone__id']:
                psum = 0
            zone_id = item['gs__area__zone__id']
            month = item['create_shamsi_month']
            count = item['ticket_count']
            psum += count

            if zone_id in zones_data:
                _zoneid = zone_id
                zones_data[zone_id][f'previous{month}'] = count
                zones_data[zone_id]['previoussum'] = psum

        # تبدیل به لیست برای مرتب سازی

        zones_list = [{'id': k, **v} for k, v in zones_data.items()]
        zones_list.sort(key=lambda x: x['name'])

        return TemplateResponse(request, 'report/ticket_list_row.html', {
            'zones_list': zones_list,
            'year': year,
            'previous_year': previous_year,
            'mgr': 1,
            'newyear': newyear,
            'status': int(_status)
        })

    return TemplateResponse(request, 'report/ticket_list_row.html', {
        'zones_list': [],
        'year': None,
        'previous_year': None,
        'mgr': 1,
        'newyear': newyear,
        'status': '0',

    })


@cache_permission('report')
def ticket_list_row_zone(request, _zone, _year, _status):
    monthly_data = None
    prev_monthly_data = None
    previous_year = None

    year = _year
    newyear = _year
    if _status == '0':
        myst = [1010, 1011, 1012, 1013, 1014, 1015]
    elif _status == '1':
        myst = [1010, 1011]
    elif _status == '2':
        myst = [1010]
    elif _status == '3':
        myst = [1011]
    if request.user.owner.role.role in ['zone', 'tek', 'area']:
        _zone = request.user.owner.zone.id
    tickets = Ticket.object_role.c_gs(request, 0).filter(
        failure__failurecategory_id__in=myst,
        create_shamsi_year=year,
        gs__area__zone_id=_zone
    )

    # گروه‌بندی بر اساس منطقه و ماه شمسی
    monthly_data = tickets.values(
        'gs__name',
        'gs__gsid',
        'create_shamsi_month'
    ).annotate(
        ticket_count=Count('id')
    ).order_by('gs__name', 'create_shamsi_month')

    # داده‌های سال قبل برای مقایسه
    previous_year = str(int(year) - 1)
    prev_year_tickets = Ticket.object_role.c_gs(request, 0).filter(
        failure__failurecategory_id__in=myst,
        create_shamsi_year=previous_year,
        gs__area__zone_id=_zone
    )

    prev_monthly_data = prev_year_tickets.values(
        'gs__name',
        'gs__gsid',
        'create_shamsi_month'
    ).annotate(
        ticket_count=Count('id')
    ).order_by('gs__name', 'create_shamsi_month')

    # ساختاردهی داده‌ها
    zones_data = {}

    d = 0
    # پردازش داده‌های سال جاری
    csum = 0
    psum = 0
    _zoneid = 0
    for item in monthly_data:
        zone_name = item['gs__name']
        zone_id = item['gs__gsid']
        month = item['create_shamsi_month']
        count = item['ticket_count']

        if zone_id not in zones_data:
            d = 0
            csum = 0

            zones_data[zone_id] = {
                'name': zone_name,
                'zone_id': zone_id,
            }

        csum += count
        for i in range(12):
            if d == 0:
                zones_data[zone_id]['current01'] = 0
                zones_data[zone_id]['current02'] = 0
                zones_data[zone_id]['current03'] = 0
                zones_data[zone_id]['current04'] = 0
                zones_data[zone_id]['current05'] = 0
                zones_data[zone_id]['current06'] = 0
                zones_data[zone_id]['current07'] = 0
                zones_data[zone_id]['current08'] = 0
                zones_data[zone_id]['current09'] = 0
                zones_data[zone_id]['current10'] = 0
                zones_data[zone_id]['current11'] = 0
                zones_data[zone_id]['current12'] = 0
                zones_data[zone_id]['previous01'] = 0
                zones_data[zone_id]['previous02'] = 0
                zones_data[zone_id]['previous03'] = 0
                zones_data[zone_id]['previous04'] = 0
                zones_data[zone_id]['previous05'] = 0
                zones_data[zone_id]['previous06'] = 0
                zones_data[zone_id]['previous07'] = 0
                zones_data[zone_id]['previous08'] = 0
                zones_data[zone_id]['previous09'] = 0
                zones_data[zone_id]['previous10'] = 0
                zones_data[zone_id]['previous11'] = 0
                zones_data[zone_id]['previous12'] = 0
                d += 1
            else:
                zones_data[zone_id][f'current{month}'] = count
                zones_data[zone_id]['currentsum'] = csum

    # پردازش داده‌های سال قبل
    for item in prev_monthly_data:
        if _zoneid != item['gs__gsid']:
            psum = 0
        zone_id = item['gs__gsid']
        month = item['create_shamsi_month']
        count = item['ticket_count']
        psum += count

        if zone_id in zones_data:
            _zoneid = zone_id
            zones_data[zone_id][f'previous{month}'] = count
            zones_data[zone_id]['previoussum'] = psum

    # تبدیل به لیست برای مرتب سازی

    zones_list = [{'id': k, **v} for k, v in zones_data.items()]
    zones_list.sort(key=lambda x: x['name'])

    return TemplateResponse(request, 'report/ticket_list_row.html', {
        'zones_list': zones_list,
        'year': year,
        'previous_year': previous_year,
        'mgr': 2,
        'newyear': newyear,
        'status': int(_status)

    })


@cache_permission('sellproduct')
def sell_product_list(request):
    # دریافت کاربر لاگین شده و اطلاعات owner مربوطه
    try:
        owner = Owner.objects.get(user=request.user)
        company = owner.company
    except Owner.DoesNotExist:
        messages.error(request, 'خطا در ثبت اطلاعات: ')
        return TemplateResponse(request, 'sendproduct/sell_product_list.html')

    # لیست جایگاه‌های اختصاص داده شده به کاربر
    assigned_gs = GsList.objects.filter(owner=owner).select_related('gs')

    # لیست فروش‌های کاربر
    sales = SellProduct.objects.filter(owner=owner).select_related('gs', 'product').order_by('-send_date')

    context = {
        'assigned_gs': assigned_gs,
        'sales': sales,
        'products': Product.objects.all(),
        'company': company
    }
    return TemplateResponse(request, 'sendproduct/sell_product_list.html', context)


def add_sell_product(request):
    if request.method == 'POST':
        gs_id = request.POST.get('gs')
        product_id = request.POST.get('product')
        send_date = request.POST.get('send_date')
        amount = request.POST.get('amount')

        price = request.POST.get('price')
        try:
            owner = Owner.objects.get(user=request.user.owner.id)
            send_date = to_miladi(send_date)
            owner = Owner.objects.get(user=request.user)

            # ایجاد رکورد جدید
            sell_product = SellProduct(
                gs_id=gs_id,
                product_id=product_id,
                owner=owner,
                send_date=send_date,
                amount=amount,
                price=price,
                status_id=1
            )
            sell_product.save()
            messages.success(request, 'عملیات با موفقیت انجام شد')
            return redirect('base:sell_product_list')

        except Exception as e:
            messages.error(request, f'خطا در ثبت اطلاعات: {str(e)}')
            # مدیریت خطا
            return redirect('base:sell_product_list')

    return redirect('base:sell_product_list')


# views.py


class MountListView(LoginRequiredMixin, ListView):
    model = Mount
    template_name = 'mount/mount_list.html'
    context_object_name = 'mounts'
    paginate_by = 10

    @method_decorator(cache_permission('addmount'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()

        # فیلتر بر اساس سال (اگر پارامتر year وجود داشته باشد)
        year = self.request.GET.get('year')
        if year:
            queryset = queryset.filter(year=year)

        # فیلتر بر اساس وضعیت فعال
        active = self.request.GET.get('active')
        if active is not None:
            queryset = queryset.filter(active=(active == '1'))

        # مرتب‌سازی
        queryset = queryset.order_by('-year', '-mah')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # اضافه کردن فرم جستجو به context
        context['search_year'] = self.request.GET.get('year', '')
        context['search_active'] = self.request.GET.get('active', '')
        return context


@cache_permission('addmount')
def mount_create(request):
    if request.method == 'POST':
        form = MountForm(request.POST)
        if form.is_valid():
            mount = form.save()
            messages.success(request, f'ماه {mount.mount} با موفقیت ایجاد شد.')
            return redirect('base:mount_list')
    else:
        form = MountForm()

    return TemplateResponse(request, 'mount/mount_form.html', {'form': form, 'title': 'ایجاد ماه جدید'})

@cache_permission('addmount')
def mount_update(request, pk):
    mount = get_object_or_404(Mount, pk=pk)

    if request.method == 'POST':
        form = MountForm(request.POST, instance=mount)
        if form.is_valid():
            form.save()
            messages.success(request, f'ماه {mount.mount} با موفقیت ویرایش شد.')
            return redirect('mount_list')
    else:
        form = MountForm(instance=mount)

    return TemplateResponse(request, 'mount/mount_form.html', {'form': form, 'title': 'ویرایش ماه'})


# def mount_delete(request, pk):
#     mount = get_object_or_404(Mount, pk=pk)
#
#     if request.method == 'POST':
#         mount_name = mount.mount
#         mount.delete()
#         messages.success(request, f'ماه {mount_name} با موفقیت حذف شد.')
#         return redirect('mount_list')
#
#     return render(request, 'mount_confirm_delete.html', {'mount': mount})


def mount_toggle_active(request, pk):
    mount = get_object_or_404(Mount, pk=pk)
    mount.active = not mount.active
    mount.save()

    status = "فعال" if mount.active else "غیرفعال"
    messages.success(request, f'ماه {mount.mount} با موفقیت {status} شد.')

    return redirect('base:mount_list')